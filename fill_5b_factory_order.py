#!/usr/bin/env python3
"""Fill 5b-订单细节 col B (工厂订单号) with INDEX/MATCH lookup from 6-工厂下单表,
and update col J (问题标注) with a formula-based annotation."""

import shutil
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("/home/user/Panse-System/畔色系统总表_已填.xlsx")
DST_VALUES = Path("/home/user/Panse-System/畔色系统总表_数值版.xlsx")

wb = load_workbook(SRC)
ws5b = wb["5b-订单细节"]
ws6  = wb["6-工厂下单表"]

# Find last data row in 6-工厂下单表 (need the exact range for MATCH)
last6 = 2
for r in range(3, ws6.max_row + 1):
    if ws6.cell(r, 1).value is not None:
        last6 = r
print(f"6-工厂下单表 last data row: {last6}")

# Find last data row in 5b-订单细节
last5b = 2
for r in range(3, ws5b.max_row + 1):
    if ws5b.cell(r, 1).value is not None:
        last5b = r
print(f"5b-订单细节 last data row: {last5b}")

# ── Write formulas for col B and col J ───────────────────────────────────────
for r in range(3, last5b + 1):

    # col B (2): 工厂订单号
    # INDEX/MATCH: find A (platform order) in 6-工厂下单表 col B, return col A
    ws5b.cell(r, 2).value = (
        f"=IFERROR(INDEX('6-工厂下单表'!$A$3:$A${last6},"
        f"MATCH(A{r},'6-工厂下单表'!$B$3:$B${last6},0)),\"\")"
    )

    # col J (10): 问题标注
    # If factory order found → blank (no problem)
    # If not found → note that no factory order is on record
    ws5b.cell(r, 10).value = (
        f"=IF(B{r}<>\"\",\"\","
        f"\"⚠️ 暂无对应工厂订单（现货/库存发货或工厂单待录入）\")"
    )

print(f"Wrote B/J formulas for rows 3–{last5b} ({last5b - 2} rows)")

# ── Update row-1 header note (col J, row 1) ──────────────────────────────────
ws5b.cell(1, 10).value = (
    "工厂订单号由公式从6-工厂下单表反推。"
    "⚠️标注表示该平台订单目前未在6-工厂下单表录入工厂生产单，"
    "可能为现货发货或待补录。"
)

wb.save(SRC)
print(f"Saved formula version → {SRC}")

shutil.copy2(SRC, DST_VALUES)
print(f"Saved values reference copy → {DST_VALUES}")
print("Done.")

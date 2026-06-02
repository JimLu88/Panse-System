#!/usr/bin/env python3
"""5-订单总表 R列(产品理论成本) 改为公式：
总出厂成本(2-定价总表 col T, 按 产品编码+SKU 匹配) × 数量(M)。
产品编码 I 是 P 前缀，定价总表是 PPS 前缀 → "PPS"&MID(I,2,99) 转换。"""

import shutil
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("/home/user/Panse-System/畔色系统总表_已填.xlsx")
DST_VALUES = Path("/home/user/Panse-System/畔色系统总表_数值版.xlsx")

wb = load_workbook(SRC)
ws = wb["5-订单总表"]

last = 2
for r in range(3, ws.max_row + 1):
    if ws.cell(r, 2).value is not None:
        last = r
print(f"5-订单总表 last data row: {last}")

for r in range(3, last + 1):
    # R(18) 产品理论成本 = 总出厂成本(单件) × 数量
    # SUMPRODUCT 按 产品编码(PPS) + SKU 匹配 2-定价总表 col T(总出厂成本)
    ws.cell(r, 18).value = (
        f"=IFERROR("
        f"SUMPRODUCT(('2-定价总表'!$A$3:$A$493=\"PPS\"&MID(I{r},2,99))*"
        f"('2-定价总表'!$C$3:$C$493=K{r})*"
        f"('2-定价总表'!$T$3:$T$493))*IF(M{r}=\"\",1,M{r}),"
        f"\"\")"
    )

print(f"Wrote 产品理论成本 formula for rows 3–{last} ({last - 2} rows)")

wb.save(SRC)
print(f"Saved → {SRC}")
shutil.copy2(SRC, DST_VALUES)
print(f"Saved values copy → {DST_VALUES}")
EOF_DONE = "Done."
print(EOF_DONE)

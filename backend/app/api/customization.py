"""尺寸微定制 API (业务需求 §2)."""
import asyncio
from decimal import Decimal
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import require_ingest_token
from app.services import customization_ai_service, customization_service

router = APIRouter(prefix="/api/customization", tags=["customization"])


class DiffLineOut(BaseModel):
    material_code: str
    material_name: Optional[str]
    original_qty: Decimal
    new_qty: Decimal
    note: Optional[str]
    requires_new_material: bool = False


class PreviewOut(BaseModel):
    base_sku_code: str
    proposed_custom_sku_code: str
    dimension_changes: dict
    diff_lines: list[DiffLineOut]
    # Plan F5: 库存预检 {in_stock, need_purchase, need_new_material, has_shortage}
    stock_check: Optional[dict] = None


class PreviewIn(BaseModel):
    base_sku_code: str = Field(..., min_length=3)
    dimension_changes: dict = Field(
        ..., description="如 {长: 2000, 宽: 400} (单位 mm), 不变的可不传"
    )


@router.post("/preview", response_model=PreviewOut)
def preview(payload: PreviewIn, db: Session = Depends(get_db)):
    try:
        r = customization_service.preview(
            db,
            base_sku_code=payload.base_sku_code,
            dimension_changes=payload.dimension_changes,
        )
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    return PreviewOut(
        base_sku_code=r.base_sku_code,
        proposed_custom_sku_code=r.proposed_custom_sku_code,
        dimension_changes=r.dimension_changes,
        diff_lines=[DiffLineOut(**d.__dict__) for d in r.diff_lines],
        stock_check=customization_service.precheck_stock(db, r.diff_lines),
    )


class ConfirmIn(BaseModel):
    base_sku_code: str
    dimension_changes: dict
    order_no: Optional[str] = None
    note: Optional[str] = None
    qty_overrides: Optional[dict[str, Decimal]] = None  # material_code → 新数量
    # Plan F5: 缺料时必须显式确认才放行 (前端分组弹窗确认后置 True)
    acknowledge_shortage: bool = False


class ConfirmOut(BaseModel):
    custom_variant_id: int
    custom_sku_code: str
    cloned_bom_lines: int


class PriceBreakdownItemOut(BaseModel):
    label: str
    amount: float
    note: str = ""


class AiQuoteOut(BaseModel):
    base_product: Optional[str]
    base_sku: Optional[str]
    base_size: Optional[str]
    changes: list[str]
    est_price: Optional[float]
    breakdown: list[PriceBreakdownItemOut]
    ai_used: bool
    model: Optional[str]
    error: Optional[str]


@router.post("/ai-quote", response_model=AiQuoteOut)
async def ai_quote(
    image: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    data = await image.read()
    mime = image.content_type or "image/jpeg"
    result = await asyncio.to_thread(customization_ai_service.ai_quote, db, data, mime)
    return AiQuoteOut(
        base_product=result.base_product,
        base_sku=result.base_sku,
        base_size=result.base_size,
        changes=result.changes,
        est_price=result.est_price,
        breakdown=[PriceBreakdownItemOut(**b.__dict__) for b in result.breakdown],
        ai_used=result.ai_used,
        model=result.model,
        error=result.error,
    )


@router.post("/confirm", response_model=ConfirmOut, status_code=201)
def confirm(payload: ConfirmIn, db: Session = Depends(get_db)):
    # Plan F5: 缺料前置检查 — 未确认缺料 → 409 返回分组明细, 前端弹窗确认后重发
    try:
        pre = customization_service.preview(
            db, base_sku_code=payload.base_sku_code,
            dimension_changes=payload.dimension_changes,
        )
        check = customization_service.precheck_stock(db, pre.diff_lines)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    if check["has_shortage"] and not payload.acknowledge_shortage:
        raise HTTPException(409, detail={
            "shortage_unacknowledged": True,
            "message": "存在缺料/需新开料, 请确认后再下定制单",
            "stock_check": check,
        })
    try:
        r = customization_service.confirm(
            db,
            base_sku_code=payload.base_sku_code,
            dimension_changes=payload.dimension_changes,
            order_no=payload.order_no,
            note=payload.note,
            qty_overrides=payload.qty_overrides,
        )
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    db.commit()
    return ConfirmOut(**r.__dict__)


# -------- 全定制报价参数 (后台可调) --------

@router.get("/quote-config", response_model=dict)
def get_quote_config(db: Session = Depends(get_db)):
    """读全定制报价参数 (利润系数/人工表/大小规则/单价/打包/投影)."""
    from app.services import custom_quote_config_service as cfg
    return cfg.get_config(db)


class QuoteConfigPatch(BaseModel):
    factory_profit_rate: Optional[float] = None
    panse_profit_rate: Optional[float] = None
    safety_rate: Optional[float] = None
    competitor_coupon_rate: Optional[float] = None
    projection_type: Optional[str] = None
    projection_rate: Optional[float] = None
    packing: Optional[list[float]] = None
    freight: Optional[list[float]] = None
    install: Optional[list[float]] = None
    labor: Optional[dict[str, list[float]]] = None
    size_rules: Optional[dict[str, list[float]]] = None
    prices: Optional[dict[str, float]] = None


@router.put("/quote-config", response_model=dict)
def update_quote_config(payload: QuoteConfigPatch, db: Session = Depends(get_db)):
    """改报价参数 (只更新传入的键), 返回完整配置."""
    from app.services import custom_quote_config_service as cfg
    patch = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    result = cfg.save_config(db, patch)
    db.commit()
    return result


# -------- 全定制: 板单 → 报价 + 工厂对比 + 投影对照 --------

class BoardIn(BaseModel):
    part: str
    material: str
    length_cm: float = 0
    width_cm: float = 0
    qty: float = 1
    unit: str = "平方米"
    is_accessory: bool = False
    is_drawer_rail: bool = False


class BoardQuoteIn(BaseModel):
    product_type: str
    length_m: float
    overall_width_m: Optional[float] = None
    overall_height_m: Optional[float] = None
    boards: list[BoardIn]
    factory_quote: Optional[float] = None   # 工厂报价(可填, 用于显示差额)


class BoardQuoteOut(BaseModel):
    wood_cost: float
    labor_fee: float
    factory_in_cost: float
    factory_profit: float
    factory_wood_total: float
    accessory_total: float
    drawer_rail_total: float
    packing_fee: float
    freight: float
    install_fee: float
    panse_cost: float
    final_quote: float
    factory_quote_compare: float
    factory_quote_conservative: float       # 工厂对比×安全系数(宁高不低)
    safety_rate: float
    projection_estimate: Optional[float]
    projection_area_m2: Optional[float]
    factory_quote: Optional[float] = None
    factory_diff: Optional[float] = None     # 工厂报价 − 我的保守对比价
    size_class: str
    wood_lines: list[dict]
    accessory_lines: list[dict]


@router.post("/board-quote", response_model=BoardQuoteOut)
def board_quote(payload: BoardQuoteIn, db: Session = Depends(get_db)):
    """按板单实时算价 (单价/人工/系数全从后台配置读)."""
    from app.services import custom_quote_service as q
    from app.services import custom_quote_config_service as cfg_svc
    specs = [q.BoardSpec(**b.model_dump()) for b in payload.boards]
    r = q.quote_from_spec(
        db, product_type=payload.product_type, length_m=payload.length_m,
        boards=specs, overall_width_m=payload.overall_width_m,
        overall_height_m=payload.overall_height_m,
    )
    cfg = cfg_svc.get_config(db)
    fq = payload.factory_quote
    return BoardQuoteOut(
        wood_cost=float(r.wood_cost), labor_fee=float(r.labor_fee),
        factory_in_cost=float(r.factory_in_cost), factory_profit=float(r.factory_profit),
        factory_wood_total=float(r.factory_wood_total), accessory_total=float(r.accessory_total),
        drawer_rail_total=float(r.drawer_rail_total), packing_fee=float(r.packing_fee),
        freight=float(r.freight), install_fee=float(r.install_fee),
        panse_cost=float(r.panse_cost), final_quote=float(r.final_quote),
        factory_quote_compare=float(r.factory_quote_compare),
        factory_quote_conservative=float(r.factory_quote_conservative),
        safety_rate=float(r.safety_rate),
        projection_estimate=float(r.projection_estimate) if r.projection_estimate is not None else None,
        projection_area_m2=float(r.projection_area_m2) if r.projection_area_m2 is not None else None,
        factory_quote=fq,
        factory_diff=(fq - float(r.factory_quote_conservative)) if fq is not None else None,
        size_class=cfg_svc.classify_size(cfg, payload.product_type, payload.length_m),
        wood_lines=r.wood_lines, accessory_lines=r.accessory_lines,
    )


@router.post("/extract-boards", response_model=dict)
async def extract_boards(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """上传设计图 → AI 拆板单 (AI 不可用时返回空, 前端转手动)."""
    raw = await file.read()
    return await asyncio.to_thread(customization_ai_service.extract_boards, db, raw, file.content_type or "image/jpeg")


# -------- 竞品 Top-10 (按匹配度) --------

class CompetitorOut(BaseModel):
    id: int
    store: Optional[str]
    category: Optional[str]
    product: Optional[str]
    link: Optional[str]
    wood: Optional[str]
    sku_name: Optional[str]
    daily_price: Optional[float]          # 我表价(叠券前)
    latest_price: Optional[float]         # 抓取/手动最新价(叠券前)
    fetch_status: Optional[str]
    latest_fetched_at: Optional[str]
    coupon_cut: float                     # 通用券减额
    after_coupon: Optional[float]         # 券后价(基于最新价, 无则用我表价)
    confidence: float


def _comp_out(db, r, conf: float) -> "CompetitorOut":
    from app.services import competitor_price_service as cps
    from app.services import custom_quote_config_service as cfg_svc
    rate = float(cfg_svc.get_config(db).get("competitor_coupon_rate", 0.08))
    base = r.latest_price if r.latest_price is not None else r.daily_price
    after, cut = cps.after_coupon(base, rate)
    return CompetitorOut(
        id=r.id, store=r.store, category=r.category, product=r.product, link=r.link,
        wood=r.wood, sku_name=r.sku_name,
        daily_price=float(r.daily_price) if r.daily_price is not None else None,
        latest_price=float(r.latest_price) if r.latest_price is not None else None,
        fetch_status=r.fetch_status,
        latest_fetched_at=r.latest_fetched_at.isoformat() if r.latest_fetched_at else None,
        coupon_cut=cut, after_coupon=after, confidence=round(conf, 2),
    )


@router.get("/competitors", response_model=list[CompetitorOut])
def competitors_top(q: str = "", limit: int = 10, db: Session = Depends(get_db)):
    """按查询词(产品名/SKU)返回竞品 Top-N, 匹配度从高到低 (中文友好相似度).

    每条含: 我表价 / 最新价(抓取或手动) / 券后价(减通用券, 披露减额)。
    """
    from sqlalchemy import select
    from app.models.competitor import CompetitorPrice
    from app.services.product_match_service import _similarity

    if not q.strip():
        return []
    rows = db.execute(select(CompetitorPrice)).scalars().all()
    scored = []
    for r in rows:
        target = " ".join(filter(None, [r.product, r.sku_name, r.wood]))
        conf = _similarity(q, target)
        if conf > 0:
            scored.append((conf, r))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [_comp_out(db, r, conf) for conf, r in scored[:limit]]


@router.post("/competitors/import")
async def import_competitors(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """竞品价库 xlsx 导入 (用户需求 2026-06-12: 页面没有导入按钮)。

    表头自动识别 (店铺/类目/产品/链接/木材/SKU/价格/最新价 及常见变体);
    去重: (店铺, SKU名) 已有则按新值更新非空字段。原文件归档。
    """
    import io

    import openpyxl

    from app.models.competitor import CompetitorPrice
    from app.services import import_storage
    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(400, "无法解析 xlsx 文件")
    ws = wb.worksheets[0]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col: dict[str, int] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        if "店" in h:
            col["store"] = i
        elif "类目" in h or "分类" in h:
            col["category"] = i
        elif "产品" in h or "商品" in h:
            col["product"] = i
        elif "链接" in h or "http" in h.lower():
            col["link"] = i
        elif "木" in h:
            col["wood"] = i
        elif "sku" in h.lower():
            col["sku_name"] = i
        elif "最新" in h:
            col["latest_price"] = i
        elif "价" in h:
            col["daily_price"] = i
    if "sku_name" not in col and "product" not in col:
        raise HTTPException(400, f"表头识别失败 (需含 SKU 或 产品 列), 实际表头: {headers[:8]}")

    def _v(r, k):
        i = col.get(k)
        if i is None or i >= len(r) or r[i] is None:
            return None
        s = str(r[i]).strip()
        return s or None

    def _money(r, k):
        from decimal import Decimal as _D
        from decimal import InvalidOperation
        s = _v(r, k)
        if s is None:
            return None
        try:
            return _D(s.replace("¥", "").replace(",", ""))
        except InvalidOperation:
            return None

    from sqlalchemy import select as _sel
    existing = {((r.store or ""), (r.sku_name or "")): r
                for r in db.execute(_sel(CompetitorPrice)).scalars().all()}
    inserted = updated = skipped = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        sku_name = _v(r, "sku_name") or _v(r, "product")
        if not sku_name:
            skipped += 1
            continue
        vals = dict(
            store=_v(r, "store"), category=_v(r, "category"),
            product=_v(r, "product"), link=_v(r, "link"), wood=_v(r, "wood"),
            sku_name=sku_name, daily_price=_money(r, "daily_price"),
            latest_price=_money(r, "latest_price"),
        )
        key = (vals["store"] or "", sku_name)
        old = existing.get(key)
        if old is not None:
            changed = False
            for k, v in vals.items():
                if v is not None and getattr(old, k) != v:
                    setattr(old, k, v)
                    changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            row = CompetitorPrice(**vals, fetch_status="manual")
            db.add(row)
            existing[key] = row
            inserted += 1
    import_storage.archive(
        db, content=data, original_name=file.filename or "竞品价库.xlsx",
        kind="generic", source="web",
        row_summary={"inserted": inserted, "updated": updated, "note": "竞品价库导入"},
    )
    db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


class CompetitorCreateIn(BaseModel):
    store: Optional[str] = None
    category: Optional[str] = None
    product: Optional[str] = None
    sku_name: Optional[str] = None
    wood: Optional[str] = None
    link: Optional[str] = None
    daily_price: Optional[float] = None      # 我表价(叠券前)
    latest_price: Optional[float] = None     # 最新价(叠券前)


@router.post("/competitors", response_model=CompetitorOut, status_code=201)
def add_competitor(payload: CompetitorCreateIn, db: Session = Depends(get_db)):
    """新增一条竞品价记录 (手动录入竞品价库)。"""
    from decimal import Decimal as _D
    from app.models.competitor import CompetitorPrice
    r = CompetitorPrice(
        store=payload.store, category=payload.category, product=payload.product,
        sku_name=payload.sku_name, wood=payload.wood, link=payload.link,
        daily_price=_D(str(payload.daily_price)) if payload.daily_price is not None else None,
        latest_price=_D(str(payload.latest_price)) if payload.latest_price is not None else None,
        fetch_status="manual",
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return _comp_out(db, r, 1.0)


@router.post("/competitors/{comp_id}/refresh", response_model=CompetitorOut)
def refresh_competitor(comp_id: int, db: Session = Depends(get_db)):
    """尽力抓取最新价 (淘宝反爬, 抓不到记 blocked, 不报错)."""
    from app.services import competitor_price_service as cps
    try:
        r = cps.refresh_one(db, comp_id)
    except ValueError as e:
        raise HTTPException(404, str(e))
    db.commit()
    return _comp_out(db, r, 1.0)


class CompetitorManualIn(BaseModel):
    latest_price: float


@router.patch("/competitors/{comp_id}", response_model=CompetitorOut)
def set_competitor_price(comp_id: int, payload: CompetitorManualIn, db: Session = Depends(get_db)):
    """手动更新竞品最新价 (抓不到时人工填)."""
    from app.services import competitor_price_service as cps
    try:
        r = cps.set_manual_price(db, comp_id, Decimal(str(payload.latest_price)))
    except ValueError as e:
        raise HTTPException(404, str(e))
    db.commit()
    return _comp_out(db, r, 1.0)


# -------- 竞品最新价批量回灌 (外部采集服务 → 本系统) --------

class CompetitorPriceItem(BaseModel):
    id: Optional[int] = None          # 我表行 id (优先)
    link: Optional[str] = None        # 或按链接精确匹配
    latest_price: float
    fetch_status: Optional[str] = None  # ok/blocked/failed; 缺省 ok
    fetched_at: Optional[str] = None    # ISO8601 抓取时间; 缺省服务器当前时间


class CompetitorBatchIn(BaseModel):
    items: list[CompetitorPriceItem] = Field(..., min_length=1)


class CompetitorBatchOut(BaseModel):
    updated: int
    not_found: list = []
    errors: list = []


@router.post("/competitors/batch-prices", response_model=CompetitorBatchOut)
def batch_competitor_prices(
    payload: CompetitorBatchIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_ingest_token),
):
    """外部采集服务一次推一批竞品最新价 (用 X-API-Key 令牌鉴权)。

    每条按 id 优先、否则按 link 精确匹配我表里的行; 逐条容错, 返回命中/未匹配/错误统计。
    """
    from app.services import competitor_price_service as cps
    r = cps.batch_update_prices(db, [i.model_dump() for i in payload.items])
    db.commit()
    return CompetitorBatchOut(**r)


@router.get("/competitors/worklist", response_model=dict)
def competitors_worklist(
    only_with_link: bool = True,
    limit: int = 1000,
    db: Session = Depends(get_db),
    _: bool = Depends(require_ingest_token),
):
    """外部采集服务拉取"待抓清单"(id + link), 抓完用 batch-prices 回推。需 X-API-Key。"""
    from sqlalchemy import select
    from app.models.competitor import CompetitorPrice
    stmt = select(CompetitorPrice)
    if only_with_link:
        stmt = stmt.where(CompetitorPrice.link.isnot(None), CompetitorPrice.link != "")
    rows = db.execute(stmt.limit(limit)).scalars().all()
    return {"items": [
        {
            "id": r.id, "link": r.link, "product": r.product, "sku_name": r.sku_name,
            "last_fetched_at": r.latest_fetched_at.isoformat() if r.latest_fetched_at else None,
            "fetch_status": r.fetch_status,
        }
        for r in rows
    ]}


# -------- 定制报价 AI 对话 (统一入口: 全定制 + 微定制 + 截图识别) -------- #

_CHAT_SONNET = "claude-sonnet-4-6"
_CHAT_OPUS = "claude-opus-4-8"

_CHAT_SYSTEM = """\
你是「畔色孚格 ERP」的定制报价顾问。用户会用文字或图片描述他们想要的产品/改造需求。

你的任务（按序）：
1. 理解需求：识别产品类型、尺寸、材质、数量、特殊要求
2. 判断类型：
   - 【全定制】：完全全新的产品规格，在现有 SKU 中找不到匹配
   - 【微定制】：已有产品的尺寸/颜色/材质微调，能找到一个「基础 SKU」作为起点
3. 给出报价方向：
   - 微定制：找出最接近的 SKU 及其日常价，说明需要调整的部分（尺寸变化/加减料）
   - 全定制：基于同品类产品价格区间，给出大致价格范围和计价逻辑
4. 列出下一步操作建议

回答格式（严格遵守）：
【需求理解】<2-3句话描述识别到的需求>
【定制类型】全定制 / 微定制
【参考价格】<价格区间或具体 SKU 日常价>
【推荐操作】<1-3个具体步骤，每步以动词开头>
【参考 SKU】<SKU编码（如果是微定制）或「暂无匹配」>

现有产品定价数据（供参考）：
{pricing_context}
"""


def _build_pricing_context(db: Session) -> str:
    from sqlalchemy import select
    from app.models.pricing import PricingSku
    from app.models.product import Product

    rows = db.execute(
        select(PricingSku.sku_code, PricingSku.sku, PricingSku.product_code,
               PricingSku.size_category, PricingSku.daily_price, PricingSku.list_price)
        .order_by(PricingSku.product_code, PricingSku.sku_code)
        .limit(200)
    ).all()

    if not rows:
        return "（暂无导入定价数据）"

    product_codes = list({r.product_code for r in rows})[:30]
    products = db.execute(
        select(Product.code, Product.name)
        .where(Product.code.in_(product_codes))
    ).all()
    name_map = {p.code: p.name for p in products}

    lines = []
    cur_prod = None
    for r in rows:
        if r.product_code != cur_prod:
            cur_prod = r.product_code
            pname = name_map.get(r.product_code, "")
            lines.append(f"\n产品 {r.product_code} {pname}:")
        price = f"¥{r.daily_price:.0f}" if r.daily_price else (f"¥{r.list_price:.0f}" if r.list_price else "—")
        lines.append(f"  {r.sku_code} [{r.sku or ''}] {r.size_category or ''} 日常价{price}")
    return "\n".join(lines)


class AiChatOut(BaseModel):
    text: str
    route_type: str  # "full_custom" | "micro_custom" | "unknown"
    suggested_sku: Optional[str] = None
    model: str
    ai_used: bool
    error: Optional[str] = None


@router.post("/ai-chat", response_model=AiChatOut)
async def ai_chat(
    message: str = Form(""),
    model_pref: str = Form("sonnet"),
    images: List[UploadFile] = File(default=[]),
    db: Session = Depends(get_db),
):
    """定制报价 AI 对话：文字 + 最多5张图 → AI 分析并给出定制类型和报价方向。"""
    from app.services import settings_service
    from app.services import ai_provider as ai_mod
    from app.services.ai_provider import AiUnavailable

    if len(images) > 5:
        raise HTTPException(400, "最多上传5张图片")

    target_model = _CHAT_OPUS if model_pref == "opus" else _CHAT_SONNET
    pricing_ctx = await asyncio.to_thread(_build_pricing_context, db)
    system = _CHAT_SYSTEM.format(pricing_context=pricing_ctx)

    cfg = settings_service.get_ai_config(db, "ocr")
    if not cfg.get("api_key"):
        return AiChatOut(
            text="AI 未配置，请在后台管理 → AI 设置 中填写 API Key。",
            route_type="unknown", model="none", ai_used=False,
            error="AI 未配置",
        )

    cfg = {**cfg, "model": target_model}
    provider = ai_mod.build_provider(cfg)

    image_data: list[tuple[bytes, str]] = []
    for img in images:
        raw = await img.read()
        image_data.append((raw, img.content_type or "image/jpeg"))

    user_msg = message.strip() or "（用户上传了图片，请分析定制需求）"

    try:
        if image_data:
            resp = await asyncio.to_thread(
                provider.chat_with_images,
                system=system, user=user_msg, images=image_data, max_tokens=1500,
            )
        else:
            resp = await asyncio.to_thread(
                provider.chat, system=system, user=user_msg, max_tokens=1500,
            )
    except AiUnavailable as e:
        return AiChatOut(
            text=str(e), route_type="unknown", model=target_model,
            ai_used=False, error=str(e),
        )

    text = resp.text
    route_type = "unknown"
    suggested_sku = None
    if "微定制" in text:
        route_type = "micro_custom"
    elif "全定制" in text:
        route_type = "full_custom"

    import re as _re
    m = _re.search(r"【参考 SKU】\s*([A-Z0-9\-]+)", text)
    if m and m.group(1) != "暂无匹配":
        suggested_sku = m.group(1)

    return AiChatOut(
        text=text, route_type=route_type, suggested_sku=suggested_sku,
        model=resp.model, ai_used=True,
    )

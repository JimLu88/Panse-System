"""竞品价库 API — 从 customization.py 拆出(2026-06-16, 纯结构拆分, 逻辑不变)。

竞品 Top-N 匹配 / xlsx 导入 / 手动增改 / 抓价刷新 / 外部采集批量回灌 + 待抓清单。
路由前缀仍是 /api/customization(与定制报价同前缀, FastAPI 自动合并), 端点路径不变。
"""
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import require_ingest_token

router = APIRouter(prefix="/api/customization", tags=["competitor"])


# -------- 竞品 Top-10 (按匹配度) --------

class CompetitorOut(BaseModel):
    id: int
    store: Optional[str]
    category: Optional[str]
    product: Optional[str]
    link: Optional[str]
    wood: Optional[str]
    sku_name: Optional[str]
    daily_price: Optional[float]          # 我表价(叠券前)
    latest_price: Optional[float]         # 抓取/手动最新价(叠券前)
    fetch_status: Optional[str]
    latest_fetched_at: Optional[str]
    coupon_cut: float                     # 通用券减额
    after_coupon: Optional[float]         # 券后价(基于最新价, 无则用我表价)
    confidence: float


def _comp_out(db, r, conf: float) -> "CompetitorOut":
    from app.services import competitor_price_service as cps
    from app.services import custom_quote_config_service as cfg_svc
    rate = float(cfg_svc.get_config(db).get("competitor_coupon_rate", 0.08))
    base = r.latest_price if r.latest_price is not None else r.daily_price
    after, cut = cps.after_coupon(base, rate)
    return CompetitorOut(
        id=r.id, store=r.store, category=r.category, product=r.product, link=r.link,
        wood=r.wood, sku_name=r.sku_name,
        daily_price=float(r.daily_price) if r.daily_price is not None else None,
        latest_price=float(r.latest_price) if r.latest_price is not None else None,
        fetch_status=r.fetch_status,
        latest_fetched_at=r.latest_fetched_at.isoformat() if r.latest_fetched_at else None,
        coupon_cut=cut, after_coupon=after, confidence=round(conf, 2),
    )


@router.get("/competitors", response_model=list[CompetitorOut])
def competitors_top(q: str = "", limit: int = 10, db: Session = Depends(get_db)):
    """按查询词(产品名/SKU)返回竞品 Top-N, 匹配度从高到低 (中文友好相似度).

    每条含: 我表价 / 最新价(抓取或手动) / 券后价(减通用券, 披露减额)。
    """
    from sqlalchemy import select
    from app.models.competitor import CompetitorPrice
    from app.services.product_match_service import _similarity

    if not q.strip():
        return []
    rows = db.execute(select(CompetitorPrice)).scalars().all()
    scored = []
    for r in rows:
        target = " ".join(filter(None, [r.product, r.sku_name, r.wood]))
        conf = _similarity(q, target)
        if conf > 0:
            scored.append((conf, r))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [_comp_out(db, r, conf) for conf, r in scored[:limit]]


@router.post("/competitors/import")
async def import_competitors(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """竞品价库 xlsx 导入 (用户需求 2026-06-12: 页面没有导入按钮)。

    表头自动识别 (店铺/类目/产品/链接/木材/SKU/价格/最新价 及常见变体);
    去重: (店铺, SKU名) 已有则按新值更新非空字段。原文件归档。
    """
    import io

    import openpyxl

    from app.models.competitor import CompetitorPrice
    from app.services import import_storage
    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(400, "无法解析 xlsx 文件")
    ws = wb.worksheets[0]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col: dict[str, int] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        if "店" in h:
            col["store"] = i
        elif "类目" in h or "分类" in h:
            col["category"] = i
        elif "产品" in h or "商品" in h:
            col["product"] = i
        elif "链接" in h or "http" in h.lower():
            col["link"] = i
        elif "木" in h:
            col["wood"] = i
        elif "sku" in h.lower():
            col["sku_name"] = i
        elif "最新" in h:
            col["latest_price"] = i
        elif "价" in h:
            col["daily_price"] = i
    if "sku_name" not in col and "product" not in col:
        raise HTTPException(400, f"表头识别失败 (需含 SKU 或 产品 列), 实际表头: {headers[:8]}")

    def _v(r, k):
        i = col.get(k)
        if i is None or i >= len(r) or r[i] is None:
            return None
        s = str(r[i]).strip()
        return s or None

    def _money(r, k):
        from decimal import Decimal as _D
        from decimal import InvalidOperation
        s = _v(r, k)
        if s is None:
            return None
        try:
            return _D(s.replace("¥", "").replace(",", ""))
        except InvalidOperation:
            return None

    from sqlalchemy import select as _sel
    existing = {((r.store or ""), (r.sku_name or "")): r
                for r in db.execute(_sel(CompetitorPrice)).scalars().all()}
    inserted = updated = skipped = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        sku_name = _v(r, "sku_name") or _v(r, "product")
        if not sku_name:
            skipped += 1
            continue
        vals = dict(
            store=_v(r, "store"), category=_v(r, "category"),
            product=_v(r, "product"), link=_v(r, "link"), wood=_v(r, "wood"),
            sku_name=sku_name, daily_price=_money(r, "daily_price"),
            latest_price=_money(r, "latest_price"),
        )
        key = (vals["store"] or "", sku_name)
        old = existing.get(key)
        if old is not None:
            changed = False
            for k, v in vals.items():
                if v is not None and getattr(old, k) != v:
                    setattr(old, k, v)
                    changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            row = CompetitorPrice(**vals, fetch_status="manual")
            db.add(row)
            existing[key] = row
            inserted += 1
    import_storage.archive(
        db, content=data, original_name=file.filename or "竞品价库.xlsx",
        kind="generic", source="web",
        row_summary={"inserted": inserted, "updated": updated, "note": "竞品价库导入"},
    )
    db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


class CompetitorCreateIn(BaseModel):
    store: Optional[str] = None
    category: Optional[str] = None
    product: Optional[str] = None
    sku_name: Optional[str] = None
    wood: Optional[str] = None
    link: Optional[str] = None
    daily_price: Optional[float] = None      # 我表价(叠券前)
    latest_price: Optional[float] = None     # 最新价(叠券前)


@router.post("/competitors", response_model=CompetitorOut, status_code=201)
def add_competitor(payload: CompetitorCreateIn, db: Session = Depends(get_db)):
    """新增一条竞品价记录 (手动录入竞品价库)。"""
    from decimal import Decimal as _D
    from app.models.competitor import CompetitorPrice
    r = CompetitorPrice(
        store=payload.store, category=payload.category, product=payload.product,
        sku_name=payload.sku_name, wood=payload.wood, link=payload.link,
        daily_price=_D(str(payload.daily_price)) if payload.daily_price is not None else None,
        latest_price=_D(str(payload.latest_price)) if payload.latest_price is not None else None,
        fetch_status="manual",
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return _comp_out(db, r, 1.0)


@router.post("/competitors/{comp_id}/refresh", response_model=CompetitorOut)
def refresh_competitor(comp_id: int, db: Session = Depends(get_db)):
    """尽力抓取最新价 (淘宝反爬, 抓不到记 blocked, 不报错)."""
    from app.services import competitor_price_service as cps
    try:
        r = cps.refresh_one(db, comp_id)
    except ValueError as e:
        raise HTTPException(404, str(e))
    db.commit()
    return _comp_out(db, r, 1.0)


class CompetitorManualIn(BaseModel):
    latest_price: float


@router.patch("/competitors/{comp_id}", response_model=CompetitorOut)
def set_competitor_price(comp_id: int, payload: CompetitorManualIn, db: Session = Depends(get_db)):
    """手动更新竞品最新价 (抓不到时人工填)."""
    from app.services import competitor_price_service as cps
    try:
        r = cps.set_manual_price(db, comp_id, Decimal(str(payload.latest_price)))
    except ValueError as e:
        raise HTTPException(404, str(e))
    db.commit()
    return _comp_out(db, r, 1.0)


# -------- 竞品最新价批量回灌 (外部采集服务 → 本系统) --------

class CompetitorPriceItem(BaseModel):
    id: Optional[int] = None          # 我表行 id (优先)
    link: Optional[str] = None        # 或按链接精确匹配
    latest_price: float
    fetch_status: Optional[str] = None  # ok/blocked/failed; 缺省 ok
    fetched_at: Optional[str] = None    # ISO8601 抓取时间; 缺省服务器当前时间


class CompetitorBatchIn(BaseModel):
    items: list[CompetitorPriceItem] = Field(..., min_length=1)


class CompetitorBatchOut(BaseModel):
    updated: int
    not_found: list = []
    errors: list = []


@router.post("/competitors/batch-prices", response_model=CompetitorBatchOut)
def batch_competitor_prices(
    payload: CompetitorBatchIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_ingest_token),
):
    """外部采集服务一次推一批竞品最新价 (用 X-API-Key 令牌鉴权)。

    每条按 id 优先、否则按 link 精确匹配我表里的行; 逐条容错, 返回命中/未匹配/错误统计。
    """
    from app.services import competitor_price_service as cps
    r = cps.batch_update_prices(db, [i.model_dump() for i in payload.items])
    db.commit()
    return CompetitorBatchOut(**r)


@router.get("/competitors/worklist", response_model=dict)
def competitors_worklist(
    only_with_link: bool = True,
    limit: int = 1000,
    db: Session = Depends(get_db),
    _: bool = Depends(require_ingest_token),
):
    """外部采集服务拉取"待抓清单"(id + link), 抓完用 batch-prices 回推。需 X-API-Key。"""
    from sqlalchemy import select
    from app.models.competitor import CompetitorPrice
    stmt = select(CompetitorPrice)
    if only_with_link:
        stmt = stmt.where(CompetitorPrice.link.isnot(None), CompetitorPrice.link != "")
    rows = db.execute(stmt.limit(limit)).scalars().all()
    return {"items": [
        {
            "id": r.id, "link": r.link, "product": r.product, "sku_name": r.sku_name,
            "last_fetched_at": r.latest_fetched_at.isoformat() if r.latest_fetched_at else None,
            "fetch_status": r.fetch_status,
        }
        for r in rows
    ]}

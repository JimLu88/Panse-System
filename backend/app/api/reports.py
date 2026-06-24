import csv
import io
from datetime import date as _date, datetime, timedelta
from decimal import Decimal
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import not_, or_, select
from sqlalchemy.orm import Session

from sqlalchemy import func

from app.database import get_db
from app.dependencies import require_role
from app.models.auth import User
from app.models.finance import AlipayFlow
from app.models.knowledge import AiKnowledge
from app.models.marketing import AfterSales, OutsourcingExpense, PromotionFlow
from app.models.order import FactoryOrder, Order
from app.services import asset_service, dashboard_monthly_service, data_freshness_service, health_report, sales_analytics, sales_rollup_service, shop_report_service

# 路由级守卫: 报表含财务数据, 统一要求登录 (此前多数端点无守卫, 外网下可被未登录读取)。
router = APIRouter(
    prefix="/api/reports",
    tags=["reports"],
    dependencies=[Depends(require_role("admin", "operator", "viewer"))],
)


def _range_for(period: str) -> tuple[_date, _date]:
    today = _date.today()
    if period == "7d":
        return today - timedelta(days=7), today
    if period == "30d":
        return today - timedelta(days=30), today
    if period == "month":
        return today.replace(day=1), today
    if period == "year":
        return today.replace(month=1, day=1), today
    if period == "last_month":  # 上个月整月
        prev_last = today.replace(day=1) - timedelta(days=1)
        return prev_last.replace(day=1), prev_last
    # 具体月份 YYYY-MM (用户拍板 2026-06-17: 时间筛选支持按月下拉)
    import re
    from calendar import monthrange
    mm = re.fullmatch(r"(\d{4})-(\d{1,2})", period or "")
    if mm:
        y, mo = int(mm.group(1)), int(mm.group(2))
        if 1 <= mo <= 12:
            return _date(y, mo, 1), _date(y, mo, monthrange(y, mo)[1])
    raise HTTPException(400, f"未知 period: {period} (允许 7d/30d/month/year/last_month/YYYY-MM)")


class ShopStatOut(BaseModel):
    shop: str
    order_count: int
    total_qty: int
    total_revenue: float


@router.get("/shops", response_model=list[ShopStatOut])
def shop_stats(
    period: Optional[str] = Query(None, description="7d/30d/month/year; 不填=全部"),
    db: Session = Depends(get_db),
):
    """分店统计: 各店铺 单数/销量/销售额 (按销售额降序, 未归属单独成桶)。"""
    start = end = None
    if period:
        start, end = _range_for(period)
    return shop_report_service.compute_shop_stats(db, start=start, end=end)


class HealthReportOut(BaseModel):
    period_start: str
    period_end: str
    exceptions: dict[str, Any]
    reconciliation: dict[str, Any]
    inventory: dict[str, Any]
    orders: dict[str, Any]
    roi: dict[str, Any]
    integrity_score: int
    headlines: list[str]


@router.get("/monthly", response_model=HealthReportOut)
def monthly_health(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
):
    try:
        r = health_report.generate(db, year, month)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    return HealthReportOut(**health_report.to_dict(r))


@router.get("/monthly/current", response_model=HealthReportOut)
def current_month(db: Session = Depends(get_db)):
    now = datetime.now()
    return monthly_health(year=now.year, month=now.month, db=db)


@router.get("/monthly-financial")
def monthly_financial(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    fmt: str = Query("json", pattern="^(json|xlsx)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin", "operator", "viewer")),
):
    """月度财务报表 (优化 #10): 营收/成本/毛利/净利/订单数; fmt=xlsx 下载 Excel。"""
    from app.services import financial_report_service
    s = financial_report_service.monthly_summary(db, year, month)
    if fmt == "xlsx":
        data = financial_report_service.build_excel(s)
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="financial-{s["period"]}.xlsx"'},
        )
    return s


class KnowledgeOut(BaseModel):
    id: int
    exception_type: str
    context_hash: str
    solution_text: str
    source_description: str | None
    model: str | None
    usage_count: int
    last_used_at: str | None
    created_at: str


# ----------------------------- Phase 4: 销售报表 ---------------------- #


class SalesSummaryOut(BaseModel):
    period_start: str
    period_end: str
    order_count: int
    revenue: float
    cost: float
    gross_profit: float
    net_profit: float
    top_products_by_profit: list[dict]
    top_products_by_profit_rate: list[dict]
    bottom_products_by_profit: list[dict] = []


def _dec(d) -> float:
    return float(Decimal(d or 0)) if d is not None else 0.0


@router.get("/sales/summary", response_model=SalesSummaryOut)
def sales_summary(
    period: str = Query("30d", description="7d / 30d / month / year"),
    platform: Optional[str] = None,
    brand: Optional[str] = Query(None, description="PS / PFG 品牌过滤 (Plan F8)"),
    db: Session = Depends(get_db),
):
    """业务需求 15: 店铺销售汇总 (销售额/成本/毛利/净利 + 利润排行 Top 10)."""
    start, end = _range_for(period)
    s = sales_analytics.summary(db, start=start, end=end, platform=platform, brand=brand)
    def _ser(rows):
        return [
            {k: (_dec(v) if isinstance(v, Decimal) else v) for k, v in r.items()}
            for r in rows
        ]
    return SalesSummaryOut(
        period_start=s.period_start.isoformat(), period_end=s.period_end.isoformat(),
        order_count=s.order_count,
        revenue=_dec(s.revenue), cost=_dec(s.cost),
        gross_profit=_dec(s.gross_profit), net_profit=_dec(s.net_profit),
        top_products_by_profit=_ser(s.top_products_by_profit),
        top_products_by_profit_rate=_ser(s.top_products_by_profit_rate),
        bottom_products_by_profit=_ser(s.bottom_products_by_profit),
    )


@router.get("/sales/breakdown")
def sales_breakdown(
    period: str = Query("30d"),
    brand: Optional[str] = Query(None, description="PS / PFG 品牌过滤 (Plan F8)"),
    db: Session = Depends(get_db),
):
    """业务需求 16: 分产品 SKU 销售明细."""
    start, end = _range_for(period)
    rows = sales_analytics.product_breakdown(db, start=start, end=end, brand=brand)
    out = []
    for r in rows:
        d = {}
        for k, v in r.items():
            d[k] = _dec(v) if isinstance(v, Decimal) else v
        out.append(d)
    return {"period_start": start.isoformat(), "period_end": end.isoformat(), "rows": out}


@router.get("/sales/cost-anomaly")
def sales_cost_anomaly(
    period: str = Query("year"),
    product_code: Optional[str] = Query(None, description="只看某产品 (如 PPS24210070901)"),
    db: Session = Depends(get_db),
):
    """诊断销售汇总成本异常: 找『实付小却背全额成本』的错配单(成本>实付), 量化它们把总成本拉高多少。
    用于解释为何某产品利润率偏低 / 总利润为负。"""
    from app.models.order import Order as _O
    from app.services import sales_analytics, order_financials as ofin
    start, end = _range_for(period)
    q = select(_O).where(_O.order_date >= start, _O.order_date <= end,
                         sales_analytics.settled_sale_clause(),  # 统一成交口径
                         _O.is_refill == False)  # noqa: E712
    if product_code:
        q = q.where(_O.product_code == product_code)
    orders = db.execute(q).scalars().all()
    tot_rev = tot_cost = Decimal("0")
    mism = []
    for o in orders:
        rev = Decimal(o.paid_amount or 0) - Decimal(o.refund_amount or 0)
        cost = ofin.physical_cost(o)   # 统一物理成本(含片段85%兜底)
        tot_rev += rev
        tot_cost += cost
        if cost > rev:
            mism.append({
                "order_no": o.order_no, "product_code": o.product_code,
                "paid": float(rev), "cost": float(cost), "excess": float(cost - rev),
                "cost_src": "actual" if o.actual_cost is not None else "theoretical",
                "status": o.status,
            })
    mism.sort(key=lambda r: r["excess"], reverse=True)
    return {
        "period_start": start.isoformat(), "period_end": end.isoformat(),
        "orders": len(orders), "total_revenue": float(tot_rev), "total_cost": float(tot_cost),
        "cost_minus_revenue": float(tot_cost - tot_rev),
        "mismatched_count": len(mism),
        "mismatched_excess_total": round(sum(m["excess"] for m in mism), 2),
        "samples": mism[:40],
    }


@router.get("/sales/ranking")
def sales_ranking(
    granularity: str = Query("month", description="month(按月) / year(按年)"),
    metric: str = Query("revenue", description="revenue(销售额) / qty(销量) / profit(利润率)"),
    period: Optional[str] = Query(None, description="指定周期 2026-04 / 2026; 缺省取最新"),
    limit: int = Query(30, le=100),
    db: Session = Depends(get_db),
):
    """销售排行榜: 按月/年 分产品 销量/销售额/利润率 排行 + 每期冠军时间线 (正式销售, 不含补单)."""
    return sales_analytics.product_ranking(
        db, granularity=granularity, metric=metric, period=period, limit=limit,
    )


@router.get("/monthly-pnl")
def monthly_pnl(db: Session = Depends(get_db)):
    """数据大盘·月度经营: 工厂口径利润 + 对账完成度(accurate/reference_only) + 推广ROI + 累计投资回收率, 逐月。"""
    return dashboard_monthly_service.monthly_pnl(db)


@router.get("/sales-mix")
def sales_mix(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    by: str = Query("product", description="product(产品) / shop(店铺)"),
    db: Session = Depends(get_db),
):
    """某月销售占比(饼图数据): 正式销售按 产品/店铺 维度, 剔除补差价邮费专链。"""
    return dashboard_monthly_service.sales_mix(db, year=year, month=month, by=by)


# ----------------------------- Phase 12: CSV 导出 -------------------- #


@router.get("/sales/breakdown.csv")
def export_sales_breakdown(
    period: str = Query("30d"),
    db: Session = Depends(get_db),
):
    """业务需求扩展: 分产品销售导出 CSV. 老板转发到企业微信群用."""
    start, end = _range_for(period)
    rows = sales_analytics.product_breakdown(db, start=start, end=end)
    buf = io.StringIO()
    buf.write("﻿")   # BOM, Excel 中文不乱码
    writer = csv.writer(buf)
    writer.writerow([
        "产品编码", "产品名", "SKU 编码", "SKU 名", "件数",
        "销售额", "成本", "毛利", "净利",
        "毛利率 %", "净利率 %",
    ])
    for r in rows:
        revenue = float(r.get("revenue") or 0)
        cost = float(r.get("cost") or 0)
        net = float(r.get("net_profit") or 0)
        writer.writerow([
            r.get("product_code") or "", r.get("product_name") or "",
            r.get("sku_code") or "", r.get("sku") or "",
            r.get("qty") or 0,
            f"{revenue:.2f}", f"{cost:.2f}",
            f"{revenue - cost:.2f}", f"{net:.2f}",
            f"{float(r.get('gross_profit_rate') or 0) * 100:.1f}",
            f"{float(r.get('net_profit_rate') or 0) * 100:.1f}",
        ])
    csv_data = buf.getvalue()
    fname = f"sales_{start.isoformat()}_{end.isoformat()}.csv"
    return StreamingResponse(
        iter([csv_data.encode("utf-8")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/sales/summary.csv")
def export_sales_summary(
    period: str = Query("30d"),
    db: Session = Depends(get_db),
):
    """导出汇总 + Top 10 利润排行."""
    start, end = _range_for(period)
    s = sales_analytics.summary(db, start=start, end=end)
    buf = io.StringIO()
    buf.write("﻿")
    writer = csv.writer(buf)
    writer.writerow(["项目", "数值"])
    writer.writerow(["周期", f"{start.isoformat()} ~ {end.isoformat()}"])
    writer.writerow(["订单数", s.order_count])
    writer.writerow(["销售额", float(s.revenue)])
    writer.writerow(["成本", float(s.cost)])
    writer.writerow(["毛利", float(s.gross_profit)])
    writer.writerow(["净利", float(s.net_profit)])
    writer.writerow([])
    writer.writerow(["Top 10 利润排行"])
    writer.writerow(["产品编码", "产品名", "订单数", "净利"])
    for r in s.top_products_by_profit:
        writer.writerow([
            r.get("product_code") or "", r.get("product_name") or "",
            r.get("order_count") or 0,
            f"{float(r.get('net_profit') or 0):.2f}",
        ])
    fname = f"summary_{start.isoformat()}_{end.isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue().encode("utf-8")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/forecast/30d")
def forecast_30d(db: Session = Depends(get_db)):
    """业务需求 7: 未来 30 天 SKU 销量预测 (移动平均 + 20% 安全系数)."""
    return {"forecast": sales_analytics.forecast_30d(db)}


@router.get("/stock-advice")
def stock_advice(db: Session = Depends(get_db)):
    """业务需求 8: 智能提前备货建议."""
    return sales_analytics.stock_advice(db)


@router.get("/slow-moving")
def slow_moving(
    long_no_sale_days: int = 60,
    overstock_ratio: float = 3.0,
    db: Session = Depends(get_db),
):
    """业务需求 8: 滞销分类 (长期未售 / 超大库存)."""
    return sales_analytics.slow_moving_split(
        db, long_no_sale_days=long_no_sale_days, overstock_ratio=overstock_ratio,
    )


# ----------------------------- 经营状况分析 (功能 5: 收支占比) -------- #


@router.get("/operating-analysis")
def operating_analysis(
    period: str = Query("30d", description="7d / 30d / month / year"),
    platform: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """业务需求扩展 (功能 5): 7天/月度/年度经营状况分析 — 各项收支占比.

    收入侧: 销售额。支出侧: 成本 / 运费等订单费用 / 推广 / 人员外包。
    每项给出金额与占销售额的百分比, 末了给净利与净利率。
    """
    start, end = _range_for(period)
    # 统一会计 P&L (用户拍板 2026-06-18: 与月度经营/逐单核对/大盘 完全同口径)。platform 过滤已并入整体口径。
    from app.services import order_financials as _ofin
    s = _ofin.accounting_summary(db, start, end)
    revenue = s["revenue"]

    def _pct(x) -> float:
        return float(Decimal(x or 0) / revenue * 100) if revenue else 0.0

    items = [
        {"name": "物理产品成本", "amount": _dec(s["goods"]), "pct": _pct(s["goods"])},
        {"name": "物流费", "amount": _dec(s["freight"]), "pct": _pct(s["freight"])},
        {"name": "安装/上楼", "amount": _dec(s["install"]), "pct": _pct(s["install"])},
        {"name": "平台扣点(实付−实收)", "amount": _dec(s["platform"]), "pct": _pct(s["platform"])},
        {"name": "税费", "amount": _dec(s["tax"]), "pct": _pct(s["tax"])},
        {"name": "额外售后(退款外)", "amount": _dec(s["aftersales"]), "pct": _pct(s["aftersales"])},
        {"name": "推广费", "amount": _dec(s["promo"]), "pct": _pct(s["promo"])},
        {"name": "人员外包", "amount": _dec(s["outsourcing"]), "pct": _pct(s["outsourcing"])},
        {"name": "固定成本(房租等)", "amount": _dec(s["fixed"]), "pct": _pct(s["fixed"])},
        {"name": "补单(刷单)成本", "amount": _dec(s["refill"]["total"]), "pct": _pct(s["refill"]["total"])},
    ]
    net = s["net"]
    return {
        "period": period,
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "revenue": _dec(revenue),
        "expense_items": items,
        "total_expense": _dec(s["total_cost"]),
        "net_profit": _dec(net),
        "net_profit_rate": float(net / revenue * 100) if revenue else 0.0,
    }


# ----------------------------- 销售汇总 (rollup 加速) ----------------- #


@router.get("/sales/rollup-summary")
def sales_rollup_summary(
    period: str = Query("30d"),
    platform: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """从 sales_daily_rollup 预聚合表快速取总览 (大数据量时比实时 SUM 快)。

    rollup 未覆盖该区间时返回 {} (前端回退到 /sales/summary 实时算)。
    """
    start, end = _range_for(period)
    return sales_rollup_service.query_summary(db, start=start, end=end, platform=platform)


@router.post("/sales/rollup-backfill")
def sales_rollup_backfill(
    period: str = Query("30d"),
    db: Session = Depends(get_db),
):
    """补算一段时间的销售日汇总 (历史回填 / rollup 缺口修补)。"""
    start, end = _range_for(period)
    r = sales_rollup_service.rollup_range(db, start, end)
    db.commit()
    return r


# ----------------------------- Phase 4: 资产 / 经营分析 -------------- #


class AssetCategoryOut(BaseModel):
    name: str
    amount: float
    detail: list[dict] = []


class AssetSummaryOut(BaseModel):
    total: float
    categories: list[AssetCategoryOut]
    formula_a: float
    formula_b: float
    diff: float
    breakdown: dict[str, float] = {}


@router.get("/assets", response_model=AssetSummaryOut)
def assets_summary(db: Session = Depends(get_db)):
    """业务需求 14: 资产总额 + 饼图分类; 业务需求 19: 公式 A/B 对比 + 逐项拆解."""
    s = asset_service.summary(db)
    return AssetSummaryOut(
        total=_dec(s.total),
        categories=[
            AssetCategoryOut(name=c.name, amount=_dec(c.amount), detail=c.detail)
            for c in s.categories
        ],
        formula_a=_dec(s.formula_a),
        formula_b=_dec(s.formula_b),
        diff=_dec(s.diff),
        breakdown=s.breakdown,
    )


@router.get("/data-freshness")
def data_freshness(db: Session = Depends(get_db)):
    """各数据源新鲜度状态: 距今天数 / 是否过期 / 提醒文字。前端可用于显示"待补录"小红点。"""
    items = data_freshness_service.check_all(db)
    return [
        {
            "source": i.source,
            "last_date": i.last_date.isoformat() if i.last_date else None,
            "days_stale": i.days_stale,
            "threshold_days": i.threshold_days,
            "overdue": i.overdue,
            "message": i.message,
        }
        for i in items
    ]


@router.post("/data-freshness/remind-now")
def remind_now(db: Session = Depends(get_db)):
    """手动触发一次新鲜度检查 + 推送 (无需等调度器)。"""
    return data_freshness_service.check_and_remind(db)


@router.get("/unmatched-flows")
def unmatched_flows(days: int = 7, db: Session = Depends(get_db)):
    """业务需求 19: 未核销异常池 — 最近 7 天 open 状态流水."""
    return {"days": days, "rows": asset_service.unmatched_recent_flows(db, days=days)}


@router.get("/assets/diff-drilldown")
def assets_diff_drilldown(db: Session = Depends(get_db)):
    """Plan L6: 两口径资金差额下钻 — 每个科目的构成明细 TopN."""
    return asset_service.diff_drilldown(db)


@router.get("/sales-by-brand")
def sales_by_brand(
    period: str = Query("30d", description="7d / 30d / month / year"),
    db: Session = Depends(get_db),
):
    """Plan F8: 按品牌 (PS 畔色 / PFG 孚格) 汇总销售额/净利/单量."""
    start, end = _range_for(period)
    out = []
    for b in ("PS", "PFG"):
        s = sales_analytics.summary(db, start=start, end=end, brand=b)
        out.append({
            "brand": b, "order_count": s.order_count,
            "revenue": _dec(s.revenue), "net_profit": _dec(s.net_profit),
        })
    return {"period": period, "brands": out}


@router.get("/unmatched-flows/classify")
def classify_unmatched(days: int = 7, limit: int = 50,
                       db: Session = Depends(get_db)):
    """Phase 11 P4-24: AI 辅助归类未核销流水."""
    from app.services import flow_classification_service
    return {"results": flow_classification_service.batch_classify(
        db, days=days, limit=limit,
    )}


@router.get("/knowledge", response_model=list[KnowledgeOut])
def list_knowledge(limit: int = Query(50, le=200), db: Session = Depends(get_db)):
    """AI 知识库内容 (plan §12.2 常见问题库)."""
    rows = db.execute(
        select(AiKnowledge).order_by(AiKnowledge.usage_count.desc()).limit(limit)
    ).scalars().all()
    return [
        KnowledgeOut(
            id=r.id, exception_type=r.exception_type, context_hash=r.context_hash,
            solution_text=r.solution_text, source_description=r.source_description,
            model=r.model, usage_count=r.usage_count,
            last_used_at=r.last_used_at.isoformat() if r.last_used_at else None,
            created_at=r.created_at.isoformat(),
        )
        for r in rows
    ]


# ─── 月度经营数据表格 ─────────────────────────────────────────────────────────


def _month_range(year: int, month: int) -> tuple[_date, _date]:
    start = _date(year, month, 1)
    if month == 12:
        end = _date(year, 12, 31)
    else:
        end = _date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def _business_month(db: Session, year: int, month: int) -> dict:
    """单月经营数据汇总。"""
    start, end = _month_range(year, month)

    def _qs(col, *wheres):
        stmt = select(func.coalesce(func.sum(col), 0))
        for w in wheres:
            stmt = stmt.where(w)
        return float(db.execute(stmt).scalar() or 0)

    def _qc(*wheres):
        stmt = select(func.count(Order.id))
        for w in wheres:
            stmt = stmt.where(w)
        return int(db.execute(stmt).scalar() or 0)

    # 不再按 is_historical 过滤: 日常导入的订单大多被标了 is_historical=True, 若过滤会把绝大多数
    # 真实订单漏掉(出现"某月 0 单"的怪数)。报表本就从 2026 起, 直接按下单日统计。
    # 真实订单口径(用户拍板 2026-06-17, 全系统统一): 只算"已付款且成交"——排除 待付款/取消/关闭
    # 与 全额退款单(原来只排取消, 把 90 个待付款也算进 1 月营收 → 净利率虚高到 73%)。
    from app.services.sales_analytics import settled_sale_clause
    _settled = settled_sale_clause()
    base = [Order.order_date >= start, Order.order_date <= end]
    real = [*base, _settled, Order.is_refill == False]  # noqa: E712

    from app.services import order_financials as _ofin
    # 统一会计 P&L (用户拍板 2026-06-18: 全系统同口径; 月度/经营状况/逐单/大盘 都走 accounting_summary)
    s = _ofin.accounting_summary(db, start, end)
    real_count = _qc(*real)
    real_revenue = round(float(s["revenue"]), 2)
    refill_count = _qc(*base, _settled, Order.is_refill == True)  # noqa: E712  # 排关闭/取消的补单
    refill_revenue = _qs(Order.paid_amount, *base, _settled, Order.is_refill == True)  # noqa: E712

    factory_bill = float(db.execute(
        select(func.coalesce(func.sum(FactoryOrder.factory_bill_amount), 0)).where(
            FactoryOrder.order_date >= start,
            FactoryOrder.order_date <= end,
            FactoryOrder.voided_at.is_(None),
        )
    ).scalar() or 0)

    cogs = float(s["goods"]); cogs_estimated = bool(s["goods_estimated"])
    freight = float(s["freight"]); install_upstairs = float(s["install"])
    platform_ded = float(s["platform"])          # 平台扣点(实付−实收, 含平台优惠券)
    tax_expense = float(s["tax"])
    aftersales_comp = float(s["aftersales"]); aftersales_count = int(s["aftersales_count"])
    promo = float(s["promo"])
    outsourcing = float(s["outsourcing"]); outsourcing_estimated = bool(s["outsourcing_estimated"])
    fixed_costs = float(s["fixed"])
    refill_cost = float(s["refill"]["total"])
    total_expense = float(s["total_cost"])
    net_profit = float(s["net"])

    total_revenue = real_revenue + refill_revenue   # 总流水 (含补单, 仅参考)
    total_orders = real_count + refill_count
    refill_ratio = round(refill_count / total_orders * 100, 1) if total_orders else 0.0
    refill_cost_ratio = round(refill_revenue / total_revenue * 100, 1) if total_revenue else 0.0
    promo_ratio = round(promo / real_revenue * 100, 1) if real_revenue else 0.0
    aftersales_rate = round(aftersales_count / real_count * 100, 1) if real_count else 0.0
    lead_time_days = None
    lt_rows = db.execute(
        select(FactoryOrder.order_date, FactoryOrder.actual_delivery).where(
            FactoryOrder.order_date >= start,
            FactoryOrder.order_date <= end,
            FactoryOrder.actual_delivery.isnot(None),
            FactoryOrder.voided_at.is_(None),
        )
    ).all()
    if lt_rows:
        deltas = [(r.actual_delivery - r.order_date).days for r in lt_rows if r.actual_delivery >= r.order_date]
        if deltas:
            lead_time_days = round(sum(deltas) / len(deltas), 1)

    return {
        "year": year,
        "month": month,
        "period": f"{year}-{month:02d}",
        "real_order_count": real_count,
        "refill_order_count": refill_count,
        "real_revenue": round(real_revenue, 2),
        "refill_revenue": round(refill_revenue, 2),
        "total_revenue": round(total_revenue, 2),
        "refill_order_ratio": refill_ratio,
        "refill_cost_ratio": refill_cost_ratio,
        "promo_expense": round(promo, 2),
        "promo_ratio": promo_ratio,
        "factory_bill": round(factory_bill, 2),
        "effective_cost": round(cogs, 2),                        # 商品成本 (Σ逐单 actual/theoretical)
        "cogs_estimated": cogs_estimated,                        # 含推演(工厂未对账)
        "freight_expense": round(freight, 2),                    # 物流费
        "install_upstairs_expense": round(install_upstairs, 2),  # 安装上楼
        "platform_deduction": round(platform_ded, 2),           # 平台扣点 (实付−实收, 含优惠券)
        "tax_expense": round(tax_expense, 2),                    # 税费
        "aftersales_compensation": round(aftersales_comp, 2),    # 额外售后 (按订单归属, 退款不重复计)
        "aftersales_count": aftersales_count,
        "aftersales_rate": aftersales_rate,
        "outsourcing_expense": round(outsourcing, 2),
        "outsourcing_estimated": outsourcing_estimated,          # 5月起按¥10000/月预估
        "fixed_costs": round(fixed_costs, 2),                    # 固定成本/管理费用 (房租等)
        "refill_cost": round(refill_cost, 2),                    # 补单(刷单)纯成本
        "total_expense": round(total_expense, 2),
        "net_profit": round(net_profit, 2),
        "net_profit_rate": round(net_profit / real_revenue * 100, 1) if real_revenue else 0.0,
        "avg_lead_time_days": lead_time_days,
    }


@router.get("/business-monthly")
def business_monthly_table(
    from_year: int = Query(2026, description="起始年份"),
    from_month: int = Query(1, ge=1, le=12, description="起始月份"),
    to_year: Optional[int] = Query(None),
    to_month: Optional[int] = Query(None, ge=1, le=12),
    db: Session = Depends(get_db),
):
    """月度经营数据表格 — 从指定月份至今, 每月一行。

    返回字段:
      period, real_order_count, refill_order_count, real_revenue, refill_revenue,
      total_revenue, refill_order_ratio (%), refill_cost_ratio (%),
      promo_expense, promo_ratio (%), factory_bill,
      aftersales_compensation, aftersales_count, aftersales_rate (%),
      outsourcing_expense, platform_fee,
      total_expense, net_profit, net_profit_rate (%),
      avg_lead_time_days
    """
    today = _date.today()
    end_year = to_year or today.year
    end_month = to_month or today.month

    months = []
    y, m = from_year, from_month
    while (y, m) <= (end_year, end_month):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1

    rows = [_business_month(db, y, m) for y, m in months]
    rows.reverse()  # 最新月在前

    # 汇总行
    def _sum(key):
        return round(sum(r[key] for r in rows if isinstance(r[key], (int, float))), 2)

    total_real = sum(r["real_order_count"] for r in rows)
    total_refill = sum(r["refill_order_count"] for r in rows)
    total_rev = _sum("total_revenue")
    total_refill_rev = _sum("refill_revenue")
    total_promo = _sum("promo_expense")
    total_exp = _sum("total_expense")
    total_net = _sum("net_profit")
    total_real_rev = _sum("real_revenue")

    def _any(key):
        return any(bool(r.get(key)) for r in rows)

    summary = {
        "period": "合计",
        "real_order_count": total_real,
        "refill_order_count": total_refill,
        "real_revenue": total_real_rev,
        "refill_revenue": total_refill_rev,
        "total_revenue": total_rev,
        "refill_order_ratio": round(total_refill / (total_real + total_refill) * 100, 1) if (total_real + total_refill) else 0.0,
        "refill_cost_ratio": round(total_refill_rev / total_rev * 100, 1) if total_rev else 0.0,
        "promo_expense": total_promo,
        "promo_ratio": round(total_promo / total_real_rev * 100, 1) if total_real_rev else 0.0,
        "factory_bill": _sum("factory_bill"),
        "effective_cost": _sum("effective_cost"),
        "cogs_estimated": _any("cogs_estimated"),
        "freight_expense": _sum("freight_expense"),
        "install_upstairs_expense": _sum("install_upstairs_expense"),
        "platform_deduction": _sum("platform_deduction"),
        "tax_expense": _sum("tax_expense"),
        "aftersales_compensation": _sum("aftersales_compensation"),
        "aftersales_count": sum(r["aftersales_count"] for r in rows),
        "aftersales_rate": round(sum(r["aftersales_count"] for r in rows) / total_real * 100, 1) if total_real else 0.0,
        "outsourcing_expense": _sum("outsourcing_expense"),
        "outsourcing_estimated": _any("outsourcing_estimated"),
        "fixed_costs": _sum("fixed_costs"),
        "refill_cost": _sum("refill_cost"),
        "total_expense": total_exp,
        "net_profit": total_net,
        "net_profit_rate": round(total_net / total_real_rev * 100, 1) if total_real_rev else 0.0,
        "avg_lead_time_days": None,
    }

    return {"rows": rows, "summary": summary}


@router.get("/per-order-reconcile")
def per_order_reconcile(
    year: int = Query(2026),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
):
    """逐单核对 (财务) — 某月每笔真实成交订单的完整成本拆解 + 支付宝覆盖/对账状态 + 问题单高亮。
    口径与经营状况一致 (order_financials 会计成本); 合计再减该月推广费、人员外包 = 本月真实净利。
    """
    start, end = _month_range(year, month)
    from app.services import order_financials as _ofin, sales_analytics
    coef = _ofin.load_coefficients(db)
    as_by_order = _ofin.extra_aftersales_by_order(db)

    orders = db.execute(
        select(Order).where(
            Order.order_date >= start, Order.order_date <= end,
            sales_analytics.settled_sale_clause(), Order.is_refill == False,  # noqa: E712
        )
    ).scalars().all()

    # 工厂成本核对 (用户 2026-06-20): 每单 预算(定价表 木作/配件/打包) vs 实际(工厂账单 actual_cost, 仅木作) + 差额。
    from app.models.pricing import PricingSku
    _codes = {o.sku_code for o in orders if o.sku_code}
    _ps_by_code = {ps.sku_code: ps for ps in db.execute(
        select(PricingSku).where(PricingSku.sku_code.in_(_codes))).scalars()} if _codes else {}
    _skus = {o.sku for o in orders if not o.sku_code and o.sku}
    _ps_by_sku = {ps.sku: ps for ps in db.execute(
        select(PricingSku).where(PricingSku.sku.in_(_skus))).scalars()} if _skus else {}

    rows = []
    SUM_KEYS = ("paid_amount", "refund_amount", "revenue", "cost_goods", "cost_freight",
                "cost_install", "cost_platform", "cost_tax", "cost_aftersales", "cost_total", "net_profit")
    sums = {k: 0.0 for k in SUM_KEYS}
    # 工厂成本核对合计 (None 当 0)。wood_diff 合计=有账单单的(实际−预算)之和。
    BSUM_KEYS = ("predicted_wood", "est_parts", "est_packaging", "actual_wood", "wood_diff")
    bsum = {k: 0.0 for k in BSUM_KEYS}
    for o in orders:
        paid = float(o.paid_amount or 0)
        refund = float(o.refund_amount or 0)
        revenue = paid - refund
        b = _ofin.cost_breakdown(o, coef, Decimal("0"))   # 售后改逐单口径, 不用人均均摊
        pb = _ofin.physical_cost_breakdown(o)             # 物理成本加法拆解 (导出逐项公式回推用)
        goods = float(b["physical"]); freight = float(b["freight"]); install = float(b["install_upstairs"])
        platform = float(b["platform"]); tax = float(b["tax"])
        aftersales = float(as_by_order.get(o.order_no, 0))
        cost_total = goods + freight + install + platform + tax + aftersales
        net = revenue - cost_total
        cost_estimated = o.actual_cost is None   # 用推演成本(未对账), 非工厂实报
        # 工厂成本核对: 预算(定价表) vs 实际(工厂账单 actual_cost, 仅木作) + 木作差额
        _ps = _ps_by_code.get(o.sku_code) or (_ps_by_sku.get(o.sku) if o.sku else None)
        _qty = int(o.qty or 1)
        pred_wood = (round(float(o.wood_cost_est), 2) if o.wood_cost_est
                     else (round(float(_ps.wood_cost) * _qty, 2) if (_ps and _ps.wood_cost) else None))
        est_parts = round(float(_ps.external_parts_cost) * _qty, 2) if (_ps and _ps.external_parts_cost) else None
        est_packaging = round(float(_ps.packaging_cost) * _qty, 2) if (_ps and _ps.packaging_cost) else None
        actual_wood = round(float(o.actual_cost), 2) if o.actual_cost is not None else None
        wood_diff = (round(actual_wood - pred_wood, 2)
                     if (actual_wood is not None and pred_wood is not None) else None)
        row = {
            "order_no": o.order_no,
            "product_name": o.product_name or o.product_code or "",
            "is_custom": bool(o.is_custom),
            "order_date": o.order_date.isoformat() if o.order_date else None,
            "paid_amount": round(paid, 2), "refund_amount": round(refund, 2), "revenue": round(revenue, 2),
            "cost_goods": round(goods, 2), "cost_freight": round(freight, 2), "cost_install": round(install, 2),
            # 物理成本(商品成本)加法拆解 — 导出逐项公式回推: 商品成本 = 工厂木作 + 定价表估算 + 打包 (或 实付×85%)
            "cost_factory_wood": round(float(pb["factory_wood"]), 2),
            "cost_estimate_part": round(float(pb["estimate_part"]), 2),
            "cost_packing": round(float(pb["packing"]), 2),
            "cost_cap_mode": pb["cap_mode"],
            "cost_platform": round(platform, 2), "cost_tax": round(tax, 2), "cost_aftersales": round(aftersales, 2),
            "cost_total": round(cost_total, 2), "net_profit": round(net, 2),
            "net_margin": round(net / revenue * 100, 1) if revenue else 0.0,
            "alipay_covered": o.alipay_flow_no is not None,    # 已配上支付宝到账流水
            "cost_reconciled": o.actual_cost is not None,      # 工厂成本已对账(非推演)
            "cost_estimated": cost_estimated,
            "is_loss": net < 0,
            # 工厂成本核对列 (预算 vs 实际): 工厂账单只含木作, 配件/打包恒为预估
            "factory_bill_recorded": o.actual_cost is not None,   # 工厂账单已入账
            "predicted_wood": pred_wood,                          # 预算木作(定价表/wood_cost_est)
            "est_parts": est_parts,                               # 预估配件(外采)
            "est_packaging": est_packaging,                       # 预估打包
            "actual_wood": actual_wood,                           # 实际木作(工厂账单 actual_cost)
            "wood_diff": wood_diff,                               # 木作差额(实际−预算; +超支/−省)
        }
        rows.append(row)
        for k in SUM_KEYS:
            sums[k] += row[k]
        for k in BSUM_KEYS:
            bsum[k] += row[k] or 0

    promo = float(db.execute(
        select(func.coalesce(func.sum(PromotionFlow.amount), 0)).where(
            PromotionFlow.flow_type == "支出",
            PromotionFlow.transaction_date >= start, PromotionFlow.transaction_date <= end)
    ).scalar() or 0)
    _os, os_est = _ofin.outsourcing_for_range(db, start, end, coef)
    outsourcing = float(_os)
    fixed_costs = float(_ofin.fixed_costs_monthly(db))            # 固定成本/管理费用 (房租等, 年度÷12)
    rc = _ofin.refill_cost(db, start, end, coef)                  # 补单=刷单 纯成本 (本金回流非收入)
    refill_c = float(rc["total"])
    # 本月真实净利 = 行净利合计 − 推广 − 人员 − 固定成本 − 补单成本(刷单是纯支出)
    period_net = sums["net_profit"] - promo - outsourcing - fixed_costs - refill_c

    # 问题单高亮: 亏损 / 支付宝未覆盖 (用推演是常态——未对账, 单列标蓝, 不算问题, 否则会全是问题)
    problem_count = sum(1 for r in rows if r["is_loss"] or not r["alipay_covered"])
    rows.sort(key=lambda r: (
        0 if r["is_loss"] else 1,
        0 if not r["alipay_covered"] else 1,
        -r["paid_amount"],
    ))

    return {
        "period": f"{year}-{month:02d}",
        "order_count": len(rows),
        "problem_count": problem_count,
        "loss_count": sum(1 for r in rows if r["is_loss"]),
        "uncovered_count": sum(1 for r in rows if not r["alipay_covered"]),
        "estimated_count": sum(1 for r in rows if r["cost_estimated"]),
        "rows": rows,
        "subtotal": {
            **{k: round(v, 2) for k, v in sums.items()},
            **{k: round(v, 2) for k, v in bsum.items()},   # 工厂成本核对合计(预算木作/配件/打包/实际木作)
            "promo_expense": round(promo, 2),
            "outsourcing_expense": round(outsourcing, 2),
            "outsourcing_estimated": os_est,
            "fixed_costs": round(fixed_costs, 2),
            "fixed_cost_items": _ofin.fixed_cost_items(db),
            "refill_count": rc["count"],
            "refill_gmv": float(rc["gmv"]),         # 刷单流水(本金, 来回滚抵销)
            "refill_cost": refill_c,                # 补单成本 = 平台扣点+税+运费+佣金
            "refill_platform": float(rc["platform"]),
            "refill_tax": float(rc["tax"]),
            "refill_commission": float(rc["commission"]),
            "period_net_profit": round(period_net, 2),
            "period_net_margin": round(period_net / sums["revenue"] * 100, 1) if sums["revenue"] else 0.0,
        },
    }


@router.get("/per-order-reconcile/export")
def per_order_reconcile_export(
    year: int = Query(2026),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
):
    """逐单核对 → 带颜色/格式的 xlsx (与页面同口径同配色, 客户自核对)。内存生成不落盘。"""
    import io
    from urllib.parse import quote as _q

    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = per_order_reconcile(year=year, month=month, db=db)
    rows, st = data["rows"], data["subtotal"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"逐单核对{data['period']}"

    GREEN, RED, BLUE, ORANGE, GREY = "389E0D", "CF1322", "1677FF", "FA8C16", "8C8C8C"
    MONEY = "¥#,##0;-¥#,##0"
    f_hdr = Font(bold=True, color="FFFFFF", size=10)
    fill_hdr = PatternFill("solid", fgColor="1F3A5F")
    fill_hdr2 = PatternFill("solid", fgColor="2E6CA8")
    fill_loss = PatternFill("solid", fgColor="FFF1F0")
    fill_sum = PatternFill("solid", fgColor="F0F0F0")
    side = Side(style="thin", color="E0E0E0")
    border = Border(left=side, right=side, top=side, bottom=side)
    ctr = Alignment(horizontal="center", vertical="center")
    rgt = Alignment(horizontal="right")

    headers = ["订单号", "产品", "订单金额", "退款", "真实收入", "商品成本", "物流", "安装",
               "平台扣点", "税", "售后", "成本合计", "净利", "净利率", "支付宝", "对账", "问题",
               "工厂账单", "预算木作", "预估配件", "预估打包", "实际木作", "木作差额"]
    widths = [20, 24, 11, 10, 11, 11, 8, 8, 10, 7, 8, 11, 11, 9, 9, 9, 10, 10, 11, 11, 11, 11, 11]
    FACTORY_FROM = 18

    ws.append(headers)
    for ci, _h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.font, c.alignment, c.border = f_hdr, ctr, border
        c.fill = fill_hdr2 if ci >= FACTORY_FROM else fill_hdr
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 22

    def _money(ri, ci, v, color=None, bold=False):
        c = ws.cell(ri, ci, float(v) if v is not None else None)
        c.number_format, c.alignment = MONEY, rgt
        if color or bold:
            c.font = Font(color=color, bold=bold, size=10)

    def _tag(ri, ci, txt, color):
        c = ws.cell(ri, ci, txt)
        c.alignment, c.font = ctr, Font(color=color, size=10)

    for r in rows:
        ri = ws.max_row + 1
        ws.cell(ri, 1, r["order_no"])
        ws.cell(ri, 2, r["product_name"])
        _money(ri, 3, r["paid_amount"])
        _money(ri, 4, -r["refund_amount"] if r["refund_amount"] else None, color=ORANGE)
        _money(ri, 5, r["revenue"], bold=True)
        _money(ri, 6, r["cost_goods"], color=(BLUE if r["cost_estimated"] else None))
        _money(ri, 7, r["cost_freight"])
        _money(ri, 8, r["cost_install"])
        _money(ri, 9, r["cost_platform"])
        _money(ri, 10, r["cost_tax"])
        _money(ri, 11, r["cost_aftersales"] or None)
        _money(ri, 12, r["cost_total"], bold=True)
        _money(ri, 13, r["net_profit"], color=(GREEN if r["net_profit"] >= 0 else RED), bold=True)
        mg = r["net_margin"] or 0
        mc = ws.cell(ri, 14, mg / 100)
        mc.number_format, mc.alignment = "0.0%", rgt
        mc.font = Font(color=(GREEN if mg >= 15 else ORANGE if mg >= 0 else RED), size=10)
        _tag(ri, 15, "已覆盖" if r["alipay_covered"] else "未覆盖", GREEN if r["alipay_covered"] else RED)
        _tag(ri, 16, "仅木作入账" if r["cost_reconciled"] else "推演", GREEN if r["cost_reconciled"] else BLUE)
        prob = (("亏损 " if r["is_loss"] else "") + ("未覆盖" if not r["alipay_covered"] else "")).strip()
        _tag(ri, 17, prob, RED if r["is_loss"] else ORANGE)
        _tag(ri, 18, "已入账" if r["factory_bill_recorded"] else "未入账",
             GREEN if r["factory_bill_recorded"] else GREY)
        _money(ri, 19, r["predicted_wood"])
        _money(ri, 20, r["est_parts"], color=GREY)
        _money(ri, 21, r["est_packaging"], color=GREY)
        _money(ri, 22, r["actual_wood"], color=(GREEN if r["actual_wood"] is not None else None))
        wd = r["wood_diff"]
        _money(ri, 23, wd, color=(RED if (wd or 0) > 0 else GREEN if wd is not None else None),
               bold=(wd is not None))
        for ci in range(1, len(headers) + 1):
            cc = ws.cell(ri, ci)
            cc.border = border
            if r["is_loss"]:
                cc.fill = fill_loss

    sr = ws.max_row + 1
    ws.cell(sr, 1, f"合计 {len(rows)} 单").font = Font(bold=True, size=10)
    sum_map = {3: "paid_amount", 5: "revenue", 6: "cost_goods", 7: "cost_freight",
               8: "cost_install", 9: "cost_platform", 10: "cost_tax", 11: "cost_aftersales",
               12: "cost_total", 13: "net_profit", 19: "predicted_wood", 20: "est_parts",
               21: "est_packaging", 22: "actual_wood", 23: "wood_diff"}
    sum_map[4] = "__refund__"
    for ci, key in sum_map.items():
        v = -st["refund_amount"] if key == "__refund__" else st.get(key)
        cc = ws.cell(sr, ci, float(v) if v else None)
        cc.number_format, cc.alignment, cc.font = MONEY, rgt, Font(bold=True, size=10)
    for ci in range(1, len(headers) + 1):
        cc = ws.cell(sr, ci)
        cc.fill, cc.border = fill_sum, border

    fr = ws.max_row + 2
    line = (f"行净利合计 ¥{st['net_profit']:,.0f}   − 推广费 ¥{st['promo_expense']:,.0f}   "
            f"− 人员成本 ¥{st['outsourcing_expense']:,.0f}   − 固定成本 ¥{st['fixed_costs']:,.0f}   "
            f"− 补单成本 ¥{st['refill_cost']:,.0f}   =   本月真实净利 ¥{st['period_net_profit']:,.0f} "
            f"({st['period_net_margin']}%)")
    fc = ws.cell(fr, 1, line)
    fc.font = Font(bold=True, size=11, color=GREEN if st["period_net_profit"] >= 0 else RED)
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=13)

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"逐单核对_{data['period']}.xlsx"
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_q(fname)}"},
    )


def _build_reconcile_workbook_all(db: Session, months: list[tuple[int, int]]):
    """逐单核对 多月工作簿: 每月一 sheet; **每个金额都可公式回推**:
    真实收入=实付−退款; 商品成本=工厂木作+定价表估算+打包(或 实付×85%); 成本合计=商品成本+物流+安装+平台+税+售后;
    净利=收入−成本合计; 净利率=净利/收入; 木作差额=实际−预算 —— 全为 Excel 公式, 改任一基础值自动重算。
    颜色+来源列+批注 标注: 绿=工厂账单实报 / 蓝=定价表推演 / 橙=实付×85%兜底封顶。纯函数, 便于测试。"""
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    GREEN, BLUE, ORANGE, GREY = "389E0D", "1677FF", "FA8C16", "8C8C8C"
    MONEY = "¥#,##0.00;-¥#,##0.00"   # 2位小数: 逐单核对要精确到分
    PCT = "0.0%"
    f_hdr = Font(bold=True, color="FFFFFF", size=10)
    fill_hdr = PatternFill("solid", fgColor="1F3A5F")
    fill_sum = PatternFill("solid", fgColor="F0F0F0")
    fill_actual = PatternFill("solid", fgColor="F6FFED")   # 绿底=实际账单
    fill_est = PatternFill("solid", fgColor="E6F4FF")      # 蓝底=定价表推演
    fill_cap = PatternFill("solid", fgColor="FFF7E6")      # 橙底=85%兜底/封顶
    side = Side(style="thin", color="E0E0E0")
    border = Border(left=side, right=side, top=side, bottom=side)
    ctr = Alignment(horizontal="center", vertical="center")
    rgt = Alignment(horizontal="right")

    HEAD = ["订单号", "产品", "下单日", "实付", "退款", "真实收入",
            "工厂木作账单", "定价表估算(配件+物流+安装)", "打包", "商品成本",
            "物流(额外)", "安装(额外)", "平台扣点", "税", "售后", "成本合计", "净利", "净利率",
            "成本来源", "预算木作", "实际木作", "木作差额"]
    WID = [20, 22, 11, 10, 9, 11, 12, 20, 8, 11, 10, 10, 10, 8, 8, 11, 11, 9, 18, 10, 10, 10]
    # 列号 (1-indexed): D实付 E退款 F收入 | G工厂木作 H定价估算 I打包 J商品成本 | K物流 L安装 M平台 N税 O售后 P成本合计 Q净利 R净利率 | S来源 T预算木作 U实际木作 V木作差额
    C_PAID, C_REFUND, C_REV = 4, 5, 6
    C_FWOOD, C_EST, C_PACK, C_GOODS = 7, 8, 9, 10
    C_FREIGHT, C_INSTALL, C_PLAT, C_TAX, C_AS = 11, 12, 13, 14, 15
    C_TOTAL, C_NET, C_MARGIN, C_SRC = 16, 17, 18, 19
    C_PRED, C_AWOOD, C_WDIFF = 20, 21, 22

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    any_sheet = False
    for (yy, mm) in months:
        data = per_order_reconcile(year=yy, month=mm, db=db)
        rows = data["rows"]
        if not rows:
            continue
        any_sheet = True
        ws = wb.create_sheet(title=f"{yy}-{mm:02d}")
        ws.cell(1, 1, "图例: 绿=工厂账单实报 · 蓝=定价表推演 · 橙=实付×85%兜底/封顶 ｜ 蓝字均为公式(改任一基础值自动重算): "
                      "真实收入=实付−退款 · 商品成本=工厂木作+定价估算+打包(或实付×85%) · 成本合计=商品成本+物流+安装+平台+税+售后 · 净利=收入−成本合计").font = Font(size=9, color=GREY)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEAD))
        for ci, h in enumerate(HEAD, 1):
            c = ws.cell(2, ci, h)
            c.font, c.alignment, c.fill, c.border = f_hdr, ctr, fill_hdr, border
            ws.column_dimensions[get_column_letter(ci)].width = WID[ci - 1]
        ws.freeze_panes = "D3"
        first = 3

        def _money(ri, ci, v, color=None, bold=False):
            c = ws.cell(ri, ci, float(v) if v is not None else None)
            c.number_format, c.alignment = MONEY, rgt
            if color or bold:
                c.font = Font(color=color, bold=bold, size=10)
            return c

        for r in rows:
            ri = ws.max_row + 1
            cap = r.get("cost_cap_mode", "none")
            if cap == "none":
                src, col, fillc = (("工厂账单实报", GREEN, fill_actual) if r.get("cost_reconciled")
                                   else ("定价表推演", BLUE, fill_est))
                goods_cell = f"=G{ri}+H{ri}+I{ri}"            # 商品成本 = 工厂木作 + 定价估算 + 打包
            elif cap.endswith("85"):
                src, col, fillc = "实付×85%(" + cap + ")", ORANGE, fill_cap
                goods_cell = f"=D{ri}*0.85"                   # 兜底/封顶 = 实付×85%
            else:   # 归零 等
                src, col, fillc = cap, GREY, fill_sum
                goods_cell = float(r["cost_goods"])           # 归零等: 直接取实值
            ws.cell(ri, 1, r["order_no"]); ws.cell(ri, 2, r["product_name"]); ws.cell(ri, 3, r["order_date"])
            _money(C_PAID, ri, r["paid_amount"])
            _money(C_REFUND, ri, r["refund_amount"] or None, color=ORANGE)
            ws.cell(ri, C_REV, f"=D{ri}-E{ri}").number_format = MONEY        # 真实收入=实付−退款
            # 商品成本拆解: 工厂木作 + 定价表估算 + 打包 (= 商品成本); 封顶单商品成本=实付×85%
            _money(C_FWOOD, ri, r.get("cost_factory_wood"), color=GREY)
            _money(C_EST, ri, r.get("cost_estimate_part"), color=GREY)
            _money(C_PACK, ri, r.get("cost_packing"), color=GREY)
            gc = ws.cell(ri, C_GOODS, goods_cell)
            gc.number_format, gc.alignment, gc.fill = MONEY, rgt, fillc
            gc.font = Font(color=col, bold=True, size=10)
            gc.comment = Comment(
                f"成本来源: {src}\n"
                f"商品成本 = 工厂木作账单(G) + 定价表估算·配件物流安装(H) + 打包(I);\n"
                f"工厂账单未到→工厂木作=0、整件走定价表估算(H); 走兜底/封顶时 商品成本=实付×0.85(G/H/I 仅供参考)。", "系统")
            _money(C_FREIGHT, ri, r["cost_freight"]); _money(C_INSTALL, ri, r["cost_install"])
            _money(C_PLAT, ri, r["cost_platform"]); _money(C_TAX, ri, r["cost_tax"]); _money(C_AS, ri, r["cost_aftersales"] or None)
            ws.cell(ri, C_TOTAL, f"=J{ri}+K{ri}+L{ri}+M{ri}+N{ri}+O{ri}").number_format = MONEY   # 成本合计
            nc = ws.cell(ri, C_NET, f"=F{ri}-P{ri}"); nc.number_format = MONEY; nc.font = Font(bold=True, size=10)  # 净利=收入−成本合计
            ws.cell(ri, C_MARGIN, f"=IF(F{ri}=0,0,Q{ri}/F{ri})").number_format = PCT   # 净利率
            sc = ws.cell(ri, C_SRC, src); sc.font = Font(color=col, size=9); sc.alignment = ctr
            _money(C_PRED, ri, r.get("predicted_wood"), color=GREY)
            _money(C_AWOOD, ri, r.get("actual_wood"), color=(GREEN if r.get("actual_wood") is not None else None))
            if r.get("predicted_wood") is not None and r.get("actual_wood") is not None:
                ws.cell(ri, C_WDIFF, f"=U{ri}-T{ri}").number_format = MONEY   # 木作差额=实际−预算
            for ci in range(1, len(HEAD) + 1):
                ws.cell(ri, ci).border = border
        last = ws.max_row

        sr = last + 1   # 合计行 (SUM 公式)
        ws.cell(sr, 1, f"合计 {len(rows)} 单").font = Font(bold=True, size=10)
        for ci in (C_PAID, C_REFUND, C_REV, C_FWOOD, C_EST, C_PACK, C_GOODS,
                   C_FREIGHT, C_INSTALL, C_PLAT, C_TAX, C_AS, C_TOTAL, C_NET):
            L = get_column_letter(ci)
            cc = ws.cell(sr, ci, f"=SUM({L}{first}:{L}{last})")
            cc.number_format, cc.font = MONEY, Font(bold=True, size=10)
        ws.cell(sr, C_MARGIN, f"=IF(F{sr}=0,0,Q{sr}/F{sr})").number_format = PCT
        for ci in range(1, len(HEAD) + 1):
            ws.cell(sr, ci).fill = fill_sum; ws.cell(sr, ci).border = border

        st = data["subtotal"]   # 本月真实净利 = 行净利合计 − 推广 − 人员 − 固定 − 补单 (公式)
        fr2 = sr + 2
        items = [("推广费", st["promo_expense"]), ("人员外包", st["outsourcing_expense"]),
                 ("固定成本", st["fixed_costs"]), ("补单成本", st["refill_cost"])]
        ws.cell(fr2, 1, "本月真实净利 = 行净利合计 − 推广 − 人员 − 固定 − 补单 :").font = Font(bold=True, size=10)
        for i, (name, val) in enumerate(items, 1):
            ws.cell(fr2 + i, 1, name)
            ws.cell(fr2 + i, 2, float(val)).number_format = MONEY
        ws.cell(fr2 + 5, 1, "本月真实净利").font = Font(bold=True, size=11)
        rn = ws.cell(fr2 + 5, 2, f"=Q{sr}-B{fr2 + 1}-B{fr2 + 2}-B{fr2 + 3}-B{fr2 + 4}")
        rn.number_format, rn.font = MONEY, Font(bold=True, size=11, color=GREEN)

    if not any_sheet:
        ws = wb.create_sheet(title="无数据")
        ws.cell(1, 1, "该时间段无成交订单")
    return wb


@router.get("/per-order-reconcile/export-all")
def per_order_reconcile_export_all(
    from_year: int = Query(2026),
    from_month: int = Query(1, ge=1, le=12),
    to_year: Optional[int] = Query(None),
    to_month: Optional[int] = Query(None, ge=1, le=12),
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin", "operator", "viewer")),
):
    """逐单核对 · 多月导出 → xlsx (每月一 sheet, 金额带公式可回推, 标注 预估/实际/85%兜底)。用户 2026-06-25。"""
    import io
    from urllib.parse import quote as _q

    today = _date.today()
    end_year = to_year or today.year
    end_month = to_month or today.month
    months: list[tuple[int, int]] = []
    y, m = from_year, from_month
    while (y, m) <= (end_year, end_month):
        months.append((y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1

    wb = _build_reconcile_workbook_all(db, months)
    buf = io.BytesIO()
    wb.save(buf)
    fname = f"逐单核对_{from_year}-{from_month:02d}_至_{end_year}-{end_month:02d}.xlsx"
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_q(fname)}"},
    )

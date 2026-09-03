"""活动生命周期 API (P1, 2026-07-17 spec: docs/活动生命周期系统_执行plan.md §五)。

GET    /api/campaigns                     计划列表
POST   /api/campaigns                     建计划 (类型点选 + 档期精确到秒)
GET    /api/campaigns/no-sales-group      动销分组 (含登记表同步结果)
POST   /api/campaigns/no-sales-group/notify  无动销名单推飞书 (spec §四.2a)
GET    /api/campaigns/{id}                计划详情
PUT    /api/campaigns/{id}                改计划
DELETE /api/campaigns/{id}                删计划 (admin)
POST   /api/campaigns/{id}/precheck       R0~R19 预检
GET    /api/campaigns/{id}/rows           行预览 (kind=signup|discount)
POST   /api/campaigns/{id}/push-discount  推单品立减 (phase=stage|commit, admin)
POST   /api/campaigns/{id}/push-signup    已禁用 (仅自动报名程序内部可执行)
POST   /api/campaigns/{id}/recon          核对 (multipart 手动上传三种导出兜底)
GET    /api/campaigns/{id}/recon-reports  核对报告列表

页面权限: /api/campaigns 挂 permKey "pricing" (活动自动填写 wizard 在定价页, 见 page_permissions)。
"""
from __future__ import annotations

from datetime import datetime
import re
from typing import Optional
from urllib.parse import quote
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import (
    ServicePrincipal,
    get_current_user,
    require_campaign_preparation_bundle_principal,
    require_campaign_prepare_principal,
    require_web_agent_plan8_v4_claim_verifier,
    require_web_agent_plan8_v5_claim_verifier,
    require_web_agent_plan8_v6_claim_verifier,
    require_web_agent_plan8_v7_claim_verifier,
    require_web_agent_plan8_v8_claim_verifier,
    require_role,
)
from app.models.auth import User
from app.models.campaign import CampaignPlan, CampaignReconReport
from app.services import campaign_service

router = APIRouter(prefix="/api/campaigns", tags=["campaigns"])


class CampaignPlanIn(BaseModel):
    name: str
    campaign_type: str
    start_at: datetime          # 档期精确到秒 (spec §四.4)
    end_at: datetime
    qn_campaign_title: Optional[str] = None
    price_protection_days: int = 19
    price_protection_rule_url: Optional[str] = None
    remark: Optional[str] = None
    workflow_key: Optional[str] = None
    platform_activity_mode: str = "fixed_window"
    platform_campaign_id: Optional[str] = None
    platform_united_activity_id: Optional[str] = None
    platform_sign_record_id: Optional[str] = None
    platform_active_until: Optional[datetime] = None


class CampaignPrepareIn(CampaignPlanIn):
    workflow_key: str
    official_all_store: Optional[bool] = None
    official_active_item_ids: list[str] = Field(default_factory=list)
    official_exempt_item_ids: list[str] = Field(default_factory=list)


class CampaignEvidenceRefreshIn(BaseModel):
    workflow_key: str
    plan_id: Optional[int] = Field(default=None, ge=1)


class CampaignPreparationBundleIn(BaseModel):
    """Read-only preparation compiler input; never permits a final write."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: Optional[str] = None
    exact_item_scope: list[str] = Field(default_factory=list)
    mode: str = Field(pattern=r"^(compile|refresh_and_compile|read_latest)$")


class CampaignOfficialExemptionsCorrectionIn(BaseModel):
    """CAS-style correction of one plan-scoped official exemption list."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_official_exempt_item_ids: list[str] = Field(default_factory=list)
    official_exempt_item_ids: list[str] = Field(default_factory=list)


class CampaignPlan7ResumeExecuteIn(BaseModel):
    """Exact CAS input for the one approved Super Reduce recovery."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    expected_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7FinalCloseoutIn(BaseModel):
    """Immutable-bundle identity for the last plan-7 safe item."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    bundle_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    expected_source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_policy_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_manifest_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_item_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7FinalCloseoutV2In(CampaignPlan7FinalCloseoutIn):
    """Second-generation recovery identity after a pre-claim export failure."""

    recovery_id: str = Field(min_length=1, max_length=80)
    expected_web_agent_commit: str = Field(pattern=r"^[0-9a-f]{40}$")


class CampaignPlan7FinalCloseoutV4In(BaseModel):
    """Fixed V4 identity after the official SKU-scope incident."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    source_bundle_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    expected_source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_policy_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_source_manifest_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_item_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    recovery_id: str = Field(min_length=1, max_length=80)
    expected_web_agent_commit: str = Field(pattern=r"^[0-9a-f]{40}$")
    official_export_record_id: str = Field(pattern=r"^[0-9]+$")
    expected_official_export_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_extra_sku_id: str = Field(pattern=r"^[0-9]+$")


class CampaignPlan8ExecuteIn(BaseModel):
    """Exact CAS input for the approved September Super88 execution."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    expected_candidate_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8SignupRecoveryIn(BaseModel):
    """Exact CAS input for the signup-only recovery after partial execution."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    expected_original_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    expected_original_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_full_signup_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_pending_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_policy_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_candidate_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8FinalRecoveryV2In(BaseModel):
    """Fixed request for the full-SKU plan-8 recovery."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    recovery_version: int = Field(ge=2, le=2)


class CampaignPlan8FinalRecoveryV4In(BaseModel):
    """Fixed request for in-place completion of six bound plan-8 drafts."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    recovery_version: int = Field(ge=4, le=4)
    mode: str = Field(pattern=r"^(execute|readback)$")
    confirmation: str
    target_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8FinalRecoveryV4ClaimVerifyIn(BaseModel):
    """Web-Agent's fixed, read-only proof request before consuming its lease."""

    model_config = ConfigDict(extra="forbid")

    attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    workflow_key: str
    plan_id: int = Field(ge=1)
    operation: str
    scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    inspect_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    reservation_token_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8FinalRecoveryV5In(BaseModel):
    """Fixed request for the independent V5 recovery."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    recovery_version: int = Field(ge=5, le=5)
    mode: str = Field(pattern=r"^(execute|readback)$")
    confirmation: str
    target_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8FinalRecoveryV5ClaimVerifyIn(
        CampaignPlan8FinalRecoveryV4ClaimVerifyIn):
    """Web-Agent V5 read-only claim proof request."""


class CampaignPlan8FinalRecoveryV6In(BaseModel):
    """Fixed request for the independent V6 recovery."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    recovery_version: int = Field(ge=6, le=6)
    mode: str = Field(pattern=r"^(execute|readback)$")
    confirmation: str
    target_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8FinalRecoveryV6ClaimVerifyIn(
        CampaignPlan8FinalRecoveryV4ClaimVerifyIn):
    """Web-Agent V6 read-only claim proof request."""


class CampaignPlan8FinalRecoveryV7In(BaseModel):
    """Fixed request for the independent V7 recovery."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    recovery_version: int = Field(ge=7, le=7)
    mode: str = Field(pattern=r"^(execute|readback)$")
    confirmation: str
    target_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8FinalRecoveryV7ClaimVerifyIn(
        CampaignPlan8FinalRecoveryV4ClaimVerifyIn):
    """Web-Agent V7 read-only claim proof request."""


class CampaignPlan8FinalRecoveryV8In(BaseModel):
    """Fixed request for the V8 publish-before-discount continuation."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    recovery_version: int = Field(ge=8, le=8)
    mode: str = Field(pattern=r"^(execute|readback)$")
    confirmation: str
    target_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan8FinalRecoveryV8ClaimVerifyIn(
        CampaignPlan8FinalRecoveryV4ClaimVerifyIn):
    """Web-Agent V8 read-only claim proof request."""


class CampaignPlan7PostSubmitVerifyIn(BaseModel):
    """Exact read-only delayed verification for the submitted plan-7 attempt."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    expected_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7DiscountAuditIn(BaseModel):
    """Exact read-only current-state audit for plan 7 single discounts."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7DiscountTimeUpdateIn(BaseModel):
    """Exact CAS input for the approved three-activity time-only update."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    activity_ids: list[str] = Field(min_length=3, max_length=3)
    expected_start_at: str
    expected_end_at: str
    target_start_at: str
    target_end_at: str


class CampaignPlan7DiscountTimePrewriteReceiptIn(BaseModel):
    """One fixed external receipt proving no activity write was reached."""

    model_config = ConfigDict(extra="forbid")

    request_id: str = Field(pattern=r"^[0-9a-f]{12}$")
    web_agent_job_id: str
    attempt_id: str | None = Field(default=None, pattern=r"^[0-9a-f]{24}$")
    platform_write: bool
    submitted: bool
    confirmed_activity_ids: list[str]


class CampaignPlan7DiscountTimeRecoveryIn(CampaignPlan7DiscountTimeUpdateIn):
    """Exact one-time recovery after the two known write-free stops."""

    failed_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    prewrite_receipts: list[CampaignPlan7DiscountTimePrewriteReceiptIn] = Field(
        min_length=2, max_length=2)


class CampaignPlan7DiscountTimeRecoveryV2ReceiptIn(BaseModel):
    """The exact first-recovery receipt proving it stopped before a claim."""

    model_config = ConfigDict(extra="forbid")

    request_id: str = Field(pattern=r"^[0-9a-f]{12}$")
    web_agent_job_id: str
    platform_write: bool
    submitted: bool
    confirmed_activity_ids: list[str]
    recovery_not_claimed: bool


class CampaignPlan7DiscountTimeRecoveryV2In(
        CampaignPlan7DiscountTimeRecoveryIn):
    """Replacement recovery bound to the exact write-free V1 receipt."""

    first_recovery_receipt: CampaignPlan7DiscountTimeRecoveryV2ReceiptIn


class CampaignPlan7DiscountTimeReadbackV3In(BaseModel):
    """Exact read-only closeout of the already-submitted V2 attempt."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    request_id: str
    web_agent_job_id: str
    external_request_id: str = Field(pattern=r"^[0-9a-f]{12}$")
    confirmed_activity_ids: list[str] = Field(min_length=3, max_length=3)


class CampaignPlan7DiscountCorrectionIn(BaseModel):
    """Immutable CAS input for the exact four-row plan-7 correction."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_snapshot_id: int = Field(ge=1)
    expected_snapshot_artifact_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    expected_missing_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7DiscountSupplementRowIn(BaseModel):
    model_config = ConfigDict(extra="forbid")

    item_id: str = Field(pattern=r"^\d+$")
    sku_id: str = Field(pattern=r"^\d+$")
    expected_deduct: str = Field(pattern=r"^\d+(?:\.\d{2})$")


class CampaignPlan7DiscountSupplementIn(BaseModel):
    """Immutable input for the exact four-SKU existing-activity supplement."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    activity_ids: list[str] = Field(min_length=3, max_length=3)
    target_activity_id: str = Field(pattern=r"^\d+$")
    item_id: str = Field(pattern=r"^\d+$")
    rows: list[CampaignPlan7DiscountSupplementRowIn] = Field(
        min_length=4, max_length=4)
    scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    readonly_artifact_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    start_at: str
    end_at: str


class CampaignPlan7SmallPromoCorrectionIn(BaseModel):
    """Immutable identity for the fixed two-item / 20-SKU correction."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    activity_ids: list[str] = Field(min_length=3, max_length=3)
    target_activity_id: str = Field(pattern=r"^\d+$")
    manifest_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    current_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    target_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    source_snapshot_id: int = Field(ge=1)
    source_snapshot_artifact_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    start_at: str
    end_at: str


class CampaignWarehouseProductPriceCorrectionIn(BaseModel):
    """Immutable identity for one warehouse-preserving SKU price correction."""

    model_config = ConfigDict(extra="forbid")

    item_id: str = Field(pattern=r"^\d+$")
    sku_id: str = Field(pattern=r"^\d+$")
    expected_title: str
    expected_product_state: str
    expected_old_price: str = Field(pattern=r"^\d+(?:\.\d{2})$")
    target_price: str = Field(pattern=r"^\d+(?:\.\d{2})$")
    expected_quantity: str = Field(pattern=r"^\d+$")
    baseline_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    target_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignFivePriceCorrectionIn(BaseModel):
    """Exact one-shot phase for the reviewed five-item correction."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    phase: str = Field(pattern=r"^(single_discount|super_reduce)$")
    manifest_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    source_export_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignFivePriceRecoveryIn(CampaignFivePriceCorrectionIn):
    """One recovery bound to the exact NAS-to-Web-Agent timeout attempt."""

    expected_failed_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    confirmed_no_platform_write: bool


class CampaignFivePriceSuperRecoveryV2In(CampaignFivePriceCorrectionIn):
    """Recovery after fresh readback proves the next-step attempt persisted none."""

    expected_failed_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    confirmed_no_persisted_change: bool


class CampaignFivePriceSuperRecoveryV3In(CampaignFivePriceCorrectionIn):
    """Recovery bound to one persisted item and three untouched items."""

    expected_failed_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    confirmed_partial_persisted_change: bool
    persisted_item_id: str = Field(pattern=r"^\d+$")


class CampaignPlan7DiscountIdentityRecoveryIn(BaseModel):
    """Exact official-export-bound SKU identity repair and one recovery."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_old_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    official_product_export_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    official_product_export_b64: str = Field(min_length=1, max_length=3_000_000)
    expected_new_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7DiscountIdentityReadbackIn(BaseModel):
    """Exact read-only closeout for the already-submitted recovery attempt."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    expected_terminal_evidence_request_id: str = Field(min_length=1)
    expected_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7RemainingSignupIn(BaseModel):
    """Exact one-shot input for the approved remaining plan-7 item scope."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_status: str
    expected_item_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    recovery_incident_id: str = Field(min_length=1)


class CampaignPlan7PartialSignupAuditIn(BaseModel):
    """Exact read-only closeout for the one partial draft import."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    expected_manifest_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlan7DraftPublishIn(BaseModel):
    """Exact one-shot publication of two existing audited platform drafts."""

    model_config = ConfigDict(extra="forbid")

    workflow_key: str
    plan_id: int = Field(ge=1)
    expected_attempt_id: str = Field(pattern=r"^[0-9a-f]{24}$")
    expected_snapshot_id: int = Field(ge=1)
    expected_scope_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


class CampaignPlanUpdate(BaseModel):
    name: Optional[str] = None
    campaign_type: Optional[str] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    qn_campaign_title: Optional[str] = None
    price_protection_days: Optional[int] = None
    price_protection_rule_url: Optional[str] = None
    remark: Optional[str] = None
    platform_activity_mode: Optional[str] = None
    platform_campaign_id: Optional[str] = None
    platform_united_activity_id: Optional[str] = None
    platform_active_until: Optional[datetime] = None


class CampaignItemExclusionIn(BaseModel):
    taobao_item_id: str
    reason: str


class SuperReduceRepairIn(BaseModel):
    item_ids: list[str]
    phase: str = "stage"


def _plan_out(p: CampaignPlan) -> dict:
    from app.services import campaign_price_protection_service
    until = campaign_price_protection_service.protection_until(p)
    return {
        "id": p.id, "name": p.name, "campaign_type": p.campaign_type, "tier": p.tier,
        "campaign_type_name": campaign_service.CAMPAIGN_TYPES.get(p.campaign_type, ("?",))[0],
        "start_at": p.start_at.isoformat(sep=" ") if p.start_at else None,
        "end_at": p.end_at.isoformat(sep=" ") if p.end_at else None,
        "qn_campaign_title": p.qn_campaign_title,
        "price_protection_days": campaign_price_protection_service.protection_days(p),
        "price_protection_rule_url": p.price_protection_rule_url,
        "price_protection_confirmed_at": (
            p.price_protection_confirmed_at.isoformat(sep=" ")
            if p.price_protection_confirmed_at else None),
        "price_protection_until": until.isoformat(sep=" ") if until else None,
        "status": p.status, "remark": p.remark,
        "workflow_key": p.workflow_key,
        "platform_activity_mode": p.platform_activity_mode,
        "platform_campaign_id": p.platform_campaign_id,
        "platform_united_activity_id": p.platform_united_activity_id,
        "platform_sign_record_id": p.platform_sign_record_id,
        "platform_active_until": (
            p.platform_active_until.isoformat(sep=" ")
            if p.platform_active_until else None),
    }


def _get_plan(db: Session, plan_id: int) -> CampaignPlan:
    plan = db.get(CampaignPlan, plan_id)
    if plan is None:
        raise HTTPException(404, f"计划 {plan_id} 不存在")
    return plan


def _xlsx_download_response(content: bytes, filename: str, *,
                            metadata: Optional[dict] = None) -> Response:
    safe_name = str(filename or "campaign_feedback.xlsx").replace("\r", "").replace("\n", "")
    headers = {
        "Content-Disposition": (
            "attachment; filename=campaign_feedback.xlsx; "
            f"filename*=UTF-8''{quote(safe_name)}"
        )
    }
    for key, value in (metadata or {}).items():
        if value is not None:
            headers[f"X-Panse-{key}"] = str(value).lower() if isinstance(value, bool) else str(value)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _validate_type_and_window(campaign_type: str, start_at, end_at) -> str:
    if campaign_type not in campaign_service.CAMPAIGN_TYPES:
        raise HTTPException(422, f"未知活动类型 {campaign_type!r}; "
                                 f"可选 {list(campaign_service.CAMPAIGN_TYPES)}")
    if start_at and end_at and end_at <= start_at:
        raise HTTPException(422, "档期结束时间必须晚于开始时间")
    return campaign_service.CAMPAIGN_TYPES[campaign_type][1]


def _validate_price_protection(days: Optional[int], url: Optional[str]) -> None:
    if days is not None and not 1 <= int(days) <= 365:
        raise HTTPException(422, "价保冷静期必须是1至365天")
    if url and not str(url).strip().lower().startswith(("http://", "https://")):
        raise HTTPException(422, "价保说明链接必须以 http:// 或 https:// 开头")


def _validate_workflow_key(value: str) -> str:
    key = str(value or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._:-]{7,127}", key):
        raise HTTPException(422, "workflow_key 必须为8至128位稳定字母数字/._:-标识")
    return key


def _validate_sign_record_identity(mode: str, value: Optional[str]) -> None:
    if value is None:
        return
    if mode != "fixed_window" or not str(value).isdigit():
        raise HTTPException(
            422,
            "platform_sign_record_id 仅可用于 fixed_window 且必须是数字",
        )


def _validate_formal_platform_identity(body: CampaignPrepareIn) -> None:
    mode = body.platform_activity_mode
    if mode not in ("fixed_window", "long_running_update"):
        raise HTTPException(422, "platform_activity_mode 必须是 fixed_window 或 long_running_update")
    if body.campaign_type == "super_reduce":
        if mode != "long_running_update":
            raise HTTPException(422, "超级立减是长期活动，必须使用 long_running_update 更新窗口")
        if body.platform_active_until is None or body.platform_active_until < body.end_at:
            raise HTTPException(422, "长期超级立减必须提供晚于更新窗口的官方有效期")
    elif mode != "fixed_window":
        raise HTTPException(422, "大促必须使用 fixed_window 精确档期")
    if mode == "fixed_window":
        for field, value in (
            ("platform_campaign_id", body.platform_campaign_id),
            ("platform_united_activity_id", body.platform_united_activity_id),
        ):
            if not str(value or "").isdigit():
                raise HTTPException(422, f"固定档期活动必须提供数字 {field}")
    _validate_sign_record_identity(mode, body.platform_sign_record_id)
    if not str(body.qn_campaign_title or "").strip():
        raise HTTPException(422, "正式准备入口必须提供千牛官方活动标题")


def _local_naive(value: Optional[datetime]) -> Optional[datetime]:
    """Normalize API timestamps to the existing Asia/Shanghai naive DB contract."""
    if value is None or value.tzinfo is None:
        return value
    return value.astimezone(ZoneInfo("Asia/Shanghai")).replace(tzinfo=None)


def _structured_prepare_remark(body: CampaignPrepareIn) -> Optional[str]:
    """Persist typed external-platform scope without asking callers to craft markers."""
    from app.services import no_sales_service

    active = no_sales_service.normalize_item_ids(body.official_active_item_ids)
    exempt = no_sales_service.normalize_item_ids(body.official_exempt_item_ids)
    if len(active) != len(set(body.official_active_item_ids)):
        raise HTTPException(422, "official_active_item_ids 含无效淘宝商品号")
    if len(exempt) != len(set(body.official_exempt_item_ids)):
        raise HTTPException(422, "official_exempt_item_ids 含无效淘宝商品号")
    if body.official_all_store is True and active:
        raise HTTPException(422, "official_all_store=true 时不得同时填写 active_item_ids")
    text = str(body.remark or "")
    for key in ("official_all_store", "official_active_items", "official_exempt_items"):
        text = re.sub(
            rf"(?:^|[;\n；])\s*{key}\s*=\s*[^;\n；]*", "", text,
            flags=re.IGNORECASE).strip(" ;\n；")
    markers = []
    if body.official_all_store is not None:
        markers.append(
            f"official_all_store={'true' if body.official_all_store else 'false'}")
    if active:
        markers.append(f"official_active_items={','.join(sorted(active))}")
    # Pydantic keeps track of whether the caller actually supplied the field.
    # An explicit [] is materially different from omission here: for an
    # all-store official discount it proves that the operator checked the live
    # activity and found no exempt items.  Persist the empty marker so R15 can
    # distinguish that evidence from an unconfigured scope.
    if "official_exempt_item_ids" in body.model_fields_set:
        markers.append(f"official_exempt_items={','.join(sorted(exempt))}")
    return "; ".join([value for value in (text, *markers) if value]) or None


@router.get("/policy")
def get_campaign_policy(_: User = Depends(get_current_user)):
    """Root policy used by every signup generator and shown in the wizard."""
    from app.services import campaign_policy_service
    try:
        return campaign_policy_service.public_policy()
    except RuntimeError as exc:
        raise HTTPException(503, str(exc)) from exc


@router.get("")
def list_plans(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    plans = db.execute(select(CampaignPlan).order_by(CampaignPlan.id.desc())).scalars().all()
    return {"items": [_plan_out(p) for p in plans],
            "types": {k: v[0] for k, v in campaign_service.CAMPAIGN_TYPES.items()}}


@router.post("")
def create_plan(body: CampaignPlanIn, db: Session = Depends(get_db),
                _: User = Depends(require_role("admin", "operator"))):
    tier = _validate_type_and_window(body.campaign_type, body.start_at, body.end_at)
    _validate_price_protection(body.price_protection_days, body.price_protection_rule_url)
    _validate_sign_record_identity(
        body.platform_activity_mode, body.platform_sign_record_id)
    plan = CampaignPlan(name=body.name, campaign_type=body.campaign_type, tier=tier,
                        start_at=body.start_at, end_at=body.end_at,
                        qn_campaign_title=body.qn_campaign_title,
                        price_protection_days=body.price_protection_days,
                        price_protection_rule_url=body.price_protection_rule_url,
                        price_protection_confirmed_at=(
                            datetime.now() if body.price_protection_rule_url else None),
                        workflow_key=body.workflow_key,
                        platform_activity_mode=body.platform_activity_mode,
                        platform_campaign_id=body.platform_campaign_id,
                        platform_united_activity_id=body.platform_united_activity_id,
                        platform_sign_record_id=body.platform_sign_record_id,
                        platform_active_until=body.platform_active_until,
                        remark=body.remark,
                        status="draft")
    db.add(plan)
    db.commit()
    return _plan_out(plan)


@router.post("/prepare")
def prepare_campaign(
        body: CampaignPrepareIn, db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Formal ERP-only campaign package; never reads ERP through a browser."""
    from app.services import campaign_workflow_service

    workflow_key = _validate_workflow_key(body.workflow_key)
    _validate_type_and_window(body.campaign_type, body.start_at, body.end_at)
    _validate_price_protection(body.price_protection_days, body.price_protection_rule_url)
    _validate_formal_platform_identity(body)
    values = body.model_dump(exclude={
        "workflow_key", "official_all_store", "official_active_item_ids",
        "official_exempt_item_ids",
    })
    values["remark"] = _structured_prepare_remark(body)
    for field in ("start_at", "end_at", "platform_active_until"):
        values[field] = _local_naive(values.get(field))
    values["tier"] = campaign_service.CAMPAIGN_TYPES[body.campaign_type][1]
    values["price_protection_confirmed_at"] = (
        datetime.now() if body.price_protection_rule_url else None)
    result = campaign_workflow_service.prepare(
        db, workflow_key=workflow_key, values=values)
    if result.get("conflict"):
        raise HTTPException(409, detail={
            "error": "workflow_key_payload_conflict",
            "workflow_key": workflow_key,
            "plan_id": result["plan"].id,
            "different_fields": result["different_fields"],
        })
    result["plan"] = _plan_out(result["plan"])
    return result


@router.post("/correct-official-exemptions")
def correct_campaign_official_exemptions(
        body: CampaignOfficialExemptionsCorrectionIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Correct only one unsubmitted plan's official exemption scope.

    This endpoint cannot alter campaign identity, prices or permanent item
    exclusions.  Replays of the desired state are idempotent; changing from a
    different current list requires the caller's exact expected-old list.
    """
    from app.services import campaign_workflow_service

    workflow_key = _validate_workflow_key(body.workflow_key)
    result = campaign_workflow_service.correct_official_exemptions(
        db,
        workflow_key=workflow_key,
        expected_plan_id=body.plan_id,
        expected_item_ids=body.expected_official_exempt_item_ids,
        desired_item_ids=body.official_exempt_item_ids,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in ("invalid_item_ids", "official_scope_not_correctable"):
            code = 422
        raise HTTPException(code, detail=result)
    result["plan"] = _plan_out(result["plan"])
    return result


@router.post("/refresh-evidence")
def refresh_campaign_evidence(
        body: CampaignEvidenceRefreshIn, db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Read-only QianNiu export for one workflow, then formal ERP preflight.

    The route cannot select any platform-write operation and has no notification
    or retry branch.  The dedicated machine identity is valid only on this exact
    path and ``/prepare``.
    """
    from app.services import campaign_workflow_service

    workflow_key = _validate_workflow_key(body.workflow_key)
    result = campaign_workflow_service.refresh_evidence_and_prepare(
        db, workflow_key=workflow_key, expected_plan_id=body.plan_id)
    if not result.get("ok"):
        code = 404 if result.get("error") == "workflow_not_found" else 409
        if result.get("error") not in (
                "workflow_not_found", "workflow_plan_mismatch"):
            code = 422
        raise HTTPException(code, detail=result)
    result["plan"] = _plan_out(result["plan"])
    return result


@router.post("/prepare-final-bundle")
def prepare_final_campaign_bundle(
        body: CampaignPreparationBundleIn, db: Session = Depends(get_db),
        principal: User | ServicePrincipal = Depends(
            require_campaign_preparation_bundle_principal)):
    """Freeze a durable package immediately before the final campaign action.

    ``refresh_and_compile`` may only perform the existing read-only evidence
    export.  None of the modes can upload, submit, edit prices, rotate SKUs,
    notify, or create a platform-write claim.
    """
    from app.services import campaign_preparation_service

    workflow_key = _validate_workflow_key(body.workflow_key)
    if body.mode == "read_latest":
        result = campaign_preparation_service.get_latest_bundle(
            db, workflow_key=workflow_key, expected_plan_id=body.plan_id)
    else:
        result = campaign_preparation_service.compile_bundle(
            db,
            workflow_key=workflow_key,
            expected_plan_id=body.plan_id,
            expected_status=body.expected_status,
            refresh_evidence=body.mode == "refresh_and_compile",
            exact_item_scope=set(body.exact_item_scope),
            prepared_by=principal.username,
        )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found", "preparation_bundle_not_found"} else 409
        raise HTTPException(code, detail=result)
    return result


@router.post("/resume-super-reduce-plan7")
def resume_super_reduce_plan7(
        body: CampaignPlan7ResumeExecuteIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Resume exactly plan 7 once from its fresh stored evidence.

    The service is fail-closed and cannot select another workflow/plan, refresh
    pre-submit evidence, rotate SKUs, change prices, or automatically retry.
    A successful platform submission is still followed by the normal exact
    post-submit export/verification so submission is not mistaken for success.
    """
    from app.services import campaign_resume_service

    workflow_key = _validate_workflow_key(body.workflow_key)
    result = campaign_resume_service.resume_super_reduce_plan7(
        db,
        workflow_key=workflow_key,
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        expected_scope_sha256=body.expected_scope_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "resume_identity_not_allowed", "resume_request_not_allowed",
                "resume_scope_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/execute-super-reduce-plan7-final-closeout")
def execute_super_reduce_plan7_final_closeout(
        body: CampaignPlan7FinalCloseoutIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Retired after the pre-claim export ambiguity incident."""
    result = {
        "ok": False,
        "error": "final_closeout_v1_retired_after_preclaim_export_ambiguity",
        "automatic_retry": False,
        "platform_write": False,
    }
    raise HTTPException(409, detail=result)


@router.post("/execute-super-reduce-plan7-final-closeout-v2")
def execute_super_reduce_plan7_final_closeout_v2(
        body: CampaignPlan7FinalCloseoutV2In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Retired because its service path was not registered in the auth scope."""
    result = {
        "ok": False,
        "error": "final_closeout_v2_retired_before_platform_read",
        "automatic_retry": False,
        "platform_write": False,
    }
    raise HTTPException(409, detail=result)


@router.post("/execute-super-reduce-plan7-final-closeout-v3")
def execute_super_reduce_plan7_final_closeout_v3(
        body: CampaignPlan7FinalCloseoutV2In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Retired after the official 14-SKU export disproved its 13-SKU scope."""
    raise HTTPException(409, detail={
        "ok": False,
        "error": "final_closeout_v3_retired_after_official_sku_scope_incident",
        "automatic_retry": False,
        "platform_write": False,
    })


@router.post("/execute-super-reduce-plan7-final-closeout-v4")
def execute_super_reduce_plan7_final_closeout_v4(
        body: CampaignPlan7FinalCloseoutV4In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Repair the exact SKU identity incident and consume one fresh bundle."""
    from app.services import campaign_plan7_final_closeout_v4_service as service

    payload = body.model_dump()
    payload["workflow_key"] = _validate_workflow_key(body.workflow_key)
    result = service.execute_plan7_final_closeout_v4(db, payload)
    if not result.get("ok"):
        code = 404 if result.get("error") in {
            "workflow_not_found", "final_closeout_v4_source_bundle_not_found",
        } else 409
        if result.get("error") == "final_closeout_v4_request_not_allowed":
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/execute-super88-plan8")
def execute_super88_plan8(
        body: CampaignPlan8ExecuteIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Execute exactly plan 8 once from fresh candidate-list evidence."""
    from app.services import campaign_plan8_execute_service

    workflow_key = _validate_workflow_key(body.workflow_key)
    result = campaign_plan8_execute_service.execute_plan8(
        db,
        workflow_key=workflow_key,
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        expected_candidate_sha256=body.expected_candidate_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_execute_request_not_allowed",
                "plan8_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-signup")
def recover_super88_plan8_signup(
        body: CampaignPlan8SignupRecoveryIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Recover only the six pending plan-8 items after discount completion."""
    from app.services import campaign_plan8_signup_recovery_service

    result = campaign_plan8_signup_recovery_service.recover_plan8_signup(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        expected_original_attempt_id=body.expected_original_attempt_id,
        expected_original_scope_sha256=body.expected_original_scope_sha256,
        expected_full_signup_scope_sha256=body.expected_full_signup_scope_sha256,
        expected_pending_scope_sha256=body.expected_pending_scope_sha256,
        expected_policy_sha256=body.expected_policy_sha256,
        expected_candidate_sha256=body.expected_candidate_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_signup_recovery_request_not_allowed",
                "plan8_signup_recovery_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v2")
def recover_super88_plan8_final_v2(
        body: CampaignPlan8FinalRecoveryV2In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Refresh evidence, add eight discounts, then enroll 6 items / 78 SKUs."""
    from app.services import campaign_plan8_final_recovery_v2_service

    result = campaign_plan8_final_recovery_v2_service.recover_plan8_final_v2(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        recovery_version=body.recovery_version,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_final_v2_request_not_allowed",
                "plan8_final_v2_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v3")
def recover_super88_plan8_final_v3_retired(
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """The failed pre-write V3 entry is permanently retired."""
    raise HTTPException(409, detail={
        "ok": False, "error": "plan8_final_v3_retired_use_v4",
        "platform_write": False, "claim_created": False,
    })


@router.post("/recover-super88-plan8-final-v4")
def recover_super88_plan8_final_v4(
        body: CampaignPlan8FinalRecoveryV4In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Complete or read back exactly the six existing plan-8 draft records."""
    from app.services import campaign_plan8_final_recovery_v3_service

    result = campaign_plan8_final_recovery_v3_service.recover_plan8_final_v3(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        recovery_version=body.recovery_version,
        mode=body.mode,
        confirmation=body.confirmation,
        target_scope_sha256=body.target_scope_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_final_v3_request_not_allowed",
                "plan8_final_v3_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v4/claim-verification")
def verify_super88_plan8_final_v4_claim(
        body: CampaignPlan8FinalRecoveryV4ClaimVerifyIn,
        db: Session = Depends(get_db),
        _: ServicePrincipal = Depends(
            require_web_agent_plan8_v4_claim_verifier)):
    """Let Web-Agent prove ERP's durable one-shot claim without a credential leak."""
    from app.services import campaign_plan8_final_recovery_v3_service

    result = campaign_plan8_final_recovery_v3_service.verify_plan8_final_v3_claim(
        db, attempt_id=body.attempt_id, workflow_key=body.workflow_key,
        plan_id=body.plan_id, operation=body.operation,
        scope_sha256=body.scope_sha256,
        inspect_scope_sha256=body.inspect_scope_sha256,
        reservation_token_sha256=body.reservation_token_sha256,
    )
    if not result.get("ok"):
        raise HTTPException(409, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v5")
def recover_super88_plan8_final_v5(
        body: CampaignPlan8FinalRecoveryV5In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Independent V5 after V4 proved it stopped before platform write."""
    from app.services import campaign_plan8_final_recovery_v5_service

    result = campaign_plan8_final_recovery_v5_service.recover_plan8_final_v3(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        recovery_version=body.recovery_version,
        mode=body.mode,
        confirmation=body.confirmation,
        target_scope_sha256=body.target_scope_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_final_v3_request_not_allowed",
                "plan8_final_v3_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v5/claim-verification")
def verify_super88_plan8_final_v5_claim(
        body: CampaignPlan8FinalRecoveryV5ClaimVerifyIn,
        db: Session = Depends(get_db),
        _: ServicePrincipal = Depends(
            require_web_agent_plan8_v5_claim_verifier)):
    """Let V5 Web-Agent prove the exact durable ERP claim."""
    from app.services import campaign_plan8_final_recovery_v5_service

    result = campaign_plan8_final_recovery_v5_service.verify_plan8_final_v3_claim(
        db, attempt_id=body.attempt_id, workflow_key=body.workflow_key,
        plan_id=body.plan_id, operation=body.operation,
        scope_sha256=body.scope_sha256,
        inspect_scope_sha256=body.inspect_scope_sha256,
        reservation_token_sha256=body.reservation_token_sha256,
    )
    if not result.get("ok"):
        raise HTTPException(409, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v6")
def recover_super88_plan8_final_v6(
        body: CampaignPlan8FinalRecoveryV6In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Use the exact activity add-product path after V5 made no submission."""
    from app.services import campaign_plan8_final_recovery_v6_service

    result = campaign_plan8_final_recovery_v6_service.recover_plan8_final_v6(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        recovery_version=body.recovery_version,
        mode=body.mode,
        confirmation=body.confirmation,
        target_scope_sha256=body.target_scope_sha256,
    )
    if not result.get("ok"):
        error = str(result.get("error") or "")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_final_v6_request_not_allowed",
                "plan8_final_v6_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v6/claim-verification")
def verify_super88_plan8_final_v6_claim(
        body: CampaignPlan8FinalRecoveryV6ClaimVerifyIn,
        db: Session = Depends(get_db),
        _: ServicePrincipal = Depends(
            require_web_agent_plan8_v6_claim_verifier)):
    """Let V6 Web-Agent prove the exact durable ERP claim."""
    from app.services import campaign_plan8_final_recovery_v6_service

    result = campaign_plan8_final_recovery_v6_service.verify_plan8_final_v6_claim(
        db, attempt_id=body.attempt_id, workflow_key=body.workflow_key,
        plan_id=body.plan_id, operation=body.operation,
        scope_sha256=body.scope_sha256,
        inspect_scope_sha256=body.inspect_scope_sha256,
        reservation_token_sha256=body.reservation_token_sha256,
    )
    if not result.get("ok"):
        raise HTTPException(409, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v7")
def recover_super88_plan8_final_v7(
        body: CampaignPlan8FinalRecoveryV7In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Create a separate 8-SKU discount activity and finish six drafts."""
    from app.services import campaign_plan8_final_recovery_v7_service

    result = campaign_plan8_final_recovery_v7_service.recover_plan8_final_v7(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        recovery_version=body.recovery_version,
        mode=body.mode,
        confirmation=body.confirmation,
        target_scope_sha256=body.target_scope_sha256,
    )
    if not result.get("ok"):
        error = str(result.get("error") or "")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_final_v7_request_not_allowed",
                "plan8_final_v7_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v7/claim-verification")
def verify_super88_plan8_final_v7_claim(
        body: CampaignPlan8FinalRecoveryV7ClaimVerifyIn,
        db: Session = Depends(get_db),
        _: ServicePrincipal = Depends(
            require_web_agent_plan8_v7_claim_verifier)):
    """Let V7 Web-Agent prove the exact durable ERP claim."""
    from app.services import campaign_plan8_final_recovery_v7_service

    result = campaign_plan8_final_recovery_v7_service.verify_plan8_final_v7_claim(
        db, attempt_id=body.attempt_id, workflow_key=body.workflow_key,
        plan_id=body.plan_id, operation=body.operation,
        scope_sha256=body.scope_sha256,
        inspect_scope_sha256=body.inspect_scope_sha256,
        reservation_token_sha256=body.reservation_token_sha256,
    )
    if not result.get("ok"):
        raise HTTPException(409, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v8")
def recover_super88_plan8_final_v8(
        body: CampaignPlan8FinalRecoveryV8In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Publish the six bound drafts, then supplement eight existing discounts."""
    from app.services import campaign_plan8_final_recovery_v8_service

    result = campaign_plan8_final_recovery_v8_service.recover_plan8_final_v8(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_status=body.expected_status,
        recovery_version=body.recovery_version,
        mode=body.mode,
        confirmation=body.confirmation,
        target_scope_sha256=body.target_scope_sha256,
    )
    if not result.get("ok"):
        error = str(result.get("error") or "")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "plan8_final_v8_request_not_allowed",
                "plan8_final_v8_identity_not_allowed"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super88-plan8-final-v8/claim-verification")
def verify_super88_plan8_final_v8_claim(
        body: CampaignPlan8FinalRecoveryV8ClaimVerifyIn,
        db: Session = Depends(get_db),
        _: ServicePrincipal = Depends(
            require_web_agent_plan8_v8_claim_verifier)):
    """Let V8 Web-Agent prove the exact durable ERP claim."""
    from app.services import campaign_plan8_final_recovery_v8_service

    result = campaign_plan8_final_recovery_v8_service.verify_plan8_final_v8_claim(
        db, attempt_id=body.attempt_id, workflow_key=body.workflow_key,
        plan_id=body.plan_id, operation=body.operation,
        scope_sha256=body.scope_sha256,
        inspect_scope_sha256=body.inspect_scope_sha256,
        reservation_token_sha256=body.reservation_token_sha256,
    )
    if not result.get("ok"):
        raise HTTPException(409, detail=result)
    return result


@router.post("/verify-super-reduce-plan7-post-submit")
def verify_super_reduce_plan7_post_submit(
        body: CampaignPlan7PostSubmitVerifyIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Fresh export-only verification after the one submitted plan-7 attempt."""
    from app.services import campaign_resume_service

    result = campaign_resume_service.verify_super_reduce_plan7_post_submit(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_attempt_id=body.expected_attempt_id,
        expected_scope_sha256=body.expected_scope_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "post_submit_verify_request_not_allowed",
                "post_submit_attempt_not_eligible",
                "post_submit_verify_state_not_allowed",
                "post_submit_verify_scope_drift"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/audit-super-reduce-plan7-discount")
def audit_super_reduce_plan7_discount(
        body: CampaignPlan7DiscountAuditIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Read and persist exact plan-7 single-discount state without writing it."""
    from app.services import campaign_discount_audit_service

    result = campaign_discount_audit_service.audit_plan7_single_discount(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_scope_sha256=body.expected_scope_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/update-super-reduce-plan7-discount-times")
def update_super_reduce_plan7_discount_times(
        body: CampaignPlan7DiscountTimeUpdateIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """CAS-read and update only the three approved activity time windows once."""
    from app.services import campaign_plan7_time_update_service

    result = campaign_plan7_time_update_service.update_plan7_single_discount_times(
        db, request_payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
            "plan7_discount_time_update_request_not_allowed",
            "plan7_discount_time_update_plan_identity_drift",
            "plan7_discount_time_update_scope_drift",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super-reduce-plan7-discount-times")
def recover_super_reduce_plan7_discount_times(
        body: CampaignPlan7DiscountTimeRecoveryIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """One recovery bound to the exact write-free failed attempt."""
    from app.services import campaign_plan7_time_update_service

    result = campaign_plan7_time_update_service.recover_plan7_single_discount_times(
        db, request_payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "plan7_discount_time_recovery_attempt_not_found",
        } else 409
        if error in {
            "plan7_discount_time_update_request_not_allowed",
            "plan7_discount_time_update_plan_identity_drift",
            "plan7_discount_time_update_scope_drift",
            "plan7_discount_time_recovery_fields_invalid",
            "plan7_discount_time_recovery_attempt_mismatch",
            "plan7_discount_time_recovery_receipts_invalid",
            "plan7_discount_time_recovery_receipts_mismatch",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super-reduce-plan7-discount-times-v2")
def recover_super_reduce_plan7_discount_times_v2(
        body: CampaignPlan7DiscountTimeRecoveryV2In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """One V2 recovery after the original and V1 calls both stopped write-free."""
    from app.services import campaign_plan7_time_update_service

    result = (
        campaign_plan7_time_update_service
        .recover_plan7_single_discount_times_v2(
            db, request_payload=body.model_dump())
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "plan7_discount_time_recovery_attempt_not_found",
        } else 409
        if error in {
            "plan7_discount_time_update_request_not_allowed",
            "plan7_discount_time_update_plan_identity_drift",
            "plan7_discount_time_update_scope_drift",
            "plan7_discount_time_recovery_fields_invalid",
            "plan7_discount_time_recovery_attempt_mismatch",
            "plan7_discount_time_recovery_receipts_invalid",
            "plan7_discount_time_recovery_receipts_mismatch",
            "plan7_discount_time_recovery_v2_fields_invalid",
            "plan7_discount_time_recovery_v2_receipt_invalid",
            "plan7_discount_time_recovery_v2_receipt_mismatch",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/closeout-super-reduce-plan7-discount-times-v3")
def closeout_super_reduce_plan7_discount_times_v3(
        body: CampaignPlan7DiscountTimeReadbackV3In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Read back only; never open an editor or write to the platform."""
    from app.services import campaign_plan7_time_update_service

    result = (
        campaign_plan7_time_update_service
        .closeout_plan7_single_discount_times_v3(
            db, request_payload=body.model_dump())
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "plan7_discount_time_readback_attempt_not_found",
        } else 409
        if error in {
            "plan7_discount_time_readback_v3_fields_invalid",
            "plan7_discount_time_readback_v3_confirmed_ids_invalid",
            "plan7_discount_time_readback_v3_identity_mismatch",
            "plan7_discount_time_readback_attempt_evidence_mismatch",
            "plan7_discount_time_readback_plan_identity_drift",
            "plan7_discount_time_readback_scope_drift",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/correct-super-reduce-plan7-discount")
def correct_super_reduce_plan7_discount(
        body: CampaignPlan7DiscountCorrectionIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Submit only snapshot-1's four missing discount rows, once.

    The service performs an exact read-before-write, one four-row import and an
    exact readback.  It cannot call campaign signup, modify prices, rotate SKUs,
    or touch the 384 rows already present in the original activity.
    """
    from app.services import campaign_discount_correction_service

    result = campaign_discount_correction_service.correct_plan7_single_discount(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_snapshot_id=body.expected_snapshot_id,
        expected_snapshot_artifact_sha256=(
            body.expected_snapshot_artifact_sha256),
        expected_missing_scope_sha256=body.expected_missing_scope_sha256,
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
                "discount_correction_request_not_allowed",
                "discount_correction_plan_identity_not_allowed",
                "discount_correction_erp_price_scope_drift",
                "discount_correction_final_price_math_mismatch"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/supplement-super-reduce-plan7-single-discount")
def supplement_super_reduce_plan7_single_discount(
        body: CampaignPlan7DiscountSupplementIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Append exactly four reviewed SKUs to one existing plan-7 activity once."""
    from app.services import campaign_plan7_discount_supplement_service

    result = (
        campaign_plan7_discount_supplement_service
        .execute_plan7_discount_supplement(
            db, request_payload=body.model_dump())
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
            "plan7_discount_supplement_request_not_allowed",
            "plan7_discount_supplement_plan_identity_drift",
            "plan7_discount_supplement_erp_scope_drift",
            "plan7_discount_supplement_xlsx_scope_drift",
            "plan7_discount_supplement_activity_scope_not_exact",
            "plan7_discount_supplement_activity_identity_drift",
            "plan7_discount_supplement_readback_scope_drift",
            "plan7_discount_supplement_platform_state_not_allowed",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/correct-super-reduce-plan7-small-promo")
def correct_super_reduce_plan7_small_promo(
        body: CampaignPlan7SmallPromoCorrectionIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Correct only the fixed 20 deductions to ERP small-promo targets."""
    from app.services import campaign_plan7_small_promo_correction_service

    result = campaign_plan7_small_promo_correction_service.execute(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
            "plan7_small_promo_request_not_allowed",
            "plan7_small_promo_plan_identity_drift",
            "plan7_small_promo_source_snapshot_drift",
            "plan7_small_promo_source_snapshot_scope_drift",
            "plan7_small_promo_erp_manifest_drift",
            "plan7_small_promo_forbidden_accessory_missing",
            "plan7_small_promo_forbidden_accessory_identity_drift",
            "plan7_small_promo_activity_identity_drift",
            "plan7_small_promo_platform_state_drift",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/correct-warehouse-product-sku-price")
def correct_warehouse_product_sku_price(
        body: CampaignWarehouseProductPriceCorrectionIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Return the permanent user-rule exclusion for this warehouse SKU."""
    from app.services import campaign_warehouse_product_price_correction_service

    result = campaign_warehouse_product_price_correction_service.execute(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
            "warehouse_product_price_request_not_allowed",
            "warehouse_product_price_plan_identity_drift",
            "warehouse_product_price_readback_drift",
            "warehouse_product_price_readback_boundary_violation",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/correct-five-price-issues")
def correct_five_price_issues(
        body: CampaignFivePriceCorrectionIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Run one no-retry exact phase; never touch the zero-sales item."""
    from app.services import campaign_five_price_correction_service

    result = campaign_five_price_correction_service.execute(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
            "five_price_request_not_allowed",
            "five_price_plan_identity_drift",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-five-price-single-discount")
def recover_five_price_single_discount(
        body: CampaignFivePriceRecoveryIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Recover only the fixed single-discount timeout; preserve its attempt."""
    from app.services import campaign_five_price_correction_service

    result = campaign_five_price_correction_service.recover_single_discount(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "five_price_recovery_failed_attempt_not_found",
        } else 409
        if error in {
            "five_price_recovery_request_not_allowed",
            "five_price_plan_identity_drift",
            "five_price_recovery_failed_attempt_not_write_free",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-five-price-super-reduce")
def recover_five_price_super_reduce(
        body: CampaignFivePriceRecoveryIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Recover only the fixed write-free price-input locator failure."""
    from app.services import campaign_five_price_correction_service

    result = campaign_five_price_correction_service.recover_super_reduce(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "five_price_super_recovery_failed_attempt_not_found",
        } else 409
        if error in {
            "five_price_super_recovery_request_not_allowed",
            "five_price_plan_identity_drift",
            "five_price_super_recovery_failed_attempt_not_write_free",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-five-price-super-reduce-v2")
def recover_five_price_super_reduce_v2(
        body: CampaignFivePriceSuperRecoveryV2In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Recover after four fresh readbacks prove the prior form was not saved."""
    from app.services import campaign_five_price_correction_service

    result = campaign_five_price_correction_service.recover_super_reduce_v2(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "five_price_super_recovery_v2_failed_attempt_not_found",
        } else 409
        if error in {
            "five_price_super_recovery_v2_request_not_allowed",
            "five_price_plan_identity_drift",
            "five_price_super_recovery_v2_failed_attempt_not_exact",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-five-price-super-reduce-v3")
def recover_five_price_super_reduce_v3(
        body: CampaignFivePriceSuperRecoveryV3In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Recover only after fresh reads prove one new and three old items."""
    from app.services import campaign_five_price_correction_service

    result = campaign_five_price_correction_service.recover_super_reduce_v3(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "five_price_super_recovery_v3_failed_attempt_not_found",
        } else 409
        if error in {
            "five_price_super_recovery_v3_request_not_allowed",
            "five_price_plan_identity_drift",
            "five_price_super_recovery_v3_failed_attempt_not_exact",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-five-price-super-reduce-v4")
def recover_five_price_super_reduce_v4(
        body: CampaignFivePriceSuperRecoveryV3In,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Recover after the V3 stale-editor failure proved zero new writes."""
    from app.services import campaign_five_price_correction_service

    result = campaign_five_price_correction_service.recover_super_reduce_v4(
        db, payload=body.model_dump())
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error in {
            "workflow_not_found",
            "five_price_super_recovery_v4_failed_attempt_not_found",
        } else 409
        if error in {
            "five_price_super_recovery_v4_request_not_allowed",
            "five_price_plan_identity_drift",
            "five_price_super_recovery_v4_failed_attempt_not_exact",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/recover-super-reduce-plan7-discount-sku-identity")
def recover_super_reduce_plan7_discount_sku_identity(
        body: CampaignPlan7DiscountIdentityRecoveryIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Repair four stale Taobao SKU identities, then permit one new import.

    The immutable official product export must prove a one-to-one merchant-code
    match.  This route cannot change any price, rotate specifications, touch
    plan 8 or retry an attempted platform write.
    """
    from app.services import campaign_discount_identity_recovery_service

    result = (
        campaign_discount_identity_recovery_service
        .recover_plan7_single_discount_identity(
            db,
            workflow_key=_validate_workflow_key(body.workflow_key),
            expected_plan_id=body.plan_id,
            expected_old_attempt_id=body.expected_old_attempt_id,
            official_product_export_sha256=(
                body.official_product_export_sha256),
            official_product_export_b64=body.official_product_export_b64,
            expected_new_scope_sha256=body.expected_new_scope_sha256,
        )
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
            "sku_identity_recovery_request_not_allowed",
            "official_product_export_base64_invalid",
            "official_product_export_size_not_allowed",
            "official_product_export_sha256_mismatch",
            "official_product_export_scope_mismatch",
            "official_product_export_old_identity_mismatch",
            "official_product_export_current_identity_mismatch",
            "official_product_export_placeholder_mismatch",
            "official_product_export_merchant_scope_mismatch",
            "sku_identity_recovery_plan_identity_not_allowed",
            "sku_identity_recovery_prior_evidence_mismatch",
            "sku_identity_recovery_full_scope_drift",
            "sku_identity_recovery_price_scope_drift",
            "sku_identity_recovery_final_price_math_mismatch",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/verify-super-reduce-plan7-discount-sku-identity-readback")
def verify_super_reduce_plan7_discount_sku_identity_readback(
        body: CampaignPlan7DiscountIdentityReadbackIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Read four submitted rows once and close the existing attempt.

    This route has no upload/submit path and cannot create a new business
    attempt.  Any navigation or row difference is terminal and remains visible.
    """
    from app.services import campaign_discount_identity_recovery_service

    result = (
        campaign_discount_identity_recovery_service
        .verify_plan7_identity_recovery_readback(
            db,
            workflow_key=_validate_workflow_key(body.workflow_key),
            expected_plan_id=body.plan_id,
            expected_attempt_id=body.expected_attempt_id,
            expected_terminal_evidence_request_id=(
                body.expected_terminal_evidence_request_id),
            expected_scope_sha256=body.expected_scope_sha256,
        )
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 422 if error in {
            "sku_identity_readback_request_not_allowed",
            "sku_identity_readback_plan_identity_not_allowed",
            "sku_identity_readback_attempt_receipt_mismatch",
            "sku_identity_readback_identity_receipt_mismatch",
            "sku_identity_recovery_full_scope_drift",
            "sku_identity_recovery_price_scope_drift",
            "sku_identity_recovery_final_price_math_mismatch",
        } else 409
        raise HTTPException(code, detail=result)
    return result


@router.post("/execute-super-reduce-plan7-remaining")
def execute_super_reduce_plan7_remaining(
        body: CampaignPlan7RemainingSignupIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Submit only the reviewed remaining plan-7 manifest, once.

    This path cannot touch single-item discounts, plan 8, real-SKU prices,
    daily prices, SKU identities, withdrawals, pauses or removals.  It claims
    every platform batch before upload and requires a terminal plus exact
    per-SKU export readback before advancing to another batch.
    """
    from app.services import campaign_plan7_remaining_signup_service

    result = (
        campaign_plan7_remaining_signup_service
        .execute_plan7_remaining_signup(
            db,
            workflow_key=_validate_workflow_key(body.workflow_key),
            expected_plan_id=body.plan_id,
            expected_status=body.expected_status,
            expected_item_scope_sha256=(
                body.expected_item_scope_sha256),
            recovery_incident_id=body.recovery_incident_id,
        )
    )
    if not result.get("ok"):
        error = result.get("error")
        code = 404 if error == "workflow_not_found" else 409
        if error in {
            "remaining_signup_request_not_allowed",
            "remaining_signup_authorized_scope_constant_invalid",
            "remaining_signup_plan_identity_not_allowed",
            "remaining_signup_official_scope_drift",
            "remaining_signup_placeholder_scope_drift",
            "remaining_signup_official_exclusion_missing",
            "remaining_signup_real_price_hold_scope_drift",
            "remaining_signup_safe_scope_incomplete",
            "remaining_signup_forbidden_item_in_manifest",
            "remaining_signup_no_sales_scope_drift",
            "remaining_signup_whole_item_exclusion_missing",
            "remaining_signup_price_source_guard_failed",
            "remaining_signup_manifest_count_drift",
            "remaining_signup_preflight_blocked",
        }:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/audit-super-reduce-plan7-partial-signup")
def audit_super_reduce_plan7_partial_signup(
        body: CampaignPlan7PartialSignupAuditIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Download official failure proof and enrolled rows; never upload/publish."""
    from app.services import campaign_plan7_remaining_signup_service

    result = campaign_plan7_remaining_signup_service.audit_plan7_partial_signup(
        db,
        workflow_key=_validate_workflow_key(body.workflow_key),
        expected_plan_id=body.plan_id,
        expected_attempt_id=body.expected_attempt_id,
        expected_manifest_sha256=body.expected_manifest_sha256,
    )
    if not result.get("ok"):
        code = 404 if result.get("error") == "workflow_not_found" else 409
        if result.get("error") in {
                "partial_signup_audit_request_not_allowed",
                "partial_signup_attempt_receipt_mismatch",
                "partial_signup_manifest_drift",
                "partial_signup_feedback_scope_mismatch",
                "partial_signup_final_price_math_blocked"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.post("/publish-super-reduce-plan7-existing-drafts")
def publish_super_reduce_plan7_existing_drafts(
        body: CampaignPlan7DraftPublishIn,
        db: Session = Depends(get_db),
        _: User | ServicePrincipal = Depends(require_campaign_prepare_principal)):
    """Retired unsafe route kept only for a structured fail-closed result."""
    from app.services import campaign_plan7_remaining_signup_service

    result = (
        campaign_plan7_remaining_signup_service
        .publish_plan7_existing_drafts(
            db,
            workflow_key=_validate_workflow_key(body.workflow_key),
            expected_plan_id=body.plan_id,
            expected_attempt_id=body.expected_attempt_id,
            expected_snapshot_id=body.expected_snapshot_id,
            expected_scope_sha256=body.expected_scope_sha256,
        )
    )
    if not result.get("ok"):
        code = 404 if result.get("error") == "workflow_not_found" else 409
        if result.get("error") in {
                "draft_publish_request_not_allowed",
                "draft_publish_plan_identity_not_allowed",
                "draft_publish_snapshot_identity_mismatch",
                "draft_publish_snapshot_scope_mismatch",
                "draft_publish_price_fingerprint_mismatch",
                "draft_publish_removed_paused_is_enrolled_state",
                "draft_publish_global_paused_scope_mismatch",
                "draft_publish_pre_read_price_or_sku_mismatch"}:
            code = 422
        raise HTTPException(code, detail=result)
    return result


@router.get("/item-exclusions")
def list_item_exclusions(db: Session = Depends(get_db),
                         _: User = Depends(get_current_user)):
    """List explicit and authoritative all-placeholder whole-item exclusions."""
    return {"items": list(campaign_service.campaign_item_exclusions(db).values())}


@router.post("/item-exclusions")
def set_item_exclusion(
        body: CampaignItemExclusionIn, db: Session = Depends(get_db),
        _: User = Depends(require_role("admin", "operator"))):
    from app.services import campaign_item_exclusion_service

    try:
        row = campaign_item_exclusion_service.upsert(
            db, item_id=body.taobao_item_id, reason=body.reason)
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    return {"ok": True, "item": {
        "taobao_item_id": row.taobao_item_id,
        "reason": row.reason,
        "source": row.source,
        "active": row.active,
    }}


@router.delete("/item-exclusions/{item_id}")
def remove_item_exclusion(
        item_id: str, db: Session = Depends(get_db),
        _: User = Depends(require_role("admin"))):
    from app.services import campaign_item_exclusion_service

    return {"ok": campaign_item_exclusion_service.deactivate(db, item_id)}


@router.get("/no-sales-group")
def no_sales_group(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """动销检查与分组 (spec §四.1): 近60天淘宝订单聚合 + no_sales 登记表同步。"""
    return campaign_service.group_by_sales(db)


@router.post("/no-sales-group/notify")
@router.post("/no-sales-group/push-feishu")     # 前端契约别名 (同一 handler)
def notify_no_sales_group(db: Session = Depends(get_db),
                          _: User = Depends(require_role("admin", "operator"))):
    """无动销名单一键飞书推送给运营促成交 (spec §四.2a)。"""
    from app.services import notify_service
    grouping = campaign_service.group_by_sales(db)
    items = grouping["无动销"]
    if not items:
        return {"ok": True, "sent": False, "message": "当前没有无动销商品"}
    names = grouping.get("item_names", {})
    lines = [f"📉 无动销商品名单（近{grouping['days']}天零成交, 共 {len(items)} 个）:"]
    lines += [f"- {iid} {names.get(iid, '')}" for iid in items]
    lines.append("请运营跟进促成交; 卖出1单后系统会提示转正(撤 nosales 立减→报名大促)。")
    result = notify_service.broadcast_text(db, "\n".join(lines), title="无动销名单", level="warning")
    return {"ok": True, "sent": True, "count": len(items), "notify_result": result}


@router.get("/no-sales-group/export.xlsx")
def export_no_sales_group(db: Session = Depends(get_db),
                          _: User = Depends(get_current_user)):
    """无动销名单一键导出 xlsx (spec §四.2a): 产品名/产品编码/淘宝商品ID/近60天单量/建议动作。"""
    import io
    import openpyxl
    from fastapi.responses import Response
    rows = campaign_service.no_sales_export_rows(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "无动销名单"
    headers = ("产品名", "产品编码", "淘宝商品ID", "近60天单量", "建议动作")
    for ci, h in enumerate(headers, start=1):
        ws.cell(1, ci, h)
    for ri, r in enumerate(rows, start=2):
        ws.cell(ri, 1, r["product_name"])
        ws.cell(ri, 2, r["product_codes"])
        ws.cell(ri, 3, r["taobao_item_id"]).number_format = "@"
        ws.cell(ri, 4, r["sales_60d"])
        ws.cell(ri, 5, r["action"])
    out = io.BytesIO()
    wb.save(out)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="no_sales_group.xlsx"'})


@router.get("/{plan_id}")
def get_plan(plan_id: int, db: Session = Depends(get_db),
             _: User = Depends(get_current_user)):
    return _plan_out(_get_plan(db, plan_id))


@router.put("/{plan_id}")
def update_plan(plan_id: int, body: CampaignPlanUpdate, db: Session = Depends(get_db),
                _: User = Depends(require_role("admin", "operator"))):
    plan = _get_plan(db, plan_id)
    data = body.model_dump(exclude_unset=True)
    if "price_protection_days" in data and data["price_protection_days"] is None:
        raise HTTPException(422, "价保冷静期不能为空")
    _validate_price_protection(
        data.get("price_protection_days"),
        data.get("price_protection_rule_url"))
    if "campaign_type" in data:
        plan.tier = _validate_type_and_window(
            data["campaign_type"], data.get("start_at", plan.start_at),
            data.get("end_at", plan.end_at))
    for k, v in data.items():
        setattr(plan, k, v)
    if "price_protection_rule_url" in data:
        plan.price_protection_rule_url = (
            str(data["price_protection_rule_url"]).strip()
            if data["price_protection_rule_url"] else None)
        plan.price_protection_confirmed_at = (
            datetime.now() if plan.price_protection_rule_url else None)
    db.commit()
    return _plan_out(plan)


@router.post("/{plan_id}/price-protection/remind")
def remind_price_protection_rule(
        plan_id: int,
        db: Session = Depends(get_db),
        _: User = Depends(require_role("admin", "operator"))):
    from app.services import campaign_price_protection_service

    plan = _get_plan(db, plan_id)
    return campaign_price_protection_service.notify_rule_link_needed(db, plan, force=True)


@router.delete("/{plan_id}")
def delete_plan(plan_id: int, db: Session = Depends(get_db),
                _: User = Depends(require_role("admin"))):
    plan = _get_plan(db, plan_id)
    db.delete(plan)
    db.commit()
    return {"ok": True}


@router.post("/{plan_id}/precheck")
def precheck(plan_id: int, db: Session = Depends(get_db),
             _: User = Depends(require_role("admin", "operator"))):
    """R0~R17 预检 (spec §三)。只有全无 error 才进入 precheck 状态。"""
    plan = _get_plan(db, plan_id)
    checks = campaign_service.preflight(db, plan)
    has_error = any(c["level"] == "error" for c in checks)
    if plan.status == "draft" and not has_error:
        plan.status = "precheck"
        db.commit()
    from app.services import campaign_policy_service
    return {
        "plan": _plan_out(plan),
        "checks": checks,
        "has_error": has_error,
        "policy": campaign_policy_service.public_policy() if checks[0]["rule"] == "R0"
                  and checks[0]["level"] == "pass" else None,
    }


@router.get("/{plan_id}/rows")
def preview_rows(plan_id: int, kind: str = Query("signup"), db: Session = Depends(get_db),
                 _: User = Depends(get_current_user)):
    """行预览: kind=signup(报名行) / discount(单品立减行)。"""
    plan = _get_plan(db, plan_id)
    if kind == "signup":
        rows, stats = campaign_service.build_signup_rows(db, plan)
    elif kind == "discount":
        rows, stats = campaign_service.build_discount_rows(db, plan)
    else:
        raise HTTPException(422, "kind 必须是 signup 或 discount")
    return {"rows": rows, "stats": stats}


@router.post("/{plan_id}/push-discount")
def push_discount(plan_id: int, phase: str = Query("stage"), db: Session = Depends(get_db),
                  _: User = Depends(require_role("admin"))):
    """推单品立减。phase=stage 挂文件停在提交前; commit ★不可逆★ (R12, 用户确认后才调)。"""
    if phase not in ("stage", "commit"):
        raise HTTPException(422, "phase 必须是 stage 或 commit")
    plan = _get_plan(db, plan_id)
    return campaign_service.push_discount(db, plan, phase=phase)


@router.post("/{plan_id}/push-signup")
def push_signup(plan_id: int, db: Session = Depends(get_db),
                _: User = Depends(require_role("admin"))):
    """Direct signup is disabled; only the scheduled campaign program may submit."""
    _get_plan(db, plan_id)
    raise HTTPException(
        409,
        "活动报名只由 ERP 自动报名程序执行；页面或 AI 直推已禁用。错误先报告并等待用户决定。",
    )


@router.get("/{plan_id}/single-discount-error-file.xlsx")
def download_single_discount_error_file(
        plan_id: int, activity_id: str = Query(..., min_length=1),
        db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """Read-only proxy for the original QianNiu single-discount error file."""
    from app.services import web_agent_service

    _get_plan(db, plan_id)
    result = web_agent_service.single_discount_error_file(db, activity_id)
    if not result.get("ok"):
        raise HTTPException(422, detail=result)
    return _xlsx_download_response(
        result["xlsx_bytes"], result.get("filename") or "单品立减错误文件.xlsx",
        metadata={
            "Activity-Id": result.get("activity_id"),
            "Operation-Time": result.get("operation_time"),
            "Success-Count": result.get("success_count"),
            "Failed-Count": result.get("failed_count"),
            "Detail-Rows": result.get("detail_rows"),
            "Empty-Detail": result.get("empty_detail"),
        },
    )


@router.get("/{plan_id}/operation-feedback.xlsx")
def download_campaign_operation_feedback(
        plan_id: int, db: Session = Depends(get_db),
        _: User = Depends(get_current_user)):
    """Download the original latest operation-feedback workbook, read-only."""
    from app.services import web_agent_service

    plan = _get_plan(db, plan_id)
    campaign_id, united_activity_id = campaign_service.plan_campaign_ids(plan)
    if campaign_id and united_activity_id:
        result = web_agent_service.campaign_feedback(
            db,
            plan.qn_campaign_title or plan.name,
            campaign_id=campaign_id,
            united_activity_id=united_activity_id,
        )
    elif plan.campaign_type == "super_reduce":
        # The long-running Super Reduce page is itself an exact activity entry;
        # this read-only fallback is needed for older plans created before the
        # immutable IDs were persisted in their remark.
        result = web_agent_service.super_reduce_feedback(db)
    else:
        raise HTTPException(422, "计划缺少千牛 campaignId / unitedActivityId，无法安全下载")
    if not result.get("ok"):
        raise HTTPException(422, detail=result)
    if not result.get("xlsx_bytes"):
        raise HTTPException(422, "Web-Agent 已解析反馈，但未返回平台原始文件；请更新本机 Web-Agent")
    feedback = result.get("feedback") or {}
    return _xlsx_download_response(
        result["xlsx_bytes"], result.get("filename") or "活动报名操作反馈.xlsx",
        metadata={
            "Failed-Rows": len(feedback.get("failed") or []),
            "Failure-Groups": len(feedback.get("by_reason") or []),
        },
    )


@router.post("/{plan_id}/repair-super-reduce-activation")
def repair_super_reduce_activation(
        plan_id: int, payload: SuperReduceRepairIn,
        db: Session = Depends(get_db), _: User = Depends(require_role("admin"))):
    """User-authorized correction through the ERP campaign program."""
    plan = _get_plan(db, plan_id)
    result = campaign_service.repair_super_reduce_early_activation(
        db,
        plan,
        payload.item_ids,
        phase=payload.phase,
        execution_source="campaign_automation_repair",
    )
    if not result.get("ok"):
        raise HTTPException(422, detail=result)
    return result


@router.post("/{plan_id}/recon")
async def recon(plan_id: int,
                activity_file: Optional[UploadFile] = File(None),
                discount_file: Optional[UploadFile] = File(None),
                product_file: Optional[UploadFile] = File(None),
                db: Session = Depends(get_db),
                _: User = Depends(require_role("admin", "operator"))):
    """核对 (spec §四.6, §五「/recon (自动+手动上传兜底)」):
    - 带文件 → 手动上传兜底 (三种导出任传);
    - 不带文件 → 自动: WA campaign_export_items 按活动标题导出「活动商品导出」再比对。"""
    from app.services import campaign_recon_service
    plan = _get_plan(db, plan_id)
    activity_bytes = await activity_file.read() if activity_file else None
    discount_bytes = await discount_file.read() if discount_file else None
    product_bytes = await product_file.read() if product_file else None
    source = "manual"
    if not any((activity_bytes, discount_bytes, product_bytes)):
        from app.services import web_agent_service
        identity = campaign_service.campaign_identity(plan)
        if not identity["ok"]:
            raise HTTPException(422, detail={
                "error": "campaign_identity_incomplete",
                "missing": identity["missing"],
            })
        exp = web_agent_service.campaign_export_items(
            db,
            identity["campaign_title"],
            campaign_id=identity["campaign_id"],
            united_activity_id=identity["united_activity_id"],
            sign_record_id=identity["sign_record_id"],
            campaign_phase=identity["campaign_phase"],
            campaign_start=identity["campaign_start"],
            campaign_end=identity["campaign_end"],
            official_rate=identity["official_rate"],
            platform_activity_mode=identity["platform_activity_mode"],
            platform_active_until=identity["platform_active_until"],
        )
        if not exp.get("ok"):
            err = exp.get("error") or exp.get("message") or "未知原因"
            raise HTTPException(422, f"WA 自动导出失败（{err}）; 请自己去千牛导出后走手动上传兜底")
        activity_bytes, source = exp["xlsx_bytes"], "auto"
    result = campaign_recon_service.reconcile(
        db, plan, activity_bytes=activity_bytes, discount_bytes=discount_bytes,
        product_bytes=product_bytes, source=source)
    if not result.get("ok"):
        raise HTTPException(422, result.get("error", "核对失败"))
    return result


@router.get("/{plan_id}/recon-reports")
def recon_reports(plan_id: int, db: Session = Depends(get_db),
                  _: User = Depends(get_current_user)):
    reports = db.execute(
        select(CampaignReconReport).where(CampaignReconReport.plan_id == plan_id)
        .order_by(CampaignReconReport.id.desc())).scalars().all()
    return {"items": [{"id": r.id, "source": r.source, "summary": r.summary,
                       "alarm_count": r.alarm_count,
                       "created_at": r.created_at.isoformat() if r.created_at else None}
                      for r in reports]}

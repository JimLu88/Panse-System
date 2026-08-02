import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict

from app.dependencies import require_role
from app.models.auth import User
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.finance import (
    AccountBalance, AlipayFlow, FactoryReconciliation, LogisticsBill,
    PackingBill, RefillRecord, WanshifuBill,
)
from app.models.marketing import PromotionFlow
from app.services import (
    alipay_backfill_service,
    alipay_flow_router_service,
    alipay_import,
    balance_service,
    bill_import_service,
    cash_flow_service,
    email_import_service,
    factory_reconciliation_service,
    import_storage,
    reconciliation_service,
    smart_matching_service,
)

router = APIRouter(prefix="/api/finance", tags=["finance"])


# -------- Alipay flows --------

class AlipayFlowOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    account: str
    transaction_no: str
    transaction_time: Optional[datetime]
    transaction_type: Optional[str]
    counterparty: Optional[str]
    amount: Decimal
    related_order_no: Optional[str]
    balance: Optional[Decimal]
    reconciliation_status: str
    reconciliation_type: Optional[str]
    remark: Optional[str]


@router.get("/alipay-flows", response_model=list[AlipayFlowOut])
def list_alipay(
    account: Optional[str] = None,
    recon_type: Optional[str] = None,
    q: Optional[str] = None,           # 搜索: 备注/对方/流水号(应对爱群号海量+无日期流水翻不到)
    only_unclassified: bool = False,   # 只看未分类(reconciliation_type 为空)
    limit: int = Query(100, le=500),
    offset: int = 0,
    db: Session = Depends(get_db),
):
    stmt = select(AlipayFlow)
    if account:
        stmt = stmt.where(AlipayFlow.account == account)
    if recon_type:
        stmt = stmt.where(AlipayFlow.reconciliation_type == recon_type)
    if only_unclassified:
        stmt = stmt.where(AlipayFlow.reconciliation_type.is_(None))
    if q and q.strip():
        from sqlalchemy import or_
        like = f"%{q.strip()}%"
        stmt = stmt.where(or_(
            AlipayFlow.remark.ilike(like),
            AlipayFlow.counterparty.ilike(like),
            AlipayFlow.transaction_no.ilike(like),
        ))
    stmt = stmt.order_by(AlipayFlow.transaction_time.desc().nulls_last(), AlipayFlow.id.desc()).limit(limit).offset(offset)
    return db.execute(stmt).scalars().all()


class AlipayFlowEditIn(BaseModel):
    reconciliation_type: Optional[str] = None   # 给值=改核销类型(空串=清空); None=不改
    amount: Optional[Decimal] = None            # 给值=改金额(爱群号丢符号: 支出应为负)
    transaction_time: Optional[datetime] = None  # 给值=改交易时间(补无日期流水)


@router.patch("/alipay-flows/{flow_id}", response_model=AlipayFlowOut)
def edit_alipay_flow(
    flow_id: int,
    payload: AlipayFlowEditIn,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """手动修正流水(系统入口, 替代手填库): 改 核销类型 / 金额 / 交易时间。
    用于爱群号等脏流水的纠正(丢符号→改金额符号、无日期→补日期、错分类→改类型)。
    每处改动记字段修改档案(可回溯), 改后触发对账重算。"""
    f = db.get(AlipayFlow, flow_id)
    if not f:
        raise HTTPException(404, "流水不存在")
    from app.services import field_change_service
    actor = ("手动修正:" + (getattr(user, "username", "") or "")).rstrip(": ")
    label = (f.counterparty or f.remark or f.transaction_no or "")[:40]
    changed = False
    if payload.reconciliation_type is not None:
        new_rt = payload.reconciliation_type.strip() or None
        if new_rt != f.reconciliation_type:
            field_change_service.record(db, table="alipay_flows", pk=str(flow_id), field="reconciliation_type",
                old=f.reconciliation_type, new=new_rt, actor=actor, source="manual",
                row_label=label, field_label="核销类型")
            f.reconciliation_type = new_rt
            changed = True
    if payload.amount is not None and payload.amount != f.amount:
        field_change_service.record(db, table="alipay_flows", pk=str(flow_id), field="amount",
            old=str(f.amount), new=str(payload.amount), actor=actor, source="manual",
            row_label=label, field_label="金额")
        f.amount = payload.amount
        changed = True
    if payload.transaction_time is not None and payload.transaction_time != f.transaction_time:
        field_change_service.record(db, table="alipay_flows", pk=str(flow_id), field="transaction_time",
            old=str(f.transaction_time), new=str(payload.transaction_time), actor=actor, source="manual",
            row_label=label, field_label="交易时间")
        f.transaction_time = payload.transaction_time
        changed = True
    db.commit()
    if changed:
        try:
            from app.services import realtime_sync_service
            realtime_sync_service.trigger(f"edit:alipay:{flow_id}")
        except Exception:  # noqa: BLE001
            pass
    db.refresh(f)
    return f


class AlipayImportResult(BaseModel):
    inserted: int
    skipped_duplicate: int
    skipped_invalid: int
    errors: list[str]
    auto_tagged: dict[str, int] = {}
    auto_untouched: int = 0
    archived_file_id: Optional[int] = None   # 归档原文件 id (导入档案可回溯)
    duplicate_upload: bool = False            # 同一文件曾上传过


@router.post("/alipay-flows/import-csv", response_model=AlipayImportResult)
async def import_alipay(
    account: str = Query(..., description="账户名 (企业号 / 私账 / 主力号 …)"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    raw = await file.read()
    arch = import_storage.archive(
        db, content=raw, original_name=file.filename or f"alipay-{account}.csv",
        kind="alipay", source="web",
    )
    from app.services import tabular
    text = tabular.to_csv_text(raw, file.filename)
    r = alipay_import.import_alipay_csv(db, text, account=account)
    # plan §12.4: 导入完跑一次智能核销
    matched = smart_matching_service.run(db, account=account)
    import_storage.update_summary(db, arch.file.id, {
        "inserted": r.inserted, "skipped_duplicate": r.skipped_duplicate,
        "skipped_invalid": r.skipped_invalid, "account": account,
    })
    db.commit()
    # 实时同步: 支付宝流水导入后立即重算对账(流水喂全部对账规则)
    from app.services import realtime_sync_service
    realtime_sync_service.trigger(f"import:alipay:{account}")
    return AlipayImportResult(
        inserted=r.inserted,
        skipped_duplicate=r.skipped_duplicate,
        skipped_invalid=r.skipped_invalid,
        errors=r.errors,
        auto_tagged=matched.tagged,
        auto_untouched=matched.untouched,
        archived_file_id=arch.file.id,
        duplicate_upload=arch.is_duplicate,
    )


class SmartMatchResult(BaseModel):
    total_scanned: int
    tagged: dict[str, int]
    untouched: int


class FactoryReconRebuildOut(BaseModel):
    periods: int
    created: int
    updated: int


@router.post("/factory-reconciliation/rebuild", response_model=FactoryReconRebuildOut)
def rebuild_factory_reconciliation(
    factory_name: Optional[str] = None, db: Session = Depends(get_db),
):
    """按自然月把工厂下单表汇总成工厂对账记录 (本期下单金额/账单金额/实付/差异)。

    导入工厂下单表后触发; 幂等, 可反复跑。
    """
    r = factory_reconciliation_service.rebuild_all_periods(db, factory_name=factory_name)
    db.commit()
    return FactoryReconRebuildOut(periods=r.periods, created=r.created, updated=r.updated)


class FactoryReconOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    factory_name: str
    period_start: Optional[date]
    period_end: Optional[date]
    order_amount: Optional[Decimal]
    bill_amount: Optional[Decimal]
    paid_amount: Optional[Decimal]
    diff_amount: Decimal
    status: str          # balanced/underpaid/overpaid/unpaid
    diff_reason: Optional[str]
    reconciled_at: Optional[date]
    alipay_flow_no: Optional[str]


@router.get("/factory-reconciliation", response_model=list[FactoryReconOut])
def list_factory_reconciliation(
    status: Optional[str] = Query(None, description="按对账状态过滤: balanced/underpaid/overpaid/unpaid"),
    unbalanced_only: bool = Query(False, description="只看对不上的 (未付清/超付)"),
    db: Session = Depends(get_db),
):
    """工厂对账逐周期清单 (含 账单/实付/差异/对账状态)。

    前端可据此显示「未付清/超付」红黄标; unbalanced_only=True 时只返回对不上的周期,
    与异常中心 factory_recon_unbalanced 一致。
    """
    stmt = select(FactoryReconciliation).order_by(
        FactoryReconciliation.factory_name, FactoryReconciliation.period_end.desc(),
    )
    if status:
        stmt = stmt.where(FactoryReconciliation.status == status)
    elif unbalanced_only:
        stmt = stmt.where(FactoryReconciliation.status.in_(["underpaid", "overpaid"]))
    return db.execute(stmt).scalars().all()


class AlipayRouteResult(BaseModel):
    aftersales_created: int = 0
    promotion_filled: int
    daily_filled: int
    outsourcing_filled: int
    purchases_created: int
    factory_flipped: int


@router.post("/alipay-flows/route", response_model=AlipayRouteResult)
def route_alipay_flows(rerun_classify: bool = True, db: Session = Depends(get_db)):
    """支付宝流水自动归类回填: 售后先建→推广/日常/外包补流水号, 未分类建采购, 工厂翻已付款。

    rerun_classify=True 时先跑一遍 smart_matching 打 reconciliation_type。
    """
    if rerun_classify:
        smart_matching_service.run(db)
    r = alipay_flow_router_service.run_all(db)
    db.commit()
    return AlipayRouteResult(
        aftersales_created=r.aftersales_created,
        promotion_filled=r.promotion_filled, daily_filled=r.daily_filled,
        outsourcing_filled=r.outsourcing_filled, purchases_created=r.purchases_created,
        factory_flipped=r.factory_flipped,
    )


@router.post("/alipay-flows/amount-match", response_model=dict)
def amount_match_alipay_flows(
    days_window: int = 3,
    include_deprecated: bool = False,
    db: Session = Depends(get_db),
):
    """对账细化(4规则): 流水里没订单号时, 用 金额唯一锁定 / 金额+日期 / 多对一·一对多 /
    账户语义(只用收入流水、默认排除爱群佳宝) 补匹配。保守: 只填空、仅唯一命中。"""
    from app.services import alipay_amount_match_service
    r = alipay_amount_match_service.match(
        db, days_window=days_window, include_deprecated=include_deprecated
    )
    db.commit()
    return {
        "candidate_orders": r.candidate_orders,
        "candidate_flows": r.candidate_flows,
        "matched": r.matched,
        "linked_flow_no": r.linked_flow_no,
        "by_rule": r.by_rule,
        "samples": r.samples,
    }


@router.get("/alipay-flows/sign-audit", response_model=dict)
def alipay_sign_audit(
    account: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """只读: 审计「历史符号脏数据」—— 金额为正但疑似支出的流水(尤其交易类型=支出却为正)。

    按账户汇总 strong(高置信错符号)/weak(文本疑似)+样例, 供人工确认。不改任何数据;
    修正(原地翻符号)需先据此清单人工确认后另行执行。"""
    from app.services import alipay_sign_audit_service
    return alipay_sign_audit_service.audit_wrong_sign(db, account=account)


@router.post("/alipay-flows/detect-refunds", response_model=dict)
def detect_refunds(db: Session = Depends(get_db)):
    """识别退款对: 同关联订单号下金额相等、方向相反的两条流水标为 refund_in/refund_out。

    打标后退款对不再被归为「重复流水」异常; 支出侧被 route 识别为售后。
    """
    from app.services import flow_refund_service
    n = flow_refund_service.detect_refunds(db)
    db.commit()
    return {"pairs_found": n, "message": f"识别到 {n} 对退款流水并已打标"}


@router.post("/factory-reconciliation/match-alipay", response_model=dict)
def match_factory_alipay(
    factory_name: Optional[str] = None, db: Session = Depends(get_db),
):
    """按工厂对账账单金额合计, 从支付宝支出流水中找等额一笔回填工厂订单 alipay_flow_no。

    命中后 flip_factory_payment 可自动翻「已付款」。
    factory_name 留空时处理全部工厂。
    """
    n = factory_reconciliation_service.match_factory_alipay_by_bill_amount(
        db, factory_name=factory_name,
    )
    db.commit()
    return {"matched_periods": n, "message": f"按账单金额匹配到 {n} 个工厂付款周期"}


@router.post("/smart-match/rerun", response_model=SmartMatchResult)
def rerun_smart_match(account: Optional[str] = None, db: Session = Depends(get_db)):
    """重新核销: 对未打标流水按 关联订单号→工厂名→关键字 三段重新打 reconciliation_type.

    导入订单/工厂下单表后再触发, 可把之前没识别的货款/客户回款挂上。
    """
    r = smart_matching_service.run(db, account=account)
    db.commit()
    return SmartMatchResult(total_scanned=r.total_scanned, tagged=r.tagged, untouched=r.untouched)


# -------- 支付宝流水 → 订单 反向匹配回填 --------

class AlipayBackfillOut(BaseModel):
    total_flows: int
    matched_orders: int
    filled_flow_no: int
    ambiguous: int
    unmatched: int
    by_rule: dict[str, int]
    samples: list[dict]
    flows_marked_matched: int = 0   # 已标记 reconciliation_status='matched' 的流水数
    filled_dates: int = 0           # 从流水号前缀回填交易时间的笔数 (企业号缺日期兜底)


def _backfill_to_out(r: "alipay_backfill_service.BackfillResult") -> AlipayBackfillOut:
    return AlipayBackfillOut(
        total_flows=r.total_flows,
        matched_orders=r.matched_orders,
        filled_flow_no=r.filled_flow_no,
        ambiguous=r.ambiguous,
        unmatched=r.unmatched,
        by_rule=r.by_rule,
        samples=r.samples,
        flows_marked_matched=r.flows_marked_matched,
    )


@router.get("/order-flow-match/analyze", response_model=AlipayBackfillOut)
def analyze_order_flow_match(account: Optional[str] = None, db: Session = Depends(get_db)):
    """只读: 自动从流水里找订单号规律, 预览能匹配多少订单 (不写库).

    返回各规律命中数 (exact/strip_prefix/tail_index) + 样本, 看清匹配逻辑。
    """
    return _backfill_to_out(alipay_backfill_service.analyze(db, account=account))


@router.post("/order-flow-match/backfill", response_model=AlipayBackfillOut)
def backfill_order_flow_no(
    account: Optional[str] = None, only_missing: bool = True, db: Session = Depends(get_db),
):
    """落库: 把匹配到的支付宝流水号回填到订单 Order.alipay_flow_no。

    自己找规律 (T200P 前缀 / 备注内订单号 / 尾号倒排), 歧义流水跳过交人工。
    顺带把交易时间为空的流水 (企业号常见) 从流水号前缀回填日期。
    """
    dates = alipay_backfill_service.backfill_transaction_time(db, account=account)
    r = alipay_backfill_service.backfill(db, account=account, only_missing=only_missing)
    db.commit()
    out = _backfill_to_out(r)
    out.filled_dates = dates.get("filled", 0)
    return out


# -------- 万师傅 / 物流账单导入 (安装费 / 物流费对账数据源) --------

class BillImportResult(BaseModel):
    inserted: int
    updated_existing: int = 0
    skipped_invalid: int
    errors: list[str]
    skipped_duplicate: int = 0
    unmapped_columns: list[str] = []


async def _read_csv(file: UploadFile) -> str:
    """读上传文件为 CSV 文本 —— CSV 直接解码, Excel(xlsx) 自动转 CSV 文本。"""
    from app.services import tabular
    raw = await file.read()
    return tabular.to_csv_text(raw, file.filename)


@router.post("/wanshifu-bills/import-csv", response_model=BillImportResult)
async def import_wanshifu(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入万师傅安装账单 CSV → 供安装费对账 (rule=install_fee) 当应付口径。"""
    text = await _read_csv(file)
    r = bill_import_service.import_wanshifu_csv(db, text)
    db.commit()
    return BillImportResult(inserted=r.inserted, updated_existing=r.updated_existing,
                            skipped_invalid=r.skipped_invalid, errors=r.errors,
                            skipped_duplicate=r.skipped_duplicate, unmapped_columns=r.unmapped_columns)


@router.get("/refill-records/settings")
def get_refill_settings(db: Session = Depends(get_db)):
    """补单导入费用设置 (设置按钮用)。"""
    from app.services import settings_service as _ss
    raw = _ss.get(db, "refill_freight_default", env_fallback=False)
    try:
        freight = float(raw) if raw else 5.0
    except (TypeError, ValueError):
        freight = 5.0
    return {"freight_default": freight}


@router.put("/refill-records/settings")
def put_refill_settings(
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """调整补单费用缺省 (留痕修改档案)。body: {"freight_default": 5}"""
    from app.services import field_change_service, settings_service as _ss
    try:
        freight = float(payload.get("freight_default"))
    except (TypeError, ValueError):
        raise HTTPException(400, "freight_default 必须是数字")
    if freight < 0 or freight > 1000:
        raise HTTPException(400, "快递费缺省须在 0~1000 之间")
    old = _ss.get(db, "refill_freight_default", env_fallback=False)
    _ss.set_value(db, "refill_freight_default", str(freight))
    field_change_service.record(
        db, table="system_settings", pk="refill_freight_default",
        field="refill_freight_default", old=old, new=str(freight),
        actor=getattr(user, "username", None),
        row_label="补单费用设置", field_label="补单快递费缺省",
    )
    db.commit()
    return {"ok": True, "freight_default": freight}


@router.post("/refill-records/import-xlsx")
async def import_refill_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """补单简表 xlsx 直接导入 (订单号/旺旺/本金/佣金/店铺; 日期取文件名如 5.31)。

    原文件按 类别(refill)+日期 归档进 工具→导入档案。
    """
    import io

    import openpyxl

    from app.services import bill_import_service as bis
    from app.services import import_storage
    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(400, "无法解析 xlsx 文件")
    refill_date = bis.refill_date_from_filename(file.filename or "")
    arch = import_storage.archive(
        db, content=data, original_name=file.filename or "补单表.xlsx",
        kind="refill", source="web", on_date=refill_date,
    )
    # 快递费缺省 ¥5 (用户拍板), settings refill_freight_default 可调
    from decimal import Decimal as _D

    from app.services import settings_service as _ss
    try:
        freight = _D(str(_ss.get(db, "refill_freight_default", env_fallback=False) or "5"))
    except Exception:
        freight = _D("5")
    rep = bis.import_refill_simple_xlsx(db, wb, refill_date=refill_date,
                                        freight_default=freight)
    if rep.errors:
        db.commit()   # 归档保留, 方便排查
        raise HTTPException(400, "; ".join(rep.errors))
    import_storage.update_summary(db, arch.file.id, {
        "inserted": rep.inserted, "skipped_duplicate": rep.skipped_duplicate,
        "skipped_invalid": rep.skipped_invalid, "note": f"补单日期 {refill_date}",
    })
    db.commit()
    # 实时同步: 补单导入后立即重算成本/对账, 待办异常即时跟上 (后台, 不阻塞)
    from app.services import realtime_sync_service
    realtime_sync_service.trigger("import:refill-xlsx")
    return {"inserted": rep.inserted, "skipped_duplicate": rep.skipped_duplicate,
            "skipped_invalid": rep.skipped_invalid, "refill_date": str(refill_date),
            "archived": not arch.is_duplicate}


# -------- 万师傅安装订单档案 (38列订单导出, 2026-06 起默认格式) --------

@router.post("/wanshifu-orders/import")
async def import_wanshifu_orders(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入万师傅「订单导出」xlsx (含客户信息), 导入后自动跑订单配对。"""
    import io

    import openpyxl

    from app.services import wanshifu_order_service as wsf
    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(400, "无法解析 xlsx 文件")
    rep = wsf.import_workbook(db, wb)
    if rep.errors:
        raise HTTPException(400, "; ".join(rep.errors))
    counts = wsf.match_orders(db)
    # 售后自动化① (用户拍板): 配对完立即把交易成功单建成售后条目
    from app.services import aftersales_auto_service
    aftersales_n = aftersales_auto_service.create_from_wanshifu(db)
    db.commit()
    from app.services import realtime_sync_service
    realtime_sync_service.trigger("import:wanshifu-orders")
    return {"parsed": rep.parsed, "inserted": rep.inserted, "updated": rep.updated,
            "verified_matched": rep.verified_matched,
            "match": counts, "aftersales_created": aftersales_n}


@router.get("/wanshifu-orders")
def list_wanshifu_orders(
    only_unmatched: bool = Query(False),
    limit: int = Query(500, le=2000),
    db: Session = Depends(get_db),
):
    from app.models.finance import WanshifuOrder
    from app.services.wanshifu_order_service import METHOD_CN
    stmt = select(WanshifuOrder)
    if only_unmatched:
        stmt = stmt.where(WanshifuOrder.matched_order_no.is_(None))
    stmt = stmt.order_by(WanshifuOrder.created_time.desc().nulls_last()).limit(limit)
    return [{
        "id": w.id, "wsf_order_no": w.wsf_order_no, "status": w.status,
        "service_type": w.service_type,
        "product_category": w.product_category, "customer_name": w.customer_name,
        "customer_phone": w.customer_phone,
        "region": "".join(x for x in (w.province, w.city, w.district) if x),
        "address": w.address,
        "net_amount": float(w.net_amount) if w.net_amount is not None else None,
        "service_fee": float(w.service_fee) if w.service_fee is not None else None,
        "created_time": w.created_time.isoformat() if w.created_time else None,
        "matched_order_no": w.matched_order_no,
        "match_method": METHOD_CN.get(w.match_method or "", w.match_method),
        "match_note": w.match_note,
    } for w in db.execute(stmt).scalars().all()]


class WanshifuMatchUpdate(BaseModel):
    matched_order_no: Optional[str] = None   # 人工指定淘宝订单号; 空 = 清除匹配


@router.patch("/wanshifu-orders/{wsf_id}")
def update_wanshifu_match(
    wsf_id: int,
    payload: WanshifuMatchUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """人工逐行确认/修正万师傅订单匹配 (方案2: 系统内校对落库, 留痕)。
    设 match_method=manual (最高权威, 重配/重导都不覆盖); 订单库暂无此单仅批注提示不拦。"""
    from app.models.finance import WanshifuOrder
    from app.models.order import Order
    w = db.get(WanshifuOrder, wsf_id)
    if not w:
        raise HTTPException(404, "万师傅订单不存在")
    new_no = (payload.matched_order_no or "").strip() or None
    old_no = w.matched_order_no
    if new_no == old_no:
        return {"ok": True, "matched_order_no": old_no, "unchanged": True}
    in_lib = bool(new_no) and db.execute(
        select(Order.order_no).where(Order.order_no == new_no).limit(1)).first() is not None
    w.matched_order_no = new_no
    w.match_method = "manual" if new_no else None
    w.match_note = (None if (not new_no or in_lib)
                    else "人工校对; 订单库暂无此单(早期单/待导入)")
    from app.services import field_change_service
    field_change_service.record(
        db, table="wanshifu_orders", pk=str(w.id),
        field="matched_order_no", old=old_no, new=new_no,
        actor=getattr(user, "username", None),
        row_label=f"万师傅单 {w.wsf_order_no}", field_label="匹配订单号(人工)",
    )
    db.commit()
    from app.services import realtime_sync_service
    realtime_sync_service.trigger("wanshifu:manual-match")
    return {"ok": True, "matched_order_no": new_no, "in_lib": in_lib}


@router.post("/wanshifu-orders/match")
def match_wanshifu_orders(
    rematch_all: bool = Query(False, description="true=全部重配 (人工指定的除外)"),
    db: Session = Depends(get_db),
):
    from app.services import wanshifu_order_service as wsf
    counts = wsf.match_orders(db, only_unmatched=not rematch_all)
    from app.services import aftersales_auto_service
    counts["aftersales_created"] = aftersales_auto_service.create_from_wanshifu(db)
    db.commit()
    return counts


@router.get("/wanshifu-orders/export-annotated")
def export_wanshifu_annotated(db: Session = Depends(get_db)):
    """档案+匹配批注 xlsx — 「对不上的告诉我」的回执表。"""
    import io

    from fastapi.responses import StreamingResponse

    from app.services import wanshifu_order_service as wsf
    wb = wsf.build_annotated_workbook(db)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=wanshifu_orders_annotated.xlsx"},
    )


@router.post("/logistics-bills/import-csv", response_model=BillImportResult)
async def import_logistics(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入物流公司月结账单 CSV → 供物流费对账 (rule=logistics_fee) 当应付口径。"""
    text = await _read_csv(file)
    r = bill_import_service.import_logistics_csv(db, text)
    db.commit()
    return BillImportResult(inserted=r.inserted, updated_existing=r.updated_existing,
                            skipped_invalid=r.skipped_invalid, errors=r.errors,
                            skipped_duplicate=r.skipped_duplicate, unmapped_columns=r.unmapped_columns)


@router.post("/logistics-bills/import-xlsx", response_model=BillImportResult)
async def import_logistics_xlsx_ep(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """物流账单 xlsx 统一导入 (用户 2026-06-15): 按文件名自动识别承运商。
       - 德邦 (文件名含「德邦」): 逐运单 + 实收运费/运费。
       - 壹米滴答 (李爱群月结): 月结总额取自文件名 (如「…账单 14540元」) → 1 条汇总。
    原文件归档进 工具→导入档案。"""
    import io

    import openpyxl

    from app.services import import_storage
    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "无法解析 xlsx 文件")
    try:
        _y, _m, _ = bill_import_service.parse_logi_bill_filename(file.filename or "")
        import datetime as _dt
        _on = _dt.date(_y, _m, 1) if _m else _dt.date.today()
        import_storage.archive(db, content=data, original_name=file.filename or "logistics.xlsx",
                               kind="logistics", source="web", on_date=_on)
    except Exception:
        pass  # 归档失败不阻断导入
    r = bill_import_service.import_logistics_xlsx(db, wb, source_name=file.filename or "")
    db.commit()
    return BillImportResult(inserted=r.inserted, updated_existing=r.updated_existing,
                            skipped_invalid=r.skipped_invalid, errors=r.errors,
                            skipped_duplicate=r.skipped_duplicate, unmapped_columns=r.unmapped_columns)


# -------- 推广记录 / 补单对账 / 账户余额 CSV 导入 --------

@router.post("/promotion-flows/import-csv", response_model=BillImportResult)
async def import_promotion_flows(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入推广记录 CSV (直通车/万相台充值+支出; 列名自动识别)。"""
    text = await _read_csv(file)
    r = bill_import_service.import_promotion_flows_csv(db, text)
    db.commit()
    from app.services import realtime_sync_service
    realtime_sync_service.trigger("import:promotion")
    return BillImportResult(inserted=r.inserted, skipped_invalid=r.skipped_invalid, errors=r.errors,
                            skipped_duplicate=r.skipped_duplicate, unmapped_columns=r.unmapped_columns)


@router.post("/refill-records/import-csv", response_model=BillImportResult)
async def import_refill_records(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入补单对账 CSV (订单号必填)。"""
    text = await _read_csv(file)
    r = bill_import_service.import_refill_records_csv(db, text)
    db.commit()
    from app.services import realtime_sync_service
    realtime_sync_service.trigger("import:refill-csv")
    return BillImportResult(inserted=r.inserted, skipped_invalid=r.skipped_invalid, errors=r.errors,
                            skipped_duplicate=r.skipped_duplicate, unmapped_columns=r.unmapped_columns)


@router.post("/accounts/import-csv", response_model=BillImportResult)
async def import_account_balances(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入账户余额 CSV (同账户同月 upsert; 账户名+年+月必填)。"""
    text = await _read_csv(file)
    r = bill_import_service.import_account_balances_csv(db, text)
    db.commit()
    from app.services import realtime_sync_service
    realtime_sync_service.trigger("import:account-balance")
    return BillImportResult(inserted=r.inserted, skipped_invalid=r.skipped_invalid, errors=r.errors,
                            skipped_duplicate=r.skipped_duplicate, unmapped_columns=r.unmapped_columns)


# -------- Account balances --------

class AccountBalanceOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    account_name: str
    account_no: Optional[str] = None
    period_year: int
    period_month: int
    as_of_date: Optional[date] = None
    opening_balance: Decimal
    income: Decimal
    expense: Decimal
    closing_balance: Decimal
    remark: Optional[str] = None


@router.get("/accounts", response_model=list[AccountBalanceOut])
def list_balances(
    account_name: Optional[str] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
):
    stmt = select(AccountBalance)
    if account_name:
        stmt = stmt.where(AccountBalance.account_name == account_name)
    if year:
        stmt = stmt.where(AccountBalance.period_year == year)
    stmt = stmt.order_by(AccountBalance.account_name, AccountBalance.period_year.desc(), AccountBalance.period_month.desc())
    return db.execute(stmt).scalars().all()


class BalanceUpsertIn(BaseModel):
    """手动录入/更新一条账户余额快照 (账户名+年+月 upsert)。"""
    account_name: str
    account_no: Optional[str] = None
    period_year: int
    period_month: int
    as_of_date: Optional[date] = None
    opening_balance: Optional[Decimal] = None
    income: Optional[Decimal] = None
    expense: Optional[Decimal] = None
    closing_balance: Decimal
    remark: Optional[str] = None


@router.post("/accounts", response_model=AccountBalanceOut)
def upsert_balance(payload: BalanceUpsertIn, db: Session = Depends(get_db)):
    """手动录入/更新账户余额快照。余额多是某天手填的, as_of_date 存「统计日期」(新鲜度据此算)。"""
    existing = db.execute(
        select(AccountBalance).where(
            AccountBalance.account_name == payload.account_name,
            AccountBalance.period_year == payload.period_year,
            AccountBalance.period_month == payload.period_month,
        )
    ).scalar_one_or_none()
    row = existing or AccountBalance(
        account_name=payload.account_name,
        period_year=payload.period_year,
        period_month=payload.period_month,
    )
    if not existing:
        db.add(row)
    if payload.account_no is not None:
        row.account_no = payload.account_no
    if payload.as_of_date is not None:
        row.as_of_date = payload.as_of_date
    if payload.opening_balance is not None:
        row.opening_balance = payload.opening_balance
    if payload.income is not None:
        row.income = payload.income
    if payload.expense is not None:
        row.expense = payload.expense
    row.closing_balance = payload.closing_balance
    if payload.remark is not None:
        row.remark = payload.remark
    db.commit()
    db.refresh(row)
    return row


@router.delete("/accounts/{balance_id}")
def delete_balance(balance_id: int, db: Session = Depends(get_db)):
    """删除一条账户余额快照 (录错时清理)。"""
    row = db.get(AccountBalance, balance_id)
    if not row:
        raise HTTPException(404, "余额记录不存在")
    db.delete(row)
    db.commit()
    return {"deleted": balance_id}


@router.delete("/accounts/by-name/all")
def delete_account_all(
    account_name: str = Query(..., min_length=1, description="要整账删除的账户名"),
    password: str = Body(..., embed=True, description="登录密码二次确认"),
    user: User = Depends(require_role("admin", "operator")),
    db: Session = Depends(get_db),
):
    """删除某账户的**全部**余额快照 (清理重复/废弃账户)。

    高危操作 (用户拍板 2026-06-17): 需登录密码二次确认 (前端还要再输一遍账户名)。
    用例: 旧手填的『企业号』与自动抓取的『支付宝-企业账号』是同一账号, 把旧的整账删掉。
    """
    from app.services import auth_service
    if not user.password_hash or not auth_service.verify_password(password, user.password_hash):
        raise HTTPException(403, "密码不正确, 整账删除已取消")
    name = account_name.strip()
    rows = db.execute(
        select(AccountBalance).where(AccountBalance.account_name == name)
    ).scalars().all()
    if not rows:
        raise HTTPException(404, f"账户『{name}』无余额记录")
    for r in rows:
        db.delete(r)
    db.commit()
    return {"deleted_account": name, "deleted_rows": len(rows)}


class RecomputeIn(BaseModel):
    account_name: str
    year: int
    month: int
    opening_balance: Optional[Decimal] = None


@router.post("/accounts/recompute", response_model=AccountBalanceOut)
def recompute(payload: RecomputeIn, db: Session = Depends(get_db)):
    try:
        row = balance_service.recompute_month(
            db,
            account=payload.account_name,
            year=payload.year,
            month=payload.month,
            opening_balance=payload.opening_balance,
        )
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    db.commit()
    db.refresh(row)
    return row


@router.get("/balances/derive-opening")
def derive_opening(
    account: str,
    target_date: date,
    db: Session = Depends(get_db),
):
    """Plan F10: 期初余额倒推 — 最近快照 − 区间Σ流水 → target_date 当日期初 (+ 缺流水提示)。"""
    return balance_service.derive_opening_balance(db, account=account, target_date=target_date)


# -------- Reconciliation --------

class DiffOut(BaseModel):
    key: str
    expected: Optional[Decimal]
    actual: Optional[Decimal]
    diff: Optional[Decimal]
    severity: str
    message: str
    related_records: list[str] = []  # 该差异涉及的明细单号(支付宝流水号/工厂单号/订单号), 供核对


class ReconciliationOut(BaseModel):
    rule: str
    total_diffs: int
    ok_count: int
    warning_count: int
    error_count: int
    unresolved_count: int = 0
    diffs: list[DiffOut]


def _to_out(r: reconciliation_service.ReconciliationResult) -> ReconciliationOut:
    return ReconciliationOut(
        rule=r.rule,
        total_diffs=r.total_diffs,
        ok_count=r.ok_count,
        warning_count=r.warning_count,
        error_count=r.error_count,
        unresolved_count=r.unresolved_count,
        diffs=[DiffOut(**d.__dict__) for d in r.diffs],
    )


class WriteoffIn(BaseModel):
    rule: str
    key: str
    reason: str


@router.post("/reconciliation/writeoff")
def writeoff_reconciliation_diff(
    payload: WriteoffIn,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """人工做平: 把某条对账差异永久豁免 (cron 不再翻出来), 带原因留痕进修改档案。"""
    from app.models.exception import DataException
    from app.services import exception_service, field_change_service
    pk = f"{payload.rule}:{payload.key}"
    exc = db.query(DataException).filter(
        DataException.source_table == "reconciliation",
        DataException.source_pk == pk,
        DataException.exception_type == "reconciliation_diff",
        DataException.status.in_(("open", "ignored")),
    ).first()
    if exc is None:
        exc = exception_service.record(
            db, source_table="reconciliation", source_pk=pk,
            exception_type="reconciliation_diff", severity="info",
            description=f"[人工做平] {payload.rule} {payload.key}",
            suggestion_action="manual_writeoff",
            context={"rule": payload.rule, "key": payload.key},
        )
    from datetime import datetime, timezone
    exc.status = "ignored"
    exc.resolved_by = getattr(user, "username", None)
    exc.resolved_at = datetime.now(timezone.utc).isoformat()
    exc.description = (exc.description or "") + f" | 做平原因: {payload.reason}"
    # 做平动作进修改档案 (谁/何时/为何把账做平, 可回溯)
    field_change_service.record(
        db, table="reconciliation", pk=pk, field="writeoff",
        old="差异未处理", new=f"已做平: {payload.reason}",
        actor=getattr(user, "username", None),
        row_label=f"对账[{payload.rule}] {payload.key}",
        field_label="人工做平",
    )
    db.commit()
    return {"ok": True, "rule": payload.rule, "key": payload.key}


@router.get("/reconciliation/writeoffs")
def list_reconciliation_writeoffs(db: Session = Depends(get_db)):
    """已做平的差异键 {rule: [key...]} + 做平金额小计 (对账建议 8: 做平多了会掩盖系统性问题)。"""
    from decimal import Decimal, InvalidOperation

    from app.models.exception import DataException
    rows = db.query(DataException).filter(
        DataException.source_table == "reconciliation",
        DataException.exception_type == "reconciliation_diff",
        DataException.status == "ignored",
    ).all()
    out: dict[str, list[str]] = {}
    totals: dict[str, float] = {}
    grand = Decimal("0")
    for r in rows:
        pk = r.source_pk or ""
        if ":" not in pk:
            continue
        rule, key = pk.split(":", 1)
        out.setdefault(rule, []).append(key)
        try:
            amt = abs(Decimal(str((r.context or {}).get("diff", "0"))))
        except (InvalidOperation, ValueError, TypeError):
            amt = Decimal("0")
        totals[rule] = totals.get(rule, 0.0) + float(amt)
        grand += amt
    # 异常池最近同步时间 (run_all 写异常时记录) — 前端显示"差异截至何时"
    from app.services import settings_service
    synced_at = settings_service.get(db, "recon_exceptions_synced_at", env_fallback=False)
    return {"keys": out, "totals": totals, "grand_total": float(grand),
            "count": len(rows), "synced_at": synced_at}


@router.get("/reconciliation/factory-aliases")
def get_factory_aliases(db: Session = Depends(get_db)):
    """工厂别名映射 {别名: 标准名} — 货款对账两侧名称先归一再比对。"""
    from app.services.reconciliation_service import _factory_aliases
    return {"aliases": _factory_aliases(db)}


@router.put("/reconciliation/factory-aliases")
def put_factory_aliases(
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """保存别名映射 (留痕修改档案)。body: {"aliases": {"**晶": "XX家具厂"}}"""
    import json

    from app.services import field_change_service, settings_service
    aliases = payload.get("aliases")
    if not isinstance(aliases, dict):
        raise HTTPException(400, "aliases 必须是 {别名: 标准名} 对象")
    from app.services.reconciliation_service import _factory_aliases
    old = _factory_aliases(db)
    settings_service.set_value(db, "factory_aliases", json.dumps(aliases, ensure_ascii=False))
    field_change_service.record(
        db, table="system_settings", pk="factory_aliases", field="factory_aliases",
        old=json.dumps(old, ensure_ascii=False), new=json.dumps(aliases, ensure_ascii=False),
        actor=getattr(user, "username", None), row_label="工厂别名映射",
        field_label="工厂别名",
    )
    db.commit()
    return {"ok": True, "count": len(aliases)}


# ── #1 账户角色注册表 + #2 内部登记中心 (用户 2026-06-29): 账户角色/内部主体后台可配 ──

@router.get("/reconciliation/account-roles")
def get_account_roles(db: Session = Depends(get_db)):
    """账户角色映射 {账户名: [角色]} — 角色: revenue/boguan_payment/internal/ledger_exempt。"""
    from app.services import account_registry_service
    return {"account_roles": account_registry_service._load_roles(db),
            "all_accounts": list(account_registry_service.all_known_accounts(db))}


@router.put("/reconciliation/account-roles")
def put_account_roles(
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """保存账户角色 (留痕 + 清缓存)。body: {"account_roles": {"爱群号": ["boguan_payment"]}}"""
    import json

    from app.services import account_registry_service, field_change_service, settings_service
    roles = payload.get("account_roles")
    if not isinstance(roles, dict):
        raise HTTPException(400, "account_roles 必须是 {账户名: [角色]} 对象")
    old = account_registry_service._load_roles(db)
    settings_service.set_value(db, "account_roles", json.dumps(roles, ensure_ascii=False))
    field_change_service.record(
        db, table="system_settings", pk="account_roles", field="account_roles",
        old=json.dumps(old, ensure_ascii=False), new=json.dumps(roles, ensure_ascii=False),
        actor=getattr(user, "username", None), row_label="账户角色注册表", field_label="账户角色",
    )
    db.commit()
    account_registry_service.invalidate()
    return {"ok": True, "count": len(roles)}


@router.get("/reconciliation/internal-entities")
def get_internal_entities(db: Session = Depends(get_db)):
    """内部主体登记 {owners, proxies, extra} + 当前生效关键词(种子∪配置)。"""
    from app.services import internal_accounts, settings_service
    import json as _json
    raw = settings_service.get(db, "internal_entities", env_fallback=False)
    cfg = _json.loads(raw) if raw else {}
    return {"internal_entities": cfg if isinstance(cfg, dict) else {},
            "effective_keywords": list(internal_accounts.internal_counterparty_keywords(db))}


@router.put("/reconciliation/internal-entities")
def put_internal_entities(
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """保存内部主体登记 (留痕)。body: {"internal_entities": {"owners":[...],"proxies":[...],"extra":[...]}}"""
    import json

    from app.services import field_change_service, settings_service
    ents = payload.get("internal_entities")
    if not isinstance(ents, dict):
        raise HTTPException(400, "internal_entities 必须是 {owners, proxies, extra} 对象")
    raw = settings_service.get(db, "internal_entities", env_fallback=False)
    settings_service.set_value(db, "internal_entities", json.dumps(ents, ensure_ascii=False))
    field_change_service.record(
        db, table="system_settings", pk="internal_entities", field="internal_entities",
        old=raw or "{}", new=json.dumps(ents, ensure_ascii=False),
        actor=getattr(user, "username", None), row_label="内部主体登记", field_label="内部主体",
    )
    db.commit()
    return {"ok": True}


@router.post("/reconciliation/match-expense-flows")
def match_expense_flows_api(
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "operator")),
):
    """经营支出自动配流水: 给缺流水号的 日常/外包/品牌 记录按金额+日期窗口配支付宝支出。

    只在唯一命中时回填 (多候选留人工), 回填记修改档案。
    """
    from app.services.expense_flow_match_service import match_expense_flows
    res = match_expense_flows(db, actor=getattr(user, "username", None))
    db.commit()
    return {"matched": res.matched, "ambiguous": res.ambiguous,
            "unmatched": res.unmatched, "details": res.details}


@router.get("/reconciliation/snapshots")
def list_recon_snapshots(
    days: int = Query(30, le=180),
    db: Session = Depends(get_db),
):
    """对账每日快照 (近 N 天) — 看各规则差异是在收敛还是恶化。"""
    from sqlalchemy import text as _sql
    try:
        rows = db.execute(_sql(
            "SELECT snap_date, rule, ok_count, warning_count, error_count, total_diff_abs "
            "FROM recon_snapshots WHERE snap_date >= CURRENT_DATE - CAST(:d AS integer) "
            "ORDER BY snap_date, rule"
        ), {"d": days}).fetchall()
    except Exception:
        return {"rows": []}
    return {"rows": [
        {"snap_date": str(r[0]), "rule": r[1], "ok": r[2], "warning": r[3],
         "error": r[4], "total_diff_abs": float(r[5] or 0)} for r in rows
    ]}


@router.get("/reconciliation/ai-diagnosis")
def reconciliation_ai_diagnosis(db: Session = Depends(get_db)):
    """对当前所有对账差异做 AI 诊断, 给出可能原因 + 建议处理优先级。

    静态路由注册在 /{rule} 之前, 避免被拦截。
    AI 未配置时返回 ai_available=false + diagnosis=None。
    """
    from app.services import ai_assistant as _ai
    from app.services.ai_assistant import collect_reconcile_findings as _cf
    findings = _cf(db)
    _log, ai = _ai.diagnose_reconciliation(db)
    return {
        "diagnosis": ai.text if ai else None,
        "findings_count": len(findings),
        "ai_available": ai is not None,
        "model": ai.model if ai else None,
    }


@router.get("/reconciliation/{rule}", response_model=ReconciliationOut)
def run_one_rule(
    rule: str,
    period_start: Optional[date] = Query(None, description="账期起 (含)"),
    period_end: Optional[date] = Query(None, description="账期止 (含)"),
    db: Session = Depends(get_db),
):
    fn = reconciliation_service.RULES.get(rule)
    if fn is None:
        raise HTTPException(404, f"unknown rule {rule!r}; available: {list(reconciliation_service.RULES)}")
    reconciliation_service.load_thresholds(db)
    r = fn(db, record_exceptions=False, period_start=period_start, period_end=period_end)
    return _to_out(r)


@router.get("/reconciliation", response_model=dict[str, ReconciliationOut])
def run_all_rules(
    period_start: Optional[date] = Query(None, description="账期起 (含)"),
    period_end: Optional[date] = Query(None, description="账期止 (含)"),
    db: Session = Depends(get_db),
):
    results = reconciliation_service.run_all(
        db, record_exceptions=False, period_start=period_start, period_end=period_end,
    )
    return {name: _to_out(r) for name, r in results.items()}


@router.post("/realtime-sync")
def realtime_sync_now():
    """手动「立即同步」: 重算成本兜底 + 14 条对账规则写异常池, 跑完返回结果。

    平时导入(补单/异步导入等)会自动后台触发同一逻辑; 这里给个手动入口,
    清理异常或改了数据后想立刻让待办/异常清单跟上时点一下。
    """
    from app.services import realtime_sync_service
    return realtime_sync_service.run_sync_blocking("manual")


@router.get("/realtime-sync/status")
def realtime_sync_status():
    from app.services import realtime_sync_service
    return realtime_sync_service.status()


@router.get("/cost-anomaly")
def cost_anomaly(period_start: Optional[date] = Query(None), period_end: Optional[date] = Query(None),
                 product_code: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """诊断销售成本异常: 错配单(成本>实付) + 成本口径。默认本年。供解释利润率偏低/总利润为负。"""
    from datetime import date as _d
    from app.models.order import Order as _O
    start = period_start or _d(_d.today().year, 1, 1)
    end = period_end or _d.today()
    q = select(_O).where(_O.order_date >= start, _O.order_date <= end,
                         _O.is_refill == False,  # 刷单是假单, 不进成本异常诊断 (2026-06-19)
                         _O.status.in_(("paid", "shipped", "signed")))
    if product_code:
        q = q.where(_O.product_code == product_code)
    orders = db.execute(q).scalars().all()
    tot_rev = tot_cost = Decimal("0")
    mism = []
    for o in orders:
        rev = Decimal(o.paid_amount or 0)
        cost = Decimal(o.actual_cost if o.actual_cost is not None else (o.theoretical_cost or 0))
        tot_rev += rev
        tot_cost += cost
        if cost > rev:
            mism.append({"order_no": o.order_no, "product_code": o.product_code,
                         "product_name": o.product_name, "paid": float(rev), "cost": float(cost),
                         "excess": float(cost - rev),
                         "cost_src": "actual" if o.actual_cost is not None else "theoretical"})
    mism.sort(key=lambda r: r["excess"], reverse=True)
    return {
        "period": [start.isoformat(), end.isoformat()], "orders": len(orders),
        "total_revenue": float(tot_rev), "total_cost": float(tot_cost),
        "cost_minus_revenue": float(tot_cost - tot_rev),
        "mismatched_count": len(mism), "mismatched_excess_total": round(sum(m["excess"] for m in mism), 2),
        "samples": mism[:40],
    }


@router.get("/refill-summary")
def refill_summary(period_start: Optional[date] = Query(None), period_end: Optional[date] = Query(None),
                   db: Session = Depends(get_db)):
    """刷单(补单)单列汇总 — 给所有算账页面统一展示「刷单 X笔 ¥Y 已单列、不计入经营数据」。

    刷单是假单(流水来回滚抵销、非真销售), 已从所有营收/利润/成交/资产/现金流/大盘/排行/预测/报表
    剔除; 本接口把被剔除的那部分单独算出来给前端单列提示。
    """
    from datetime import date as _d
    from app.services import order_financials
    start = period_start or _d(_d.today().year, 1, 1)
    end = period_end or _d.today()
    coef = order_financials.load_coefficients(db)
    r = order_financials.refill_cost(db, start, end, coef)
    return {
        "period": [start.isoformat(), end.isoformat()],
        "count": r["count"],
        "gmv": float(r["gmv"]),        # 刷单流水(订单额)总额
        "cost": float(r["total"]),     # 刷单真实成本(平台扣点+税+运费+佣金)
        "note": "刷单(补单)为假单, 已从上方所有经营/财务数据中剔除、单列于此",
    }


@router.post("/refill-records/purge-pre-2026")
def purge_pre_2026_refills(db: Session = Depends(get_db)):
    """删除 2025 及以前的补单记录 (系统从 2026 起算, 用户拍板 2026-06-17)。"""
    from datetime import date as _d
    from app.models.finance import RefillRecord as _RR
    rows = db.execute(select(_RR).where(_RR.refill_date < _d(2026, 1, 1))).scalars().all()
    n = len(rows)
    for r in rows:
        db.delete(r)
    db.commit()
    return {"deleted_pre_2026_refills": n}


@router.post("/refill-records/purge-unmatched")
def purge_unmatched_refills(db: Session = Depends(get_db)):
    """删除「其它店铺」补单记录 = 订单号在主订单总表里找不到的补单 (孚格家居/小红书等)。

    主店(畔色淘宝)补单的订单号都在系统里; 其它店铺的对不上 → 就是 refill_unmatched 异常那批。
    删除后下次同步复核会自动销账对应异常 (用户拍板 2026-06-17: 先全删, 加第二店铺后再议)。
    """
    from app.models.finance import RefillRecord as _RR
    from app.models.order import Order as _O
    known = {o for (o,) in db.execute(select(_O.order_no)).all()}
    rows = db.execute(select(_RR)).scalars().all()
    victims = [r for r in rows if r.order_no not in known]
    samples = [r.order_no for r in victims[:20]]
    for r in victims:
        db.delete(r)
    db.commit()
    return {"deleted_unmatched_refills": len(victims), "remaining_refills": len(rows) - len(victims),
            "sample_order_nos": samples}


@router.get("/exception-audit")
def exception_audit(db: Session = Depends(get_db)):
    """诊断 cost_missing_estimated 异常构成(关闭/非产品/真缺) + 补单按年分布。"""
    from sqlalchemy import extract, func as _f
    from app.models.exception import DataException
    from app.models.order import Order as _O
    from app.models.finance import RefillRecord as _RR
    from app.services import order_cost_service as ocs
    status_counts: dict = {}
    skip_e = zero_e = real_e = 0
    for ex in db.execute(select(DataException).where(
        DataException.exception_type == "cost_missing_estimated",
        DataException.status == "open")).scalars().all():
        o = db.execute(select(_O).where(_O.order_no == ex.source_pk)).scalar_one_or_none()
        if o is None:
            continue
        status_counts[o.status or "?"] = status_counts.get(o.status or "?", 0) + 1
        if ocs._skip_cost_estimate(o) is not None:
            skip_e += 1
        elif ocs.zero_cost_reason(o) is not None or not o.theoretical_cost:
            zero_e += 1
        else:
            real_e += 1
    ry: dict = {}
    for y, c in db.execute(select(extract("year", _RR.refill_date), _f.count())
                           .group_by(extract("year", _RR.refill_date))).all():
        ry[str(int(y)) if y else "null"] = c
    return {
        "cost_missing_status_counts": status_counts,
        "cme_skip_eligible(关闭/退款/旧)": skip_e, "cme_zero_nonproduct": zero_e,
        "cme_real_missing": real_e, "refill_by_year": ry,
    }


@router.get("/orders-missing-code")
def orders_missing_code(db: Session = Depends(get_db)):
    """诊断: 2026 销售订单里 product_code 为空的, 看能否经 sku_code→定价表 解出产品编码+短名。"""
    from datetime import date as _d
    from app.models.order import Order as _O
    from app.models.pricing import PricingSku
    from app.models.product import Product as _P
    orders = db.execute(select(_O).where(
        _O.product_code.is_(None), _O.order_date >= _d(2026, 1, 1),
        _O.is_refill == False,  # 刷单本就无产品编码, 不算缺编码异常 (2026-06-19)
        _O.status.in_(("paid", "shipped", "signed")))).scalars().all()
    sku2code = {s: c for s, c in db.execute(select(PricingSku.sku_code, PricingSku.product_code)).all() if s}
    code2name = {c: n for c, n in db.execute(select(_P.code, _P.name)).all()}
    out = []
    resolvable = 0
    for o in orders:
        via = sku2code.get(o.sku_code) if o.sku_code else None
        if via:
            resolvable += 1
        out.append({"order_no": o.order_no, "product_name": (o.product_name or "")[:36],
                    "sku_code": o.sku_code, "sku": (o.sku or "")[:24],
                    "resolved_code": via, "resolved_name": code2name.get(via) if via else None})
    return {"total_missing_code": len(orders), "resolvable_via_sku": resolvable, "samples": out[:30]}


@router.get("/order-payment-diagnosis")
def order_payment_diagnosis(order_nos: str = Query(..., description="逗号分隔订单号"),
                            db: Session = Depends(get_db)):
    """诊断「订单缺支付宝收款流水」: 这几单现在到底有没有收款凭据, 异常该不该清。"""
    from app.services import data_quality_service
    nos = [x.strip() for x in order_nos.split(",") if x.strip()]
    return {"rows": data_quality_service.order_payment_diagnosis(db, nos)}


@router.get("/alipay-balance-gaps")
def alipay_balance_gaps(account_kw: str = Query("企业号"), db: Session = Depends(get_db)):
    """诊断支付宝余额断链: 每条断链给出前后相邻流水 + 是否前驱余额为空(假断链) vs 真漏一笔。"""
    from app.services import data_quality_service
    rows = data_quality_service.balance_gap_details(db, account_kw=account_kw)
    false_alarm = sum(1 for r in rows if r["null_balance_flow_nearby"])
    return {
        "account_kw": account_kw, "gap_count": len(rows),
        "likely_false_alarm": false_alarm, "likely_real_missing": len(rows) - false_alarm,
        "gaps": rows,
    }


@router.get("/reconciliation-accuracy")
def reconciliation_accuracy(db: Session = Depends(get_db)):
    """按月『对账准确度』: 哪些月份财务已核准(该月有订单且无 open 财务对账异常)。
    用户拍板 2026-06-15: 体现哪几个月财务真实准确。"""
    rows = reconciliation_service.reconciliation_accuracy_by_month(db)
    return {
        "months": rows,
        "accurate_months": [r["month"] for r in rows if r["accurate"]],
        "accurate_count": sum(1 for r in rows if r["accurate"]),
        "total_months": len(rows),
    }


# -------- CSV 模板下载 --------

def _csv_template(headers: list[str], filename: str) -> StreamingResponse:
    buf = io.StringIO()
    buf.write("﻿")  # BOM for Excel
    csv.writer(buf).writerow(headers)
    return StreamingResponse(
        iter([buf.getvalue().encode("utf-8")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# -------- 万师傅 / 物流账单 / 补单记录 列表 --------

class WanshifuBillOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    bill_date: Optional[date]
    order_no: Optional[str]
    service_type: Optional[str]
    amount: Optional[Decimal] = None   # 防 amount 为 null 时 500 (序列化报错)
    status: Optional[str]
    remark: Optional[str]


class LogisticsBillOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    bill_date: Optional[date]
    carrier: Optional[str]
    tracking_no: Optional[str]
    order_no: Optional[str]
    weight_kg: Optional[Decimal]
    freight_amount: Decimal
    remark: Optional[str]
    recipient_name: Optional[str] = None
    destination: Optional[str] = None
    match_method: Optional[str] = None
    match_note: Optional[str] = None
    row_type: str = "line"
    # 匹配到的订单的客户名/收货地址 — 供人工核对"收货人/目的地"是否真的对得上(查匹配错误)
    order_customer_name: Optional[str] = None
    order_customer_address: Optional[str] = None
    # 订单号在订单库里是否真的存在 — 前端曾拿"客户名为空"当"订单库无此单",
    # 把「有此单但订单没存客户名」误报成无此单(2026-07-12, 23条全是误报), 必须单独给存在性
    order_exists: bool = False


class PromotionFlowOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    transaction_date: Optional[date]
    flow_type: Optional[str]
    amount: Decimal
    alipay_flow_no: Optional[str]
    remark: Optional[str]


class RefillRecordOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    order_no: str
    buyer_nick: Optional[str]
    refill_date: Optional[date]
    product_code: Optional[str]
    product_name: Optional[str]
    sku: Optional[str]
    qty: int
    order_amount: Optional[Decimal]
    refill_cost: Optional[Decimal]
    total_cost: Optional[Decimal]


@router.get("/wanshifu-bills", response_model=list[WanshifuBillOut])
def list_wanshifu_bills(
    year: Optional[int] = None,
    limit: int = Query(500, le=2000),
    db: Session = Depends(get_db),
):
    stmt = select(WanshifuBill)
    if year:
        from sqlalchemy import extract
        stmt = stmt.where(extract("year", WanshifuBill.bill_date) == year)
    stmt = stmt.order_by(WanshifuBill.bill_date.desc().nulls_last()).limit(limit)
    return db.execute(stmt).scalars().all()


@router.get("/logistics-bills", response_model=list[LogisticsBillOut])
def list_logistics_bills(
    carrier: Optional[str] = None,
    year: Optional[int] = None,
    limit: int = Query(500, le=2000),
    db: Session = Depends(get_db),
):
    stmt = select(LogisticsBill)
    if carrier:
        stmt = stmt.where(LogisticsBill.carrier == carrier)
    if year:
        from sqlalchemy import extract
        stmt = stmt.where(extract("year", LogisticsBill.bill_date) == year)
    stmt = stmt.order_by(LogisticsBill.bill_date.desc().nulls_last()).limit(limit)
    bills = db.execute(stmt).scalars().all()
    return _enrich_logistics_bills(db, bills)


def _enrich_logistics_bills(db: Session, bills: list) -> list[LogisticsBillOut]:
    """给每条逐单行补上"匹配到的订单"的客户名/收货地址, 供前端核对收货人/目的地是否真对得上。"""
    from app.models.order import Order
    onos = {b.order_no for b in bills if b.order_no}
    omap: dict[str, tuple] = {}
    if onos:
        for ono, nm, addr in db.execute(
            select(Order.order_no, Order.customer_name, Order.customer_address)
            .where(Order.order_no.in_(onos))
        ).all():
            omap[ono] = (nm, addr)
    out: list[LogisticsBillOut] = []
    for b in bills:
        d = LogisticsBillOut.model_validate(b)
        if b.order_no and b.order_no in omap:
            d.order_exists = True   # 库里有此单 (客户名可能为空 — 存在性与有没有名字是两回事)
            d.order_customer_name, d.order_customer_address = omap[b.order_no]
        out.append(d)
    return out


class LogisticsBillMatchIn(BaseModel):
    order_no: Optional[str] = None   # 填订单号=人工指定匹配; 空=取消匹配


@router.patch("/logistics-bills/{bill_id}/match", response_model=LogisticsBillOut)
def set_logistics_bill_match(bill_id: int, payload: LogisticsBillMatchIn,
                             db: Session = Depends(get_db)):
    """人工核对: 改某条逐单行的订单号。填=manual(自动配单不再覆盖), 空=取消(none)。
    改后回填该订单实际物流费 (与自动配单一致)。"""
    b = db.get(LogisticsBill, bill_id)
    if not b:
        raise HTTPException(404, "账单行不存在")
    prev_no = b.order_no                      # 改配单前的旧订单 — 定点回退它的实际物流费
    ono = (payload.order_no or "").strip()
    if ono:
        b.order_no = ono
        b.match_method = "manual"
        b.match_note = None
    else:
        b.order_no = None
        b.match_method = "none"
        b.match_note = "人工取消匹配"
    db.commit()
    # 定点回填新旧两单 (2026-07-11): 只重算受影响订单, 取消配单时旧单按"对齐"语义正确回退;
    # 不再全量 sync —— 全量会重刷所有单的 est_*, 且旧全量语义曾会清空手工合并的实际费用。
    affected = [x for x in {prev_no, b.order_no} if x]
    if affected:
        try:
            from app.services import order_fee_actual_service
            order_fee_actual_service.sync_fee_components(db, order_nos=affected)
            db.commit()
        except Exception:  # noqa: BLE001 — 回填失败不阻断改匹配
            pass
    return _enrich_logistics_bills(db, [b])[0]


@router.get("/logistics-bills/{bill_id}/match-candidates")
def logistics_bill_match_candidates(
    bill_id: int,
    limit: int = Query(5, le=20),
    name: Optional[str] = None,   # 传入则按这个(改过的)收货人名算相似度
    db: Session = Depends(get_db),
):
    """按收货人名相似度列候选订单(供下拉自选), 匹配度高→低取前 5;
    同分时目的地对上/下单日近账单日的排前 (用户 2026-07-12, 与打包费核对同款)。"""
    from app.services import logistics_bill_match
    return logistics_bill_match.match_candidates(db, bill_id, limit=limit, name_override=name)


@router.post("/logistics-bills/match")
def match_logistics_bills_ep(only_unmatched: bool = True, loose: bool = False,
                             db: Session = Depends(get_db)):
    """物流费账单逐单行 → 淘宝订单 自动配对 (运单号 / 收货人+省市)。
    loose=True 加宽松档(收货人姓名在订单地址里+省市对上, 唯一才配)。
    返回 {matched, multi, none}; 配不到的行 match_method='none' (前端显示「未能自动匹配」)。"""
    from app.services import logistics_bill_match
    counts = logistics_bill_match.match_logistics_bills(
        db, only_unmatched=only_unmatched, loose=loose)
    db.commit()
    return counts


@router.post("/packing-bills/rematch")
def rematch_packing_bills_ep(loose: bool = True, db: Session = Depends(get_db)):
    """打包费账单未配单行重跑配单 (loose=True 加宽松档: 客户名在订单地址+省份对上)。
    返回 {matched, multi, none}。"""
    from app.services import packing_bill_service
    counts = packing_bill_service.rematch_packing_bills(db, loose=loose)
    db.commit()
    try:  # 批量配单后, 把实际打包费回填到订单(覆盖预估), 与物流同一套 (用户 2026-06-24)
        from app.services import order_fee_actual_service
        order_fee_actual_service.sync_fee_components(db)
        db.commit()
    except Exception:
        db.rollback()
    return counts


@router.post("/orders/sync-fee-components")
def sync_fee_components_ep(db: Session = Depends(get_db)):
    """回填订单的 预估/实际 打包+物流费分量(供 physical_cost 实际替预估)。返回回填计数。"""
    from app.services import order_fee_actual_service
    r = order_fee_actual_service.sync_fee_components(db)
    db.commit()
    return r


def _fee_variance(db: Session, est_attr: str, actual_attr: str) -> dict:
    """逐单 实际 vs 预估 偏差。区分:
       rows  = 有预估(参与成本替换)的单 — 总计/偏差只算这些(=真实成本影响);
       gaps  = 配到实际但 SKU 缺定价预估(未替换)的单 — 单列提醒补价。"""
    from app.models.order import Order
    from app.services.sales_analytics import SETTLED_SALE_STATUSES
    stmt = select(Order.order_no, Order.customer_name, Order.product_name,
                  getattr(Order, est_attr), getattr(Order, actual_attr)).where(
        getattr(Order, actual_attr).isnot(None),
        Order.status.in_(SETTLED_SALE_STATUSES),
        Order.is_refill == False,  # noqa: E712
    )
    rows, gaps = [], []
    tot_est = tot_act = gap_act = Decimal("0")
    for no, cust, prod, est, act in db.execute(stmt).all():
        act_d = act or Decimal("0")
        if est is None:   # SKU 缺定价预估 → 未替换, 单列
            gaps.append({"order_no": no, "customer_name": cust,
                         "product_name": prod, "actual": float(act_d)})
            gap_act += act_d
            continue
        est_d = Decimal(str(est))
        diff = act_d - est_d
        rows.append({
            "order_no": no, "customer_name": cust, "product_name": prod,
            "est": float(est_d), "actual": float(act_d), "diff": float(diff),
            "diff_pct": (round(float(diff / est_d * 100), 1) if est_d else None),
        })
        tot_est += est_d
        tot_act += act_d
    rows.sort(key=lambda r: abs(r["diff"]), reverse=True)
    tot_diff = tot_act - tot_est
    return {
        "rows": rows, "count": len(rows),
        "total_est": float(tot_est), "total_actual": float(tot_act),
        "total_diff": float(tot_diff),
        "diff_pct": (round(float(tot_diff / tot_est * 100), 1) if tot_est else None),
        "gaps": gaps, "gap_count": len(gaps), "gap_actual": float(gap_act),
    }


@router.get("/logistics-bills/variance")
def logistics_fee_variance(db: Session = Depends(get_db)):
    """物流费 实际(德邦逐单) vs 预估(定价表) 逐单偏差 + 总计。"""
    return _fee_variance(db, "est_logistics", "actual_logistics")


@router.get("/packing-bills/variance")
def packing_fee_variance(db: Session = Depends(get_db)):
    """打包费 实际(手写账单) vs 预估(定价表) 逐单偏差 + 总计。"""
    return _fee_variance(db, "est_packing", "actual_packing")


@router.get("/install/variance")
def install_fee_variance(db: Session = Depends(get_db)):
    """安装费 实际(订单 install_fee+upstairs_fee) vs 预估(定价表) 逐单偏差 + 总计。"""
    return _fee_variance(db, "est_install", "actual_install")


@router.get("/factory-wood/variance")
def factory_wood_variance(db: Session = Depends(get_db)):
    """工厂木作 实际(工厂对账单 actual_cost) vs 预估(定价表 wood_cost_est) 逐单偏差 + 总计。"""
    return _fee_variance(db, "wood_cost_est", "actual_cost")


class PackingBillOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    bill_month: Optional[str]
    row_date: Optional[date]
    customer_name: Optional[str]
    order_no: Optional[str]
    matched_order_no: Optional[str]
    match_method: Optional[str]
    match_note: Optional[str]
    product: Optional[str]
    packing_fee: Optional[Decimal]
    excluded: bool = False
    exclude_reason: Optional[str]
    confidence: Optional[Decimal]
    note: Optional[str]


@router.get("/packing-bills", response_model=list[PackingBillOut])
def list_packing_bills(
    bill_month: Optional[str] = None,
    limit: int = Query(1000, le=5000),
    db: Session = Depends(get_db),
):
    """打包费手写账单逐行列表 (按账期/最新优先)。"""
    stmt = select(PackingBill)
    if bill_month:
        stmt = stmt.where(PackingBill.bill_month == bill_month)
    stmt = stmt.order_by(PackingBill.bill_month.desc().nulls_last(),
                         PackingBill.row_date.desc().nulls_last(),
                         PackingBill.id.desc()).limit(limit)
    return db.execute(stmt).scalars().all()


@router.get("/packing-bills/summary")
def packing_bills_summary(bill_month: Optional[str] = None, db: Session = Depends(get_db)):
    """当月打包费应付：有效明细合计 / 剔除额 / 未配单数。"""
    from app.services import packing_bill_service
    return packing_bill_service.month_summary(db, bill_month)


class PackingPaymentAllocationIn(BaseModel):
    flow_id: int
    bill_month: str
    amount: Decimal
    note: Optional[str] = None


@router.get("/packing-bills/payment-reconciliation")
def packing_payment_reconciliation(
    bill_month: str = Query(..., pattern=r"^\d{4}-(0[1-9]|1[0-2])$"),
    db: Session = Depends(get_db),
):
    """月度打包应付明细合计 ↔ 已分配的支付宝打包费付款。"""
    from app.services import packing_payment_service
    return packing_payment_service.month_summary(db, bill_month)


@router.post("/packing-bills/payment-allocations/auto")
def auto_allocate_packing_payments(
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin", "operator")),
):
    """自动分配明确写有“打包费”且只含一个费用账期的支付流水。"""
    from app.services import packing_payment_service
    result = packing_payment_service.auto_allocate(db)
    db.commit()
    try:
        from app.services import realtime_sync_service
        realtime_sync_service.trigger("packing-payment:auto-allocate")
    except Exception:  # noqa: BLE001
        pass
    return result


@router.post("/packing-bills/payment-allocations")
def allocate_packing_payment(
    payload: PackingPaymentAllocationIn,
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin", "operator")),
):
    """把一笔打包费付款全部或部分分配到指定费用账期，支持跨月拆分。"""
    from app.services import packing_payment_service
    try:
        row = packing_payment_service.create_allocation(
            db, flow_id=payload.flow_id, bill_month=payload.bill_month,
            amount=payload.amount, note=payload.note,
        )
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    db.commit()
    try:
        from app.services import realtime_sync_service
        realtime_sync_service.trigger(f"packing-payment:allocate:{row.id}")
    except Exception:  # noqa: BLE001
        pass
    return {"allocation_id": row.id, "flow_id": row.alipay_flow_id,
            "bill_month": row.bill_month, "amount": float(row.amount)}


@router.delete("/packing-bills/payment-allocations/{allocation_id}")
def remove_packing_payment_allocation(
    allocation_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin", "operator")),
):
    from app.services import packing_payment_service
    if not packing_payment_service.delete_allocation(db, allocation_id):
        raise HTTPException(404, "打包费付款分配不存在")
    db.commit()
    try:
        from app.services import realtime_sync_service
        realtime_sync_service.trigger(f"packing-payment:unallocate:{allocation_id}")
    except Exception:  # noqa: BLE001
        pass
    return {"deleted": allocation_id}


class PackingBillPatch(BaseModel):
    customer_name: Optional[str] = None
    packing_fee: Optional[Decimal] = None
    matched_order_no: Optional[str] = None   # 手动指定订单号; 空串 = 清空配单
    excluded: Optional[bool] = None
    note: Optional[str] = None
    bill_month: Optional[str] = None         # 改账期 YYYY-MM (手写本错填月份/OCR错识别月份时挪正确账期)
    rematch: bool = False                    # 改完客户名后按名自动重配


@router.patch("/packing-bills/{bill_id}", response_model=PackingBillOut)
def update_packing_bill(bill_id: int, payload: PackingBillPatch, db: Session = Depends(get_db)):
    """手动编辑一行打包费账单: 改客户名/打包费/手动配单 (用户 2026-06-24)。只更新请求里出现的字段。"""
    from app.services import packing_bill_service
    existing = db.get(PackingBill, bill_id)
    old_match = existing.matched_order_no if existing else None
    fields = payload.model_dump(exclude_unset=True)
    rematch = bool(fields.pop("rematch", False))
    if fields.get("bill_month"):
        import re as _re
        if not _re.fullmatch(r"\d{4}-\d{2}", str(fields["bill_month"]).strip()):
            raise HTTPException(400, "账期格式应为 YYYY-MM")
    b = packing_bill_service.update_row(db, bill_id, rematch=rematch, **fields)
    if b is None:
        raise HTTPException(404, "打包费账单行不存在")
    db.commit()          # get_db 不自动提交, 必须显式 commit, 否则手动编辑不落库 (修 2026-06-24)
    db.refresh(b)
    # 配单变化 → 把实际打包费回填到对应订单(actual_packing 覆盖 est_packing, 进而覆盖物理成本/利润),
    # 与物流账单/木作账单同一套实际覆盖预估机制 (用户 2026-06-24)。回填失败不阻断改匹配。
    affected = [o for o in {old_match, b.matched_order_no} if o]
    if affected:
        try:
            from app.services import order_fee_actual_service
            order_fee_actual_service.sync_fee_components(db, order_nos=affected)
            db.commit()
        except Exception:
            db.rollback()
    return b


@router.delete("/packing-bills/{bill_id}")
def delete_packing_bill(bill_id: int, db: Session = Depends(get_db)):
    """删除一行打包费账单 (用户 2026-06-29: 清理重复导入的账册行)。
    删前记下已配的订单号, 删后回退该订单的 actual_packing(回到 est_packing, 正确物理成本/利润)。"""
    b = db.get(PackingBill, bill_id)
    if b is None:
        raise HTTPException(404, "打包费账单行不存在")
    affected = b.matched_order_no
    db.delete(b)
    db.commit()
    if affected:
        try:
            from app.services import order_fee_actual_service
            order_fee_actual_service.sync_fee_components(db, order_nos=[affected])
            db.commit()
        except Exception:
            db.rollback()
    return {"deleted": bill_id, "affected_order_no": affected}


@router.get("/packing-bills/{bill_id}/match-candidates")
def packing_bill_match_candidates(
    bill_id: int,
    limit: int = Query(5, le=20),
    name: Optional[str] = None,   # 传入则按这个(改过的)客户名算相似度
    db: Session = Depends(get_db),
):
    """按客户名相似度列候选订单(供下拉自选), 匹配度高→低, 默认前 5 (用户 2026-06-24)。"""
    from app.services import packing_bill_service
    return packing_bill_service.match_candidates(db, bill_id, limit=limit, name_override=name)


@router.get("/promotion-flows", response_model=list[PromotionFlowOut])
def list_promotion_flows(
    flow_type: Optional[str] = None,
    year: Optional[int] = None,
    limit: int = Query(500, le=2000),
    db: Session = Depends(get_db),
):
    stmt = select(PromotionFlow)
    if flow_type:
        stmt = stmt.where(PromotionFlow.flow_type == flow_type)
    if year:
        from sqlalchemy import extract
        stmt = stmt.where(extract("year", PromotionFlow.transaction_date) == year)
    stmt = stmt.order_by(PromotionFlow.transaction_date.desc().nulls_last()).limit(limit)
    return db.execute(stmt).scalars().all()


@router.get("/refill-records", response_model=list[RefillRecordOut])
def list_refill_records(
    year: Optional[int] = None,
    limit: int = Query(500, le=2000),
    db: Session = Depends(get_db),
):
    stmt = select(RefillRecord)
    if year:
        from sqlalchemy import extract
        stmt = stmt.where(extract("year", RefillRecord.refill_date) == year)
    stmt = stmt.order_by(RefillRecord.refill_date.desc().nulls_last()).limit(limit)
    return db.execute(stmt).scalars().all()


@router.get("/alipay-flows/template.csv")
def alipay_template():
    """下载支付宝流水导入模板 (空白 CSV, 含正确列名)。"""
    return _csv_template(
        ["交易时间", "交易流水号", "交易类型", "交易对象", "收支金额", "关联订单号", "余额", "备注"],
        "alipay_flows_template.csv",
    )


@router.get("/wanshifu-bills/template.csv")
def wanshifu_template():
    """下载万师傅安装账单导入模板。"""
    return _csv_template(
        ["日期", "订单号", "服务类型", "金额", "状态", "备注"],
        "wanshifu_bills_template.csv",
    )


@router.get("/logistics-bills/template.csv")
def logistics_template():
    """下载物流费账单导入模板。"""
    return _csv_template(
        ["日期", "承运商", "运单号", "订单号", "重量(kg)", "运费", "备注"],
        "logistics_bills_template.csv",
    )


@router.get("/promotion-flows/template.csv")
def promotion_template():
    """下载推广记录导入模板。"""
    return _csv_template(
        ["日期", "类型", "金额", "支付宝流水号", "备注"],
        "promotion_flows_template.csv",
    )


@router.get("/refill-records/template.csv")
def refill_records_template():
    """下载补单对账导入模板。"""
    return _csv_template(
        ["订单号", "买家", "补单日期", "产品编码", "产品名", "SKU", "订单金额", "数量",
         "补单成本", "补发运费", "平台费", "佣金", "总成本"],
        "refill_records_template.csv",
    )


@router.get("/accounts/template.csv")
def account_balances_template():
    """下载账户余额导入模板。"""
    return _csv_template(
        ["账户名", "年", "月", "期初余额", "收入", "支出", "期末余额", "备注"],
        "account_balances_template.csv",
    )


# -------- 账户余额 CSV 预览/确认 (两步导入) --------

class CsvPreviewRow(BaseModel):
    row: int
    data: dict[str, Any]
    valid: bool
    reason: Optional[str] = None


class CsvPreviewResult(BaseModel):
    total: int
    valid_count: int
    invalid_count: int
    preview_rows: list[CsvPreviewRow]


def _parse_account_balances_preview(text: str) -> CsvPreviewResult:
    """解析账户余额 CSV, 返回预览行 (不写库)。"""
    from app.services.bill_import_service import _BALANCE_MAP, _decimal, _rows
    rows = _rows(text, _BALANCE_MAP)
    preview: list[CsvPreviewRow] = []
    valid = 0
    for i, rec in enumerate(rows, start=2):
        account_name = (rec.get("account_name") or "").strip()
        try:
            year = int(rec.get("period_year") or 0)
            month = int(rec.get("period_month") or 0)
        except (ValueError, TypeError):
            preview.append(CsvPreviewRow(row=i, data=rec, valid=False, reason="年/月不是数字"))
            continue
        if not account_name or not year or not month:
            preview.append(CsvPreviewRow(row=i, data=rec, valid=False,
                                         reason="账户名/年/月为必填项"))
            continue
        closing = _decimal(rec.get("closing_balance"))
        preview.append(CsvPreviewRow(
            row=i,
            data={
                "account_name": account_name, "year": year, "month": month,
                "opening_balance": str(_decimal(rec.get("opening_balance")) or ""),
                "closing_balance": str(closing or ""),
                "income": str(_decimal(rec.get("income")) or ""),
                "expense": str(_decimal(rec.get("expense")) or ""),
            },
            valid=True,
        ))
        valid += 1
    return CsvPreviewResult(
        total=len(preview),
        valid_count=valid,
        invalid_count=len(preview) - valid,
        preview_rows=preview,
    )


@router.post("/accounts/parse-csv", response_model=CsvPreviewResult)
async def parse_account_balances(file: UploadFile = File(...)):
    """第一步: 解析账户余额 CSV, 返回预览 (不写库)。
    确认后调 /accounts/confirm-csv 提交。"""
    text = await _read_csv(file)
    return _parse_account_balances_preview(text)


@router.post("/accounts/confirm-csv", response_model=BillImportResult)
async def confirm_account_balances(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """第二步: 正式导入账户余额 CSV (同账户同月 upsert)。"""
    text = await _read_csv(file)
    r = bill_import_service.import_account_balances_csv(db, text)
    db.commit()
    return BillImportResult(inserted=r.inserted, skipped_invalid=r.skipped_invalid, errors=r.errors,
                            skipped_duplicate=r.skipped_duplicate, unmapped_columns=r.unmapped_columns)


# -------- 支付宝流水 CSV 预览 --------

class AlipayPreviewRow(BaseModel):
    row: int
    transaction_time: Optional[str]
    transaction_no: Optional[str]
    amount: Optional[str]
    counterparty: Optional[str]
    related_order_no: Optional[str]
    valid: bool
    reason: Optional[str] = None


class AlipayPreviewResult(BaseModel):
    total: int
    valid_count: int
    duplicate_count: int
    invalid_count: int
    preview_rows: list[AlipayPreviewRow]


@router.post("/alipay-flows/parse-csv", response_model=AlipayPreviewResult)
async def parse_alipay(
    account: str = Query(..., description="账户名"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """第一步: 解析支付宝流水 CSV, 返回预览 (不写库)。
    包含重复检测 (transaction_no 已存在的标注 duplicate)。"""
    import csv as _csv
    from io import StringIO
    from app.services.alipay_import import COLUMN_MAP, _decimal, _datetime

    text = await _read_csv(file)
    reader = _csv.DictReader(StringIO(text))
    field_map: dict[str, str] = {}
    for raw in (reader.fieldnames or []):
        norm = (raw or "").strip()
        if norm in COLUMN_MAP:
            field_map[raw] = COLUMN_MAP[norm]

    existing_nos: set[str] = set(
        db.execute(select(AlipayFlow.transaction_no)).scalars().all()
    )

    preview: list[AlipayPreviewRow] = []
    valid = dup = inv = 0
    for i, raw_row in enumerate(reader, start=2):
        rec: dict[str, Any] = {}
        for k, v in raw_row.items():
            fn = field_map.get(k)
            if fn:
                rec[fn] = v
        tx_no = (rec.get("transaction_no") or "").strip()
        amt = _decimal(rec.get("amount"))
        if not tx_no or amt is None:
            preview.append(AlipayPreviewRow(
                row=i, transaction_time=None, transaction_no=tx_no or None,
                amount=None, counterparty=None, related_order_no=None,
                valid=False, reason="流水号或金额缺失",
            ))
            inv += 1
            continue
        if tx_no in existing_nos:
            preview.append(AlipayPreviewRow(
                row=i,
                transaction_time=str(rec.get("transaction_time") or ""),
                transaction_no=tx_no,
                amount=str(amt),
                counterparty=rec.get("counterparty"),
                related_order_no=rec.get("related_order_no"),
                valid=False, reason="重复 (已存在)",
            ))
            dup += 1
            continue
        preview.append(AlipayPreviewRow(
            row=i,
            transaction_time=str(rec.get("transaction_time") or ""),
            transaction_no=tx_no,
            amount=str(amt),
            counterparty=rec.get("counterparty"),
            related_order_no=rec.get("related_order_no"),
            valid=True,
        ))
        valid += 1

    return AlipayPreviewResult(
        total=len(preview),
        valid_count=valid,
        duplicate_count=dup,
        invalid_count=inv,
        preview_rows=preview,
    )


# -------- 邮箱 IMAP 配置 & 手动触发 --------

class EmailConfigOut(BaseModel):
    host: str
    port: int
    ssl: bool
    user: str
    password_set: bool
    folder: str
    subject_filter: str
    sender_filter: str
    alipay_account: str


class EmailConfigIn(BaseModel):
    host: Optional[str] = None
    port: Optional[int] = None
    ssl: Optional[bool] = None
    user: Optional[str] = None
    password: Optional[str] = None
    folder: Optional[str] = None
    subject_filter: Optional[str] = None
    sender_filter: Optional[str] = None
    alipay_account: Optional[str] = None


class EmailPollResult(BaseModel):
    scanned: int
    imported: int
    skipped: int
    errors: list[str]


@router.get("/email-poll/config", response_model=EmailConfigOut)
def get_email_config(db: Session = Depends(get_db)):
    """查看邮箱 IMAP 轮询配置。"""
    return email_import_service.get_config(db)


@router.post("/email-poll/config", response_model=EmailConfigOut)
def update_email_config(payload: EmailConfigIn, db: Session = Depends(get_db)):
    """更新邮箱 IMAP 轮询配置 (只传需要修改的字段)。"""
    kwargs: dict[str, Any] = {}
    if payload.host is not None:
        kwargs["email_imap_host"] = payload.host
    if payload.port is not None:
        kwargs["email_imap_port"] = payload.port
    if payload.ssl is not None:
        kwargs["email_imap_ssl"] = payload.ssl
    if payload.user is not None:
        kwargs["email_username"] = payload.user
    if payload.password is not None:
        kwargs["email_password"] = payload.password
    if payload.folder is not None:
        kwargs["email_folder"] = payload.folder
    if payload.subject_filter is not None:
        kwargs["email_subject_filter"] = payload.subject_filter
    if payload.sender_filter is not None:
        kwargs["email_sender_filter"] = payload.sender_filter
    if payload.alipay_account is not None:
        kwargs["email_alipay_account"] = payload.alipay_account
    email_import_service.save_config(db, **kwargs)
    db.commit()
    return email_import_service.get_config(db)


@router.post("/email-poll/trigger", response_model=EmailPollResult)
def trigger_email_poll(db: Session = Depends(get_db)):
    """立即执行一次邮箱轮询 + 导入 (无需等调度器)。"""
    r = email_import_service.poll_and_import(db)
    return EmailPollResult(scanned=r.scanned, imported=r.imported,
                           skipped=r.skipped, errors=r.errors)


# -------- 剩余流水（可用资金）测算 --------

class CashFlowLineOut(BaseModel):
    key: str
    label: str
    amount: Decimal
    manual: bool
    source: str
    detail: Optional[list] = None   # 明细 (如 待缴税费 的逐季度 [{quarter,tax,is_current,paid}])


class CashFlowFreshnessOut(BaseModel):
    source: str
    as_of: Optional[str]
    days_ago: Optional[int]
    status: str  # fresh / aging / stale / unknown


class CashFlowSummaryOut(BaseModel):
    total: Decimal
    total_additions: Decimal
    total_subtractions: Decimal
    additions: list[CashFlowLineOut]
    subtractions: list[CashFlowLineOut]
    investment: dict | None = None   # 投资回收: 总投资/累计总利润/回收率 (单列, 不进可用资金)
    other_account_balance: Decimal
    freshness: list[CashFlowFreshnessOut]
    manual: dict | None = None       # 手动常量 + 税季度明细 (tax_quarters/tax_paid_quarters 供前端手选已缴)
    generated_at: str


@router.get("/cash-flow", response_model=CashFlowSummaryOut)
def get_cash_flow(db: Session = Depends(get_db)):
    """实时测算剩余流水（可用资金）+ 各数据源新鲜度。"""
    return cash_flow_service.compute_summary(db)


class CashFlowSettingsIn(BaseModel):
    shop_deposit: Optional[Decimal] = None
    total_investment: Optional[Decimal] = None
    factory_settlement_days: Optional[int] = None   # 工厂结算周期(天), 工厂欠款回填规则B用
    factory_advance_balance: Optional[Decimal] = None
    factory_advance_target_month: Optional[str] = None
    factory_advance_note: Optional[str] = None
    tax_paid_quarters: Optional[list[str]] = None   # 已缴税季度(手选), 如 ["2026-Q1"]; 不计入减项


@router.put("/cash-flow/settings", response_model=CashFlowSummaryOut)
def update_cash_flow_settings(payload: CashFlowSettingsIn, db: Session = Depends(get_db)):
    """更新手动常量（店铺保证金 / 总投资费用 / 工厂结算周期 / 已缴税季度），返回重新测算后的结果。"""
    try:
        cash_flow_service.update_manual(
            db,
            shop_deposit=payload.shop_deposit,
            total_investment=payload.total_investment,
            factory_settlement_days=payload.factory_settlement_days,
            factory_advance_balance=payload.factory_advance_balance,
            factory_advance_target_month=payload.factory_advance_target_month,
            factory_advance_note=payload.factory_advance_note,
            tax_paid_quarters=payload.tax_paid_quarters,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    db.commit()
    return cash_flow_service.compute_summary(db)


@router.post("/factory-payment/backfill")
def backfill_factory_payment(
    settlement_days: Optional[int] = Query(None, ge=0, description="结算周期天数; 缺省读后台配置(默认45)"),
    apply_settled_inference: bool = Query(True, description="是否启用已结算推断(规则B)"),
    dry_run: bool = Query(False, description="只统计不落库"),
    db: Session = Depends(get_db),
):
    """工厂订单付款状态对账回填 — 消除"工厂欠款虚高"(payment_status 默认 unpaid 从未回填)。"""
    from app.services import factory_payment_service
    result = factory_payment_service.backfill_payment_status(
        db, settlement_days=settlement_days,
        apply_settled_inference=apply_settled_inference, dry_run=dry_run,
    )
    if not dry_run:
        db.commit()
    return result


# -------- 对账差异 AI 诊断 --------

class ReconciliationDiagnosisOut(BaseModel):
    diagnosis: Optional[str]
    findings_count: int
    ai_available: bool
    model: Optional[str] = None


# -------- 财务系数设置 (会计成本费率; 用户拍板 2026-06-17) --------

@router.get("/financial-coefficients")
def get_financial_coefficients(db: Session = Depends(get_db)):
    """读财务系数 (平台手续费率/活动抽成率/生效日/税率)。会计成本/利润全系统用它。"""
    from app.services import order_financials as ofin, settings_service as _ss
    return {k: (_ss.get(db, k, env_fallback=False) or ofin.DEFAULTS[k]) for k in ofin.DEFAULTS}


@router.put("/financial-coefficients")
def put_financial_coefficients(
    payload: dict = Body(...),
    user: User = Depends(require_role("admin", "operator")),
    db: Session = Depends(get_db),
):
    """改财务系数 (高危: 影响全系统所有利润口径)。需登录密码二次确认 (前端再 2 次严重警告)。"""
    import re as _re

    from app.services import auth_service, order_financials as ofin, settings_service as _ss
    pwd = str(payload.get("password") or "")
    if not user.password_hash or not auth_service.verify_password(pwd, user.password_hash):
        raise HTTPException(403, "密码不正确, 财务系数未修改")
    _DATE_KEYS = ("fin_platform_activity_since", "fin_platform_activity_until", "fin_outsourcing_est_since")
    _MONEY_KEYS = ("fin_outsourcing_monthly",)   # 金额(元), 非 0~1 费率
    _FLAG_KEYS = ("fin_custom_cost_v2",)          # 0/1 灰度开关, 非 0~1 费率 (否则"1"被 <1 校验误拒)
    _TRUE = ("1", "on", "true", "yes")
    changed: dict = {}
    for k in ofin.DEFAULTS:
        if k in payload and payload[k] not in (None, ""):
            v = str(payload[k]).strip()
            if k in _DATE_KEYS:
                if not _re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
                    raise HTTPException(400, f"{ofin.COEF_LABELS.get(k, k)} 格式应为 YYYY-MM-DD")
            elif k in _MONEY_KEYS:
                try:
                    if float(v) < 0:
                        raise ValueError
                except ValueError:
                    raise HTTPException(400, f"{ofin.COEF_LABELS.get(k, k)} 应为 ≥0 的金额 (元)")
            elif k in _FLAG_KEYS:
                if v.lower() not in _TRUE + ("0", "off", "false", "no"):
                    raise HTTPException(400, f"{ofin.COEF_LABELS.get(k, k)} 应为开关 0 或 1")
                v = "1" if v.lower() in _TRUE else "0"   # 归一化为 0/1
            else:
                try:
                    fv = float(v)
                    if not (0 <= fv < 1):
                        raise ValueError
                except ValueError:
                    raise HTTPException(400, f"{ofin.COEF_LABELS.get(k, k)} 应为 0~1 的小数 (如 0.02 = 2%)")
            _ss.set_value(db, k, v, description=f"财务系数: {ofin.COEF_LABELS.get(k, k)}")
            changed[k] = v
    db.commit()
    return {"changed": changed, "coefficients": get_financial_coefficients(db)}


@router.get("/fixed-cost-items")
def get_fixed_cost_items(db: Session = Depends(get_db)):
    """自定义固定成本/管理费用项 (房租/水电/软件/折旧…) + 每月合计 (年度项÷12)。"""
    from app.services import order_financials as ofin
    return {"items": ofin.fixed_cost_items(db), "monthly_total": float(ofin.fixed_costs_monthly(db))}


@router.put("/fixed-cost-items")
def put_fixed_cost_items(
    payload: dict = Body(...),
    user: User = Depends(require_role("admin", "operator")),
    db: Session = Depends(get_db),
):
    """改自定义固定成本项。items=[{name, amount, period('monthly'|'yearly'), active}] (用户可自由增删)。"""
    import json

    from app.services import order_financials as ofin, settings_service as _ss
    items = payload.get("items")
    if not isinstance(items, list):
        raise HTTPException(400, "items 应为数组")
    clean = []
    for it in items:
        name = str(it.get("name") or "").strip()
        if not name:
            continue
        try:
            amt = float(it.get("amount") or 0)
        except (TypeError, ValueError):
            raise HTTPException(400, f"{name} 金额无效")
        if amt < 0:
            raise HTTPException(400, f"{name} 金额应 ≥0")
        period = "yearly" if str(it.get("period")) == "yearly" else "monthly"
        clean.append({"name": name, "amount": amt, "period": period, "active": bool(it.get("active", True))})
    _ss.set_value(db, "fin_fixed_cost_items", json.dumps(clean, ensure_ascii=False),
                  description="财务: 自定义固定成本/管理费用项 (房租等)")
    db.commit()
    return {"items": ofin.fixed_cost_items(db), "monthly_total": float(ofin.fixed_costs_monthly(db))}

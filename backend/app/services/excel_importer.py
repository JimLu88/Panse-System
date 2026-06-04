"""通用 Excel importer (业务需求扩展).

流程:
    1) preview_excel(file_bytes)  → 解析每个 sheet 的 header + 前 5 行
    2) infer_mapping(preview, entity_type)  → 调 AI 推断列映射 (后台可改 mapping)
    3) commit(rows, mapping, entity_type, ...)  → 按 mapping 批量入库 + 自动跑订单匹配

未知供应商默认自动创建 (supplier_type=other), 用户可事后到供应商页补类型 / 关键字。
"""
from __future__ import annotations

import io
import itertools
import json
import math
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Optional

# (done_rows, total_rows) -> None; 调用方用来更新 ImportJob.processed_rows
ProgressCallback = Callable[[int, int], None]
# 返回 True 表示外部要求取消 (用户点了取消按钮); worker 据此抛 CancelledImport
CancelCallback = Callable[[], bool]


class CancelledImport(RuntimeError):
    """用户在 UI 点了取消按钮, worker 干净退出 (rollback)."""

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.finance import AlipayFlow
from app.models.order import FactoryOrder
from app.models.supplier import DeliveryNote, DeliveryNoteLine, Supplier
from app.services import delivery_matcher, settings_service
from app.services.ai_provider import AiUnavailable, build_provider
from app.services.excel_schemas import ENTITY_SCHEMAS, get_schema


# ----------------------------- 数据结构 -------------------------- #


@dataclass
class SheetPreview:
    sheet_name: str
    row_count: int                 # 总行数 (不含 header)
    column_names: list[str]        # header
    sample_rows: list[list[Any]]   # 前 5 行原值
    suggested_entity: Optional[str] = None
    suggested_mapping: dict[str, str] = field(default_factory=dict)
    # mapping: target_field -> excel_column
    notes: list[str] = field(default_factory=list)


@dataclass
class ImportReport:
    entity_type: str
    sheet_name: str
    total_rows: int
    inserted_parents: int
    inserted_children: int          # 仅 delivery_note 有子项
    skipped_rows: int
    auto_created_suppliers: list[str] = field(default_factory=list)
    matched_lines: int = 0           # delivery_note 自动匹配命中数
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    # 重导冲突: 行匹配到已有记录但字段值不同 (on_conflict != overwrite 时不自动覆盖)
    # [{source_table, source_pk, diffs: [{field, old, new}]}]
    conflicts: list[dict] = field(default_factory=list)
    # Excel 列中未被映射到任何系统字段的列名 (用于提示用户哪些内容被忽略)
    unmapped_columns: list[str] = field(default_factory=list)


class ImporterError(RuntimeError):
    """importer 调用方传参错误 (前端可直接显示给用户)."""


# ----------------------------- preview --------------------------- #

# 检测真实表头时, 向下扫描的最大行数 (表头一般在前几行内)。
_HEADER_SCAN_ROWS = 12


def _row_nonempty_count(row: Any) -> int:
    if not row:
        return 0
    return sum(1 for c in row if c is not None and str(c).strip() != "")


def _detect_header_index(buffer: list) -> int:
    """在前若干行里定位真正的表头行。

    业务表格常在第一行放一个合并单元格大标题 (如「1-产品总表」, 整行只有 1 个非空格),
    真正的列名 (产品编码 / SKU编码 ...) 在第二行。openpyxl 读 read_only 时合并标题只在
    左上角有值, 其余为 None, 若直接拿第一行当表头, 全部列名都会错位, 导致每行都报
    「缺产品编码 / 缺 sku_code」被跳过。

    策略: 取前几行里"非空单元格最多"的那一行宽度为基准, 返回第一个达到其一半宽度
    (且 >=2 列) 的行作为表头 — 这样能跳过 1 格的标题横幅, 又不会误伤本就以表头开头的表。
    """
    if not buffer:
        return 0
    counts = [_row_nonempty_count(r) for r in buffer]
    best = max(counts)
    threshold = max(2, (best + 1) // 2)
    for i, c in enumerate(counts):
        if c >= threshold:
            return i
    return 0


def preview_excel(file_bytes: bytes, *, sample_rows: int = 5) -> list[SheetPreview]:
    """解析 Excel, 每个 sheet 返回 header + 前 N 行."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ImporterError(f"无法解析 Excel: {e}") from e

    previews: list[SheetPreview] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        # 先缓冲前几行, 自动定位真正的表头行 (跳过合并单元格大标题横幅)
        buffer: list = []
        for _ in range(_HEADER_SCAN_ROWS):
            try:
                buffer.append(next(rows_iter))
            except StopIteration:
                break
        if not buffer:
            previews.append(SheetPreview(
                sheet_name=ws.title, row_count=0, column_names=[], sample_rows=[],
                notes=["空 sheet, 已跳过"],
            ))
            continue
        h_idx = _detect_header_index(buffer)
        header_row = buffer[h_idx]
        column_names = [str(c).strip() if c is not None else f"col{i + 1}"
                        for i, c in enumerate(header_row)]
        # 过滤全空列名 (Excel 尾部 None)
        last_nonempty = 0
        for i, n in enumerate(column_names):
            if n and not n.startswith("col"):
                last_nonempty = i + 1
        column_names = column_names[:last_nonempty]
        notes: list[str] = []
        if h_idx > 0:
            notes.append(f"已跳过前 {h_idx} 行 (合并标题/空行), 第 {h_idx + 1} 行识别为表头")
        sample: list[list[Any]] = []
        total = 0
        # 表头之后的数据行 = 缓冲区剩余 + 迭代器剩余
        for r in buffer[h_idx + 1:]:
            total += 1
            if len(sample) < sample_rows:
                sample.append([_clean_value(c) for c in (r or [])][:last_nonempty])
        for r in rows_iter:
            total += 1
            if len(sample) < sample_rows:
                sample.append([_clean_value(c) for c in (r or [])][:last_nonempty])
        previews.append(SheetPreview(
            sheet_name=ws.title,
            row_count=total,
            column_names=column_names,
            sample_rows=sample,
            notes=notes,
        ))
    wb.close()
    return previews


def _clean_value(v: Any) -> Any:
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


# ----------------------------- AI 推断 mapping ------------------- #


_AI_SYSTEM_PROMPT = """你是 Excel → ERP 字段映射助手。给你一个 sheet 的列名 + 前几行数据,
请输出严格 JSON, 包括:
{
  "entity_type": "供给的 supported_entities 中选一个 key, 拿不准填 unknown",
  "mapping": { "目标字段名": "Excel 列名", ... },
  "skipped_columns": ["完全不需要的列名", ...],
  "warnings": ["数据脏的提示", ...]
}
规则:
- 只在提供的 "支持的目标字段" 里选 target 字段名
- 同义词应识别 (如 "供应商" → supplier_name, "送货日期" → delivery_date)
- 没法对应的 Excel 列扔进 skipped_columns
- 必填字段没找到对应列 → 写进 warnings
- 仅输出 JSON, 不要任何解释文字"""


def infer_mapping(
    db: Session, *, preview: SheetPreview, entity_type: Optional[str] = None,
) -> SheetPreview:
    """让 AI 给一个 sheet 推荐 entity_type + column mapping. 失败时返回原 preview."""
    cfg = settings_service.get_ai_config(db, "diagnose")
    try:
        provider = build_provider(cfg)
    except AiUnavailable:
        preview.notes.append("AI 未配置, 跳过推断 — 请在 UI 手动配 mapping")
        return preview

    if entity_type and entity_type != "auto":
        schemas = {entity_type: get_schema(entity_type)}
    else:
        schemas = ENTITY_SCHEMAS

    schema_doc = {
        et: {
            "label": s["label"],
            "fields": {
                fn: {
                    "required": f.get("required", False),
                    "desc": f.get("desc", ""),
                    "aliases": f.get("aliases", []),
                }
                for fn, f in s["fields"].items()
            },
        }
        for et, s in schemas.items()
    }
    user_msg = json.dumps({
        "supported_entities": schema_doc,
        "sheet": {
            "name": preview.sheet_name,
            "columns": preview.column_names,
            "sample_rows": preview.sample_rows,
        },
    }, ensure_ascii=False)

    try:
        resp = provider.chat(system=_AI_SYSTEM_PROMPT, user=user_msg, max_tokens=800)
    except AiUnavailable as e:
        preview.notes.append(f"AI 调用失败, 跳过推断: {e}")
        return preview

    try:
        data = _extract_json(resp.text)
    except ValueError as e:
        preview.notes.append(f"AI 返回无法解析: {e}")
        return preview

    et = data.get("entity_type")
    if et in ENTITY_SCHEMAS:
        preview.suggested_entity = et
    elif entity_type and entity_type != "auto":
        preview.suggested_entity = entity_type
    mapping = data.get("mapping") or {}
    if isinstance(mapping, dict):
        # 过滤: 只保留目标字段在 schema 里, Excel 列在 columns 里
        if preview.suggested_entity:
            valid_fields = set(get_schema(preview.suggested_entity)["fields"])
            preview.suggested_mapping = {
                k: v for k, v in mapping.items()
                if k in valid_fields and v in preview.column_names
            }
    for w in data.get("warnings") or []:
        preview.notes.append(f"AI 提示: {w}")
    return preview


_FENCE_RE = re.compile(r"^```(?:json)?\s*|\s*```$", re.MULTILINE)


def _extract_json(text: str) -> dict:
    cleaned = _FENCE_RE.sub("", text).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    start, end = cleaned.find("{"), cleaned.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"不是 JSON: {text[:200]}")
    return json.loads(cleaned[start: end + 1])


# ----------------------------- 提交入库 -------------------------- #


def commit_sheet(
    db: Session,
    *,
    file_bytes: Optional[bytes] = None,
    file_path: Optional[str] = None,
    sheet_name: str,
    entity_type: str,
    mapping: dict[str, str],          # target_field -> excel_column
    auto_create_suppliers: bool = True,
    auto_match_orders: bool = True,
    dry_run: bool = False,
    on_conflict: str = "overwrite",
    sheet_account: Optional[str] = None,
    progress_callback: Optional[ProgressCallback] = None,
    cancel_callback: Optional[CancelCallback] = None,
    import_batch_id: Optional[int] = None,
) -> ImportReport:
    """按 mapping 把一个 sheet 的所有行入库 (业务需求 6: 进度回调 + 取消).

    on_conflict: overwrite (默认, 直接覆盖已有记录) / keep (保留原值) /
                 ask (记录差异到 report.conflicts 但不覆盖, 等用户裁决).
    sheet_account: 支付宝流水专用 — 当 sheet 没有账户列时, 用这个账户名填充每行.
    """
    if file_bytes is None and file_path is None:
        raise ImporterError("必须提供 file_bytes 或 file_path 之一")
    if file_bytes is None:
        with open(file_path, "rb") as f:  # type: ignore[arg-type]
            file_bytes = f.read()

    schema = get_schema(entity_type)
    # 支付宝流水允许账户名走 sheet_account 注入, 不强制映射 account 列
    externally_satisfied = {"account"} if (entity_type == "alipay_flow" and sheet_account) else set()
    _validate_mapping(schema, mapping, satisfied=externally_satisfied)

    rows = _read_all_rows(file_bytes, sheet_name)
    report = ImportReport(
        entity_type=entity_type, sheet_name=sheet_name,
        total_rows=len(rows), inserted_parents=0, inserted_children=0, skipped_rows=0,
    )
    # 计算未被映射的 Excel 列 (mapping 的 values 是 excel 列名)
    if rows:
        all_excel_cols = set(rows[0].keys())
        mapped_cols = set(mapping.values())
        unmapped = sorted(all_excel_cols - mapped_cols)
        if unmapped:
            report.unmapped_columns = unmapped
            report.warnings.append(
                f"以下 {len(unmapped)} 个 Excel 列未被映射，其中的数据未导入系统：{', '.join(unmapped[:10])}"
                + ("…（更多见 unmapped_columns）" if len(unmapped) > 10 else "")
                + "。如需导入，请在字段映射配置里为这些列选择对应的系统字段。"
            )
    if not rows:
        if progress_callback:
            progress_callback(0, 0)
        return report
    if progress_callback:
        progress_callback(0, len(rows))

    if entity_type == "delivery_note":
        _commit_delivery_notes(
            db, rows=rows, mapping=mapping, schema=schema, report=report,
            auto_create_suppliers=auto_create_suppliers,
            auto_match_orders=auto_match_orders,
            on_conflict=on_conflict,
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
            import_batch_id=import_batch_id,
        )
    elif entity_type == "factory_order":
        _commit_factory_orders(db, rows=rows, mapping=mapping, report=report,
                               on_conflict=on_conflict,
                               progress_callback=progress_callback,
                               cancel_callback=cancel_callback,
                               import_batch_id=import_batch_id)
    elif entity_type == "alipay_flow":
        _commit_alipay_flows(db, rows=rows, mapping=mapping, report=report,
                             sheet_account=sheet_account,
                             on_conflict=on_conflict,
                             progress_callback=progress_callback,
                             cancel_callback=cancel_callback,
                             import_batch_id=import_batch_id)
    elif entity_type in ("product", "material", "bom_line", "product_inventory",
                          "part_inventory", "order", "account_balance", "pricing_sku",
                          "refill_record", "factory_reconciliation",
                          "outsourcing_expense", "aftersales", "competitor_price",
                          "daily_operations", "order_details", "wood_loss", "sample",
                          "promotion_flow"):
        _commit_generic(
            db, rows=rows, mapping=mapping, entity_type=entity_type, report=report,
            on_conflict=on_conflict,
            progress_callback=progress_callback, cancel_callback=cancel_callback,
            import_batch_id=import_batch_id,
        )
    else:  # pragma: no cover
        raise ImporterError(f"暂不支持 {entity_type} 的入库")

    if progress_callback:
        progress_callback(len(rows), len(rows))

    if dry_run:
        db.rollback()
    else:
        db.flush()
    return report


def _validate_mapping(schema, mapping: dict[str, str], *, satisfied: set = frozenset()) -> None:
    missing = [
        fn for fn, f in schema["fields"].items()
        if f.get("required") and fn not in mapping and fn not in satisfied
    ]
    if missing:
        raise ImporterError(
            f"必填字段未映射: {missing}. 请到前端把每个必填字段都选一个 Excel 列。"
        )


def _read_all_rows(file_bytes: bytes, sheet_name: str) -> list[dict[str, Any]]:
    """读完整 sheet 转为 [{column_name: value, ...}, ...]."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ImporterError(f"sheet {sheet_name!r} 不存在; 可选: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    # 缓冲前几行自动定位真实表头 (与 preview 一致, 跳过合并标题横幅)
    buffer: list = []
    for _ in range(_HEADER_SCAN_ROWS):
        try:
            buffer.append(next(rows_iter))
        except StopIteration:
            break
    if not buffer:
        wb.close()
        return []
    h_idx = _detect_header_index(buffer)
    header_row = buffer[h_idx]
    columns = [str(c).strip() if c is not None else f"col{i + 1}"
               for i, c in enumerate(header_row)]
    rows: list[dict[str, Any]] = []
    # 表头之后的数据 = 缓冲剩余 + 迭代器剩余
    for r in itertools.chain(buffer[h_idx + 1:], rows_iter):
        if r is None:
            continue
        row_dict = {col: r[i] if i < len(r) else None for i, col in enumerate(columns)}
        if all(v is None or v == "" for v in row_dict.values()):
            continue
        rows.append(row_dict)
    wb.close()
    return rows


# ----------------------------- type coercion --------------------- #

# 数值字段里的"未知/待补"占位符 — 当 None 处理 (区别于 0)。
_NUMERIC_PLACEHOLDERS = {
    "？", "?", "待补", "待定", "未知", "未定", "-", "—", "/", "N/A", "n/a", "NA", "na", "无",
}


def _coerce(value: Any, field_type: str, *, label: str) -> Any:
    if value is None or value == "":
        return None
    # 数值占位符 → 空值(待补): 木作成本/配件价格等暂时未知时, 用户填 ？/待补/— 表示
    # "未知"(区别于 0=确实免费)。统一当 None 入库, 不报错也不臆造 0, 由下游标记「成本不完整」。
    if field_type in ("int", "decimal") and not isinstance(value, bool):
        if str(value).strip() in _NUMERIC_PLACEHOLDERS:
            return None
    if field_type == "str":
        return str(value).strip()
    if field_type == "int":
        if isinstance(value, (int,)):
            return value
        try:
            return int(float(str(value).replace(",", "").strip()))
        except (ValueError, InvalidOperation) as e:
            raise ImporterError(f"{label} 不是整数: {value!r}") from e
    if field_type == "decimal":
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        try:
            s = str(value).replace(",", "").replace("¥", "").replace("元", "").strip()
            return Decimal(s) if s else None
        except InvalidOperation as e:
            raise ImporterError(f"{label} 不是数字: {value!r}") from e
    if field_type == "date":
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        # Excel 日期序列号 (如 46140) → 真实日期 (Excel 纪元 1899-12-30)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            try:
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
            except (ValueError, OverflowError):
                pass
        s = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(s.replace("年", "-").replace("月", "-")
                                          .replace("日", "").replace("/", "-")
                                          .replace(".", "-"), "%Y-%m-%d").date()
            except ValueError:
                continue
        raise ImporterError(f"{label} 日期格式无法识别: {value!r}")
    if field_type == "datetime":
        if isinstance(value, datetime):
            return value
        # Excel 日期序列号 → datetime (含小数=时间部分)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            try:
                return datetime(1899, 12, 30) + timedelta(days=float(value))
            except (ValueError, OverflowError):
                pass
        return None
    if field_type == "bool":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("是", "true", "yes", "1", "y")
    return value


def _project(
    row: dict[str, Any], mapping: dict[str, str], schema,
) -> tuple[dict[str, Any], list[str]]:
    """row + mapping → {target_field: typed_value}. 收集行级错误.

    非必填字段转换失败时置 None (不杀整行);必填字段失败才加入 errors.
    """
    out: dict[str, Any] = {}
    errors: list[str] = []
    for target_field, excel_col in mapping.items():
        raw = row.get(excel_col)
        field_def = schema["fields"].get(target_field)
        if field_def is None:
            continue
        # 跳过 Excel 公式错误值 (#DIV/0! 等)
        if isinstance(raw, str) and raw.strip().startswith("#"):
            if field_def.get("required"):
                errors.append(f"{target_field} 含 Excel 错误值: {raw!r}")
            else:
                out[target_field] = None
            continue
        try:
            out[target_field] = _coerce(raw, field_def.get("type", "str"),
                                        label=target_field)
        except ImporterError as e:
            if field_def.get("required"):
                errors.append(str(e))
            else:
                out[target_field] = None  # 非必填字段转换失败 → null,不杀整行
    return out, errors


# ----------------------------- delivery_note --------------------- #


def _get_or_create_supplier(
    db: Session, name: str, *, auto_create: bool,
    created_log: list[str],
) -> Optional[Supplier]:
    s = db.execute(
        select(Supplier).where(Supplier.name == name)
    ).scalar_one_or_none()
    if s is not None:
        return s
    if not auto_create:
        return None
    s = Supplier(name=name, supplier_type="other", is_active=True)
    db.add(s)
    db.flush()
    created_log.append(name)
    return s


_PROGRESS_TICK = 50   # 每 N 行向 callback 报一次进度


def _commit_delivery_notes(
    db: Session, *, rows: list[dict], mapping: dict[str, str], schema,
    report: ImportReport, auto_create_suppliers: bool, auto_match_orders: bool,
    on_conflict: str = "ask",
    progress_callback: Optional[ProgressCallback] = None,
    cancel_callback: Optional[CancelCallback] = None,
    import_batch_id: Optional[int] = None,
) -> None:
    """同 (supplier_name, note_no) 的行聚合为一张 DeliveryNote + 多 DeliveryNoteLine."""
    groups: dict[tuple[str, str], list[tuple[dict, dict]]] = {}
    # (parent_fields, line_fields) 按 group_by 聚合
    total = len(rows)
    for i, raw_row in enumerate(rows, start=1):
        if i % _PROGRESS_TICK == 0:
            if progress_callback:
                progress_callback(i, total)
            if cancel_callback and cancel_callback():
                raise CancelledImport(f"用户取消, 已处理 {i}/{total} 行")
        projected, errs = _project(raw_row, mapping, schema)
        if errs:
            report.errors.append(f"第 {i + 1} 行: " + "; ".join(errs))
            report.skipped_rows += 1
            continue
        supplier_name = projected.get("supplier_name") or ""
        if not supplier_name:
            report.skipped_rows += 1
            report.errors.append(f"第 {i + 1} 行: 供应商名为空")
            continue
        note_no = projected.get("note_no") or f"__NO_NOTE_NO__{i}"
        key = (supplier_name, str(note_no))
        parent = {k: projected.get(k) for k in schema["parent_fields"]}
        line = {k: v for k, v in projected.items() if k not in schema["parent_fields"]}
        groups.setdefault(key, []).append((parent, line))

    for (sup_name, note_no), entries in groups.items():
        supplier = _get_or_create_supplier(
            db, sup_name, auto_create=auto_create_suppliers,
            created_log=report.auto_created_suppliers,
        )
        if supplier is None:
            report.skipped_rows += len(entries)
            report.errors.append(f"供应商 {sup_name!r} 不存在 (未启用自动创建)")
            continue
        # parent 取第一行的 (delivery_date / total_amount / status / remark)
        first_parent = entries[0][0]
        # 重复单号 → 对比字段决定是否更新
        existing = db.execute(select(DeliveryNote).where(
            DeliveryNote.supplier_id == supplier.id,
            DeliveryNote.note_no == (note_no if not note_no.startswith("__NO_NOTE_NO__") else None),
        )).scalar_one_or_none() if not note_no.startswith("__NO_NOTE_NO__") else None
        if existing is not None:
            first_parent = entries[0][0]
            dn_payload = {k: v for k, v in first_parent.items() if v is not None}
            ctx = _GenericCtx(report=report, on_conflict=on_conflict)
            action = _apply_update(existing, dn_payload, ctx, "delivery_notes", note_no, db)
            if action == "conflict":
                report.skipped_rows += len(entries)
            continue

        first_parent = entries[0][0]
        n = DeliveryNote(
            supplier_id=supplier.id,
            note_no=None if note_no.startswith("__NO_NOTE_NO__") else note_no,
            delivery_date=first_parent.get("delivery_date"),
            total_amount=first_parent.get("total_amount"),
            status=first_parent.get("status") or "confirmed",  # 历史数据默认已确认
            remark=first_parent.get("remark"),
            ocr_warnings=[], ocr_model="excel_import",
        )
        db.add(n)
        db.flush()
        if import_batch_id:
            n.import_job_id = import_batch_id
        report.inserted_parents += 1

        line_total = Decimal("0")
        for j, (_, line_d) in enumerate(entries, start=1):
            qty = line_d.get("qty") or Decimal("0")
            unit_price = line_d.get("unit_price")
            amount = line_d.get("amount")
            if amount is None and unit_price is not None and qty:
                amount = (unit_price * qty).quantize(Decimal("0.01"))
            line = DeliveryNoteLine(
                delivery_note_id=n.id, line_no=j,
                item_name=line_d.get("item_name") or "",
                spec=line_d.get("spec") or "",
                unit=line_d.get("unit") or "",
                qty=qty,
                unit_price=unit_price,
                amount=amount,
                ocr_raw_text="", ocr_warnings=[],
            )
            if auto_match_orders:
                try:
                    cands = delivery_matcher.match_line(
                        db, item_name=line.item_name, spec=line.spec, qty=qty,
                        delivery_date=n.delivery_date,
                        enable_ai_tiebreaker=False,  # 导入量大时跳 AI 省钱
                    )
                    delivery_matcher.apply_candidates_to_line(line, cands)
                    if line.matched_order_no:
                        report.matched_lines += 1
                except Exception as e:  # pragma: no cover
                    line.remark = f"匹配失败: {e}"
            if import_batch_id:
                line.import_job_id = import_batch_id
            db.add(line)
            report.inserted_children += 1
            if amount is not None:
                line_total += amount
        # 没读到 total_amount 的话, 用行金额求和
        if n.total_amount is None and line_total > 0:
            n.total_amount = line_total


# ----------------------------- factory_order --------------------- #


def _next_internal_order_no(db: Session, year: int) -> str:
    """生成下一个内部单号, 格式: Panse{YYYY}{NNNN}, 如 Panse20260001."""
    from sqlalchemy import func, text
    prefix = f"Panse{year}"
    # 查找当年最大序号
    result = db.execute(
        select(FactoryOrder.internal_order_no).where(
            FactoryOrder.internal_order_no.like(f"{prefix}%")
        ).order_by(FactoryOrder.internal_order_no.desc()).limit(1)
    ).scalar_one_or_none()
    if result:
        try:
            seq = int(result[len(prefix):]) + 1
        except (ValueError, IndexError):
            seq = 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _commit_factory_orders(
    db: Session, *, rows: list[dict], mapping: dict[str, str], report: ImportReport,
    on_conflict: str = "ask",
    progress_callback: Optional[ProgressCallback] = None,
    cancel_callback: Optional[CancelCallback] = None,
    import_batch_id: Optional[int] = None,
) -> None:
    schema = get_schema("factory_order")
    total = len(rows)
    current_year = datetime.now().year
    for i, raw_row in enumerate(rows, start=1):
        if i % _PROGRESS_TICK == 0:
            if progress_callback:
                progress_callback(i, total)
            if cancel_callback and cancel_callback():
                raise CancelledImport(f"用户取消, 已处理 {i}/{total} 行")
        projected, errs = _project(raw_row, mapping, schema)
        if errs:
            report.errors.append(f"第 {i + 1} 行: " + "; ".join(errs))
            report.skipped_rows += 1
            continue
        from app.services import factory_order_service
        fo_no = projected.get("factory_order_no")
        if not fo_no:
            # 留空 → 自动生成 畔色0001 序列 (业务确认用中文)
            fo_no = factory_order_service.next_factory_order_no(db)
        # 自动生成内部单号 (若行内未提供)
        internal_no = projected.get("internal_order_no") or _next_internal_order_no(db, current_year)
        qty = int(projected.get("qty") or 1)
        product_code = projected.get("product_code")
        sku = projected.get("sku")
        # 产品预期金额: 行内有就用, 否则按定价表「总出厂成本」自动算
        expected_amount = projected.get("expected_amount")
        if expected_amount is None:
            expected_amount = factory_order_service.expected_amount_for(
                db, product_code, sku, qty)
        # 付款状态: 有支付宝流水号即视为已付款 (与对账逻辑一致)
        flow_no = projected.get("alipay_flow_no")
        payment_status = projected.get("payment_status")
        if not payment_status:
            payment_status = "paid" if flow_no else "unpaid"
        existing = db.execute(
            select(FactoryOrder).where(FactoryOrder.factory_order_no == fo_no)
        ).scalar_one_or_none()
        if existing is not None:
            fo_payload = {
                "factory_order_no": fo_no, "internal_order_no": internal_no,
                "platform_order_no": projected.get("platform_order_no"),
                "factory_name": projected.get("factory_name") or "玉山县博冠家具有限公司",
                "order_date": projected.get("order_date"),
                "expected_delivery": projected.get("expected_delivery"),
                "actual_delivery": projected.get("actual_delivery"),
                "product_code": product_code, "sku": sku, "qty": qty,
                "unit_price": projected.get("unit_price"),
                "factory_bill_amount": projected.get("factory_bill_amount"),
                "expected_amount": expected_amount,
                "payment_method": projected.get("payment_method") or "月结",
                "payment_status": payment_status,
                "payment_date": projected.get("payment_date"),
                "carrier": projected.get("carrier"),
                "tracking_no": projected.get("tracking_no"),
                "alipay_flow_no": flow_no,
                "remark": projected.get("remark"),
            }
            ctx = _GenericCtx(report=report, on_conflict=on_conflict, import_batch_id=import_batch_id)
            action = _apply_update(existing, {k: v for k, v in fo_payload.items() if v is not None},
                                   ctx, "factory_orders", fo_no, db)
            if action == "conflict":
                report.skipped_rows += 1
            continue
        fo = FactoryOrder(
            factory_order_no=fo_no,
            internal_order_no=internal_no,
            platform_order_no=projected.get("platform_order_no"),
            factory_name=projected.get("factory_name") or "玉山县博冠家具有限公司",
            order_date=projected.get("order_date"),
            expected_delivery=projected.get("expected_delivery"),
            actual_delivery=projected.get("actual_delivery"),
            product_code=product_code,
            sku=sku,
            qty=qty,
            unit_price=projected.get("unit_price"),
            factory_bill_amount=projected.get("factory_bill_amount"),
            expected_amount=expected_amount,
            payment_method=projected.get("payment_method") or "月结",
            payment_status=payment_status,
            payment_date=projected.get("payment_date"),
            carrier=projected.get("carrier"),
            tracking_no=projected.get("tracking_no"),
            alipay_flow_no=flow_no,
            remark=projected.get("remark"),
        )
        db.add(fo)
        db.flush()  # flush so next _next_internal_order_no can see this row
        if import_batch_id:
            fo.import_job_id = import_batch_id
        report.inserted_parents += 1


# ----------------------------- alipay_flow ----------------------- #


def _commit_alipay_flows(
    db: Session, *, rows: list[dict], mapping: dict[str, str], report: ImportReport,
    sheet_account: Optional[str] = None,
    on_conflict: str = "ask",
    progress_callback: Optional[ProgressCallback] = None,
    cancel_callback: Optional[CancelCallback] = None,
    import_batch_id: Optional[int] = None,
) -> None:
    """支付宝流水: (account, transaction_no) 唯一. 自动跑 smart_matching_service.run()
    给新进来的流水打标签 (factory_payment/promotion/etc).

    sheet_account: 当 sheet 没有账户列 (账户名写在 sheet 名/标题) 时, 用它填充每行账户。
    """
    schema = get_schema("alipay_flow")
    fresh_ids: list[int] = []
    total = len(rows)
    counterparty_filled = 0   # 交易对象为空被置'待确认'的行数
    ctx = _GenericCtx(report=report, on_conflict=on_conflict, import_batch_id=import_batch_id)
    for i, raw_row in enumerate(rows, start=1):
        if i % _PROGRESS_TICK == 0:
            if progress_callback:
                progress_callback(i, total)
            if cancel_callback and cancel_callback():
                raise CancelledImport(f"用户取消, 已处理 {i}/{total} 行")
        projected, errs = _project(raw_row, mapping, schema)
        if errs:
            report.errors.append(f"第 {i + 1} 行: " + "; ".join(errs))
            report.skipped_rows += 1
            continue
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in projected.values()):
            continue
        # 清洗: 流水号/关联订单号去全部空格; 交易对象(公司名)/交易账户(邮箱)去内部空格
        projected["transaction_no"] = _strip_all_ws(projected.get("transaction_no"))
        projected["related_order_no"] = _strip_all_ws(projected.get("related_order_no"))
        if projected.get("counterparty"):
            projected["counterparty"] = _strip_all_ws(projected["counterparty"])
        if projected.get("counterparty_account"):
            projected["counterparty_account"] = _strip_all_ws(projected["counterparty_account"])
        account = projected.get("account") or sheet_account
        tx_no = projected.get("transaction_no")
        amount = projected.get("amount")
        if not account or not tx_no or amount is None:
            report.skipped_rows += 1
            report.errors.append(f"第 {i + 1} 行: 账户/流水号/金额 任一为空")
            continue
        # 交易对象为空 → 置'待确认' (爱群号等待补填的表), 汇总计数避免逐行刷异常
        if not projected.get("counterparty"):
            projected["counterparty"] = "待确认"
            counterparty_filled += 1
        existing = db.execute(
            select(AlipayFlow).where(
                AlipayFlow.account == account, AlipayFlow.transaction_no == tx_no,
            )
        ).scalar_one_or_none()
        if existing is not None:
            flow_payload = {
                "account": account, "transaction_no": tx_no,
                "transaction_time": projected.get("transaction_time"),
                "transaction_type": projected.get("transaction_type"),
                "counterparty": projected.get("counterparty"),
                "counterparty_account": projected.get("counterparty_account"),
                "amount": amount,
                "balance": projected.get("balance"),
                "related_order_no": projected.get("related_order_no"),
                "remark": projected.get("remark"),
            }
            action = _apply_update(existing, flow_payload, ctx, "alipay_flows", tx_no, db)
            if action not in ("updated", "conflict"):
                pass  # skipped (identical)
            else:
                report.skipped_rows += 1 if action == "conflict" else 0
            continue
        flow = AlipayFlow(
            account=account, transaction_no=tx_no,
            transaction_time=projected.get("transaction_time"),
            transaction_type=projected.get("transaction_type"),
            counterparty=projected.get("counterparty"),
            counterparty_account=projected.get("counterparty_account"),
            amount=amount,
            balance=projected.get("balance"),
            related_order_no=projected.get("related_order_no"),
            remark=projected.get("remark"),
            reconciliation_status="open",
        )
        db.add(flow)
        db.flush()
        if import_batch_id:
            flow.import_job_id = import_batch_id
        fresh_ids.append(flow.id)
        report.inserted_parents += 1

    if counterparty_filled:
        report.warnings.append(f"{counterparty_filled} 条流水交易对象为空, 已统一置'待确认', 请后续核对补填")

    # 入完一次性跑智能标签 (factory_payment / promotion / logistics / salary)
    if fresh_ids:
        try:
            from app.services.smart_matching_service import run as smart_tag
            tag_result = smart_tag(db)
            tagged_total = sum(tag_result.tagged.values())
            if tagged_total > 0:
                report.warnings.append(
                    f"已为 {tagged_total} 条新流水自动打标签: "
                    + ", ".join(f"{k}={v}" for k, v in tag_result.tagged.items())
                )
        except Exception as e:  # pragma: no cover
            report.warnings.append(f"自动打标签失败 (不影响入库): {e}")


# ----------------------------- 通用 entity 入库 (扩展) ----------------- #


@dataclass
class _GenericCtx:
    """传给 generic handler 的上下文 (冲突策略 + 收集差异)."""
    report: ImportReport
    on_conflict: str = "overwrite"   # overwrite / keep / ask
    import_batch_id: Optional[int] = None


def _jsonable(v: Any) -> Any:
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    return v


def _diff_fields(existing, payload: dict) -> list[dict]:
    """对比已有记录与新值, 只返回真正变化的字段 [{field, old, new}]."""
    diffs: list[dict] = []
    for f, new in payload.items():
        if not hasattr(existing, f):
            continue
        old = getattr(existing, f)
        if isinstance(old, Decimal) and new is not None:
            try:
                if Decimal(str(new)) == old:
                    continue
            except (InvalidOperation, ValueError):
                pass
        if old == new:
            continue
        diffs.append({"field": f, "old": _jsonable(old), "new": _jsonable(new)})
    return diffs


def _apply_update(existing, payload: dict, ctx: Optional["_GenericCtx"],
                  source_table: str, source_pk, db=None) -> str:
    """upsert 命中已有记录时的统一处理: 比 diff + 按 on_conflict 决定覆盖/记冲突."""
    diffs = _diff_fields(existing, payload)
    if not diffs:
        return "skipped"   # 值完全相同, 不算更新
    if ctx is None or ctx.on_conflict == "overwrite":
        for f, v in payload.items():
            if hasattr(existing, f):
                setattr(existing, f, v)
        return "updated"
    if ctx.on_conflict == "keep":
        return "skipped"   # 保留原值, 不动
    # ask: 记录差异到 report.conflicts + DataException, 不覆盖, 等用户裁决
    pk_str = str(source_pk) if source_pk is not None else None
    diff_summary = "; ".join(
        f"{d['field']}: {d['old']!r} → {d['new']!r}" for d in diffs[:5]
    )
    ctx.report.conflicts.append({
        "source_table": source_table,
        "source_pk": pk_str,
        "diffs": diffs,
    })
    if db is not None:
        from app.services import exception_service
        exception_service.record(
            db,
            source_table=source_table,
            source_pk=pk_str,
            exception_type="import_conflict",
            severity="warning",
            description=f"导入数据与已有记录不同，需确认使用哪个版本。差异：{diff_summary}",
            suggestion_action="resolve_import_conflict",
            context={
                "diffs": diffs,
                "new_values": {k: _jsonable(v) for k, v in payload.items()},
            },
        )
    return "conflict"


def _commit_generic(
    db: Session, *, rows: list[dict], mapping: dict[str, str],
    entity_type: str, report: ImportReport,
    on_conflict: str = "overwrite",
    progress_callback: Optional[ProgressCallback] = None,
    cancel_callback: Optional[CancelCallback] = None,
    import_batch_id: Optional[int] = None,
) -> None:
    """7 类简单 entity (产品/物料/BOM/库存/订单/账户余额) 统一入库.

    每行 = 一条记录, 按"唯一字段"去重 upsert.
    on_conflict 控制重导命中已有记录且值不同时的行为 (见 commit_sheet).
    """
    schema = get_schema(entity_type)
    total = len(rows)
    uniqueness_field = _UNIQUENESS_FIELD.get(entity_type)
    ctx = _GenericCtx(report=report, on_conflict=on_conflict, import_batch_id=import_batch_id)
    inserted, updated, skipped, conflicted = 0, 0, 0, 0
    # 合并单元格向下填充: 某些表 (如工厂对账) 只在首行写工厂名, 其余行因合并单元格读出空,
    # 这里按上一非空行的值补齐, 还原合并单元格的真实语义.
    ff_fields = _FORWARD_FILL.get(entity_type, [])
    ff_last: dict[str, Any] = {}
    # 同一次提交内, 源表常有重复唯一键 (如产品总表按 SKU 逐行列同一产品码).
    # 全局 autoflush=False 会让逐行 select 看不到本批已 add 但未 flush 的行,
    # 导致重复键堆到最后一次 flush 撞 UNIQUE 约束崩溃. 此处临时开 autoflush,
    # 让每行的存在性检查先 flush 上一行, 重复键即走正常 upsert 路径.
    prev_autoflush = db.autoflush
    db.autoflush = True
    try:
        for i, raw_row in enumerate(rows, start=1):
            if i % _PROGRESS_TICK == 0:
                if progress_callback:
                    progress_callback(i, total)
                if cancel_callback and cancel_callback():
                    raise CancelledImport(f"用户取消, 已处理 {i}/{total} 行")
            projected, errs = _project(raw_row, mapping, schema)
            if errs:
                report.errors.append(f"第 {i + 1} 行: " + "; ".join(errs))
                skipped += 1
                continue
            # 跳过 MAPPED 列全为空的行 (模板行/辅助列行):
            # 有些表末尾有大量「导入校验=✅」但业务列全空的占位行, 避免误报必填字段缺失。
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in projected.values()):
                continue
            for f in ff_fields:
                if projected.get(f):
                    ff_last[f] = projected[f]
                elif ff_last.get(f):
                    projected[f] = ff_last[f]
            try:
                kind, did = _GENERIC_HANDLERS[entity_type](db, projected, uniqueness_field, ctx)
            except Exception as e:
                report.errors.append(f"第 {i + 1} 行: {type(e).__name__}: {e}")
                skipped += 1
                continue
            if did == "inserted":
                inserted += 1
            elif did == "updated":
                updated += 1
            elif did == "conflict":
                conflicted += 1
            else:
                skipped += 1
    finally:
        db.autoflush = prev_autoflush
    report.inserted_parents += inserted
    report.skipped_rows += skipped + conflicted
    if updated > 0:
        report.warnings.append(f"{updated} 行更新已有记录")
    if conflicted > 0:
        report.warnings.append(f"{conflicted} 行与库内数据不同, 待人工确认 (见 conflicts)")


# 需要合并单元格向下填充的字段 (entity_type -> [字段名])
_FORWARD_FILL = {
    "factory_reconciliation": ["factory_name"],
}


_UNIQUENESS_FIELD = {
    "product": "code",
    "material": "code",
    "product_inventory": None,         # (warehouse, product_code, sku)
    "part_inventory": None,             # (warehouse, material_code)
    "bom_line": None,                   # 无 upsert, 直接 add
    "order": "order_no",
    "account_balance": None,            # (account_name, year, month)
    "pricing_sku": "sku_code",
    "refill_record": "order_no",
    "factory_reconciliation": None,     # (factory_name, period_end)
    "outsourcing_expense": None,        # alipay_flow_no 去重
    "aftersales": "platform_order_no",
    "competitor_price": None,           # (store, sku_name) 去重
    "daily_operations": None,           # 无唯一键, 直接 add
    "order_details": None,              # 无唯一键, 直接 add
    "wood_loss": None,                  # 无唯一键, 直接 add
    "sample": "sample_no",
    "promotion_flow": None,             # 无唯一键, 直接 add
}


_WS_RE = re.compile(r"\s+")
_NEWLINE_RE = re.compile(r"[\r\n]+")


def _strip_all_ws(v: Any) -> Optional[str]:
    """去除字符串中全部空白 (含内部空格/制表/换行). 用于流水号/订单号/邮箱等不应含空格的字段."""
    if v is None:
        return None
    s = _WS_RE.sub("", str(v))
    return s or None


def _join_multiline(v: Any, sep: str = ",") -> Optional[str]:
    """单元格内多行 (换行分隔多个值) → 用 sep 拼成一个值. 用于补发物流单号等."""
    if v is None:
        return None
    parts = [p.strip() for p in _NEWLINE_RE.split(str(v)) if p.strip()]
    return sep.join(parts) or None


def _clean_text_no(v: Any) -> Optional[str]:
    """数字型账号/编号 → 干净文本 (去掉浮点尾巴 .0, 防止手机号被存成科学计数/丢精度)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s or None


def _sum_opt(*vals: Any) -> Optional[Decimal]:
    """对若干可选 Decimal 求和; 全部为 None 时返回 None (不强行写 0)."""
    nums = [v for v in vals if v is not None]
    if not nums:
        return None
    total = Decimal("0")
    for n in nums:
        total += Decimal(str(n))
    return total


def _try_parse_date(v: Any) -> Optional[date]:
    """尽力把值解析成 date; 失败返回 None (不抛异常). 用于'备注列误填日期'的搬移."""
    if v is None or v == "":
        return None
    try:
        return _coerce(v, "date", label="date")
    except ImporterError:
        return None


def _normalize_product(data: dict) -> dict:
    """导入前规范化: 统一空值填充 + 产品名称拆分.

    规则 (均为用户在导入分析中确认的业务规则):
      - taobao_sku_id 空 → '待定'
      - accessory_desc / accessory_remark 空 → '-'
      - size_value 空 → '待定'
      - size_confirmed 空 → '待确定'
      - name 含竖线 (|/丨) → 拆成 name + sub_name
    """
    d = dict(data)
    # 空值填充
    if not d.get("taobao_sku_id"):
        d["taobao_sku_id"] = "待定"
    if not d.get("accessory_desc"):
        d["accessory_desc"] = "-"
    if not d.get("accessory_remark"):
        d["accessory_remark"] = "-"
    if not d.get("size_value"):
        d["size_value"] = "待定"
    if not d.get("size_confirmed"):
        d["size_confirmed"] = "待确定"
    # 产品名称拆分: "肤色榉木无边床丨榉木主属腿床" → name + sub_name
    name = d.get("name") or ""
    for sep in ("|", "丨", "｜"):
        if sep in name:
            parts = name.split(sep, 1)
            d["name"] = parts[0].strip()
            d["sub_name"] = parts[1].strip()
            break
    return d


def _h_product(db, data, key_field, ctx=None):
    from app.models.product import Product
    from app.models.pricing import PricingSku

    code = data.get("code")
    if not code:
        raise ImporterError("缺产品编码")

    data = _normalize_product(data)
    sku_code = data.get("sku_code")
    sku = data.get("sku")

    # 只写 Product 模型认识的字段 (过滤掉 pricing_sku 专属字段等)
    _PRODUCT_FIELDS = {c.key for c in Product.__table__.columns}
    payload = {k: v for k, v in data.items() if k in _PRODUCT_FIELDS and v is not None}

    existing = db.execute(select(Product).where(Product.code == code)).scalar_one_or_none()
    if existing:
        action = _apply_update(existing, payload, ctx, "products", code, db)
    else:
        db.add(Product(**payload))
        action = "inserted"

    # 若行内含 sku_code → 同步 upsert pricing_sku (整张产品总表含 SKU 列时触发)
    if sku_code:
        ps_existing = db.execute(
            select(PricingSku).where(PricingSku.sku_code == sku_code)
        ).scalar_one_or_none()
        ps_payload: dict = {"product_code": code, "sku_code": sku_code}
        if sku:
            ps_payload["sku"] = sku
        if ps_existing:
            for k, v in ps_payload.items():
                if v is not None:
                    setattr(ps_existing, k, v)
        else:
            if _is_custom_code(db, sku_code, code):
                _flag_custom(db, "pricing_sku", sku_code, sku_code)
            db.add(PricingSku(**ps_payload))

    return "product", action


def _h_material(db, data, key_field, ctx=None):
    from app.models.material import Material
    code = data.get("code")
    name = data.get("name") or code
    if not code:
        raise ImporterError("缺物料编码")
    existing = db.execute(select(Material).where(Material.code == code)).scalar_one_or_none()
    if existing:
        return "material", _apply_update(
            existing, {k: v for k, v in data.items() if v is not None}, ctx, "materials", code, db)
    # name UNIQUE: 重名加 code 后缀
    same_name = db.execute(select(Material).where(Material.name == name)).scalar_one_or_none()
    if same_name and same_name.code != code:
        name = f"{name} ({code})"
    payload = {k: v for k, v in data.items() if v is not None}
    payload["name"] = name
    db.add(Material(**payload))
    return "material", "inserted"


def _h_bom(db, data, key_field, ctx=None):
    from app.models.bom import BomLine
    from app.models.material import Material
    product_code = data.get("product_code")
    material_code = data.get("material_code")
    if not product_code or not material_code:
        raise ImporterError("缺 product_code 或 material_code")
    # 物料不存在 → 自动建占位
    if not db.execute(select(Material).where(Material.code == material_code)).scalar_one_or_none():
        db.add(Material(code=material_code, name=f"占位 ({material_code})"))
        db.flush()
    db.add(BomLine(**{k: v for k, v in data.items() if v is not None}))
    return "bom_line", "inserted"


def _h_product_inv(db, data, key_field, ctx=None):
    from app.models.inventory import ProductInventory
    warehouse = data.get("warehouse") or "江西仓库"
    product_code = data.get("product_code")
    sku = data.get("sku")
    if not product_code:
        raise ImporterError("缺 product_code")
    existing = db.execute(select(ProductInventory).where(
        ProductInventory.warehouse == warehouse,
        ProductInventory.product_code == product_code,
        ProductInventory.sku == sku,
    )).scalar_one_or_none()
    payload = {k: v for k, v in data.items() if v is not None}
    payload["warehouse"] = warehouse
    payload.setdefault("locked_qty", 0)   # 锁定库存默认 0，避免 NULL
    if existing:
        return "product_inv", _apply_update(
            existing, payload, ctx, "product_inventory",
            f"{warehouse}|{product_code}|{sku}", db)
    db.add(ProductInventory(**payload))
    return "product_inv", "inserted"


def _h_part_inv(db, data, key_field, ctx=None):
    from app.models.inventory import PartInventory
    from app.models.material import Material
    warehouse = data.get("warehouse") or "江西仓库"
    material_code = data.get("material_code")
    if not material_code:
        # 配件编码为空时: 尝试用名称从 Material 表反查编码, 否则用名称拼占位编码
        name = data.get("material_name") or data.get("spec") or ""
        if name:
            mat = db.execute(select(Material).where(Material.name == name)).scalar_one_or_none()
            material_code = mat.code if mat else f"TMP-{name[:12]}"
        else:
            raise ImporterError("缺 material_code")
    if not db.execute(select(Material).where(Material.code == material_code)).scalar_one_or_none():
        db.add(Material(code=material_code, name=f"占位 ({material_code})"))
        db.flush()
    existing = db.execute(select(PartInventory).where(
        PartInventory.warehouse == warehouse,
        PartInventory.material_code == material_code,
    )).scalar_one_or_none()
    payload = {k: v for k, v in data.items() if v is not None}
    payload["warehouse"] = warehouse
    if existing:
        return "part_inv", _apply_update(
            existing, payload, ctx, "part_inventory", f"{warehouse}|{material_code}", db)
    db.add(PartInventory(**payload))
    return "part_inv", "inserted"


def _is_custom_code(db, sku_code, product_code) -> bool:
    """定制编码识别: SKU 尾部后缀 >= 阈值 (默认 90, 含 99/98/97)."""
    if not sku_code:
        return False
    from app.services import sku_utils
    threshold = db.info.setdefault("_custom_sku_threshold", sku_utils.get_threshold(db))
    return sku_utils.is_custom_sku_code(sku_code, product_code, threshold)


def _flag_custom(db, source_table: str, source_pk, sku_code) -> None:
    from app.services import exception_service
    exception_service.record(
        db,
        source_table=source_table,
        source_pk=str(source_pk) if source_pk is not None else None,
        exception_type="custom_sku_detected",
        severity="info",
        description=f"识别到定制编码 {sku_code} (后缀达定制阈值), 已自动标记定制, 请复核.",
        suggestion_action="view",
        context={"sku_code": sku_code},
    )


def _record_exc(db, source_table: str, source_pk, exc_type: str,
                description: str, severity: str = "warning") -> None:
    """导入时记录一条数据异常 (统一入口, 供各 handler 标记问题数据)."""
    from app.services import exception_service
    exception_service.record(
        db,
        source_table=source_table,
        source_pk=str(source_pk) if source_pk is not None else None,
        exception_type=exc_type,
        severity=severity,
        description=description,
        suggestion_action="view",
    )


def _normalize_order(data: dict) -> dict:
    """导入前规范化订单数据.

    规则:
      - platform_fee 向上取整 (ceil), 按业务惯例
      - ship_date 为空但 tracking_no 不为空 → 用 order_date 补填 (已发货)
    """
    d = dict(data)
    # 订单号/快递单号去全部空格 (跨表关联匹配靠它, 空格会导致匹配失败)
    if d.get("order_no"):
        d["order_no"] = _strip_all_ws(d["order_no"])
    if d.get("tracking_no"):
        d["tracking_no"] = _strip_all_ws(d["tracking_no"])
    # 平台服务费向上取整
    if d.get("platform_fee") is not None:
        d["platform_fee"] = Decimal(str(math.ceil(float(d["platform_fee"]))))
    # 发货日期推断: 有运单号但无发货日期 → 用下单日期作为最保守估计
    if not d.get("ship_date") and d.get("tracking_no"):
        if d.get("order_date"):
            d["ship_date"] = d["order_date"]
    return d


def _h_order(db, data, key_field, ctx=None):
    from app.models.order import Order
    data = _normalize_order(data)
    order_no = data.get("order_no")
    if not order_no:
        raise ImporterError("缺 order_no")

    payload = {k: v for k, v in data.items() if v is not None}
    payload.setdefault("platform", "淘宝")
    payload.setdefault("qty", 1)
    payload.setdefault("status", "signed")
    payload.setdefault("is_historical", True)   # 通用导入默认标历史
    # 定制编码自动识别 (后缀 >= 阈值, 如 99/98/97)
    if _is_custom_code(db, payload.get("sku_code"), payload.get("product_code")):
        payload["is_custom"] = True
        _flag_custom(db, "orders", order_no, payload.get("sku_code"))
    # 标记为补单时, 交叉核验补单记录表
    if payload.get("is_refill"):
        from app.models.finance import RefillRecord
        from app.services import exception_service
        refill_exists = db.execute(
            select(RefillRecord).where(RefillRecord.order_no == order_no)
        ).scalar_one_or_none()
        if not refill_exists:
            exception_service.record(
                db,
                source_table="orders",
                source_pk=order_no,
                exception_type="refill_record_missing",
                severity="warning",
                description=f"订单 {order_no} 标记为补单 (是否补单=是), 但补单记录表中未找到对应记录, 请补录。",
                suggestion_action="view",
                context={"order_no": order_no},
            )
    # 发货仓库: 行内未指定时, 样块/补单→杭州, 其余→江西仓库
    if not payload.get("warehouse"):
        from app.services import order_cost_service
        payload["warehouse"] = order_cost_service.default_warehouse_for(
            payload.get("product_name"), payload.get("sku"),
            bool(payload.get("is_refill")))

    existing = db.execute(select(Order).where(Order.order_no == order_no)).scalar_one_or_none()
    if existing:
        return "order", _apply_update(existing, payload, ctx, "orders", order_no, db)

    obj = Order(**payload)
    if ctx and ctx.import_batch_id:
        obj.import_job_id = ctx.import_batch_id
    db.add(obj)
    return "order", "inserted"


def _h_balance(db, data, key_field, ctx=None):
    from app.models.finance import AccountBalance
    from datetime import date as _date, datetime as _dt
    data = dict(data)
    name = data.get("account_name")
    # 账户名整行为空 → 静默跳过 (用户要求, 不报错)
    if not name:
        return "balance", "skipped"
    # 账户号是手机号被存成数字 → 强制干净文本 (去 .0, 防丢精度)
    if data.get("account_no") is not None:
        data["account_no"] = _clean_text_no(data["account_no"])
    year = data.get("year")
    month = data.get("month")
    # 支持 "统计日期" 单列格式：从日期自动提取年月
    period_date = data.get("period_date")
    if period_date and not (year and month):
        if isinstance(period_date, (_date, _dt)):
            year = period_date.year
            month = period_date.month
    if not (name and year and month):
        raise ImporterError("缺账户名/年/月 (或 统计日期)")
    payload = {k: v for k, v in data.items() if v is not None and k != "period_date"}
    payload["period_year"] = payload.pop("year", year)
    payload["period_month"] = payload.pop("month", month)
    existing = db.execute(select(AccountBalance).where(
        AccountBalance.account_name == name,
        AccountBalance.period_year == year,
        AccountBalance.period_month == month,
    )).scalar_one_or_none()
    if existing:
        return "balance", _apply_update(existing, payload, ctx, "account_balances",
                                        f"{name}|{year}|{month}", db)
    db.add(AccountBalance(**payload))
    return "balance", "inserted"


def _h_pricing_sku(db, data, key_field, ctx=None):
    from app.models.pricing import PricingSku
    sku_code = data.get("sku_code")
    if not sku_code:
        raise ImporterError("缺 sku_code")
    existing = db.execute(select(PricingSku).where(
        PricingSku.sku_code == sku_code,
    )).scalar_one_or_none()
    payload = {k: v for k, v in data.items() if v is not None}
    if existing:
        return "pricing_sku", _apply_update(existing, payload, ctx, "pricing_sku", sku_code, db)
    if _is_custom_code(db, sku_code, payload.get("product_code")):
        _flag_custom(db, "pricing_sku", sku_code, sku_code)
    db.add(PricingSku(**payload))
    return "pricing_sku", "inserted"


def _h_refill_record(db, data, key_field, ctx=None):
    from app.models.finance import RefillRecord
    from app.models.order import Order
    order_no = data.get("order_no")
    if not order_no:
        raise ImporterError("缺 order_no")
    # 自动从 orders 表补填 product_code / product_name / sku (若行内未提供)
    data = dict(data)
    if not (data.get("product_code") and data.get("product_name") and data.get("sku")):
        matched = db.execute(
            select(Order).where(Order.order_no == order_no)
        ).scalar_one_or_none()
        if matched:
            data.setdefault("product_code", matched.product_code)
            data.setdefault("product_name", matched.product_name)
            data.setdefault("sku", matched.sku)
    existing = db.execute(select(RefillRecord).where(RefillRecord.order_no == order_no)).scalar_one_or_none()
    payload = {k: v for k, v in data.items() if v is not None}
    if existing:
        return "refill_record", _apply_update(existing, payload, ctx, "refill_records", order_no, db)
    db.add(RefillRecord(**payload))
    return "refill_record", "inserted"


def _h_factory_reconciliation(db, data, key_field, ctx=None):
    from app.models.finance import AlipayFlow, FactoryReconciliation
    factory = data.get("factory_name")
    if not factory:
        raise ImporterError("缺 factory_name")
    data = dict(data)
    # 绑定支付宝流水号 → 自动确认实付/对账日期/对账状态=completed
    flow_no = data.get("alipay_flow_no")
    if flow_no:
        flow = db.execute(
            select(AlipayFlow).where(AlipayFlow.transaction_no == flow_no)
        ).scalar_one_or_none()
        if flow is not None:
            if data.get("paid_amount") is None and flow.amount is not None:
                data["paid_amount"] = abs(flow.amount)
            if data.get("reconciled_at") is None and flow.transaction_time is not None:
                data["reconciled_at"] = flow.transaction_time.date()
            data.setdefault("status", "completed")
        else:
            _record_exc(db, "factory_reconciliations", flow_no, "flow_not_found",
                        f"工厂对账绑定的支付宝流水号 {flow_no} 在流水表中未找到, 请核对.", "warning")
    # 差异金额缺省 0 (模型已有 default, 这里显式兜底)
    data.setdefault("diff_amount", Decimal("0"))
    period_end = data.get("period_end")
    existing = db.execute(
        select(FactoryReconciliation).where(
            FactoryReconciliation.factory_name == factory,
            FactoryReconciliation.period_end == period_end,
        )
    ).scalar_one_or_none() if period_end else None
    payload = {k: v for k, v in data.items() if v is not None}
    if existing:
        return "factory_reconciliation", _apply_update(
            existing, payload, ctx, "factory_reconciliations", f"{factory}|{period_end}", db)
    rec = FactoryReconciliation(**payload)
    if ctx and ctx.import_batch_id:
        rec.import_job_id = ctx.import_batch_id
    db.add(rec)
    return "factory_reconciliation", "inserted"


def _h_outsourcing_expense(db, data, key_field, ctx=None):
    from app.models.marketing import OutsourcingExpense
    amount = data.get("amount")
    if amount is None:
        raise ImporterError("缺 amount")
    data = dict(data)
    # 支付日期误填在备注列 → 搬到 payment_date, 备注清空
    if not data.get("payment_date") and data.get("remark"):
        moved = _try_parse_date(data["remark"])
        if moved is not None:
            data["payment_date"] = moved
            data["remark"] = None
    # 关联平台订单号为空: 工资类统一填 N/A; 其余报异常待补
    if not data.get("related_order_no"):
        proj = f"{data.get('project') or ''}{data.get('cost_category') or ''}"
        if "工资" in proj:
            data["related_order_no"] = "N/A"
        else:
            _record_exc(db, "outsourcing_expenses", data.get("alipay_flow_no") or data.get("payee"),
                        "related_order_no_missing",
                        f"外包费用 (收款人 {data.get('payee')}) 关联平台订单号为空, 请补填以便成本追溯.", "info")
    payload = {k: v for k, v in data.items() if v is not None}
    # 支付宝流水号为空 → 报异常待回填
    if not data.get("alipay_flow_no"):
        _record_exc(db, "outsourcing_expenses", data.get("payee"),
                    "alipay_flow_no_missing",
                    f"外包费用 (收款人 {data.get('payee')}) 支付宝流水号为空, 无法与流水对账, 请回填.", "warning")
    # OutsourcingExpense 无天然唯一键 → 按 alipay_flow_no 去重(若有)
    flow_no = payload.get("alipay_flow_no")
    if flow_no:
        existing = db.execute(
            select(OutsourcingExpense).where(OutsourcingExpense.alipay_flow_no == flow_no)
        ).scalar_one_or_none()
        if existing:
            return "outsourcing_expense", _apply_update(
                existing, payload, ctx, "outsourcing_expenses", flow_no, db)
    db.add(OutsourcingExpense(**payload))
    return "outsourcing_expense", "inserted"


def _h_aftersales(db, data, key_field, ctx=None):
    from app.models.marketing import AfterSales
    order_no = data.get("platform_order_no")
    if not order_no:
        raise ImporterError("缺 platform_order_no")
    data = dict(data)
    # 补发物流单号: 单元格内含换行/多个单号 → 英文逗号拼接
    if data.get("refill_tracking_no"):
        data["refill_tracking_no"] = _join_multiline(data["refill_tracking_no"])
    # 售后成本公式自动核算 (为空时):
    #   平台内售后总成本 = 订单赔付费 + 好评/差价返
    #   平台外售后总成本 = 直接赔付客户 + 二次上门维修费 + 返厂打包运费
    if data.get("in_platform_total") is None:
        v = _sum_opt(data.get("compensation_fee"), data.get("good_review_refund"))
        if v is not None:
            data["in_platform_total"] = v
    if data.get("out_platform_total") is None:
        v = _sum_opt(data.get("direct_compensation"), data.get("second_visit_fee"),
                     data.get("return_pack_freight"))
        if v is not None:
            data["out_platform_total"] = v
    existing = db.execute(
        select(AfterSales).where(AfterSales.platform_order_no == order_no)
    ).scalar_one_or_none()
    payload = {k: v for k, v in data.items() if v is not None}
    # 支付宝流水号为空 → 报异常, 让同事回填 (售后表行数有限, 逐行标记可接受)
    if not data.get("alipay_flow_no"):
        _record_exc(db, "after_sales", order_no, "alipay_flow_no_missing",
                    f"售后单 {order_no} 支付宝流水号为空, 无法与流水核销, 请回填.", "warning")
    if existing:
        return "aftersales", _apply_update(existing, payload, ctx, "after_sales", order_no, db)
    db.add(AfterSales(**payload))
    return "aftersales", "inserted"


def _h_competitor(db, data, key_field, ctx=None):
    from app.models.competitor import CompetitorPrice
    sku_name = data.get("sku_name")
    if not sku_name:
        raise ImporterError("缺 SKU 名")
    payload = {k: v for k, v in data.items() if v is not None}
    # 按 (store, sku_name) 去重 upsert
    store = data.get("store")
    existing = db.execute(
        select(CompetitorPrice).where(
            CompetitorPrice.store == store, CompetitorPrice.sku_name == sku_name
        )
    ).scalar_one_or_none()
    if existing:
        return "competitor_price", _apply_update(existing, payload, ctx, "competitor_prices", sku_name, db)
    db.add(CompetitorPrice(**payload))
    return "competitor_price", "inserted"


def _h_daily_operation(db, data, key_field, ctx=None):
    from app.models.marketing import DailyOperation
    payload = {k: v for k, v in data.items() if v is not None}
    if not payload:
        return "daily_operation", "skipped"
    db.add(DailyOperation(**payload))
    return "daily_operation", "inserted"


def _h_order_detail(db, data, key_field, ctx=None):
    from app.models.order import OrderDetail
    order_no = data.get("order_no")
    factory_order_no = data.get("factory_order_no")
    product_code = data.get("product_code")
    if not (order_no or factory_order_no or product_code):
        return "order_detail", "skipped"
    payload = {k: v for k, v in data.items() if v is not None}
    db.add(OrderDetail(**payload))
    return "order_detail", "inserted"


def _h_wood_loss(db, data, key_field, ctx=None):
    from app.models.marketing import WoodLoss
    payload = {k: v for k, v in data.items() if v is not None}
    if not payload:
        return "wood_loss", "skipped"
    db.add(WoodLoss(**payload))
    return "wood_loss", "inserted"


def _h_promotion_flow(db, data, key_field, ctx=None):
    from app.models.marketing import PromotionFlow
    data = dict(data)
    # 流水号去全部空格
    if data.get("alipay_flow_no"):
        data["alipay_flow_no"] = _strip_all_ws(data["alipay_flow_no"])
        if data["alipay_flow_no"] is None:
            data.pop("alipay_flow_no")
    # 无流水号 → 现金消耗 (用户规则)
    if not data.get("alipay_flow_no"):
        data["alipay_flow_no"] = "现金消耗"
    payload = {k: v for k, v in data.items() if v is not None}
    if not payload:
        return "promotion_flow", "skipped"
    payload.setdefault("amount", 0)   # amount NOT NULL
    db.add(PromotionFlow(**payload))
    return "promotion_flow", "inserted"


def _h_sample(db, data, key_field, ctx=None):
    from app.models.marketing import Sample
    sample_no = data.get("sample_no")
    if not sample_no:
        raise ImporterError("缺样品编号")
    data = dict(data)
    # 用途为空 → 统一 '待处理'
    if not data.get("usage"):
        data["usage"] = "待处理"
    # 产品名称用斜杠拼了多个 (一个样品含多把椅子混放) → 取第一个为主名, 其余进备注
    name = data.get("product_name") or ""
    if "/" in name:
        parts = [p.strip() for p in name.split("/") if p.strip()]
        if parts:
            data["product_name"] = parts[0]
            if len(parts) > 1:
                extra = "含: " + "/".join(parts[1:])
                data["remark"] = f"{data['remark']}; {extra}" if data.get("remark") else extra
    from sqlalchemy import select as _select
    existing = db.execute(_select(Sample).where(Sample.sample_no == sample_no)).scalar_one_or_none()
    payload = {k: v for k, v in data.items() if v is not None}
    if existing:
        return "sample", _apply_update(existing, payload, ctx, "samples", sample_no, db)
    db.add(Sample(**payload))
    return "sample", "inserted"


_GENERIC_HANDLERS = {
    "product": _h_product, "material": _h_material, "bom_line": _h_bom,
    "product_inventory": _h_product_inv, "part_inventory": _h_part_inv,
    "order": _h_order, "account_balance": _h_balance,
    "pricing_sku": _h_pricing_sku,
    "refill_record": _h_refill_record,
    "factory_reconciliation": _h_factory_reconciliation,
    "outsourcing_expense": _h_outsourcing_expense,
    "aftersales": _h_aftersales,
    "competitor_price": _h_competitor,
    "daily_operations": _h_daily_operation,
    "order_details": _h_order_detail,
    "wood_loss": _h_wood_loss,
    "sample": _h_sample,
    "promotion_flow": _h_promotion_flow,
}

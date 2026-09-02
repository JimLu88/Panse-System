"""工厂对账单导入 (用户拍板 2026-06-15)。

工厂每月发来的「对账单/下单发货明细表」格式很随便:
  - 一个工作簿可能有多个 sheet (如「26年1月」「26年 对账单」), 任意一个都可能单发来;
  - 每个 sheet 顶部有标题行 (「畔色 2026年 下单发货明细表」), 中间可能再插「2026年4月结账单」
    这种分段标题 + 重复表头;
  - 行里混着: 正常订单行 / 备货行(无订单号) / 售后单 / 小计行(「3月结账1-14止」「优惠后」)/
    材料费、整年结清等大额调整行;
  - 价格列可能是数字、0(用现货发)、负数(退补/材料费)、或文字(「已结算」「没生产」「#NAME?」)。

本服务: 按表头定位列 → 逐行识别 → 凡有「订单号」(含追加订单号) 的行, 按订单号匹配系统里
未作废的工厂单, 把「价格」写进 factory_bill_amount(工厂实际)。匹配不上的(备货/售后/查无订单/
价格非数字)只报告、不强建 (用户: 没对上的不搞)。幂等: 重导同值无副作用。

公开:
    parse_workbook(file_bytes) -> (lines, skipped, subtotals)         # 全部 sheet
    parse_sheet_rows(rows) -> (lines, skipped, subtotals)             # 单 sheet 原始行
    import_bill(db, file_bytes, *, dry_run) -> dict                    # 入口①: 整份文件
    commit_sheet_bill(db, file_bytes, sheet_name, *, dry_run) -> dict  # 入口②: smart importer 单 sheet
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.order import FactoryOrder

_log = logging.getLogger("panse.factory_bill")

_MIN_ORDER_NO_LEN = 15   # 淘宝订单号是长数字串


@dataclass
class BillLine:
    order_no: str
    extra_nos: list[str] = field(default_factory=list)   # 追加订单号
    product: Optional[str] = None
    qty: Optional[int] = None
    price: Optional[Decimal] = None       # 工厂价格(实收); None=价格非数字
    price_raw: Any = None
    sheet: Optional[str] = None


def _is_order_no(v: Any) -> bool:
    return isinstance(v, str) and v.strip().isdigit() and len(v.strip()) >= _MIN_ORDER_NO_LEN


def _norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_decimal(v: Any) -> Optional[Decimal]:
    """价格转 Decimal; 文字('已结算''没生产''#NAME?')或空 → None。"""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except (InvalidOperation, ValueError):
            return None
    s = str(v).strip().replace(",", "").replace("¥", "").replace("元", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _build_colmap(header: list[Any]) -> dict[str, int]:
    """表头列名 → 逻辑字段 (子串匹配, 兼容「详 情」「下单时间」等)。"""
    cmap: dict[str, int] = {}
    for i, c in enumerate(header):
        cc = (_norm(c) or "").replace(" ", "")
        # 部分工厂表把追加单直接写成「订单号2/订单2」；它是同一订单补差，
        # 账单价格只属于订单1，不能再给订单2生成一笔工厂待付。
        if any(token in cc for token in ("订单号2", "订单2")):
            cmap["extra1" if "extra1" not in cmap else "extra2"] = i
        elif "订单号" in cc and "追加" not in cc and "order_no" not in cmap:
            cmap["order_no"] = i
        elif "追加" in cc:
            cmap["extra1" if "extra1" not in cmap else "extra2"] = i
        elif ("详情" in cc or "货物" in cc or "详" in cc) and "product" not in cmap:
            cmap["product"] = i
        elif "数量" in cc and "qty" not in cmap:
            cmap["qty"] = i
        elif "价格" in cc and "price" not in cmap:
            cmap["price"] = i
    return cmap


def _looks_like_header(row: list[Any]) -> bool:
    cells = {(_norm(c) or "").replace(" ", "") for c in row}
    return any("订单号" in c for c in cells) and any("价格" in c for c in cells)


def parse_sheet_rows(rows: list[list[Any]]) -> tuple[list[BillLine], list[dict], list[dict]]:
    """解析单个 sheet 的原始行。返回 (订单行, 跳过行明细, 小计/结算行)。"""
    lines: list[BillLine] = []
    skipped: list[dict] = []     # {order_no?/product, price, reason}
    subtotals: list[dict] = []   # {label, amount}
    colmap: dict[str, int] = {}
    for row in rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        if _looks_like_header(row):
            colmap = _build_colmap(row)   # 表头(可能多次出现, 每次刷新列位)
            continue
        if not colmap:
            continue   # 表头之前的标题行
        # 工厂有时把下期延期订单附在「账单尾款」之后作为备忘；这些行不属于本期账单。
        # 以明确尾款标记为边界，避免把下期订单误写进本期工厂实际。
        row_text = " ".join(str(cell).strip() for cell in row if cell is not None)
        if any(marker in row_text for marker in ("账单尾款", "本期尾款")):
            break
        oi = colmap.get("order_no")
        pi = colmap.get("price")
        prod_i = colmap.get("product")
        order_no = _norm(row[oi]) if (oi is not None and oi < len(row)) else None
        price_raw = row[pi] if (pi is not None and pi < len(row)) else None
        product = _norm(row[prod_i]) if (prod_i is not None and prod_i < len(row)) else None
        tag = product or ""
        if not _is_order_no(order_no):
            # 小计/结算标记行 vs 备货/调整行
            if any(k in tag for k in ("结账", "结算", "优惠", "截止", "账单", "止")):
                subtotals.append({"label": tag[:40], "amount": str(_to_decimal(price_raw) or "")})
            elif tag or price_raw is not None:
                skipped.append({"order_no": order_no, "product": tag[:40],
                                "price": str(price_raw)[:20] if price_raw is not None else None,
                                "reason": "无订单号(备货/售后/调整)"})
            continue
        extras = []
        for k in ("extra1", "extra2"):
            ci = colmap.get(k)
            if ci is not None and ci < len(row) and _is_order_no(_norm(row[ci])):
                extras.append(_norm(row[ci]).strip())
        qi = colmap.get("qty")
        qty = None
        if qi is not None and qi < len(row) and row[qi] is not None:
            try:
                qty = int(float(str(row[qi])))
            except (ValueError, TypeError):
                qty = None
        lines.append(BillLine(order_no=order_no.strip(), extra_nos=extras,
                              product=product, qty=qty,
                              price=_to_decimal(price_raw), price_raw=price_raw))
    return lines, skipped, subtotals


def parse_workbook(file_bytes: bytes, sheet_name: Optional[str] = None):
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    all_lines: list[BillLine] = []
    all_skipped: list[dict] = []
    all_sub: list[dict] = []
    for ws in wb.worksheets:
        if sheet_name is not None and ws.title != sheet_name:
            continue
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        lines, skipped, sub = parse_sheet_rows(rows)
        for ln in lines:
            ln.sheet = ws.title
        all_lines.extend(lines)
        all_skipped.extend(skipped)
        all_sub.extend(sub)
    wb.close()
    return all_lines, all_skipped, all_sub


def _apply(db: Session, lines: list[BillLine]) -> dict:
    """按订单号(含追加号)匹配未作废工厂单, 汇总写 factory_bill_amount。

    同一平台订单可能在账单里拆成多行(例如两个 SKU/两件产品)，必须先求和后写入；
    逐行覆盖会只留下最后一行，造成工厂实际少计。
    """
    fo_by_no: dict[str, FactoryOrder] = {}
    for fo in db.execute(select(FactoryOrder).where(FactoryOrder.voided_at.is_(None))).scalars().all():
        if fo.platform_order_no:
            fo_by_no.setdefault(fo.platform_order_no, fo)
    updated = unchanged = non_numeric = topup_linked = 0
    unmatched: list[dict] = []
    matched_totals: dict[int, Decimal] = {}
    matched_orders: dict[int, FactoryOrder] = {}
    for ln in lines:
        if ln.price is None:
            non_numeric += 1
            unmatched.append({"order_no": ln.order_no, "product": (ln.product or "")[:30],
                              "reason": f"价格非数字({str(ln.price_raw)[:16]})"})
            continue
        primary_fo = fo_by_no.get(ln.order_no)
        fo = primary_fo
        if fo is None:
            for ex in ln.extra_nos:
                fo = fo_by_no.get(ex)
                if fo is not None:
                    break
        if fo is None:
            unmatched.append({"order_no": ln.order_no, "product": (ln.product or "")[:30],
                              "reason": "系统无对应工厂单"})
            continue
        # 当订单1在系统中存在时，工厂表的订单2/追加订单号只作为关联凭证：
        # 该平台单仍保留，但单独工厂费用为 0，不再进入待付、缺账单或异常统计。
        if primary_fo is not None:
            for extra_no in ln.extra_nos:
                extra_fo = fo_by_no.get(extra_no)
                if extra_fo is None or extra_fo.id == primary_fo.id:
                    continue
                if (
                    (extra_fo.factory_cost_type or "normal") != "same_order_topup"
                    or extra_fo.related_primary_order_no != ln.order_no
                    or Decimal(str(extra_fo.factory_bill_amount or 0)) != Decimal("0")
                ):
                    topup_linked += 1
                extra_fo.factory_cost_type = "same_order_topup"
                extra_fo.related_primary_order_no = ln.order_no
                extra_fo.factory_bill_amount = Decimal("0")
        matched_orders[fo.id] = fo
        matched_totals[fo.id] = matched_totals.get(fo.id, Decimal("0")) + ln.price

    for fo_id, total in matched_totals.items():
        fo = matched_orders[fo_id]
        if fo.factory_bill_amount is not None and Decimal(str(fo.factory_bill_amount)) == total:
            unchanged += 1
            continue
        fo.factory_bill_amount = total
        updated += 1
    return {"updated": updated, "unchanged": unchanged, "topup_linked": topup_linked,
            "non_numeric": non_numeric, "unmatched": unmatched}


def import_bill(db: Session, file_bytes: bytes, *, sheet_name: Optional[str] = None,
                dry_run: bool = False) -> dict:
    """入口①: 解析整份(或指定 sheet)工厂对账单 → 写工厂实际。"""
    lines, skipped, subtotals = parse_workbook(file_bytes, sheet_name=sheet_name)
    res = _apply(db, lines)
    if dry_run:
        db.rollback()
    else:
        db.commit()
    res.update({
        "order_lines": len(lines),
        "stock_or_aftersales_skipped": len(skipped),
        "subtotals": subtotals,
        "unmatched_count": len(res["unmatched"]),
        "unmatched": res["unmatched"][:50],
        "dry_run": dry_run,
    })
    _log.info("import_bill: 行=%d 更新=%d 不变=%d 关联补差=%d 非数字=%d 未匹配=%d dry=%s",
              len(lines), res["updated"], res["unchanged"], res["topup_linked"], res["non_numeric"],
              res["unmatched_count"], dry_run)
    return res


def commit_sheet_bill(db: Session, *, file_bytes: bytes, sheet_name: str,
                      dry_run: bool = False) -> dict:
    """入口②: smart importer 对单个 sheet 调用 (entity_type=factory_bill)。"""
    res = import_bill(db, file_bytes, sheet_name=sheet_name, dry_run=dry_run)
    warn = []
    if res["unmatched_count"]:
        warn.append(f"{res['unmatched_count']} 行未匹配系统工厂单(备货/售后/查无订单/价格非数字)")
    return {
        "sheet_name": sheet_name, "entity_type": "factory_bill",
        "total_rows": res["order_lines"],
        "inserted_parents": res["updated"], "inserted_children": 0,
        "skipped_rows": res["unchanged"] + res["unmatched_count"],
        "errors": [], "conflicts": [], "warnings": warn,
        "unmapped_columns": [],
    }

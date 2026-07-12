"""导入「已生效活动价」(校验底价) — 千牛活动商品导出(已报商品列表) → PricingSkuPromo.enrolled_floor_price。

淘宝规则: 活动券后价不得高于校验期内最低普惠券后价 → 上一场已生效活动价就是硬底。
用途 (2026-07-12 第二场超级88导入62件全失败根因):
  1. 占位/定制 SKU 报名价自动封顶到此价 (data_export builder);
  2. 预检"券后价超线"红字: 任何 SKU 计划报名价 > 此价 → 会被淘宝拦, 提前暴露。
表结构 = 千牛「活动商品导出」: sheet[已报商品列表], 前3行表头(第2行列名:
  商品ID/商品名称/营销ID/商品状态/SKUID/SKU名称/活动价/...), 数据第4行起。
匹配: SKUID ↔ promo.taobao_sku_id(或 alt); 重复导入取 min(已存, 新值) —— 底价只会更低不会抬高。
"""
from __future__ import annotations

import io
from decimal import Decimal

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.pricing_ext import PricingSkuPromo


def import_from_xlsx_bytes(db: Session, raw: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["已报商品列表"] if "已报商品列表" in wb.sheetnames else wb.worksheets[-1]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 4:
        return {"ok": False, "error": "表里没有数据行(需千牛「活动商品导出」原表)"}
    header = [str(c) if c is not None else "" for c in rows[1]]
    try:
        i_sid = header.index("SKUID")
        i_price = header.index("活动价")
    except ValueError:
        return {"ok": False, "error": f"没找到 SKUID/活动价 列, 实际表头: {[h for h in header if h][:8]}"}

    floor_by_sid: dict[str, Decimal] = {}
    for row in rows[3:]:
        if not row or row[i_sid] is None:
            continue
        sid = str(row[i_sid]).strip()
        try:
            price = Decimal(str(row[i_price]))
        except Exception:  # noqa: BLE001
            continue
        if not sid or price <= 0:
            continue
        # 同 SKUID 多行取最低(底价从严)
        if sid not in floor_by_sid or price < floor_by_sid[sid]:
            floor_by_sid[sid] = price

    promos = db.execute(select(PricingSkuPromo)).scalars().all()
    updated = matched = 0
    unmatched = set(floor_by_sid)
    for p in promos:
        ids = [str(p.taobao_sku_id).strip()] if p.taobao_sku_id else []
        ids += [str(a).strip() for a in (p.alt_taobao_sku_ids or []) if a]
        hit = next((i for i in ids if i in floor_by_sid), None)
        if hit is None:
            continue
        matched += 1
        unmatched.discard(hit)
        new = floor_by_sid[hit]
        old = p.enrolled_floor_price
        if old is None or new < old:      # 底价只降不抬
            p.enrolled_floor_price = new
            updated += 1
    db.commit()
    return {"ok": True, "file_rows": len(floor_by_sid), "matched_sku": matched,
            "updated": updated, "unmatched_count": len(unmatched),
            "unmatched_sample": sorted(unmatched)[:10]}

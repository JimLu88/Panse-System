# -*- coding: utf-8 -*-
"""异常中心导出 (用户需求 2026-06-11 重申): 异常要随「对应的源表」一起导出。

不是导一张异常清单, 而是: 订单有异常 → 导出订单表, 在出问题的那一行
最后一列写「异常批注」。每个 source_table 一个 sheet:
    - 能定位到源表模型 → 整行数据 + 末列批注 (同一行多条异常编号拼接)
    - 定位不到的异常 (如 reconciliation 这种虚拟源) → 该 sheet 直接列异常本身
只导 open 异常 (可选含 ignored), resolved 的不再打扰。
"""
from __future__ import annotations

from typing import Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.base import Base
from app.models.exception import DataException

# 源表业务键候选 (按序尝试; 都没有就拿主键 id 按数字匹配)
_KEY_CANDIDATES = ("order_no", "sku_code", "code", "purchase_no", "transaction_no",
                   "wsf_order_no", "factory_order_no", "sample_no", "platform_order_no")

_SEVERITY_CN = {"info": "提示", "warning": "警告", "error": "严重"}
_NUMS = "①②③④⑤⑥⑦⑧⑨⑩"

# 对账规则前缀 → 中文 (业务键如 factory_payment:玉山县… 翻成 货款对账 · 玉山县…)
_RULE_CN = {
    "factory_payment": "货款对账", "install_fee": "安装费收支", "promotion": "推广支出",
    "refill_compensation": "补单赔实付", "inventory_value": "库存资产", "logistics_fee": "物流费销项",
    "revenue_alipay": "收入对账", "operating_expense": "经营支出", "purchase_payment": "采购付款",
    "refill_commission_payout": "补单佣金代付", "refill_express_payout": "补单快递代付",
    "aftersales_payout": "售后赔付代付", "refund_reconciliation": "退款进出对账",
}
# 源表英文名 → 中文 sheet 名 (虚拟源/兜底)
_TABLE_CN = {
    "reconciliation": "对账差异", "orders": "销售订单", "products": "产品",
    "factory_orders": "工厂订单", "factory_reconciliations": "工厂对账",
    "alipay_flows": "支付宝流水", "promotion_flows": "推广流水", "after_sales": "售后",
    "refill_records": "补单记录", "pricing_skus": "定价SKU", "materials": "物料",
    "part_inventory": "配件库存", "product_inventory": "成品库存",
}


def _humanize_key(pk) -> str:
    """业务键 'factory_payment:玉山县…' → '货款对账 · 玉山县…' (前缀转中文, 让人看懂)。"""
    if pk is None:
        return ""
    s = str(pk)
    pre, sep, rest = s.partition(":")
    cn = _RULE_CN.get(pre)
    if sep and cn:
        return f"{cn} · {rest}" if rest else cn
    return s


def _sheet_label(table_name: str) -> str:
    """源表 → 中文 sheet 名: 先查 ENTITY_MODELS 标签, 再兜底 _TABLE_CN, 最后原名。"""
    try:
        from app.api.table_explorer import ENTITY_MODELS
        for cfg in ENTITY_MODELS.values():
            if getattr(cfg.get("model"), "__tablename__", None) == table_name:
                return (cfg.get("label") or table_name)[:28]
    except Exception:
        pass
    return _TABLE_CN.get(table_name, table_name)[:28]


def _model_for_table(table_name: str):
    for mapper in Base.registry.mappers:
        if getattr(mapper.class_, "__tablename__", None) == table_name:
            return mapper.class_
    return None


def _key_column(model):
    cols = {c.key for c in model.__table__.columns}
    for k in _KEY_CANDIDATES:
        if k in cols:
            return k
    return "id"


def _cell(v):
    """openpyxl 能写的标量; Decimal→float、date/datetime 原生透传(让 Excel 当数字/日期),
    其余 dict/list 等转 str。修: 旧版把 Decimal/日期一并 str() 导致全量导出数字变文本。
    注意: datetime/time 必须剥掉 tzinfo —— openpyxl 不支持带时区的时间(TimestampMixin 是 tz-aware)。"""
    from datetime import date as _date, datetime as _dt, time as _time
    from decimal import Decimal as _Dec
    if v is None or isinstance(v, (int, float, str, bool)):
        return v
    if isinstance(v, _Dec):
        return float(v)
    if isinstance(v, _dt):            # datetime 是 date 子类, 必须先判
        return v.replace(tzinfo=None)
    if isinstance(v, _time):
        return v.replace(tzinfo=None)
    if isinstance(v, _date):
        return v   # 纯日期无时区, openpyxl 原生写成日期, 由调用方设 number_format
    return str(v)


def _join_notes(ns: list[str]) -> str:
    if len(ns) == 1:
        return ns[0]
    return " ".join(f"{_NUMS[i] if i < len(_NUMS) else i+1}{n}" for i, n in enumerate(ns))


_STATUS_CN = {"open": "未处理", "ignored": "已忽略", "resolved": "已处理"}


def build_export_workbook(db: Session, *, source_table: Optional[str] = None,
                          include_ignored: bool = False):
    """异常随源表导出 (批注表)。全中文 + 美化(深蓝表头/边框/冻结/筛选/隔行底色)。
    复用 data_export_service 的表头翻译/枚举翻译/样式 (延迟导入避免循环)。"""
    import openpyxl
    # 延迟导入: data_export_service 顶层 import 本模块, 这里运行期再反向取, 不构成循环
    from app.services.data_export_service import (
        _COMMON_HEADER_CN, _apply_table_style, _cn_header, _col_type, _num_fmt, _translate,
    )
    try:
        from app.api.table_explorer import ENTITY_MODELS
        _table2ekey = {getattr(c.get("model"), "__tablename__", None): k
                       for k, c in ENTITY_MODELS.items()}
    except Exception:
        _table2ekey = {}

    statuses = ("open", "ignored") if include_ignored else ("open",)
    stmt = select(DataException).where(DataException.status.in_(statuses))
    if source_table:
        stmt = stmt.where(DataException.source_table == source_table)
    excs = db.execute(stmt.order_by(DataException.source_table, DataException.id)).scalars().all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    by_table: dict[str, list[DataException]] = {}
    for e in excs:
        by_table.setdefault(e.source_table, []).append(e)

    if not by_table:
        ws = wb.create_sheet("无未处理异常")
        ws.append(["当前没有未处理 (open) 的异常。"])
        return wb

    def _new_sheet(table: str):
        name = _sheet_label(table) or "异常"
        base, i = name, 2
        while name in used:
            sfx = f"({i})"
            name = base[:28 - len(sfx)] + sfx
            i += 1
        used.add(name)
        return wb.create_sheet(name)

    for table, items in by_table.items():
        model = _model_for_table(table)

        if model is None:
            # 虚拟源 (对账差异等无真实源表): 列异常本身, 业务键/严重度/状态 全中文
            ws = _new_sheet(table)
            headers = ["业务键", "严重度", "异常说明", "状态", "记录时间"]
            ws.append(headers)
            for e in items:
                ws.append([
                    _humanize_key(e.source_pk),
                    _SEVERITY_CN.get(e.severity, e.severity),
                    e.description,
                    _STATUS_CN.get(e.status, e.status),
                    e.created_at.replace(tzinfo=None) if e.created_at else None,
                ])
            _apply_table_style(ws, headers, exc_col_idx=3,
                               col_fmts={5: "yyyy-mm-dd hh:mm"}, data_end_row=ws.max_row)
            continue

        key_col = _key_column(model)
        notes: dict[str, list[str]] = {}
        for e in items:
            if e.source_pk:
                notes.setdefault(str(e.source_pk), []).append(
                    f"[{_SEVERITY_CN.get(e.severity, e.severity)}] {e.description}")
        rows = []
        located: set[str] = set()
        if notes:
            col = getattr(model, key_col)
            keys = list(notes.keys())
            if key_col == "id":
                int_keys = [int(k) for k in keys if k.isdigit()]
                found = (db.execute(select(model).where(col.in_(int_keys))).scalars().all()
                         if int_keys else [])
            else:
                found = db.execute(select(model).where(col.in_(keys))).scalars().all()
            for r in found:
                k = str(getattr(r, key_col))
                located.add(k)
                rows.append((r, _join_notes(notes[k])))

        ws = _new_sheet(table)
        ekey = _table2ekey.get(table)
        model_cols = list(model.__table__.columns)
        col_names = [c.key for c in model_cols]
        # 表头中文化 (有 entity schema 用其 desc, 否则常用映射)
        cn_headers = [(_cn_header(ekey, c) if ekey else _COMMON_HEADER_CN.get(c, c))
                      for c in col_names]
        headers = cn_headers + ["异常批注"]
        ws.append(headers)
        # 列号 → number_format (金额/日期/百分比)
        col_fmts: dict[int, str] = {}
        for i, c in enumerate(model_cols, start=1):
            fmt = _num_fmt(c.key, _col_type(c))
            if fmt:
                col_fmts[i] = fmt
        note_rows: set[int] = set()
        for r, note in rows:
            ws.append([_translate(c, _cell(getattr(r, c))) for c in col_names] + [note])
            note_rows.add(ws.max_row)
        # 定位不到源行的异常 (行已删/键变了) 也不能丢
        orphans = [(k, ns) for k, ns in notes.items() if k not in located]
        data_end = ws.max_row
        exc_idx = len(col_names) + 1
        _apply_table_style(ws, headers, exc_col_idx=exc_idx, note_rows=note_rows,
                           col_fmts=col_fmts, data_end_row=data_end)
        if orphans:
            from openpyxl.styles import Font
            ws.append([])
            tip = ws.cell(row=ws.max_row + 1, column=1,
                          value="── 以下异常在源表里找不到对应行 (可能已删除/键已变) ──")
            tip.font = Font(bold=True, color="C0392B")
            for k, ns in orphans:
                ws.append([_humanize_key(k)] + [None] * (len(col_names) - 1) + ["; ".join(ns)])
    return wb

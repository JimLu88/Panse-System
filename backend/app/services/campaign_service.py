"""活动生命周期 P1 引擎 (2026-07-17 权威 spec: docs/活动生命周期系统_执行plan.md)。

职责:
- group_by_sales      动销检查与分组 (spec §四.1) + no_sales 登记表同步
- build_signup_rows   报名行 builder: 报名价=日常价 / 占位=min(现行, floor(线/(1−lev))) (spec §二.1, R3/R4)
- build_discount_rows 单品立减 builder: spec §二 立减公式逐字 (官方立减向上取整到元 R9 /
                      ERP目标+已授权微调 / 无动销=日常−中促 / 10% ceil 留开关)
- preflight           平台规则库 R0~R19 静态可查项逐条输出 (spec §三)
- push_discount/push_signup  推送编排 (复用 web_agent_service upload_file→wait_job,
                      与 activity_upload_service 同模式)
- target_prices       核对器用的逐 skuId 目标到手 (campaign_recon_service 消费)

铁则 (spec §二, 用户 2026-07-17 拍板):
  报名价 = ERP 日常价, 永不再变; 中促 = 大促 × 1.03 (就地计算, 不写 mid_buyer 字段——那是任务#22);
  无动销到手 = ERP 中促价（精确相等，不再额外加 1 元）; ERP 价是唯一标准。
  平台最低普惠券后价会计入已生效的店铺其他优惠，因此资格门校验
  「报名价−官方立减−同期单品立减 ≤ 近15天最低普惠券后价」；
  最终到手仍必须等于 ERP 目标，只有用户逐 SKU 授权的 1 元内微调可继续向下贴线。
只读 PricingSku / PricingSkuPromo, 绝不改其字段。
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal, ROUND_CEILING, ROUND_FLOOR, ROUND_HALF_UP
import hashlib
import json
import re
import secrets
from typing import Optional

from sqlalchemy import func, select
from sqlalchemy.orm import Session

# 活动类型 → (人话名, 档位)。档位: mid=超级立减10%→中促到手 / big=12%→大促到手 / big618=15%→大促到手
CAMPAIGN_TYPES = {
    "super_reduce": ("超级立减", "mid"),
    "big88":        ("88VIP大促", "big"),
    "big38":        ("38大促", "big"),
    "big_other":    ("其他大促", "big"),
    "big618":       ("618大促", "big618"),
    "big11":        ("双11大促", "big618"),
}
TIER_LEVERAGE = {"mid": Decimal("0.10"), "big": Decimal("0.12"), "big618": Decimal("0.15")}
MID_OVER_BIG_RATIO = Decimal("1.03")       # 任务#22: 中促 = 大促 × 1.03 (系统统一系数, 就地算)
LINE_CONCESSION_MAX_YUAN = Decimal("1")    # 贴线让幅 > 1 元 → 暂缓该商品并提醒人工决策 (R2)
PLACEHOLDER_LINE_FALLBACK_RATIO = Decimal("0.8")   # 占位无券后线 → 日常×0.8 保守线(行备注标注)
OFFICIAL_CEIL_KEY = "campaign_official_ceil"       # 10% 官方立减是否向上取整(待7-20实证), 默认真
PLACEHOLDER_LIVE_PRICE_KEY_PREFIX = "campaign_placeholder_live_prices_v1_plan_"
_CENT = Decimal("0.01")
# spec §四.1 说剔 closed; 本库订单状态机把交易关闭存成 cancelled → 两个都剔 (口径决定, 见交付说明)
_CLOSED_STATUSES = ("closed", "cancelled")


def _d(x) -> Optional[Decimal]:
    if x is None:
        return None
    try:
        return Decimal(str(x))
    except Exception:  # noqa: BLE001 — 脏数据当缺失, 不炸主流程
        return None


def plan_tier(plan) -> str:
    """计划档位: 优先已固化的 plan.tier, 缺省由活动类型派生。未知类型显式报错。"""
    tier = getattr(plan, "tier", None)
    if tier in TIER_LEVERAGE:
        return tier
    ctype = getattr(plan, "campaign_type", None)
    if ctype not in CAMPAIGN_TYPES:
        raise ValueError(f"未知活动类型 {ctype!r}; 可选 {list(CAMPAIGN_TYPES)}")
    return CAMPAIGN_TYPES[ctype][1]


def official_scope_for_plan(plan) -> dict:
    """Read the per-item official-discount scope captured from the live activity page.

    Supported remark markers:
    - official_active_items=1,2,3
    - official_all_store=true; official_exempt_items=4,5

    The markers are deliberately explicit.  Guessing that an item has no official
    discount can make the generated single-item discount too large.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")

    def _ids(key: str) -> tuple[bool, set[str]]:
        matched = re.search(
            rf"(?:^|[;\n；])\s*{re.escape(key)}\s*=\s*([^;\n；]*)",
            text,
            flags=re.IGNORECASE,
        )
        if not matched:
            return False, set()
        return True, set(re.findall(r"\d+", matched.group(1)))

    all_store_match = re.search(
        r"(?:^|[;\n；])\s*official_all_store\s*=\s*"
        r"(true|false|1|0|yes|no|on|off)\b",
        text,
        flags=re.IGNORECASE,
    )
    all_store_present = bool(all_store_match)
    all_store = bool(
        all_store_match
        and all_store_match.group(1).lower() in ("true", "1", "yes", "on")
    )
    active_present, active_items = _ids("official_active_items")
    exempt_present, exempt_items = _ids("official_exempt_items")
    errors: list[str] = []
    if all_store and active_present:
        errors.append("official_all_store=true 与 official_active_items 不能同时配置")
    if exempt_present and not all_store:
        errors.append("official_exempt_items 仅可与 official_all_store=true 同时使用")
    if all_store and not exempt_present:
        errors.append("全店官方立减必须显式配置 official_exempt_items（允许空名单）")
    if all_store_present and not all_store and not active_present:
        errors.append("official_all_store=false 时必须显式配置 official_active_items")
    if active_items & exempt_items:
        errors.append("同一商品不能同时处于官方立减生效与豁免名单")
    configured = (all_store and exempt_present) or active_present
    return {
        "configured": configured and not errors,
        "all_store": all_store,
        "active_items": active_items,
        "exempt_items": exempt_items,
        "errors": errors,
    }


def _official_applies(item_id: str, scope: dict) -> bool:
    if not scope.get("configured"):
        return False
    if scope.get("all_store"):
        return item_id not in scope.get("exempt_items", set())
    return item_id in scope.get("active_items", set())


def _super_reduce_plan_exempt_item(plan, item_id: str, scope: dict) -> bool:
    """A long-running Super Reduce exemption means no rows in this update plan.

    Other campaign types retain the existing no-sales single-item-discount
    fallback semantics for official exemptions.
    """
    return bool(
        str(getattr(plan, "campaign_type", "")) == "super_reduce"
        and str(getattr(plan, "platform_activity_mode", ""))
        == "long_running_update"
        and scope.get("all_store")
        and item_id in scope.get("exempt_items", set())
    )


def current_activity_prices_for_plan(plan) -> dict[str, Decimal]:
    """Read SKU-level activity prices captured from this plan's live export."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*current_activity_prices\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    out: dict[str, Decimal] = {}
    for sku_id, raw_price in re.findall(
            r"(\d+)\s*:\s*(\d+(?:\.\d+)?)", matched.group(1)):
        price = _d(raw_price)
        if price is not None and price > 0:
            out[sku_id] = price.quantize(_CENT)
    return out


def _set_plan_decimal_map_marker(plan, key: str,
                                 values: dict[str, Decimal]) -> None:
    """Idempotently replace a compact ``id:decimal`` plan marker."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    pattern = rf"(?:^|[;\n；])\s*{re.escape(key)}\s*=\s*[^;\n；]*"
    text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip(" ;\n；")
    value = ",".join(
        f"{sku_id}:{price.quantize(_CENT)}"
        for sku_id, price in sorted(values.items())
    )
    plan.remark = f"{text}; {key}={value}" if text else f"{key}={value}"


def sync_live_activity_evidence(db: Session, plan,
                                active_records: list[dict]) -> dict:
    """Refresh known-item official scope and prices from a live full export.

    Items already known from qualification/no-sales/failure evidence may be
    promoted when the platform now proves them active.  Entirely unknown active
    items remain outside scope so reconciliation still alarms on them.
    """
    from app.services import no_sales_service

    prior_qualified = platform_qualified_items(plan)
    prior_no_sales = platform_no_sales_items(plan)
    prior_hard_failed = platform_hard_failed_items(plan)
    registered_no_sales = no_sales_service.get_no_sales(db)
    known_items = (
        prior_qualified | prior_no_sales | prior_hard_failed
        | registered_no_sales
    )
    active_items = {
        str(row.get("item_id") or "").strip()
        for row in active_records
        if str(row.get("item_id") or "").strip()
    }
    if not known_items:
        return {
            "active_items": sorted(active_items),
            "accepted_items": [],
            "promoted_items": [],
            "unknown_active_items": sorted(active_items),
            "live_price_skus": [],
        }
    accepted_items = active_items & known_items
    promoted_items = accepted_items - prior_qualified
    live_prices: dict[str, Decimal] = {}
    for row in active_records:
        item_id = str(row.get("item_id") or "").strip()
        sku_id = str(row.get("sku_id") or "").strip()
        price = _d(row.get("activity_price"))
        if item_id in accepted_items and sku_id and price is not None and price > 0:
            live_prices[sku_id] = price.quantize(_CENT)

    _sync_official_scope_after_platform_result(plan, accepted_items)
    _set_plan_item_marker(
        plan, "platform_qualified_items", prior_qualified | accepted_items)
    _set_plan_item_marker(
        plan, "platform_no_sales_items", prior_no_sales - accepted_items)
    _set_plan_item_marker(
        plan, "platform_hard_failed_items", prior_hard_failed - accepted_items)
    _set_plan_decimal_map_marker(plan, "current_activity_prices", live_prices)
    no_sales_service.remove_no_sales(db, accepted_items)
    return {
        "active_items": sorted(active_items),
        "accepted_items": sorted(accepted_items),
        "promoted_items": sorted(promoted_items),
        "unknown_active_items": sorted(active_items - known_items),
        "live_price_skus": sorted(live_prices),
    }


def placeholder_live_prices_for_plan(db: Session, plan) -> dict[str, Decimal]:
    """Read exact, plan-scoped platform prices for placeholder SKU IDs.

    New read-only exports persist these observations outside the immutable plan
    remark. The remark parser remains a compatibility bridge for older plans.
    """
    import re

    from app.services import settings_service

    plan_id = str(getattr(plan, "id", "") or "").strip()
    if plan_id.isdigit():
        raw = settings_service.get(
            db, f"{PLACEHOLDER_LIVE_PRICE_KEY_PREFIX}{plan_id}",
            env_fallback=False)
        if raw is not None:
            try:
                payload = json.loads(raw)
            except (TypeError, ValueError):
                payload = {}
            out: dict[str, Decimal] = {}
            if isinstance(payload, dict):
                for sid, raw_price in payload.items():
                    value = _d(raw_price)
                    if str(sid).isdigit() and value is not None and value > 0:
                        out[str(sid)] = value
            return out

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*placeholder_live_prices\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    out: dict[str, Decimal] = {}
    for sid, price in re.findall(r"(\d+)\s*:\s*(\d+(?:\.\d+)?)", matched.group(1)):
        value = _d(price)
        if value is not None and value > 0:
            out[sid] = value.quantize(_CENT)
    return out


def record_placeholder_live_prices(
        db: Session, plan, floor_rows: list[dict],
        candidate_sku_ids: list[str]) -> dict:
    """Replace placeholder prices from one exact current platform export.

    Only candidate placeholder SKU IDs are accepted. Missing candidates are
    deliberately omitted so R16 hard-stops them; stale observations are never
    carried forward and real SKU prices never enter this setting.
    """
    from app.services import settings_service

    plan_id = str(getattr(plan, "id", "") or "").strip()
    candidates = {str(sid) for sid in candidate_sku_ids if str(sid).isdigit()}
    if not plan_id.isdigit():
        return {"observed": 0, "candidate_count": len(candidates),
                "missing_sku_ids": sorted(candidates), "scope": None}
    observed: dict[str, float] = {}
    for row in floor_rows:
        sid = str(row.get("sku_id") or "").strip()
        if sid not in candidates:
            continue
        value = _d(row.get("list_price"))
        if value is not None and value > 0:
            observed[sid] = float(value)
    key = f"{PLACEHOLDER_LIVE_PRICE_KEY_PREFIX}{plan_id}"
    settings_service.set_value(
        db, key,
        json.dumps(observed, ensure_ascii=False, sort_keys=True,
                   separators=(",", ":")),
        description=f"淘宝活动计划{plan_id}占位SKU平台当前一口价证据",
    )
    db.flush()
    return {
        "observed": len(observed),
        "candidate_count": len(candidates),
        "missing_sku_ids": sorted(candidates - set(observed)),
        "scope": key,
    }


def placeholder_price_protection_expired(plan) -> bool:
    """Return whether placeholder price protection was explicitly confirmed expired.

    ``price_protection_days`` is only a reminder horizon; it does not identify when
    each historical placeholder low price was created.  Lowering a placeholder
    campaign price therefore requires an explicit per-plan confirmation in remark.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*placeholder_price_protection_expired\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return False
    return matched.group(1).strip().lower() in (
        "1", "true", "yes", "on", "已到期", "是",
    )


def placeholder_price_lowering_authorized(plan) -> bool:
    """Current-plan user decision allowing custom placeholders to use safe caps.

    This does not lower any real SKU or authorize extra real-SKU concessions.
    It only removes the stale-price hold for placeholder
    consultation entries whose commercial price is explicitly non-binding.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*placeholder_price_lowering_authorized\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return bool(matched and matched.group(1).strip().lower() in (
        "1", "true", "yes", "on", "是", "已授权",
    ))


def new_item_no_history_authorized_items(plan) -> set[str]:
    """Item IDs explicitly confirmed as new links with no platform history yet."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*new_item_no_history_authorized\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return set(re.findall(r"\d{8,}", matched.group(1))) if matched else set()


def authorized_supplement_items(plan) -> set[str]:
    """Return the exact item scope approved for one corrective program run.

    Marker format::

        supplement_items_authorized=1007407909979

    The marker narrows the generated upload; it never expands the normal
    official-discount scope or bypasses price, completeness, and live-state
    checks.  A named item missing from the safe generated rows hard-stops.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*supplement_items_authorized\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return set(re.findall(r"\d{8,}", matched.group(1))) if matched else set()


def authorized_sku_refresh_items(plan) -> set[str]:
    """SKU rotation is disabled for campaign enrollment."""
    return set()


def requested_sku_refresh_items(plan) -> set[str]:
    """Return legacy/requested rotation markers so preflight can block them."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*sku_refresh_items_authorized\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return set(re.findall(r"\d{8,}", matched.group(1))) if matched else set()


def authorized_signup_shipping_days(plan) -> dict[str, int]:
    """Exact item-level overrides for the policy default shipping days.

    Marker format::

        signup_shipping_days_authorized=793084818113:30

    Every row receives the policy default (currently 30 days). This marker may
    override only exact item IDs without changing the other rows.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*signup_shipping_days_authorized\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    shipping_days: dict[str, int] = {}
    for item_id, raw_days in re.findall(
            r"(\d{8,})\s*:\s*(\d{1,3})", matched.group(1)):
        days = int(raw_days)
        if 1 <= days <= 365:
            shipping_days[item_id] = days
    return shipping_days


def signup_shipping_days(plan, rows: list[dict]) -> dict[str, int]:
    """Resolve a non-empty relative shipping day value for every target item."""
    from app.services import campaign_policy_service

    default_days = campaign_policy_service.default_shipping_days()
    overrides = authorized_signup_shipping_days(plan)
    item_ids = {
        str(row.get("taobao_item_id") or "").strip() for row in rows
        if str(row.get("taobao_item_id") or "").strip()
    }
    return {item_id: overrides.get(item_id, default_days) for item_id in item_ids}


def platform_qualified_items(plan) -> set[str]:
    """Items accepted by this campaign's latest pre-publish platform probe."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*platform_qualified_items\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return set(re.findall(r"\d{8,}", matched.group(1))) if matched else set()


def platform_terminal_accepted_items(plan) -> set[str]:
    """Items the platform accepted in the latest terminal qualification run.

    This is deliberately narrower than ``platform_qualified_items``: an item
    provisionally allowed because the planned single-item discount can clear a
    coupon floor is *not* a terminal platform acceptance yet.
    """
    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*platform_terminal_accepted_items\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return set(re.findall(r"\d{8,}", matched.group(1))) if matched else set()


def _plan_marker_value(plan, key: str) -> str:
    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        rf"(?:^|[;\n；])\s*{re.escape(key)}\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return str(matched.group(1) or "").strip() if matched else ""


def _set_plan_value_marker(plan, key: str, value: str) -> None:
    """Idempotently replace one scalar semicolon-delimited plan marker."""
    text = str(getattr(plan, "remark", None) or "")
    pattern = rf"(?:^|[;\n；])\s*{re.escape(key)}\s*=\s*[^;\n；]*"
    text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip(" ;\n；")
    plan.remark = f"{text}; {key}={value}" if text else f"{key}={value}"


def _signup_rows_digest(rows: list[dict], item_ids: set[str]) -> str:
    """Fingerprint exact item/SKU/signup-price tuples accepted by platform."""
    payload = sorted(
        (
            str(row.get("taobao_item_id") or "").strip(),
            str(row.get("taobao_sku_id") or "").strip(),
            str(_d(row.get("price")) or ""),
            bool(row.get("is_placeholder")),
        )
        for row in rows
        if str(row.get("taobao_item_id") or "").strip() in item_ids
    )
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _record_terminal_platform_acceptance(
        plan, rows: list[dict], item_ids: set[str]) -> None:
    """Persist fresh exact-price platform acceptance evidence for R17."""
    _record_terminal_item_evidence(
        plan, rows, "platform_terminal_accepted", item_ids)


def _record_terminal_item_evidence(
        plan, rows: list[dict], prefix: str, item_ids: set[str]) -> None:
    """Persist one fresh, exact-row terminal platform evidence category."""
    accepted = set(item_ids)
    _set_plan_item_marker(plan, f"{prefix}_items", accepted)
    _set_plan_value_marker(
        plan, f"{prefix}_digest",
        _signup_rows_digest(rows, accepted),
    )
    _set_plan_value_marker(
        plan, f"{prefix}_observed_at",
        datetime.now(timezone.utc).isoformat(),
    )


def _record_terminal_coupon_floor_qualification(
        plan, rows: list[dict], item_ids: set[str]) -> None:
    """Record items rejected solely on coupon floor before planned discount."""
    _record_terminal_item_evidence(
        plan, rows, "platform_terminal_coupon_floor", item_ids)


def _fresh_terminal_platform_acceptance(
        plan, signup_rows: list[dict], max_age_hours: float) -> dict:
    """Return accepted items only when time and row fingerprint still match."""
    return _fresh_terminal_item_evidence(
        plan, signup_rows, max_age_hours, "platform_terminal_accepted")


def _fresh_terminal_coupon_floor_qualification(
        plan, signup_rows: list[dict], max_age_hours: float) -> dict:
    """Return fresh exact coupon-floor-only terminal qualification items."""
    return _fresh_terminal_item_evidence(
        plan, signup_rows, max_age_hours, "platform_terminal_coupon_floor")


def _fresh_terminal_item_evidence(
        plan, signup_rows: list[dict], max_age_hours: float, prefix: str) -> dict:
    item_ids = set(re.findall(
        r"\d{8,}", _plan_marker_value(plan, f"{prefix}_items")))
    observed_raw = _plan_marker_value(plan, f"{prefix}_observed_at")
    expected_digest = _plan_marker_value(plan, f"{prefix}_digest")
    age_hours = None
    if observed_raw:
        try:
            observed = datetime.fromisoformat(observed_raw.replace("Z", "+00:00"))
            if observed.tzinfo is None:
                observed = observed.replace(tzinfo=timezone.utc)
            age_hours = max(0.0, (
                datetime.now(timezone.utc) - observed.astimezone(timezone.utc)
            ).total_seconds() / 3600)
        except ValueError:
            age_hours = None
    actual_digest = _signup_rows_digest(signup_rows, item_ids) if item_ids else ""
    valid = bool(
        item_ids and expected_digest and actual_digest == expected_digest
        and age_hours is not None and age_hours <= max_age_hours
    )
    return {
        "item_ids": item_ids if valid else set(),
        "observed_at": observed_raw or None,
        "age_hours": age_hours,
        "digest_matches": bool(expected_digest and actual_digest == expected_digest),
    }


def platform_no_sales_items(plan) -> set[str]:
    """Items rejected only for no-sales by the latest platform probe."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*platform_no_sales_items\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return set(re.findall(r"\d{8,}", matched.group(1))) if matched else set()


def platform_hard_failed_items(plan) -> set[str]:
    """Items isolated for price/SKU/listing or other non-no-sales failures."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*platform_hard_failed_items\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    return set(re.findall(r"\d{8,}", matched.group(1))) if matched else set()


def platform_scope_present(plan) -> bool:
    """Whether this plan already has a terminal platform qualification scope."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    return bool(re.search(
        r"(?:^|[;\n；])\s*platform_(?:qualified|no_sales|hard_failed)_items\s*=",
        text,
        flags=re.IGNORECASE,
    ))


def _set_plan_item_marker(plan, key: str, item_ids: set[str]) -> None:
    """Idempotently replace one semicolon-delimited plan marker."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    pattern = rf"(?:^|[;\n；])\s*{re.escape(key)}\s*=\s*[^;\n；]*"
    text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip(" ;\n；")
    value = ",".join(sorted(item_ids))
    plan.remark = f"{text}; {key}={value}" if text else f"{key}={value}"


def _remove_plan_marker(plan, key: str) -> None:
    """Remove one plan marker completely, including its delimiter."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    pattern = rf"(?:^|[;\n；])\s*{re.escape(key)}\s*=\s*[^;\n；]*"
    plan.remark = re.sub(pattern, "", text, flags=re.IGNORECASE).strip(" ;\n；")


def _sync_official_scope_after_platform_result(
        plan, accepted_items: set[str]) -> None:
    """Record observed items without destroying the plan's requested scope.

    ``official_all_store`` and its explicit exemption list are operator-owned
    plan inputs.  Platform qualification and terminal submission receipts may
    refine runtime evidence, but they must not turn an all-store plan into an
    active-item plan.  Explicit active-item plans keep the historical behavior.
    """
    scope = official_scope_for_plan(plan)
    if scope.get("configured") and scope.get("all_store"):
        _remove_plan_marker(plan, "official_active_items")
        return
    _remove_plan_marker(plan, "official_all_store")
    _remove_plan_marker(plan, "official_exempt_items")
    _set_plan_item_marker(plan, "official_active_items", accepted_items)


def _apply_authorized_supplement_scope(plan, rows: list[dict], stats: dict) -> tuple[
        list[dict], list[str]]:
    """Narrow a retry upload to the explicitly authorized item IDs."""
    scope = authorized_supplement_items(plan)
    if not scope:
        return rows, []
    present = {str(row.get("taobao_item_id") or "") for row in rows}
    missing = sorted(scope - present)
    scoped = [row for row in rows if str(row.get("taobao_item_id") or "") in scope]
    stats["authorized_supplement_items"] = sorted(scope)
    stats["supplement_scope_rows"] = len(scoped)
    stats["supplement_scope_missing_items"] = missing
    return scoped, missing


def authorized_line_concessions(plan) -> dict[str, Decimal]:
    """Return explicit per-SKU sub-yuan final-price concessions for one plan.

    Marker format::

        line_concession_authorized=6287431318352:0.27,6287431318353:0.06

    The marker is deliberately narrow: every entry names a platform SKU, the
    amount must be positive and strictly below one yuan, and invalid entries
    are ignored.  Signup prices remain unchanged; the concession is added only
    to that SKU's single-item reduction and is exposed in audit stats.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*line_concession_authorized\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    out: dict[str, Decimal] = {}
    for sku_id, raw_amount in re.findall(
        r"(\d{8,})\s*:\s*([0-9]+(?:\.[0-9]+)?)", matched.group(1)
    ):
        amount = _d(raw_amount)
        if amount is not None and Decimal("0") < amount <= Decimal("1"):
            out[sku_id] = amount.quantize(_CENT)
    return out


def authorized_custom_line_concessions(plan) -> dict[str, Decimal]:
    """Return exact per-SKU concessions for explicitly custom-priced SKUs.

    This is deliberately separate from the generic sub-yuan allowance.  The
    marker only names the requested platform SKU and amount; callers must also
    verify that the ERP SKU itself is a custom/consultation SKU before applying
    it.  Signup price is never changed.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*custom_line_concession_authorized\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    out: dict[str, Decimal] = {}
    for sku_id, raw_amount in re.findall(
        r"(\d{8,})\s*:\s*([0-9]+(?:\.[0-9]+)?)", matched.group(1)
    ):
        amount = _d(raw_amount)
        if amount is not None and amount > 0:
            out[sku_id] = amount.quantize(_CENT)
    return out


def _is_explicit_custom_price_sku(s) -> bool:
    """Narrow data guard for above-one-yuan user-authorized concessions."""
    if bool(getattr(s, "is_custom_placeholder", False)):
        return True
    text = " ".join((
        str(getattr(s, "sku", None) or ""),
        str(getattr(s, "product_name", None) or ""),
    ))
    return any(token in text for token in ("定制", "咨询"))


def _authorized_concession_for_sku(plan, s, sku_id: str) -> tuple[Decimal, str | None]:
    generic = authorized_line_concessions(plan).get(str(sku_id))
    if generic is not None:
        return generic, "named_sub_yuan"
    custom = authorized_custom_line_concessions(plan).get(str(sku_id))
    if custom is not None and _is_explicit_custom_price_sku(s):
        return custom, "named_custom_price_sku"
    return Decimal("0"), None


def official_ceil_enabled(db: Session) -> bool:
    """超级立减10% 官方立减是否向上取整到元 (spec §二: 待 7-20 实证, 默认按取整)。"""
    from app.services import settings_service
    raw = settings_service.get(db, OFFICIAL_CEIL_KEY, env_fallback=False)
    if raw is None or str(raw).strip() == "":
        return True
    return str(raw).strip().lower() not in ("0", "false", "off", "no")


def official_deduction(daily: Decimal, lev: Decimal, ceil_on: bool = True) -> Decimal:
    """官方立减金额 = 日常价 × 场次力度, 向上取整到元 (R9 平台实测 339.3→340、260.1→261)。
    ceil_on=False (仅 10% 场留的开关) → 精确到分不取整。"""
    exact = daily * lev
    if ceil_on:
        return exact.to_integral_value(rounding=ROUND_CEILING)
    return exact.quantize(_CENT, ROUND_HALF_UP)


def mid_buyer_inplace(promo) -> Optional[Decimal]:
    """中促到手 = 大促到手 × 1.03, 就地计算 (任务#22 口径; 不读不写 mid_buyer_price 字段本身)。"""
    big = _d(getattr(promo, "big_buyer_price", None))
    if big is None or big <= 0:
        return None
    return (big * MID_OVER_BIG_RATIO).quantize(_CENT, ROUND_HALF_UP)


def _expand_sku_ids(promo) -> list[str]:
    """一码多SKU: [主SKUID, *alt] 去重去空 (与 activity_upload_service._expand_ids 同口径)。"""
    ids: list[str] = []
    for sid in [promo.taobao_sku_id, *(promo.alt_taobao_sku_ids or [])]:
        s = str(sid).strip() if sid else ""
        if s and s not in ids:
            ids.append(s)
    return ids


def _mapped_pairs(db: Session) -> list[tuple]:
    """全部已映射 (PricingSku, PricingSkuPromo) 对, 按 product_code/sku_code 稳定排序。"""
    from app.models.pricing import PricingSku
    from app.models.pricing_ext import PricingSkuPromo
    skus = db.execute(select(PricingSku).order_by(
        PricingSku.product_code, PricingSku.sku_code)).scalars().all()
    promo_by_sku = {p.sku_code: p for p in db.execute(select(PricingSkuPromo)).scalars().all()}
    out = []
    for s in skus:
        p = promo_by_sku.get(s.sku_code)
        if p is not None and p.taobao_item_id:
            out.append((s, p))
    return out


def campaign_item_exclusions(db: Session, pairs: Optional[list[tuple]] = None) -> dict[str, dict]:
    """Effective whole-item exclusions without name/keyword inference."""
    from app.services import campaign_item_exclusion_service

    return campaign_item_exclusion_service.effective_items(
        db, pairs if pairs is not None else _mapped_pairs(db))


def price_hold_items(db: Session, plan) -> list[dict]:
    """已知历史价格线与ERP目标冲突的整品暂缓清单。

    这里只使用按 SKUID 采集且带时间戳的平台证据，不猜平台价保订单。暂缓商品不进入报名表
    或同期单品立减表；历史价格线只做资格判断，绝不参与最终到手价或单品立减计算。
    """
    from app.services import (
        campaign_price_floor_service,
        delisted_sku_service,
        no_sales_service,
    )

    tier = plan_tier(plan)
    lev = TIER_LEVERAGE[tier]
    ceil_on = official_ceil_enabled(db) if tier == "mid" else True
    evidence = campaign_price_floor_service.evidence_map(db, plan=plan)
    delisted = delisted_sku_service.get_delisted(db)
    no_sales = no_sales_service.get_no_sales(db)
    by_item: dict[str, dict] = {}
    for s, p in _mapped_pairs(db):
        if bool(getattr(s, "is_custom_placeholder", False)):
            continue
        item_id = str(getattr(p, "taobao_item_id", "") or "").strip()
        if not item_id or item_id in no_sales:
            continue
        daily = _d(getattr(s, "daily_price", None))
        if daily is None or daily <= 0:
            continue
        reasons = []
        low_price_exact = daily < Decimal("100")
        official = official_deduction(daily, lev, ceil_on and not low_price_exact)
        erp_target = (
            mid_buyer_inplace(p)
            if tier == "mid"
            else _d(getattr(p, "big_buyer_price", None))
        )
        for sid in _expand_sku_ids(p):
            # A rotated/de-bound physical SKU is no longer part of the current
            # listing. Its historical floor evidence must not hold the whole
            # replacement item after the new SKU mapping has been confirmed.
            if str(sid) in delisted:
                continue
            entry = evidence.get(str(sid)) if isinstance(evidence.get(str(sid)), dict) else {}
            min_list = _d(entry.get("min_list_price"))
            min_coupon = _d(entry.get("min_coupon_line"))
            eligible_activity_ceiling = None
            if (str(entry.get("source") or "").startswith(
                    "campaign_candidate_selectable:")
                    and entry.get("max_eligible_activity_price_observed")):
                eligible_activity_ceiling = _d(
                    entry.get("max_eligible_activity_price"))
            if min_list is not None and daily > min_list + Decimal("0.005"):
                reasons.append({
                    "type": "signup_floor",
                    "taobao_sku_id": str(sid),
                    "erp_signup_price": float(daily),
                    "platform_history_line": float(min_list),
                    "difference": float((daily - min_list).quantize(_CENT)),
                    "evidence_source": entry.get("source"),
                    "evidence_observed_at": entry.get("observed_at"),
                })
            if (eligible_activity_ceiling is not None
                    and daily > eligible_activity_ceiling + Decimal("0.005")):
                reasons.append({
                    "type": "official_candidate_activity_price_ceiling",
                    "taobao_sku_id": str(sid),
                    "erp_signup_price": float(daily),
                    "platform_eligible_activity_price": float(
                        eligible_activity_ceiling),
                    "difference": float(
                        (daily - eligible_activity_ceiling).quantize(_CENT)),
                    "evidence_source": entry.get("source"),
                    "evidence_observed_at": entry.get("observed_at"),
                })
            concession, _authorization = _authorized_concession_for_sku(
                plan, s, str(sid))
            platform_coupon_after = (
                (erp_target - concession).quantize(_CENT)
                if erp_target is not None else None
            )
            if (
                eligible_activity_ceiling is None
                and min_coupon is not None
                and platform_coupon_after is not None
                and platform_coupon_after > min_coupon + Decimal("0.005")
            ):
                planned_discount = (
                    daily - official - platform_coupon_after
                ).quantize(_CENT)
                reasons.append({
                    "type": "coupon_floor",
                    "taobao_sku_id": str(sid),
                    "erp_signup_price": float(daily),
                    "official_rate": float(lev),
                    "official_deduction": float(official),
                    "platform_coupon_after": float(platform_coupon_after),
                    "platform_history_line": float(min_coupon),
                    "difference": float((platform_coupon_after - min_coupon).quantize(_CENT)),
                    "planned_single_item_discount": float(planned_discount),
                    "authorized_concession": float(concession),
                    "single_item_discount_included_by_platform": True,
                    "evidence_source": entry.get("source"),
                    "evidence_observed_at": entry.get("observed_at"),
                })
        if not reasons:
            continue
        entry = by_item.setdefault(item_id, {
            "taobao_item_id": item_id,
            "product": s.product_name or s.product_code or "",
            "skus": [],
            "action": (
                "整品暂缓；同期单品立减计入后仍高于历史券后线；"
                "仅允许逐SKU授权的1元内微调，较大差额需等待价格线或人工决定轮换"),
        })
        entry["skus"].append({"sku_code": s.sku_code, "reasons": reasons})
    return [by_item[k] for k in sorted(by_item)]


# ── 1. 动销检查与分组 (spec §四.1) ─────────────────────────────────────────────

def group_by_sales(db: Session, days: int = 60) -> dict:
    """近{days}天淘宝订单(剔关闭单, 含刷单=平台视角)按 product_code→taobao_item_id 聚合
    → {有动销/无动销}; 并与 no_sales_service 登记表同步: 新零动销自动登记;
    已登记但出了单的只进 promote_candidates 提示转正, **不自动移除** (R6 单行道)。"""
    from datetime import date, timedelta
    from app.models.order import Order
    from app.services import no_sales_service
    from app.services.product_coder import brand_variants

    cutoff = date.today() - timedelta(days=days)
    sold_rows = db.execute(select(Order.product_code).where(
        Order.platform == "淘宝",
        Order.order_date >= cutoff,
        Order.product_code.isnot(None),
        Order.status.notin_(_CLOSED_STATUSES),
    )).scalars().all()
    sold_codes: set[str] = set()
    for pc in sold_rows:
        sold_codes |= brand_variants(pc)

    registry_cleanup = no_sales_service.sanitize_registry(db)
    mapped_pairs = _mapped_pairs(db)
    whole_item_exclusions = campaign_item_exclusions(db, mapped_pairs)
    item_codes: dict[str, set] = defaultdict(set)
    item_names: dict[str, str] = {}
    invalid_item_ids: set[str] = set()
    for s, p in mapped_pairs:
        iid = str(p.taobao_item_id).strip()
        if not no_sales_service.is_valid_item_id(iid):
            invalid_item_ids.add(iid)
            continue
        if iid in whole_item_exclusions:
            continue
        item_codes[iid].add(s.product_code or "")
        item_names.setdefault(iid, s.product_name or s.product_code or "")
    active, inactive = [], []
    for iid in sorted(item_codes):
        has_sale = any(brand_variants(c) & sold_codes for c in item_codes[iid] if c)
        (active if has_sale else inactive).append(iid)

    registered = no_sales_service.get_no_sales(db)
    newly = sorted(set(inactive) - registered)
    if newly:
        no_sales_service.add_no_sales(db, newly)     # 新零动销自动登记
    promote = sorted(registered & set(active))       # 出单了 → 提示转正, 不自动移除
    return {"有动销": active, "无动销": inactive, "days": days,
            "newly_registered": newly, "promote_candidates": promote,
            "registered": sorted(no_sales_service.get_no_sales(db)),
            "invalid_item_ids_ignored": sorted(invalid_item_ids),
            "registry_cleanup": registry_cleanup,
            "excluded_whole_items": list(whole_item_exclusions.values()),
            "item_names": item_names}


def no_sales_export_rows(db: Session, days: int = 60) -> list[dict]:
    """无动销名单导出行 (spec §四.2a 一键导出): 每个无动销淘宝商品一行
    {product_name, product_codes, taobao_item_id, sales_60d, action}。
    促成交名单以「无动销组」为准 (与飞书推送同口径); 已出单的登记品 (promote_candidates)
    也带上并标「建议转正」, 运营一张表看全。"""
    from collections import Counter
    from datetime import date, timedelta
    from app.models.order import Order
    from app.services.product_coder import brand_variants

    grouping = group_by_sales(db, days)
    cutoff = date.today() - timedelta(days=days)
    sold_rows = db.execute(select(Order.product_code).where(
        Order.platform == "淘宝",
        Order.order_date >= cutoff,
        Order.product_code.isnot(None),
        Order.status.notin_(_CLOSED_STATUSES),
    )).scalars().all()
    counts: Counter = Counter(sold_rows)

    item_codes: dict[str, set] = defaultdict(set)
    for s, p in _mapped_pairs(db):
        item_codes[str(p.taobao_item_id).strip()].add(s.product_code or "")

    def _sales_of(iid: str) -> int:
        variants: set[str] = set()
        for c in item_codes.get(iid, set()):
            if c:
                variants |= brand_variants(c)
        return sum(n for pc, n in counts.items() if pc and (brand_variants(pc) & variants))

    names = grouping["item_names"]
    rows: list[dict] = []
    for iid in grouping["无动销"]:
        rows.append({
            "product_name": names.get(iid, ""),
            "product_codes": "、".join(sorted(c for c in item_codes.get(iid, set()) if c)),
            "taobao_item_id": iid,
            "sales_60d": _sales_of(iid),
            "action": "促成交; 到手=ERP中促价; 勿撤在场报名(动销门单行道 R6)",
        })
    for iid in grouping["promote_candidates"]:
        rows.append({
            "product_name": names.get(iid, ""),
            "product_codes": "、".join(sorted(c for c in item_codes.get(iid, set()) if c)),
            "taobao_item_id": iid,
            "sales_60d": _sales_of(iid),
            "action": "已出单→建议转正: 撤 nosales 立减 → 报名大促",
        })
    return rows


# ── 2. 报名行 builder (spec §二.1, R3/R4) ─────────────────────────────────────

def _placeholder_signup_price(s, p, lev: Decimal) -> tuple[Optional[float], Optional[str]]:
    """占位SKU报名价 = min(现行值, floor(线/(1−lev)))。
    现行值 = 日常×0.9 → 500 顶 → enrolled_floor 封顶 (与现网占位口径一致);
    线 = coupon_floor_price; 无线 → 日常×0.8 保守值并在行备注标注。返回 (price|None, remark|None)。"""
    daily = _d(s.daily_price)
    if daily is None or daily <= 0:
        return None, None
    current = min((daily * Decimal("0.9")).quantize(_CENT, ROUND_HALF_UP), Decimal("500"))
    fl = _d(getattr(p, "enrolled_floor_price", None))
    if fl is not None and fl > 0:
        current = min(current, fl.quantize(_CENT, ROUND_HALF_UP))
    line = _d(getattr(p, "coupon_floor_price", None))
    remark = None
    if line is None or line <= 0:
        line = daily * PLACEHOLDER_LINE_FALLBACK_RATIO
        remark = "无券后线, 按日常价×0.8保守值封顶"
    cap = (line / (Decimal("1") - lev)).to_integral_value(rounding=ROUND_FLOOR)
    return float(min(current, cap)), remark


def signup_price_for_sku(s, p, tier: str) -> tuple[Optional[float], Optional[str]]:
    """单个 SKU 的活动报名价唯一口径。

    真 SKU 永远填 ERP 日常价；定制占位 SKU 才走既有保护价。这个函数同时供自动报名、
    定价页和下载表调用，避免各入口再次自行反推报名价。
    """
    if tier not in TIER_LEVERAGE:
        raise ValueError(f"未知活动档位 {tier!r}; 可选 {list(TIER_LEVERAGE)}")
    if bool(getattr(s, "is_custom_placeholder", False)):
        return _placeholder_signup_price(s, p, TIER_LEVERAGE[tier])
    daily = _d(getattr(s, "daily_price", None))
    return (float(daily), None) if daily is not None and daily > 0 else (None, None)


def _item_signup_rows(item_id: str, pairs: list, lev: Decimal, stats: dict) -> tuple[list, list]:
    """单商品报名行收集: 返回 (rows, missing_sku_codes)。真SKU 报名价 = 日常价 (铁则1)。"""
    rows, missing = [], []
    for s, p in pairs:
        placeholder = bool(getattr(s, "is_custom_placeholder", False))
        tier = next(k for k, v in TIER_LEVERAGE.items() if v == lev)
        price, remark = signup_price_for_sku(s, p, tier)
        if placeholder:
            if remark:
                stats["placeholder_no_line"].append({"sku_code": s.sku_code, "remark": remark})
        if price is None or price <= 0:
            missing.append(f"{s.sku_code}（{s.sku or s.product_name or '?'}）")
            continue
        for sid in _expand_sku_ids(p):
            rows.append({"taobao_item_id": item_id, "taobao_sku_id": sid,
                         "sku_code": s.sku_code, "price": round(price, 2),
                         "is_placeholder": placeholder, "remark": remark})
    return rows, missing


def _erp_listed_product_codes(db: Session) -> set[str] | None:
    """Return the ERP in-sale scope, or None for legacy/test databases with no products."""
    from app.models.product import Product

    total = db.execute(select(func.count()).select_from(Product)).scalar_one()
    if total == 0:
        return None
    return set(db.execute(
        select(Product.code).where(Product.listing_status == "在售")
    ).scalars().all())


def build_signup_rows(
        db: Session, plan, *, enforce_price_holds: bool = True,
        allow_placeholder_safe_lowering: bool = False,
) -> tuple[list[dict], dict]:
    """报名行 builder: 报名价=日常价; 过滤下架(R4)+坏价; 整品全SKU完整性断言(R3):
    任一在售已映射SKU算不出价 → 整品剔除并记 incomplete_items (半套必拒, 绝不静默)。
    返回 (rows, stats); 行 = {taobao_item_id, taobao_sku_id, sku_code, price, is_placeholder, remark}。"""
    from app.services import campaign_policy_service, delisted_sku_service, no_sales_service
    from app.services.activity_preflight_service import bad_price_product_codes

    # The repository-root contract is a runtime dependency, not documentation.
    # Missing/malformed policy must stop every generator that can create a signup file.
    campaign_policy_service.require_policy()

    lev = TIER_LEVERAGE[plan_tier(plan)]
    placeholder_live_prices = placeholder_live_prices_for_plan(db, plan)
    placeholder_expired = placeholder_price_protection_expired(plan)
    placeholder_lowering = bool(
        allow_placeholder_safe_lowering
        or placeholder_price_lowering_authorized(plan)
    )
    delisted = delisted_sku_service.get_delisted(db)
    registered_no_sales = no_sales_service.get_no_sales(db)
    bad_pc = bad_price_product_codes(db)
    holds = price_hold_items(db, plan) if enforce_price_holds else []
    held_item_ids = {x["taobao_item_id"] for x in holds}
    listed_codes = _erp_listed_product_codes(db)
    mapped_pairs = _mapped_pairs(db)
    whole_item_exclusions = campaign_item_exclusions(db, mapped_pairs)
    official_scope = official_scope_for_plan(plan)
    stats = {"rows": 0, "skipped_no_skuid": 0, "skipped_delisted": 0,
             "skipped_bad_price": 0,
             "skipped_bad_price_items": [], "incomplete_items": [], "placeholder_no_line": [],
             "excluded_price_hold_items": holds,
             "placeholder_live_prices": {
                 sid: float(price) for sid, price in placeholder_live_prices.items()},
             "placeholder_missing_live_price": [],
             "placeholder_price_protection_expired": placeholder_expired,
             "placeholder_price_lowering_authorized": placeholder_lowering,
             "placeholder_candidate_sku_ids": [],
             "placeholder_price_blocked_items": [],
             "placeholder_price_lowered": [],
             "registered_no_sales_items_included": [],
             "excluded_whole_items": list(whole_item_exclusions.values()),
             "excluded_official_exempt_items": [],
             "skipped_not_erp_listed": 0}
    stats["excluded_no_sales_items"] = []
    by_item: dict[str, list] = defaultdict(list)
    for s, p in mapped_pairs:
        if listed_codes is not None and (s.product_code or "") not in listed_codes:
            stats["skipped_not_erp_listed"] += 1
            continue
        if not p.taobao_sku_id:
            stats["skipped_no_skuid"] += 1
            continue
        if str(p.taobao_sku_id) in delisted:          # R4: 下架SKU不出报名行 (不进完整性统计)
            stats["skipped_delisted"] += 1
            continue
        by_item[str(p.taobao_item_id).strip()].append((s, p))

    rows: list[dict] = []
    for item_id, pairs in sorted(by_item.items()):
        if item_id in whole_item_exclusions:
            continue
        if _super_reduce_plan_exempt_item(plan, item_id, official_scope):
            if item_id not in stats["excluded_official_exempt_items"]:
                stats["excluded_official_exempt_items"].append(item_id)
            continue
        if item_id in held_item_ids:
            continue
        # Registered no-sales items never enter a campaign signup workbook.
        if item_id in registered_no_sales:
            stats["excluded_no_sales_items"].append(item_id)
            continue
        if all((s.product_code or "") in bad_pc for s, _ in pairs):
            stats["skipped_bad_price_items"].append(item_id)      # 坏价整品排除
            stats["skipped_bad_price"] += len(pairs)
            continue
        item_rows, missing = _item_signup_rows(item_id, pairs, lev, stats)
        if missing:                                    # R3 整品完整性: 缺一个SKU=整品拒 → 整品剔除
            stats["incomplete_items"].append({
                "taobao_item_id": item_id,
                "product": (pairs[0][0].product_name or pairs[0][0].product_code or "")[:30],
                "ok_skus": len(item_rows), "missing_skus": missing[:10]})
            continue
        stats["placeholder_candidate_sku_ids"].extend(
            str(row["taobao_sku_id"])
            for row in item_rows if row.get("is_placeholder"))
        blocked_placeholders = []
        for row in item_rows:
            if not row.get("is_placeholder"):
                continue
            sid = str(row["taobao_sku_id"])
            live_price = placeholder_live_prices.get(sid)
            if live_price is None:
                detail = {
                    "taobao_item_id": item_id,
                    "taobao_sku_id": sid,
                    "sku_code": row["sku_code"],
                    "safe_cap": row["price"],
                    "current_live_price": None,
                }
                # The current-plan authorization covers only a *known* live
                # price that is above the already-computed safe cap.  It must
                # never turn missing platform evidence into an inferred price.
                stats["placeholder_missing_live_price"].append(detail)
                continue
            generated = _d(row["price"]) or Decimal("0")
            if live_price > generated:
                detail = {
                    "taobao_item_id": item_id,
                    "taobao_sku_id": sid,
                    "sku_code": row["sku_code"],
                    "safe_cap": float(generated),
                    "current_live_price": float(live_price),
                }
                if placeholder_expired or placeholder_lowering:
                    row["remark"] = (
                        "用户已授权定制咨询规格使用保护报名价"
                        if placeholder_lowering else
                        "价保已确认到期，按最低普惠券后价安全上限报名"
                    )
                    stats["placeholder_price_lowered"].append({
                        **detail,
                        "authorization": (
                            "current_plan_user_decision"
                            if placeholder_lowering else
                            "price_protection_expired"
                        ),
                    })
                else:
                    blocked_placeholders.append(detail)
        if blocked_placeholders:
            stats["placeholder_price_blocked_items"].append({
                "taobao_item_id": item_id,
                "product": (pairs[0][0].product_name or pairs[0][0].product_code or "")[:30],
                "placeholders": blocked_placeholders,
                "action": "价保到期未确认，整品暂缓；禁止用高保护价覆盖券后安全上限",
            })
            continue
        rows.extend(item_rows)
    stats["rows"] = len(rows)
    stats["placeholder_candidate_sku_ids"] = sorted(set(
        stats["placeholder_candidate_sku_ids"]))
    return rows, stats


# ── 3. 单品立减 builder (spec §二 立减公式逐字) ────────────────────────────────

def _campaign_discount_row(s, p, tier: str, lev: Decimal, ceil_on: bool,
                           stats: dict, *,
                           base_price: Optional[Decimal] = None) -> Optional[dict]:
    """有动销 SKU 立减只对齐 ERP 场次目标价。

    最低普惠券后价只用于资格校验；单品立减默认只对齐 ERP 目标，绝不能
    未经授权用历史价格线压低。逐 SKU 的 1 元内授权由 build_discount_rows 追加。
    """
    daily = _d(s.daily_price)
    base = base_price if base_price is not None else daily
    target0 = mid_buyer_inplace(p) if tier == "mid" else _d(getattr(p, "big_buyer_price", None))
    if target0 is None or target0 <= 0:
        stats["skipped_no_target"] += 1
        return None
    target, concession = target0, Decimal("0")
    # 2026-07-24 平台实证：低价 SKU 的官方立减按精确比例计算到分。
    # 狂暑季 ¥30×12%=¥3.60；超级立减 ¥25×10%=¥2.50，均不向上取整到元。
    # 普通价位仍沿用整元向上取整规则。
    low_price_exact = base < Decimal("100")
    official = official_deduction(base, lev, ceil_on and not low_price_exact)
    if low_price_exact:
        stats["official_low_price_exact"] += 1
    deduct = (base - official - target).quantize(_CENT)
    if deduct <= 0:
        stats["skipped_no_deduct"] += 1                   # 官方立减已够 → 不出行(不给假数)
        return None
    return {"deduct": float(deduct), "kind": "campaign", "target_price": float(target),
            "official": float(official), "concession": float(concession),
            "calculation_base": float(base)}


def _nosales_discount_row(s, p, stats: dict, tier: str = "mid",
                          official: Decimal = Decimal("0"), *,
                          base_price: Optional[Decimal] = None) -> Optional[dict]:
    """无动销 SKU：不报名活动，只靠单品立减直达到当前场次目标。

    大促档位直接到 ERP 大促买家价；超级立减档位精确使用 ERP 中促价。
    是否叠加官方立减由活动实时范围决定；不套券后线。
    """
    daily = _d(s.daily_price)
    base = base_price if base_price is not None else daily
    if tier in ("big", "big618"):
        target = _d(getattr(p, "big_buyer_price", None))
    else:
        mid = mid_buyer_inplace(p)
        target = mid.quantize(_CENT) if mid is not None else None
    if target is None or target <= 0:
        stats["skipped_no_target"] += 1
        return None
    deduct = (base - official - target).quantize(_CENT)
    if deduct <= 0:
        stats["skipped_no_deduct"] += 1
        return None
    return {"deduct": float(deduct), "kind": "nosales", "target_price": float(target),
            "official": float(official), "concession": 0.0,
            "calculation_base": float(base)}


def discount_for_sku(db: Session, s, p, tier: str,
                     no_sales_items: Optional[set[str]] = None, *,
                     official_applies: Optional[bool] = None) -> Optional[dict]:
    """单个真 SKU 的单品立减计算口径，供下载参考表复用。

    自动上传仍由 build_discount_rows 负责映射、下架、坏价等资格过滤；本函数只负责价格数学。
    """
    if tier not in TIER_LEVERAGE:
        raise ValueError(f"未知活动档位 {tier!r}; 可选 {list(TIER_LEVERAGE)}")
    if bool(getattr(s, "is_custom_placeholder", False)):
        return None
    daily = _d(getattr(s, "daily_price", None))
    if daily is None or daily <= 0:
        return None
    from app.services import no_sales_service
    stats = {"skipped_no_target": 0, "skipped_no_deduct": 0,
             "rotation_suggested": [], "line_concessions": [],
             "official_low_price_exact": 0}
    item_id = str(getattr(p, "taobao_item_id", "") or "").strip()
    nosales = no_sales_items if no_sales_items is not None else no_sales_service.get_no_sales(db)
    if item_id and item_id in nosales:
        # Generic pricing pages/downloads do not know the live activity scope.
        # A blank is safer than publishing a deceptively precise discount amount.
        if official_applies is None:
            return None
        official = Decimal("0")
        if official_applies:
            low_price_exact = daily < Decimal("100")
            lev = TIER_LEVERAGE[tier]
            ceil_on = True if tier != "mid" else official_ceil_enabled(db)
            official = official_deduction(
                daily, lev, ceil_on and not low_price_exact)
        return _nosales_discount_row(s, p, stats, tier, official=official)
    lev = TIER_LEVERAGE[tier]
    ceil_on = True if tier != "mid" else official_ceil_enabled(db)
    return _campaign_discount_row(s, p, tier, lev, ceil_on, stats)


def build_discount_rows(
        db: Session, plan, *, no_sales_items: Optional[set[str]] = None
) -> tuple[list[dict], dict]:
    """单品立减 builder (spec §二):
      大促12% / 618双11 15%: 立减 = 日常 − ceil(日常×lev) − 大促到手 − 已授权微调
      超级立减10%:            立减 = 日常 − ceil(日常×10%) − 中促到手 − 已授权微调
      无动销(登记表):         立减 = 日常 − ERP中促到手, 占位不出行
    返回 (rows, stats); 行含 taobao_item_id/taobao_sku_id/sku_code/deduct/target_price/kind。"""
    from app.services import delisted_sku_service, no_sales_service
    from app.services.activity_preflight_service import bad_price_product_codes

    tier = plan_tier(plan)
    lev = TIER_LEVERAGE[tier]
    ceil_on = True if lev != TIER_LEVERAGE["mid"] else official_ceil_enabled(db)
    official_scope = official_scope_for_plan(plan)
    live_activity_prices = current_activity_prices_for_plan(plan)
    nosales = (
        no_sales_service.get_no_sales(db)
        if no_sales_items is None else set(no_sales_items)
    )
    delisted = delisted_sku_service.get_delisted(db)
    bad_pc = bad_price_product_codes(db)
    holds = price_hold_items(db, plan)
    held_item_ids = {x["taobao_item_id"] for x in holds}
    listed_codes = _erp_listed_product_codes(db)
    mapped_pairs = _mapped_pairs(db)
    whole_item_exclusions = campaign_item_exclusions(db, mapped_pairs)
    stats = {"tier": tier, "official_ceil": ceil_on, "rows": 0, "skipped_no_skuid": 0,
             "skipped_delisted": 0, "skipped_bad_price": 0, "skipped_placeholder": 0,
             "skipped_price_hold": 0, "excluded_price_hold_items": holds,
             "skipped_no_daily": 0, "skipped_no_target": 0, "skipped_no_deduct": 0,
             "line_concessions": [], "rotation_suggested": [],
             "official_low_price_exact": 0,
             "live_activity_price_overrides": [],
             "excluded_whole_items": list(whole_item_exclusions.values()),
             "excluded_official_exempt_items": [],
             "skipped_not_erp_listed": 0,
             "official_scope": {
                 "configured": official_scope["configured"],
                 "all_store": official_scope["all_store"],
                 "active_items": sorted(official_scope["active_items"]),
                 "exempt_items": sorted(official_scope["exempt_items"]),
                 "errors": official_scope["errors"],
             }}
    rows: list[dict] = []
    for s, p in mapped_pairs:
        if listed_codes is not None and (s.product_code or "") not in listed_codes:
            stats["skipped_not_erp_listed"] += 1
            continue
        if not p.taobao_sku_id:
            stats["skipped_no_skuid"] += 1
            continue
        if str(p.taobao_sku_id) in delisted:
            stats["skipped_delisted"] += 1
            continue
        if (s.product_code or "") in bad_pc:
            stats["skipped_bad_price"] += 1
            continue
        if getattr(s, "is_custom_placeholder", False):   # 占位不出行 (spec §二.4)
            stats["skipped_placeholder"] += 1
            continue
        daily = _d(s.daily_price)
        if daily is None or daily <= 0:
            stats["skipped_no_daily"] += 1
            continue
        item_id = str(p.taobao_item_id).strip()
        if item_id in whole_item_exclusions:
            continue
        if _super_reduce_plan_exempt_item(plan, item_id, official_scope):
            if item_id not in stats["excluded_official_exempt_items"]:
                stats["excluded_official_exempt_items"].append(item_id)
            continue
        if item_id in held_item_ids:
            stats["skipped_price_hold"] += 1
            continue
        for sid in _expand_sku_ids(p):
            sid = str(sid)
            base = live_activity_prices.get(sid, daily)
            if base != daily:
                stats["live_activity_price_overrides"].append({
                    "taobao_item_id": item_id,
                    "taobao_sku_id": sid,
                    "daily_price": float(daily),
                    "activity_price": float(base),
                })
            if item_id in nosales:
                official = Decimal("0")
                if _official_applies(item_id, official_scope):
                    low_price_exact = base < Decimal("100")
                    official = official_deduction(
                        base, lev, ceil_on and not low_price_exact)
                    if low_price_exact:
                        stats["official_low_price_exact"] += 1
                core = _nosales_discount_row(
                    s, p, stats, tier, official=official, base_price=base)
            else:
                core = _campaign_discount_row(
                    s, p, tier, lev, ceil_on, stats, base_price=base)
            if core is None:
                continue
            row = {"taobao_item_id": item_id, "taobao_sku_id": sid,
                   "sku_code": s.sku_code, **core}
            concession, authorization = _authorized_concession_for_sku(
                plan, s, str(sid))
            if concession > 0 and concession < _d(row["target_price"]):
                deduct = (_d(row["deduct"]) + concession).quantize(_CENT)
                target = (_d(row["target_price"]) - concession).quantize(_CENT)
                row.update({
                    "deduct": float(deduct),
                    "target_price": float(target),
                    "concession": float(concession),
                })
                stats["line_concessions"].append({
                    "taobao_item_id": item_id,
                    "taobao_sku_id": str(sid),
                    "sku_code": s.sku_code,
                    "amount": float(concession),
                    "erp_target": float(_d(core["target_price"])),
                    "authorized_target": float(target),
                    "authorization": authorization,
                })
            rows.append(row)
    stats["rows"] = len(rows)
    return rows, stats


# ── 4. preflight (spec §三 R1~R12 静态可查项) ─────────────────────────────────

_STATIC_REMINDERS = [
    ("R5", "warn", "已报名非草稿的品批量导入必被拒 — 推送前 wizard 卡点确认该品已在千牛撤销"),
    ("R7", "pass", "活动报名禁止 SKU 轮换；SKUID 变化必须先修正正式映射并重新取证"),
    ("R8", "info", "刷新 SKU 映射必须同事务清线 (coupon_floor/enrolled_floor 挂编码、线跟 sid)"),
    ("R10", "info", "回执真相以千牛「批量操作记录」最新一条为准, WA published 回执不可信"),
    ("R11", "warn", "同品同时只能一个单品立减生效 — 推送前先在千牛删除在场旧批, 否则新批不生效"),
    ("R12", "warn", "单品立减导入即生效、报名导入即成功, 均无草稿不可逆 — 每步确认后再推"),
]


def _check_r1(db: Session, plan) -> dict:
    """R1 静态代理: 报名价(=日常价) > 已生效活动价硬底(enrolled_floor_price) → 必被
    "≤近15天最低标价/已生效价"拦, 提示轮换。(真实15天标价窗口在平台侧, 离线取不到 — 交回执自愈。)"""
    items = []
    for item in price_hold_items(db, plan):
        skus = [
            sku for sku in item["skus"]
            if any(r["type"] == "signup_floor" for r in sku["reasons"])
        ]
        if skus:
            items.append({**item, "skus": skus})
    return {"rule": "R1", "level": "warn" if items else "pass",
            "title": "报名价历史线冲突：相关整品已暂缓，不轮换、不降价迁就", "items": items}


def _check_price_math(db: Session, plan, signup_rows: list[dict],
                      discount_rows: list[dict]) -> dict:
    """逐SKU验算报名价与到手公式，0.01元偏差也阻断。"""
    pair_by_sid: dict[str, tuple] = {}
    for s, p in _mapped_pairs(db):
        for sid in _expand_sku_ids(p):
            pair_by_sid[sid] = (s, p)
    errors = []
    tier = plan_tier(plan)
    for row in signup_rows:
        pair = pair_by_sid.get(str(row.get("taobao_sku_id")))
        if row.get("is_placeholder"):
            safe_cap = None
            if pair:
                safe_cap, _remark = _placeholder_signup_price(
                    pair[0], pair[1], TIER_LEVERAGE[tier])
            price = _d(row.get("price"))
            if (safe_cap is None or price is None
                    or price > _d(safe_cap) + Decimal("0.005")):
                errors.append({
                    "sku_id": row.get("taobao_sku_id"),
                    "sku_code": row.get("sku_code"),
                    "check": "placeholder_signup_within_coupon_floor_cap",
                    "safe_cap": safe_cap,
                    "signup_price": float(price) if price is not None else None,
                })
            continue
        daily = _d(getattr(pair[0], "daily_price", None)) if pair else None
        price = _d(row.get("price"))
        if daily is None or price is None or abs(daily - price) > Decimal("0.005"):
            errors.append({
                "sku_id": row.get("taobao_sku_id"), "sku_code": row.get("sku_code"),
                "check": "signup_price_equals_daily",
                "daily": float(daily) if daily is not None else None,
                "signup_price": float(price) if price is not None else None,
            })
    for row in discount_rows:
        pair = pair_by_sid.get(str(row.get("taobao_sku_id")))
        daily = _d(getattr(pair[0], "daily_price", None)) if pair else None
        base = _d(row.get("calculation_base")) or daily
        official = _d(row.get("official")) or Decimal("0")
        deduct = _d(row.get("deduct"))
        target = _d(row.get("target_price"))
        if pair and bool(getattr(pair[0], "is_custom_placeholder", False)):
            errors.append({
                "sku_id": row.get("taobao_sku_id"), "sku_code": row.get("sku_code"),
                "check": "placeholder_must_not_have_discount",
            })
            continue
        landing = (
            (base - official - deduct).quantize(_CENT)
            if base is not None and deduct is not None else None
        )
        if landing is None or target is None or abs(landing - target) > Decimal("0.005"):
            errors.append({
                "sku_id": row.get("taobao_sku_id"), "sku_code": row.get("sku_code"),
                "check": "activity_base_minus_official_minus_discount_equals_target",
                "daily": float(daily) if daily is not None else None,
                "calculation_base": float(base) if base is not None else None,
                "official": float(official), "deduct": float(deduct) if deduct is not None else None,
                "target": float(target) if target is not None else None,
                "calculated_landing": float(landing) if landing is not None else None,
            })
    return {
        "rule": "R13",
        "level": "error" if errors else "pass",
        "title": ("逐SKU最终价格验算（不代替R2报名资格门）：报名价=ERP日常价；"
                  "本期实际活动价（未回读时为日常价）−官方立减−单品立减=ERP目标价"),
        "items": errors[:100],
        "checked": {"signup_rows": len(signup_rows), "discount_rows": len(discount_rows)},
    }


def _check_official_scope(
        db: Session, plan, *, no_sales_items: Optional[set[str]] = None) -> dict:
    """R15: no-sales rows require live evidence of official-discount applicability."""
    from app.services import no_sales_service

    no_sales = sorted(
        no_sales_service.get_no_sales(db)
        if no_sales_items is None else set(no_sales_items)
    )
    scope = official_scope_for_plan(plan)
    errors = list(scope["errors"])
    if no_sales and not scope["configured"] and not errors:
        errors.append(
            "存在无动销商品，但计划未记录官方立减生效/豁免范围；禁止按0元官方立减猜测")
    return {
        "rule": "R15",
        "level": "error" if errors else "pass",
        "title": "官方立减逐商品范围：必须来自当前活动页，缺失即停止单品立减",
        "items": [{
            "errors": errors,
            "no_sales_items": no_sales,
            "all_store": scope["all_store"],
            "active_items": sorted(scope["active_items"]),
            "exempt_items": sorted(scope["exempt_items"]),
        }],
    }


def _check_super_reduce_publish_window(
        plan, *, now: Optional[datetime] = None, enforce: bool = False) -> dict:
    """R18: plan dates never imply delaying or withdrawing an authorized signup."""
    return {
        "rule": "R18", "level": "pass",
        "title": "长期超级立减按报名授权执行；计划日期不得推导撤销或延迟",
        "items": [],
    }


def _authorized_withdrawal_items(plan) -> set[str]:
    """Exact current-user authorization marker for a destructive withdrawal."""
    remark = str(getattr(plan, "remark", "") or "")
    match = re.search(
        r"(?:^|[;\n；])\s*user_authorized_campaign_withdrawal\s*=\s*([^;\n；]*)",
        remark,
    )
    if not match:
        return set()
    return {value.strip() for value in match.group(1).split(",") if value.strip()}


def _check_super_reduce_discount_coverage(
        db: Session, plan, signup_rows: list[dict], discount_rows: list[dict]) -> dict:
    """R19: every official-Super-Reduce real SKU must land on the ERP target."""
    if str(getattr(plan, "campaign_type", "")) != "super_reduce":
        return {
            "rule": "R19", "level": "pass",
            "title": "超级立减官方范围与单品立减逐SKU完整配对",
            "items": [],
        }
    qualified = platform_qualified_items(plan)
    signup_by_sid = {
        str(row.get("taobao_sku_id") or ""): row
        for row in signup_rows
        if not row.get("is_placeholder")
        and (not qualified or str(row.get("taobao_item_id") or "") in qualified)
    }
    discount_sids = {
        str(row.get("taobao_sku_id") or "") for row in discount_rows
    }
    pair_by_sid: dict[str, tuple] = {}
    for sku, promo in _mapped_pairs(db):
        for sku_id in _expand_sku_ids(promo):
            pair_by_sid[str(sku_id)] = (sku, promo)
    problems: list[dict] = []
    lev = TIER_LEVERAGE["mid"]
    ceil_on = official_ceil_enabled(db)
    for sku_id, row in sorted(signup_by_sid.items()):
        pair = pair_by_sid.get(sku_id)
        signup_price = _d(row.get("price"))
        target = mid_buyer_inplace(pair[1]) if pair else None
        if signup_price is None or target is None:
            problems.append({
                "taobao_item_id": row.get("taobao_item_id"),
                "taobao_sku_id": sku_id,
                "sku_code": row.get("sku_code"),
                "check": "missing_signup_price_or_target",
            })
            continue
        low_price_exact = signup_price < Decimal("100")
        official = official_deduction(
            signup_price, lev, ceil_on and not low_price_exact)
        after_official = (signup_price - official).quantize(_CENT)
        if target > after_official + Decimal("0.005"):
            problems.append({
                "taobao_item_id": row.get("taobao_item_id"),
                "taobao_sku_id": sku_id,
                "sku_code": row.get("sku_code"),
                "check": "official_discount_already_below_erp_target",
                "after_official": float(after_official),
                "erp_target": float(target),
            })
        elif target < after_official - Decimal("0.005") and sku_id not in discount_sids:
            problems.append({
                "taobao_item_id": row.get("taobao_item_id"),
                "taobao_sku_id": sku_id,
                "sku_code": row.get("sku_code"),
                "check": "missing_paired_single_item_discount",
                "required_deduct": float((after_official - target).quantize(_CENT)),
                "after_official": float(after_official),
                "erp_target": float(target),
            })
    return {
        "rule": "R19", "level": "error" if problems else "pass",
        "title": "超级立减逐SKU配对：官方10%与同期单品立减必须共同精确落到ERP目标",
        "items": problems[:500],
        "checked": len(signup_by_sid),
    }


def _check_placeholder_live_prices(signup_stats: dict) -> dict:
    missing = signup_stats.get("placeholder_missing_live_price") or []
    blocked = signup_stats.get("placeholder_price_blocked_items") or []
    return {
        "rule": "R16",
        "level": "error" if missing else ("warn" if blocked else "pass"),
        "title": (
            "占位SKU保护价：缺平台当前价即停止；高保护价不得覆盖券后安全上限"
            if not blocked else
            f"占位SKU价保到期未确认：已整品暂缓{len(blocked)}品，其余安全行可继续"
        ),
        "items": missing,
        "blocked_items": blocked,
        "lowered": signup_stats.get("placeholder_price_lowered") or [],
        "price_protection_expired": bool(
            signup_stats.get("placeholder_price_protection_expired")),
    }


def _check_campaign_policy() -> dict:
    from app.services import campaign_policy_service
    try:
        policy = campaign_policy_service.public_policy()
    except Exception as exc:  # noqa: BLE001 - surfaced as a structured hard gate
        return {
            "rule": "R0",
            "level": "error",
            "title": "根目录活动报名规则缺失或无效，程序已停止",
            "items": [{"error": str(exc)}],
        }
    return {
        "rule": "R0",
        "level": "pass",
        "title": "已加载根目录活动报名唯一规则（程序自动执行；AI禁止提交/改价/重试）",
        "items": [{
            "policy_id": policy["policy_id"],
            "version": policy.get("version"),
            "sha256": policy.get("sha256"),
        }],
    }


def _check_price_floor_evidence(db: Session, plan, signup_rows: list[dict]) -> dict:
    """Require fresh, per-SKUID evidence for both platform qualification lines."""
    from app.services import campaign_policy_service, campaign_price_floor_service

    max_age = campaign_policy_service.floor_evidence_max_age_hours()
    evidence = campaign_price_floor_service.evidence_map(db, plan=plan)
    terminal_acceptance = _fresh_terminal_platform_acceptance(
        plan, signup_rows, max_age)
    terminal_items = terminal_acceptance["item_ids"]
    coupon_floor_qualification = _fresh_terminal_coupon_floor_qualification(
        plan, signup_rows, max_age)
    coupon_floor_items = coupon_floor_qualification["item_ids"]
    problems: list[dict] = []
    authorized_new_items = (
        new_item_no_history_authorized_items(plan)
        | authorized_sku_refresh_items(plan)
    )
    authorized_new_rows: list[dict] = []
    terminal_accepted_rows: list[dict] = []
    terminal_coupon_floor_rows: list[dict] = []
    candidate_ceiling_rows: list[dict] = []
    sku_by_id = {
        mapped_sid: sku
        for sku, promo in _mapped_pairs(db)
        for mapped_sid in _expand_sku_ids(promo)
    }
    seen: set[str] = set()
    for row in signup_rows:
        if row.get("is_placeholder"):
            continue
        sid = str(row.get("taobao_sku_id") or "").strip()
        if not sid or sid in seen:
            continue
        seen.add(sid)
        entry = evidence.get(sid) if isinstance(evidence.get(sid), dict) else {}
        item_id = str(row.get("taobao_item_id") or "")
        if item_id in terminal_items:
            terminal_accepted_rows.append({
                "taobao_item_id": item_id,
                "taobao_sku_id": sid,
                "sku_code": row.get("sku_code"),
                "signup_price": row.get("price"),
                "source": "fresh_terminal_platform_qualification",
            })
            continue
        if item_id in coupon_floor_items:
            terminal_coupon_floor_rows.append({
                "taobao_item_id": item_id,
                "taobao_sku_id": sid,
                "sku_code": row.get("sku_code"),
                "signup_price": row.get("price"),
                "source": "fresh_terminal_coupon_floor_only_qualification",
            })
            continue
        if not entry and item_id in authorized_new_items:
            sku = sku_by_id.get(sid)
            erp_list_price = _d(getattr(sku, "list_price", None))
            signup_price = _d(row.get("price"))
            if (erp_list_price is None or signup_price is None
                    or signup_price > erp_list_price):
                problems.append({
                    "taobao_item_id": item_id,
                    "taobao_sku_id": sid,
                    "sku_code": row.get("sku_code"),
                    "missing": ["current_erp_list_price_ceiling"],
                    "signup_price": float(signup_price) if signup_price is not None else None,
                    "erp_current_list_price": (
                        float(erp_list_price) if erp_list_price is not None else None),
                    "source": "explicit_new_item_without_platform_history",
                })
                continue
            authorized_new_rows.append({
                "taobao_item_id": item_id,
                "taobao_sku_id": sid,
                "sku_code": row.get("sku_code"),
                "signup_price": float(signup_price),
                "erp_current_list_price": float(erp_list_price),
                "reason": (
                    "user_confirmed_rotated_sku_without_platform_history"
                    if item_id in authorized_sku_refresh_items(plan)
                    else "user_confirmed_new_link_without_platform_history"
                ),
            })
            continue
        candidate_ceiling = (
            str(entry.get("source") or "").startswith(
                "campaign_candidate_selectable:")
            and entry.get("min_list_price") is not None
            and entry.get("min_list_price_observed")
            and entry.get("max_eligible_activity_price") is not None
            and entry.get("max_eligible_activity_price_observed")
        )
        missing = [] if candidate_ceiling else [
            key for key in ("min_list_price", "min_coupon_line")
            if entry.get(key) is None and not entry.get(f"{key}_observed")]
        age = campaign_price_floor_service.evidence_age_hours(entry)
        stale = age is None or age > max_age
        if missing or stale:
            problems.append({
                "taobao_item_id": row.get("taobao_item_id"),
                "taobao_sku_id": sid,
                "sku_code": row.get("sku_code"),
                "missing": missing,
                "observed_at": entry.get("observed_at"),
                "age_hours": round(age, 2) if age is not None else None,
                "max_age_hours": max_age,
                "source": entry.get("source"),
            })
        elif candidate_ceiling:
            candidate_ceiling_rows.append({
                "taobao_item_id": item_id,
                "taobao_sku_id": sid,
                "sku_code": row.get("sku_code"),
                "signup_price": row.get("price"),
                "min_list_price": entry.get("min_list_price"),
                "max_eligible_activity_price": entry.get(
                    "max_eligible_activity_price"),
                "source": entry.get("source"),
            })
    return {
        "rule": "R17",
        "level": "error" if problems else "pass",
        "title": (
            "逐SKUID平台价格证据：已报名商品须有最低标价和最低普惠券后价；"
            "未报名候选须有最低标价和平台合格建议活动价上限；缺失/过期即停止"
        ),
        "items": problems[:500],
        "checked": len(seen),
        "max_age_hours": max_age,
        "authorized_new_item_rows": authorized_new_rows,
        "platform_terminal_accepted_rows": terminal_accepted_rows,
        "platform_terminal_coupon_floor_rows": terminal_coupon_floor_rows,
        "platform_candidate_ceiling_rows": candidate_ceiling_rows,
        "platform_terminal_acceptance": {
            "observed_at": terminal_acceptance["observed_at"],
            "age_hours": (
                round(terminal_acceptance["age_hours"], 2)
                if terminal_acceptance["age_hours"] is not None else None
            ),
            "digest_matches": terminal_acceptance["digest_matches"],
            "accepted_item_count": len(terminal_items),
        },
        "platform_terminal_coupon_floor_qualification": {
            "observed_at": coupon_floor_qualification["observed_at"],
            "age_hours": (
                round(coupon_floor_qualification["age_hours"], 2)
                if coupon_floor_qualification["age_hours"] is not None else None
            ),
            "digest_matches": coupon_floor_qualification["digest_matches"],
            "item_count": len(coupon_floor_items),
        },
    }


def preflight(
        db: Session, plan, *, no_sales_items: Optional[set[str]] = None,
        exact_item_scope: Optional[set[str]] = None,
        allow_placeholder_safe_lowering: bool = False,
) -> list[dict]:
    """R0~R21 read-only pre-submit checks; never uploads to QianNiu."""
    from app.services import no_sales_service
    from app.services import campaign_price_protection_service

    policy_check = _check_campaign_policy()
    if policy_check["level"] == "error":
        return [policy_check]
    _srows, sstats = build_signup_rows(
        db, plan,
        allow_placeholder_safe_lowering=allow_placeholder_safe_lowering,
    )
    _drows, dstats = build_discount_rows(
        db, plan, no_sales_items=no_sales_items)
    supplement_scope = authorized_supplement_items(plan)
    exact_scope = {
        str(item_id).strip() for item_id in (exact_item_scope or set())
        if str(item_id).strip()
    }
    if exact_scope:
        # A plan-resume service can deliberately preflight one immutable,
        # operator-approved scope.  It must not inherit an empty/stale platform
        # qualification marker that would otherwise make R16/R17 check zero
        # rows and create a false pass.
        _srows = [row for row in _srows
                  if str(row.get("taobao_item_id") or "") in exact_scope]
        _drows = [row for row in _drows
                  if str(row.get("taobao_item_id") or "") in exact_scope]
        # Builders intentionally collect diagnostics for the whole plan.  A
        # one-shot resume must scope the item-level R3/R16 diagnostics too;
        # otherwise unrelated held products can block an immutable reviewed
        # item even though they cannot be uploaded by this execution.
        for key in (
            "incomplete_items",
            "placeholder_missing_live_price",
            "placeholder_price_blocked_items",
            "placeholder_price_lowered",
        ):
            sstats[key] = [
                row for row in (sstats.get(key) or [])
                if str(row.get("taobao_item_id") or "") in exact_scope
            ]
        sstats["exact_preflight_scope_items"] = sorted(exact_scope)
        sstats["exact_preflight_scope_rows"] = len(_srows)
        dstats["exact_preflight_scope_items"] = sorted(exact_scope)
        dstats["exact_preflight_scope_rows"] = len(_drows)
    elif platform_scope_present(plan):
        qualified = platform_qualified_items(plan)
        no_sales = platform_no_sales_items(plan)
        signup_allowed = qualified | supplement_scope
        discount_allowed = qualified | no_sales | supplement_scope
        _srows = [row for row in _srows
                  if str(row.get("taobao_item_id") or "") in signup_allowed]
        _drows = [row for row in _drows
                  if str(row.get("taobao_item_id") or "") in discount_allowed]
        sstats["platform_preflight_scope_items"] = sorted(signup_allowed)
        sstats["platform_preflight_scope_rows"] = len(_srows)
        dstats["platform_preflight_scope_items"] = sorted(discount_allowed)
        dstats["platform_preflight_scope_rows"] = len(_drows)
    if supplement_scope:
        # Corrective runs must preflight the exact same item set that will be
        # uploaded.  An unrelated qualified item with stale/missing evidence
        # may be held, but must not block a fully evidenced scoped retry.
        _srows = [
            row for row in _srows
            if str(row.get("taobao_item_id") or "") in supplement_scope
        ]
        _drows = [
            row for row in _drows
            if str(row.get("taobao_item_id") or "") in supplement_scope
        ]
        sstats["supplement_preflight_scope_items"] = sorted(supplement_scope)
        sstats["supplement_preflight_scope_rows"] = len(_srows)
        dstats["supplement_preflight_scope_items"] = sorted(supplement_scope)
        dstats["supplement_preflight_scope_rows"] = len(_drows)
    nosales = sorted(no_sales_service.get_no_sales(db))
    holds = price_hold_items(db, plan)
    coupon_holds = []
    for item in holds:
        skus = [
            sku for sku in item["skus"]
            if any(r["type"] == "coupon_floor" for r in sku["reasons"])
        ]
        if skus:
            coupon_holds.append({**item, "skus": skus})
    checks = [
        policy_check,
        _check_r1(db, plan),
        {"rule": "R2", "level": "warn" if coupon_holds else "pass",
         "title": ("报名资格硬门：活动价−官方立减−同期单品立减必须≤近15天最低普惠券后价；"
                   "任一SKU仍冲突则整品暂缓"),
         "items": coupon_holds, "audit": dstats["line_concessions"]},
        {"rule": "R3", "level": "error" if sstats["incomplete_items"] else "pass",
         "title": "报名整品全SKU完整性 (缺SKU=整品拒)", "items": sstats["incomplete_items"]},
        {"rule": "R4", "level": "info", "title": "下架SKU已过滤不出行 (回执自愈登记)",
         "items": [{"skipped_delisted_signup": sstats["skipped_delisted"],
                    "skipped_delisted_discount": dstats["skipped_delisted"]}]},
        {"rule": "R6", "level": "warn" if nosales else "pass",
         "title": "无动销报名硬排除（不进入活动报名；同期单品立减兜底）",
         "items": nosales,
         "excluded_signup_items": sstats.get("excluded_no_sales_items") or []},
        {"rule": "R9", "level": "pass",
         "title": "官方立减向上取整到元已内建 (10%场开关 campaign_official_ceil)",
         "items": [{"official_ceil": dstats["official_ceil"]}]},
        _check_price_math(db, plan, _srows, _drows),
        campaign_price_protection_service.rule_check(plan),
        _check_official_scope(db, plan, no_sales_items=no_sales_items),
        _check_placeholder_live_prices(sstats),
        _check_price_floor_evidence(db, plan, _srows),
        _check_super_reduce_publish_window(plan),
        _check_super_reduce_discount_coverage(db, plan, _srows, _drows),
        _check_campaign_identity(plan),
        _check_signup_shipping(plan, _srows),
    ]
    checks += [{"rule": r, "level": lv, "title": t, "items": []} for r, lv, t in _STATIC_REMINDERS]
    requested_rotation = requested_sku_refresh_items(plan)
    if requested_rotation:
        r7 = next(check for check in checks if check["rule"] == "R7")
        r7.update({
            "level": "error",
            "title": "禁止 SKU 轮换；发现旧轮换标记即停止并要求修正正式映射",
            "items": sorted(requested_rotation),
        })
    checks.sort(key=lambda c: int(c["rule"][1:]))
    return checks


# ── 5. 推送编排 (复用 web_agent_service upload_file → wait_job) ────────────────

def _build_discount_xlsx(rows: list[dict]) -> bytes:
    """单品立减上传表 (表头与淘宝模板逐字一致, 复用 data_export_service._TB_DISCOUNT_HEADERS)。"""
    import io
    import openpyxl
    from app.services.data_export_service import _TB_DISCOUNT_HEADERS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "单品立减"
    for ci, h in enumerate(_TB_DISCOUNT_HEADERS, start=1):
        ws.cell(1, ci, h)
    r, seen = 2, set()
    for row in rows:
        sid = row["taobao_sku_id"]
        if sid in seen:
            continue                                   # 重复映射保首行 (平台"存在重复的SKUID"整品拒)
        seen.add(sid)
        ws.cell(r, 1, str(row["taobao_item_id"])).number_format = "@"
        ws.cell(r, 2, sid).number_format = "@"
        ws.cell(r, 3, float(row["deduct"])).number_format = "0.00"
        r += 1
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_signup_xlsx(
        rows: list[dict], shipping_days_by_item: dict[str, int] | None = None) -> bytes:
    """大促报名上传表: 官方模板 promo_signup_sku.xlsx (保留说明sheet+前3行表头, 数据从第4行)。"""
    import io
    from pathlib import Path
    import openpyxl
    tpl = Path(__file__).resolve().parent.parent / "assets" / "taobao_templates" / "promo_signup_sku.xlsx"
    wb = openpyxl.load_workbook(tpl)
    ws = wb["商品SKU导入列表"]
    from app.services import campaign_policy_service

    default_shipping_days = campaign_policy_service.default_shipping_days()
    shipping_days_by_item = shipping_days_by_item or {}
    if ws.max_row >= 4:                                # 清模板示例数据行, 保留前3行表头
        ws.delete_rows(4, ws.max_row - 3)
    r, seen = 4, set()
    for row in rows:
        sid = row["taobao_sku_id"]
        if sid in seen:
            continue
        seen.add(sid)
        item_id = str(row["taobao_item_id"])
        ws.cell(r, 1, item_id).number_format = "@"
        ws.cell(r, 2, sid).number_format = "@"
        ws.cell(r, 3, float(row["price"])).number_format = "0.00"
        shipping_days = shipping_days_by_item.get(item_id, default_shipping_days)
        if not isinstance(shipping_days, int) or not 1 <= shipping_days <= 365:
            raise ValueError(f"商品 {item_id} 的发货天数非法")
        # The official template accepts relative-time text (for example
        # "30天"), not a bare numeric day count. Every row is required.
        ws.cell(r, 5, f"{shipping_days}天").number_format = "@"
        r += 1
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_super_signup_xlsx(rows: list[dict]) -> bytes:
    """超级立减官方14列表：真实报名价仍为日常价，只填10%让利比例。"""
    import io
    from pathlib import Path
    import openpyxl

    tpl = (Path(__file__).resolve().parent.parent / "assets" / "taobao_templates"
           / "super_reduce_import.xlsx")
    wb = openpyxl.load_workbook(tpl)
    ws = wb["商品SKU导入列表"] if "商品SKU导入列表" in wb.sheetnames else wb.worksheets[-1]
    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)
    r, seen = 4, set()
    for row in rows:
        sid = str(row["taobao_sku_id"])
        if sid in seen:
            continue
        seen.add(sid)
        ws.cell(r, 1, str(row["taobao_item_id"])).number_format = "@"
        ws.cell(r, 2, sid).number_format = "@"
        ws.cell(r, 3, float(row["price"])).number_format = "0.00"
        ws.cell(r, 5, "包邮")
        ws.cell(r, 13, 10)
        r += 1
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _fmt_dt(dt) -> Optional[str]:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else None


def _plan_campaign_ids(plan) -> tuple[Optional[str], Optional[str]]:
    """Read typed immutable IDs first; fall back to legacy remark markers."""
    typed_cid = str(getattr(plan, "platform_campaign_id", None) or "").strip()
    typed_uid = str(getattr(plan, "platform_united_activity_id", None) or "").strip()
    if typed_cid or typed_uid:
        return typed_cid or None, typed_uid or None
    import re
    text = str(getattr(plan, "remark", None) or "")
    cid = re.search(r"(?:campaignId|campaign_id)\s*[:=]\s*(\d+)", text)
    uid = re.search(r"(?:unitedActivityId|united_activity_id)\s*[:=]\s*(\d+)", text)
    return (cid.group(1) if cid else None, uid.group(1) if uid else None)


def _campaign_identity(plan) -> dict:
    """Return the immutable identity required before any campaign I/O."""
    title = str(
        getattr(plan, "qn_campaign_title", None)
        or getattr(plan, "name", None)
        or ""
    ).strip()
    campaign_id, united_activity_id = _plan_campaign_ids(plan)
    start = getattr(plan, "start_at", None)
    end = getattr(plan, "end_at", None)
    activity_mode = str(
        getattr(plan, "platform_activity_mode", None) or "fixed_window")
    sign_record_id = str(
        getattr(plan, "platform_sign_record_id", None) or "").strip()
    active_until = getattr(plan, "platform_active_until", None)
    long_running = (
        activity_mode == "long_running_update"
        and str(getattr(plan, "campaign_type", "")) == "super_reduce"
    )
    missing = []
    if not title:
        missing.append("campaign_title")
    if not long_running:
        if not campaign_id or not campaign_id.isdigit():
            missing.append("campaign_id")
        if not united_activity_id or not united_activity_id.isdigit():
            missing.append("united_activity_id")
        if sign_record_id and not sign_record_id.isdigit():
            missing.append("sign_record_id")
    elif sign_record_id:
        missing.append("sign_record_id_not_allowed_for_long_running")
    elif active_until is None:
        missing.append("platform_active_until")
    if start is None:
        missing.append("campaign_start")
    if end is None:
        missing.append("campaign_end")
    if start is not None and end is not None and end <= start:
        missing.append("campaign_window_order")
    if long_running and end is not None and active_until is not None and active_until < end:
        missing.append("platform_active_until_before_update_window")
    return {
        "ok": not missing,
        "missing": missing,
        "campaign_title": title,
        "campaign_id": campaign_id,
        "united_activity_id": united_activity_id,
        "sign_record_id": sign_record_id or None,
        "campaign_start": _fmt_dt(start),
        "campaign_end": _fmt_dt(end),
        "platform_activity_mode": activity_mode,
        "platform_active_until": _fmt_dt(active_until),
        "campaign_phase": (
            str(getattr(plan, "name", None) or "").strip()
            if str(getattr(plan, "name", None) or "").strip() != title else ""
        ),
        "official_rate": f"{int(TIER_LEVERAGE[plan_tier(plan)] * 100)}%",
    }


def _check_campaign_identity(plan) -> dict:
    identity = _campaign_identity(plan)
    return {
        "rule": "R20",
        "level": "pass" if identity["ok"] else "error",
        "title": (
            "活动唯一身份：长期超级立减锁定标题、更新窗口和官方有效期；"
            "固定大促锁定标题、双ID和秒级档期"
        ),
        "items": [] if identity["ok"] else [{"missing": identity["missing"]}],
        "identity": identity,
    }


def _check_signup_shipping(plan, signup_rows: list[dict]) -> dict:
    resolved = signup_shipping_days(plan, signup_rows)
    item_ids = {
        str(row.get("taobao_item_id") or "").strip()
        for row in signup_rows
        if str(row.get("taobao_item_id") or "").strip()
    }
    missing = sorted(item_ids - set(resolved))
    invalid = sorted(
        item_id for item_id, days in resolved.items()
        if not isinstance(days, int) or not 1 <= days <= 365
    )
    return {
        "rule": "R21",
        "level": "error" if missing or invalid else "pass",
        "title": "报名表逐行相对发货时效：每行必须为有效的N天文本",
        "items": ([{"missing_items": missing, "invalid_items": invalid}]
                  if missing or invalid else []),
        "item_count": len(item_ids),
        "resolved": resolved,
    }


def plan_campaign_ids(plan) -> tuple[Optional[str], Optional[str]]:
    """Public read-only accessor used by feedback/export API routes."""
    return _plan_campaign_ids(plan)


def campaign_identity(plan) -> dict:
    """Public immutable campaign identity for read-only exports and receipts."""
    return _campaign_identity(plan)


def _plan_single_discount_activity_id(plan) -> Optional[str]:
    """Read the verified existing single-item-discount activity for a repair."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*single_discount_activity_id\s*=\s*(\d+)",
        text,
        flags=re.IGNORECASE,
    )
    return matched.group(1) if matched else None


def _plan_single_discount_activity_ids(plan) -> dict[str, str]:
    """Read verified per-item existing single-discount activity bindings.

    Marker format:: ``single_discount_activity_ids=item_id:activity_id,...``.
    A QianNiu activity ID belongs to one item row in the SKU-level editor and
    must never be applied as a store-wide ID.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*single_discount_activity_ids\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    return {
        item_id: activity_id
        for item_id, activity_id in re.findall(
            r"(\d{8,})\s*:\s*(\d+)", matched.group(1))
    }


def _set_plan_single_discount_activity_ids(plan, mapping: dict[str, str]) -> None:
    """Idempotently persist exact item-to-discount-activity bindings."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    pattern = r"(?:^|[;\n；])\s*single_discount_activity_ids\s*=\s*[^;\n；]*"
    text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip(" ;\n；")
    value = ",".join(
        f"{item_id}:{mapping[item_id]}" for item_id in sorted(mapping)
    )
    plan.remark = (
        f"{text}; single_discount_activity_ids={value}"
        if text else f"single_discount_activity_ids={value}"
    )


def _plan_single_discount_sku_activity_ids(plan) -> dict[str, str]:
    """Read verified per-SKU single-discount activity bindings.

    One Taobao item may have physical SKUs spread across multiple historical
    single-item-discount activities.  The older item-level marker cannot
    represent that state, so exact SKU bindings take precedence when present.
    """
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*single_discount_sku_activity_ids\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    return {
        sku_id: activity_id
        for sku_id, activity_id in re.findall(
            r"(\d{8,})\s*:\s*(\d+)", matched.group(1)
        )
    }


def _set_plan_single_discount_sku_activity_ids(
        plan, mapping: dict[str, str]) -> None:
    """Idempotently persist exact SKU-to-discount-activity bindings."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    pattern = (
        r"(?:^|[;\n；])\s*single_discount_sku_activity_ids\s*=\s*[^;\n；]*"
    )
    text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip(" ;\n；")
    value = ",".join(
        f"{sku_id}:{mapping[sku_id]}" for sku_id in sorted(mapping)
    )
    plan.remark = (
        f"{text}; single_discount_sku_activity_ids={value}"
        if text else f"single_discount_sku_activity_ids={value}"
    )


def _plan_single_discount_refreshed_activity_ids(plan) -> dict[str, str]:
    """Read bindings proven after an authorized physical-SKU rotation."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    matched = re.search(
        r"(?:^|[;\n；])\s*single_discount_refreshed_activity_ids\s*=\s*([^;\n；]*)",
        text,
        flags=re.IGNORECASE,
    )
    if not matched:
        return {}
    return {
        item_id: activity_id
        for item_id, activity_id in re.findall(
            r"(\d{8,})\s*:\s*(\d+)", matched.group(1))
    }


def _set_plan_single_discount_refreshed_activity_ids(
        plan, mapping: dict[str, str]) -> None:
    """Persist exact post-rotation item/activity evidence idempotently."""
    import re

    text = str(getattr(plan, "remark", None) or "")
    pattern = (
        r"(?:^|[;\n；])\s*single_discount_refreshed_activity_ids\s*=\s*[^;\n；]*"
    )
    text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip(" ;\n；")
    value = ",".join(
        f"{item_id}:{mapping[item_id]}" for item_id in sorted(mapping)
    )
    plan.remark = (
        f"{text}; single_discount_refreshed_activity_ids={value}"
        if text else f"single_discount_refreshed_activity_ids={value}"
    )


def _single_discount_existing_activity_sku_conflicts(
        result: dict, by_item: dict[str, list[dict]],
        known_activity_ids: set[str], *,
        allow_unrecognized_activity_ids: bool = False) -> dict[str, str]:
    """Return complete SKU bindings proven by the platform failure workbook.

    Every expected SKU must be present exactly once with an "already joined"
    activity ID.  Different SKUs of one item may legitimately point at
    different historical activities.  A newly observed ID is only a candidate:
    the exact existing-activity editor must still prove item, period and SKU
    rows before the binding becomes durable.
    """
    import re

    final_import = result.get("final_import")
    if not isinstance(final_import, dict):
        return {}
    failed_rows = final_import.get("failed_rows")
    if not isinstance(failed_rows, list) or not failed_rows:
        return {}
    if final_import.get("failed") != len(failed_rows):
        return {}

    expected = {
        (str(item_id), str(row.get("taobao_sku_id") or "").strip())
        for item_id, rows in by_item.items()
        for row in rows
    }
    if not expected or any(not sku_id for _, sku_id in expected):
        return {}
    observed: set[tuple[str, str]] = set()
    bindings: dict[str, str] = {}
    pattern = re.compile(r"已经参加了单品立减活动\s*[，,]?\s*id\s*[：:]\s*(\d+)")
    for row in failed_rows:
        item_id = str((row or {}).get("item_id") or "").strip()
        sku_id = str((row or {}).get("sku_id") or "").strip()
        matched = pattern.search(str((row or {}).get("reason") or ""))
        key = (item_id, sku_id)
        if key not in expected or key in observed or not matched:
            return {}
        activity_id = matched.group(1)
        if (activity_id not in known_activity_ids
                and not allow_unrecognized_activity_ids):
            return {}
        observed.add(key)
        bindings[sku_id] = activity_id
    if observed != expected:
        return {}
    return bindings


def _single_discount_existing_activity_conflicts(
        result: dict, by_item: dict[str, list[dict]],
        known_activity_ids: set[str], *,
        allow_unrecognized_activity_ids: bool = False) -> dict[str, str]:
    """Return complete uniform item bindings from platform failure rows.

    This compatibility view can also identify one fully failed item inside a
    partial import.  Split-activity items are handled by the SKU-level helper.
    """
    import re

    final_import = result.get("final_import")
    if not isinstance(final_import, dict):
        return {}
    failed_rows = final_import.get("failed_rows")
    if not isinstance(failed_rows, list) or not failed_rows:
        return {}
    if final_import.get("failed") != len(failed_rows):
        return {}

    grouped: dict[str, list[dict]] = defaultdict(list)
    bindings: dict[str, str] = {}
    pattern = re.compile(r"已经参加了单品立减活动\s*[，,]?\s*id\s*[：:]\s*(\d+)")
    for row in failed_rows:
        item_id = str((row or {}).get("item_id") or "").strip()
        sku_id = str((row or {}).get("sku_id") or "").strip()
        matched = pattern.search(str((row or {}).get("reason") or ""))
        if item_id not in by_item or not sku_id or not matched:
            return {}
        activity_id = matched.group(1)
        if (activity_id not in known_activity_ids
                and not allow_unrecognized_activity_ids):
            return {}
        if item_id in bindings and bindings[item_id] != activity_id:
            return {}
        bindings[item_id] = activity_id
        grouped[item_id].append(row)

    for item_id, failed_item_rows in grouped.items():
        expected_skus = {
            str(row.get("taobao_sku_id") or "").strip()
            for row in by_item[item_id]
        }
        failed_skus = {
            str((row or {}).get("sku_id") or "").strip()
            for row in failed_item_rows
        }
        if not expected_skus or failed_skus != expected_skus:
            return {}
    return bindings


def _upload_and_wait(db: Session, channel: str, phase: str, xlsx: bytes,
                     start_dt: Optional[str], end_dt: Optional[str], *,
                     plan=None, expected_rows: Optional[int] = None,
                     expected_items: Optional[int] = None,
                     discount_activity_id: Optional[str] = None,
                     ignore_plan_discount_activity: bool = False) -> dict:
    """WA 上传编排 (与 activity_upload_service 同模式: upload_file → wait_job)。"""
    from app.services import web_agent_service
    extra = {}
    if channel == "promo_signup" and plan is not None:
        identity = _campaign_identity(plan)
        if not identity["ok"]:
            return {
                "ok": False,
                "step": "campaign_identity_guard",
                "error": "campaign_identity_incomplete",
                "missing": identity["missing"],
            }
        extra = {
            "campaign_title": identity["campaign_title"],
            "campaign_phase": identity["campaign_phase"],
            "campaign_start": identity["campaign_start"],
            "campaign_end": identity["campaign_end"],
            "official_rate": identity["official_rate"],
            "campaign_id": identity["campaign_id"],
            "united_activity_id": identity["united_activity_id"],
        }
    elif channel == "single_item_discount" and plan is not None:
        existing_id = discount_activity_id
        if existing_id is None and not ignore_plan_discount_activity:
            existing_id = _plan_single_discount_activity_id(plan)
        if existing_id:
            extra = {"campaign_id": existing_id}
    j = web_agent_service.upload_file(
        db, channel, phase, xlsx, f"campaign_{channel}.xlsx",
        start_dt=start_dt, end_dt=end_dt, expected_rows=expected_rows, **extra)
    if not j.get("ok") or not j.get("job"):
        return {"ok": False, "error": j.get("error", "取数服务(:8500)未响应, 无法上传")}
    final = web_agent_service.wait_job(db, j["job"], timeout_s=200)
    res = final.get("result") or {}
    if res.get("need_scan"):
        return {"ok": False, "need_scan": True, "message": "淘宝登录态过期, 请先扫码后再上传"}
    validation = res.get("validation")
    if channel in ("promo_signup", "super_reduce"):
        total = validation.get("total_items") if isinstance(validation, dict) else None
        ok_count = validation.get("ok") if isinstance(validation, dict) else None
        failed = validation.get("failed") if isinstance(validation, dict) else None
        terminal = bool(
            isinstance(total, int) and isinstance(ok_count, int) and isinstance(failed, int)
            and total > 0 and ok_count + failed == total
        )
        submitted = bool(res.get("submitted"))
        success = bool(
            res.get("ok") and submitted and terminal and failed == 0 and ok_count == total
            and (expected_items is None or total == expected_items)
        )
        error = res.get("error") or res.get("message")
        if not terminal:
            error = error or "批量操作记录未进入终态（不能把附件已挂上当成报名成功）"
        elif expected_items is not None and total != expected_items:
            error = f"批量操作范围不符：平台{total}品，预期{expected_items}品"
        elif failed:
            error = f"批量操作终态失败：成功{ok_count}品/失败{failed}品"
        elif not res.get("submitted"):
            error = "平台未返回已提交写入回执，不能视为报名完成"
    else:
        error = res.get("error") or res.get("message")
        if channel == "single_item_discount" and phase == "commit":
            final_import = res.get("final_import")
            complete = bool(
                isinstance(final_import, dict)
                and final_import.get("state") == "complete"
                and int(final_import.get("failed") or 0) == 0
                and (expected_rows is None
                     or int(final_import.get("ok") or -1) == expected_rows)
            )
            success = bool(res.get("ok") and res.get("submitted") and complete)
            if not isinstance(final_import, dict):
                error = error or "单品立减平台终态缺失，禁止按成功处理"
            elif not complete:
                error = error or (
                    "单品立减终态未完整成功："
                    f"成功{final_import.get('ok')}行/失败{final_import.get('failed')}行"
                )
        else:
            success = bool(res.get("submitted") if phase == "commit" else res.get("ok"))
    out = {"ok": success, "error": error,
           "job": j["job"], "validation": validation,
           "submitted": res.get("submitted"),
           "attached": bool(res.get("attached")),
           "published": bool(res.get("published")),
           "platform_write_observed": bool(
               res.get("platform_write_observed") or res.get("attached")
               or res.get("submitted")),
           "operation_semantics": res.get("operation_semantics"),
           "mode": res.get("mode"),
           "stopped_before": res.get("stopped_before"),
           "draft_import_success_count": res.get("draft_import_success_count"),
           "draft_import_failed_count": res.get("draft_import_failed_count"),
           "final_import": res.get("final_import"),
           "final_step_error": res.get("final_step_error"),
           "screenshot_base64": res.get("screenshot_base64")}
    if channel == "single_item_discount" and phase == "commit" and plan is not None:
        try:
            from app.services import campaign_discount_audit_service
            out["evidence_request_id"] = (
                campaign_discount_audit_service.persist_single_discount_terminal(
                    plan_id=plan.id,
                    workflow_key=str(plan.workflow_key or ""),
                    job_id=j.get("job"),
                    target_xlsx=xlsx,
                    result=out,
                )
            )
            if not out["evidence_request_id"]:
                raise RuntimeError("terminal_evidence_not_persisted")
            failed_artifact = (out.get("final_import") or {}).get(
                "failed_artifact")
            if isinstance(failed_artifact, dict):
                failed_artifact.pop("xlsx_b64", None)
                failed_artifact.pop("content_b64", None)
        except Exception as exc:  # noqa: BLE001 - platform may already be written
            out["ok"] = False
            out["error"] = (
                "平台终态已返回，但完整回执保存失败；禁止自动重试: "
                f"{type(exc).__name__}: {exc}"
            )
            out["evidence_persist_error"] = f"{type(exc).__name__}: {exc}"
    return out


def _learn_from_validation(db: Session, plan, validation) -> dict:
    """Record platform facts for diagnosis; never adjust price, scope or retry.

    Delisted/no-sales facts keep their existing registries.  Exact platform
    floor numbers are stored as evidence for a later *user-approved* program
    run.  The current failed plan is still stopped and marked ``alarmed``.
    """
    if not validation:
        return {"recorded": False}
    try:
        from app.services import (
            campaign_price_floor_service,
            delisted_sku_service,
            no_sales_service,
        )
        failed = validation.get("failed_items") if isinstance(validation, dict) else None
        ids = delisted_sku_service.extract_delisted_from_feedback(failed)
        if ids:
            delisted_sku_service.add_delisted(db, ids)
        items = no_sales_service.extract_no_sales_only_from_feedback(failed)
        if items:
            no_sales_service.add_no_sales(db, items)
        floors = campaign_price_floor_service.record_failed_feedback(
            db, failed, source="campaign_signup_failed_feedback", plan=plan)
        return {
            "recorded": True,
            "delisted_sku_ids": sorted(ids),
            "no_sales_item_ids": sorted(items),
            "price_floor_evidence": floors,
        }
    except Exception:  # noqa: BLE001 — 自愈失败不影响主流程
        return {"recorded": False, "error": "failed_feedback_fact_recording_failed"}


def qualify_signup_scope(db: Session, plan) -> dict:
    """Blocked public compatibility entry.

    QianNiu creates a real batch signup operation as soon as the workbook is
    attached. There is no read-only upload qualification phase.
    """
    return {
        "ok": False,
        "step": "unsafe_write_probe_disabled",
        "error": "上传报名表本身就是平台写入；资格检查已并入唯一一次正式报名",
        "platform_write_performed": False,
        "automatic_retry": False,
    }


def _legacy_write_probe_qualify_signup_scope(db: Session, plan) -> dict:
    """Historical write-probe implementation retained only for regression tests.

    Historical no-sales registrations do not narrow the probe. Only failures
    explicitly classified by platform feedback as no-sales are a normal
    fallback. Price, SKU, listing-state, and unknown failures remain hard stops.
    """
    from app.services import no_sales_service

    current = refresh_floor_evidence_from_current_activity(db, plan)
    if not current.get("ok"):
        return {"ok": False, "step": "qualification_current_export",
                "error": current.get("error")}
    rows, stats = build_signup_rows(db, plan)
    supplement_scope = authorized_supplement_items(plan)
    rows, missing_scope = _apply_authorized_supplement_scope(plan, rows, stats)
    if missing_scope:
        return {
            "ok": False,
            "step": "qualification_supplement_scope_guard",
            "error": "补报授权商品未全部出现在安全报名行中，已停止平台资格检查",
            "missing_items": missing_scope,
            "stats": stats,
        }
    prior_qualified = platform_qualified_items(plan)
    prior_no_sales = platform_no_sales_items(plan)
    prior_hard_failed = platform_hard_failed_items(plan)
    qualification_rows = current["rows"]
    if str(getattr(plan, "campaign_type", "")) == "super_reduce":
        from app.services.campaign_recon_service import ACTIVITY_IN_CAMPAIGN_STATUSES
        # Paused rows are eligible for a fresh probe/re-import.  Treating them
        # as already published made qualification disagree with push_signup.
        qualification_rows = [
            row for row in qualification_rows
            if row.get("status") in (None, "", *ACTIVITY_IN_CAMPAIGN_STATUSES)
        ]
    live_by_sku = {str(row["sku_id"]): row for row in qualification_rows}
    sku_refresh_scope = authorized_sku_refresh_items(plan)
    by_item: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_item[str(row["taobao_item_id"])].append(row)
    already_correct: set[str] = set()
    wrong_existing: list[dict] = []
    for item_id, item_rows in by_item.items():
        seen = [live_by_sku.get(str(row["taobao_sku_id"])) for row in item_rows]
        if all(current_row is not None for current_row in seen):
            mismatches = [
                str(row["taobao_sku_id"])
                for row, current_row in zip(item_rows, seen)
                if not row.get("is_placeholder")
                and (current_row.get("activity_price") is None
                or abs(float(current_row["activity_price"]) - float(row["price"])) > 0.005
                )
            ]
            if mismatches and item_id not in sku_refresh_scope:
                wrong_existing.append({"item_id": item_id, "mismatched_skus": mismatches})
            elif mismatches:
                # The user-confirmed physical SKU rotation can leave those new
                # skuIds attached to stale activity prices before the complete
                # item is re-imported.  Keep the item in the platform probe;
                # push_signup still performs the exact full-item refresh and
                # all normal ERP price/pairing gates remain in force.
                pass
            else:
                already_correct.add(item_id)
        elif (any(current_row is not None for current_row in seen)
              and item_id not in sku_refresh_scope):
            wrong_existing.append({"item_id": item_id, "error": "partial_existing_item"})
    wrong_existing_ids = {str(row["item_id"]) for row in wrong_existing}

    probe_rows = [
        row for row in rows
        if str(row["taobao_item_id"]) not in already_correct
        and str(row["taobao_item_id"]) not in wrong_existing_ids
    ]
    candidate_items = {str(row["taobao_item_id"]) for row in probe_rows}
    if not rows:
        return {"ok": False, "step": "qualification_empty", "stats": stats,
                "error": "no_safe_listed_items_for_platform_qualification"}

    if not probe_rows:
        qualified = set(already_correct)
        if supplement_scope:
            qualified |= prior_qualified
        no_sales_service.remove_no_sales(db, qualified)
        _set_plan_item_marker(plan, "platform_qualified_items", qualified)
        _set_plan_item_marker(
            plan, "platform_no_sales_items",
            (prior_no_sales - supplement_scope) if supplement_scope else set())
        _set_plan_item_marker(
            plan, "platform_hard_failed_items",
            (prior_hard_failed - supplement_scope) if supplement_scope else set())
        _set_plan_item_marker(plan, "platform_existing_wrong_items", wrong_existing_ids)
        _sync_official_scope_after_platform_result(plan, qualified)
        _record_terminal_platform_acceptance(plan, rows, already_correct)
        _record_terminal_coupon_floor_qualification(plan, rows, set())
        if not supplement_scope:
            _remove_plan_marker(plan, "supplement_items_authorized")
        db.commit()
        return {"ok": True, "no_change": True,
                "qualified_item_ids": sorted(qualified),
                "no_sales_item_ids": [],
                "wrong_existing_items": wrong_existing, "stats": stats}

    channel = "super_reduce" if str(plan.campaign_type) == "super_reduce" else "promo_signup"
    upload = (_build_super_signup_xlsx(probe_rows) if channel == "super_reduce"
              else _build_signup_xlsx(
                  probe_rows, signup_shipping_days(plan, probe_rows)))
    probe = _upload_and_wait(
        db, channel, "stage", upload,
        _fmt_dt(plan.start_at), _fmt_dt(plan.end_at), plan=plan,
        expected_rows=len(probe_rows), expected_items=len(candidate_items),
    )
    validation = probe.get("validation") if isinstance(probe, dict) else None
    if not isinstance(validation, dict):
        return {"ok": False, "step": "platform_qualification",
                "error": probe.get("error") if isinstance(probe, dict) else None,
                "validation": validation, "stats": stats}
    total, ok_count, failed_count = (
        validation.get("total_items"), validation.get("ok"), validation.get("failed"))
    terminal = (
        all(isinstance(value, int) for value in (total, ok_count, failed_count))
        and total > 0 and ok_count + failed_count == total
        and total == len(candidate_items)
    )
    if not terminal:
        return {"ok": False, "step": "platform_qualification_terminal",
                "error": "platform_qualification_not_terminal_or_scope_mismatch",
                "validation": validation, "stats": stats}

    failed_rows = validation.get("failed_items") or []
    nested_feedback = validation.get("failed_reasons")
    if not failed_rows and isinstance(nested_feedback, dict):
        failed_rows = nested_feedback.get("failed") or []
        validation["failed_items"] = failed_rows
        validation["failed_reasons"] = nested_feedback.get("by_reason") or []
    feedback_refresh = None
    if failed_count and not failed_rows:
        # 平台导入页偶发拿不到瞬时反馈下载；按不可变活动 ID 从最近一次
        # 批量操作记录只读补取。没有逐商品原因时绝不根据数量猜测分类。
        from app.services import web_agent_service
        if channel == "super_reduce":
            feedback_refresh = web_agent_service.super_reduce_feedback(db)
        else:
            campaign_id, united_activity_id = _plan_campaign_ids(plan)
            feedback_refresh = web_agent_service.campaign_feedback(
                db,
                str(getattr(plan, "qn_campaign_title", None) or plan.name or ""),
                campaign_id=campaign_id or "",
                united_activity_id=united_activity_id or "",
            )
        if feedback_refresh.get("ok"):
            feedback = feedback_refresh.get("feedback") or {}
            failed_rows = feedback.get("failed") or []
            validation["failed_items"] = failed_rows
            validation["failed_reasons"] = feedback.get("by_reason") or []
    failed_ids = {
        str((row or {}).get("item_id") or "").strip() for row in failed_rows
    } - {""}
    no_sales_ids = no_sales_service.extract_no_sales_only_from_feedback(failed_rows)
    if failed_count and len(failed_ids) != failed_count:
        return {"ok": False, "step": "platform_qualification_non_sales_failure",
                "error": "qualification_failed_item_details_incomplete",
                "failed_item_ids": sorted(failed_ids),
                "no_sales_item_ids": sorted(no_sales_ids),
                "validation": validation, "feedback_refresh": feedback_refresh,
                "stats": stats}

    # The platform's staged qualification evaluates the campaign signup price
    # before this plan's single-item discount has been committed.  Therefore a
    # coupon-floor-only rejection is not a real hard failure when our fresh
    # per-SKU evidence proves that the planned discount (including an explicitly
    # authorized sub-yuan concession) lands on or below the same history line.
    # Do not relax no-sales, missing-SKU, list-price, mixed, or still-held items.
    held_ids = {
        str(item.get("taobao_item_id") or "")
        for item in price_hold_items(db, plan)
    } - {""}
    failed_by_item: dict[str, list[dict]] = defaultdict(list)
    for failed_row in failed_rows:
        item_id = str((failed_row or {}).get("item_id") or "").strip()
        if item_id:
            failed_by_item[item_id].append(failed_row or {})

    def _coupon_floor_only(rows: list[dict]) -> bool:
        if not rows:
            return False
        forbidden = (
            "动销", "销量", "sku已失效", "sku缺失", "缺skuid", "sku不全",
            "一口价", "最低标价", "活动价格需小于等于", "下架", "不存在",
        )
        for row in rows:
            reason = str(row.get("reason") or "").strip().lower()
            raw = str(row.get("raw") or "").strip().lower()
            if not ("券后价超线" in reason or "最低普惠券后价" in reason):
                return False
            if any(token in reason or token in raw for token in forbidden):
                return False
        return True

    planned_discount_qualified_ids = {
        item_id for item_id, rows_for_item in failed_by_item.items()
        if item_id not in held_ids
        and item_id not in no_sales_ids
        and _coupon_floor_only(rows_for_item)
    }
    hard_failed_ids = failed_ids - no_sales_ids - planned_discount_qualified_ids
    terminal_accepted_ids = already_correct | (candidate_items - failed_ids)
    qualified = (
        already_correct
        | (candidate_items - failed_ids)
        | planned_discount_qualified_ids
    )
    marker_qualified = qualified | prior_qualified if supplement_scope else qualified
    marker_no_sales = (
        (prior_no_sales - supplement_scope) | no_sales_ids
        if supplement_scope else no_sales_ids
    )
    marker_hard_failed = (
        (prior_hard_failed - supplement_scope) | hard_failed_ids
        if supplement_scope else hard_failed_ids
    )
    no_sales_service.add_no_sales(db, no_sales_ids)
    no_sales_service.remove_no_sales(db, qualified)
    _set_plan_item_marker(plan, "platform_qualified_items", marker_qualified)
    _set_plan_item_marker(plan, "platform_no_sales_items", marker_no_sales)
    _set_plan_item_marker(plan, "platform_hard_failed_items", marker_hard_failed)
    _set_plan_item_marker(plan, "platform_existing_wrong_items", wrong_existing_ids)
    _sync_official_scope_after_platform_result(plan, marker_qualified)
    # Coupon-floor-only items are provisional until final signup succeeds.
    _record_terminal_platform_acceptance(plan, rows, terminal_accepted_ids)
    _record_terminal_coupon_floor_qualification(
        plan, rows, planned_discount_qualified_ids)
    # A one-off corrective retry marker must not survive a new full-scope probe.
    if not supplement_scope:
        _remove_plan_marker(plan, "supplement_items_authorized")
    db.commit()
    return {"ok": True, "qualified_item_ids": sorted(qualified),
            "no_sales_item_ids": sorted(no_sales_ids),
            "hard_failed_item_ids": sorted(hard_failed_ids),
            "planned_discount_qualification_item_ids": sorted(
                planned_discount_qualified_ids),
            "terminal_accepted_item_ids": sorted(terminal_accepted_ids),
            "hard_failed_items": [row for row in failed_rows
                                  if str((row or {}).get("item_id") or "") in hard_failed_ids],
            "validation": validation, "feedback_refresh": feedback_refresh,
            "already_correct_item_ids": sorted(already_correct),
            "wrong_existing_items": wrong_existing, "stats": stats}


def push_discount(
        db: Session, plan, phase: str = "stage", *,
        item_scope: Optional[set[str]] = None,
        no_sales_items: Optional[set[str]] = None,
        terminal_no_sales_fallback: bool = False,
        update_plan_status: bool = True,
        _existing_conflict_retry: bool = False,
) -> dict:
    """推单品立减 (channel single_item_discount, 带计划档期精确到秒)。
    phase='stage' 挂文件停在提交前; 'commit' ★不可逆★ 真提交 (仅用户确认后调, R12)。"""
    scope_check = _check_official_scope(
        db, plan, no_sales_items=no_sales_items)
    if scope_check["level"] == "error" and not terminal_no_sales_fallback:
        return {
            "ok": False,
            "error": "官方立减逐商品范围未通过安全门，已停止上传",
            "check": scope_check,
        }
    rows, stats = build_discount_rows(
        db, plan, no_sales_items=no_sales_items)
    rows, missing_scope = _apply_authorized_supplement_scope(plan, rows, stats)
    if missing_scope:
        return {
            "ok": False,
            "step": "supplement_scope_guard",
            "error": "补报授权商品未全部出现在安全立减行中，已停止上传",
            "missing_items": missing_scope,
            "stats": stats,
        }
    qualified_scope = platform_qualified_items(plan)
    no_sales_scope = platform_no_sales_items(plan)
    supplement_scope = authorized_supplement_items(plan)
    if platform_scope_present(plan):
        # A prior platform probe is historical evidence.  An exact supplement
        # marker is current-turn authorization to retry those items through all
        # normal price, SKU-completeness, and final-feedback gates.  Do not
        # silently erase their already-safe rows merely because the previous
        # probe classified them as hard failures.
        allowed = qualified_scope | no_sales_scope | supplement_scope
        rows = [
            row for row in rows
            if str(row.get("taobao_item_id") or "") in allowed
        ]
        stats["platform_discount_scope_items"] = sorted(allowed)
        stats["platform_discount_scope_rows"] = len(rows)
    if item_scope is not None:
        exact_scope = {str(value or "").strip() for value in item_scope}
        rows = [
            row for row in rows
            if str(row.get("taobao_item_id") or "") in exact_scope
        ]
        stats["exact_item_scope"] = sorted(exact_scope)
        stats["exact_item_scope_rows"] = len(rows)
    stats["terminal_no_sales_fallback"] = bool(terminal_no_sales_fallback)
    if not rows:
        return {"ok": False, "error": "无可推送的立减行", "stats": stats}
    by_item: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_item[str(row["taobao_item_id"])].append(row)
    expected_item_ids = set(by_item)
    sku_activity_ids = _plan_single_discount_sku_activity_ids(plan)
    sku_bound_items: set[str] = set()
    item_results: list[dict] = []
    # Exact SKU bindings override the older item-level marker.  This is needed
    # when one item was historically split across two single-discount
    # activities.  Only a complete per-item SKU binding is used; partial
    # evidence falls through to the normal guarded discovery path.
    for item_id in sorted(by_item):
        item_rows = by_item[item_id]
        if not item_rows or not all(
            str(row.get("taobao_sku_id") or "").strip() in sku_activity_ids
            for row in item_rows
        ):
            continue
        rows_by_activity: dict[str, list[dict]] = defaultdict(list)
        for row in item_rows:
            sku_id = str(row.get("taobao_sku_id") or "").strip()
            rows_by_activity[sku_activity_ids[sku_id]].append(row)
        for activity_id in sorted(rows_by_activity):
            activity_rows = rows_by_activity[activity_id]
            item_result = _upload_and_wait(
                db, "single_item_discount", phase,
                _build_discount_xlsx(activity_rows),
                _fmt_dt(plan.start_at), _fmt_dt(plan.end_at),
                plan=plan, expected_rows=len(activity_rows),
                discount_activity_id=activity_id,
            )
            item_results.append({
                "item_id": item_id,
                "activity_id": activity_id,
                "mode": "sku_activity_binding",
                **item_result,
            })
            if not item_result.get("ok"):
                stats["single_discount_execution_mode"] = "per_sku_activity"
                stats["single_discount_expected_items"] = len(expected_item_ids)
                stats["single_discount_sku_bound_items"] = sorted(sku_bound_items)
                stats["single_discount_processed_items"] = len(sku_bound_items)
                return {
                    "ok": False,
                    "step": "single_item_discount_per_sku_activity",
                    "error": item_result.get("error") or "逐SKU活动修改失败",
                    "failed_item_id": item_id,
                    "failed_activity_id": activity_id,
                    "completed_item_ids": sorted(sku_bound_items),
                    "item_results": item_results,
                    "stats": stats,
                }
        sku_bound_items.add(item_id)
    if sku_bound_items:
        by_item = defaultdict(list, {
            item_id: item_rows
            for item_id, item_rows in by_item.items()
            if item_id not in sku_bound_items
        })
        rows = [row for item_rows in by_item.values() for row in item_rows]
    if not by_item:
        res = {
            "ok": True,
            "submitted": phase == "commit",
            "processed_items": len(expected_item_ids),
            "item_results": item_results,
        }
        stats["single_discount_execution_mode"] = "per_sku_activity"
        stats["single_discount_expected_items"] = len(expected_item_ids)
        stats["single_discount_sku_bound_items"] = sorted(sku_bound_items)
        stats["single_discount_existing_items"] = sorted(sku_bound_items)
        stats["single_discount_new_items"] = []
        stats["single_discount_processed_items"] = len(expected_item_ids)
        res["stats"] = stats
        if phase == "commit":
            if update_plan_status:
                plan.status = "discount_pushed"
            db.commit()
        return res
    activity_ids = _plan_single_discount_activity_ids(plan)
    legacy_id = _plan_single_discount_activity_id(plan)
    if not activity_ids and legacy_id:
        if len(by_item) != 1:
            return {
                "ok": False,
                "step": "single_discount_activity_binding_guard",
                "error": "既有单品立减活动ID缺少商品绑定，拒绝把一个活动ID应用到多个商品",
                "legacy_activity_id": legacy_id,
                "target_item_ids": sorted(by_item),
                "stats": stats,
            }
        activity_ids = {next(iter(by_item)): legacy_id}
    if activity_ids:
        # 千牛“修改优惠”是逐商品抽屉：同一活动可包含多商品，但每次只能
        # 打开一个商品并逐 SKU 回读。绝不能把全店行一次交给该入口，否则
        # Web-Agent 只能修改/证明一个商品。逐商品执行也让中断后的重跑保持幂等。
        # A rotated physical skuId cannot be found in the old activity's edit
        # drawer even though the item itself has an activity binding.  For the
        # exact user-authorized refresh scope, keep the old activity untouched
        # and submit the complete new physical-SKU set as a new batch.  Platform
        # rejection remains terminal; this never withdraws or rewrites old rows.
        refreshed_activity_ids = _plan_single_discount_refreshed_activity_ids(plan)
        rotated_new_batch_items = {
            item_id
            for item_id in set(by_item) & authorized_sku_refresh_items(plan)
            if refreshed_activity_ids.get(item_id) != activity_ids.get(item_id)
        }
        editable_activity_ids = {
            item_id: activity_id
            for item_id, activity_id in activity_ids.items()
            if item_id not in rotated_new_batch_items
        }
        existing_items = sorted(set(by_item) & set(editable_activity_ids))
        for item_id in existing_items:
            item_rows = by_item[item_id]
            item_result = _upload_and_wait(
                db, "single_item_discount", phase, _build_discount_xlsx(item_rows),
                _fmt_dt(plan.start_at), _fmt_dt(plan.end_at),
                plan=plan, expected_rows=len(item_rows),
                discount_activity_id=editable_activity_ids[item_id],
            )
            item_results.append({"item_id": item_id, **item_result})
            if not item_result.get("ok"):
                res = {
                    "ok": False,
                    "step": "single_item_discount_per_item",
                    "error": item_result.get("error") or "逐商品单品立减失败",
                    "failed_item_id": item_id,
                    "completed_item_ids": [
                        row["item_id"] for row in item_results[:-1] if row.get("ok")
                    ],
                    "item_results": item_results,
                }
                break
        else:
            new_item_ids = sorted(set(by_item) - set(editable_activity_ids))
            new_rows = [row for item_id in new_item_ids for row in by_item[item_id]]
            if new_rows:
                new_result = _upload_and_wait(
                    db, "single_item_discount", phase, _build_discount_xlsx(new_rows),
                    _fmt_dt(plan.start_at), _fmt_dt(plan.end_at),
                    plan=plan, expected_rows=len(new_rows),
                    ignore_plan_discount_activity=True,
                )
                item_results.append({"item_ids": new_item_ids, "mode": "new_batch", **new_result})
                if not new_result.get("ok"):
                    discovered_skus = (
                        _single_discount_existing_activity_sku_conflicts(
                            new_result, by_item, set(activity_ids.values()),
                            allow_unrecognized_activity_ids=True)
                        if phase == "commit" and not _existing_conflict_retry
                        else {}
                    )
                    discovered = (
                        _single_discount_existing_activity_conflicts(
                            new_result, by_item, set(activity_ids.values()),
                            allow_unrecognized_activity_ids=True)
                        if phase == "commit" and not _existing_conflict_retry
                        else {}
                    )
                    if discovered_skus or discovered:
                        final_import = new_result.get("final_import") or {}
                        if int(final_import.get("ok") or 0) > 0:
                            # Successful rows in a partial import are already
                            # live, but their activity ID is absent from the
                            # failure workbook.  Never resend them blindly.
                            res = {
                                "ok": False,
                                "step": "single_item_discount_partial_import",
                                "error": "单品立减部分成功已生效，必须先回读活动后再继续",
                                "committed_rows": int(final_import.get("ok") or 0),
                                "failed_rows": int(final_import.get("failed") or 0),
                                "discovered_existing_activity_ids": discovered,
                                "discovered_existing_sku_activity_ids": discovered_skus,
                                "completed_item_ids": existing_items,
                                "failed_item_ids": new_item_ids,
                                "item_results": item_results,
                                "stats": stats,
                            }
                            return res
                        uniform_items: dict[str, str] = {}
                        for item_id, item_rows in by_item.items():
                            item_activity_ids = {
                                discovered_skus.get(str(
                                    row.get("taobao_sku_id") or "").strip())
                                for row in item_rows
                            }
                            if None not in item_activity_ids and len(item_activity_ids) == 1:
                                uniform_items[item_id] = next(iter(item_activity_ids))
                        merged_activity_ids = {
                            **activity_ids, **discovered, **uniform_items,
                        }
                        refreshed_ids = {
                            **_plan_single_discount_refreshed_activity_ids(plan),
                            **discovered, **uniform_items,
                        }
                        merged_sku_activity_ids = {
                            **_plan_single_discount_sku_activity_ids(plan),
                            **discovered_skus,
                        }
                        _set_plan_single_discount_activity_ids(
                            plan, merged_activity_ids)
                        _set_plan_single_discount_refreshed_activity_ids(
                            plan, refreshed_ids)
                        _set_plan_single_discount_sku_activity_ids(
                            plan, merged_sku_activity_ids)
                        retry_result = push_discount(
                            db, plan, phase=phase,
                            item_scope=item_scope,
                            no_sales_items=no_sales_items,
                            terminal_no_sales_fallback=terminal_no_sales_fallback,
                            update_plan_status=update_plan_status,
                            _existing_conflict_retry=True)
                        if not retry_result.get("ok"):
                            # Candidate bindings become durable only after the
                            # editor proves exact item, exact period and all SKUs.
                            db.rollback()
                            retry_result["candidate_single_discount_activity_ids"] = discovered
                            retry_result["candidate_single_discount_sku_activity_ids"] = discovered_skus
                        retry_result["reconciled_single_discount_activity_ids"] = discovered
                        retry_result["reconciled_single_discount_sku_activity_ids"] = discovered_skus
                        return retry_result
                    res = {
                        "ok": False,
                        "step": "single_item_discount_new_batch",
                        "error": new_result.get("error") or "新建单品立减批次失败",
                        "completed_item_ids": existing_items,
                        "failed_item_ids": new_item_ids,
                        "item_results": item_results,
                    }
                else:
                    res = {"ok": True, "submitted": phase == "commit",
                           "processed_items": len(by_item), "item_results": item_results}
            else:
                res = {"ok": True, "submitted": phase == "commit",
                       "processed_items": len(by_item), "item_results": item_results}
        stats["single_discount_execution_mode"] = "per_item_existing_then_new_batch"
        stats["single_discount_expected_items"] = len(by_item)
        stats["single_discount_existing_items"] = existing_items
        stats["single_discount_new_items"] = sorted(
            set(by_item) - set(editable_activity_ids))
        stats["single_discount_rotated_new_batch_items"] = sorted(
            rotated_new_batch_items)
        stats["single_discount_processed_items"] = (
            len(by_item) if res.get("ok") else len(existing_items))
    else:
        res = _upload_and_wait(
            db, "single_item_discount", phase, _build_discount_xlsx(rows),
            _fmt_dt(plan.start_at), _fmt_dt(plan.end_at),
            plan=plan, expected_rows=len(rows),
        )
        stats["single_discount_execution_mode"] = "new_activity_batch"
        if (not res.get("ok") and phase == "commit"
                and not _existing_conflict_retry):
            discovered_skus = _single_discount_existing_activity_sku_conflicts(
                res, by_item, set(), allow_unrecognized_activity_ids=True)
            discovered = _single_discount_existing_activity_conflicts(
                res, by_item, set(), allow_unrecognized_activity_ids=True)
            if discovered_skus or discovered:
                final_import = res.get("final_import") or {}
                if int(final_import.get("ok") or 0) > 0:
                    res = {
                        "ok": False,
                        "step": "single_item_discount_partial_import",
                        "error": "单品立减部分成功已生效，必须先回读活动后再继续",
                        "committed_rows": int(final_import.get("ok") or 0),
                        "failed_rows": int(final_import.get("failed") or 0),
                        "discovered_existing_activity_ids": discovered,
                        "discovered_existing_sku_activity_ids": discovered_skus,
                    }
                else:
                    uniform_items: dict[str, str] = {}
                    for item_id, item_rows in by_item.items():
                        item_activity_ids = {
                            discovered_skus.get(str(
                                row.get("taobao_sku_id") or "").strip())
                            for row in item_rows
                        }
                        if None not in item_activity_ids and len(item_activity_ids) == 1:
                            uniform_items[item_id] = next(iter(item_activity_ids))
                    discovered = {**discovered, **uniform_items}
                    _set_plan_single_discount_activity_ids(plan, discovered)
                    _set_plan_single_discount_refreshed_activity_ids(
                        plan, discovered)
                    _set_plan_single_discount_sku_activity_ids(
                        plan, {
                            **_plan_single_discount_sku_activity_ids(plan),
                            **discovered_skus,
                        })
                    retry_result = push_discount(
                        db, plan, phase=phase,
                        item_scope=item_scope,
                        no_sales_items=no_sales_items,
                        terminal_no_sales_fallback=terminal_no_sales_fallback,
                        update_plan_status=update_plan_status,
                        _existing_conflict_retry=True)
                    if not retry_result.get("ok"):
                        db.rollback()
                        retry_result["candidate_single_discount_activity_ids"] = discovered
                        retry_result["candidate_single_discount_sku_activity_ids"] = discovered_skus
                    retry_result["reconciled_single_discount_activity_ids"] = discovered
                    retry_result["reconciled_single_discount_sku_activity_ids"] = discovered_skus
                    return retry_result
    res["stats"] = stats
    if res.get("ok") and phase == "commit":
        if update_plan_status:
            plan.status = "discount_pushed"
        db.commit()
    return res


def _signup_failure_signature(plan, result: dict) -> str:
    import hashlib
    import json
    payload = {
        "plan_id": getattr(plan, "id", None),
        "step": result.get("step"),
        "error": result.get("error"),
        "validation": result.get("validation"),
        "wrong_published_items": result.get("wrong_published_items"),
    }
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def _notify_signup_failure(db: Session, plan, result: dict) -> dict:
    """同一失败内容只发一次运营通知；原因变化后会重新通知。"""
    import json
    from app.services import campaign_notification_service as notify_service, settings_service

    key = f"campaign_signup_failure_{getattr(plan, 'id', 'unknown')}"
    signature = _signup_failure_signature(plan, result)
    if settings_service.get(db, key, env_fallback=False) == signature:
        return {"deduped": True}
    validation = result.get("validation") or {}
    lines = [
        f"活动：{getattr(plan, 'name', '')}",
        f"千牛活动：{getattr(plan, 'qn_campaign_title', None) or getattr(plan, 'name', '')}",
        f"失败步骤：{result.get('step') or '批量导入/终态核对'}",
        f"原因：{result.get('error') or '未知错误'}",
    ]
    if isinstance(validation, dict):
        if any(k in validation for k in ("total_items", "ok", "failed")):
            lines.append(
                f"平台终态：总{validation.get('total_items')}品，"
                f"成功{validation.get('ok')}品，失败{validation.get('failed')}品"
            )
        reasons = validation.get("failed_reasons") or validation.get("failed_items")
        if reasons:
            lines.append("失败明细：" + json.dumps(
                reasons[:8] if isinstance(reasons, list) else reasons,
                ensure_ascii=False, default=str)[:1800])
    wrong = result.get("wrong_published_items")
    if wrong:
        lines.append("已发布错价/缺SKU：" + json.dumps(
            wrong[:8], ensure_ascii=False, default=str)[:1800])
    lines.append("系统已停止本次自动报名；不会盲目全量重推。")
    lines.append("AI 仅可读取并解释错误，不得改价、改范围、轮换 SKU 或重试；等待用户决定。")
    delivered = notify_service.broadcast_text(
        db, "\n".join(lines), title="活动自动报名失败", level="error")
    if any(v is True for v in delivered.values()):
        settings_service.set_value(
            db, key, signature, description="活动自动报名失败运营通知去重签名")
        db.commit()
    return delivered


def _clear_signup_failure_dedupe(db: Session, plan) -> None:
    from app.services import settings_service
    settings_service.set_value(
        db, f"campaign_signup_failure_{getattr(plan, 'id', 'unknown')}", "")
    db.commit()


def _notify_placeholder_price_blocks(db: Session, plan, blocked: list[dict]) -> dict:
    """Warn once when placeholders are held because price protection is unconfirmed."""
    import hashlib
    import json
    from app.services import campaign_notification_service as notify_service, settings_service

    details = []
    for item in blocked:
        placeholders = item.get("placeholders") or []
        details.append({
            "item_id": item.get("taobao_item_id"),
            "product": item.get("product"),
            "placeholders": [{
                "sku_id": row.get("taobao_sku_id"),
                "sku_code": row.get("sku_code"),
                "current_live_price": row.get("current_live_price"),
                "safe_cap": row.get("safe_cap"),
            } for row in placeholders],
        })
    payload = json.dumps(details, ensure_ascii=False, sort_keys=True, default=str)
    signature = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    key = f"campaign_placeholder_price_hold_{getattr(plan, 'id', 'unknown')}"
    if settings_service.get(db, key, env_fallback=False) == signature:
        return {"deduped": True}

    lines = [
        f"活动：{getattr(plan, 'name', '')}",
        f"暂缓：{len(blocked)} 个商品（仅定制占位 SKU）",
        "原因：平台当前占位保护价高于最低普惠券后价反算出的安全报名上限，"
        "且本场尚未明确确认价保到期。",
        "处理：整品未导入；未轮换 SKU、未强制降价，其他安全商品可继续。",
        "明细：" + payload[:2600],
        "如已确认价保到期，请在活动计划中明确确认后再补报；系统届时只用安全上限。",
    ]
    delivered = notify_service.broadcast_text(
        db, "\n".join(lines), title="活动占位SKU因价保暂缓", level="warning")
    if any(value is True for value in delivered.values()):
        settings_service.set_value(
            db,
            key,
            signature,
            description="活动占位SKU价保暂缓飞书通知去重签名",
        )
        db.commit()
    return delivered


def _notify_coupon_floor_blocks(db: Session, plan, blocked: list[dict]) -> dict:
    """报名资格线命中时飞书告警；相同活动、相同明细只发送一次。"""
    import hashlib
    import json
    from app.services import campaign_notification_service as notify_service, settings_service

    coupon_blocked = []
    for item in blocked:
        skus = []
        for sku in item.get("skus") or []:
            reasons = [r for r in sku.get("reasons") or []
                       if r.get("type") == "coupon_floor"]
            if reasons:
                skus.append({"sku_code": sku.get("sku_code"), "reasons": reasons})
        if skus:
            coupon_blocked.append({
                "item_id": item.get("taobao_item_id"),
                "product": item.get("product"),
                "skus": skus,
            })
    if not coupon_blocked:
        return {"skipped": True}

    payload = json.dumps(
        coupon_blocked, ensure_ascii=False, sort_keys=True, default=str)
    signature = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    key = f"campaign_coupon_floor_hold_{getattr(plan, 'id', 'unknown')}"
    if settings_service.get(db, key, env_fallback=False) == signature:
        return {"deduped": True}

    lines = [
        f"活动：{getattr(plan, 'name', '')}",
        f"暂缓：{len(coupon_blocked)} 个商品（任一SKU冲突则整品排除）",
        "报名资格硬门：活动报名价−官方立减−同期单品立减必须≤近15天最低普惠券后价。",
        "仅允许用户逐SKU授权的1元内微调；大额差异不会自动压低。",
        "处理：未强降、未自动轮换；其他安全商品可继续报名。",
        "明细：" + payload[:2600],
        "请等待平台价格线解除后刷新校验，或由人工决定是否建立独立新商品。",
    ]
    delivered = notify_service.broadcast_text(
        db, "\n".join(lines), title="活动报名资格价格线暂缓", level="warning")
    if any(value is True for value in delivered.values()):
        settings_service.set_value(
            db, key, signature, description="活动报名资格价格线暂缓飞书通知去重签名")
        db.commit()
    return delivered


def _record_signup_execution_receipt(db: Session, plan, result: dict) -> dict:
    """Persist a bounded, non-secret receipt for each final signup outcome."""
    from app.services import settings_service

    identity = _campaign_identity(plan)
    validation = result.get("validation")
    terminal_counts = result.get("terminal_counts")
    if terminal_counts is None and isinstance(validation, dict):
        terminal_counts = {
            "total": validation.get("total_items"),
            "ok": validation.get("ok"),
            "failed": validation.get("failed"),
        }
    classification = result.get("terminal_classification") or {}
    post_export = result.get("post_submit_export_evidence") or {}
    verification = result.get("post_submit_verification")
    receipt = {
        "observed_at": datetime.now(timezone.utc).isoformat(),
        "plan_id": getattr(plan, "id", None),
        "status": "success" if result.get("ok") else "failed",
        "step": result.get("step") or "signup_complete",
        "campaign_title": identity.get("campaign_title"),
        "campaign_id": identity.get("campaign_id"),
        "united_activity_id": identity.get("united_activity_id"),
        "campaign_start": identity.get("campaign_start"),
        "campaign_end": identity.get("campaign_end"),
        "job_id": result.get("job"),
        "submitted": bool(result.get("submitted")),
        "terminal_counts": terminal_counts,
        "accepted_item_ids": classification.get("accepted_item_ids") or [],
        "no_sales_item_ids": classification.get("no_sales_item_ids") or [],
        "hard_failed_item_ids": classification.get("hard_failed_item_ids") or [],
        "feedback_artifact": (
            classification.get("feedback_artifact")
            or (validation.get("feedback_artifact") if isinstance(validation, dict) else None)
        ),
        "fresh_export_sha256": post_export.get("sha256"),
        "per_sku_verification": verification,
        "no_sales_fallback_ok": (
            (result.get("no_sales_fallback") or {}).get("ok")
            if isinstance(result.get("no_sales_fallback"), dict) else None
        ),
        "error": result.get("error"),
    }
    key = f"campaign_execution_receipts_{getattr(plan, 'id', 'unknown')}"
    try:
        existing_raw = settings_service.get(db, key, env_fallback=False)
        existing = json.loads(existing_raw) if existing_raw else []
        if not isinstance(existing, list):
            existing = []
    except (TypeError, ValueError, json.JSONDecodeError):
        existing = []
    history = (existing + [receipt])[-20:]
    settings_service.set_value(
        db, key, json.dumps(history, ensure_ascii=False, default=str),
        description="活动报名终态结构化回执（最多20条，不含凭据）")
    db.commit()
    return receipt


def _stop_signup(db: Session, plan, result: dict) -> dict:
    """Put a failed plan in a non-retrying state and notify with explicit boundaries."""
    result.setdefault("ok", False)
    result["requires_user_decision"] = True
    result["automatic_retry"] = False
    result["ai_may_adjust_or_resubmit"] = False
    plan.status = "alarmed"
    db.commit()
    result["execution_receipt"] = _record_signup_execution_receipt(
        db, plan, result)
    result["notification"] = _notify_signup_failure(db, plan, result)
    return result


def _halt_prewrite(db: Session, plan, result: dict) -> dict:
    """Retry only a proven transient read failure before any platform claim."""
    from app.services import campaign_execution_service

    if not campaign_execution_service.failure_is_retryable_prewrite(result):
        return _stop_signup(db, plan, result)
    result.setdefault("ok", False)
    result.update({
        "requires_user_decision": False,
        "automatic_retry": True,
        "retry_boundary": "read_only_before_platform_write_claim",
        "ai_may_adjust_or_resubmit": False,
        "notification": {"skipped": "transient_prewrite_retry_next_scheduler_window"},
    })
    # Keep draft/precheck/discount_pushed unchanged.  The scheduler may safely
    # retry the read-only phase next hour because no write claim exists.
    result["execution_receipt"] = _record_signup_execution_receipt(
        db, plan, result)
    return result


def _refresh_official_product_sku_identity(
        db: Session, pending_rows: list[dict]) -> dict:
    """Read the official current product export and require exact SKU sets.

    This is the last read-only guard before the one-shot write claim.  It
    prevents an ERP mapping that omits a newly added live SKU from creating a
    partial whole-item upload.
    """
    from app.services import campaign_recon_service, web_agent_service

    exported = web_agent_service.export_product_prices(db, timeout_s=420)
    if not exported.get("ok"):
        return {
            "ok": False,
            "step": "official_product_sku_export",
            "error": exported.get("error") or exported.get("message")
                     or "official_product_export_failed",
            "need_scan": bool(exported.get("need_scan")),
        }
    content = exported.get("xlsx_bytes") or b""
    records = campaign_recon_service.parse_product_batch_export(content)
    actual_by_item: dict[str, set[str]] = defaultdict(set)
    for record in records:
        item_id = str(record.get("item_id") or "").strip()
        sku_id = str(record.get("sku_id") or "").strip()
        if item_id.isdigit() and sku_id.isdigit():
            actual_by_item[item_id].add(sku_id)
    expected_by_item: dict[str, set[str]] = defaultdict(set)
    for row in pending_rows:
        item_id = str(row.get("taobao_item_id") or "").strip()
        sku_id = str(row.get("taobao_sku_id") or "").strip()
        if item_id.isdigit() and sku_id.isdigit():
            expected_by_item[item_id].add(sku_id)
    mismatches = []
    for item_id, expected in sorted(expected_by_item.items()):
        actual = actual_by_item.get(item_id, set())
        if actual != expected:
            mismatches.append({
                "item_id": item_id,
                "expected_sku_ids": sorted(expected),
                "official_current_sku_ids": sorted(actual),
                "missing_in_erp_manifest": sorted(actual - expected),
                "missing_on_platform": sorted(expected - actual),
            })
    return {
        "ok": not mismatches,
        "step": "official_product_sku_identity",
        "error": "official_product_sku_scope_mismatch" if mismatches else None,
        "mismatches": mismatches,
        "artifact": {
            "filename": exported.get("filename"),
            "size": len(content),
            "sha256": hashlib.sha256(content).hexdigest(),
        },
        "checked_items": len(expected_by_item),
        "checked_skus": sum(len(value) for value in expected_by_item.values()),
    }


def refresh_floor_evidence_from_current_activity(
        db: Session, plan, *, allow_placeholder_safe_lowering: bool = False,
) -> dict:
    """Refresh enrolled H/I plus exact not-yet-enrolled candidate price ceilings."""
    from app.services import (
        campaign_price_floor_service,
        campaign_recon_service,
        web_agent_service,
    )

    identity = _campaign_identity(plan)
    if not identity["ok"]:
        return {
            "ok": False,
            "error": "campaign_identity_incomplete",
            "missing": identity["missing"],
        }
    evidence_scope_rows, evidence_scope_stats = build_signup_rows(
        db, plan, enforce_price_holds=False,
        allow_placeholder_safe_lowering=allow_placeholder_safe_lowering)
    exported = web_agent_service.campaign_export_items(
        db,
        identity["campaign_title"],
        campaign_id=identity["campaign_id"],
        united_activity_id=identity["united_activity_id"],
        campaign_phase=(
            "" if identity["platform_activity_mode"] == "long_running_update"
            else identity["campaign_phase"]),
        campaign_start=(
            "" if identity["platform_activity_mode"] == "long_running_update"
            else identity["campaign_start"]),
        campaign_end=(
            "" if identity["platform_activity_mode"] == "long_running_update"
            else identity["campaign_end"]),
        official_rate=identity["official_rate"],
        sign_record_id=identity["sign_record_id"],
        platform_activity_mode=identity["platform_activity_mode"],
        platform_active_until=identity["platform_active_until"],
    )
    if not exported.get("ok"):
        return {
            "ok": False,
            "step": exported.get("step") or "current_activity_export",
            "error": exported.get("error") or exported.get("message")
                     or "无法可靠取得当前活动生效集合",
            "job_id": exported.get("job_id"),
            "detail": exported.get("detail"),
        }
    live_rows = campaign_recon_service.parse_activity_items_export(
        exported["xlsx_bytes"],
        include_paused=str(getattr(plan, "campaign_type", "")) == "super_reduce",
    )
    floor_rows = campaign_recon_service.parse_activity_floor_evidence_export(
        exported["xlsx_bytes"])
    refresh = campaign_price_floor_service.record_activity_export(
        db,
        floor_rows,
        source=f"campaign_pre_submit_export:plan={getattr(plan, 'id', '')}",
        plan=plan,
    )
    covered_sku_ids = {
        str(row.get("sku_id") or "").strip() for row in floor_rows
        if str(row.get("sku_id") or "").strip().isdigit()
    }
    missing_scope: dict[str, set[str]] = defaultdict(set)
    for row in evidence_scope_rows:
        item_id = str(row.get("taobao_item_id") or "").strip()
        sku_id = str(row.get("taobao_sku_id") or "").strip()
        if (item_id.isdigit() and sku_id.isdigit()
                and sku_id not in covered_sku_ids):
            missing_scope[item_id].add(sku_id)

    candidate_result = None
    candidate_refresh = None
    candidate_records: list[dict] = []
    if (missing_scope
            and identity["platform_activity_mode"] == "fixed_window"
            and identity.get("sign_record_id")):
        candidate_scope = [
            {"item_id": item_id, "sku_ids": sorted(sku_ids)}
            for item_id, sku_ids in sorted(missing_scope.items())
        ]
        candidate_result = web_agent_service.campaign_candidate_price_evidence(
            db,
            identity["campaign_title"],
            campaign_id=identity["campaign_id"],
            united_activity_id=identity["united_activity_id"],
            sign_record_id=identity["sign_record_id"],
            campaign_phase=identity["campaign_phase"],
            campaign_start=identity["campaign_start"],
            campaign_end=identity["campaign_end"],
            official_rate=identity["official_rate"],
            candidate_scope=candidate_scope,
        )
        if not candidate_result.get("ok"):
            return {
                "ok": False,
                "step": candidate_result.get("step")
                        or "candidate_price_evidence",
                "error": candidate_result.get("error")
                         or "候选商品价格证据读取失败",
                "job_id": candidate_result.get("job_id"),
                "detail": candidate_result.get("detail"),
            }
        candidate_records = candidate_result.get("records") or []
        candidate_refresh = (
            campaign_price_floor_service.record_candidate_price_evidence(
                db,
                candidate_records,
                source=(
                    f"campaign_candidate_selectable:plan={getattr(plan, 'id', '')}:"
                    f"sha256={candidate_result.get('sha256') or 'missing'}"),
                plan=plan,
            )
        )

    # Replace placeholder current-price observations from the union of the
    # enrolled export and the exact candidate response.  Missing candidates are
    # deliberately omitted so R16 remains a hard stop.
    placeholder_price_rows = list(floor_rows)
    placeholder_price_rows.extend({
        "sku_id": row.get("sku_id"),
        "list_price": row.get("current_list_price"),
    } for row in candidate_records)
    placeholder_refresh = record_placeholder_live_prices(
        db, plan, placeholder_price_rows,
        evidence_scope_stats.get("placeholder_candidate_sku_ids") or [])
    return {
        "ok": True,
        "rows": live_rows,
        "floor_refresh": refresh,
        "candidate_floor_refresh": candidate_refresh,
        "candidate_evidence": ({
            "sha256": candidate_result.get("sha256"),
            "job_id": candidate_result.get("job_id"),
            "requested_sku_count": candidate_result.get(
                "requested_sku_count"),
            "observed_sku_count": candidate_result.get("observed_sku_count"),
            "missing_sku_ids": candidate_result.get("missing_sku_ids") or [],
            "candidate_items_scanned": candidate_result.get(
                "candidate_items_scanned"),
            "page_count": candidate_result.get("page_count"),
            "identity": candidate_result.get("identity"),
            "selection_guard": candidate_result.get("selection_guard"),
            "execution_boundary": candidate_result.get("execution_boundary"),
        } if candidate_result else None),
        "placeholder_price_refresh": placeholder_refresh,
        "export_evidence": {
            "filename": exported.get("filename"),
            "size": len(exported["xlsx_bytes"]),
            "sha256": hashlib.sha256(exported["xlsx_bytes"]).hexdigest(),
            "job_id": exported.get("job_id"),
            "identity": identity,
        },
    }


def _verify_signup_rows(expected_rows: list[dict], live_rows: list[dict]) -> dict:
    """Require every real SKU from the submitted scope to have the exact P price.

    Platform operation feedback is only an import receipt.  A product can still
    be labelled ``活动中`` because one legacy/custom SKU survived while newly
    rotated physical SKUs have blank activity prices.  Accept any matching live
    marketing record for an exact item/SKU pair, but never accept item-level
    status as proof for the whole SKU set.
    """
    from app.services.campaign_recon_service import ACTIVITY_IN_CAMPAIGN_STATUSES

    live_by_pair: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in live_rows:
        if row.get("status") not in ACTIVITY_IN_CAMPAIGN_STATUSES:
            continue
        pair = (str(row.get("item_id") or ""), str(row.get("sku_id") or ""))
        live_by_pair[pair].append(row)

    failures: list[dict] = []
    checked = 0
    for expected in expected_rows:
        if expected.get("is_placeholder"):
            continue
        checked += 1
        item_id = str(expected.get("taobao_item_id") or "")
        sku_id = str(expected.get("taobao_sku_id") or "")
        price = float(expected["price"])
        candidates = live_by_pair.get((item_id, sku_id), [])
        if any(
            row.get("activity_price") is not None
            and abs(float(row["activity_price"]) - price) <= 0.005
            for row in candidates
        ):
            continue
        paused_candidates = [
            row for row in live_rows
            if row.get("status") == "暂停"
            and str(row.get("item_id") or "") == item_id
            and str(row.get("sku_id") or "") == sku_id
        ]
        paused_exact = [
            row for row in paused_candidates
            if row.get("activity_price") is not None
            and abs(float(row["activity_price"]) - price) <= 0.005
        ]
        if paused_exact:
            failures.append({
                "item_id": item_id,
                "sku_id": sku_id,
                "expected_activity_price": price,
                "actual_activity_prices": [
                    row.get("activity_price") for row in paused_exact
                ],
                "actual_statuses": ["暂停"],
                "error": "活动价已导入但商品仍为暂停",
            })
            continue
        failures.append({
            "item_id": item_id,
            "sku_id": sku_id,
            "expected_activity_price": price,
            "actual_activity_prices": [
                row.get("activity_price") for row in candidates
            ],
            "error": "活动中记录缺失" if not candidates else "活动价为空或不一致",
        })
    return {
        "ok": not failures,
        "checked_real_skus": checked,
        "failed_real_skus": len(failures),
        "failures": failures,
    }


def _classify_final_signup(
        db: Session, plan, result: dict, pending_rows: list[dict],
        correct_items: set[str]) -> dict:
    """Classify the one real signup's terminal result without retrying it."""
    from app.services import no_sales_service, web_agent_service

    validation = result.get("validation")
    if not isinstance(validation, dict):
        return {"ok": False, "error": "signup_terminal_validation_missing"}
    total = validation.get("total_items")
    ok_count = validation.get("ok")
    failed_count = validation.get("failed")
    expected_items = {
        str(row.get("taobao_item_id") or "").strip()
        for row in pending_rows
        if str(row.get("taobao_item_id") or "").strip()
    }
    if (not all(isinstance(value, int) for value in (total, ok_count, failed_count))
            or total <= 0 or ok_count + failed_count != total
            or total != len(expected_items)):
        return {
            "ok": False,
            "error": "signup_terminal_counts_invalid_or_scope_mismatch",
            "expected_items": sorted(expected_items),
            "terminal_counts": {
                "total": total, "ok": ok_count, "failed": failed_count,
            },
        }

    failed_rows = validation.get("failed_items") or []
    nested = validation.get("failed_reasons")
    if not failed_rows and isinstance(nested, dict):
        failed_rows = nested.get("failed") or []
        validation["failed_items"] = failed_rows
        validation["failed_reasons"] = nested.get("by_reason") or []
    feedback_refresh = None
    if failed_count and not failed_rows:
        campaign_id, united_activity_id = _plan_campaign_ids(plan)
        if (str(getattr(plan, "campaign_type", "")) == "super_reduce"
                and not (campaign_id and united_activity_id)):
            feedback_refresh = web_agent_service.super_reduce_feedback(db)
        else:
            feedback_refresh = web_agent_service.campaign_feedback(
                db,
                str(getattr(plan, "qn_campaign_title", None) or plan.name or ""),
                campaign_id=campaign_id or "",
                united_activity_id=united_activity_id or "",
            )
        if feedback_refresh.get("ok"):
            feedback = feedback_refresh.get("feedback") or {}
            failed_rows = feedback.get("failed") or []
            validation["failed_items"] = failed_rows
            validation["failed_reasons"] = feedback.get("by_reason") or []
            if feedback_refresh.get("filename"):
                validation["feedback_artifact"] = {
                    "filename": feedback_refresh.get("filename"),
                    "size": len(feedback_refresh.get("xlsx_bytes") or b""),
                    "sha256": hashlib.sha256(
                        feedback_refresh.get("xlsx_bytes") or b"").hexdigest(),
                }

    failed_ids = {
        str((row or {}).get("item_id") or "").strip() for row in failed_rows
        if str((row or {}).get("item_id") or "").strip()
    }
    unexpected_failed_ids = failed_ids - expected_items
    if (failed_count != len(failed_ids) or unexpected_failed_ids
            or len(expected_items - failed_ids) != ok_count):
        return {
            "ok": False,
            "error": "signup_failed_item_details_incomplete_or_scope_mismatch",
            "failed_item_ids": sorted(failed_ids),
            "unexpected_failed_item_ids": sorted(unexpected_failed_ids),
            "terminal_counts": {
                "total": total, "ok": ok_count, "failed": failed_count,
            },
            "feedback_refresh": feedback_refresh,
        }
    no_sales_ids = no_sales_service.extract_no_sales_only_from_feedback(failed_rows)
    hard_failed_ids = failed_ids - no_sales_ids
    accepted_items = expected_items - failed_ids
    if not result.get("submitted"):
        return {
            "ok": False,
            "error": "signup_partial_enrollment_import_paused",
            "platform_write_observed": bool(
                result.get("platform_write_observed") or result.get("attached")),
            "enrollment_record_created": True,
            "active": False,
            "published": None,
            "stopped_before": None,
            "terminal_counts": {
                "total": total, "ok": ok_count, "failed": failed_count,
            },
            "accepted_item_ids": sorted(accepted_items),
            "enrolled_paused_item_ids": sorted(accepted_items),
            "draft_imported_item_ids": [],
            "failed_item_ids": sorted(failed_ids),
            "no_sales_item_ids": sorted(no_sales_ids),
            "hard_failed_item_ids": sorted(hard_failed_ids),
            "failed_items": failed_rows,
            "feedback_artifact": validation.get("feedback_artifact"),
            "feedback_refresh": feedback_refresh,
            "state_interpretation_version": 2,
        }
    qualified_items = set(correct_items) | accepted_items

    no_sales_service.remove_no_sales(db, qualified_items)
    no_sales_service.add_no_sales(db, no_sales_ids)
    _set_plan_item_marker(plan, "platform_qualified_items", qualified_items)
    _set_plan_item_marker(plan, "platform_no_sales_items", no_sales_ids)
    _set_plan_item_marker(plan, "platform_hard_failed_items", hard_failed_ids)
    _sync_official_scope_after_platform_result(plan, qualified_items)
    _record_terminal_platform_acceptance(
        plan, pending_rows, accepted_items)
    db.commit()
    return {
        "ok": True,
        "terminal_counts": {
            "total": total, "ok": ok_count, "failed": failed_count,
        },
        "accepted_item_ids": sorted(accepted_items),
        "qualified_item_ids": sorted(qualified_items),
        "no_sales_item_ids": sorted(no_sales_ids),
        "hard_failed_item_ids": sorted(hard_failed_ids),
        "failed_items": failed_rows,
        "feedback_artifact": validation.get("feedback_artifact"),
        "feedback_refresh": feedback_refresh,
    }


def repair_super_reduce_early_activation(
        db: Session, plan, item_ids: list[str], *, phase: str = "stage",
        execution_source: str | None = None) -> dict:
    """Withdraw exact prematurely-active items with three independent proofs.

    Proof 1 is a fresh activity export, proof 2 is Web-Agent exact-row stage/
    post-click readback, and proof 3 is another fresh export.  The function is
    deliberately separate from normal signup so it cannot broaden its target
    from the current activity contents.
    """
    from app.services import campaign_recon_service, web_agent_service

    if execution_source != "campaign_automation_repair":
        return {
            "ok": False,
            "step": "execution_policy_guard",
            "error": "超级立减纠正只允许 ERP 活动程序执行",
        }
    if str(getattr(plan, "campaign_type", "")) != "super_reduce":
        return {"ok": False, "step": "plan_guard", "error": "仅支持超级立减计划"}
    if phase not in ("stage", "commit"):
        return {"ok": False, "step": "args", "error": "phase must be stage or commit"}
    targets = sorted({str(value or "").strip() for value in (item_ids or [])})
    if not targets or len(targets) > 100 or any(not value.isdigit() for value in targets):
        return {"ok": False, "step": "args", "error": "必须提供1至100个精确数字商品ID"}
    authorized = _authorized_withdrawal_items(plan)
    if set(targets) != authorized:
        return {
            "ok": False,
            "step": "explicit_withdrawal_authorization_guard",
            "error": "未经用户对本次精确商品清单的当前明确授权，禁止撤销、暂停或移除活动报名",
            "requested_item_ids": targets,
            "authorized_item_ids": sorted(authorized),
        }

    before = refresh_floor_evidence_from_current_activity(db, plan)
    if not before.get("ok"):
        return {"ok": False, "step": "before_export", "error": before.get("error")}
    rows = before["rows"]
    visible_items = {str(row.get("item_id") or "") for row in rows}
    missing = sorted(set(targets) - visible_items)
    if missing:
        return {
            "ok": False,
            "step": "before_scope_guard",
            "error": "纠正清单中有商品未出现在最新超级立减导出",
            "missing_items": missing,
        }
    active_statuses = set(campaign_recon_service.ACTIVITY_IN_CAMPAIGN_STATUSES)
    active_targets = sorted({
        str(row.get("item_id") or "") for row in rows
        if str(row.get("item_id") or "") in targets
        and row.get("status") in active_statuses
    })
    before_states = {
        item_id: sorted({
            str(row.get("status") or "") for row in rows
            if str(row.get("item_id") or "") == item_id
        })
        for item_id in targets
    }
    if not active_targets:
        if phase == "commit":
            plan.status = "discount_pushed"
            marker = "super_reduce_early_activation_already_clear=" + ",".join(targets)
            if marker not in str(getattr(plan, "remark", "") or ""):
                plan.remark = (
                    f"{getattr(plan, 'remark', '') or ''}; {marker}"
                ).strip("; ")
            db.commit()
        return {
            "ok": True, "no_change": True, "phase": phase,
            "proof_1_before_export": before_states,
            "message": "指定商品均已不在活动中",
            "plan_status": getattr(plan, "status", None),
        }

    staged = web_agent_service.withdraw_super_reduce_items(
        db, active_targets, phase="stage")
    if not staged.get("ok"):
        return {
            "ok": False, "step": "exact_row_stage",
            "error": staged.get("error") or "精确商品行预演失败",
            "proof_1_before_export": before_states,
            "stage": staged,
        }
    staged_ids = sorted({
        str(row.get("item_id") or "")
        for row in staged.get("item_results") or []
        if row.get("result") == "ready"
    })
    if staged_ids != active_targets:
        return {
            "ok": False, "step": "exact_row_stage_guard",
            "error": "Web-Agent预演清单或在场状态与最新导出不一致",
            "expected_active_items": active_targets,
            "staged_ready_items": staged_ids,
            "stage": staged,
        }
    if phase == "stage":
        return {
            "ok": True, "phase": "stage", "ready": True,
            "target_item_ids": targets,
            "active_target_item_ids": active_targets,
            "proof_1_before_export": before_states,
            "proof_2_exact_row_stage": staged.get("item_results") or [],
        }

    committed = web_agent_service.withdraw_super_reduce_items(
        db, active_targets, phase="commit")
    if not committed.get("ok"):
        return {
            "ok": False, "step": "exact_row_commit",
            "error": committed.get("error") or "精确商品撤出失败",
            "proof_1_before_export": before_states,
            "proof_2_exact_row_stage": staged.get("item_results") or [],
            "commit": committed,
        }
    after = refresh_floor_evidence_from_current_activity(db, plan)
    if not after.get("ok"):
        return {
            "ok": False, "step": "after_export",
            "error": after.get("error") or "撤出后无法取得新导出",
            "commit": committed,
        }
    remaining_active = sorted({
        str(row.get("item_id") or "") for row in after["rows"]
        if str(row.get("item_id") or "") in active_targets
        and row.get("status") in active_statuses
    })
    if remaining_active:
        return {
            "ok": False, "step": "after_export_guard",
            "error": "撤出后最新导出仍有目标商品处于活动中",
            "remaining_active_items": remaining_active,
            "commit": committed,
        }

    plan.status = "discount_pushed"
    marker = "super_reduce_early_activation_withdrawn=" + ",".join(active_targets)
    if marker not in str(getattr(plan, "remark", "") or ""):
        plan.remark = f"{getattr(plan, 'remark', '') or ''}; {marker}".strip("; ")
    db.commit()
    return {
        "ok": True, "phase": "commit", "submitted": True,
        "target_item_ids": targets,
        "withdrawn_item_ids": active_targets,
        "proof_1_before_export": before_states,
        "proof_2_exact_row_commit": committed.get("item_results") or [],
        "proof_3_remaining_active_items": remaining_active,
        "plan_status": plan.status,
    }


def push_signup(
        db: Session, plan, *, execution_source: str | None = None,
        reuse_fresh_plan_evidence: bool = False,
        exact_item_scope: Optional[set[str]] = None,
        allow_terminal_no_sales_fallback: bool = True,
) -> dict:
    """推大促报名 (channel promo_signup)。R12: 报名导入即报名成功 (stage 即生效, 无 commit 步)。
    只允许活动自动化程序调用；失败记录事实并停在 alarmed，绝不自动改价或重试。"""
    from app.services import (
        campaign_policy_service,
    )

    try:
        policy = campaign_policy_service.require_policy()
    except Exception as exc:  # noqa: BLE001 - policy is a hard runtime dependency
        return _stop_signup(db, plan, {
            "step": "policy_guard",
            "error": str(exc),
        })
    allowed_source = execution_source in {
        "campaign_automation", "campaign_super_reduce_plan7_resume",
    }
    resume_source = execution_source == "campaign_super_reduce_plan7_resume"
    if not allowed_source:
        return {
            "ok": False,
            "step": "execution_policy_guard",
            "error": "活动报名只允许 ERP 自动报名程序执行；页面或 AI 直推已禁用",
            "requires_user_decision": False,
            "automatic_retry": False,
            "ai_may_adjust_or_resubmit": False,
            "policy_version": policy.get("version"),
        }
    if resume_source and (
            getattr(plan, "status", None) != "resume_executing"
            or getattr(plan, "id", None) != 7
            or getattr(plan, "workflow_key", None)
            != "campaign:super-reduce:2026-09-01"
            or getattr(plan, "campaign_type", None) != "super_reduce"
            or not reuse_fresh_plan_evidence
            or exact_item_scope != {"797294092429"}):
        return {
            "ok": False,
            "step": "resume_execution_policy_guard",
            "error": "计划7恢复执行上下文不完整，拒绝进入平台写入",
            "requires_user_decision": True,
            "automatic_retry": False,
            "ai_may_adjust_or_resubmit": False,
        }
    if getattr(plan, "status", None) == "alarmed":
        return {
            "ok": False,
            "step": "waiting_user_decision",
            "error": "本计划已因报名错误停止，等待用户决定；程序、页面和 AI 均不得自动重试",
            "requires_user_decision": True,
            "automatic_retry": False,
            "ai_may_adjust_or_resubmit": False,
        }

    # Ordinary automation performs a current-state export before final
    # preflight.  The one approved plan-7 recovery instead reuses the already
    # persisted, plan-scoped fresh evidence.  Its wrapper has an exact row hash,
    # evidence-age and CAS guard; no second pre-submit platform read is allowed.
    if resume_source:
        current = {
            "ok": True,
            "rows": [],
            "floor_refresh": {
                "reused": True,
                "scope": f"campaign_price_floor_evidence_v2_plan_{plan.id}",
                "platform_read": False,
            },
            "export_evidence": None,
        }
    else:
        current = refresh_floor_evidence_from_current_activity(db, plan)
        if not current.get("ok"):
            return _halt_prewrite(db, plan, {
                "step": "current_state_export",
                "error": current.get("error") or "无法可靠取得当前活动生效集合",
                "detail": current.get("detail"),
            })
    live_rows = current["rows"]
    floor_refresh = current["floor_refresh"]
    from app.services.campaign_recon_service import ACTIVITY_IN_CAMPAIGN_STATUSES
    active_live_rows = [
        row for row in live_rows
        if row.get("status") in ACTIVITY_IN_CAMPAIGN_STATUSES
    ]
    live_activity_evidence = None
    if active_live_rows:
        live_activity_evidence = sync_live_activity_evidence(
            db, plan, active_live_rows)
        db.commit()

    # Registered no-sales items are excluded before any platform upload.
    checks = preflight(db, plan, exact_item_scope=exact_item_scope)
    critical = [check for check in checks if check.get("level") == "error"]
    if critical:
        return _stop_signup(db, plan, {
            "step": "mandatory_preflight",
            "error": f"程序最终预检发现 {len(critical)} 条阻塞规则，未生成或上传报名表",
            "checks": critical,
            "price_floor_refresh": floor_refresh,
        })

    rows, stats = build_signup_rows(db, plan)
    rows, missing_scope = _apply_authorized_supplement_scope(plan, rows, stats)
    if missing_scope:
        return _stop_signup(db, plan, {
            "step": "supplement_scope_guard",
            "error": "补报授权商品未全部出现在安全报名行中，已停止上传",
            "missing_items": missing_scope,
            "stats": stats,
        })
    exact_scope = {
        str(item_id).strip() for item_id in (exact_item_scope or set())
        if str(item_id).strip()
    }
    if exact_scope:
        rows = [row for row in rows
                if str(row.get("taobao_item_id") or "") in exact_scope]
        stats["exact_resume_scope_items"] = sorted(exact_scope)
        stats["exact_resume_scope_rows"] = len(rows)
    qualified_scope = platform_qualified_items(plan)
    supplement_scope = authorized_supplement_items(plan)
    if platform_scope_present(plan) and not exact_scope:
        allowed = qualified_scope | supplement_scope
        rows = [
            row for row in rows
            if str(row.get("taobao_item_id") or "") in allowed
        ]
        stats["platform_qualified_items"] = sorted(allowed)
        stats["platform_qualified_rows"] = len(rows)

    super_reduce = str(getattr(plan, "campaign_type", "")) == "super_reduce"
    if super_reduce:
        expected_items = {str(row.get("taobao_item_id") or "") for row in rows}
        active_outside_upload = {
            str(row.get("item_id") or "") for row in live_rows
            if row.get("status") in ACTIVITY_IN_CAMPAIGN_STATUSES
            and str(row.get("item_id") or "") not in expected_items
        } - {""}
        # A corrective supplement deliberately uploads only the exact authorized
        # item.  Existing active rows in the already-qualified plan scope must be
        # preserved and audited, not treated as an instruction to withdraw them
        # or as a reason to block the supplement.  Unknown active rows remain a
        # hard stop because they have not passed this plan's platform probe.
        preserved_active_items: list[str] = []
        if supplement_scope:
            preserved_active_items = sorted(active_outside_upload & qualified_scope)
            unexpected_active_items = sorted(
                active_outside_upload - qualified_scope
            )
            stats["preserved_active_items_outside_supplement"] = (
                preserved_active_items
            )
        else:
            unexpected_active_items = sorted(active_outside_upload)
        if unexpected_active_items:
            return _stop_signup(db, plan, {
                "step": "super_reduce_unexpected_active_scope_guard",
                "error": (
                    "长期超级立减仍有本计划范围外商品处于生效状态；"
                    "必须保留现场并纳入完整价格校验；未经精确清单授权不得撤出"
                ),
                "unexpected_active_items": unexpected_active_items,
            })
    stats["policy_version"] = policy.get("version")
    stats["policy_sha256"] = policy.get("_sha256")
    stats["price_floor_refresh"] = floor_refresh
    stats["live_activity_evidence"] = live_activity_evidence
    price_holds = stats.get("excluded_price_hold_items") or []
    if price_holds:
        stats["coupon_floor_hold_notification"] = _notify_coupon_floor_blocks(
            db, plan, price_holds)
    placeholder_check = _check_placeholder_live_prices(stats)
    if placeholder_check["level"] == "error":
        res = {
            "ok": False,
            "step": "placeholder_price_guard",
            "error": "未取得全部占位SKU的平台当前保护价，已停止活动导入",
            "check": placeholder_check,
            "stats": stats,
        }
        return _stop_signup(db, plan, res)
    if placeholder_check["level"] == "warn":
        stats["placeholder_hold_notification"] = _notify_placeholder_price_blocks(
            db, plan, placeholder_check.get("blocked_items") or [])
    if not rows:
        res = {
            "ok": False,
            "step": "price_eligibility_guard",
            "error": "全部目标商品均被报名资格/历史价格线安全门暂缓，无可推送报名行",
            "stats": stats,
        }
        return _stop_signup(db, plan, res)

    # 已完成只读导出：整品全部 SKU 已发布且活动价一致才视为正确；
    # 正确品不重复导入。若发现“已发布但错价”，批量导入无法安全修正，立即停并报告。
    live_by_sku = {
        str(r["sku_id"]): r for r in live_rows
        if not super_reduce
        or r.get("status") in ACTIVITY_IN_CAMPAIGN_STATUSES
    }
    expected_by_item: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        expected_by_item[str(row["taobao_item_id"])].append(row)
    correct_items: set[str] = set()
    wrong_published: list[dict] = []
    authorized_refresh_existing: list[dict] = []
    sku_refresh_scope = authorized_sku_refresh_items(plan)
    for item_id, item_rows in expected_by_item.items():
        seen = [live_by_sku.get(str(r["taobao_sku_id"])) for r in item_rows]
        if all(x is not None for x in seen):
            mismatches = [
                {"sku_id": row["taobao_sku_id"], "expected": row["price"],
                 "actual": current.get("activity_price")}
                for row, current in zip(item_rows, seen)
                if not row.get("is_placeholder")
                and (current.get("activity_price") is None
                or abs(float(current["activity_price"]) - float(row["price"])) > 0.005
                )
            ]
            if mismatches and item_id not in sku_refresh_scope:
                wrong_published.append({"item_id": item_id, "mismatches": mismatches[:20]})
            elif mismatches:
                authorized_refresh_existing.append({
                    "item_id": item_id, "mismatches": mismatches[:20],
                })
            else:
                correct_items.add(item_id)
        elif any(x is not None for x in seen) and item_id not in sku_refresh_scope:
            wrong_published.append({
                "item_id": item_id,
                "error": "已发布集合缺 SKU",
                "missing_skus": [
                    r["taobao_sku_id"] for r, current in zip(item_rows, seen)
                    if current is None
                ][:20],
            })
        elif any(x is not None for x in seen):
            authorized_refresh_existing.append({
                "item_id": item_id,
                "error": "用户确认SKU轮换，允许完整SKU集合原位重导",
                "missing_skus": [
                    r["taobao_sku_id"] for r, current in zip(item_rows, seen)
                    if current is None
                ][:20],
            })
    stats["correct_items_excluded"] = sorted(correct_items)
    stats["wrong_published_items"] = wrong_published
    stats["authorized_sku_refresh_existing_items"] = authorized_refresh_existing
    if wrong_published:
        res = {"ok": False, "step": "published_price_guard",
               "error": f"发现 {len(wrong_published)} 个已发布商品错价/缺SKU，拒绝用新增导入覆盖",
               "wrong_published_items": wrong_published, "stats": stats}
        return _stop_signup(db, plan, res)

    pending = [r for r in rows if str(r["taobao_item_id"]) not in correct_items]
    pending_items = {str(r["taobao_item_id"]) for r in pending}
    stats["pending_items"] = sorted(pending_items)
    stats["pending_rows"] = len(pending)
    if not pending:
        verification = _verify_signup_rows(rows, live_rows)
        if not verification.get("ok"):
            return _stop_signup(db, plan, {
                "step": "already_registered_sku_verification",
                "error": "当前活动导出中的逐SKU活动价不一致",
                "post_submit_verification": verification,
                "post_submit_export_evidence": current.get("export_evidence"),
                "stats": stats,
            })
        plan.status = "signup_pushed"
        _remove_plan_marker(plan, "supplement_items_authorized")
        _remove_plan_marker(plan, "sku_refresh_items_authorized")
        db.commit()
        _clear_signup_failure_dedupe(db, plan)
        result = {
            "ok": True, "no_change": True, "stats": stats,
            "message": "当前活动中所有目标商品已发布且价格一致，无需重复导入",
            "post_submit_verification": verification,
            "post_submit_export_evidence": current.get("export_evidence"),
        }
        result["execution_receipt"] = _record_signup_execution_receipt(
            db, plan, result)
        return result

    identity = _campaign_identity(plan)
    from app.services import campaign_execution_service
    manifest_sha256 = campaign_execution_service.scope_sha256(
        identity=identity,
        rows=pending,
        policy_sha256=str(policy.get("_sha256") or ""),
    )
    attempt, _created = campaign_execution_service.ensure_attempt(
        db,
        plan=plan,
        scope_sha256_value=manifest_sha256,
        result_summary={
            "item_ids": sorted(pending_items),
            "row_count": len(pending),
            "policy_sha256": policy.get("_sha256"),
            "execution_boundary": {
                "platform_write": False,
                "account_action": False,
                "automatic_retry": False,
            },
        },
    )
    stats["execution_attempt_id"] = attempt.id
    stats["immutable_manifest_sha256"] = manifest_sha256
    if attempt.write_claimed or attempt.state in (
            "write_claimed", "platform_terminal", "completed",
            "failed_no_retry", "unknown_no_retry"):
        return _stop_signup(db, plan, {
            "step": "durable_one_shot_guard",
            "error": "campaign_execution_already_claimed_no_retry",
            "execution_attempt_id": attempt.id,
            "execution_attempt_state": attempt.state,
            "stats": stats,
        })

    official_identity = _refresh_official_product_sku_identity(db, pending)
    stats["official_product_sku_identity"] = official_identity
    if not official_identity.get("ok"):
        retryable = campaign_execution_service.failure_is_retryable_prewrite(
            official_identity)
        campaign_execution_service.record_prewrite_failure(
            db, attempt,
            step=str(official_identity.get("step") or "official_product_sku_identity"),
            error_code=str(official_identity.get("error") or "official_product_sku_guard"),
            retryable=retryable,
            detail=official_identity,
        )
        result = {
            "step": official_identity.get("step"),
            "error": official_identity.get("error"),
            "official_product_sku_identity": official_identity,
            "execution_attempt_id": attempt.id,
            "stats": stats,
        }
        return _halt_prewrite(db, plan, result) if retryable else _stop_signup(
            db, plan, result)

    channel = "super_reduce" if super_reduce else "promo_signup"
    # Uploading the promo workbook itself is a real platform write. There is no
    # safe stage phase; both campaign channels therefore use one commit only.
    phase = "commit"
    upload_xlsx = (_build_super_signup_xlsx(pending) if super_reduce
                   else _build_signup_xlsx(
                       pending, signup_shipping_days(plan, pending)))
    request_id = f"campaign-signup-{secrets.token_hex(12)}"
    try:
        attempt = campaign_execution_service.claim_platform_write(
            db, attempt.id, request_id=request_id)
    except ValueError as exc:
        return _stop_signup(db, plan, {
            "step": "durable_one_shot_claim",
            "error": str(exc),
            "execution_attempt_id": attempt.id,
            "stats": stats,
        })
    res = _upload_and_wait(
        db, channel, phase, upload_xlsx,
        _fmt_dt(plan.start_at), _fmt_dt(plan.end_at), plan=plan,
        expected_rows=len(pending), expected_items=len(pending_items))
    res["request_id"] = request_id
    res["execution_attempt_id"] = attempt.id
    observed_write = (
        True if any(bool(res.get(key)) for key in (
            "submitted", "platform_write_observed", "attached"))
        else None
    )
    terminal_state = (
        "platform_terminal"
        if isinstance(res.get("validation"), dict)
        else "unknown_no_retry"
    )
    campaign_execution_service.record_platform_terminal(
        db, attempt,
        state=terminal_state,
        platform_write_observed=observed_write,
        step=str(res.get("step") or "platform_terminal"),
        error_code=str(res.get("error") or "") or None,
        job_id=str(res.get("job") or res.get("job_id") or "") or None,
        result_summary={
            "submitted": res.get("submitted"),
            "validation": res.get("validation"),
            "execution_boundary": {
                "platform_write": observed_write,
                "automatic_retry": False,
            },
        },
    )

    def fail_claimed_attempt(step: str, error_code: str,
                             summary: dict | None = None) -> None:
        campaign_execution_service.record_platform_terminal(
            db, attempt, state="failed_no_retry",
            platform_write_observed=observed_write,
            step=step, error_code=error_code,
            job_id=str(res.get("job") or res.get("job_id") or "") or None,
            result_summary=summary or {},
        )

    res["stats"] = stats
    res["recorded_platform_facts"] = _learn_from_validation(
        db, plan, res.get("validation"))
    classification = _classify_final_signup(
        db, plan, res, pending, correct_items)
    res["terminal_classification"] = classification
    if not classification.get("ok"):
        campaign_execution_service.record_platform_terminal(
            db, attempt, state="failed_no_retry",
            platform_write_observed=observed_write,
            step="signup_terminal_classification",
            error_code=str(classification.get("error") or "signup_terminal_failed"),
            job_id=str(res.get("job") or res.get("job_id") or "") or None,
            result_summary={"classification": classification},
        )
        res.update({
            "ok": False,
            "step": "signup_terminal_classification",
            "error": res.get("error") or classification.get("error")
                     or "正式报名终态无法可靠分类",
        })
        return _stop_signup(db, plan, res)

    no_sales_ids = set(classification.get("no_sales_item_ids") or [])
    if no_sales_ids and not allow_terminal_no_sales_fallback:
        fail_claimed_attempt(
            "terminal_no_sales_requires_new_decision",
            "terminal_no_sales_requires_new_decision",
            {"no_sales_item_ids": sorted(no_sales_ids)},
        )
        res.update({
            "ok": False,
            "step": "terminal_no_sales_requires_new_decision",
            "error": (
                "平台把计划7目标商品判定为无动销；本次恢复只允许一次报名写入，"
                "未自动创建或修改单品立减"
            ),
        })
        return _stop_signup(db, plan, res)
    if no_sales_ids:
        fallback = push_discount(
            db, plan, phase="commit",
            item_scope=no_sales_ids,
            no_sales_items=no_sales_ids,
            terminal_no_sales_fallback=True,
            update_plan_status=False,
        )
        res["no_sales_fallback"] = fallback
        if not fallback.get("ok"):
            fail_claimed_attempt(
                "terminal_no_sales_fallback",
                str(fallback.get("error") or "terminal_no_sales_fallback_failed"),
                {"fallback": fallback},
            )
            res.update({
                "ok": False,
                "step": "terminal_no_sales_fallback",
                "error": fallback.get("error")
                         or "平台判定无销量的商品自动转单品立减失败",
            })
            return _stop_signup(db, plan, res)

    # The operation record proves submission; only a fresh, exact-ID export can
    # prove which SKU prices are actually present in this campaign.
    post_submit = refresh_floor_evidence_from_current_activity(db, plan)
    if not post_submit.get("ok"):
        fail_claimed_attempt(
            "post_submit_export",
            str(post_submit.get("error") or "post_submit_export_failed"),
            {"post_submit": post_submit},
        )
        res.update({
            "ok": False,
            "step": "post_submit_export",
            "error": post_submit.get("error")
                     or "平台已返回终态，但无法取得精确活动的新导出",
        })
        return _stop_signup(db, plan, res)
    res["post_submit_export_evidence"] = post_submit.get("export_evidence")
    accepted_ids = set(classification.get("accepted_item_ids") or [])
    accepted_rows = [
        row for row in pending
        if str(row.get("taobao_item_id") or "") in accepted_ids
    ]
    verification = _verify_signup_rows(
        accepted_rows, post_submit.get("rows") or [])
    res["post_submit_verification"] = verification
    if not verification.get("ok"):
        campaign_execution_service.record_platform_terminal(
            db, attempt, state="failed_no_retry",
            platform_write_observed=True,
            step="post_submit_sku_verification",
            error_code="post_submit_sku_verification_failed",
            job_id=str(res.get("job") or res.get("job_id") or "") or None,
            result_summary={"verification": verification},
        )
        res.update({
            "ok": False,
            "step": "post_submit_sku_verification",
            "error": (
                "平台终态已返回，但逐SKU新导出发现活动价为空/不一致；"
                "不得视为报名完成"
            ),
        })
        return _stop_signup(db, plan, res)

    hard_failed = classification.get("hard_failed_item_ids") or []
    if hard_failed:
        fail_claimed_attempt(
            "signup_hard_failures_isolated",
            "signup_hard_failures_isolated",
            {"hard_failed_item_ids": hard_failed},
        )
        res.update({
            "ok": False,
            "step": "signup_hard_failures_isolated",
            "error": "安全商品已报名并核验，但仍有平台硬失败商品等待人工决定",
            "hard_failed_item_ids": hard_failed,
        })
        return _stop_signup(db, plan, res)

    res["ok"] = True
    res.pop("error", None)
    plan.status = "signup_pushed"
    _remove_plan_marker(plan, "supplement_items_authorized")
    _remove_plan_marker(plan, "sku_refresh_items_authorized")
    db.commit()
    _clear_signup_failure_dedupe(db, plan)
    res["execution_receipt"] = _record_signup_execution_receipt(db, plan, res)
    campaign_execution_service.record_platform_terminal(
        db, attempt, state="completed", platform_write_observed=True,
        step="completed",
        job_id=str(res.get("job") or res.get("job_id") or "") or None,
        result_summary={
            "classification": classification,
            "verification": verification,
            "execution_receipt": res.get("execution_receipt"),
        },
    )
    return res


# ── 6. 核对器用的目标到手 (campaign_recon_service 消费) ────────────────────────

def target_prices(db: Session, plan) -> dict[str, dict]:
    """逐 skuId 的 {sku_code, target(目标到手, 不贴线), line, daily, signup_price,
    is_placeholder, kind}。目标 = 大促到手 / 中促到手(×1.03就地) / 无动销=ERP中促价;
    贴线判定交核对器 (它要区分"一分不差"与"贴线让X")。"""
    from app.services import no_sales_service
    tier = plan_tier(plan)
    lev = TIER_LEVERAGE[tier]
    nosales = no_sales_service.get_no_sales(db)
    out: dict[str, dict] = {}
    for s, p in _mapped_pairs(db):
        if not p.taobao_sku_id:
            continue
        placeholder = bool(getattr(s, "is_custom_placeholder", False))
        item_id = str(p.taobao_item_id).strip()
        daily = float(s.daily_price) if s.daily_price else None
        if placeholder:
            price, _remark = _placeholder_signup_price(s, p, lev)
            entry = {"sku_code": s.sku_code, "target": None, "line": None, "daily": daily,
                     "signup_price": price, "is_placeholder": True, "kind": "placeholder"}
        else:
            mid = mid_buyer_inplace(p)
            if item_id in nosales:
                if tier in ("big", "big618"):
                    no_sales_target = _d(getattr(p, "big_buyer_price", None))
                else:
                    no_sales_target = mid.quantize(_CENT) if mid else None
                target = float(no_sales_target) if no_sales_target else None
                kind = "nosales"
            elif tier == "mid":
                target, kind = (float(mid) if mid else None), "campaign"
            else:
                big = _d(getattr(p, "big_buyer_price", None))
                target, kind = (float(big) if big and big > 0 else None), "campaign"
            line = _d(getattr(p, "coupon_floor_price", None))
            entry = {"sku_code": s.sku_code, "target": target,
                     "line": float(line) if line else None, "daily": daily,
                     "signup_price": daily, "is_placeholder": False, "kind": kind}
        for sid in _expand_sku_ids(p):
            out[sid] = entry
    return out

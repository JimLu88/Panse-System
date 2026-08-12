"""工厂系统下单表 → 飞书多维表格。

这张表是给内部运营与木作工厂共看的只读业务投影：
- 一行对应一个平台订单，主键为订单号；
- 常规单同步 SKU 木作成本单价；定制单复用系统定制核价结果，并明确标记待人工核验；
- 不同步总出厂价、订单金额或其它成本；
- 每次系统订单更新、下单图推送完成后立即增量覆盖；
- 保留飞书模板原有的表格 / 状态看板 / 甘特视图所依赖字段。
"""
from __future__ import annotations

import io
import json
import logging
import re
import time
from collections import Counter
from datetime import date, datetime, time as dt_time
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from PIL import Image
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.models.import_file import ImportedFile
from app.models.order import Order, OrderDetail
from app.models.pricing import PricingSku
from app.models.product import Product
from app.services import (
    custom_order_reconcile_service,
    data_quality_service,
    factory_sheet,
    feishu_client,
    import_storage,
    order_flags,
    order_service,
    order_sheet_archive_service,
    settings_service,
)

_logger = logging.getLogger("panse.factory_dispatch_feishu")

SYSTEM_TABLE = "factory_dispatch"
DEFAULT_APP_TOKEN = "TaOebxGaTaU9WVsfA8Pc5vZgnEb"
DEFAULT_TABLE_ID = "tblqn9PDPO0S69wZ"
EXPECTED_VIEWS = {
    "vewaQAjTUH": ("订单管理表", "grid"),
    "vewsN92OdL": ("订单状态看板", "kanban"),
    "vewrj6iAGf": ("订单下单&发货时间", "gantt"),
    "vew4sUBEFl": ("产品图片总览", "gallery"),
}
APP_TOKEN_KEY = "factory_dispatch_feishu_app_token"
TABLE_ID_KEY = "factory_dispatch_feishu_table_id"
IMAGE_CACHE_KEY = "factory_dispatch_feishu_image_tokens"
IMAGE_BINDINGS_KEY = "factory_dispatch_feishu_image_bindings"
AUTO_ENABLED_KEY = "factory_dispatch_feishu_auto_enabled"
INCLUDE_IMAGES_KEY = "factory_dispatch_feishu_include_images"
_CN_TZ = ZoneInfo("Asia/Shanghai")

# 飞书单选字段内置 55 套颜色。10 是低干扰的灰白底/浅色字方案，
# 用于已经退出生产排程的终态，和仍需工厂关注的彩色交期标签明确区分。
TERMINAL_URGENCY_COLOR = 10
TERMINAL_URGENCY_LABELS = frozenset({"完成", "取消", "售后处理", "待核实"})

# 字段 type 见飞书 Bitable：1文本、2数字、3单选、5日期、7复选框、17附件。
# 客户联系方式必须用文本：真实订单可能是固话、多个号码或带文字说明，
# 不能让飞书 Phone 字段的格式校验阻断整单同步。
FIELD_SPECS: tuple[tuple[str, int], ...] = (
    # 字段顺序就是清空重建后的飞书底表顺序。主字段必须是工厂下单号，
    # 工厂在表格、看板和画册中看到的第一项才不会再是平台订单号。
    ("工厂下单号", 1),
    ("商品名称", 1),
    ("订单状态", 3),
    ("客户名称", 1),
    ("客户联系方式", 1),
    ("订购数量", 2),
    ("木作成本价", 2),
    ("下单日期", 5),
    ("预计发货日期", 5),
    ("产品编码", 1),
    ("订单号", 1),
    ("子订单号", 1),
    ("SKU编码", 1),
    ("SKU规格", 1),
    ("工厂下单图", 17),
    ("尺寸", 1),
    ("发货安排", 3),
    ("客户延期单", 7),
    ("客户通知拍照", 7),
    ("订单备注", 1),
    ("客户地址", 1),
    ("物流单号", 1),
    ("系统更新时间", 5),
    ("下单分组", 3),
    ("系统排序键", 1),
    ("订单提醒", 1),
    ("定制标识", 3),
    ("木作成本说明", 1),
    # 生产表已上线后新增的字段必须放尾部，飞书字段 API 不能可靠重排既有列。
    ("交期紧急度", 3),
)

EXPECTED_FIELD_ORDER: tuple[str, ...] = tuple(name for name, _type in FIELD_SPECS)
MAIN_VIEW_LAYOUT = {
    "primary_field": "工厂下单号",
    "group_by": "下单分组",
    "sort_by": "系统排序键",
    "hidden_fields": ["系统排序键", "客户延期单", "客户通知拍照", "系统更新时间"],
}

EXPORT_FIELDS: tuple[str, ...] = (
    "工厂下单号",
    "下单分组",
    "交期紧急度",
    "商品名称",
    "产品编码",
    "SKU编码",
    "SKU规格",
    "尺寸",
    "订购数量",
    "木作成本价",
    "定制标识",
    "木作成本说明",
    "订单状态",
    "发货安排",
    "订单提醒",
    "下单日期",
    "预计发货日期",
    "订单号",
    "子订单号",
    "客户名称",
    "客户联系方式",
    "客户地址",
    "订单备注",
    "物流单号",
    "工厂下单图",
)

# 用户给的是飞书订单模板。沿用原字段 ID 改名，可保留三个视图的列位置和显示配置。
LEGACY_RENAMES: dict[str, tuple[str, int]] = {
    "订单金额": ("木作成本价", 2),
    "销售负责人": ("客户联系方式", 1),
    "库存情况": ("产品编码", 1),
    "负责人销售记录": ("工厂下单号", 1),
    "下单序号": ("系统排序键", 1),
    # 原来这里放的是图库产品图。保留字段 ID 和各视图位置，只改成每天推给
    # 工厂群的同一张下单图，避免飞书视图重新排版。
    "产品图": ("工厂下单图", 17),
}

_PHOTO_KW = (
    "拍照", "拍照片", "拍图片", "拍几张照片", "拍一张照片",
    "通知拍照", "拍照通知", "发货前拍照", "出厂前拍照", "做好拍照",
    "拍照确认", "拍照发", "发照片", "发图片", "发相片", "发图", "传图",
    "提供照片", "提供图片", "照片确认", "图片确认", "看图确认",
    "录视频", "拍视频", "发视频", "传视频", "视频确认", "视频验货",
    "线上验货", "远程验货", "成品图", "完工图", "验货图", "出厂图",
)
_PHOTO_NEGATIVE_KW = (
    "无需拍照", "不用拍照", "不需拍照", "不需要拍照", "不必拍照", "不要拍照", "勿拍照",
    "免拍照", "取消拍照", "无需发图", "不用发图", "不需发图", "不要发图",
    "不需要发图", "无需传图", "不用传图", "不需要传图",
    "无需发照片", "不用发照片", "不需发照片", "不需要发照片", "不要发照片",
    "无需发图片", "不用发图片", "不需发图片", "不需要发图片", "不要发图片",
    "无需提供照片", "不用提供照片", "不需要提供照片",
    "无需提供图片", "不用提供图片", "不需要提供图片",
    "无需视频", "不用视频", "不需视频", "不需要视频", "不要视频",
    "无需录视频", "不用录视频", "不需录视频", "不需要录视频", "不要录视频",
    "无需拍视频", "不用拍视频", "不需拍视频", "不需要拍视频", "不要拍视频",
    "没有拍照要求", "无拍照要求", "不要求拍照",
)
_NON_FACTORY_NAME_KW = (
    "补差", "差价", "补拍", "补邮费", "邮费补拍", "补运费", "补款",
    "尾款", "专拍", "专链", "加价链接",
)
_CUSTOM_SEMANTIC_KW = (
    "定制", "微定制", "补差", "专拍",
    "改尺寸", "尺寸修改", "变更尺寸",
    "改长度", "改宽度", "改高度", "改深度",
    "改材质", "换材质", "改颜色", "换颜色",
    "加长", "加高", "缩短",
)


def _target(db: Session) -> tuple[str, str]:
    return (
        settings_service.get(db, APP_TOKEN_KEY, env_fallback=False) or DEFAULT_APP_TOKEN,
        settings_service.get(db, TABLE_ID_KEY, env_fallback=False) or DEFAULT_TABLE_ID,
    )


def _setting_bool(db: Session, key: str, default: bool) -> bool:
    raw = settings_service.get(db, key, env_fallback=False)
    if raw is None or str(raw).strip() == "":
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


def get_sync_settings(db: Session) -> dict[str, Any]:
    """订单页使用的工厂下单表同步设置；方向固定为 ERP → 飞书。"""
    app_token, table_id = _target(db)
    return {
        "auto_enabled": _setting_bool(db, AUTO_ENABLED_KEY, True),
        "include_images": _setting_bool(db, INCLUDE_IMAGES_KEY, True),
        "direction": "out",
        "direction_label": "仅 ERP → 飞书",
        "app_token": app_token,
        "table_id": table_id,
    }


def save_sync_settings(
    db: Session,
    *,
    auto_enabled: Optional[bool] = None,
    include_images: Optional[bool] = None,
) -> dict[str, Any]:
    if auto_enabled is not None:
        settings_service.set_value(
            db,
            AUTO_ENABLED_KEY,
            "1" if auto_enabled else "0",
            description="工厂系统下单表：订单更新后自动单向同步到飞书",
        )
    if include_images is not None:
        settings_service.set_value(
            db,
            INCLUDE_IMAGES_KEY,
            "1" if include_images else "0",
            description="工厂系统下单表：同步每天推给工厂的下单图到飞书",
        )
    db.commit()
    return get_sync_settings(db)


def _date_ms(value: Optional[date]) -> Optional[int]:
    if value is None:
        return None
    return int(datetime.combine(value, dt_time.min, tzinfo=_CN_TZ).timestamp() * 1000)


def _now_ms() -> int:
    return int(datetime.now(_CN_TZ).timestamp() * 1000)


def _order_notes(order: Order) -> str:
    seen: set[str] = set()
    parts: list[str] = []
    for label, value in (
        ("买家", order.buyer_message),
        ("卖家", order.seller_memo),
        ("ERP", order.remark),
        ("生产", order.production_note),
    ):
        text = str(value or "").strip()
        if text and text not in seen:
            parts.append(f"{label}：{text}")
            seen.add(text)
    return "\n".join(parts)


def _photo_requested(order: Order) -> bool:
    text = "".join(order_flags.order_text(order).split())
    for phrase in _PHOTO_NEGATIVE_KW:
        text = text.replace(phrase, "")
    return any(k in text for k in _PHOTO_KW)


def _is_non_factory_order(order: Order) -> bool:
    text = " ".join(str(x or "") for x in (order.product_name, order.sku, order.sku_code))
    return any(k in text for k in _NON_FACTORY_NAME_KW)


def _status(order: Order, *, remote: bool, refunded: bool) -> str:
    normalized = order_service.normalize_status(order.status)
    if normalized == "cancelled" or refunded:
        return "已作废"
    if normalized == "signed":
        return "已签收"
    if normalized == "shipped":
        return "已发货"
    if remote:
        return "等客户通知"
    if order.is_customer_delayed:
        return "客户延期"
    if normalized == "aftersales":
        return "售后中"
    if order.factory_no:
        return "生产中"
    return "待制单"


def _urgency_label(order: Order, *, refunded: bool, schedule: dict[str, Any]) -> str:
    """交期列始终给出原因：在制单显示排程，终态显示低干扰的结果标签。"""
    normalized = order_service.normalize_status(order.status)
    if refunded or normalized == "cancelled":
        return "取消"
    if normalized in {"shipped", "signed"}:
        return "完成"
    if normalized == "aftersales":
        return "售后处理"
    # 工厂在制总口径会把任意退款金额都排除，但下单表只把全额/接近全额退款视为作废。
    # 部分退款（差价、运费等）仍是正常生产单，交期必须继续按排程显示。
    if normalized == "paid":
        return str(schedule["urgency_label"])
    return "待核实"


def _ship_plan(order: Order, *, remote: bool, photo_requested: bool = False) -> str:
    if photo_requested:
        return "需拍照后通知爱群"
    if remote:
        return "之后发货（等通知）"
    if order.is_customer_delayed:
        return "客户延期（继续生产）"
    return "做好直接发货"


def _wood_unit_price(order: Order, pricing: Optional[PricingSku]) -> Optional[Decimal]:
    qty = Decimal(str(order.qty or 1))
    if order.wood_cost_est is not None and qty > 0:
        return (Decimal(str(order.wood_cost_est)) / qty).quantize(Decimal("0.01"))
    if pricing is not None and pricing.wood_cost is not None:
        return Decimal(str(pricing.wood_cost)).quantize(Decimal("0.01"))
    return None


def _custom_order_info(
    order: Order,
    pricing: Optional[PricingSku],
) -> tuple[bool, str]:
    """飞书工厂表的定制判定：结构化标识、定制 SKU、占位 SKU、备注语义共用。"""
    if bool(order.is_custom):
        return True, "订单定制标识"
    if data_quality_service.is_custom_order(order):
        return True, "定制SKU"
    if pricing is not None and bool(pricing.is_custom_placeholder):
        return True, "定制占位SKU"
    text = " ".join(str(v or "") for v in (
        order.product_name,
        order.sku,
        order.sku_code,
        order_flags.order_text(order),
    ))
    hit = next((keyword for keyword in _CUSTOM_SEMANTIC_KW if keyword in text), None)
    return (True, f"备注关键词：{hit}") if hit else (False, "")


def _production_qty(order: Order, *, is_custom: bool) -> int:
    """定制凑价 SKU 的平台件数不是生产件数；4 件以上按一个生产任务展示。"""
    raw = max(int(order.qty or 1), 1)
    if is_custom and raw >= 4:
        return 1
    return raw


def _custom_wood_unit_price(
    db: Session,
    order: Order,
    pricing: Optional[PricingSku],
    *,
    production_qty: int,
) -> tuple[Optional[Decimal], str]:
    """复用定制核价结果估算木作成本，并对明显不适合木作的兜底回退基础木作价。"""
    qty = Decimal(str(max(production_qty, 1)))
    if order.actual_cost is not None:
        total = Decimal(str(order.actual_cost))
        return (total / qty).quantize(Decimal("0.01")), "工厂实际木作成本"

    base_total: Optional[Decimal] = None
    if order.wood_cost_est is not None:
        base_total = Decimal(str(order.wood_cost_est))
    elif pricing is not None and pricing.wood_cost is not None:
        base_total = Decimal(str(pricing.wood_cost)) * qty

    remark = custom_order_reconcile_service.remark_text(order)
    explicit = custom_order_reconcile_service._r_cost_keyword(db, order, remark)
    if explicit and explicit.get("cost") is not None:
        total = Decimal(str(explicit["cost"]))
        return (total / qty).quantize(Decimal("0.01")), "备注写明成本"

    if order.custom_surcharge is not None and base_total is not None:
        total = base_total + Decimal(str(order.custom_surcharge))
        return (total / qty).quantize(Decimal("0.01")), "基础木作成本+定制加价"

    resolved = custom_order_reconcile_service._display_resolve(db, order, remark)
    resolved_cost = resolved.get("cost")
    method = str(resolved.get("method") or "系统定制估算")
    source = str(resolved.get("source") or "")

    # 85% 与纯插座分支依赖平台实付；定制凑价链接会严重失真，木作表优先保留基础木作成本。
    if base_total is not None and (source in {"fallback", "socket"} or method.startswith("85%")):
        return (base_total / qty).quantize(Decimal("0.01")), "基础木作成本兜底"

    if resolved_cost is not None:
        projected_total = Decimal(str(resolved_cost))
        nonwood_values = (
            order.est_parts,
            order.est_packing,
            order.est_logistics,
            order.est_install,
        )
        known_nonwood = [Decimal(str(value)) for value in nonwood_values if value is not None]
        if known_nonwood:
            total = projected_total - sum(known_nonwood, Decimal("0"))
            method = f"{method}，已扣除非木作成本"
        elif (
            pricing is not None
            and pricing.physical_cost is not None
            and pricing.wood_cost is not None
        ):
            nonwood = max(
                Decimal(str(pricing.physical_cost)) - Decimal(str(pricing.wood_cost)),
                Decimal("0"),
            )
            total = projected_total - nonwood * qty
            method = f"{method}，已扣除非木作成本"
        elif base_total is not None:
            # 没有配件/包装/物流/安装拆分时，不能把物理总成本冒充木作成本。
            return (base_total / qty).quantize(Decimal("0.01")), "基础木作成本兜底"
        else:
            total = projected_total

        # 低于基础木作成本 60% 的结果通常来自凑价/补差金额，不作为木作报价。
        if total <= 0:
            if base_total is None:
                return None, "缺少可用成本依据"
            return (base_total / qty).quantize(Decimal("0.01")), "基础木作成本兜底"
        if base_total is not None and total < base_total * Decimal("0.60"):
            return (base_total / qty).quantize(Decimal("0.01")), "基础木作成本兜底"
        return (total / qty).quantize(Decimal("0.01")), method

    if base_total is not None:
        return (base_total / qty).quantize(Decimal("0.01")), "基础木作成本兜底"
    return None, "缺少可用成本依据"


def build_rows(db: Session) -> list[dict[str, Any]]:
    """按工厂下单图同口径构造飞书行，不改变订单或价格数据。"""
    auto_since = order_sheet_archive_service.AUTO_SINCE
    orders = db.execute(
        select(Order).where(
            Order.is_refill.is_(False),
            or_(
                Order.order_date >= auto_since,
                Order.factory_no.isnot(None),
                Order.remote_seq.isnot(None),
            ),
        ).order_by(Order.order_date.asc().nulls_last(), Order.id.asc())
    ).scalars().all()

    codes = {o.product_code for o in orders if o.product_code}
    product_names = {
        code: name
        for code, name in db.execute(
            select(Product.code, Product.name).where(Product.code.in_(codes))
        ).all()
    } if codes else {}
    sku_codes = {o.sku_code for o in orders if o.sku_code}
    pricing = {
        row.sku_code: row
        for row in db.execute(
            select(PricingSku).where(PricingSku.sku_code.in_(sku_codes))
        ).scalars().all()
    } if sku_codes else {}
    sheet_images = _factory_sheet_images(db)
    now_ms = _now_ms()
    out: list[dict[str, Any]] = []
    for order in orders:
        paid = order_sheet_archive_service._is_paid(order)
        refunded = order_sheet_archive_service._is_refunded(order)
        current = (order.status or "") != "cancelled" and not refunded
        if not paid or (order.status or "") == "pending_payment":
            continue
        if order_sheet_archive_service._is_sample_order(order):
            continue
        topup, _reason = order_sheet_archive_service._is_parts_topup(db, order)
        if topup or _is_non_factory_order(order):
            continue
        if not (order.sku_code or order.sku):
            continue
        line_rows = db.execute(
            select(OrderDetail).where(
                OrderDetail.order_no == order.order_no,
                OrderDetail.source == "import",
                OrderDetail.factory_delivery_required.is_(True),
                OrderDetail.sub_order_no.isnot(None),
            ).order_by(OrderDetail.id.asc())
        ).scalars().all()
        # 子订单链路一行一件；退款但未曾送达的子商品不进入工厂表。
        if line_rows:
            from app.services import order_line_delivery_service
            line_sent = order_line_delivery_service.sent_line_evidence(db)
            for line in line_rows:
                sub_order_no = str(line.sub_order_no or "")
                line_refunded = order_line_delivery_service.line_is_refunded(line)
                if line_refunded and sub_order_no not in line_sent:
                    continue
                ps_line = db.execute(
                    select(PricingSku).where(PricingSku.sku_code == line.sku_code)
                ).scalar_one_or_none()
                current_line = not line_refunded
                factory_no = line.factory_no
                factory_label = f"畔色{factory_no}单" if factory_no else "待编号"
                image_slots = sheet_images.get(sub_order_no) or {}
                image_slot = "void" if line_refunded else "sent"
                sheet_image = image_slots.get(image_slot)
                if (
                    factory_no is None
                    or sheet_image is None
                    or int(sheet_image.get("factory_no_at_render") or 0) != int(factory_no)
                    or str(sheet_image.get("factory_label_at_render") or "") != factory_label
                    or int(sheet_image.get("render_width") or 0) != 1684
                ):
                    sheet_image = None
                schedule = order_flags.factory_schedule(order)
                is_custom, custom_reason = _custom_order_info(order, ps_line)
                production_qty = max(int(line.qty or 1), 1)
                unit_wood = (
                    Decimal(str(ps_line.wood_cost)).quantize(Decimal("0.01"))
                    if ps_line is not None and ps_line.wood_cost is not None else None
                )
                if unit_wood is None and len(line_rows) == 1:
                    unit_wood = _wood_unit_price(order, None)
                photo_requested = _photo_requested(order)
                alerts = ["📷 通知拍照"] if photo_requested else []
                out.append({
                    "订单号": order.order_no,
                    "子订单号": sub_order_no,
                    "工厂下单号": factory_label,
                    "下单分组": "工厂正式单" if factory_no else "待编号",
                    "系统排序键": (
                        f"1-{factory_no:06d}" if factory_no else f"3-{order.id:010d}-{line.id:010d}"
                    ),
                    "商品名称": line.product_name or product_names.get(line.product_code or "") or "",
                    "产品编码": line.product_code or "",
                    "SKU编码": line.sku_code or "",
                    "SKU规格": line.sku_name or "",
                    "尺寸": (ps_line.size_info if ps_line else None) or "",
                    "订购数量": production_qty,
                    "木作成本价": float(unit_wood) if unit_wood is not None else None,
                    "定制标识": "定制单" if is_custom else "常规单",
                    "木作成本说明": (
                        f"定制成本需人工核验｜{custom_reason}" if is_custom else ""
                    ),
                    "下单日期": _date_ms(order.order_date),
                    "预计发货日期": _date_ms(schedule["effective_deadline"]),
                    "订单状态": "已作废" if line_refunded else ("生产中" if factory_no else "待制单"),
                    "交期紧急度": (
                        "取消" if line_refunded else _urgency_label(order, refunded=False, schedule=schedule)
                    ),
                    "发货安排": _ship_plan(order, remote=False, photo_requested=photo_requested),
                    "客户延期单": bool(order.is_customer_delayed),
                    "客户通知拍照": photo_requested,
                    "订单提醒": " · ".join(alerts),
                    "订单备注": _order_notes(order),
                    "客户名称": order.customer_name or "",
                    "客户联系方式": order.customer_phone or "",
                    "客户地址": order.customer_address or "",
                    "物流单号": order.tracking_no or "",
                    "系统更新时间": now_ms,
                    "_order_id": order.id,
                    "_line_id": line.id,
                    "_sheet_path": sheet_image["path"] if sheet_image else None,
                    "_sheet_signature": sheet_image["signature"] if sheet_image else None,
                    "_sheet_name": sheet_image["name"] if sheet_image else None,
                })
            continue
        # 取消/退款但曾经拿过正式或远期编号的单保留，状态明确显示作废；从未进入工厂链路的不展示。
        if not (order.factory_no or order.remote_seq or current):
            continue

        ps = pricing.get(order.sku_code or "")
        remote = order_flags.is_remote(order)
        schedule = order_flags.factory_schedule(order)
        urgency = _urgency_label(order, refunded=refunded, schedule=schedule)
        is_custom, custom_reason = _custom_order_info(order, ps)
        production_qty = _production_qty(order, is_custom=is_custom)
        if is_custom:
            unit_wood, cost_method = _custom_wood_unit_price(
                db,
                order,
                ps,
                production_qty=production_qty,
            )
        else:
            unit_wood = _wood_unit_price(order, ps)
            cost_method = ""
        factory_label = order_flags.factory_label(order)
        if not factory_label and current:
            factory_label = "待编号"
        image_slots = sheet_images.get(order.order_no) or {}
        image_slot = "void" if refunded else "sent"
        sheet_image = image_slots.get(image_slot)
        # 图片里的编号、表格第一列和工厂群发送标题必须逐字对应。没有正式号、
        # 编号不同或还是旧版窄图时一律不挂图，避免工厂按错单生产。
        if (
            order.factory_no is None
            or sheet_image is None
            or int(sheet_image.get("factory_no_at_render") or 0) != int(order.factory_no)
            or str(sheet_image.get("factory_label_at_render") or "") != factory_label
            or int(sheet_image.get("render_width") or 0) != 1684
        ):
            sheet_image = None
        if remote:
            order_group = "远期单"
            order_sequence = int(order.remote_seq) if order.remote_seq is not None else None
            system_sort_key = (
                f"2-{order_sequence:06d}"
                if order_sequence is not None
                else f"2-999999-{order.id:010d}"
            )
        elif order.factory_no:
            order_group = "工厂正式单"
            order_sequence = int(order.factory_no)
            system_sort_key = f"1-{order_sequence:06d}"
        else:
            order_group = "待编号"
            order_sequence = None
            order_day = order.order_date.isoformat() if order.order_date else "9999-12-31"
            system_sort_key = f"3-{order_day}-{order.id:010d}"
        photo_requested = _photo_requested(order)
        alerts = []
        if order.is_customer_delayed:
            alerts.append("⏳ 远期等通知" if remote else "⏳ 客户延期（已开始制作）")
        if photo_requested:
            alerts.append("📷 通知拍照")
        out.append({
            "订单号": order.order_no,
            "子订单号": "",
            "工厂下单号": factory_label,
            "下单分组": order_group,
            "系统排序键": system_sort_key,
            "商品名称": product_names.get(order.product_code or "") or order.product_name or "",
            "产品编码": order.product_code or "",
            "SKU编码": order.sku_code or "",
            "SKU规格": order.sku or "",
            "尺寸": (ps.size_info if ps else None) or "",
            "订购数量": production_qty,
            "木作成本价": float(unit_wood) if unit_wood is not None else None,
            "定制标识": "定制单" if is_custom else "常规单",
            "木作成本说明": (
                f"定制成本需人工核验｜{cost_method}｜{custom_reason}"
                if is_custom else ""
            ),
            "下单日期": _date_ms(order.order_date),
            "预计发货日期": _date_ms(schedule["effective_deadline"]),
            "订单状态": _status(order, remote=remote, refunded=refunded),
            "交期紧急度": urgency,
            "发货安排": _ship_plan(
                order,
                remote=remote,
                photo_requested=photo_requested,
            ),
            "客户延期单": bool(order.is_customer_delayed),
            "客户通知拍照": photo_requested,
            "订单提醒": " · ".join(alerts),
            "订单备注": _order_notes(order),
            "客户名称": order.customer_name or "",
            "客户联系方式": order.customer_phone or "",
            "客户地址": order.customer_address or "",
            "物流单号": order.tracking_no or "",
            "系统更新时间": now_ms,
            "_order_id": order.id,
            "_sheet_path": sheet_image["path"] if sheet_image else None,
            "_sheet_signature": sheet_image["signature"] if sheet_image else None,
            "_sheet_name": sheet_image["name"] if sheet_image else None,
        })
    out.sort(key=lambda row: str(row.get("系统排序键") or "9"))
    return out


def preview_summary(db: Session) -> dict[str, Any]:
    rows = build_rows(db)
    return {
        "rows": len(rows),
        "urgency_counts": dict(Counter(
            str(row.get("交期紧急度") or "待核实")
            for row in rows
        )),
        "group_counts": dict(Counter(str(row.get("下单分组") or "未分组") for row in rows)),
        "custom_count": sum(row.get("定制标识") == "定制单" for row in rows),
        "photo_notice_count": sum(
            row.get("发货安排") == "需拍照后通知爱群" for row in rows
        ),
    }


def export_workbook(db: Session, *, include_images: bool = True) -> bytes:
    """把当前系统下单表导出为 Excel；只读 ERP，不访问或读取飞书记录。"""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = build_rows(db)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "系统下单表"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_FIELDS))}{max(len(rows) + 1, 2)}"

    for col, field in enumerate(EXPORT_FIELDS, start=1):
        cell = sheet.cell(row=1, column=col, value=field)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    image_buffers: list[io.BytesIO] = []
    date_fields = {"下单日期", "预计发货日期"}
    image_col = EXPORT_FIELDS.index("工厂下单图") + 1
    for row_no, row in enumerate(rows, start=2):
        for col, field in enumerate(EXPORT_FIELDS, start=1):
            if field == "工厂下单图":
                continue
            value = row.get(field)
            if field in date_fields and isinstance(value, (int, float)):
                value = datetime.fromtimestamp(value / 1000, tz=_CN_TZ).date()
            cell = sheet.cell(row=row_no, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if field in date_fields and value:
                cell.number_format = "yyyy-mm-dd"
        if include_images and row.get("_sheet_path"):
            try:
                content, _name = _sheet_image_bytes(
                    str(row["_sheet_path"]),
                    str(row.get("_sheet_name") or "工厂下单图.jpg"),
                )
                buffer = io.BytesIO(content)
                image_buffers.append(buffer)
                image = ExcelImage(buffer)
                image.width = 176
                image.height = 105
                sheet.add_image(image, f"{get_column_letter(image_col)}{row_no}")
                sheet.row_dimensions[row_no].height = 80
            except (FileNotFoundError, OSError, ValueError):
                sheet.cell(row=row_no, column=image_col, value="下单图不可用")

    widths = {
        "工厂下单号": 15, "下单分组": 13, "交期紧急度": 13, "商品名称": 28,
        "产品编码": 18, "SKU编码": 20, "SKU规格": 28, "尺寸": 24,
        "订购数量": 10, "木作成本价": 13, "定制标识": 11, "木作成本说明": 32,
        "订单状态": 12, "发货安排": 20, "订单提醒": 22, "下单日期": 13,
        "预计发货日期": 15, "订单号": 22, "子订单号": 22,
        "客户名称": 12, "客户联系方式": 20,
        "客户地址": 38, "订单备注": 45, "物流单号": 22, "工厂下单图": 28,
    }
    for col, field in enumerate(EXPORT_FIELDS, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = widths.get(field, 16)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _ensure_schema(db: Session, app_token: str, table_id: str) -> dict[str, dict]:
    fields = feishu_client.list_table_fields(db, app_token, table_id)
    by_name = {f.get("field_name"): f for f in fields}

    # 先把模板旧列改造成业务列，保留字段 ID 与视图布局。
    for old_name, (new_name, new_type) in LEGACY_RENAMES.items():
        old = by_name.get(old_name)
        if old is None or new_name in by_name:
            continue
        try:
            feishu_client.update_field(
                db, app_token, table_id, old["field_id"],
                field_name=new_name, field_type=new_type,
            )
        except feishu_client.FeishuError:
            _logger.warning("飞书模板字段改造失败 %s→%s，改为新建字段", old_name, new_name, exc_info=True)
            feishu_client.create_field(db, app_token, table_id, new_name, new_type)
        fields = feishu_client.list_table_fields(db, app_token, table_id)
        by_name = {f.get("field_name"): f for f in fields}

    for name, field_type in FIELD_SPECS:
        field = by_name.get(name)
        if field is None:
            feishu_client.create_field(db, app_token, table_id, name, field_type)
            fields = feishu_client.list_table_fields(db, app_token, table_id)
            by_name = {f.get("field_name"): f for f in fields}
            continue
        if field.get("is_primary"):
            continue
        if int(field.get("type") or 0) != field_type:
            feishu_client.update_field(
                db, app_token, table_id, field["field_id"],
                field_name=name, field_type=field_type,
            )
            fields = feishu_client.list_table_fields(db, app_token, table_id)
            by_name = {f.get("field_name"): f for f in fields}

    # 模板原有「订单金额」不是木作成本价。若字段类型无法直接改造而走了新建兜底，
    # 只删除这一个已确认的模板旧列，避免工厂继续看到订单总金额。
    legacy_total = by_name.get("订单金额")
    if legacy_total and by_name.get("木作成本价") and not legacy_total.get("is_primary"):
        feishu_client.delete_field(db, app_token, table_id, legacy_total["field_id"])
        fields = feishu_client.list_table_fields(db, app_token, table_id)
        by_name = {f.get("field_name"): f for f in fields}

    obsolete_inspection_count = by_name.get("验货图片数")
    if obsolete_inspection_count and not obsolete_inspection_count.get("is_primary"):
        feishu_client.delete_field(
            db,
            app_token,
            table_id,
            obsolete_inspection_count["field_id"],
        )
        fields = feishu_client.list_table_fields(db, app_token, table_id)
        by_name = {f.get("field_name"): f for f in fields}

    return by_name


def _ensure_urgency_option_styles(
    db: Session,
    app_token: str,
    table_id: str,
    *,
    fields: Optional[list[dict]] = None,
) -> bool:
    """把交期终态统一成灰白浅色；完整保留其它选项、ID 和字段属性。"""
    if fields is None:
        fields = feishu_client.list_table_fields(db, app_token, table_id)
    field = next(
        (item for item in fields if item.get("field_name") == "交期紧急度"),
        None,
    )
    if not field:
        return False

    property_ = dict(field.get("property") or {})
    options = [dict(option) for option in property_.get("options") or []]
    changed = False
    for option in options:
        if (
            str(option.get("name") or "") in TERMINAL_URGENCY_LABELS
            and int(option.get("color", -1)) != TERMINAL_URGENCY_COLOR
        ):
            option["color"] = TERMINAL_URGENCY_COLOR
            changed = True
    if not changed:
        return False

    property_["options"] = options
    feishu_client.update_field(
        db,
        app_token,
        table_id,
        str(field["field_id"]),
        field_name="交期紧急度",
        field_type=3,
        property_=property_,
        ui_type=str(field.get("ui_type") or "SingleSelect"),
    )
    return True


def _schema_layout_errors(fields: list[dict]) -> list[str]:
    """校验底表结构；除主字段外允许运营在飞书里自由拖动列顺序。"""
    errors: list[str] = []
    names = [str(field.get("field_name") or "") for field in fields]
    primary = next((field for field in fields if field.get("is_primary")), None)
    if not primary or primary.get("field_name") != "工厂下单号":
        errors.append("第一列必须是主字段「工厂下单号」")
    elif not names or names[0] != "工厂下单号":
        errors.append("第一列必须是「工厂下单号」")
    if "下单序号" in names:
        errors.append("旧字段「下单序号」仍存在")
    by_name = {field.get("field_name"): field for field in fields}
    for field_name, expected_type in FIELD_SPECS:
        field = by_name.get(field_name)
        if field is None:
            errors.append(f"缺少字段「{field_name}」")
        elif field.get("type") != expected_type:
            errors.append(f"字段「{field_name}」类型不正确")
    sort_field = by_name.get("系统排序键")
    if not sort_field or sort_field.get("type") != 1:
        errors.append("「系统排序键」必须为文本字段")
    return errors


def _load_image_cache(db: Session) -> dict[str, str]:
    raw = settings_service.get(db, IMAGE_CACHE_KEY, env_fallback=False)
    try:
        value = json.loads(raw or "{}")
        return value if isinstance(value, dict) else {}
    except json.JSONDecodeError:
        return {}


def _load_image_bindings(db: Session) -> dict[str, str]:
    """实体子订单号 → 飞书当前附件对应的工厂图签名。"""
    raw = settings_service.get(db, IMAGE_BINDINGS_KEY, env_fallback=False)
    try:
        value = json.loads(raw or "{}")
        return value if isinstance(value, dict) else {}
    except json.JSONDecodeError:
        return {}


def _factory_sheet_images(db: Session) -> dict[str, dict[str, dict[str, Any]]]:
    """子订单号（旧记录回退主订单号）→ 已发送最终图 / 作废图。

    草稿归档 ``order_sheet`` 生成时可能还没有分配工厂编号，不能用于工厂下单表。
    这里只认实际发送版 ``order_sheet_sent`` 和作废版，并携带生成时编号做逐行校验。
    """
    records = db.execute(
        select(ImportedFile).where(
            ImportedFile.kind.in_(("order_sheet_sent", "order_sheet_void"))
        ).order_by(ImportedFile.id.asc())
    ).scalars().all()
    result: dict[str, dict[str, dict[str, Any]]] = {}
    for record in records:
        summary = record.row_summary or {}
        if record.kind == "order_sheet_sent":
            if summary.get("pushed") is not True:
                continue
            order_no = str(summary.get("order_no") or "").strip() or None
            slot = "sent"
        else:
            order_no = order_sheet_archive_service._void_order_no_from_name(
                record.original_filename
            )
            slot = "void"
        if not order_no:
            continue
        entity_no = str(summary.get("sub_order_no") or "").strip() or order_no
        signature = record.file_hash or f"imported-file:{record.id}:{record.size_bytes or 0}"
        result.setdefault(entity_no, {})[slot] = {
            "path": record.stored_path,
            "signature": f"factory-sheet:{signature}",
            "name": record.original_filename or f"工厂下单图_{order_no}.jpg",
            "factory_no_at_render": summary.get("factory_no_at_render"),
            "factory_label_at_render": summary.get("factory_label_at_render"),
            "render_width": summary.get("render_width"),
        }
    return result


def backfill_factory_sheet_snapshots(db: Session, *, limit: int = 500) -> dict[str, Any]:
    """把历史正式工厂单重制为当前标准横版，并写入可信最终图档案。

    这是幂等的生产迁移工具：已有相同工厂编号、1684px 标准图的订单会跳过。
    不发送飞书消息，只为下单表补齐与历史工厂群编号一致的最终图片。
    """
    rows = build_rows(db)
    snapshots = _factory_sheet_images(db)
    result: dict[str, Any] = {
        "eligible": 0,
        "pending": 0,
        "attempted": 0,
        "generated": 0,
        "void_generated": 0,
        "skipped": 0,
        "errors": [],
        "order_nos": [],
    }
    candidates: list[tuple[Order, bool]] = []
    for row in rows:
        order = db.get(Order, int(row["_order_id"]))
        if order is None or order.factory_no is None:
            continue
        result["eligible"] += 1
        refunded = order_sheet_archive_service._is_refunded(order)
        slot = "void" if refunded else "sent"
        current = (snapshots.get(order.order_no) or {}).get(slot)
        if (
            current
            and int(current.get("factory_no_at_render") or 0) == int(order.factory_no)
            and int(current.get("render_width") or 0) == 1684
        ):
            result["skipped"] += 1
            continue
        candidates.append((order, refunded))

    result["pending"] = len(candidates)
    for order, refunded in candidates[: max(0, int(limit))]:
        result["attempted"] += 1
        try:
            sheet = factory_sheet.build(db, order.id)
            if refunded:
                content = order_sheet_archive_service.render_void_png(sheet)
                d = order.refund_date or date.today()
                import_storage.archive(
                    db,
                    content=content,
                    original_name=f"{d.isoformat()}_{order.order_no}_已作废.jpg",
                    kind="order_sheet_void",
                    source="factory_backfill",
                    on_date=d,
                    row_summary={
                        "note": f"退款作废 ¥{order.refund_amount or 0}",
                        "order_no": order.order_no,
                        "factory_no_at_render": int(order.factory_no),
                        "factory_label_at_render": f"畔色{order.factory_no}单",
                        "render_width": 1684,
                        "backfilled": True,
                    },
                )
                result["void_generated"] += 1
            else:
                content = order_sheet_archive_service.render_png(sheet)
                order_sheet_archive_service.archive_sent_snapshot(
                    db,
                    order,
                    content,
                    source="factory_backfill",
                    backfilled=True,
                )
                result["generated"] += 1
            db.commit()
            result["order_nos"].append(order.order_no)
        except Exception as exc:  # noqa: BLE001 - 单张失败不阻断其它订单
            db.rollback()
            result["errors"].append(
                f"{order.order_no}: {type(exc).__name__}: {exc}"
            )
    result["ok"] = not result["errors"]
    return result


def _sheet_image_bytes(stored_path: str, original_name: str) -> tuple[bytes, str]:
    content = import_storage.read(stored_path)
    with Image.open(io.BytesIO(content)) as image:
        image = image.convert("RGB")
        image.thumbnail((1800, 1800))
        buf = io.BytesIO()
        image.save(buf, format="JPEG", quality=86, optimize=True)
    safe_name = Path(original_name).stem[:210] or "工厂下单图"
    return buf.getvalue(), safe_name + ".jpg"


def _attachment_value(
    db: Session,
    app_token: str,
    stored_path: Optional[str],
    signature: Optional[str],
    original_name: Optional[str],
    cache: dict[str, str],
) -> list[dict]:
    if not stored_path or not signature:
        return []
    token = cache.get(signature)
    if not token:
        content, name = _sheet_image_bytes(
            stored_path,
            original_name or "工厂下单图.jpg",
        )
        token = feishu_client.upload_bitable_image(db, app_token, content, name)
        if not token:
            raise feishu_client.FeishuError(
                f"工厂下单图上传未返回 file_token: {original_name or stored_path}"
            )
        cache[signature] = token
        # 飞书素材接口 5 QPS；顺序上传时留一点间隔，首轮上传后只传新增下单图。
        time.sleep(0.22)
    return [{"file_token": token}]


def _remote_attachment_value(fields: dict[str, Any]) -> list[dict[str, str]]:
    """把飞书附件读回值收敛成可再次写入的最小结构。"""
    raw = fields.get("工厂下单图")
    if not isinstance(raw, list):
        return []
    return [
        {"file_token": str(item["file_token"])}
        for item in raw
        if isinstance(item, dict) and item.get("file_token")
    ]


def _remote_attachment_matches(
    fields: dict[str, Any],
    *,
    signature: str,
    expected_name: str,
    bound_signature: str = "",
) -> bool:
    """确认远端附件就是当前工厂图，允许旧数据按精确文件名安全自举。"""
    attachment = _remote_attachment_value(fields)
    if not attachment:
        return False
    if bound_signature:
        return bound_signature == signature
    # 旧记录还没有本地签名绑定；实体键已精确匹配后，再以正式发送图文件名
    # 一致作为一次性迁移证据。同步成功后写入系统设置，之后只按签名判断。
    names = {
        str(item.get("name") or "").strip()
        for item in (fields.get("工厂下单图") or [])
        if isinstance(item, dict)
    }
    return bool(expected_name and expected_name in names)


def _norm(value: Any) -> Any:
    # 飞书附件字段清空后读回可能是 None，也可能是 []。两者都代表无附件，
    # 必须归一化为同一个值，避免远期/待编号订单每天被重复更新。
    if value in (None, "") or value == []:
        return None
    if isinstance(value, list):
        if all(isinstance(x, dict) and x.get("file_token") for x in value):
            return sorted(x.get("file_token") for x in value)
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    return str(value)


def _equivalent(remote: Any, expected: Any) -> bool:
    # 飞书数字列的读取结果可能是字符串（例如 "1"、"3000"），而写入值是
    # int/float。按数值比较，避免每天把全部订单误判为有变化。
    if isinstance(expected, (int, float, Decimal)) and not isinstance(expected, bool):
        try:
            return Decimal(str(remote)) == Decimal(str(expected))
        except (ArithmeticError, ValueError):
            return False
    return _norm(remote) == _norm(expected)


def _same(remote: dict, expected: dict) -> bool:
    # 系统更新时间只在订单有真实字段变化时随更新载荷写入，不能反过来成为
    # “每次同步都更新全表”的变化来源。
    return all(
        k == "系统更新时间" or _equivalent(remote.get(k), v)
        for k, v in expected.items()
    )


def _is_template_demo(record: dict, expected_order_nos: set[str]) -> bool:
    fields = record.get("fields") or {}
    order_no = str(fields.get("订单号") or "")
    if not order_no or order_no in expected_order_nos:
        return False
    if re.fullmatch(r"F18\d{5}", order_no):
        return True
    raw_date = fields.get("下单日期")
    return isinstance(raw_date, (int, float)) and raw_date < 1704067200000  # 2024-01-01


def sync(db: Session, *, include_images: Optional[bool] = None) -> dict:
    """单向增量同步系统下单表（只写飞书，绝不把飞书内容回写 ERP）。

    成功不发消息；错误由调用方纳入订单自动化告警与重试。
    """
    if include_images is None:
        include_images = _setting_bool(db, INCLUDE_IMAGES_KEY, True)
    app_token, table_id = _target(db)
    result: dict[str, Any] = {
        "table_id": table_id,
        "direction": "out",
        "include_images": include_images,
        "rows": 0,
        "created": 0,
        "updated": 0,
        "deleted_demo": 0,
        "urgency_style_updated": False,
        "missing_wood_cost": [],
        "missing_factory_sheet_image": [],
        "deferred_image_uploads": [],
        "errors": [],
        "views": {},
        "view_layout": MAIN_VIEW_LAYOUT,
    }
    try:
        _ensure_schema(db, app_token, table_id)
        fields = feishu_client.list_table_fields(db, app_token, table_id)
        layout_errors = _schema_layout_errors(fields)
        if layout_errors:
            result["errors"].extend(layout_errors)
            result["ok"] = False
            return result
        views = feishu_client.list_views(db, app_token, table_id)
        view_map = {v.get("view_id"): v for v in views}
        for view_id, (name, kind) in EXPECTED_VIEWS.items():
            got = view_map.get(view_id)
            # 视图名称允许运营在飞书里调整；ID 和类型才决定链接及展示能力。
            ok = bool(got and got.get("view_type") == kind)
            result["views"][view_id] = {
                "ok": ok,
                "name": got.get("view_name") if got else None,
                "type": got.get("view_type") if got else None,
            }
            if not ok:
                result["errors"].append(
                    f"飞书视图不匹配: {view_id} 预期类型{kind}"
                    f"（建议名称{name}）"
                )
        if result["errors"]:
            result["ok"] = False
            return result

        rows = build_rows(db)
        result["rows"] = len(rows)
        result["missing_wood_cost"] = [
            str(row["订单号"]) for row in rows if row.get("木作成本价") is None
        ]
        if result["missing_wood_cost"]:
            result["errors"].append(
                "木作成本价缺失 "
                f"{len(result['missing_wood_cost'])} 单: "
                + ",".join(result["missing_wood_cost"][:10])
            )
            result["ok"] = False
            return result
        cache = _load_image_cache(db)
        cache_before = dict(cache)
        image_bindings = _load_image_bindings(db)
        image_bindings_before = dict(image_bindings)
        remote_rows = feishu_client.list_records(db, app_token, table_id)
        remote_by_no: dict[str, dict] = {}
        legacy_remote_by_order: dict[str, list[dict]] = {}
        for rec in remote_rows:
            fields = rec.get("fields") or {}
            order_no = str(fields.get("订单号") or "").strip()
            sub_order_no = str(fields.get("子订单号") or "").strip()
            entity_key = sub_order_no or order_no
            if entity_key and entity_key not in remote_by_no:
                remote_by_no[entity_key] = rec
            if order_no and not sub_order_no:
                legacy_remote_by_order.setdefault(order_no, []).append(rec)

        creates: list[dict] = []
        create_image_bindings: list[tuple[str, str]] = []
        updates: list[dict] = []
        update_image_bindings: dict[str, tuple[str, str]] = {}
        claimed_remote_ids: set[str] = set()
        expected_order_nos = {str(row["订单号"]) for row in rows}
        for row in rows:
            order_no = str(row["订单号"])
            entity_key = str(row.get("子订单号") or order_no)
            payload = {k: v for k, v in row.items() if not k.startswith("_")}
            sheet_path = row.get("_sheet_path")
            sheet_signature = row.get("_sheet_signature")
            sheet_name = str(row.get("_sheet_name") or "工厂下单图.jpg")
            attachment_confirmed = False
            if not sheet_path:
                result["missing_factory_sheet_image"].append(order_no)

            remote = remote_by_no.get(entity_key)
            if remote is None and row.get("子订单号"):
                # 0139 上线前飞书一行=主订单，没有子订单号。用“主订单+原工厂号”
                # 唯一匹配并原地升级，避免切换后重复新建一行。
                matches = [
                    item for item in legacy_remote_by_order.get(order_no, [])
                    if str(item.get("record_id") or "") not in claimed_remote_ids
                    and str((item.get("fields") or {}).get("工厂下单号") or "").strip()
                    == str(row.get("工厂下单号") or "").strip()
                ]
                if len(matches) == 1:
                    remote = matches[0]
            remote_fields = (remote or {}).get("fields") or {}
            if include_images:
                if not sheet_path:
                    # 字段由「产品图」原位改名而来。没有真正下单图的远期/待编号单
                    # 必须清空旧产品图，不能让工厂误以为那张缩略图就是生产下单图。
                    payload["工厂下单图"] = []
                elif remote is not None and _remote_attachment_matches(
                    remote_fields,
                    signature=str(sheet_signature or ""),
                    expected_name=sheet_name,
                    bound_signature=str(image_bindings.get(entity_key) or ""),
                ):
                    payload["工厂下单图"] = _remote_attachment_value(remote_fields)
                    image_bindings[entity_key] = str(sheet_signature or "")
                    attachment_confirmed = True
                else:
                    try:
                        payload["工厂下单图"] = _attachment_value(
                            db,
                            app_token,
                            str(sheet_path),
                            str(sheet_signature or ""),
                            sheet_name,
                            cache,
                        )
                        attachment_confirmed = True
                    except Exception as e:  # noqa: BLE001 - 附件局部待重试，不否定订单送达
                        # 已有旧附件时原样保留；没有附件时只跳过附件字段，绝不能写 []
                        # 清掉工厂仍在使用的图片。其余业务字段继续增量同步。
                        existing = _remote_attachment_value(remote_fields)
                        if existing:
                            payload["工厂下单图"] = existing
                        else:
                            payload.pop("工厂下单图", None)
                        result["deferred_image_uploads"].append({
                            "order_no": order_no,
                            "sub_order_no": str(row.get("子订单号") or ""),
                            "reason": f"{type(e).__name__}: {e}",
                        })
            else:
                # 关闭图片同步时只更新业务字段，不改变附件签名绑定。
                pass
            if remote is None:
                creates.append(payload)
                create_image_bindings.append(
                    (entity_key, str(sheet_signature or ""))
                    if attachment_confirmed and sheet_signature else ("", "")
                )
            elif not _same(remote.get("fields") or {}, payload):
                record_id = str(remote.get("record_id") or "")
                claimed_remote_ids.add(record_id)
                updates.append({"record_id": remote["record_id"], "fields": payload})
                if attachment_confirmed and sheet_signature:
                    update_image_bindings[record_id] = (
                        entity_key, str(sheet_signature)
                    )

        if creates:
            ids = feishu_client.batch_create_records(db, app_token, table_id, creates)
            result["created"] = sum(bool(x) for x in ids)
            for record_id, binding in zip(ids, create_image_bindings):
                if record_id and binding[0]:
                    image_bindings[binding[0]] = binding[1]
            if result["created"] != len(creates):
                result["errors"].append(
                    f"飞书下单表新建失败 {len(creates) - result['created']}/{len(creates)} 条"
                )
        if updates:
            failed = feishu_client.batch_update_records(db, app_token, table_id, updates)
            result["updated"] = len(updates) - len(failed)
            failed_ids = set(failed)
            for record_id, binding in update_image_bindings.items():
                if record_id not in failed_ids:
                    image_bindings[binding[0]] = binding[1]
            if failed:
                result["errors"].append(f"飞书下单表更新失败 {len(failed)}/{len(updates)} 条")

        # 首次写入新的终态文本时，飞书会自动生成单选项。记录写完后再统一修成
        # 灰白浅色；之后每次同步也会校正，避免运营误改或模板重建后颜色漂移。
        style_fields = feishu_client.list_table_fields(db, app_token, table_id)
        result["urgency_style_updated"] = _ensure_urgency_option_styles(
            db,
            app_token,
            table_id,
            fields=style_fields,
        )

        # 只清飞书模板自带的 2022 年示例行；不删除任何无法确认来源的人工行。
        demo_ids = [
            rec["record_id"] for rec in remote_rows
            if _is_template_demo(rec, expected_order_nos)
        ]
        if rows and not result["errors"] and demo_ids:
            result["deleted_demo"] = feishu_client.batch_delete_records(
                db, app_token, table_id, demo_ids
            )

        if cache != cache_before:
            # 防缓存无限增长：保留最近使用和最新插入的至多 500 个 token。
            if len(cache) > 500:
                used: set[str] = set()
                for row in rows:
                    signature = row.get("_sheet_signature")
                    if signature:
                        used.add(str(signature))
                cache = {k: v for k, v in cache.items() if k in used}
            settings_service.set_value(
                db, IMAGE_CACHE_KEY, json.dumps(cache, ensure_ascii=False),
                description="工厂系统下单表下单图飞书素材 token 缓存",
            )
        if image_bindings != image_bindings_before:
            # 只保留当前系统投影仍存在的实体，防历史记录无限增长。
            current_entities = {
                str(row.get("子订单号") or row.get("订单号") or "")
                for row in rows
            }
            image_bindings = {
                key: value for key, value in image_bindings.items()
                if key in current_entities
            }
            settings_service.set_value(
                db,
                IMAGE_BINDINGS_KEY,
                json.dumps(image_bindings, ensure_ascii=False),
                description="工厂系统下单表实体与当前附件签名绑定",
            )
        settings_service.set_value(
            db, APP_TOKEN_KEY, app_token, description="工厂系统下单表飞书 app_token"
        )
        settings_service.set_value(
            db, TABLE_ID_KEY, table_id, description="工厂系统下单表飞书 table_id"
        )
        db.commit()
    except Exception as e:  # noqa: BLE001 - 交给订单自动化告警与后续重试
        db.rollback()
        result["errors"].append(f"{type(e).__name__}: {e}")
        _logger.exception("工厂系统下单表同步失败")
    result["ok"] = not result["errors"]
    return result


def sync_if_enabled(db: Session) -> dict:
    """自动化入口：关闭自动同步时静默跳过；手动同步仍可直接调用 sync。"""
    if not _setting_bool(db, AUTO_ENABLED_KEY, True):
        return {
            "ok": True,
            "skipped": "auto_disabled",
            "direction": "out",
            "rows": 0,
            "created": 0,
            "updated": 0,
            "errors": [],
        }
    return sync(db)

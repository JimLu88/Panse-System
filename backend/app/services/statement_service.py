"""月度对账表 (业务需求: 木作 / 岩板 / 玻璃 / 未来自定义).

build_statement_data(db, supplier_id, year, month) → StatementData
render_excel(data) → bytes (openpyxl)
render_html(data) → str (浏览器打印为 PDF; 业务需求里 PDF 也要)

汇总规则:
    - 按 delivery_date 落在 (year, month) 范围内的 delivery_notes
    - 一行单据展开: 单号 / 日期 / 商品 / 规格 / 数量 / 单价 / 金额 / 状态 / 关联订单号
    - 底部统计: 单据数 / 总金额 / 已付 / 未付
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from typing import Optional

from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from app.models.supplier import DeliveryNote, DeliveryNoteLine, Supplier


@dataclass
class StatementRow:
    note_id: int
    note_no: Optional[str]
    delivery_date: Optional[date]
    line_no: int
    item_name: str
    spec: str
    unit: str
    qty: Decimal
    unit_price: Optional[Decimal]
    amount: Optional[Decimal]
    matched_order_no: Optional[str]
    match_confidence: Optional[Decimal]
    note_status: str
    paid_at: Optional[date]


@dataclass
class StatementData:
    supplier: Supplier
    year: int
    month: int
    rows: list[StatementRow]
    note_count: int
    total_amount: Decimal
    paid_amount: Decimal
    unpaid_amount: Decimal
    warnings: list[str] = field(default_factory=list)


def build_statement_data(
    db: Session, *, supplier_id: int, year: int, month: int,
) -> StatementData:
    supplier = db.get(Supplier, supplier_id)
    if supplier is None:
        raise ValueError(f"supplier {supplier_id} not found")

    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    notes = list(db.execute(
        select(DeliveryNote).where(and_(
            DeliveryNote.supplier_id == supplier_id,
            DeliveryNote.delivery_date >= start,
            DeliveryNote.delivery_date < end,
        )).order_by(DeliveryNote.delivery_date, DeliveryNote.id)
    ).scalars().all())

    rows: list[StatementRow] = []
    paid_amount = Decimal("0")
    total_amount = Decimal("0")
    warnings: list[str] = []
    for n in notes:
        if n.total_amount:
            total_amount += n.total_amount
            if n.status == "paid":
                paid_amount += n.total_amount
        if not n.note_no:
            warnings.append(f"单据 #{n.id} 缺少单号")
        lines = list(db.execute(
            select(DeliveryNoteLine).where(DeliveryNoteLine.delivery_note_id == n.id)
            .order_by(DeliveryNoteLine.line_no)
        ).scalars().all())
        if not lines:
            rows.append(StatementRow(
                note_id=n.id, note_no=n.note_no, delivery_date=n.delivery_date,
                line_no=0, item_name="(空单据)", spec="", unit="",
                qty=Decimal("0"), unit_price=None, amount=n.total_amount,
                matched_order_no=None, match_confidence=None,
                note_status=n.status, paid_at=n.paid_at.date() if n.paid_at else None,
            ))
            continue
        for ln in lines:
            rows.append(StatementRow(
                note_id=n.id, note_no=n.note_no, delivery_date=n.delivery_date,
                line_no=ln.line_no, item_name=ln.item_name or "",
                spec=ln.spec or "", unit=ln.unit or "",
                qty=ln.qty, unit_price=ln.unit_price, amount=ln.amount,
                matched_order_no=ln.matched_order_no,
                match_confidence=ln.match_confidence,
                note_status=n.status, paid_at=n.paid_at.date() if n.paid_at else None,
            ))

    return StatementData(
        supplier=supplier, year=year, month=month, rows=rows,
        note_count=len(notes), total_amount=total_amount,
        paid_amount=paid_amount, unpaid_amount=total_amount - paid_amount,
        warnings=warnings,
    )


# ----------------------------- Excel ----------------------------------- #


def render_excel(data: StatementData) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"{data.year}-{data.month:02d}"

    title = f"{data.supplier.name} {data.year}年{data.month}月对账单"
    ws.merge_cells("A1:K1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["单据号", "送货日期", "行号", "品名", "规格", "单位",
               "数量", "单价", "金额", "状态", "关联订单"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_label = {
        "pending_review": "待审", "confirmed": "已确认",
        "billed": "已开账", "paid": "已付款", "disputed": "争议中",
    }
    for i, row in enumerate(data.rows, start=3):
        ws.cell(row=i, column=1, value=row.note_no or "-")
        ws.cell(row=i, column=2, value=row.delivery_date.isoformat() if row.delivery_date else "")
        ws.cell(row=i, column=3, value=row.line_no)
        ws.cell(row=i, column=4, value=row.item_name)
        ws.cell(row=i, column=5, value=row.spec)
        ws.cell(row=i, column=6, value=row.unit)
        ws.cell(row=i, column=7, value=float(row.qty))
        ws.cell(row=i, column=8, value=float(row.unit_price) if row.unit_price else None)
        ws.cell(row=i, column=9, value=float(row.amount) if row.amount else None)
        ws.cell(row=i, column=10, value=status_label.get(row.note_status, row.note_status))
        order_text = row.matched_order_no or ""
        if row.match_confidence is not None and row.matched_order_no:
            order_text = f"{row.matched_order_no} ({float(row.match_confidence):.0f}%)"
        ws.cell(row=i, column=11, value=order_text)
        for c in range(1, 12):
            ws.cell(row=i, column=c).border = border

    # 汇总
    total_row = len(data.rows) + 4
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=8, value="").font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=float(data.total_amount)).font = Font(bold=True)
    ws.cell(row=total_row + 1, column=1, value="已付款").font = Font(color="00AA00", bold=True)
    ws.cell(row=total_row + 1, column=9, value=float(data.paid_amount)).font = Font(color="00AA00")
    ws.cell(row=total_row + 2, column=1, value="未付款").font = Font(color="CC0000", bold=True)
    ws.cell(row=total_row + 2, column=9, value=float(data.unpaid_amount)).font = Font(color="CC0000")

    widths = [16, 12, 6, 28, 18, 6, 8, 10, 12, 10, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------- HTML (打印 PDF) ------------------------- #


def render_html(data: StatementData) -> str:
    """打印友好 HTML; 业务需求里 PDF 也要 — 浏览器 print → PDF 即得."""
    status_label = {
        "pending_review": "待审", "confirmed": "已确认",
        "billed": "已开账", "paid": "已付款", "disputed": "争议中",
    }
    rows_html = []
    for r in data.rows:
        match_html = ""
        if r.matched_order_no:
            conf = f" ({float(r.match_confidence):.0f}%)" if r.match_confidence else ""
            match_html = f"{r.matched_order_no}{conf}"
        rows_html.append(
            "<tr>"
            f"<td>{_esc(r.note_no) if r.note_no else '-'}</td>"
            f"<td>{r.delivery_date.isoformat() if r.delivery_date else ''}</td>"
            f"<td>{r.line_no}</td>"
            f"<td>{_esc(r.item_name)}</td>"
            f"<td>{_esc(r.spec)}</td>"
            f"<td>{_esc(r.unit)}</td>"
            f"<td class='num'>{r.qty}</td>"
            f"<td class='num'>{r.unit_price or ''}</td>"
            f"<td class='num'>{r.amount or ''}</td>"
            f"<td>{status_label.get(r.note_status, r.note_status)}</td>"
            f"<td>{_esc(match_html)}</td>"
            "</tr>"
        )
    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8">
<title>{_esc(data.supplier.name)} {data.year}-{data.month:02d} 对账单</title>
<style>
  @page {{ size: A4; margin: 14mm; }}
  body {{ font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif;
         font-size: 11px; color: #222; }}
  h1 {{ font-size: 18px; margin: 0 0 4px; text-align: center; }}
  .meta {{ text-align: center; color: #666; margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th, td {{ border: 1px solid #bbb; padding: 4px 6px; }}
  th {{ background: #4472C4; color: #fff; font-weight: 600; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .totals {{ margin-top: 14px; }}
  .totals .row {{ display: flex; justify-content: flex-end; gap: 16px; padding: 2px 0; }}
  .totals .row.bold {{ font-weight: 700; }}
  .totals .paid {{ color: #00aa00; }}
  .totals .unpaid {{ color: #cc0000; }}
  .warnings {{ margin-top: 10px; padding: 8px; background: #fff4e0; border: 1px solid #f5c97a;
               color: #884400; font-size: 11px; }}
  .actions {{ margin: 8px 0; text-align: right; }}
  .actions button {{ font-size: 12px; padding: 6px 14px; }}
  @media print {{ .actions {{ display: none; }} }}
</style></head><body>
<div class="actions"><button onclick="window.print()">打印 / 保存为 PDF</button></div>
<h1>{_esc(data.supplier.name)} 对账单</h1>
<div class="meta">{data.year} 年 {data.month} 月 · 单据 {data.note_count} 张</div>
<table>
  <thead><tr>
    <th>单据号</th><th>日期</th><th>行</th><th>品名</th><th>规格</th><th>单位</th>
    <th>数量</th><th>单价</th><th>金额</th><th>状态</th><th>关联订单</th>
  </tr></thead>
  <tbody>{''.join(rows_html) or '<tr><td colspan="11" style="text-align:center;color:#999">本月无单据</td></tr>'}</tbody>
</table>
<div class="totals">
  <div class="row bold"><span>合计金额</span><span>¥ {data.total_amount}</span></div>
  <div class="row paid"><span>已付款</span><span>¥ {data.paid_amount}</span></div>
  <div class="row unpaid bold"><span>未付款</span><span>¥ {data.unpaid_amount}</span></div>
</div>
{('<div class="warnings"><b>提醒</b>: ' + ' / '.join(_esc(w) for w in data.warnings) + '</div>') if data.warnings else ''}
</body></html>"""


def _esc(s) -> str:
    if s is None:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))

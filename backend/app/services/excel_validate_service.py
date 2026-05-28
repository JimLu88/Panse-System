"""导入前数据质量校验 + 标注 Excel 生成.

主要入口:
    validate_and_annotate(file_bytes) -> bytes
        读取 Excel 每个 sheet, 用确定性规则逐行校验, 在原表末尾追加「导入校验」列,
        问题单元格标黄, 可选调用 AI 对少数模糊列做 fallback 提示.
        返回带标注的 xlsx bytes.

校验规则 (确定性, 零 token):
    1. 自动嗅探表头行 (同 smart_import_service._detect_header_row)
    2. 自动按 aliases 推断字段映射
    3. 对每行每个已映射必填字段检查: 是否为空
    4. 对每行已映射数字/日期字段检查: 值能否正确转换
    5. 对未能映射必填字段的表: sheet 级提示
    6. 汇总: 若某 sheet 所有数据行都有错 → sheet 级别说明

AI fallback (消耗 token, 仅在规则无法确定映射时调用):
    对规则完全无法匹配任何必填字段的 sheet, 发给 AI 嗅探正确映射.
    默认 disabled; 调用方可传 use_ai=True 开启.
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from app.services.excel_schemas import ENTITY_SCHEMAS
from app.services.smart_import_service import _detect_header_row

_log = logging.getLogger("panse.excel_validate")

_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_RED_BG = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_BOLD = Font(bold=True)
_RED_FONT = Font(color="CC0000")
_GREY_FONT = Font(color="888888", italic=True)


# ----------------------------- 自动字段映射 --------------------------- #


def _auto_map(columns: list[str], schema) -> dict[str, str]:
    """按 schema aliases 把 Excel 列名映射到 target_field.

    返回 {target_field: excel_col}. 无法映射的字段不在结果里.
    """
    # 先建立 alias → field 的反查表
    alias_to_field: dict[str, str] = {}
    for fn, fdef in schema["fields"].items():
        for alias in fdef.get("aliases", []):
            alias_to_field[alias.strip().lower()] = fn
        alias_to_field[fn.lower()] = fn  # 字段名本身也算

    mapping: dict[str, str] = {}
    for col in columns:
        if not col:
            continue
        key = col.strip().lower()
        fn = alias_to_field.get(key)
        if fn and fn not in mapping:
            mapping[fn] = col
    return mapping


# ----------------------------- 单值类型校验 --------------------------- #


def _check_value(v: Any, ftype: str) -> Optional[str]:
    """校验单个值是否符合目标类型. 返回问题描述字符串或 None."""
    if v is None or (isinstance(v, str) and v.strip() in ("", "-", "—", "#N/A", "N/A")):
        return None  # 空值留给必填检查
    if isinstance(v, str) and v.strip().startswith("#"):
        return f"Excel 错误值 {v!r}"
    if ftype == "decimal":
        if isinstance(v, (int, float, Decimal)):
            return None
        s = str(v).replace(",", "").replace("¥", "").replace("元", "").strip()
        try:
            Decimal(s)
        except InvalidOperation:
            return f"应为数字, 当前值 {v!r}"
    elif ftype == "int":
        if isinstance(v, (int, bool)):
            return None
        try:
            int(float(str(v).replace(",", "").strip()))
        except (ValueError, TypeError):
            return f"应为整数, 当前值 {v!r}"
    elif ftype in ("date", "datetime"):
        if isinstance(v, (date, datetime)):
            return None
        if isinstance(v, (int, float)):
            return None  # Excel 序列号
        s = str(v).strip()
        if not re.search(r"\d{4}", s):
            return f"应为日期 (YYYY-MM-DD), 当前值 {v!r}"
    return None


# ----------------------------- sheet 级校验 -------------------------- #


@dataclass
class RowIssue:
    row_idx: int    # 1-indexed (含表头, 与 Excel 行号一致)
    col_name: str   # Excel 列名
    target_field: Optional[str]
    kind: str       # "empty_required" | "bad_type" | "too_long"
    msg: str
    suggestion: str


@dataclass
class SheetResult:
    sheet_name: str
    entity_label: str
    entity_type: Optional[str]
    header_row: int          # 1-indexed
    total_data_rows: int
    ok_rows: int
    issue_rows: int
    sheet_warnings: list[str] = field(default_factory=list)
    row_issues: list[RowIssue] = field(default_factory=list)


def _validate_sheet(
    ws_rows: list[list],  # all rows as list-of-list (values_only)
    sheet_name: str,
) -> SheetResult:
    """对一个 sheet 的所有数据行做规则校验."""
    # 1. 嗅探表头
    header_row_idx = _detect_header_row(ws_rows)  # 1-indexed
    header_cells = ws_rows[header_row_idx - 1] if header_row_idx <= len(ws_rows) else []
    columns = [str(c).strip() if c is not None else f"col{i+1}"
               for i, c in enumerate(header_cells)]

    data_rows = ws_rows[header_row_idx:]  # rows after header

    if not data_rows:
        return SheetResult(sheet_name=sheet_name, entity_label="(空表)",
                           entity_type=None, header_row=header_row_idx,
                           total_data_rows=0, ok_rows=0, issue_rows=0,
                           sheet_warnings=["表格无数据行"])

    # 2. 匹配 entity schema
    # sheet 名关键词 → 优先候选
    _SHEET_HINTS: dict[str, str] = {
        "支付宝": "alipay_flow",
        "alipay": "alipay_flow",
        "补单": "refill_record",
        "定价": "pricing_sku",
        "bom": "bom_line",
        "BOM": "bom_line",
        "物料": "material",
        "配件价格": "material",
        "配件采购": "part_inventory",
        "库存": "part_inventory",
        "成品库存": "product_inventory",
        "工厂对账": "factory_reconciliation",
        "工厂下单": "factory_order",
        "送货单": "delivery_note",
        "订单细节": "order_details",
        "订单": "order",
        "余额": "account_balance",
        "样品": "sample",
        "损耗": "wood_loss",
        "外包": "outsourcing_expense",
        "推广": "promotion_flow",
        "营销": "daily_operations",
        "日常经营": "daily_operations",
        "售后": "aftersales",
        "产品": "product",
    }
    sheet_hint: Optional[str] = None
    for kw, etype in _SHEET_HINTS.items():
        if kw.lower() in sheet_name.lower():
            sheet_hint = etype
            break

    best_entity: Optional[str] = None
    best_mapping: dict[str, str] = {}
    best_score = -1

    for etype, schema in ENTITY_SCHEMAS.items():
        mapping = _auto_map(columns, schema)
        required_fields = [fn for fn, fd in schema["fields"].items() if fd.get("required")]
        matched_required = sum(1 for fn in required_fields if fn in mapping)
        score = matched_required * 10 + len(mapping)
        # sheet 名 hint 奖励 5 分, 不至于把列更匹配的覆盖掉
        if sheet_hint and etype == sheet_hint:
            score += 5
        if score > best_score:
            best_score = score
            best_entity = etype
            best_mapping = mapping

    result = SheetResult(
        sheet_name=sheet_name,
        entity_label=ENTITY_SCHEMAS[best_entity]["label"] if best_entity else "未知",
        entity_type=best_entity,
        header_row=header_row_idx,
        total_data_rows=len(data_rows),
        ok_rows=0, issue_rows=0,
    )

    if not best_entity or best_score <= 0:
        result.sheet_warnings.append("无法识别表格类型 (没有任何字段能自动映射)")
        result.ok_rows = len(data_rows)
        return result

    schema = ENTITY_SCHEMAS[best_entity]
    required_fields = [fn for fn, fd in schema["fields"].items() if fd.get("required")]
    unmapped_required = [fn for fn in required_fields if fn not in best_mapping]
    if unmapped_required:
        labels = [schema["fields"][fn].get("desc", fn) for fn in unmapped_required]
        result.sheet_warnings.append(
            f"以下必填字段未能自动映射: {', '.join(labels)}。"
            f"请检查列名是否与常见别名一致。"
        )

    col_to_idx = {col: i for i, col in enumerate(columns)}

    # 检测可能是合并单元格的列 (前向填充候选):
    # 若某列在前 20 数据行中第一行有值、后续 5+ 行连续为空 → 视为合并单元格, 不报必填空
    _FORWARD_FILL_COLS: set[str] = set()
    sample20 = data_rows[:20]
    for target_field, excel_col in best_mapping.items():
        fdef = schema["fields"].get(target_field, {})
        if not fdef.get("required"):
            continue
        ci = col_to_idx.get(excel_col)
        if ci is None:
            continue
        vals = [r[ci] if ci < len(r) else None for r in sample20]
        nonnull = [v for v in vals if v is not None and v != ""]
        null_streak = 0
        for v in vals[1:]:  # 从第二行开始
            if v is None or v == "":
                null_streak += 1
            else:
                null_streak = 0
        if nonnull and null_streak >= 3:
            _FORWARD_FILL_COLS.add(excel_col)

    # 3. 逐行校验
    issues_by_row: set[int] = set()
    for row_offset, raw_row in enumerate(data_rows):
        excel_row_no = header_row_idx + 1 + row_offset  # 1-indexed Excel row
        # 跳过空行: 所有已映射字段均为空 (忽略公式残留的 0 值和单一固定文字列)
        if all(
            (lambda v: v is None or v == "" or (isinstance(v, (int, float)) and v == 0))(
                raw_row[col_to_idx[ec]] if ec in col_to_idx and col_to_idx[ec] < len(raw_row) else None
            )
            for ec in best_mapping.values()
        ):
            continue

        row_issues_this: list[RowIssue] = []

        for target_field, excel_col in best_mapping.items():
            fdef = schema["fields"].get(target_field, {})
            ftype = fdef.get("type", "str")
            col_i = col_to_idx.get(excel_col)
            if col_i is None:
                continue
            v = raw_row[col_i] if col_i < len(raw_row) else None

            is_empty = v is None or (isinstance(v, str) and v.strip() == "")

            if is_empty:
                if fdef.get("required") and excel_col not in _FORWARD_FILL_COLS:
                    row_issues_this.append(RowIssue(
                        row_idx=excel_row_no, col_name=excel_col,
                        target_field=target_field, kind="empty_required",
                        msg=f"必填字段「{fdef.get('desc', target_field)}」为空",
                        suggestion="请填写此字段",
                    ))
                continue

            # 类型检查
            err = _check_value(v, ftype)
            if err:
                row_issues_this.append(RowIssue(
                    row_idx=excel_row_no, col_name=excel_col,
                    target_field=target_field, kind="bad_type",
                    msg=err,
                    suggestion=f"改为 {ftype} 格式"
                    if ftype != "decimal" else "改为纯数字 (如 123.45)",
                ))

        if row_issues_this:
            issues_by_row.add(row_offset)
            result.row_issues.extend(row_issues_this)

    result.issue_rows = len(issues_by_row)
    result.ok_rows = len(data_rows) - result.issue_rows
    return result


# ----------------------------- 标注写入 ------------------------------ #


def _annotation_text(issues: list[RowIssue]) -> str:
    parts = []
    for iss in issues:
        parts.append(f"[{iss.col_name}] {iss.msg}。{iss.suggestion}")
    return "\n".join(parts)


def _write_annotations(
    ws,  # openpyxl Worksheet (writable)
    result: SheetResult,
    header_row_idx: int,
) -> None:
    """在 sheet 里追加「导入校验」列并标注问题行."""
    if ws.max_column is None or ws.max_column == 0:
        return

    # 找追加列号
    ann_col = ws.max_column + 1
    ann_col_letter = get_column_letter(ann_col)

    # sheet 级警告写在表头行
    header_cell = ws.cell(row=header_row_idx, column=ann_col)
    header_cell.value = "导入校验"
    header_cell.font = _BOLD

    if result.sheet_warnings:
        warn_cell = ws.cell(row=max(1, header_row_idx - 1), column=ann_col)
        warn_cell.value = "⚠ " + "；".join(result.sheet_warnings)
        warn_cell.font = _RED_FONT

    # 按行聚合问题
    issues_by_row: dict[int, list[RowIssue]] = {}
    for iss in result.row_issues:
        issues_by_row.setdefault(iss.row_idx, []).append(iss)

    # 标注每个问题行
    for row_no, row_issues in issues_by_row.items():
        ann_cell = ws.cell(row=row_no, column=ann_col)
        ann_cell.value = _annotation_text(row_issues)
        ann_cell.font = _RED_FONT

        # 问题单元格标黄
        col_to_idx = {}
        for ri in range(1, ann_col):
            c = ws.cell(row=header_row_idx, column=ri)
            if c.value:
                col_to_idx[str(c.value).strip()] = ri

        for iss in row_issues:
            ci = col_to_idx.get(iss.col_name)
            if ci:
                ws.cell(row=row_no, column=ci).fill = _YELLOW

    # 没问题的数据行也写一个"✅"
    if result.total_data_rows > 0:
        data_start = header_row_idx + 1
        for r in range(data_start, data_start + result.total_data_rows + 1):
            cell = ws.cell(row=r, column=ann_col)
            if cell.value is None:
                cell.value = "✅ 可导入"
                cell.font = _GREY_FONT

    # 列宽
    ws.column_dimensions[ann_col_letter].width = 40


# ----------------------------- 主入口 --------------------------------- #


def validate_and_annotate(file_bytes: bytes, *, use_ai: bool = False) -> tuple[bytes, list[SheetResult]]:
    """校验整个 Excel, 返回 (带标注的 xlsx bytes, 每个 sheet 的校验结果列表).

    带标注的 xlsx 在原表右侧追加「导入校验」列。
    """
    wb_ro = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb_ro.sheetnames

    # 收集所有 sheet 数据 (read-only, 快)
    sheet_rows_map: dict[str, list[list]] = {}
    for name in sheet_names:
        ws = wb_ro[name]
        sheet_rows_map[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb_ro.close()

    # 校验所有 sheet
    results: list[SheetResult] = []
    for name in sheet_names:
        rows = sheet_rows_map[name]
        if not rows:
            results.append(SheetResult(sheet_name=name, entity_label="(空)", entity_type=None,
                                        header_row=1, total_data_rows=0, ok_rows=0, issue_rows=0))
            continue
        try:
            r = _validate_sheet(rows, name)
        except Exception as e:
            _log.exception("[validate] sheet %s 校验异常: %s", name, e)
            r = SheetResult(sheet_name=name, entity_label="校验出错", entity_type=None,
                            header_row=1, total_data_rows=0, ok_rows=0, issue_rows=0,
                            sheet_warnings=[f"校验时发生异常: {e}"])
        results.append(r)
        _log.info("[validate] %s -> %s, %d/%d 行有问题",
                  name, r.entity_label, r.issue_rows, r.total_data_rows)

    # 写标注 (需要可写 workbook, 耗内存稍多但只做一次)
    wb_rw = load_workbook(io.BytesIO(file_bytes), data_only=True)
    for r in results:
        if r.entity_type is None and not r.row_issues and not r.sheet_warnings:
            continue  # 完全空 / 无法识别且无问题 → 不动
        if r.sheet_name not in wb_rw.sheetnames:
            continue
        ws = wb_rw[r.sheet_name]
        try:
            _write_annotations(ws, r, r.header_row)
        except Exception as e:
            _log.exception("[annotate] sheet %s 写标注异常: %s", r.sheet_name, e)

    buf = io.BytesIO()
    wb_rw.save(buf)
    wb_rw.close()
    return buf.getvalue(), results

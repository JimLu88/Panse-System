# -*- coding: utf-8 -*-
"""全类目 Excel 导出 + 存档轮转 (用户需求 2026-06-12, 2026-06-17 大改)。

- 按 table_explorer.ENTITY_MODELS 每个类目一个 Sheet, 全行导出。
- 每个 Sheet 末列「异常批注」: 该行若有未处理(open)异常, 写进去 + 加单元格批注(Comment)。
- 导出后归档到「资料存档库」(ImportedFile kind=full_export); 超过 MAX_KEEP 份自动删最早(轮转)。

2026-06-17 用户大改:
- **数字按数字**: 复用修好的 exceptions_export_service._cell (Decimal→float、日期原生) +
  金额/百分比/日期 number_format。
- **产品总表按 SKU 展开**: Product ⨝ PricingSku 每 SKU 一行; 价格/成本列用 VLOOKUP 关联「定价总表」
  (改定价总表自动联动) + 「毛利率验算」公式列 (=1-会计成本/大促价)。
- **美化**: 细边框 / 冻结首行 / 自动筛选 / 隔行底色 / 深蓝表头白字 / 合理列宽。
- **类目分色**: 有「类目」列的表, 该列按品类底色区隔。
- **英文转中文**: 表头(原有) + 常见枚举值(状态/类型/重要程度…)转中文。
复用 exceptions_export_service 的源表键/取值助手, 不重复实现。
"""
from __future__ import annotations

import hashlib
import io
import re
from datetime import date
from typing import Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.table_explorer import ENTITY_MODELS
from app.models.exception import DataException
from app.models.import_file import ImportedFile
from app.models.pricing import PricingSku
from app.models.product import Product
from app.services import import_storage
from app.services.exceptions_export_service import (
    _SEVERITY_CN, _cell, _join_notes, _key_column,
)

MAX_KEEP = 30          # 资料存档库里全量导出最多留几份, 超出删最早 (用户拍板)
_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")

# 常见英文字段 → 中文表头兜底 (ENTITY_SCHEMAS 没覆盖的表/列用这个; 内容不翻, 仅表头)
_COMMON_HEADER_CN = {
    "id": "ID", "code": "编码", "name": "名称", "remark": "备注", "unit": "单位",
    "spec": "规格", "price": "单价", "amount": "金额", "qty": "数量", "status": "状态",
    "created_at": "创建时间", "updated_at": "更新时间", "warehouse": "仓库",
    "material_code": "物料编码", "material_name": "物料名称", "product_code": "产品编码",
    "product_name": "产品名称", "sku": "SKU", "sku_code": "SKU编码", "order_no": "订单号",
    "customer_name": "客户", "customer_phone": "电话", "platform": "平台",
    "order_date": "下单日期", "ship_date": "发货日期", "paid_amount": "实付金额",
    "tracking_no": "物流单号", "carrier": "快递公司", "is_custom": "是否定制",
    "is_refill": "是否补单", "lead_time_days": "提前期(天)", "priority": "优先级",
    "physical_qty": "物理库存", "locked_qty": "锁定库存", "safety_stock": "安全库存",
    "size_type": "尺寸类型", "is_discontinued": "已停产", "base_material_code": "基础物料",
    "primary_supplier_id": "主供应商", "alt_supplier_ids": "备选供应商",
    "area": "面积", "width_mm": "宽(mm)", "height_mm": "高(mm)",
    "sub_name": "副名称", "image_url": "图片链接", "category": "类目", "brand": "品牌",
    "listing_status": "上架状态", "main_material": "主材", "aux_material": "辅材",
    "import_job_id": "导入批次", "import_batch_id": "导入批次", "is_factory_provided": "工厂提供",
    "qty_required": "需求数量", "purchase_no": "采购单号", "self_delivered": "自送",
    "order_id": "订单ID", "factory_order_no": "工厂单号", "platform_order_no": "平台订单号",
    "transaction_no": "交易流水号", "transaction_time": "交易时间", "balance": "余额",
    "counterparty": "对方", "account": "账户", "bill_date": "账单日期", "service_type": "服务类型",
    "flow_type": "类型", "transaction_date": "交易日期", "gross_margin_rate": "毛利率",
    "accounting_cost": "会计成本", "big_promo": "大促价", "mid_promo": "中促价",
    "small_promo": "小促价", "daily_price": "日常价", "list_price": "标价",
    "factory_cost": "工厂成本", "wood_cost": "木作成本", "size_category": "尺寸类型",
    "payment_status": "付款状态", "compensation_fee": "赔付金额", "refund_amount": "退款金额",
    "refund_status": "退款状态", "expense_type": "费用类型",
}

# ── 枚举值 → 中文 (英文转中文, 让普通人看懂) ──────────────────────────────────
_VALUE_CN: dict[str, dict[str, str]] = {
    "status": {
        "pending_payment": "待付款", "unpaid": "待付款", "paid": "已付款",
        "production": "生产中", "producing": "生产中", "in_production": "生产中",
        "shipped": "已发货", "signed": "已签收", "completed": "已完成", "done": "已完成",
        "aftersales": "售后中", "refunding": "退款中", "refunded": "已退款",
        "cancelled": "已取消", "canceled": "已取消", "closed": "已关闭",
        "open": "未处理", "resolved": "已处理", "ignored": "已忽略", "pending": "待处理",
        "processing": "处理中", "matched": "已匹配", "unmatched": "未匹配",
        "active": "启用", "inactive": "停用", "settled": "已结清", "overdue": "逾期",
    },
    "payment_status": {"unpaid": "未付款", "paid": "已付款", "partial": "部分付款",
                       "overdue": "逾期", "settled": "已结清", "pending": "待付款"},
    "priority": {"high": "高", "mid": "中", "medium": "中", "low": "低"},
    "flow_type": {"recharge": "充值", "expense": "支出", "refund": "退款",
                  "income": "收入", "topup": "充值"},
    "expense_type": {"fixed": "固定成本", "variable": "变动成本", "fixed_cost": "固定成本",
                     "variable_cost": "变动成本"},
    "supplier_type": {"factory": "工厂", "material": "材料商", "logistics": "物流商",
                      "service": "服务商", "hardware": "五金", "accessory": "配件"},
    "sample_type": {"photo": "拍摄", "display": "陈列", "test": "测试", "gift": "赠送"},
    "tier": {"vip": "VIP", "normal": "普通", "new": "新客", "old": "老客"},
    "project_type": {"brand": "品牌", "performance": "效果", "content": "内容"},
    "size_category": {"small": "小型", "mid": "中型", "medium": "中型", "large": "大型"},
    "severity": dict(_SEVERITY_CN),
}

# 以小数存储的「比率」列 → 百分比格式 (gross_margin_rate=0.15 → 15.00%)
_PCT_COLS = {
    "gross_margin_rate", "platform_fee_rate",
    "mid_platform_discount", "big_platform_discount", "xhs_promo_discount",
}

# 类目分色调色板 (柔和, 不刺眼)
_CAT_PALETTE = [
    "FCE4D6", "E2EFDA", "DDEBF7", "FFF2CC", "EAD1DC",
    "D9E1F2", "FBE5D6", "E2F0D9", "FFE699", "D6DCE4",
    "F8CBAD", "C6E0B4", "BDD7EE", "FFD966", "D5A6BD",
]


def _cn_header(entity_key: str, col: str) -> str:
    """英文字段名 → 中文表头: 优先取 ENTITY_SCHEMAS 字段 desc(去括号), 再兜底常用映射, 最后原样。"""
    from app.services.excel_schemas import ENTITY_SCHEMAS
    sch = ENTITY_SCHEMAS.get(entity_key)
    if sch and col in sch.get("fields", {}):
        desc = sch["fields"][col].get("desc")
        if desc:
            return re.split(r"[（(]", desc)[0].strip() or col
    return _COMMON_HEADER_CN.get(col, col)


def _safe_sheet_name(label: str, used: set[str]) -> str:
    """openpyxl Sheet 名: 去非法字符 []:*?/\\, ≤31 字符, 去重。"""
    name = _INVALID_SHEET.sub("·", (label or "表").strip())[:31] or "表"
    base = name
    i = 2
    while name in used:
        suffix = f"({i})"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _col_type(col) -> str:
    t = str(col.type).lower()
    if "bool" in t:
        return "bool"
    if "int" in t:
        return "int"
    if "numeric" in t or "float" in t or "decimal" in t:
        return "decimal"
    if "date" in t and "time" in t:
        return "datetime"
    if "date" in t or "time" in t:
        return "date"
    return "str"


def _num_fmt(col: str, ctype: str) -> Optional[str]:
    """列 → Excel number_format。金额千分位两位、比率百分比、整数、日期。"""
    if col in _PCT_COLS:
        return "0.00%"
    if ctype == "date":
        return "yyyy-mm-dd"
    if ctype == "datetime":
        return "yyyy-mm-dd hh:mm"
    if col == "id" or col.endswith("_id") or col.endswith("_job_id"):
        return "0"
    if ctype == "int":
        return "#,##0"
    if ctype == "decimal":
        return "#,##0.00"
    return None


def _translate(col: str, v):
    """枚举值英文 → 中文 (仅当列在 _VALUE_CN 且值命中)。"""
    if isinstance(v, str):
        m = _VALUE_CN.get(col)
        if m and v in m:
            return m[v]
    return v


def _cat_fill(cat):
    """类目 → 稳定底色 (md5 → 调色板, 跨次运行一致)。"""
    from openpyxl.styles import PatternFill
    if cat in (None, ""):
        return None
    idx = int(hashlib.md5(str(cat).encode("utf-8")).hexdigest(), 16) % len(_CAT_PALETTE)
    return PatternFill("solid", fgColor=_CAT_PALETTE[idx])


def _exception_notes(db: Session, table_name: str) -> dict[str, list[str]]:
    """该源表所有 open 异常 → {source_pk字符串: [批注...]}。"""
    excs = db.execute(
        select(DataException).where(
            DataException.source_table == table_name,
            DataException.status == "open",
        )
    ).scalars().all()
    out: dict[str, list[str]] = {}
    for e in excs:
        if e.source_pk:
            out.setdefault(str(e.source_pk), []).append(
                f"[{_SEVERITY_CN.get(e.severity, e.severity)}] {e.description}")
    return out


# ── 样式 ─────────────────────────────────────────────────────────────────────
def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin", color="D9D9D9")
    return {
        "head_fill": PatternFill("solid", fgColor="1F4E79"),
        "head_font": Font(bold=True, color="FFFFFF", size=11),
        "exc_head_fill": PatternFill("solid", fgColor="C0392B"),
        "exc_row_fill": PatternFill("solid", fgColor="FFF3CD"),
        "zebra_fill": PatternFill("solid", fgColor="F4F7FB"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center"),
    }


def _apply_table_style(ws, headers, *, exc_col_idx=None, note_rows=None,
                       col_fmts=None, cat_col_idx=None, data_end_row=None):
    """表头样式 + 边框 + 隔行底色 + number_format + 类目分色 + 冻结 + 自动筛选 + 列宽。"""
    from openpyxl.utils import get_column_letter

    s = _styles()
    note_rows = note_rows or set()
    col_fmts = col_fmts or {}
    n = len(headers)
    last_data = data_end_row or ws.max_row

    # 表头
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill = s["exc_head_fill"] if (exc_col_idx and ci == exc_col_idx) else s["head_fill"]
        c.font = s["head_font"]
        c.alignment = s["center"]
        c.border = s["border"]
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(h)) * 2 + 4, 10), 42)
    ws.row_dimensions[1].height = 20

    # 数据行
    for ri in range(2, last_data + 1):
        is_note = ri in note_rows
        zebra = (ri % 2 == 0) and not is_note
        for ci in range(1, n + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = s["border"]
            if is_note:
                cell.fill = s["exc_row_fill"]
            elif zebra:
                cell.fill = s["zebra_fill"]
            fmt = col_fmts.get(ci)
            if fmt and cell.value is not None:
                cell.number_format = fmt
        # 类目分色 (覆盖隔行底色)
        if cat_col_idx:
            cv = ws.cell(row=ri, column=cat_col_idx).value
            fill = _cat_fill(cv)
            if fill is not None:
                ws.cell(row=ri, column=cat_col_idx).fill = fill

    ws.freeze_panes = "A2"
    if last_data >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(n)}{max(last_data, 1)}"


# ── 通用类目表 (产品总表以外) ─────────────────────────────────────────────────
def _build_entity_sheet(db: Session, wb, key: str, cfg: dict, used: set[str]):
    from openpyxl.comments import Comment

    model = cfg["model"]
    label = cfg.get("label", key)
    table_name = model.__tablename__
    ws = wb.create_sheet(_safe_sheet_name(label, used))
    cols = [c.key for c in model.__table__.columns]
    headers = [_cn_header(key, c) for c in cols] + ["异常批注"]
    ws.append(headers)

    # 列 → number_format + 类目列下标
    model_cols = {c.key: c for c in model.__table__.columns}
    col_fmts: dict[int, str] = {}
    cat_col_idx = None
    for i, c in enumerate(cols, start=1):
        fmt = _num_fmt(c, _col_type(model_cols[c]))
        if fmt:
            col_fmts[i] = fmt
        if c == "category":
            cat_col_idx = i

    notes = _exception_notes(db, table_name)
    key_col = _key_column(model)
    consumed: set[str] = set()
    exc_col_idx = len(cols) + 1
    note_rows: set[int] = set()

    for r in db.execute(select(model)).scalars().all():
        cand = {str(getattr(r, "id", "") or "")}
        if key_col != "id":
            cand.add(str(getattr(r, key_col, "") or ""))
        matched: list[str] = []
        for k in cand:
            if k and k in notes:
                matched.extend(notes[k])
                consumed.add(k)
        note = _join_notes(matched) if matched else None
        ws.append([_translate(c, _cell(getattr(r, c, None))) for c in cols] + [note])
        if note:
            rid = ws.max_row
            note_rows.add(rid)
            ws.cell(row=rid, column=exc_col_idx).comment = Comment(note, "异常中心")

    data_end = ws.max_row
    _apply_table_style(ws, headers, exc_col_idx=exc_col_idx, note_rows=note_rows,
                       col_fmts=col_fmts, cat_col_idx=cat_col_idx, data_end_row=data_end)

    # 源表里定位不到行的异常 (行已删/键已变) 附在表尾
    orphans = [(k, ns) for k, ns in notes.items() if k not in consumed]
    if orphans:
        from openpyxl.styles import Font
        ws.append([])
        tip = ws.cell(row=ws.max_row + 1, column=1,
                      value="── 以下异常的关联行在本表已找不到 (行已删 / 业务键已变 / 异常已修复待复核) ──")
        tip.font = Font(bold=True, color="C0392B")
        for k, ns in orphans:
            ws.append([k] + [None] * (len(cols) - 1) + ["; ".join(ns)])
    return ws


# ── 产品总表: 按 SKU 展开 + 价格 VLOOKUP 关联定价总表 + 毛利率验算公式 ─────────
def _build_product_sku_sheet(db: Session, wb, used: set[str], pricing_sheet: Optional[str]):
    """Product ⨝ PricingSku 每 SKU 一行。价格/成本列 = VLOOKUP 关联「定价总表」(改定价总表自动联动);
    末加「毛利率验算」公式列 (=1-会计成本/大促价)。无 SKU 的产品保留一行(价格列空)。"""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(_safe_sheet_name("产品总表", used))
    headers = [
        "产品编码", "产品名称", "副名称", "品牌", "类目", "上架状态", "重要程度",
        "主材介绍", "辅材介绍",
        "SKU编码", "SKU名称", "尺寸类型",
        "会计成本", "毛利率", "大促利润", "标价", "日常价", "小促价", "中促价", "大促价",
        "工厂成本", "木作成本", "毛利率验算(公式)", "备注",
    ]
    ws.append(headers)

    # 定价总表列布局 (VLOOKUP 用): sku_code 起到末列, 各目标列相对偏移
    pcols = [c.key for c in PricingSku.__table__.columns]
    sku_i = pcols.index("sku_code")            # 0-based
    sku_letter = get_column_letter(sku_i + 1)
    last_letter = get_column_letter(len(pcols))

    def off(field: str) -> int:
        return pcols.index(field) - sku_i + 1

    pname = pricing_sheet or "定价总表 (全列)"
    pref = pname.replace("'", "''")            # 公式里单引号转义

    # 产品总表中 第13..22 列 → 对应定价总表字段 (VLOOKUP)
    vlookup_fields = {
        13: "accounting_cost", 14: "gross_margin_rate", 15: "big_promo_margin",
        16: "list_price", 17: "daily_price", 18: "small_promo", 19: "mid_promo",
        20: "big_promo", 21: "factory_cost", 22: "wood_cost",
    }

    # 产品 → 其 SKU 列表
    skus_by_code: dict[str, list[PricingSku]] = {}
    for s in db.execute(select(PricingSku).order_by(PricingSku.sku_code)).scalars().all():
        skus_by_code.setdefault(s.product_code, []).append(s)

    notes = _exception_notes(db, "products")
    note_rows: set[int] = set()

    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    r = 1
    for p in products:
        plist = skus_by_code.get(p.code) or [None]
        pnote = notes.get(str(p.code)) or notes.get(str(p.id))
        for s in plist:
            r += 1
            has_sku = s is not None
            row = [
                p.code, p.name, p.sub_name, p.brand, p.category, p.listing_status,
                _translate("priority", p.priority), p.main_material, p.aux_material,
                (s.sku_code if has_sku else None), (s.sku if has_sku else None),
                _translate("size_category", s.size_category) if has_sku else None,
            ]
            # 13..22 价格/成本: VLOOKUP 关联定价总表 (有 SKU 才填公式)
            for ci in range(13, 23):
                if has_sku:
                    row.append(
                        f"=IFERROR(VLOOKUP($J{r},'{pref}'!${sku_letter}:${last_letter},"
                        f"{off(vlookup_fields[ci])},FALSE),\"\")"
                    )
                else:
                    row.append(None)
            # 23 毛利率验算 = 1 - 会计成本(M) / 大促价(T)
            row.append(f"=IFERROR(1-M{r}/T{r},\"\")" if has_sku else None)
            # 24 备注 (SKU 备注; 产品有异常则并入)
            rmk = (s.remark if has_sku else None) or ""
            if pnote:
                rmk = (rmk + "  ⚠" + _join_notes(pnote)).strip()
                note_rows.add(r)
            row.append(rmk or None)
            ws.append(row)

    data_end = ws.max_row
    # number_format: 13..22 金额、14 毛利率%、23 验算%
    col_fmts: dict[int, str] = {ci: "#,##0.00" for ci in range(13, 23)}
    col_fmts[14] = "0.00%"
    col_fmts[23] = "0.00%"
    _apply_table_style(ws, headers, exc_col_idx=None, note_rows=note_rows,
                       col_fmts=col_fmts, cat_col_idx=5, data_end_row=data_end)
    return ws


# ── 定价总表 sheet: 派生列改活公式 (编辑成本→价格/利润 Excel 内自动重算) ─────────────
def _apply_pricing_formulas(ws, *, col_offset: int = 0, data_start_row: int = 2) -> None:
    """把定价总表 sheet 的派生列(物理/各档价/平台费/税/会计/利润/毛利率)改成活公式,
    口径复刻 pricing_calc_service.recompute (2026-07-01 用户拍板对齐 Excel):
      物理 = 工厂 + 物流 + 安装 ; 会计基准 = 物理 / (1 − 2.6%)
      标价/小促/中促/大促 = ROUNDUP(会计基准 / 对应基数, −1) ; 日常 = 标价 × 0.75
      平台费 = 大促×0.6% ; 税 = 大促×2% ; 会计 = 物理+平台费+税 ; 利润 = 大促−会计 ; 毛利率 = 利润/大促
    价格公式仅在该行有对应基数(base_*)时写(无基数保原值, 与引擎一致); 利润链恒随大促联动。
    → 导出的 Excel 里改工厂成本/物流/安装/基数, 价格与利润会像用户原表一样自动联动。

    col_offset: 字段列整体右移几列 (图册 sheet 首列是产品图, 传 1); data_start_row: 数据起始行
    (图册有 分类色带+表头 两行, 传 3)。默认 0/2 = 定价总表原布局。"""
    from openpyxl.utils import get_column_letter
    cols = [c.key for c in PricingSku.__table__.columns]
    idx = {c: i + 1 + col_offset for i, c in enumerate(cols)}
    if "physical_cost" not in idx or "big_promo" not in idx:
        return
    def L(f):
        return get_column_letter(idx[f])
    RATE = "0.026"
    fac, log, ins, phys = L("factory_cost"), L("logistics_cost"), L("install_cost"), L("physical_cost")
    bl, bs, bm, bb = L("base_list"), L("base_small"), L("base_mid"), L("base_big")
    lp, dp, sp, mp, bp = L("list_price"), L("daily_price"), L("small_promo"), L("mid_promo"), L("big_promo")
    plat, tax, acct, marg, rate = (L("platform_fee_rate"), L("tax"), L("accounting_cost"),
                                   L("big_promo_margin"), L("gross_margin_rate"))
    for r in range(data_start_row, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.startswith("──"):
            break  # 到达表尾"孤儿异常"提示区, 停止
        has = lambda col: ws[f"{col}{r}"].value not in (None, "")
        if not has(fac):
            continue
        ws[f"{phys}{r}"] = f"=SUM({fac}{r},{log}{r},{ins}{r})"
        base_expr = f"({phys}{r}/(1-{RATE}))"          # 会计基准 = Excel 成本总计
        if has(bl):
            ws[f"{lp}{r}"] = f"=ROUNDUP({base_expr}/{bl}{r},-1)"
            ws[f"{dp}{r}"] = f"={lp}{r}*0.75"
        if has(bs):
            ws[f"{sp}{r}"] = f"=ROUNDUP({base_expr}/{bs}{r},-1)"
        if has(bm):
            ws[f"{mp}{r}"] = f"=ROUNDUP({base_expr}/{bm}{r},-1)"
        if has(bb):
            ws[f"{bp}{r}"] = f"=ROUNDUP({base_expr}/{bb}{r},-1)"
        if has(bp):                                    # 大促(公式或原值)在 → 利润链联动
            ws[f"{plat}{r}"] = f"={bp}{r}*0.006"
            ws[f"{tax}{r}"] = f"={bp}{r}*0.02"
            ws[f"{acct}{r}"] = f"={phys}{r}+{plat}{r}+{tax}{r}"
            ws[f"{marg}{r}"] = f"={bp}{r}-{acct}{r}"
            ws[f"{rate}{r}"] = f'=IFERROR({marg}{r}/{bp}{r},"")'


def _coupon_if(cellref: str, tiers) -> str:
    """88VIP 消费券阶梯 → 嵌套 IF (降序匹配最高满足档)。tiers: [[阈值, 减额], ...]。"""
    from decimal import Decimal as D
    expr = "0"
    for thr, ded in sorted(tiers, key=lambda t: D(str(t[0]))):
        expr = f"IF({cellref}>={float(thr):g},{float(ded):g},{expr})"
    return expr


def _apply_promo_formulas(ws, pos: dict, *, data_start_row: int,
                          mid_tiers, big_tiers) -> None:
    """活动价倒推链改活公式 (2026-07-02 复刻用户 Excel「活动价」表, 锚 = 各档店铺实收 = 小/中/大促价):
      日常价 = 标价×0.75; 小/中/大促价(=店铺实收) 由售价档位列驱动(成本加成或手填)。
      小促: 买家到手 = 小促价; 单品立减系数 = 小促价 ÷ 日常                    (Excel Q = R/P)
      中促: 买家到手 = 中促价 ÷ (1−中促佣金); 单品立减系数 = 买家到手 ÷ (日常×(1−中促立减)) (Excel U)
            店铺到手 = 中促价(实收); VIP到手 = 买家到手 − 88VIP消费券(嵌套IF阶梯)
      大促: 同中促, 换大促价/大促佣金/大促立减                                (Excel AB)
    → 改 小/中/大促价 (或日常价), 买家到手/系数/店铺到手/VIP 全自动重算; 系数就是要填进淘宝单品立减的数。
    另: 淘宝活动价/小红书标价 = 日常; 小红书促销价 = 活动价×(1−折扣)。"""
    from openpyxl.utils import get_column_letter
    def L(f):
        i = pos.get(f)
        return get_column_letter(i) if i else None
    daily = L("daily_price")
    if not daily:
        return
    C = {f: L(f) for f in (
        "taobao_activity_price", "xhs_list_price", "small_promo", "shop_internal_final", "shop_promo_rate",
        "mid_promo", "mid_buyer_price", "mid_platform_discount", "mid_shop_rate", "mid_vip_commission",
        "mid_shop_receipt", "mid_vip_final", "big_promo", "big_buyer_price", "big_platform_discount",
        "big_shop_rate", "big_vip_commission", "big_shop_receipt", "big_vip_final",
        "xhs_promo_price", "xhs_activity_price", "xhs_promo_discount")}
    def has(col, r):
        return bool(col) and ws[f"{col}{r}"].value not in (None, "")
    def put(f, r, formula):
        if C.get(f):
            ws[f"{C[f]}{r}"] = formula
    for r in range(data_start_row, ws.max_row + 1):
        if not has(daily, r):
            continue
        put("taobao_activity_price", r, f"={daily}{r}")
        put("xhs_list_price", r, f"={daily}{r}")
        # 小促: 店内到手 = 小促价 (单品立减「折/降价金额」改用派生列 _PRICING_DISCOUNT_FIELDS, 不再输出乘法系数)
        if has(C["small_promo"], r):
            put("shop_internal_final", r, f'={C["small_promo"]}{r}')
        # 中促: 买家到手 = 中促价÷(1−佣金); 店铺到手 = 中促价; VIP = 买家 − 消费券
        if has(C["mid_promo"], r):
            mp, mbp = f'{C["mid_promo"]}{r}', f'{C["mid_buyer_price"]}{r}'
            put("mid_buyer_price", r, f'=IFERROR({mp}/(1-{C["mid_vip_commission"]}{r}),"")')
            put("mid_shop_receipt", r, f'={mp}')
            put("mid_vip_final", r, f'={mbp}-({_coupon_if(mbp, mid_tiers)})')
        # 大促: 同理
        if has(C["big_promo"], r):
            bp, bbp = f'{C["big_promo"]}{r}', f'{C["big_buyer_price"]}{r}'
            put("big_buyer_price", r, f'=IFERROR({bp}/(1-{C["big_vip_commission"]}{r}),"")')
            put("big_shop_receipt", r, f'={bp}')
            put("big_vip_final", r, f'={bbp}-({_coupon_if(bbp, big_tiers)})')
        if has(C["xhs_activity_price"], r):
            put("xhs_promo_price", r, f'={C["xhs_activity_price"]}{r}*(1-{C["xhs_promo_discount"]}{r})')


# ── 定价总表: 中文表头 + 平台活动价(淘宝/小红书)补全 + 分类配色 (用户 2026-07-01) ──────
# PricingSkuPromo 追加列 (顺序即列序): 淘宝/店内 → 无国补中促 → 无国补大促 → 小红书
_PRICING_PROMO_FIELDS = [
    "taobao_item_id", "taobao_url", "taobao_sku_id", "taobao_activity_price",
    "shop_internal_promo", "shop_internal_final",
    "mid_platform_discount", "mid_buyer_price",
    "mid_vip_commission", "mid_shop_receipt", "mid_vip_final",
    "big_platform_discount", "big_buyer_price",
    "big_vip_commission", "big_shop_receipt", "big_vip_final",
    "xhs_item_id", "xhs_sku_name", "xhs_sku_id", "xhs_list_price",
    "xhs_activity_price", "xhs_promo_discount", "xhs_promo_price",
]
# 字段 → 中文表头
_PRICING_CN: dict[str, str] = {
    "id": "ID", "product_code": "产品编码", "product_name": "产品名称", "taobao_title": "淘宝标题",
    "sku": "SKU描述", "sku_code": "SKU编码", "size_category": "大小分类", "size_info": "成品尺寸",
    "list_price": "标价", "daily_price": "日常价/单品宝", "small_promo": "小促价",
    "mid_promo": "中促价", "big_promo": "大促价",
    "big_promo_margin": "大促利润", "gross_margin_rate": "毛利率",
    "accounting_cost": "会计总成本", "platform_fee_rate": "平台费率", "tax": "税费",
    "physical_cost": "物理总成本", "logistics_cost": "物流成本", "install_cost": "安装成本",
    "factory_cost": "总出厂成本", "wood_cost": "木作成本", "packaging_cost": "包装成本",
    "external_parts_cost": "外采配件成本合计", "factory_cost_override": "工厂成本手动覆盖",
    "base_list": "标价基数", "base_small": "小促基数", "base_mid": "中促基数", "base_big": "大促基数",
    "image_url": "图片链接", "remark": "备注", "created_at": "创建时间", "updated_at": "更新时间",
    # 淘宝 / 店内活动
    "taobao_item_id": "淘宝商品ID", "taobao_url": "淘宝链接", "taobao_sku_id": "淘宝SKU ID",
    "taobao_activity_price": "淘宝活动价", "shop_promo_rate": "单品立减系数",
    "shop_internal_promo": "单品立减设置价", "shop_internal_final": "店内到手价(小促)",
    # 无国补中促
    "mid_platform_discount": "中促平台立减", "mid_shop_rate": "中促店铺系数", "mid_buyer_price": "中促买家价",
    "mid_vip_commission": "中促88VIP佣金", "mid_shop_receipt": "中促店铺到手", "mid_vip_final": "中促VIP到手价",
    # 无国补大促
    "big_platform_discount": "大促平台立减", "big_shop_rate": "大促店铺系数", "big_buyer_price": "大促买家价",
    "big_vip_commission": "大促88VIP佣金", "big_shop_receipt": "大促店铺到手", "big_vip_final": "大促VIP到手价",
    # 报名价 (派生, 填淘宝超级立减/官方大促报名表)
    "report_price": "88VIP大促报名价", "report_price_618": "超大促报名价(618/双11)",
    # 单品立减 (派生, 加法口径: 淘宝该填的 折 + 降价金额, 每档不同)
    "mid_disc_zhe": "中促单品立减(折)", "mid_disc_amt": "中促降价金额(元)",
    "big_disc_zhe": "大促单品立减(折)", "big_disc_amt": "大促降价金额(元)",
    "big618_disc_zhe": "超大促单品立减(折)", "big618_disc_amt": "超大促降价金额(元)",
    # 小红书
    "xhs_item_id": "小红书商品ID", "xhs_sku_name": "小红书SKU名", "xhs_sku_id": "小红书SKU ID",
    "xhs_list_price": "小红书标价", "xhs_activity_price": "小红书活动价",
    "xhs_promo_discount": "小红书折扣率", "xhs_promo_price": "小红书促销价",
    # 配件成本明细 (PricingSkuCosts, 工厂成本拆到每个配件)
    "rock_slab": "岩板", "drawer_rail": "抽屉轨道", "led_strip": "灯带", "glass": "玻璃",
    "electric_rail": "电力轨道", "packing_sheet": "打包纸片", "iron_pin": "铁销", "connector": "连接片",
    "aluminum_rail": "铝合金轨道", "plastic_rail": "塑料轨道", "mini_handle": "mini把手",
    "nail_free_glue": "免钉胶", "engraving": "雕刻", "acrylic_strip": "亚克力条",
    "embedded_sleeve": "预埋套杆", "cable_mgmt": "理线架+插排", "back_panel": "背板",
    "stainless_trim": "装饰条(不锈钢)", "leg": "腿部", "soft_pack": "软包", "bed_board": "床铺板",
    "other_cost": "其他配件", "other_desc": "外配件说明", "parts_remark": "配件备注",
}
# 配件成本明细字段 (PricingSkuCosts, 仅图册导出展开工厂成本用); 前 22 为数值配件, 后 2 为文本
_PRICING_COST_NUM_FIELDS = [
    "rock_slab", "drawer_rail", "led_strip", "glass", "electric_rail", "packing_sheet",
    "iron_pin", "connector", "aluminum_rail", "plastic_rail", "mini_handle", "nail_free_glue",
    "engraving", "acrylic_strip", "embedded_sleeve", "cable_mgmt", "back_panel", "stainless_trim",
    "leg", "soft_pack", "bed_board", "other_cost",
]
_PRICING_COST_FIELDS = _PRICING_COST_NUM_FIELDS + ["other_desc", "parts_remark"]
# 报名价 (派生, 由 report_prices 计算, 填淘宝超级立减/官方大促报名表)
_PRICING_REPORT_FIELDS = ["report_price", "report_price_618"]
# 单品立减 (派生, 加法口径 single_item_discounts): 只出【降价金额】(淘宝单品立减/单品补贴直接填这个数,
#   SKU级别不支持打折, 故不再出折), 中促/大促/超大促。替代旧乘法系数。
_PRICING_DISCOUNT_FIELDS = ["mid_disc_amt", "big_disc_amt", "big618_disc_amt"]
# 分类 → (表头底色, 字段列表); 表头按分类上色, 排版一眼分区
_PRICING_CATEGORIES: list[tuple[str, str, list[str]]] = [
    ("标识", "1F4E79", ["id", "product_code", "product_name", "taobao_title", "sku", "sku_code", "size_category", "size_info"]),
    ("售价档位", "2E7D32", ["list_price", "daily_price", "small_promo", "mid_promo", "big_promo"]),
    ("利润", "00838F", ["big_promo_margin", "gross_margin_rate"]),
    ("成本", "6A1B9A", ["accounting_cost", "platform_fee_rate", "tax", "physical_cost", "logistics_cost", "install_cost", "factory_cost", "wood_cost", "packaging_cost", "external_parts_cost"]),
    ("加成系数", "B8860B", ["factory_cost_override", "base_list", "base_small", "base_mid", "base_big"]),
    ("备注/时间", "607D8B", ["image_url", "remark", "created_at", "updated_at"]),
    ("淘宝/店内", "E65100", ["taobao_item_id", "taobao_url", "taobao_sku_id", "taobao_activity_price", "shop_internal_promo", "shop_internal_final"]),
    ("淘宝中促", "EF6C00", ["mid_platform_discount", "mid_buyer_price", "mid_vip_commission", "mid_shop_receipt", "mid_vip_final"]),
    ("淘宝大促", "F57F17", ["big_platform_discount", "big_buyer_price", "big_vip_commission", "big_shop_receipt", "big_vip_final"]),
    ("报名价", "1565C0", _PRICING_REPORT_FIELDS),
    ("单品立减(淘宝填)", "0D9488", _PRICING_DISCOUNT_FIELDS),
    ("小红书", "AD1457", ["xhs_item_id", "xhs_sku_name", "xhs_sku_id", "xhs_list_price", "xhs_activity_price", "xhs_promo_discount", "xhs_promo_price"]),
    ("配件成本明细", "5D4037", _PRICING_COST_FIELDS),   # 工厂成本拆到每个配件 (图册导出)
]
_PRICING_FIELD_COLOR: dict[str, str] = {f: color for _n, color, fs in _PRICING_CATEGORIES for f in fs}


def _build_pricing_sheet(db: Session, wb, used: set[str]):
    """定价总表(全列): PricingSku 全列(模型顺序, 保 VLOOKUP/公式列位不变) + 追加平台活动价
    (淘宝/店内/中促大促/小红书) + 中文表头 + 按分类给表头上色。"""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from app.models.pricing_ext import PricingSkuPromo

    label = ENTITY_MODELS.get("pricing_sku", {}).get("label", "定价总表 (全列)")
    ws = wb.create_sheet(_safe_sheet_name(label, used))

    from app.services import pricing_calc_service
    promo_params = pricing_calc_service.get_promo_params(db)
    base_cols = [c.key for c in PricingSku.__table__.columns]   # 模型顺序 → 与 VLOOKUP/公式列位一致
    all_fields = base_cols + _PRICING_PROMO_FIELDS + _PRICING_REPORT_FIELDS + _PRICING_DISCOUNT_FIELDS
    headers = [_PRICING_CN.get(f) or _cn_header("pricing_sku", f) for f in all_fields] + ["异常批注"]
    ws.append(headers)

    promo_by_sku = {p.sku_code: p for p in db.execute(select(PricingSkuPromo)).scalars().all()}

    # number_format
    model_cols = {c.key: c for c in PricingSku.__table__.columns}
    promo_cols = {c.key: c for c in PricingSkuPromo.__table__.columns}
    col_fmts: dict[int, str] = {}
    for i, f in enumerate(all_fields, start=1):
        col = model_cols.get(f)
        if col is None:
            col = promo_cols.get(f)
        if col is not None:
            fmt = _num_fmt(f, _col_type(col))
            if fmt:
                col_fmts[i] = fmt
        elif f in _PRICING_REPORT_FIELDS:                    # 报名价(派生)→金额
            col_fmts[i] = "#,##0.00"
        elif f in _PRICING_DISCOUNT_FIELDS:                  # 单品立减(派生): 折 / 降价金额
            col_fmts[i] = "0.00" if f.endswith("_zhe") else "#,##0.00"

    notes = _exception_notes(db, "pricing_sku")
    exc_col_idx = len(all_fields) + 1
    note_rows: set[int] = set()
    for s in db.execute(select(PricingSku).order_by(PricingSku.sku_code)).scalars().all():
        p = promo_by_sku.get(s.sku_code)
        row = [_translate(f, _cell(getattr(s, f, None))) for f in base_cols]
        row += [_cell(getattr(p, f, None)) if p is not None else None for f in _PRICING_PROMO_FIELDS]
        rp = pricing_calc_service.report_prices(p, promo_params) if p is not None else {}
        sid = (pricing_calc_service.single_item_discounts(p, s.daily_price, promo_params)
               if p is not None else {})
        row += [float(rp.get(f)) if rp.get(f) is not None else None for f in _PRICING_REPORT_FIELDS]
        _dv = {"mid_disc_zhe": sid.get("mid_discount"), "mid_disc_amt": sid.get("mid_deduct"),
               "big_disc_zhe": sid.get("big_discount"), "big_disc_amt": sid.get("big_deduct"),
               "big618_disc_zhe": sid.get("big618_discount"), "big618_disc_amt": sid.get("big618_deduct")}
        row += [(round(float(_dv[f]) * 10, 2) if f.endswith("_zhe") else float(_dv[f]))
                if _dv[f] is not None else None for f in _PRICING_DISCOUNT_FIELDS]
        matched: list[str] = []
        for k in {str(getattr(s, "id", "") or ""), str(s.sku_code or "")}:
            if k and k in notes:
                matched.extend(notes[k])
        note = _join_notes(matched) if matched else None
        row.append(note)
        ws.append(row)
        if note:
            rid = ws.max_row
            note_rows.add(rid)
            ws.cell(rid, exc_col_idx).comment = Comment(note, "异常中心")

    _apply_table_style(ws, headers, exc_col_idx=exc_col_idx, note_rows=note_rows,
                       col_fmts=col_fmts, data_end_row=ws.max_row)

    # 分类配色: 覆盖表头底色 (_apply_table_style 默认深蓝; 这里按分类改)
    white = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    for ci, f in enumerate(all_fields, start=1):
        color = _PRICING_FIELD_COLOR.get(f)
        if color:
            c = ws.cell(1, ci)
            c.fill = PatternFill("solid", fgColor=color)
            c.font = white
            c.alignment = center
    return ws


# ── 定价图册 (带产品图的 Excel, 用户 2026-07-01「批量导出带图」要 Excel 不要 HTML) ──────
_CATALOG_IMG_PX = 88          # 产品图在表内显示的最长边 (px)
_CATALOG_IMG_ROW_PT = 70      # 每个产品组首行行高 (pt, 容得下图)
_CATALOG_DATA_ROW_PT = 20     # 其余 SKU 行行高
_CATALOG_WIDE = {             # 文本列加宽
    "product_name": 22, "taobao_title": 26, "sku": 18, "size_info": 16,
    "remark": 18, "taobao_url": 22, "xhs_sku_name": 16, "image_url": 22,
    "other_desc": 18, "parts_remark": 18,
}


def build_catalog_xlsx(db: Session):
    """定价图册 (Excel, 带产品图): 一 SKU 一行, 首列产品图(同编码多 SKU 只放一张、纵向合并),
    全字段 + 中文表头 + 分类色带 (行1 分类 / 行2 表头, 按分类上色)。返回 BytesIO。
    选图: 本地图库主图优先 → 淘宝 CDN 兜底 (见 pricing_catalog_service.product_image_map)。"""
    import io as _io
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from app.models.pricing import PricingSku
    from app.models.pricing_ext import PricingSkuCosts, PricingSkuPromo
    from app.models.product import Product
    from app.services.pricing_catalog_service import _is_real_product, product_image_map

    base_cols = [c.key for c in PricingSku.__table__.columns]
    all_fields = (base_cols + _PRICING_PROMO_FIELDS + _PRICING_REPORT_FIELDS
                  + _PRICING_DISCOUNT_FIELDS + _PRICING_COST_FIELDS)
    IMG_COL = 1
    FIRST_DATA_COL = 2
    field_pos = {f: FIRST_DATA_COL + i for i, f in enumerate(all_fields)}

    # 数据: 按 产品编码 → SKU 分组 (剔除作废/服务/占位类非产品 SKU)
    skus = db.execute(
        select(PricingSku).order_by(PricingSku.product_code, PricingSku.sku_code)
    ).scalars().all()
    prod_by_code = {p.code: p for p in db.execute(select(Product)).scalars().all()}
    promo_by_sku = {p.sku_code: p for p in db.execute(select(PricingSkuPromo)).scalars().all()}
    costs_by_sku = {c.sku_code: c for c in db.execute(select(PricingSkuCosts)).scalars().all()}

    groups: list[tuple[str, list]] = []
    cur = None
    for s in skus:
        prod = prod_by_code.get(s.product_code)
        nm = (prod.name if prod else None) or s.product_name or ""
        if not _is_real_product(nm, s.product_code):
            continue
        if s.product_code != cur:
            groups.append((s.product_code, []))
            cur = s.product_code
        groups[-1][1].append(s)
    groups = [g for g in groups if g[1]]

    # 每产品兜底图 URL (本地图库优先由 product_image_map 内部处理)
    url_by_code = {
        c: ((prod_by_code.get(c).image_url if prod_by_code.get(c) else None)
            or next((x.image_url for x in gs if x.image_url), None))
        for c, gs in groups
    }
    img_map = product_image_map([c for c, _ in groups], url_by_code)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "定价图册"

    # 数字格式
    model_cols = {c.key: c for c in PricingSku.__table__.columns}
    promo_cols = {c.key: c for c in PricingSkuPromo.__table__.columns}
    cost_cols = {c.key: c for c in PricingSkuCosts.__table__.columns}
    col_fmt: dict[int, str] = {}
    for f, ci in field_pos.items():
        col = model_cols.get(f)
        if col is None:
            col = promo_cols.get(f)
        if col is None:
            col = cost_cols.get(f)
        if col is not None:
            fmt = _num_fmt(f, _col_type(col))
            if fmt:
                col_fmt[ci] = fmt
    for f in _PRICING_REPORT_FIELDS:                       # 报名价(派生, 非模型列)→ 金额格式
        if f in field_pos:
            col_fmt[field_pos[f]] = "#,##0.00"
    for f in _PRICING_DISCOUNT_FIELDS:                     # 单品立减(派生): 折 2位 / 降价金额 金额
        if f in field_pos:
            col_fmt[field_pos[f]] = "0.00" if f.endswith("_zhe") else "#,##0.00"

    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    white = Font(bold=True, color="FFFFFF", size=11)
    band_font = Font(bold=True, color="FFFFFF", size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    vcenter = Alignment(vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # 行1: 产品图表头 (A1:A2 合并, 深色)
    ws.merge_cells(start_row=1, start_column=IMG_COL, end_row=2, end_column=IMG_COL)
    a = ws.cell(1, IMG_COL, "产品图")
    a.fill = PatternFill("solid", fgColor="0F172A"); a.font = white; a.alignment = center; a.border = border

    # 行1: 分类色带 (按列序合并连续同类, 兼容任意列顺序)
    cat_of = {f: n for n, _c, fs in _PRICING_CATEGORIES for f in fs}
    color_of = {n: c for n, c, _fs in _PRICING_CATEGORIES}
    col_cat = [(field_pos[f], cat_of.get(f)) for f in all_fields]   # 已按列序
    j = 0
    while j < len(col_cat):
        pos, cat = col_cat[j]
        k = j
        while k + 1 < len(col_cat) and col_cat[k + 1][1] == cat and col_cat[k + 1][0] == col_cat[k][0] + 1:
            k += 1
        c0, c1 = col_cat[j][0], col_cat[k][0]
        if cat:
            if c1 > c0:
                ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
            bc = ws.cell(1, c0, cat)
            bc.fill = PatternFill("solid", fgColor=color_of[cat]); bc.font = band_font; bc.alignment = center
        for ci in range(c0, c1 + 1):
            ws.cell(1, ci).border = border
        j = k + 1

    # 行2: 中文表头 (按分类上色)
    for f in all_fields:
        ci = field_pos[f]
        cc = ws.cell(2, ci, _PRICING_CN.get(f) or _cn_header("pricing_sku", f))
        cc.fill = PatternFill("solid", fgColor=_PRICING_FIELD_COLOR.get(f, "334155"))
        cc.font = white; cc.alignment = center; cc.border = border

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28

    # 数据行 + 产品图
    from app.services import pricing_calc_service
    promo_params = pricing_calc_service.get_promo_params(db)   # 平台立减/佣金口径 (回填空系数用)
    gray = Font(color="94A3B8")
    keep_alive: list = []
    override_by_row: dict[int, bool] = {}
    r = 3
    for code, gs in groups:
        r0 = r
        for s in gs:
            p = promo_by_sku.get(s.sku_code)
            costs = costs_by_sku.get(s.sku_code)
            rp = pricing_calc_service.report_prices(p, promo_params) if p is not None else {}
            sid = (pricing_calc_service.single_item_discounts(p, s.daily_price, promo_params)
                   if p is not None else {})
            disc_vals = {"mid_disc_zhe": sid.get("mid_discount"), "mid_disc_amt": sid.get("mid_deduct"),
                         "big_disc_zhe": sid.get("big_discount"), "big_disc_amt": sid.get("big_deduct"),
                         "big618_disc_zhe": sid.get("big618_discount"), "big618_disc_amt": sid.get("big618_deduct")}
            override_by_row[r] = bool(getattr(s, "factory_cost_override", False))
            for f in all_fields:
                ci = field_pos[f]
                if f in _PRICING_REPORT_FIELDS:            # 报名价(派生) = 大促到手 ÷ 0.88 / ÷ 0.85
                    rv = rp.get(f)
                    v = float(rv) if rv is not None else None
                elif f in _PRICING_DISCOUNT_FIELDS:        # 单品立减(派生): 折(×10) / 降价金额(元)
                    dv = disc_vals.get(f)
                    v = (round(float(dv) * 10, 2) if f.endswith("_zhe") else float(dv)) if dv is not None else None
                elif f in model_cols:
                    v = _translate(f, _cell(getattr(s, f, None)))
                elif f in promo_cols:
                    v = _cell(getattr(p, f, None)) if p is not None else None
                else:                                      # 配件成本明细 (PricingSkuCosts)
                    v = _cell(getattr(costs, f, None)) if costs is not None else None
                cell = ws.cell(r, ci, v)
                cell.border = border
                if ci in col_fmt:
                    cell.number_format = col_fmt[ci]
                    cell.alignment = right
                else:
                    cell.alignment = vcenter
            # 部分促销系数在库里未持久化(空) → 回填当前口径值, 让派生列都能做活公式且口径正确
            if p is not None:
                def _bf(field, val):
                    bci = field_pos.get(field)
                    if bci and getattr(p, field, None) is None and val is not None:
                        ws.cell(r, bci, float(val))
                # 单品立减系数由公式反推 (店内到手÷日常), 不再回填隐含值; 但补上反推分母要用的立减/佣金
                if p.mid_shop_rate is not None or p.mid_buyer_price is not None:   # 参与中促 → 补平台立减/佣金
                    _bf("mid_platform_discount", promo_params.get("mid_platform_discount"))
                    _bf("mid_vip_commission", promo_params.get("mid_vip_commission"))
                if p.big_shop_rate is not None or p.big_buyer_price is not None:   # 参与大促
                    _bf("mid_platform_discount", promo_params.get("mid_platform_discount"))
                    _bf("mid_vip_commission", promo_params.get("mid_vip_commission"))
                if p.big_shop_rate is not None:       # 参与大促
                    _bf("big_platform_discount", promo_params.get("big_platform_discount"))
                    _bf("big_vip_commission", promo_params.get("big_vip_commission"))
                if p.xhs_activity_price is not None:  # 上小红书 → 默认折扣 15%
                    _bf("xhs_promo_discount", 0.15)
            ws.row_dimensions[r].height = _CATALOG_IMG_ROW_PT if r == r0 else _CATALOG_DATA_ROW_PT
            r += 1
        r1 = r - 1
        if r1 > r0:
            ws.merge_cells(start_row=r0, start_column=IMG_COL, end_row=r1, end_column=IMG_COL)
        acell = ws.cell(r0, IMG_COL)
        acell.border = border
        png = img_map.get(code)
        if png:
            acell.alignment = Alignment(horizontal="center", vertical="top")
            bio = _io.BytesIO(png)
            img = XLImage(bio)
            scale = min(_CATALOG_IMG_PX / img.width, _CATALOG_IMG_PX / img.height, 1.0)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            ws.add_image(img, f"{get_column_letter(IMG_COL)}{r0}")
            keep_alive.append(bio)          # BytesIO 需存活到 wb.save()
        else:
            acell.value = "暂无图片"; acell.font = gray
            acell.alignment = Alignment(horizontal="center", vertical="center")
        for rr in range(r0, r1 + 1):        # 图列每行都描边
            ws.cell(rr, IMG_COL).border = border

    # 列宽 + 冻结 (锁产品图列 + 两行表头)
    ws.column_dimensions[get_column_letter(IMG_COL)].width = 15
    for f in all_fields:
        ws.column_dimensions[get_column_letter(field_pos[f])].width = _CATALOG_WIDE.get(f, 12)
    ws.freeze_panes = "B3"

    # ── 全链活公式 (改任一配件/成本/系数 → 外配件/工厂/物理/价格/利润/平台价 Excel 内全联动) ──
    # 1) 配件明细 → 外配件(SUM) → 工厂成本(木作+包装+外配件); override 行保留手改工厂值。先写, 让物理链认它。
    acc_cols = [field_pos[f] for f in _PRICING_COST_NUM_FIELDS if f in field_pos]
    ext_i = field_pos.get("external_parts_cost")
    fac_i = field_pos.get("factory_cost")
    wood_i = field_pos.get("wood_cost")
    pack_i = field_pos.get("packaging_cost")

    def _present(rr, ci):
        return bool(ci) and ws.cell(rr, ci).value not in (None, "")

    for rr in range(3, ws.max_row + 1):
        if ext_i and acc_cols and any(_present(rr, c) for c in acc_cols):
            refs = ",".join(f"{get_column_letter(c)}{rr}" for c in acc_cols)
            ws.cell(rr, ext_i).value = f"=SUM({refs})"
        if (fac_i and wood_i and pack_i and ext_i and not override_by_row.get(rr)
                and any(_present(rr, c) for c in (wood_i, pack_i, ext_i))):
            ws.cell(rr, fac_i).value = (f"={get_column_letter(wood_i)}{rr}+"
                                        f"{get_column_letter(pack_i)}{rr}+{get_column_letter(ext_i)}{rr}")
    # 2) 物理成本 → 各档价 → 会计/利润/毛利率 (图册首列是产品图→字段右移1列; 数据从第3行)
    _apply_pricing_formulas(ws, col_offset=1, data_start_row=3)
    # 3) 平台活动价 (淘宝/店内/中促/大促/小红书) 派生列 → 活公式 (含 88VIP 消费券阶梯)
    _apply_promo_formulas(ws, field_pos, data_start_row=3,
                          mid_tiers=promo_params.get("mid_coupon_tiers"),
                          big_tiers=promo_params.get("big_coupon_tiers"))

    out = _io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── 活动报名表 (2026-07-06 用户: 给同事填淘宝活动价用的精简表) ────────────────────────────
def build_signup_form_xlsx(db: Session):
    """活动报名表 (Excel, 带产品图): 一 SKU 一行, 只留填淘宝活动【必要】列 ——
      产品图 / 产品名 / 规格 + 一口价 / 日常价(活动价) + 各档目标到手 +
      报名价(88VIP大促 / 超大促618双11) + 单品立减(折 + 立减金额, 三档场次力度 10/12/15%)。
    去掉 ID / 淘宝标题 / SKU编码 / 产品编码 / 大小 / 成品尺寸 / 小促价 / 成本全块 / 小红书 / 配件明细 /
    旧乘法系数 等无关列 (用户 2026-07-06 指定)。数值口径 = single_item_discounts + report_prices (加法, 对齐淘宝)。"""
    import io as _io
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from app.models.pricing import PricingSku
    from app.models.pricing_ext import PricingSkuPromo
    from app.models.product import Product
    from app.services import pricing_calc_service
    from app.services.pricing_catalog_service import _is_real_product, product_image_map

    def _f(v):
        return float(v) if v is not None else None

    def _zhe(v):                          # 折扣小数 → 折 (0.792 → 7.92)
        return round(float(v) * 10, 2) if v is not None else None

    params = pricing_calc_service.get_promo_params(db)
    skus = db.execute(
        select(PricingSku).order_by(PricingSku.product_code, PricingSku.sku_code)).scalars().all()
    prod_by_code = {p.code: p for p in db.execute(select(Product)).scalars().all()}
    promo_by_sku = {p.sku_code: p for p in db.execute(select(PricingSkuPromo)).scalars().all()}

    # 分组(剔除作废/服务/占位非产品), 供产品图纵向合并
    groups: list[tuple[str, list]] = []
    cur = None
    for s in skus:
        prod = prod_by_code.get(s.product_code)
        nm = (prod.name if prod else None) or s.product_name or ""
        if not _is_real_product(nm, s.product_code):
            continue
        if s.product_code != cur:
            groups.append((s.product_code, [])); cur = s.product_code
        groups[-1][1].append(s)
    groups = [g for g in groups if g[1]]
    url_by_code = {
        c: ((prod_by_code.get(c).image_url if prod_by_code.get(c) else None)
            or next((x.image_url for x in gs if x.image_url), None))
        for c, gs in groups}
    img_map = product_image_map([c for c, _ in groups], url_by_code)

    headers = ["产品图", "产品名称", "规格", "一口价", "日常价(活动价)",
               "中促到手", "大促到手", "88VIP大促报名价", "超大促报名价(618/双11)",
               "中促降价金额(元)", "大促降价金额(元)", "超大促降价金额(元)"]
    money = "#,##0.00"
    money_cols = {4, 5, 6, 7, 8, 9, 10, 11, 12}     # 价格 + 降价金额列 (单品立减全用减金额, 不再出折)
    zhe_cols: set[int] = set()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "活动报名表"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    vcenter = Alignment(vertical="center")

    # 行1: 口径说明
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tip = ws.cell(1, 1,
                  "淘宝加法口径: 到手 = 活动价 − 官方立减 − 单品立减 (官方立减 日常10%/88VIP大促12%/618双11 15%, 平台自动扣)。"
                  "★两种填法【二选一, 千万别混】: "
                  "【① 常规·推荐】活动价填『日常价』, 再填『单品立减(各档降价金额, 元)』→ 到手=各档「到手」列。 "
                  "【② 报名价】活动价直接填『报名价』(这数已把官方立减算进去了), 单品立减就填0/不叠, 主要给618换SKU用。 "
                  "⚠️绝不能『活动价填报名价』又『叠单品立减』—— 那是打两次折, 价格会砸穿(到手远低于目标)!")
    tip.font = Font(bold=True, color="B45309")
    tip.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 62

    # 行2: 表头
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(2, ci, h)
        c.fill = head_fill; c.font = head_font; c.alignment = center; c.border = border
    ws.row_dimensions[2].height = 34

    keep_alive: list = []
    r = 3
    for code, gs in groups:
        r0 = r
        prod = prod_by_code.get(code)
        pname = (prod.name if prod else None) or (gs[0].product_name if gs else None) or code
        for s in gs:
            promo = promo_by_sku.get(s.sku_code)
            rp = pricing_calc_service.report_prices(promo, params) if promo is not None else {}
            sid = (pricing_calc_service.single_item_discounts(promo, s.daily_price, params)
                   if promo is not None else {})
            row_vals = [
                None, pname, s.sku,
                _f(s.list_price), _f(s.daily_price),
                _f(getattr(promo, "mid_buyer_price", None)), _f(getattr(promo, "big_buyer_price", None)),
                _f(rp.get("report_price")), _f(rp.get("report_price_618")),
                _f(sid.get("mid_deduct")), _f(sid.get("big_deduct")), _f(sid.get("big618_deduct")),
            ]
            for ci, v in enumerate(row_vals, start=1):
                cell = ws.cell(r, ci, v)
                cell.border = border
                if ci in money_cols and v is not None:
                    cell.number_format = money; cell.alignment = right
                elif ci in zhe_cols:
                    cell.alignment = right
                else:
                    cell.alignment = vcenter
            ws.row_dimensions[r].height = _CATALOG_IMG_ROW_PT if r == r0 else _CATALOG_DATA_ROW_PT
            r += 1
        r1 = r - 1
        if r1 > r0:
            ws.merge_cells(start_row=r0, start_column=1, end_row=r1, end_column=1)
        acell = ws.cell(r0, 1)
        acell.border = border
        png = img_map.get(code)
        if png:
            acell.alignment = Alignment(horizontal="center", vertical="top")
            bio = _io.BytesIO(png)
            img = XLImage(bio)
            scale = min(_CATALOG_IMG_PX / img.width, _CATALOG_IMG_PX / img.height, 1.0)
            img.width = int(img.width * scale); img.height = int(img.height * scale)
            ws.add_image(img, f"A{r0}")
            keep_alive.append(bio)
        else:
            acell.value = "暂无图片"; acell.font = Font(color="94A3B8")
            acell.alignment = Alignment(horizontal="center", vertical="center")
        for rr in range(r0, r1 + 1):
            ws.cell(rr, 1).border = border

    widths = [15, 24, 16, 11, 13, 11, 11, 15, 17, 15, 15, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"

    out = _io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── 淘宝「单品立减」批量上传表 (2026-07-06 用户: 各档力度一张, 可直接上传淘宝) ──────────────
# 表头与淘宝官方「单品立减批量模板」逐字一致 (含换行), 上传器按列位解析 → 生成的表可直接导入。
_TB_DISCOUNT_HEADERS = [
    "商品id\n(请使用文本格式)",
    "SKU_ID\n(活动为SKU级别必填，商品级别不填,一行只能填写一个sku)",
    ("优惠值：\n1.优惠方式为减钱时，填写该商品立减金额，单位元，如20（元）\n"
     "2.优惠方式为打折时，填写该商品折扣，如9（折）\n注意：sku级别的活动不支持打折！"),
    ("打折对应的优惠金额取值方式：\n不填：不抹分取整\n1：抹分（即向上取整到角,例如18.43变为18.5）请填写数字1\n"
     "2：取整（即向上取整到元,例如18.43变为19）请填写数字2\n注意：\n"
     "1.仅打折商品级活动支持优惠金额的抹分、取整\n2.打折是在标价的基础上打折，折扣实际对应的减钱金额会根据标价的变动而变动"),
    "提醒:\n（1）填写表格时，请删除示例数据；\n（2）单元格不支持使用公式；",
]
# tier → (人看的档位名, single_item_discounts 里对应的【立减金额】字段)
_TB_DISCOUNT_TIERS = {
    "mid":    ("超级立减10%",          "mid_deduct"),
    "big":    ("88VIP大促12%",         "big_deduct"),
    "big618": ("大促15%(618双11)",     "big618_deduct"),
}


def build_single_item_discount_upload_xlsx(db: Session, tier: str):
    """淘宝『单品立减』批量上传表 (SKU 级别, **减钱口径** —— 模板明确注: sku级别不支持打折, 故填立减金额)。
    列 = 商品id / SKU_ID / 优惠值(=立减金额, 元) / 取值方式(留空) / 提醒(留空); 表头与淘宝模板逐字一致, 可直接上传。
    tier: mid(超级立减10%) / big(88VIP大促12%) / big618(大促15% 618双11)。
    只出「有淘宝商品id + SKU_ID + 该档立减金额」的行(缺 SKU_ID / 官方立减已够的跳过)。
    数值 = single_item_discounts 加法口径, **每次下载实时算**(成本/售价一变即变)。返回 (BytesIO, 统计dict)。"""
    import io as _io
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from app.models.pricing import PricingSku
    from app.models.pricing_ext import PricingSkuPromo
    from app.services import pricing_calc_service

    if tier not in _TB_DISCOUNT_TIERS:
        raise ValueError(f"未知档位 {tier}; 可选 {list(_TB_DISCOUNT_TIERS)}")
    _tier_name, deduct_field = _TB_DISCOUNT_TIERS[tier]
    params = pricing_calc_service.get_promo_params(db)
    skus = db.execute(
        select(PricingSku).order_by(PricingSku.product_code, PricingSku.sku_code)).scalars().all()
    promo_by_sku = {p.sku_code: p for p in db.execute(select(PricingSkuPromo)).scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "单品立减"
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for ci, h in enumerate(_TB_DISCOUNT_HEADERS, start=1):
        c = ws.cell(1, ci, h)
        c.fill = head_fill; c.font = head_font; c.alignment = wrap
    ws.row_dimensions[1].height = 78

    stats = {"tier": tier, "rows": 0, "skipped_no_skuid": 0, "skipped_no_deduct": 0}
    r = 2
    for s in skus:
        p = promo_by_sku.get(s.sku_code)
        if p is None or not p.taobao_item_id:
            continue
        if not p.taobao_sku_id:                       # SKU 级别必须有 SKU_ID
            stats["skipped_no_skuid"] += 1
            continue
        sid = pricing_calc_service.single_item_discounts(p, s.daily_price, params)
        d = sid.get(deduct_field)
        if d is None:                                 # 官方立减已够 / 缺买家价 → 跳过
            stats["skipped_no_deduct"] += 1
            continue
        ws.cell(r, 1, str(p.taobao_item_id)).number_format = "@"   # 长号必须文本, 防科学计数
        ws.cell(r, 2, str(p.taobao_sku_id)).number_format = "@"
        ws.cell(r, 3, float(d)).number_format = "0.00"             # 立减金额(元)
        r += 1
    stats["rows"] = r - 2

    widths = [22, 30, 30, 26, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out = _io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, stats


def build_full_export_workbook(db: Session):
    """全类目工作簿: 产品总表(按SKU展开+公式) 置顶, 定价总表次之, 其余每类目一 Sheet。"""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    for key, cfg in ENTITY_MODELS.items():
        if key in ("product", "pricing_sku"):
            continue                      # 产品总表/定价总表 特殊处理 (专用 builder)
        _build_entity_sheet(db, wb, key, cfg, used)

    # 定价总表: 专用 builder (补全平台活动价 + 中文表头 + 分类配色), 再改派生列为活公式
    ws_p = _build_pricing_sheet(db, wb, used)
    pricing_name: Optional[str] = ws_p.title
    _apply_pricing_formulas(ws_p)         # 定价总表派生列改活公式(改成本→价格/利润联动)

    _build_product_sku_sheet(db, wb, used, pricing_name)

    # 排序: 产品总表 第一, 定价总表 第二
    order_titles = [t for t in ("产品总表", pricing_name) if t and t in wb.sheetnames]
    for i, t in enumerate(order_titles):
        sh = wb[t]
        wb._sheets.remove(sh)
        wb._sheets.insert(i, sh)

    if not wb.sheetnames:
        ws = wb.create_sheet("空")
        ws.append(["系统当前没有可导出的类目。"])
    return wb


def rotate_full_exports(db: Session, *, keep: int = MAX_KEEP) -> int:
    """资料存档库里 kind=full_export 只留最新 keep 份, 超出删最早(连磁盘文件)。返回删除数。"""
    recs = db.execute(
        select(ImportedFile).where(ImportedFile.kind == "full_export")
        .order_by(ImportedFile.id.desc())
    ).scalars().all()
    removed = 0
    for old in recs[keep:]:
        if import_storage.delete_record(db, old.id):
            removed += 1
    return removed


def run_full_export(db: Session, *, uploaded_by: Optional[str] = None) -> dict:
    """生成全类目 Excel → 存「资料存档库」→ 轮转(留≤30)。返回 {content, filename, ...}。"""
    wb = build_full_export_workbook(db)
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    today = date.today()
    filename = f"全量导出_{today.isoformat()}.xlsx"
    res = import_storage.archive(
        db, content=content, original_name=filename, kind="full_export",
        source="web", uploaded_by=uploaded_by,
        row_summary={"sheets": len(wb.sheetnames), "exported_at": today.isoformat()},
    )
    removed = rotate_full_exports(db)
    db.commit()
    return {
        "content": content, "filename": filename,
        "file_id": res.file.id, "sheets": len(wb.sheetnames), "rotated_removed": removed,
    }

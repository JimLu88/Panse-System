# -*- coding: utf-8 -*-
"""配件(外采)对账服务 — 配件 epic (用户 2026-06-26; 方向1: 分类 + BOM 驱动)。

三件事:
  A. aggregate_related_purchases —— 填了 related_order_no 的配件采购单按订单汇总写 Order.actual_parts。
  B. bulk_material_recon —— **按 Material.category 分组、BOM 驱动** 的对账: 每分类每发货月
     历史平均 | 预估(Σ发货单BOM里该类外采配件 price×qty) | 实际(工厂月度对账总额) | 差异%。
     **铁律: 消费窗口按订单「发货日期 ship_date」圈定(生产周期~30天)。** 取代旧硬编码关键词登记表。
  C. 工厂月度对账总额录入(PartsMonthlyRecon, material_key 存「分类」)。
  D. export_shipped_orders —— 导当月发货单清单(全部 / 按分类逐单展开 BOM 部位+预设尺寸)给工厂对账。

口径: 总账准(差异落当期); 逐单这块是有界/可校准的近似(物理限制绕不开)。
"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import date
from decimal import Decimal
from typing import Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.material import Material
from app.models.order import Order, PartPurchase, PartsMonthlyRecon
from app.services import sales_analytics

_CENTS = Decimal("0.01")

# 工厂自备/木作/木材/人工(WD-/MW-/MP-)在 BOM 里是工厂自备 → 不进 BOM 驱动的配件「标准消耗」对账。
_FACTORY_PREFIXES = ("WD", "MW", "MP")
# 但「支付宝自购采购」侧只排纯工厂服务(木作 WD / 人工 MP); 自购木材(MW, 如榉木自采、工厂未含)与
# 特殊件按用户口径(2026-06-29)当零星采购照常记入(不在 BOM 的会标 not_in_bom 待逐单核对)。
_PURCHASE_EXCLUDE_PREFIXES = ("WD", "MP")
_UNCATEGORIZED = "未分类"

# 结算模式 (用户 2026-06-27): 五金都在五金店买 → 月结(月度总额对账, 工厂/供应商填);
# 其余(杂项/洞石饰面板/木皮/大宗料…)= 零星采购(用了才买, 支付宝备注/导入逐项归账, 实际=Σ真实采购)。
# 月结类才"导清单给对方填总额"; 零星类没"对方", 实际从采购单(PartPurchase)来。
# 玻璃 2026-06-27 改月结(用户: 玻璃也有固定供应商按月对账)。
_MONTHLY_SETTLE_CATEGORIES = ("五金", "电力轨道", "岩板", "玻璃")   # 有固定供应商、按月对账的; 杂项/洞石饰面板/木皮等=零星


def _settle_mode(category: str) -> str:
    return "月结" if category in _MONTHLY_SETTLE_CATEGORIES else "零星"

# 非配件采购(代扣/理财/服务费/淘天扣款…)排除关键词 — 与 purchases.list_purchases 一致。
_NON_PART_KW = ("代扣", "代付", "资金扣回", "消费券", "理财", "申购",
                "服务费", "手续费", "余额宝", "转入", "转出", "单次转", "转账")


def _d(v) -> Decimal:
    return Decimal(str(v)) if v is not None else Decimal("0")


def _ym(d: Optional[date]) -> Optional[str]:
    return f"{d.year:04d}-{d.month:02d}" if d else None


def _size_area(size: Optional[str]) -> Decimal:
    """从尺寸备注取前两个数相乘当"面积"(定制单同料合并时取最大者偏保守)。无数→0。"""
    if not size:
        return Decimal("0")
    nums = re.findall(r"\d+(?:\.\d+)?", size)
    if len(nums) >= 2:
        return Decimal(nums[0]) * Decimal(nums[1])
    return Decimal(nums[0]) if nums else Decimal("0")


def _looks_non_part(p: PartPurchase) -> bool:
    name = (p.material_name or "")
    if any(k in name for k in _NON_PART_KW):
        return True
    if p.supplier and "淘天" in p.supplier:
        return True
    # 纯工厂服务(木作 WD / 人工 MP)= 工厂账单覆盖, 不是自购配件 → 不进 actual_parts(防货款双算);
    # 自购木材(MW, 榉木自采)/特殊件按用户口径当零星照常记(用户 2026-06-29), 不在此排除。
    if (p.material_code or "").split("-", 1)[0].upper() in _PURCHASE_EXCLUDE_PREFIXES:
        return True
    return False


def _purchase_amount(p: PartPurchase) -> Optional[Decimal]:
    if p.total_amount is not None:
        return _d(p.total_amount)
    if p.amount is not None:
        return _d(p.amount)
    return None


# ── 零星采购 → 订单(多单按 BOM 占比分摊) 的共用拆分逻辑 ───────────────────────
def _split_order_nos(s: Optional[str]) -> list[str]:
    """related_order_no 拆成订单号 token 列表(一笔零星采购可对应多个平台订单号, \n/空格/逗号分隔)。

    按 token 精确比对 Order.order_no(不抽数字), 故 26-28 位支付宝商户号会作单 token 比对、对不上→
    自然落 unmatched, 不会误抠出假订单号; 同时兼容非纯数字订单号(测试/历史)。去重保序。
    """
    if not s:
        return []
    seen: set = set()
    out: list[str] = []
    for tok in re.split(r"[\s,;]+", str(s).strip()):
        tok = tok.strip()
        if tok and tok not in seen:
            seen.add(tok)
            out.append(tok)
    return out


def _purchase_category(p: PartPurchase, mat_info: dict) -> str:
    """该采购的配件分类: 料号→分类优先; 无料号/无分类则按物料名关键词推断。"""
    if p.material_code:
        cat = (mat_info.get(p.material_code) or {}).get("category")
        if cat:
            return cat
    from app.services.material_category_service import _ac_category
    return _ac_category(p.material_name or "") or _UNCATEGORIZED


def _resolve_purchase_cat(p: PartPurchase, cons_o: dict, mat_info: dict) -> tuple[str, bool]:
    """把一笔采购在某订单上锚定到该单 BOM 的料/分类, 返回 (分类, 是否在该单BOM内)。

    防分类错配致双算(用户红线): 真实采购必须覆盖该单 BOM 里"对应那一类"的预估, 而非凭独立推断
    新增一类(否则 real(X)+est(Y) 同一笔配件钱算两次)。
      1) 料号命中该单 BOM 某料 → 用该料在 BOM 的分类(保证 real 覆盖对应 est);
      2) 否则关键词分类若也在该单 BOM 分类集合里 → 用之(可覆盖);
      3) 都不在 → (关键词分类, False): 该料不在该单 BOM, 标记待人工确认(不静默与其它类 est 叠加误读)。
    """
    if p.material_code:
        for cat, c in cons_o.items():
            for m in c.get("materials", []):
                if m.get("material_code") == p.material_code:
                    return cat, True
    pc = _purchase_category(p, mat_info)
    return (pc, True) if pc in cons_o else (pc, False)


def _monthly_supplier_names(db: Session) -> set:
    """payment_terms 含「月结」的供应商名(用于把零星采购里偶发的月结供应商排除出"零星覆盖")。"""
    from app.models.supplier import Supplier
    out: set = set()
    for name, terms in db.execute(select(Supplier.name, Supplier.payment_terms)).all():
        if name and terms and "月结" in terms:
            out.add(name)
    return out


def _category_usage(cons_cat: Optional[dict]) -> Decimal:
    """该订单某分类的"用量"权重(多单分摊用): 面积料(尺寸可解析)按面积, 计数料按数量, ×该行件数。

    实现用户口径「按各单该配件 BOM 用量(面积/数量)占比分摊」: 同一笔采购的同种料跨单, 面积大的单分得多。
    """
    if not cons_cat:
        return Decimal("0")
    total = Decimal("0")
    for m in cons_cat.get("materials", []):
        area = _size_area(m.get("size_note"))
        unit = area if area > 0 else Decimal("1")
        total += unit * _d(m.get("qty"))
    return total


def _allocate_purchases(db: Session, mat_info, bom_by_pcsku, bom_by_pc):
    """把所有有效配件采购单按平台订单号拆分(多单按该分类 BOM 用量占比分摊金额), 返回:
      order_real     {order_no: {category: Decimal}}   每单每类的真实采购成本(已分摊, 含月结+零星)
      order_objs     {order_no: Order}
      sporadic_cover {(order_no, category): [evidence...]}  非月结供应商的零星覆盖
                     (供月结导出/对账扣除 + 红字提示; evidence 含 流水号/金额/供应商)
      unmatched      [平台订单号都对不上真实订单的采购]
    """
    monthly = _monthly_supplier_names(db)
    rows = db.execute(
        select(PartPurchase).where(
            PartPurchase.related_order_no.isnot(None),
            PartPurchase.related_order_no != "",
        )
    ).scalars().all()

    order_objs: dict[str, Order] = {}

    def _order(no: str):
        if no not in order_objs:
            order_objs[no] = db.execute(
                select(Order).where(Order.order_no == no)).scalar_one_or_none()
        return order_objs[no]

    cons_cache: dict[str, dict] = {}

    def _cons(o: Order):
        if o.order_no not in cons_cache:
            cons_cache[o.order_no] = _order_category_consumption(o, mat_info, bom_by_pcsku, bom_by_pc)
        return cons_cache[o.order_no]

    order_real: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    sporadic_cover: dict[tuple, list] = defaultdict(list)
    unmatched: list[dict] = []

    for p in rows:
        if _looks_non_part(p):
            continue
        amt = _purchase_amount(p)
        if amt is None or amt <= 0:
            continue
        amt = amt.quantize(_CENTS)
        valid = [(no, _order(no)) for no in _split_order_nos(p.related_order_no)]
        valid = [(no, o) for no, o in valid if o is not None]
        if not valid:
            unmatched.append({"purchase_no": p.purchase_no, "related_order_no": p.related_order_no,
                              "amount": float(amt), "material_name": p.material_name,
                              "supplier": p.supplier})
            continue
        is_monthly = bool(p.supplier and p.supplier in monthly)
        # 逐单把采购锚定到该单 BOM 的料/分类(防分类错配致双算), 并取该类用量作分摊权重
        per: list[tuple] = []   # (order, cat, in_bom, usage)
        for no, o in valid:
            cons_o = _cons(o)
            cat, in_bom = _resolve_purchase_cat(p, cons_o, mat_info)
            per.append((o, cat, in_bom, _category_usage(cons_o.get(cat))))
        # 分摊金额: 单单全额; 多单按用量占比(无用量→均分), 余额给用量最大的单, 保证 Σ=amt 守恒。
        n = len(per)
        if n == 1:
            shares = [amt]
        else:
            weights = [w for (_, _, _, w) in per]
            tot = sum(weights, Decimal("0"))
            if tot > 0:
                shares = [(amt * w / tot).quantize(_CENTS) for w in weights]
                lead = max(range(n), key=lambda i: weights[i])
            else:
                each = (amt / n).quantize(_CENTS)
                shares = [each] * n
                lead = 0
            shares[lead] += amt - sum(shares, Decimal("0"))   # 余额给主单, 守恒不漏分
        for (o, cat, in_bom, _w), sh in zip(per, shares):
            if sh <= 0:
                continue
            order_real[o.order_no][cat] += sh
            if not is_monthly:
                sporadic_cover[(o.order_no, cat)].append({
                    "purchase_no": p.purchase_no, "flow_no": p.alipay_flow_no,
                    "supplier": p.supplier, "amount": float(sh),
                    "material_name": p.material_name, "in_bom": in_bom,
                })
    return order_real, order_objs, sporadic_cover, unmatched


def _sporadic_covered(db: Session, mat_info, bom_by_pcsku, bom_by_pc) -> dict:
    """{(order_no, category): [evidence...]} —— 非月结(零星/现付)供应商已覆盖的(订单×分类)。"""
    return _allocate_purchases(db, mat_info, bom_by_pcsku, bom_by_pc)[2]


def _sporadic_note(cover: list) -> str:
    """红字提示文案: 含已付金额 + 支付宝流水号证据, 让月结供应商核对、勿重复计入。"""
    total = sum((_d(e.get("amount")) for e in cover), Decimal("0")).quantize(_CENTS)
    flows = [str(e.get("flow_no")) for e in cover if e.get("flow_no")]
    tail = (" 流水" + "/".join(flows[:3])) if flows else ""
    return f"查看是否为零星采购,非月结付款 — 已走支付宝现付 ¥{total}{tail},请勿计入月结"


# ── A. 逐单配件采购 → actual_parts 汇总 ──────────────────────────────────────
def aggregate_related_purchases(db: Session, *, apply: bool = False) -> dict:
    """填了 related_order_no 的配件采购单 → 按订单(多单按 BOM 占比分摊)写 Order.actual_parts。

    - 多单: 一笔零星采购对应多个平台订单号 → 按各单该分类 BOM 用量占比分摊金额(无 BOM 用量→均分)。
    - 逐类覆盖: actual_parts = Σ各分类[有真实零星采购→真实金额(覆盖该类预估), 否则→该类 BOM 预估],
      即只覆盖"对应的零星配件预估", 其它分类保留 BOM 预估(用户 2026-06-28 口径)。
    apply=False(默认): 只算预览, 不落库; 返回每单 physical_cost 变化 + 逐类明细供人工核对。
    apply=True: 写 actual_parts 并 commit(该单 physical_cost 转「逐项真实计价」)。
    """
    from app.services.order_financials import physical_cost, physical_cost_breakdown

    mat_info, bom_by_pcsku, bom_by_pc = _load_bom_and_materials(db)
    order_real, order_objs, sporadic_cover, unmatched = _allocate_purchases(
        db, mat_info, bom_by_pcsku, bom_by_pc)

    items: list[dict] = []
    skipped: list[dict] = []
    applied = 0
    for ono in sorted(order_real):
        o = order_objs.get(ono)
        real = order_real[ono]
        if o is None:
            items.append({"order_no": ono, "matched": False,
                          "parts_total": float(sum(real.values(), Decimal("0")))})
            continue
        # 安全门(用户 2026-06-29 抽查纠错): 只对【正常成交 + 非定制 + 非补单 + 成本未被封顶】的订单覆盖。
        # 取消/未付款=非成交; 定制单 BOM 是模板不可靠; 定金/片段/兜底单(实付远低于整件)用真实配件会虚高
        # (逐项真实计价分支去掉了实付×85%封顶)→ 一律跳过, 不动这些单的成本。
        bd = physical_cost_breakdown(o)   # 当前 actual_parts 为空 → 该单"自然"封顶状态
        reason = None
        if (o.status or "") in ("cancelled", "pending_payment"):
            reason = "非成交单(%s)" % o.status
        elif bool(getattr(o, "is_refill", False)):
            reason = "补单"
        elif bool(getattr(o, "is_custom", False)):
            reason = "定制单(BOM为模板)"
        elif bd.get("cap_mode") not in (None, "", "none"):
            reason = "成本被封顶(%s)" % bd.get("cap_mode")
        if reason:
            skipped.append({"order_no": ono, "reason": reason,
                            "real_parts": float(sum(real.values(), Decimal("0")))})
            continue
        cons = _order_category_consumption(o, mat_info, bom_by_pcsku, bom_by_pc)
        # 基线 = 该单原本的配件预估 est_parts(缺则退回 BOM 估总额)。只把"被真实采购覆盖的分类"换成真实值,
        # 其余配件仍按原预估 —— 不用 BOM 全类汇总(否则会比定价预估大很多 → 整单成本虚高)。
        base = _d(o.est_parts) if o.est_parts is not None else sum(
            (_d(c.get("amount")) for c in cons.values()), Decimal("0"))
        real_total = sum(real.values(), Decimal("0"))
        est_covered = sum((_d(cons[c].get("amount")) for c in real if c in cons), Decimal("0"))
        new_parts = (base - est_covered + real_total).quantize(_CENTS)
        if new_parts < 0:
            new_parts = real_total.quantize(_CENTS)
        cat_rows: list[dict] = []
        not_in_bom_n = 0
        for c in sorted(real):
            in_bom = c in cons
            if not in_bom:
                not_in_bom_n += 1
            cat_rows.append({
                "category": c, "real": float(real[c]),
                "est": float(_d(cons.get(c, {}).get("amount"))),
                "not_in_bom": (not in_bom),
                "evidence": sporadic_cover.get((ono, c), []),
            })
        old_parts = o.actual_parts
        old_phys = physical_cost(o)
        o.actual_parts = new_parts
        new_phys = physical_cost(o)
        # 健全性门: 归账是"用真实配件替换配件预估", 整单物理成本最多只该上升≈新增的真实配件额。
        # 若成本上升远超真实配件额(多因 wood_cost_est 与 theoretical_cost 数据不一致 → actual_parts 分支
        # 换算木作基线致整单虚高, 如 theo=2108 却 wood_est=5200)→ 还原并跳过, 不动该单成本, 列出供人工核对。
        if (new_phys - old_phys) > real_total + Decimal("100"):
            o.actual_parts = old_parts
            skipped.append({
                "order_no": ono,
                "reason": "成本异常上升%.0f(>真实配件%.0f, 疑 wood_est/理论成本不一致)" % (
                    float(new_phys - old_phys), float(real_total)),
                "real_parts": float(real_total),
            })
            continue
        if apply:
            applied += 1
        else:
            o.actual_parts = old_parts   # dry-run 还原
        items.append({
            "order_no": ono, "matched": True, "categories_count": len(real),
            "product_name": o.product_name, "is_custom": bool(o.is_custom),
            "est_parts_base": float(base), "real_total": float(real_total),
            "old_actual_parts": float(_d(old_parts)) if old_parts is not None else None,
            "new_actual_parts": float(new_parts),
            "old_physical_cost": float(old_phys), "new_physical_cost": float(new_phys),
            "physical_delta": float((new_phys - old_phys).quantize(_CENTS)),
            "not_in_bom_categories": not_in_bom_n,
            "categories": cat_rows,
        })

    if apply:
        db.commit()

    matched = [i for i in items if i.get("matched")]
    return {
        "applied": apply,
        "applied_count": applied,
        "matched_orders": len(matched),
        # 跳过的订单(取消/定制/补单/成本被封顶) — 不覆盖成本, 列出原因供人工核对
        "skipped_orders": len(skipped),
        "skipped": skipped,
        # unmatched = 平台订单号全部对不上真实订单的"采购单"数(非订单数)
        "unmatched_purchases_count": len(unmatched),
        "unmatched_purchases": unmatched,
        # 真实采购分类不在该单 BOM 内、需人工核对的订单数(防分类错配高估)
        "flagged_not_in_bom_orders": sum(1 for i in matched if i.get("not_in_bom_categories")),
        "total_parts_amount": float(sum((_d(i.get("new_actual_parts")) for i in matched), Decimal("0"))),
        "items": items,
    }


# ── BOM 驱动的物料消耗 (取代硬编码关键词; 用户 2026-06-26 方向1) ──────────────
def _settled_shipped_orders(db: Session) -> list[Order]:
    """成交 + 已发货(ship_date 非空) + 非补单 的订单(对账消费窗口基底, 按发货日期)。"""
    return db.execute(
        select(Order).where(
            sales_analytics.settled_sale_clause(),
            Order.is_refill.is_(False),
            Order.ship_date.isnot(None),
        )
    ).scalars().all()


def _product_code_variants(code: Optional[str]) -> list[str]:
    """订单 product_code 用 P+11, BOM 用 PPS+11 → 两形态都试(镜像 accessory_checklist)。"""
    if not code:
        return []
    code = code.strip()
    out = {code}
    if code.startswith("PPS"):
        out.add("P" + code[3:])
    elif code.startswith("P"):
        out.add("PPS" + code[1:])
    return list(out)


def _load_bom_and_materials(db: Session):
    """一次性载入: 物料(category/price/name/unit) + BOM(按 (product_code,sku_code) 与 product_code 两键)。"""
    from app.models.bom import BomLine
    mat_info: dict[str, dict] = {}
    for code, name, unit, price, cat in db.execute(
        select(Material.code, Material.name, Material.unit, Material.price, Material.category)
    ).all():
        mat_info[code] = {"name": name, "unit": unit, "price": _d(price), "category": cat}
    bom_by_pcsku: dict[tuple, list] = defaultdict(list)
    bom_by_pc: dict[str, list] = defaultdict(list)
    for pc, sc, mcode, qty, remark, mname in db.execute(
        select(BomLine.product_code, BomLine.sku_code, BomLine.material_code,
               BomLine.qty_per_product, BomLine.remark, BomLine.material_name)
    ).all():
        rec = (mcode, _d(qty), remark, mname)
        bom_by_pcsku[(pc, sc)].append(rec)
        bom_by_pc[pc].append(rec)
    return mat_info, bom_by_pcsku, bom_by_pc


def _order_bom_lines(o: Order, bom_by_pcsku, bom_by_pc) -> list:
    """该订单的 BOM 行(镜像 _bom_rows_for_order: 先 product_code+sku_code 精确, 再 product_code 回退)。"""
    pcs = _product_code_variants(o.product_code)
    for pc in pcs:
        if (pc, o.sku_code) in bom_by_pcsku:
            return bom_by_pcsku[(pc, o.sku_code)]
    for pc in pcs:
        if pc in bom_by_pc:
            return bom_by_pc[pc]
    return []


def _order_is_custom(o: Order) -> bool:
    if bool(getattr(o, "is_custom", False)):
        return True
    from app.services import sku_utils
    try:
        return bool(sku_utils.is_custom_sku_code(o.sku_code, o.product_code))
    except Exception:  # noqa: BLE001
        return False


_SIZE_RE = re.compile(r"\d+\.\d+")   # 尺寸(小数, 如 2.05/1.75); 子型号整数(U80/T25)保留


def _family_key(name: Optional[str]) -> str:
    """物料"族"键: 名字去掉尺寸(小数)后剩下的部分。同一物理件的不同尺寸变体(不同料号)归为一族,
    供定制单 BOM"同料只取一行"去重 —— 修电力轨道(AC-0162/63/64/65)/床铺板等尺寸变体绑成多个料号
    致对账预估虚高。保留子型号整数(U80/T25/K59)和"2插座", 避免误并不同件
    (2026-07-05 验证设置类目 20 族均为同件尺寸变体, 松木/榉木、U80/T25 均分开)。"""
    return _SIZE_RE.sub("", name or "").strip()


def _order_category_consumption(o: Order, mat_info, bom_by_pcsku, bom_by_pc) -> dict[str, dict]:
    """该订单按分类的外采配件消耗: {category: {amount, materials:[{...}]}}。排木作/工厂自备(WD/MW/MP);
    qty = qty_per_product × 订单件数。

    两步:
    ① 先按 (料号, 尺寸) 去掉 BOM 里重复列的完全相同行(有些产品 BOM 把同一尺寸重复列了 3-4 遍)。
    ② 是否再"按族合一行":
       - 定制单, 或 **BOM 走了产品级模板回退**(该 SKU 没有精确 BOM → 落到含【全部尺寸变体】的产品级
         模板) → 一单实际只用一种尺寸 → **同一物理件的尺寸变体合成一行**(面积最大者, 偏保守), 标
         `size_uncertain`+`alt_size_count` 提示人工确认。修电力轨道/床铺板等尺寸变体绑多料号致【预估虚高】。
       - SKU 精确命中 BOM(该 SKU 的真实用料) → 尊重, **不合并**(可能真用到多件/多尺寸)。
    只影响【预估(供应商应付)】侧, 不碰实际对账(工厂录入)与产品利润。
    """
    qmul = Decimal(int(o.qty or 1))
    is_custom = _order_is_custom(o)
    # BOM 来源判定: 该 SKU 有精确 BOM 吗? 没有 → 落产品级模板(含全部尺寸变体) → 一单只用一种 → 需族合并。
    pcs = _product_code_variants(o.product_code)
    sku_specific = any((pc, o.sku_code) in bom_by_pcsku for pc in pcs)
    use_family = is_custom or not sku_specific
    by_cat: dict[str, dict] = {}
    seen: set = set()                   # ① 精确去重 (mcode, size): 先合掉 BOM 重复列的相同行
    fam_pick: dict[tuple, dict] = {}    # ② 按族选中料(面积最大) + alt(被合并的其它尺寸数)

    def _emit(mcode, q, price, size, cat, name, info, alt=0):
        amt = (q * price).quantize(_CENTS)
        c = by_cat.setdefault(cat, {"amount": Decimal("0"), "materials": []})
        c["amount"] += amt
        m = {
            "material_code": mcode, "part_name": name, "category": cat,
            "qty": float(q), "unit": info.get("unit"), "price": float(price),
            "amount": float(amt), "size_note": size,
        }
        if alt > 0:
            m["size_uncertain"] = True
            m["alt_size_count"] = alt
        c["materials"].append(m)

    for mcode, qty, remark, mname in _order_bom_lines(o, bom_by_pcsku, bom_by_pc):
        if (mcode or "").split("-", 1)[0].upper() in _FACTORY_PREFIXES:
            continue
        info = mat_info.get(mcode) or {}
        cat = info.get("category") or _UNCATEGORIZED
        name = info.get("name") or mname or mcode
        size = (remark or "").strip() or None
        q = qty * qmul
        price = info.get("price", Decimal("0"))
        dk = (mcode, "".join(size.split()) if size else "")
        if dk in seen:                  # ① BOM 里重复列的同料同尺寸行 → 只算一次
            continue
        seen.add(dk)
        if not use_family:
            _emit(mcode, q, price, size, cat, name, info)
            continue
        # ② 按族(类目 + 名字去尺寸)合一行: 同一物理件的尺寸变体(常是不同料号)只留面积最大者
        area = _size_area(size)
        fk = (cat, _family_key(name))
        prev = fam_pick.get(fk)
        if prev is None:
            fam_pick[fk] = {"mcode": mcode, "q": q, "price": price, "size": size, "cat": cat,
                            "name": name, "info": info, "area": area, "alt": 0}
        else:
            prev["alt"] += 1
            if area > prev["area"]:
                fam_pick[fk] = {"mcode": mcode, "q": q, "price": price, "size": size, "cat": cat,
                                "name": name, "info": info, "area": area, "alt": prev["alt"]}

    for _fk, p in fam_pick.items():
        _emit(p["mcode"], p["q"], p["price"], p["size"], p["cat"], p["name"], p["info"], alt=p["alt"])
    return by_cat


def _ac_categories(db: Session) -> list[str]:
    """配件库里 AC(外采)物料用到的分类(去重, 排工厂自备前缀)。"""
    cats: set = set()
    for code, cat in db.execute(
        select(Material.code, Material.category).where(Material.category.isnot(None))
    ).all():
        if (code or "").split("-", 1)[0].upper() in _FACTORY_PREFIXES:
            continue
        if cat:
            cats.add(cat)
    return sorted(cats)


# ── B. 分类 × 发货月 对账 (BOM 驱动, 发货日期口径) ───────────────────────────
def bulk_material_recon(db: Session, *, granularity: str = "month") -> dict:
    """配件对账(方向1, 分类+BOM 驱动): 每分类每发货月 历史平均 | 预估 | 实际(工厂月度) | 差异%。

    - 预估 standard_consume = Σ(发货成交单 BOM 里该分类外采配件 price×qty)(按 ship_date 分月)。
    - 实际 factory_actual = 工厂月度对账总额(PartsMonthlyRecon, key=分类, 多供应商求和; 没录=None)。
    - 历史平均 historical_avg = 过去已对账月「每单实际」均值 × 本月发货单数(无历史→回退预估)。
    """
    mat_info, bom_by_pcsku, bom_by_pc = _load_bom_and_materials(db)
    orders = _settled_shipped_orders(db)
    sporadic_cover = _sporadic_covered(db, mat_info, bom_by_pcsku, bom_by_pc)

    std: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    cnt: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    spor: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    for o in orders:
        ym = _ym(o.ship_date)
        if ym is None:
            continue
        for cat, c in _order_category_consumption(o, mat_info, bom_by_pcsku, bom_by_pc).items():
            cover = sporadic_cover.get((o.order_no, cat))
            # 月结类里, 该单该类已走零星采购(支付宝现付)的部分 → 从月结预估「按金额净额」扣除(防月结多付),
            # 已现付额单列 sporadic_excluded; 剩余额(该类还有别的料没现付)仍按月结计入。零星类不受影响。
            if cover and _settle_mode(cat) == "月结":
                cover_total = sum((_d(e.get("amount")) for e in cover), Decimal("0"))
                spor[cat][ym] += cover_total
                remain = _d(c["amount"]) - cover_total
                if remain > 0:                       # 仅扣已现付, 剩余仍进月结预估(不丢整类)
                    std[cat][ym] += remain
                    cnt[cat][ym] += 1
                continue
            std[cat][ym] += c["amount"]
            cnt[cat][ym] += 1

    fact_all: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    for key, ym, amt in db.execute(
        select(PartsMonthlyRecon.material_key, PartsMonthlyRecon.year_month, PartsMonthlyRecon.actual_total)
    ).all():
        fact_all[key][ym] += _d(amt)

    # 零星类"实际" = 真实采购单(支付宝备注/导入)按分类×采购月汇总 (用户 2026-06-27)。
    # 经 PartPurchase.material_code → 配件库分类; 未编码的采购映射不到分类(诚实留空, 进未分类不强塞)。
    purch_all: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    for mcode, amt, pdate in db.execute(
        select(PartPurchase.material_code, PartPurchase.amount, PartPurchase.purchase_date)
        .where(PartPurchase.amount.isnot(None))
    ).all():
        cat = (mat_info.get(mcode) or {}).get("category") if mcode else None
        ym = _ym(pdate)
        # 自购木材(MW)等零星采购计入对账; 仅排纯工厂服务(木作 WD/人工 MP)(用户 2026-06-29)。
        if cat and ym and (mcode or "").split("-", 1)[0].upper() not in _PURCHASE_EXCLUDE_PREFIXES:
            purch_all[cat][ym] += _d(abs(amt))

    out_cats = []
    all_keys = sorted(set(_ac_categories(db)) | set(std.keys()) | set(fact_all.keys()) | set(purch_all.keys()))
    for cat in all_keys:
        mode = _settle_mode(cat)   # 月结(五金/电力轨道) | 零星(其余)
        std_p, cnt_p = std.get(cat, {}), cnt.get(cat, {})
        fact_p, purch_p = fact_all.get(cat, {}), purch_all.get(cat, {})
        spor_p = spor.get(cat, {})
        periods = sorted(set(std_p) | set(cnt_p) | set(fact_p) | set(purch_p) | set(spor_p))
        rows = []
        t_std = t_fact = t_purch = t_actual = t_spor = Decimal("0")
        hist_rates: list[Decimal] = []
        for ym in periods:
            s = std_p.get(ym, Decimal("0")).quantize(_CENTS)
            c = cnt_p.get(ym, 0)
            has_fact = ym in fact_p
            fact = fact_p.get(ym, Decimal("0")).quantize(_CENTS) if has_fact else None
            has_purch = ym in purch_p
            purch = purch_p.get(ym, Decimal("0")).quantize(_CENTS) if has_purch else None
            # 月结类"实际"=工厂月度对账总额; 零星类"实际"=真实采购单(支付宝备注/导入)
            actual = fact if mode == "月结" else purch
            has_actual = has_fact if mode == "月结" else has_purch
            if hist_rates and c > 0:
                hist = ((sum(hist_rates, Decimal("0")) / Decimal(len(hist_rates))) * c).quantize(_CENTS)
            else:
                hist = s
            var = (actual - s).quantize(_CENTS) if (has_actual and actual is not None) else None
            var_pct = float((var / s * 100).quantize(_CENTS)) if (var is not None and s > 0) else None
            sp = spor_p.get(ym, Decimal("0")).quantize(_CENTS)
            rows.append({
                "period": ym, "historical_avg": float(hist), "standard_consume": float(s),
                "factory_actual": float(fact) if has_fact else None, "has_factory_actual": has_fact,
                "actual_purchase": float(purch) if has_purch else None, "has_actual_purchase": has_purch,
                "settle_mode": mode,
                "actual": float(actual) if (has_actual and actual is not None) else None, "has_actual": has_actual,
                "variance": float(var) if var is not None else None, "variance_pct": var_pct,
                "order_count": c,
                # 月结类里已走零星采购、已从月结预估扣除的金额(防双算多付; 见 export 红字提示)
                "sporadic_excluded": float(sp),
            })
            t_std += s
            t_spor += sp
            if has_fact:
                t_fact += fact
            if has_purch:
                t_purch += purch
            if has_actual and actual is not None:
                t_actual += actual
                if c > 0:
                    hist_rates.append(actual / Decimal(c))
        t_var = (t_actual - t_std).quantize(_CENTS)
        out_cats.append({
            "key": cat, "name": cat, "settle_mode": mode, "periods": rows,
            "total_standard": float(t_std),
            "total_factory_actual": float(t_fact), "total_actual_purchase": float(t_purch),
            "total_actual": float(t_actual), "total_variance": float(t_var),
            "total_sporadic_excluded": float(t_spor),
            "total_variance_pct": float((t_var / t_std * 100).quantize(_CENTS)) if t_std > 0 else None,
        })
    return {"granularity": granularity, "ship_date_basis": True, "category_driven": True, "materials": out_cats}


# ── C. 工厂月度对账总额 录入/查询/删除 (material_key 存「分类」) ──────────────
def list_monthly_recon(db: Session, *, material_key: Optional[str] = None,
                       year_month: Optional[str] = None) -> list[dict]:
    q = select(PartsMonthlyRecon)
    if material_key:
        q = q.where(PartsMonthlyRecon.material_key == material_key)
    if year_month:
        q = q.where(PartsMonthlyRecon.year_month == year_month)
    rows = db.execute(
        q.order_by(PartsMonthlyRecon.year_month.desc(), PartsMonthlyRecon.id.desc())
    ).scalars().all()
    return [{
        "id": r.id, "material_key": r.material_key, "material_name": r.material_key,
        "year_month": r.year_month, "supplier": r.supplier,
        "actual_total": float(_d(r.actual_total)), "note": r.note,
    } for r in rows]


def save_monthly_recon(db: Session, *, material_key: str, year_month: str, actual_total,
                       supplier: Optional[str] = None, note: Optional[str] = None,
                       recon_id: Optional[int] = None) -> dict:
    """录入/更新 某分类某月工厂返回的对账总额。recon_id 给了=更新, 否则新增(同分类同月可多供应商)。"""
    material_key = (material_key or "").strip()
    if not material_key:
        raise ValueError("分类不能为空")
    amt = _d(actual_total)
    if recon_id is not None:
        row = db.get(PartsMonthlyRecon, recon_id)
        if row is None:
            raise ValueError(f"对账记录 {recon_id} 不存在")
        row.material_key, row.year_month = material_key, year_month
        row.supplier, row.actual_total, row.note = (supplier or None), amt, (note or None)
    else:
        row = PartsMonthlyRecon(material_key=material_key, year_month=year_month,
                                supplier=supplier or None, actual_total=amt, note=note or None)
        db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "material_key": row.material_key, "material_name": row.material_key,
            "year_month": row.year_month, "supplier": row.supplier,
            "actual_total": float(_d(row.actual_total)), "note": row.note}


def delete_monthly_recon(db: Session, recon_id: int) -> bool:
    row = db.get(PartsMonthlyRecon, recon_id)
    if row is None:
        return False
    db.delete(row)
    db.commit()
    return True


# ── D. 当月「已发货」订单清单导出 (发给工厂对账) ──────────────────────────────
def _orders_shipped_in(db: Session, year_month: str) -> list[Order]:
    return [o for o in _settled_shipped_orders(db) if _ym(o.ship_date) == year_month]


def _orders_shipped_between(db: Session, d_from: date, d_to: date) -> list[Order]:
    """同 _orders_shipped_in, 但按发货日期区间 [d_from, d_to] 圈定(含两端)。口径仍是 ship_date。"""
    return [o for o in _settled_shipped_orders(db)
            if o.ship_date and d_from <= o.ship_date <= d_to]


def _parse_ymd(v) -> Optional[date]:
    """'YYYY-MM-DD' / 'YYYY-MM'(取月初) / date → date; 空 → None。"""
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    try:
        return date.fromisoformat(s)
    except ValueError:
        parts = s.split("-")
        if len(parts) == 2:   # 允许传 'YYYY-MM' → 当月 1 号
            return date(int(parts[0]), int(parts[1]), 1)
        raise


def _resolve_shipped_orders(db: Session, year_month: Optional[str],
                            date_from, date_to) -> tuple[list[Order], str]:
    """按 (date_from,date_to) 区间优先, 否则按 year_month 单月, 圈定发货单; 返回 (订单列表, 期间标签)。"""
    df, dt = _parse_ymd(date_from), _parse_ymd(date_to)
    if df or dt:
        df = df or date(1970, 1, 1)
        dt = dt or date(9999, 12, 31)
        if df > dt:
            df, dt = dt, df
        return _orders_shipped_between(db, df, dt), f"{df.isoformat()}~{dt.isoformat()}"
    if year_month:
        return _orders_shipped_in(db, year_month), year_month
    raise ValueError("导出需给 year_month 或 date_from/date_to")


def _product_name_map(db: Session) -> dict:
    """{产品编码: 我们自己的产品名(Product.name 短名)} — 导出用短名替换淘宝长标题, 工厂看得清。"""
    from app.models.product import Product
    return {p.code: p.name for p in db.execute(select(Product)).scalars() if p.code and p.name}


def export_shipped_orders(db: Session, *, year_month: Optional[str] = None,
                          date_from=None, date_to=None,
                          material_key: Optional[str] = None) -> dict:
    """导已发货成交订单清单(发货月单月 year_month, 或发货日区间 date_from~date_to, 二选一)。

    material_key(=分类)空 = 全部发货单(基础列, 工厂自挑); 给了 = 只列 BOM 里有该分类外采配件的单,
    逐单展开该分类的配件部位 + 预设尺寸(去重、排木作)。按发货日期口径(ship_date)。
    """
    orders, period = _resolve_shipped_orders(db, year_month, date_from, date_to)
    prod_map = _product_name_map(db)   # 产品编码 → 我们自己的短名(替换淘宝长标题, 工厂看得清)

    def _disp(o) -> str:
        return prod_map.get(o.product_code) or (o.product_name or o.product_code or "")[:24]

    out_orders: list[dict] = []
    t_est = Decimal("0")

    if material_key:
        from app.services import sku_utils
        mat_info, bom_by_pcsku, bom_by_pc = _load_bom_and_materials(db)
        sporadic_cover = _sporadic_covered(db, mat_info, bom_by_pcsku, bom_by_pc)
        for o in orders:
            cons = _order_category_consumption(o, mat_info, bom_by_pcsku, bom_by_pc).get(material_key)
            if not cons:
                continue
            parts = sorted(cons["materials"], key=lambda p: p["part_name"])
            est = cons["amount"]
            t_est += est
            # 该单该类已走零星采购(支付宝现付) → 红字提示工厂勿计入月结(防多付); 带流水号证据。
            cover = sporadic_cover.get((o.order_no, material_key))
            out_orders.append({
                "order_no": o.order_no,
                "order_date": o.order_date.isoformat() if o.order_date else None,
                "ship_date": o.ship_date.isoformat() if o.ship_date else None,
                "customer_name": o.customer_name,
                "product_name": o.product_name,
                "product_display": _disp(o),
                "sku": o.sku,
                "est_parts": float(est),
                "sporadic": bool(cover),
                "sporadic_note": _sporadic_note(cover) if cover else None,
                "is_custom": bool(getattr(o, "is_custom", False))
                or sku_utils.is_custom_sku_code(o.sku_code, o.product_code),
                "bom_parts": [{
                    "part_name": p["part_name"], "material_code": p["material_code"],
                    "category": p["category"], "qty": p["qty"], "unit": p["unit"],
                    "size_note": p["size_note"],
                    # 单价(Material.price) + 总价(=qty×price), _emit 已算好直接透传
                    "unit_price": p.get("price"), "total_price": p.get("amount"),
                    "size_uncertain": p.get("size_uncertain", False),
                    "alt_size_count": p.get("alt_size_count", 0),
                } for p in parts],
            })
    else:
        for o in orders:
            est = _d(o.est_parts)
            t_est += est
            out_orders.append({
                "order_no": o.order_no,
                "order_date": o.order_date.isoformat() if o.order_date else None,
                "ship_date": o.ship_date.isoformat() if o.ship_date else None,
                "customer_name": o.customer_name,
                "product_name": o.product_name,
                "product_display": _disp(o),
                "sku": o.sku,
                "est_parts": float(est),
            })

    out_orders.sort(key=lambda x: (x["ship_date"] or "", x["order_no"]))
    return {
        "year_month": year_month,
        "period": period,
        "material_key": material_key,
        "material_name": material_key,
        "order_count": len(out_orders),
        "total_est_parts": float(t_est.quantize(_CENTS)),
        "orders": out_orders,
    }


# ── 导出样式 (工厂对账用: 按月分组 + 配色区隔 + 高行距, 简单清晰) ──────────────────
_C_HEADER = "1A73E8"      # 表头: 谷歌蓝底白字
_C_MONTH = "D2E3FC"       # 月份分组标题: 浅蓝
_C_ZEBRA = "F1F5FB"       # 斑马纹: 极浅蓝灰
_C_SUBTOTAL = "FFF3CD"    # 月小计: 浅黄
_C_TOTAL = "FCE3B4"       # 总计: 琥珀
_C_BORDER = "C6CDD5"      # 边框: 浅灰


def _write_grouped_sheet(ws, d: dict, *, category: bool, show_est_price: bool = False) -> None:
    """按发货月分组的漂亮清单页: 表头蓝底 + 每月分组标题 + 逐单斑马纹 + 月小计 + 总计, 全表边框、
    高行距、产品用我们的短名(小字)。category=False 为『全部发货单』基础列; True 为某分类逐单展开 BOM。
    工厂对账用(category 且 show_est_price=False)单价/金额留空给工厂填。"""
    from itertools import groupby

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color=_C_BORDER)
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    A_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    A_RIGHT = Alignment(horizontal="right", vertical="center")

    if category:
        headers = ["订单号", "发货日", "产品", "部位 / 材料", "数量", "预设尺寸", "单价", "金额"]
        widths = [22, 12, 15, 18, 7, 20, 10, 12]
        money_cols, prod_col = [7, 8], 3
    else:
        headers = ["订单号", "发货日", "客户", "产品", "预估配件金额"]
        widths = [22, 12, 12, 18, 15]
        money_cols, prod_col = [5], 4
    ncol = len(headers)
    show_money = (not category) or show_est_price   # 工厂填价的分类页: 金额列留空

    def _row(vals, *, fill=None, bold=False, fontcolor=None, height=21, prod_small=False):
        ws.append(vals)
        r = ws.max_row
        ws.row_dimensions[r].height = height
        for ci in range(1, ncol + 1):
            c = ws.cell(row=r, column=ci)
            c.border = BORDER
            c.font = Font(bold=bold, color=fontcolor,
                          size=10 if (prod_small and ci == prod_col) else 11)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            if ci == 1:
                c.number_format = "@"        # 19位订单号防科学计数
                c.alignment = A_LEFT
            elif ci in money_cols:
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0.00"
                c.alignment = A_RIGHT
            elif ci == prod_col:
                c.alignment = A_LEFT
            else:
                c.alignment = A_CENTER
        return r

    _row(headers, fill=_C_HEADER, bold=True, fontcolor="FFFFFF", height=26)

    grand = Decimal("0")
    for ym, grp in groupby(d["orders"], key=lambda o: (o.get("ship_date") or "")[:7] or "无发货日"):
        grp = list(grp)
        hr = _row([f"📅 {ym}　发货 {len(grp)} 单"] + [None] * (ncol - 1),
                  fill=_C_MONTH, bold=True, height=24)
        ws.merge_cells(start_row=hr, start_column=1, end_row=hr, end_column=ncol)
        ws.cell(row=hr, column=1).alignment = A_LEFT
        sub = Decimal("0")
        for oi, o in enumerate(grp):
            prod = o.get("product_display") or ""
            tags = []
            if o.get("is_custom"):
                tags.append("定制")
            if o.get("sporadic"):
                tags.append("⚠已现付·勿计")
            if tags:
                prod = f"{prod}  【{'·'.join(tags)}】"
            zebra = _C_ZEBRA if (oi % 2 == 1) else None
            if category:
                for pi, p in enumerate(o.get("bom_parts") or [{}]):
                    size = p.get("size_note") or "—"
                    if p.get("size_uncertain"):
                        size = f"{size}(尺寸估)"
                    up = tp = None
                    if show_est_price:
                        up, tp = p.get("unit_price"), p.get("total_price")
                        if tp is not None and not o.get("sporadic"):
                            sub += _d(tp)
                    _row([o["order_no"] if pi == 0 else "",
                          (o.get("ship_date") or "") if pi == 0 else "",
                          prod if pi == 0 else "",
                          p.get("part_name") or "", p.get("qty"), size, up, tp],
                         fill=zebra, prod_small=True)
            else:
                sub += _d(o.get("est_parts"))
                _row([o["order_no"], o.get("ship_date") or "", o.get("customer_name") or "—",
                      prod, o.get("est_parts")], fill=zebra, prod_small=True)
        srow = [None] * ncol
        srow[0] = f"↳ {ym} 小计"
        if show_money:
            srow[money_cols[-1] - 1] = float(sub.quantize(_CENTS))
        _row(srow, fill=_C_SUBTOTAL, bold=True, height=20)
        grand += sub

    trow = [None] * ncol
    trow[0] = "总计（系统预估）" if show_money else "总计（工厂核对填写）"
    if show_money:
        trow[money_cols[-1] - 1] = float(grand.quantize(_CENTS))
    _row(trow, fill=_C_TOTAL, bold=True, height=26)

    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _write_all_orders_ws(ws, d: dict) -> None:
    """『全部发货单』按月分组: 订单号/发货日/客户/产品(我们短名)/预估配件金额。"""
    _write_grouped_sheet(ws, d, category=False)


def _write_category_ws(ws, d: dict, *, show_est_price: bool = False) -> None:
    """某分类逐单展开 BOM 按月分组。show_est_price=True 自己对账(填系统预估, 扣零星);
    False 发给工厂(单价/金额留空给工厂填)。"""
    _write_grouped_sheet(ws, d, category=True, show_est_price=show_est_price)


def build_shipped_orders_xlsx(db: Session, *, year_month: Optional[str] = None,
                              date_from=None, date_to=None,
                              material_key: Optional[str] = None):
    """导清单 → xlsx(扁平表格)。单月(year_month)或发货日区间(date_from~date_to)。
    发给工厂对账用: 材料单价/总价留空给工厂填。返回 (openpyxl.Workbook, data_dict)。
    """
    import openpyxl
    d = export_shipped_orders(db, year_month=year_month, date_from=date_from,
                              date_to=date_to, material_key=material_key)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (material_key or "全部发货单")[:28]
    if material_key:
        _write_category_ws(ws, d, show_est_price=False)
    else:
        _write_all_orders_ws(ws, d)
    return wb, d


def build_bulk_recon_workbook(db: Session, *, date_from=None, date_to=None,
                              year_month: Optional[str] = None):
    """一份多 sheet 对账工作簿(自己对账所有月结账户用):
      · sheet「全部发货单」= 区间内所有已发货成交订单(基础列 + 预估配件合计);
      · 每个『月结账户』(五金/电力轨道/岩板/玻璃)各一页 = 逐单展开 BOM + 系统预估单价/总价/合计。
    让用户一次对完所有月结账户的应付预估。按发货日期 ship_date 圈定。返回 (Workbook, 摘要dict)。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    d_all = export_shipped_orders(db, year_month=year_month, date_from=date_from,
                                  date_to=date_to, material_key=None)
    ws = wb.active
    ws.title = "全部发货单"
    _write_all_orders_ws(ws, d_all)

    cats: list[dict] = []
    for cat in _MONTHLY_SETTLE_CATEGORIES:
        dc = export_shipped_orders(db, year_month=year_month, date_from=date_from,
                                   date_to=date_to, material_key=cat)
        cats.append({"category": cat, "order_count": dc["order_count"],
                     "total_est_parts": dc["total_est_parts"]})
        if not dc["orders"]:
            continue   # 该月结账户区间内无发货单 → 跳过空页
        wsx = wb.create_sheet(title=cat[:28])
        _write_category_ws(wsx, dc, show_est_price=True)

    return wb, {
        "period": d_all["period"],
        "all_order_count": d_all["order_count"],
        "all_total_est_parts": d_all["total_est_parts"],
        "monthly_categories": cats,
    }


# ── E. 双算自检: 月结分类里被零星采购覆盖的(订单×分类) ──────────────────────────
def detect_sporadic_monthly_overlap(db: Session) -> list[dict]:
    """列出「月结分类却已走零星采购」的(订单×分类) —— 这些已从月结预估扣除/在导出标红,
    供人工核对、防工厂月结重复计费多付。空列表=无重叠(干净)。
    """
    mat_info, bom_by_pcsku, bom_by_pc = _load_bom_and_materials(db)
    sporadic_cover = _sporadic_covered(db, mat_info, bom_by_pcsku, bom_by_pc)
    out: list[dict] = []
    for (ono, cat), cover in sporadic_cover.items():
        if _settle_mode(cat) != "月结":
            continue
        o = db.execute(select(Order).where(Order.order_no == ono)).scalar_one_or_none()
        total = sum((_d(e.get("amount")) for e in cover), Decimal("0")).quantize(_CENTS)
        out.append({
            "order_no": ono, "category": cat,
            "ship_month": _ym(o.ship_date) if o else None,
            "product_name": o.product_name if o else None,
            "sporadic_amount": float(total),
            "evidence": cover,
        })
    out.sort(key=lambda x: (x.get("ship_month") or "", x["category"], x["order_no"]))
    return out

"""Compatibility importer for a one-column platform coupon-floor feedback file.

The imported value is diagnostic evidence only.  It never caps a real SKU's
signup price and never changes single-item discount.  A later campaign run still
requires a fresh minimum-list-price line as well (R17).
"""
from __future__ import annotations

import io
from decimal import Decimal
from typing import Optional

from sqlalchemy.orm import Session

_SID_ALIASES = ("商品SKU ID", "商品SKUID", "SKUID", "SKU ID", "sku_id")
_LINE_ALIASES = (
    "平台历史线(最低普惠券后)", "平台历史线", "最低普惠券后价",
    "历史最低普惠券后价", "校验期最低普惠券后价", "最低券后价",
)


def _find_col(header: list[str], aliases: tuple[str, ...]) -> Optional[int]:
    for alias in aliases:
        if alias in header:
            return header.index(alias)
    for alias in aliases:
        for index, value in enumerate(header):
            if value and alias in value:
                return index
    return None


def import_from_xlsx_bytes(db: Session, raw: bytes, sheet: Optional[str] = None,
                           header_row: int = 0) -> dict:
    import openpyxl
    from app.services import campaign_price_floor_service

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) <= header_row + 1:
        return {"ok": False, "error": "表里没有数据行"}
    header = [str(cell).strip() if cell is not None else "" for cell in rows[header_row]]
    sid_col = _find_col(header, _SID_ALIASES)
    line_col = _find_col(header, _LINE_ALIASES)
    if sid_col is None or line_col is None:
        return {"ok": False, "error": f"没找到 SKUID/最低普惠券后价列；实际表头：{[h for h in header if h][:12]}"}

    by_sid: dict[str, Decimal] = {}
    for row in rows[header_row + 1:]:
        if not row or sid_col >= len(row) or line_col >= len(row):
            continue
        sid = str(row[sid_col] or "").strip()
        try:
            line = Decimal(str(row[line_col]))
        except Exception:  # noqa: BLE001
            continue
        if sid and line > 0:
            by_sid[sid] = min(line, by_sid.get(sid, line))
    result = campaign_price_floor_service.record_partial_evidence(
        db,
        [{"sku_id": sid, "min_coupon_line": value} for sid, value in by_sid.items()],
        source="manual_coupon_floor_feedback",
    )
    db.commit()
    return {
        "ok": True,
        "file_rows": len(by_sid),
        "evidence": result,
        "note": "仅更新资格证据；真实SKU报名价仍等于ERP日常价，且缺最低标价时仍会被R17阻塞",
    }

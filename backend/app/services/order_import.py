"""订单 CSV 导入 (plan §10 Phase 3：淘宝 CSV)。

最小实现：列名 → 字段映射；订单编号重复跳过；新订单默认 pending_payment。
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import StringIO
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.order import Order
from app.services import import_clean, order_cost_service
from app.services.taobao_order_import import _map_status, apply_refill_flags  # 状态映射 + 补单回标 (单一来源)

# 列名别名表：CSV 头 (规范化后) → Order 字段
COLUMN_ALIASES: dict[str, str] = {
    "平台": "platform",
    "订单编号": "order_no",
    "订单号": "order_no",
    "是否补单": "is_refill",
    "下单日期": "order_date",
    "下单时间": "order_date",
    "客户姓名": "customer_name",
    "客户": "customer_name",
    "联系电话": "customer_phone",
    "手机": "customer_phone",
    "收货地址": "customer_address",
    "地址": "customer_address",
    "发货日期": "ship_date",
    "产品编码": "product_code",
    "产品名称": "product_name",
    "sku": "sku",
    "是否定制": "is_custom",
    "数量": "qty",
    "物流公司": "carrier",
    "物流单号": "tracking_no",
    "实付金额": "paid_amount",
    "买家实付金额": "paid_amount",
    "买家实际支付金额": "paid_amount",
    "买家应付货款": "buyer_payable_amount",   # 应付≠实付, 分开存 (逐笔对账/现金流都要)
    "应付金额": "buyer_payable_amount",
    "打款商家金额": "shop_received_amount",    # 店铺实收 (历史CSV有此列)
    "卖家服务费": "platform_fee",
    "平台服务费": "platform_fee",
    "退款金额": "refund_amount",
    "店铺名称": "shop",
    "店铺": "shop",
    "订单状态": "status",                      # 淘宝订单状态 → 修复"全部pending_payment"病根
    "发货时间": "ship_date",
    "主订单编号": "order_no",
    "订单创建时间": "order_date",
    "宝贝标题": "product_name",
    "商品标题": "product_name",
    "宝贝总数量": "qty",
    "购买数量": "qty",
    "收货人姓名": "customer_name",
    "联系手机": "customer_phone",
}


@dataclass
class ImportReport:
    inserted: int = 0
    backfilled: int = 0
    skipped_duplicate: int = 0
    skipped_invalid: int = 0
    errors: list[str] = None  # type: ignore[assignment]

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


def _to_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Excel 日期序列号 (46175 → 2026-06-08): 经 Excel 转存的 CSV 常见 (C14)
    return import_clean.excel_serial_to_date(s)


def _to_bool(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"是", "yes", "y", "true", "1", "✓"}


def _to_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _to_decimal(v: Any) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    s = str(v).replace(",", "").replace("¥", "").replace("元", "").strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def import_orders_from_csv(db: Session, csv_text: str) -> ImportReport:
    report = ImportReport()
    reader = csv.DictReader(StringIO(csv_text))

    # 标准化表头
    field_map: dict[str, str] = {}
    for raw in reader.fieldnames or []:
        norm = (raw or "").strip().lower()
        for alias_key, field_name in COLUMN_ALIASES.items():
            if alias_key.lower() == norm:
                field_map[raw] = field_name
                break

    if "order_no" not in field_map.values():
        report.errors.append("CSV 缺少『订单编号』列，无法导入")
        return report

    seen_in_batch: set[str] = set()
    for line_no, row in enumerate(reader, start=2):
        payload: dict[str, Any] = {}
        for raw_col, field_name in field_map.items():
            payload[field_name] = row.get(raw_col)

        order_no = (payload.get("order_no") or "").strip() if payload.get("order_no") else ""
        if not order_no:
            report.skipped_invalid += 1
            continue

        # 本批已处理过 → 跳过
        if order_no in seen_in_batch:
            report.skipped_duplicate += 1
            continue
        seen_in_batch.add(order_no)
        # DB 已存在: 只回填空缺(金额/客户), 不覆盖已有数据, 也不重复插入
        existing = db.execute(
            select(Order).where(Order.order_no == order_no)
        ).scalar_one_or_none()
        if existing is not None:
            changed = False
            _payable = _to_decimal(payload.get("buyer_payable_amount"))
            _paid = _to_decimal(payload.get("paid_amount")) or _payable
            if _paid and not existing.paid_amount:
                existing.paid_amount = _paid
                changed = True
            if _payable and not existing.buyer_payable_amount:
                existing.buyer_payable_amount = _payable
                changed = True
            for fld in ("shop_received_amount", "platform_fee", "refund_amount"):
                v = _to_decimal(payload.get(fld))
                if v is not None and getattr(existing, fld) is None:
                    setattr(existing, fld, v)
                    changed = True
            # 订单状态: 有『订单状态』列就以淘宝导出为准刷新 (修复历史单卡 pending_payment)
            if payload.get("status"):
                new_status = _map_status(payload.get("status"))
                if new_status != existing.status:
                    existing.status = new_status
                    changed = True
            sd = _to_date(payload.get("ship_date"))
            if sd and not existing.ship_date:
                existing.ship_date = sd
                changed = True
            for fld in ("customer_name", "customer_phone", "shop"):
                v = payload.get(fld)
                if fld == "customer_phone":
                    v = import_clean.clean_phone(v)   # 去淘宝虚拟分机后缀 -NNNN
                if v and not getattr(existing, fld):
                    setattr(existing, fld, v)
                    changed = True
            if changed:
                report.backfilled += 1
            else:
                report.skipped_duplicate += 1
            continue

        _is_refill = _to_bool(payload.get("is_refill"))
        _product_name = payload.get("product_name")
        _sku = payload.get("sku")
        order = Order(
            platform=(payload.get("platform") or "淘宝").strip(),
            order_no=order_no,
            is_refill=_is_refill,
            order_date=_to_date(payload.get("order_date")),
            ship_date=_to_date(payload.get("ship_date")),
            customer_name=payload.get("customer_name"),
            customer_phone=import_clean.clean_phone(payload.get("customer_phone")),
            customer_address=payload.get("customer_address"),
            product_code=payload.get("product_code"),
            product_name=_product_name,
            sku=_sku,
            is_custom=_to_bool(payload.get("is_custom")),
            qty=_to_int(payload.get("qty"), default=1),
            carrier=payload.get("carrier"),
            tracking_no=payload.get("tracking_no"),
            buyer_payable_amount=_to_decimal(payload.get("buyer_payable_amount")),
            paid_amount=_to_decimal(payload.get("paid_amount")) or _to_decimal(payload.get("buyer_payable_amount")),
            shop_received_amount=_to_decimal(payload.get("shop_received_amount")),
            platform_fee=_to_decimal(payload.get("platform_fee")),
            refund_amount=_to_decimal(payload.get("refund_amount")),
            shop=payload.get("shop"),
            # 订单状态映射 (淘宝导出); 无『订单状态』列时 _map_status 返回 pending_payment (原默认)
            status=_map_status(payload.get("status")),
            warehouse=order_cost_service.default_warehouse_for(_product_name, _sku, _is_refill),
        )
        db.add(order)
        report.inserted += 1

    db.flush()
    apply_refill_flags(db)   # 用补单对账回标 is_refill (导入后立即匹配, 优先级最高)
    db.commit()
    return report


def import_orders_from_xlsx(db: Session, content: bytes) -> ImportReport:
    """Excel 订单表导入: 取第一个含『订单编号/主订单编号』表头的工作表 → 转 CSV 文本复用 CSV 导入。

    淘宝销售明细 xlsx 含 订单报表/销售明细/发货报表 三表; 取订单报表(单级, 不重复计) 最稳。
    """
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    for ws in wb.worksheets:
        header: Optional[list] = None
        hr = 0
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=12, min_col=1, max_col=40, values_only=True), start=1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(c in ("订单编号", "主订单编号", "订单号") for c in cells):
                header = cells
                hr = ri
                break
        if not header:
            continue
        width = len(header)
        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow(header)
        for row in ws.iter_rows(min_row=hr + 1, max_row=200000, min_col=1, max_col=width, values_only=True):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            writer.writerow([
                "" if c is None else (c.isoformat() if hasattr(c, "isoformat") else str(c))
                for c in row[:width]
            ])
        return import_orders_from_csv(db, buf.getvalue())
    rep = ImportReport()
    rep.errors.append("Excel 中未找到含『订单编号 / 主订单编号』的工作表")
    return rep

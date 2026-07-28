"""木作工厂月结销账服务 (用户 2026-07-01)。

核心: 把"已开账单未付"的工厂单(FactoryOrder, factory_bill_amount 非空、unpaid)按【结算月】分组,
声明驱动地整月翻成已付(payment_status='paid') → 现金流"工厂结算(已开账单未付)"随之下降。

结算月 = FactoryOrder.settlement_month(工厂账单说的月), 缺省时按 order_date 月推断
(用户口径: "X月货款按下单时间对", 故未导账单也能直接按下单月销账)。

销账触发(不卡金额, 工厂常有减免/加费):
  - manual: 月结页/异常列表一键「已付清」
  - keyword: 支付宝备注「5月已付清/已结清…」自动识别(P2)
每次销账建 FactorySettlementPayment 记录(可一键撤销, 反查 settlement_payment_id 回滚本批翻过的单)。
"""
from __future__ import annotations

import re
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Optional

from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.models.factory_settlement import (
    DEFAULT_WOOD_SUPPLIER,
    FactorySettlementPayment,
    FactorySupplierAlias,
)
from app.models.order import FactoryOrder
from app.services import factory_advance_service

_Q = Decimal("0.01")
# 默认木作供应商别名(博冠货款走个人账户; 匹配时去星号+双向包含, 故全名即可覆盖打码流水)
_DEFAULT_ALIASES = ["博冠", "玉山", "伟男", "程卫燕"]


def _d(v) -> Decimal:
    return Decimal("0") if v is None else Decimal(str(v))


def _order_month(fo: FactoryOrder) -> Optional[str]:
    """结算归属月: settlement_month 优先, 否则按 order_date 推断 (YYYY-MM)。"""
    if fo.settlement_month:
        return fo.settlement_month
    if fo.order_date:
        return fo.order_date.strftime("%Y-%m")
    return None


# ── 月度欠款台账 ──────────────────────────────────────────────
def _apply_product_search(stmt, q: Optional[str]):
    """按 产品名/SKU/产品编码 模糊搜索过滤工厂单 (q 为空则不过滤)。"""
    if q and q.strip():
        pq = f"%{q.strip()}%"
        stmt = stmt.where(or_(
            FactoryOrder.product_name.ilike(pq),
            FactoryOrder.sku.ilike(pq),
            FactoryOrder.product_code.ilike(pq),
        ))
    return stmt


def write_settlement_ws(wb, rows: Optional[list] = None, *, db: Optional[Session] = None) -> None:
    """把木作工厂月结逐单明细写进 openpyxl workbook 的新 sheet「木作工厂月结」(两个导出共用, 用户 2026-07-08)。
    复用 settlement_detail_rows(db)(默认木作供应商)。列: 结算月/工厂单号/平台订单号/产品/数量/账单金额/已付金额/付款状态。"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    if rows is None:
        rows = settlement_detail_rows(db) if db is not None else []
    ws = wb.create_sheet("木作工厂月结")
    heads = ["结算月", "工厂单号", "平台订单号", "产品", "数量", "账单金额", "已付金额", "付款状态"]
    head_fill = PatternFill("solid", fgColor="1F4E79")
    for ci, h in enumerate(heads, start=1):
        c = ws.cell(1, ci, h)
        c.fill = head_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
    paid_fill = PatternFill("solid", fgColor="E8F5E9")
    unpaid_fill = PatternFill("solid", fgColor="FFF3E0")
    r = 2
    total = 0.0
    for row in rows:
        ws.cell(r, 1, row.get("settlement_month") or "")
        ws.cell(r, 2, str(row.get("factory_order_no") or "")).number_format = "@"
        ws.cell(r, 3, str(row.get("platform_order_no") or "")).number_format = "@"
        ws.cell(r, 4, (row.get("product_name") or row.get("sku") or "")[:28])
        ws.cell(r, 5, row.get("qty") or 0)
        amt = float(row.get("bill_amount") or 0)
        ws.cell(r, 6, amt).number_format = "0.00"
        ws.cell(r, 7, float(row.get("paid_amount") or 0)).number_format = "0.00"
        st = row.get("payment_status") or "未付"
        sc = ws.cell(r, 8, st)
        sc.fill = paid_fill if st == "已付" else unpaid_fill
        total += amt
        r += 1
    if rows:
        ws.cell(r + 1, 4, "合计").font = Font(bold=True)
        tc = ws.cell(r + 1, 6, round(total, 2))
        tc.number_format = "0.00"
        tc.font = Font(bold=True)
    for i, w in enumerate([10, 18, 20, 28, 6, 12, 12, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def month_breakdown(db: Session, supplier: str = DEFAULT_WOOD_SUPPLIER,
                    q: Optional[str] = None) -> dict:
    """按结算月汇总该供应商「已开账单」工厂单: 应付/已付/未付/状态。

    返回 {supplier, months:[{month, billed, paid, unpaid, order_count, paid_count, status}], total_*}。
    status: paid(已付清) / unpaid(未付清) / partial(部分付清)。
    q: 产品名/SKU/产品编码 模糊搜索, 只汇总匹配的单 (用户 2026-07-03)。
    """
    rows = db.execute(
        _apply_product_search(
            select(FactoryOrder).where(
                FactoryOrder.factory_name == supplier,
                FactoryOrder.voided_at.is_(None),
                FactoryOrder.factory_bill_amount.isnot(None),
            ), q)
    ).scalars().all()
    agg: dict[str, dict] = {}
    for fo in rows:
        mk = _order_month(fo) or "(无日期)"
        a = agg.setdefault(mk, {"month": mk, "billed": Decimal("0"), "paid": Decimal("0"),
                                "unpaid": Decimal("0"), "order_count": 0, "paid_count": 0})
        amt = _d(fo.factory_bill_amount)
        a["billed"] += amt
        a["order_count"] += 1
        if (fo.payment_status or "") == "paid":
            a["paid"] += amt
            a["paid_count"] += 1
        else:
            a["unpaid"] += amt
    months = []
    tot_billed = tot_paid = tot_unpaid = Decimal("0")
    for mk in sorted(agg.keys()):
        a = agg[mk]
        if a["unpaid"] <= 0:
            status = "paid"
        elif a["paid"] > 0:
            status = "partial"
        else:
            status = "unpaid"
        months.append({**a, "billed": a["billed"].quantize(_Q), "paid": a["paid"].quantize(_Q),
                       "unpaid": a["unpaid"].quantize(_Q), "status": status})
        tot_billed += a["billed"]; tot_paid += a["paid"]; tot_unpaid += a["unpaid"]
    return {
        "supplier": supplier,
        "months": months,
        "total_billed": tot_billed.quantize(_Q),
        "total_paid": tot_paid.quantize(_Q),
        "total_unpaid": tot_unpaid.quantize(_Q),
    }


# ── 销账 / 撤销 ───────────────────────────────────────────────
def settle_month(db: Session, *, supplier: str = DEFAULT_WOOD_SUPPLIER, month: str,
                 trigger: str = "manual", flow_no: Optional[str] = None,
                 paid_amount: Optional[Decimal] = None, by: Optional[str] = None,
                 note: Optional[str] = None) -> dict:
    """把某供应商某结算月「已开账单未付」的工厂单整月翻成已付, 建销账记录。

    幂等: 该月已无未付单 → flipped=0, 不建空记录。返回 {month, flipped, payment_id, billed_total}。
    """
    rows = db.execute(
        select(FactoryOrder).where(
            FactoryOrder.factory_name == supplier,
            FactoryOrder.voided_at.is_(None),
            FactoryOrder.factory_bill_amount.isnot(None),
            FactoryOrder.payment_status == "unpaid",
        )
    ).scalars().all()
    targets = [fo for fo in rows if _order_month(fo) == month]
    if not targets:
        return {"month": month, "flipped": 0, "payment_id": None, "billed_total": "0.00"}

    rec = FactorySettlementPayment(
        supplier=supplier, settlement_month=month, trigger=trigger,
        alipay_flow_no=flow_no, paid_amount=paid_amount, created_by=by, note=note,
        flipped_count=0,
    )
    db.add(rec)
    db.flush()   # 拿到 rec.id

    today = date.today()
    billed_total = Decimal("0")
    for fo in targets:
        fo.payment_status = "paid"
        if fo.payment_date is None:
            fo.payment_date = today
        if flow_no and not fo.alipay_flow_no:
            fo.alipay_flow_no = flow_no
        fo.settlement_payment_id = rec.id
        billed_total += _d(fo.factory_bill_amount)
    rec.flipped_count = len(targets)
    advance = factory_advance_service.apply_for_settlement(
        db,
        payment_id=rec.id,
        month=month,
        billed_total=billed_total,
        by=by,
    )
    advance_used = advance["used"]
    if advance_used > 0:
        suffix = f"预付款自动抵扣 ¥{advance_used}"
        rec.note = f"{rec.note}；{suffix}" if rec.note else suffix
    db.flush()
    return {"month": month, "flipped": len(targets), "payment_id": rec.id,
            "billed_total": billed_total.quantize(_Q),
            "advance_used": advance_used,
            "advance_remaining": advance["remaining"],
            "net_cash_payable": max(Decimal("0"), billed_total - advance_used).quantize(_Q)}


def reverse_settlement(db: Session, payment_id: int, *, by: Optional[str] = None) -> dict:
    """撤销一笔销账: 把本记录翻过的工厂单恢复 unpaid, 标记记录 reversed。

    只回滚 settlement_payment_id==本记录 的单(精确回滚本批), 不动别批/历史已付单。
    """
    rec = db.get(FactorySettlementPayment, payment_id)
    if rec is None:
        return {"reverted": 0, "error": "记录不存在"}
    if rec.reversed_at is not None:
        return {"reverted": 0, "error": "已撤销过"}
    rows = db.execute(
        select(FactoryOrder).where(FactoryOrder.settlement_payment_id == payment_id)
    ).scalars().all()
    n = 0
    for fo in rows:
        fo.payment_status = "unpaid"
        fo.payment_date = None
        if rec.alipay_flow_no and fo.alipay_flow_no == rec.alipay_flow_no:
            fo.alipay_flow_no = None
        fo.settlement_payment_id = None
        n += 1
    rec.reversed_at = datetime.now(timezone.utc)
    rec.reversed_by = by
    advance = factory_advance_service.reverse_for_settlement(
        db, payment_id=payment_id, by=by,
    )
    db.flush()
    return {
        "reverted": n,
        "payment_id": payment_id,
        "advance_restored": advance["restored"],
        "advance_remaining": advance["remaining"],
    }


def list_payments(db: Session, supplier: Optional[str] = None) -> list[dict]:
    """销账记录(含已撤销), 最新在前。"""
    stmt = select(FactorySettlementPayment).order_by(FactorySettlementPayment.id.desc())
    if supplier:
        stmt = stmt.where(FactorySettlementPayment.supplier == supplier)
    advance_by_payment = factory_advance_service.applied_by_payment(db)
    out = []
    for r in db.execute(stmt).scalars().all():
        out.append({
            "id": r.id, "supplier": r.supplier, "settlement_month": r.settlement_month,
            "trigger": r.trigger, "alipay_flow_no": r.alipay_flow_no,
            "paid_amount": str(r.paid_amount) if r.paid_amount is not None else None,
            "flipped_count": r.flipped_count, "created_by": r.created_by, "note": r.note,
            "advance_used": str(advance_by_payment.get(r.id, Decimal("0.00"))),
            "reversed_at": r.reversed_at.isoformat() if r.reversed_at else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return out


# ── 供应商别名 (账户/对手方 → 供应商) ────────────────────────────
def _strip_mask(s: Optional[str]) -> str:
    return re.sub(r"[\s*＊·]", "", s or "")


def seed_default_aliases(db: Session, supplier: str = DEFAULT_WOOD_SUPPLIER) -> int:
    """幂等种入默认木作供应商别名(博冠/玉山/伟男/程卫燕)。返回新增条数。"""
    existing = {(_strip_mask(a.alias), a.supplier) for a in db.execute(
        select(FactorySupplierAlias)).scalars().all()}
    n = 0
    for al in _DEFAULT_ALIASES:
        if (_strip_mask(al), supplier) in existing:
            continue
        db.add(FactorySupplierAlias(supplier=supplier, alias=al, note="默认种入"))
        n += 1
    if n:
        db.flush()
    return n


def list_aliases(db: Session, supplier: Optional[str] = None) -> list[dict]:
    stmt = select(FactorySupplierAlias).order_by(FactorySupplierAlias.id.asc())
    if supplier:
        stmt = stmt.where(FactorySupplierAlias.supplier == supplier)
    return [{"id": a.id, "supplier": a.supplier, "alias": a.alias, "note": a.note}
            for a in db.execute(stmt).scalars().all()]


def add_alias(db: Session, *, supplier: str, alias: str, note: Optional[str] = None) -> dict:
    a = FactorySupplierAlias(supplier=supplier, alias=alias.strip(), note=note)
    db.add(a)
    db.flush()
    return {"id": a.id, "supplier": a.supplier, "alias": a.alias, "note": a.note}


def delete_alias(db: Session, alias_id: int) -> bool:
    a = db.get(FactorySupplierAlias, alias_id)
    if a is None:
        return False
    db.delete(a)
    db.flush()
    return True


def match_supplier(db: Session, counterparty: Optional[str]) -> Optional[str]:
    """支付宝对手方名 → 木作供应商 (去星号 + 双向包含, 兼容打码 **男/**英)。无匹配返回 None。"""
    cpn = _strip_mask(counterparty)
    if not cpn:
        return None
    for a in db.execute(select(FactorySupplierAlias)).scalars().all():
        aln = _strip_mask(a.alias)
        if aln and (aln in cpn or cpn in aln):
            return a.supplier
    return None


# ── 支付宝备注关键词解析 (P2 用; 杂费排除 → 否定优先 → 肯定) ──────────────────────
# 杂费(打包/运费/配件采购/样品/定金/加工…): 付的不是货款, 别拿它销货款账;
# 例外: 明确"货款 + 结算/付清/结清"(整月款, 哪怕含打包费一起结)仍按货款销。
_FEE_KEYWORDS = ("打包费", "运费", "到付", "叉车", "搬运", "配件", "采购",
                 "玻璃", "灯带", "轨道", "样品", "样块", "定金", "材料费",
                 "加工", "贴皮", "封边")
_SETTLE_STRONG = ("结算", "付清", "结清")
_NEG_KEYWORDS = ("未付清", "还没付清", "没付清", "未结清", "未结算", "没结算", "还没结",
                 "先付", "部分", "付一部分", "欠", "差", "待付", "未付款", "没付款")
# 肯定(整月已结)信号。"货款"单独不算 —— 可能只是一笔部分付款, 故不入表(见 test_parse_remark)。
# 扩充点(用户 2026-07-02): 加入"结算/已结算/结算完/货款已付/已付款/已结/款已结",
# 让"X月货款…结算""X月货款已付"等真实备注(如'挚乐1月货款2025结算')也能自动销账。
_POS_KEYWORDS = ("已付清", "已结清", "全部付清", "全款付清", "结清", "付清", "全款",
                 "结算", "已结算", "结算完", "货款已付", "已付款", "已结", "款已结")
_MONTH_RE = re.compile(r"(\d{1,2})\s*月")
_CN_NUM = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6,
           "七": 7, "八": 8, "九": 9, "十": 10, "十一": 11, "十二": 12}
_CN_MONTH_RE = re.compile(r"([一二三四五六七八九十]{1,3})\s*月")


def parse_settlement_remark(text: Optional[str], *, year: int) -> dict:
    """解析支付宝货款备注 → {action, months}。

    action: 'settle'(肯定整月结, 应销账) / 'unsettle'(否定, 别自动销且保持未结)
            / 'fee'(杂费, 非货款结算, 不销货款) / None(无信号)。
    顺序: 杂费排除(除非明确货款结算) → 否定优先(防"还没付清"误判) → 肯定。
    months=["YYYY-MM", ...](按 year 补全, 可多月)。
    """
    t = text or ""
    months: list[str] = []
    for m in _MONTH_RE.findall(t):
        mm = int(m)
        if 1 <= mm <= 12:
            months.append(f"{year:04d}-{mm:02d}")
    for cn in _CN_MONTH_RE.findall(t):
        mm = _CN_NUM.get(cn)
        if mm:
            months.append(f"{year:04d}-{mm:02d}")
    months = sorted(set(months))
    # 杂费付的不是货款 → 不销货款账; 但"货款 + 结算/付清/结清"明确整月结清的仍销。
    is_goods_settle = "货款" in t and any(s in t for s in _SETTLE_STRONG)
    if any(k in t for k in _FEE_KEYWORDS) and not is_goods_settle:
        return {"action": "fee", "months": months}
    if any(k in t for k in _NEG_KEYWORDS):
        return {"action": "unsettle", "months": months}
    if any(k in t for k in _POS_KEYWORDS):
        return {"action": "settle", "months": months}
    return {"action": None, "months": months}


# ── P2: 支付宝出账自动识别木作货款 + 关键词自动销账 ──────────────
def route_alipay_settlements(db: Session, *, default_year: Optional[int] = None) -> dict:
    """扫支付宝出账流水(amount<0), 识别木作供应商货款(别名匹配), 自动处理:

    1. 纠正归类: reconciliation_type → 'factory_payment'(根治货款被误判成 customer_payment, 审计 C16/口径)。
    2. 备注含「X月已付清/已结清…」(肯定, 否定优先排除"还没付清") → 自动 settle_month(trigger=keyword)。
    幂等: 同 (flow_no, month) 已有未撤销销账记录则跳过。返回 {flagged, settled_months, flipped}。
    """
    from app.models.finance import AlipayFlow
    from app.services import field_change_service as _fcs
    flows = db.execute(select(AlipayFlow).where(AlipayFlow.amount < 0)).scalars().all()
    done = {(p.alipay_flow_no, p.settlement_month) for p in db.execute(
        select(FactorySettlementPayment).where(FactorySettlementPayment.reversed_at.is_(None))
    ).scalars().all() if p.alipay_flow_no}
    # 人工锁 (2026-07-12): 核销类型被人改过的流水(修改档案有记录), 机器归类一律绕行 ——
    # 退款护栏只认"退款"特征, 这把锁兜住所有形态(19365 曾被无护栏旧镜像翻回, 对账假差复发)。
    _locked = _fcs.human_pks(db, table="alipay_flows", field="reconciliation_type")
    flagged = flipped = 0
    settled_months: list[str] = []
    for f in flows:
        sup = match_supplier(db, f.counterparty)
        if not sup:
            continue
        if str(f.id) in _locked:
            continue
        # 退款护栏 (2026-07-10): 客户退款流水绝不当工厂货款 —— 即使打码对手方名去星号后恰好【子串】命中
        # 工厂别名(实测: 给客户「山**」的退款去星号=「山」, 命中「玉山」别名 → 被误改 factory_payment,
        # 每天覆盖手工修复、逐笔退款对账反复报差)。判据: 类型/备注含退款/退货, 或已归 refund 家族。
        _rt = (f.reconciliation_type or "").lower()
        _txt = f"{f.transaction_type or ''}{f.remark or ''}"
        if _rt in ("refund", "refund_out", "refund_in", "aftersales") or "退款" in _txt or "退货" in _txt:
            continue
        if f.reconciliation_type != "factory_payment":
            f.reconciliation_type = "factory_payment"   # 纠正: 这是付工厂货款, 非客户回款
            flagged += 1
        yr = (f.transaction_time.year if f.transaction_time else None) or default_year or date.today().year
        parsed = parse_settlement_remark(f.remark, year=yr)
        if parsed["action"] != "settle" or not parsed["months"]:
            continue
        for m in parsed["months"]:
            if (f.transaction_no, m) in done:
                continue
            r = settle_month(db, supplier=sup, month=m, trigger="keyword",
                             flow_no=f.transaction_no, paid_amount=abs(f.amount or Decimal("0")),
                             note=(f.remark or "")[:200])
            if r.get("flipped"):
                flipped += r["flipped"]
                settled_months.append(m)
                done.add((f.transaction_no, m))
    db.flush()
    return {"flagged": flagged, "settled_months": settled_months, "flipped": flipped}


# ── P4: 漏单检测 (工厂账单没覆盖到的已发货单) ────────────────────
def _is_sample(o) -> bool:
    text = (getattr(o, "product_name", "") or "") + (getattr(o, "sku", "") or "")
    return "样块" in text or "样品" in text


def missing_orders(db: Session, *, up_to_month: Optional[str] = None,
                   supplier: str = DEFAULT_WOOD_SUPPLIER) -> dict:
    """漏单: 已发货真实成交单中, 没被任何工厂账单覆盖的 (按【发货月 ship_date】累计到 up_to_month)。

    覆盖 = 订单号在 FactoryOrder(factory_bill_amount 非空) 或 导入的工厂对账单 FactoryReconItem 里。
    剔除: 补单/刷单(is_refill)、取消/全额退款(settled_sale_clause)、样块(杭州发货)、未发货(无 ship_date)。
    up_to_month("YYYY-MM"): 工厂出X月账单就查 1→X 月全部已发货漏单; None=不设上限。
    返回 {up_to_month, count, total_paid, orders:[...]}。
    """
    from app.models.factory_recon_item import FactoryReconItem
    from app.models.order import FactoryOrder, Order
    from app.services.sales_analytics import settled_sale_clause

    covered: set[str] = set()
    for (ono,) in db.execute(
        select(FactoryOrder.platform_order_no).where(
            FactoryOrder.factory_bill_amount.isnot(None),
            FactoryOrder.platform_order_no.isnot(None),
        )
    ).all():
        if ono:
            covered.add(ono)
    for (ono,) in db.execute(
        select(FactoryReconItem.order_no).where(FactoryReconItem.order_no.isnot(None))
    ).all():
        if ono:
            covered.add(ono)

    rows = db.execute(
        select(Order).where(
            Order.status.in_(("shipped", "signed")),   # 已发货(含已签收)才到开账时点
            Order.is_refill == False,                    # noqa: E712 补单不开账
            Order.is_historical == False,                # noqa: E712
            Order.ship_date.isnot(None),
            settled_sale_clause(),                       # 真实成交(实付>0、非全额退款)
        )
    ).scalars().all()

    orders = []
    total_paid = Decimal("0")
    for o in rows:
        if o.order_no in covered or _is_sample(o):
            continue
        m = o.ship_date.strftime("%Y-%m")
        if up_to_month and m > up_to_month:
            continue
        paid = _d(o.paid_amount)
        total_paid += paid
        orders.append({
            "order_no": o.order_no,
            "product_name": o.product_name,
            "sku": o.sku,
            "qty": o.qty,
            "ship_date": o.ship_date.isoformat() if o.ship_date else None,
            "ship_month": m,
            "order_date": o.order_date.isoformat() if o.order_date else None,
            "paid_amount": str(paid),
            "customer_name": o.customer_name,
        })
    orders.sort(key=lambda x: (x["ship_month"], x["order_no"] or ""))
    return {"supplier": supplier, "up_to_month": up_to_month,
            "count": len(orders), "total_paid": total_paid.quantize(_Q), "orders": orders}


def missing_orders_xlsx_bytes(db: Session, *, up_to_month: Optional[str] = None,
                              supplier: str = DEFAULT_WOOD_SUPPLIER) -> bytes:
    """漏单导出 Excel (财务对账中心 / 工厂对账单 下载用)。"""
    import openpyxl
    from io import BytesIO

    data = missing_orders(db, up_to_month=up_to_month, supplier=supplier)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工厂漏单"
    headers = ["发货月", "订单号", "产品", "SKU", "数量", "发货日", "下单日", "实付", "客户"]
    ws.append(headers)
    for o in data["orders"]:
        ws.append([o["ship_month"], o["order_no"], o["product_name"], o["sku"], o["qty"],
                   o["ship_date"], o["order_date"], o["paid_amount"], o["customer_name"]])
    ws.append([])
    ws.append([f"合计 {data['count']} 单", "", "", "", "", "", "", str(data["total_paid"]), ""])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def settlement_detail_rows(db: Session, supplier: str = DEFAULT_WOOD_SUPPLIER,
                           q: Optional[str] = None) -> list[dict]:
    """逐单明细: 该供应商每张「已开账单」工厂单一行。口径与月度台账完全一致——
    应付(账单) = Σ factory_bill_amount; 已付 = Σ (payment_status='paid' 的 factory_bill_amount)。
    q: 产品名/SKU/产品编码 模糊搜索 (用户 2026-07-03)。"""
    rows = db.execute(
        _apply_product_search(
            select(FactoryOrder).where(
                FactoryOrder.factory_name == supplier,
                FactoryOrder.voided_at.is_(None),
                FactoryOrder.factory_bill_amount.isnot(None),
            ), q)
    ).scalars().all()
    out = []
    for fo in rows:
        amt = _d(fo.factory_bill_amount)
        is_paid = (fo.payment_status or "") == "paid"
        out.append({
            "settlement_month": _order_month(fo) or "(无日期)",
            "factory_order_no": fo.factory_order_no,
            "platform_order_no": fo.platform_order_no,
            "product_name": fo.product_name,
            "product_code": fo.product_code,
            "sku": fo.sku,
            "qty": fo.qty,
            "bill_amount": amt.quantize(_Q),
            "payment_status": "已付" if is_paid else "未付",
            "paid_amount": (amt if is_paid else Decimal("0")).quantize(_Q),
            "payment_date": fo.payment_date.isoformat() if fo.payment_date else None,
            "alipay_flow_no": fo.alipay_flow_no,
            "order_date": fo.order_date.isoformat() if fo.order_date else None,
            "remark": fo.remark,
        })
    out.sort(key=lambda x: (x["settlement_month"], x["factory_order_no"] or ""))
    return out


def settlement_detail_xlsx_bytes(db: Session, supplier: str = DEFAULT_WOOD_SUPPLIER) -> bytes:
    """月结明细导出: Sheet1 月度汇总(应付/已付/未付/单数, 对齐台账口径); Sheet2 逐单明细(每张账单+已付),
    让用户看清「应付 = 该月所有账单额之和; 已付 = 其中已销账(付清)那些单的账单额之和」。"""
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill

    detail = settlement_detail_rows(db, supplier)
    bd = month_breakdown(db, supplier)
    _STAT = {"paid": "已付清", "partial": "部分付清", "unpaid": "未付清"}
    head_fill = PatternFill("solid", fgColor="1F4E78")

    def _style_header(ws, row_idx):
        for cell in ws[row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "月度汇总"
    ws1.append([f"工厂月结 · {supplier}  (应付=该月账单总额; 已付=其中已销账单的账单额之和; 未付=应付−已付)"])
    ws1.append(["结算月", "应付(账单)", "已付", "未付", "单数", "已付单数", "状态"])
    _style_header(ws1, 2)
    for m in bd["months"]:
        ws1.append([m["month"], float(m["billed"]), float(m["paid"]), float(m["unpaid"]),
                    m["order_count"], m["paid_count"], _STAT.get(m["status"], m["status"])])
    ws1.append([])
    ws1.append(["合计", float(bd["total_billed"]), float(bd["total_paid"]), float(bd["total_unpaid"])])
    for i, w in enumerate([12, 14, 12, 12, 8, 10, 10], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws2 = wb.create_sheet("逐单明细")
    hdr = ["结算月", "工厂单号", "平台订单号", "产品", "SKU", "数量",
           "账单金额", "付款状态", "已付金额", "付款日", "销账流水号", "下单日", "备注"]
    ws2.append(hdr)
    _style_header(ws2, 1)
    for r in detail:
        ws2.append([r["settlement_month"], r["factory_order_no"], r["platform_order_no"],
                    r["product_name"], r["sku"], r["qty"], float(r["bill_amount"]),
                    r["payment_status"], float(r["paid_amount"]), r["payment_date"],
                    r["alipay_flow_no"], r["order_date"], r["remark"]])
    ws2.freeze_panes = "A2"
    for i, w in enumerate([10, 18, 20, 22, 18, 6, 12, 10, 12, 12, 22, 12, 24], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

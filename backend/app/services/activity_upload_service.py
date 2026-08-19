"""活动上传编排 (ERP 侧, 2026-07-11 用户拍板): 生成表 → Web-Agent 挂到千牛(stage) →
在系统里出【比对表】(系统要的价 vs 千牛校验) → 用户在系统里确认 → commit 真提交(不可逆)。

- stage(): 不可逆动作前的预演 —— 挂文件+读千牛校验+建比对表, 停在提交前;
- commit(): ★不可逆★ 真点『确认设置/提交』, 只在用户看过比对表点确认后调。

渠道: single_item_discount(单品立减, 已通) / promo_signup(大促报名) / super_reduce(超级立减活动)。
"""
from __future__ import annotations

from sqlalchemy import select
from sqlalchemy.orm import Session

_CHANNELS = {
    "single_item_discount": "单品立减",
    "promo_signup": "大促活动报名",
    "super_reduce": "超级立减活动",
}


def _gen_xlsx(db: Session, channel: str, tier: str) -> tuple[bytes, dict]:
    from app.services import data_export_service as de
    if channel == "single_item_discount":
        if tier == "nosales":                      # 无动销平替档: 到手精确等于 ERP 中促价
            bio, stats = de.build_nosales_single_item_discount_xlsx(db)
        else:
            bio, stats = de.build_single_item_discount_upload_xlsx(db, tier)
    elif channel == "promo_signup":
        normalized = "big" if tier in ("big88", "big88p") else tier
        bio, stats = de.build_promo_signup_upload_xlsx(db, normalized)
    elif channel == "super_reduce":
        bio, stats = de.build_super_reduce_signup_upload_xlsx(db)
    else:
        raise ValueError(f"未知渠道 {channel}")
    return bio.getvalue(), stats


def _f(x):
    return None if x is None else float(x)


def _expand_ids(p) -> list[str]:
    """一码多SKU: [主SKUID, *alt] 去重去空 —— 比对表覆盖每个【真上传】的 SKUID(与 builder 同口径)。"""
    ids: list[str] = []
    for _sid in [p.taobao_sku_id, *(p.alt_taobao_sku_ids or [])]:
        if _sid and str(_sid).strip() not in ids:
            ids.append(str(_sid).strip())
    return ids


def _compare_rows(db: Session, channel: str, tier: str) -> list[dict]:
    """系统侧每个要上传 SKU 的目标值 —— 比对表左半(系统要的价), 右半(千牛校验)由 stage 汇总补。
    promo_signup/super_reduce 直接走 data_export.collect_signup_rows(与 builder 同源: 占位封顶到
    已生效价 + 整商品完整性剔除), 行集与真上传 xlsx 恒等; single_item 逐字镜像其 builder。"""
    from types import SimpleNamespace
    from app.services import campaign_service

    rows: list[dict] = []
    normalized = "big" if tier in ("big88", "big88p") else tier
    if channel in ("promo_signup", "super_reduce"):
        normalized = "mid" if channel == "super_reduce" else normalized
        entries, _stats = campaign_service.build_signup_rows(
            db, SimpleNamespace(tier=normalized))
        targets = campaign_service.target_prices(db, SimpleNamespace(tier=normalized))
        for entry in entries:
            target = targets.get(str(entry["taobao_sku_id"]), {}).get("target")
            rows.append({
                "sku_code": entry["sku_code"],
                "taobao_sku_id": str(entry["taobao_sku_id"]),
                "name": entry["sku_code"],
                "value_label": "活动价(日常价/占位保护价)",
                "system_value": _f(entry["price"]),
                "target_shoudao": _f(target),
            })
        return rows

    if channel == "single_item_discount":
        entries, _stats = campaign_service.build_discount_rows(
            db, SimpleNamespace(tier=normalized))
        for entry in entries:
            rows.append({
                "sku_code": entry["sku_code"],
                "taobao_sku_id": str(entry["taobao_sku_id"]),
                "name": entry["sku_code"],
                "value_label": "立减金额",
                "system_value": _f(entry["deduct"]),
                "target_shoudao": _f(entry["target_price"]),
            })
    return rows


def _parse_uploaded_values(channel: str, xlsx_bytes: bytes) -> dict[str, float]:
    """从【真正要上传的那份 xlsx】里读出每个 SKUID 的上传值(立减金额/报名价/补贴金额)。
    比对表以此为"上传值"列 = 所见即所传, 消除比对表与真实上传文件之间的任何算法漂移。"""
    import io
    import openpyxl
    # super_reduce 现填平台真实模版(两 sheet + 3 行表头), 比对列 = 活动价(col 3), 数据从第 4 行。
    val_col = {"single_item_discount": 3, "promo_signup": 3, "super_reduce": 3}.get(channel)
    data_start = 4 if channel in ("promo_signup", "super_reduce") else 2   # 前 3 行是表头
    if val_col is None:
        return {}
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = (wb["商品SKU导入列表"] if channel in ("promo_signup", "super_reduce")
          and "商品SKU导入列表" in wb.sheetnames else wb.worksheets[0])
    out: dict[str, float] = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or len(row) < val_col:
            continue
        skuid, val = row[1], row[val_col - 1]
        if skuid is None or val is None:
            continue
        try:
            out[str(skuid).strip()] = float(val)
        except (TypeError, ValueError):
            continue
    wb.close()
    return out


# 上传值 vs 系统应填值 判"出入"的浮点保护(1 分钱内视为一致; 超出即严格标红) —— 用户: 必须按系统价
_PRICE_MATCH_EPS = 0.005


def _learn_delisted(db: Session, validation) -> None:
    """从千牛报名回执自动登记下架SKU(自愈, 用户铁律: 在售全报、下架不报 → 下次报名不带它)。"""
    if not validation:
        return
    try:
        from app.services import delisted_sku_service
        ids = delisted_sku_service.extract_delisted_from_feedback(validation.get("failed_items"))
        if ids:
            delisted_sku_service.add_delisted(db, ids)
    except Exception:  # noqa: BLE001  自愈是尽力而为, 失败不影响主流程
        pass


def stage(db: Session, channel: str, tier: str = "big",
          start_dt: str | None = None, end_dt: str | None = None) -> dict:
    """挂文件到千牛(不提交) + 建比对表(上传值 vs 系统应填值, 0 容差核对)。
    start_dt/end_dt(单品立减专用, 'YYYY-MM-DD HH:MM:SS'): 给了则把千牛『活动时间』填成该精确档期。
    返回 {ok, compare_rows, price_match_ok, mismatch_count, validation, screenshot_base64, ...}。"""
    from app.services import web_agent_service
    if channel not in _CHANNELS:
        return {"ok": False, "error": f"未知渠道 {channel}"}
    if channel in ("promo_signup", "super_reduce"):
        return {
            "ok": False,
            "step": "execution_policy_guard",
            "error": "活动报名手工上传入口已禁用；只允许活动计划自动程序执行",
            "ai_may_adjust_or_resubmit": False,
        }
    xlsx, stats = _gen_xlsx(db, channel, tier)
    # ★核对: system_value=系统应填(独立重算) vs uploaded_value=真正要上传那份 xlsx 里的数; 差>1分即标出入。
    rows = _compare_rows(db, channel, tier)
    uploaded = _parse_uploaded_values(channel, xlsx)
    mismatch_n = 0
    for row in rows:
        up = uploaded.get(row["taobao_sku_id"])
        sysv = row.get("system_value")
        bad = (up is None) or (sysv is None) or (abs(up - sysv) > _PRICE_MATCH_EPS)
        row["uploaded_value"] = up
        row["mismatch"] = bool(bad)
        if bad:
            mismatch_n += 1
    # 反向兜底: 任何【真上传了却没进比对表】的 SKUID 也当"出入"红行标出(两路未来再分叉也兜得住, commit 不可逆)。
    covered = {row["taobao_sku_id"] for row in rows}
    for skuid, up in uploaded.items():
        if skuid not in covered:
            rows.append({
                "sku_code": "?", "taobao_sku_id": skuid, "name": "⚠️ 仅上传·未进比对(请核实此SKUID)",
                "value_label": "", "system_value": None, "target_shoudao": None,
                "uploaded_value": up, "mismatch": True,
            })
            mismatch_n += 1
    j = web_agent_service.upload_file(db, channel, "stage", xlsx, f"{channel}.xlsx",
                                      start_dt=start_dt, end_dt=end_dt)
    if not j.get("ok") or not j.get("job"):
        return {"ok": False, "error": j.get("error", "取数服务(:8500)未响应, 无法上传")}
    final = web_agent_service.wait_job(db, j["job"], timeout_s=200)
    res = final.get("result") or {}
    if res.get("need_scan"):
        return {"ok": False, "need_scan": True, "message": "淘宝登录态过期, 请先扫码后再上传"}
    _learn_delisted(db, res.get("validation"))   # 自愈: 千牛回"已下架SKU=X"→登记, 下次报名不带它
    return {
        "ok": bool(res.get("ok")), "channel": channel,
        "channel_name": _CHANNELS[channel], "gen_stats": stats,
        "validation": res.get("validation"),
        "screenshot_base64": res.get("screenshot_base64"),
        "stopped_before": res.get("stopped_before"),
        "compare_rows": rows,
        "price_match_ok": mismatch_n == 0,
        "mismatch_count": mismatch_n,
        "compare_total": len(rows),
    }


def commit(db: Session, channel: str, tier: str = "big",
           start_dt: str | None = None, end_dt: str | None = None) -> dict:
    """★不可逆★ 真提交到千牛。只在用户看过比对表、系统里点确认后调用。
    start_dt/end_dt(单品立减专用): 给了则真提交时把千牛『活动时间』填成该精确档期。
    super_reduce 是逐商品原地改价(30+商品×~25s, 远超HTTP等待) → 发起后立即返回
    {async_job}, 前端用 /activity-upload/commit-status 轮询取结果。"""
    from app.services import web_agent_service
    if channel not in _CHANNELS:
        return {"ok": False, "error": f"未知渠道 {channel}"}
    xlsx, _stats = _gen_xlsx(db, channel, tier)
    j = web_agent_service.upload_file(db, channel, "commit", xlsx, f"{channel}.xlsx",
                                      start_dt=start_dt, end_dt=end_dt)
    if not j.get("ok") or not j.get("job"):
        return {"ok": False, "error": j.get("error", "取数服务(:8500)未响应")}
    if channel == "super_reduce":
        return {"ok": True, "channel": channel, "channel_name": _CHANNELS[channel],
                "async_job": j["job"],
                "note": "已开始逐商品原地改价(首商品金丝雀失败即中止), 请轮询状态"}
    final = web_agent_service.wait_job(db, j["job"], timeout_s=200)
    res = final.get("result") or {}
    return {
        "ok": bool(res.get("submitted")), "channel": channel,
        "channel_name": _CHANNELS[channel], "submitted": res.get("submitted"),
        "validation": res.get("validation"),
        "screenshot_base64": res.get("screenshot_base64"),
    }


def commit_status(db: Session, job_id: str) -> dict:
    """轮询超级立减逐商品改价 job: 返回 {status: running|done|error, result?}。"""
    from app.services import web_agent_service
    j = web_agent_service.get_job(db, job_id)
    status = j.get("status") or ("done" if j.get("result") else "running")
    out = {"status": status}
    if status == "error":
        out["error"] = j.get("error")
    res = j.get("result") or {}
    if res:
        out["result"] = {
            "ok": res.get("ok"), "submitted": res.get("submitted"),
            "message": res.get("message"),
            "validation": res.get("validation"),
            "results": res.get("results"),
            "screenshot_base64": res.get("screenshot_base64"),
        }
    return out


def product_price_auto_push(db: Session) -> dict:
    """★全自动推标价编排 (2026-07-14 全通): WA触发千牛「excel商品批量导出」→ 从下载中心下载发布模版
    → 系统把一口价改成 ERP日常价÷0.75(单品宝配套) → WA上传千牛 excel商品批量编辑(import tab)
    → 停在提交前(最终"提交"用户点)。返回 {ok, step, modify_stats, attached, screenshot_base64, note}。"""
    from app.services import data_export_service as de
    from app.services import web_agent_service
    exp = web_agent_service.export_product_prices(db)
    if not exp.get("ok"):
        return {"ok": False, "step": "export", "need_scan": exp.get("need_scan"),
                "error": exp.get("error") or exp.get("message"),
                "screenshot_base64": exp.get("screenshot_base64")}
    bio, stats = de.build_product_price_upload_from_export(db, exp["xlsx_bytes"])
    if stats.get("rows", 0) < 20:   # 疑似抓到旧的小导出(如样块8行) → 提示先手动导出全部商品
        return {"ok": False, "step": "stale_export", "modify_stats": stats,
                "error": f"下载到的千牛导出只有 {stats.get('rows')} 行(疑似旧的小导出)。"
                         "请先在千牛「导出全部商品」等生成好, 再点本按钮; 系统会抓最新那份改价推回。",
                "export_filename": exp.get("filename")}
    if stats.get("changed", 0) == 0:
        return {"ok": True, "step": "no_change", "modify_stats": stats,
                "note": "千牛一口价已全部=日常价÷0.75, 无需改动"}
    up = web_agent_service.upload_file(db, "product_prices", "stage", bio.getvalue(),
                                       exp.get("filename") or "product_prices.xlsx")
    if not up.get("ok") or not up.get("job"):
        return {"ok": False, "step": "upload", "error": up.get("error"), "modify_stats": stats}
    final = web_agent_service.wait_job(db, up["job"], timeout_s=200)
    res = final.get("result") or {}
    return {"ok": bool(res.get("ok") or res.get("attached")), "step": "staged",
            "modify_stats": stats, "attached": res.get("attached"),
            "validation": res.get("validation"),
            "screenshot_base64": res.get("screenshot_base64"), "note": res.get("note")}

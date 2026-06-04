"""全量数据备份服务.

功能:
  1. 把系统所有主要数据表导出为一个 Excel 文件 (每张表一个 Sheet)
  2. 保存到本地目录 (默认 /data/backups), 保留最新 MAX_BACKUPS 份, 自动删除旧文件
  3. 可选: 上传到 S3 兼容存储 (需设置环境变量 PANSE_BACKUP_S3_*)

备份计划 (scheduler.py 注册):
  - 每 6 天凌晨 02:00 执行一次
  - 60 份 × 6 天 = 360 天覆盖 (约一年)

文件大小估算:
  - 每份 Excel: 1-5 MB (视数据量)
  - 60 份合计: 60-300 MB, 对群晖 NAS 完全无压力

环境变量:
  PANSE_BACKUP_DIR          本地备份目录, 默认 /data/backups
  PANSE_BACKUP_MAX          保留份数, 默认 60
  PANSE_BACKUP_S3_ENDPOINT  S3 endpoint URL (如 https://s3.ap-east-1.amazonaws.com)
  PANSE_BACKUP_S3_BUCKET    S3 bucket 名称
  PANSE_BACKUP_S3_KEY       Access key ID
  PANSE_BACKUP_S3_SECRET    Secret access key
  PANSE_BACKUP_S3_PREFIX    S3 key 前缀, 默认 panse-erp-backups/
"""
from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session

_logger = logging.getLogger("panse.backup")

BACKUP_DIR = os.environ.get("PANSE_BACKUP_DIR", "/data/backups")
MAX_BACKUPS = int(os.environ.get("PANSE_BACKUP_MAX", "60"))


# ─────────────────────────── Excel 导出 ─────────────────────────── #


def _rows_from_model(db: Session, model_cls) -> tuple[list[str], list[list]]:
    """从 SQLAlchemy 模型查出所有行, 返回 (列名列表, 数据行列表)。"""
    rows = db.query(model_cls).all()
    if not rows:
        cols = [c.name for c in model_cls.__table__.columns]
        return cols, []
    cols = [c.name for c in model_cls.__table__.columns]
    data = [[getattr(r, c) for c in cols] for r in rows]
    return cols, data


def _add_sheet(wb, sheet_name: str, cols: list[str], data: list[list]) -> None:
    import openpyxl.styles as _styles
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name <= 31 chars
    header_font = _styles.Font(bold=True)
    ws.append(cols)
    for cell in ws[1]:
        cell.font = header_font
    for row in data:
        # Convert non-serializable types to string
        safe_row = []
        for v in row:
            if v is None:
                safe_row.append("")
            elif isinstance(v, (int, float, str, bool)):
                safe_row.append(v)
            else:
                safe_row.append(str(v))
        ws.append(safe_row)


def export_all(db: Session, output_dir: str = BACKUP_DIR) -> Path:
    """导出所有主要数据表到一个 Excel 文件, 返回文件路径."""
    import openpyxl

    from app.models.auth import User
    from app.models.bom import BomLine
    from app.models.customer import Customer
    from app.models.finance import (
        AccountBalance, AlipayFlow, FactoryReconciliation,
        LogisticsBill, RefillRecord, WanshifuBill,
    )
    from app.models.inventory import PartInventory, ProductInventory
    from app.models.marketing import (
        AfterSales, BrandMarketing, DailyOperation,
        OutsourcingExpense, PromotionFlow, Sample, WoodLoss,
    )
    from app.models.material import Material
    from app.models.order import FactoryOrder, Order, OrderDetail, PartPurchase
    from app.models.pricing import PricingSku
    from app.models.product import Product
    from app.models.supplier import Supplier

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"panse_backup_{ts}.xlsx"
    filepath = Path(output_dir) / filename

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Table → sheet name mapping
    sheet_defs: list[tuple[str, object]] = [
        ("产品", Product),
        ("定价SKU", PricingSku),
        ("BOM", BomLine),
        ("物料", Material),
        ("成品库存", ProductInventory),
        ("配件库存", PartInventory),
        ("销售订单", Order),
        ("工厂下单", FactoryOrder),
        ("工厂对账", FactoryReconciliation),
        ("订单细节", OrderDetail),
        ("配件采购", PartPurchase),
        ("支付宝流水", AlipayFlow),
        ("账户余额", AccountBalance),
        ("物流账单", LogisticsBill),
        ("万师傅账单", WanshifuBill),
        ("补单对账", RefillRecord),
        ("样品", Sample),
        ("木材损耗", WoodLoss),
        ("品牌营销", BrandMarketing),
        ("推广记录", PromotionFlow),
        ("日常经营", DailyOperation),
        ("人员外包", OutsourcingExpense),
        ("售后", AfterSales),
        ("客户", Customer),
        ("供应商", Supplier),
        ("用户", User),
    ]

    total_rows = 0
    for sheet_name, model_cls in sheet_defs:
        try:
            cols, data = _rows_from_model(db, model_cls)
            _add_sheet(wb, sheet_name, cols, data)
            total_rows += len(data)
            _logger.debug("备份表 %s: %d 行", sheet_name, len(data))
        except Exception as e:
            _logger.warning("备份表 %s 失败, 跳过: %s", sheet_name, e)

    # Metadata sheet
    meta_ws = wb.create_sheet(title="备份信息", index=0)
    meta_ws.append(["字段", "值"])
    meta_ws.append(["备份时间", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    meta_ws.append(["系统版本", "Panse ERP"])
    meta_ws.append(["总行数", total_rows])
    meta_ws.append(["包含表数", len(sheet_defs)])

    wb.save(filepath)
    _logger.info("备份已保存: %s (共 %d 行)", filepath, total_rows)
    return filepath


# ─────────────────────────── 文件轮换 ─────────────────────────── #


def rotate(backup_dir: str = BACKUP_DIR, max_count: int = MAX_BACKUPS) -> int:
    """删除超出 max_count 的最旧备份文件, 返回删除数量."""
    files = sorted(Path(backup_dir).glob("panse_backup_*.xlsx"), key=lambda p: p.stat().st_mtime)
    to_delete = files[: max(0, len(files) - max_count)]
    for f in to_delete:
        try:
            f.unlink()
            _logger.info("删除旧备份: %s", f.name)
        except Exception as e:
            _logger.warning("删除备份失败 %s: %s", f.name, e)
    return len(to_delete)


# ─────────────────────────── 云端上传 ─────────────────────────── #


def _upload_s3(filepath: Path) -> bool:
    """上传到 S3 兼容存储 (需配置 PANSE_BACKUP_S3_* 环境变量). 失败仅记录警告, 不中断备份。"""
    endpoint = os.environ.get("PANSE_BACKUP_S3_ENDPOINT", "")
    bucket = os.environ.get("PANSE_BACKUP_S3_BUCKET", "")
    key_id = os.environ.get("PANSE_BACKUP_S3_KEY", "")
    secret = os.environ.get("PANSE_BACKUP_S3_SECRET", "")
    prefix = os.environ.get("PANSE_BACKUP_S3_PREFIX", "panse-erp-backups/")

    if not all([endpoint, bucket, key_id, secret]):
        return False  # 未配置, 静默跳过

    try:
        import boto3  # type: ignore[import]
        from botocore.client import Config  # type: ignore[import]

        s3 = boto3.client(
            "s3",
            endpoint_url=endpoint,
            aws_access_key_id=key_id,
            aws_secret_access_key=secret,
            config=Config(signature_version="s3v4"),
        )
        s3_key = f"{prefix}{filepath.name}"
        s3.upload_file(str(filepath), bucket, s3_key)
        _logger.info("备份已上传 S3: s3://%s/%s", bucket, s3_key)
        return True
    except ImportError:
        _logger.warning("boto3 未安装, 跳过 S3 上传 (pip install boto3 即可启用)")
        return False
    except Exception as e:
        _logger.warning("S3 上传失败 (本地备份仍保留): %s", e)
        return False


# ─────────────────────────── 主入口 ─────────────────────────── #


def run(db: Session, output_dir: Optional[str] = None) -> dict:
    """调度器 / 手动触发的主入口: 导出 → 轮换 → 可选 S3 上传."""
    out = output_dir or BACKUP_DIR
    filepath = export_all(db, output_dir=out)
    deleted = rotate(backup_dir=out)
    uploaded = _upload_s3(filepath)
    size_mb = round(filepath.stat().st_size / 1024 / 1024, 2)
    return {
        "file": filepath.name,
        "size_mb": size_mb,
        "deleted_old": deleted,
        "uploaded_s3": uploaded,
    }


def list_backups(backup_dir: Optional[str] = None) -> list[dict]:
    """返回本地备份文件列表 (供 API 查询)."""
    backup_dir = backup_dir or BACKUP_DIR
    d = Path(backup_dir)
    if not d.exists():
        return []
    files = sorted(
        d.glob("panse_backup_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return [
        {
            "filename": f.name,
            "size_mb": round(f.stat().st_size / 1024 / 1024, 2),
            "created_at": datetime.fromtimestamp(f.stat().st_mtime, tz=timezone.utc).isoformat(),
        }
        for f in files
    ]


# ─────────────────────────── 可配置的自动备份 ─────────────────────────── #
#
# 配置存 system_settings, 后台可改 (默认每 7 天):
#   backup_auto_enabled   "1"/"0"   是否开启自动备份 (默认开)
#   backup_interval_days  int       间隔天数 (默认 7)
#   backup_dir            str       备份保存目录 (默认 PANSE_BACKUP_DIR / /data/backups)
#   backup_start_date     YYYY-MM-DD 起始日期 (可选, 自动备份不早于此日)
#   backup_last_run_at    ISO       上次自动备份时间 (系统维护, 勿手填)

_CFG_DEFAULTS = {
    "backup_auto_enabled": "1",
    "backup_interval_days": "7",
    "backup_dir": BACKUP_DIR,
    "backup_start_date": "",
    "backup_last_run_at": "",
}


def get_config(db: Session) -> dict:
    from app.services import settings_service
    g = lambda k: settings_service.get(db, k, env_fallback=False) or _CFG_DEFAULTS[k]
    try:
        interval = max(1, int(g("backup_interval_days")))
    except (TypeError, ValueError):
        interval = 7
    enabled = g("backup_auto_enabled") not in ("0", "false", "False", "")
    bdir = settings_service.get(db, "backup_dir", env_fallback=False) or BACKUP_DIR
    last = settings_service.get(db, "backup_last_run_at", env_fallback=False) or None
    start = settings_service.get(db, "backup_start_date", env_fallback=False) or None
    return {
        "auto_enabled": enabled,
        "interval_days": interval,
        "dir": bdir,
        "start_date": start,
        "last_run_at": last,
        "next_run_at": _next_run_at(last, start, interval),
        "max_backups": MAX_BACKUPS,
    }


def set_config(db: Session, *, auto_enabled: Optional[bool] = None,
               interval_days: Optional[int] = None, dir: Optional[str] = None,
               start_date: Optional[str] = None) -> dict:
    from app.services import settings_service
    if auto_enabled is not None:
        settings_service.set_value(db, "backup_auto_enabled", "1" if auto_enabled else "0")
    if interval_days is not None:
        settings_service.set_value(db, "backup_interval_days", str(max(1, int(interval_days))))
    if dir is not None:
        settings_service.set_value(db, "backup_dir", dir.strip())
    if start_date is not None:
        settings_service.set_value(db, "backup_start_date", start_date.strip())
    db.commit()
    return get_config(db)


def _next_run_at(last_iso: Optional[str], start: Optional[str], interval: int) -> Optional[str]:
    from datetime import timedelta
    base = None
    if last_iso:
        try:
            base = datetime.fromisoformat(last_iso)
        except ValueError:
            base = None
    if base is None and start:
        try:
            base = datetime.fromisoformat(start)
        except ValueError:
            base = None
    if base is None:
        return None
    return (base + timedelta(days=interval)).isoformat()


def run_if_due(db: Session) -> dict:
    """调度器每日调用: 仅当距上次自动备份已满 interval_days 才执行。"""
    from datetime import date

    cfg = get_config(db)
    if not cfg["auto_enabled"]:
        return {"skipped": "自动备份已关闭"}

    today = datetime.now(timezone.utc).date()
    # 起始日期之前不跑
    if cfg["start_date"]:
        try:
            sd = datetime.fromisoformat(cfg["start_date"]).date()
            if today < sd:
                return {"skipped": f"未到起始日期 {cfg['start_date']}"}
        except ValueError:
            pass

    last_iso = cfg["last_run_at"]
    if last_iso:
        try:
            last_date = datetime.fromisoformat(last_iso).date()
            if (today - last_date).days < cfg["interval_days"]:
                return {"skipped": f"距上次备份不足 {cfg['interval_days']} 天",
                        "last_run_at": last_iso}
        except ValueError:
            pass

    result = run(db, output_dir=cfg["dir"])
    from app.services import settings_service
    settings_service.set_value(db, "backup_last_run_at",
                               datetime.now(timezone.utc).isoformat())
    db.commit()
    result["ran"] = True
    return result

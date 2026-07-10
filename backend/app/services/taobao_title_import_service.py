"""淘宝商品导出 → 定价表 taobao_title 回填 (用户拍板 2026-06-18)。

痛点: 部分订单导入只带宝贝长标题、不带商家编码 (product_code/sku_code 全空), 成本引擎
无法对到定价表 → 只能按 实付×类目成本率 估算, 比例靠不住。用户提供「畔色产品导出.xlsx」
(淘宝后台商品导出: 宝贝标题 ↔ 商家编码 ↔ SKU商家编码), 把宝贝标题写进定价表 PricingSku.taobao_title,
订单即可按 product_name == taobao_title 精确匹配回填编码, 走定价表真实成本。

导出表列 (第 3 行为表头, 数据从第 4 行起):
  col0  商品Id        col3  宝贝标题       col6  商家编码(宝贝级=product_code)
  col11 skuId         col15 商家编码(SKU级=sku_code)

回填策略 (幂等):
  1) 精确: export.sku_code → pricing_sku.sku_code 命中的行写 taobao_title;
  2) 兜底: 仍空的 pricing_sku 行按 product_code 取该宝贝标题 (标题是宝贝级, 同产品共用)。
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.pricing import PricingSku

_logger = logging.getLogger("panse.taobao_title")

# 导出表列下标 (0-based), 数据起始行 (1-based, 表头占前 3 行)
_COL_PRODUCT_CODE = 6
_COL_TITLE = 3
_COL_SKU_CODE = 15
_DATA_START_ROW = 4


@dataclass
class TitleRow:
    product_code: str
    sku_code: str
    title: str


@dataclass
class ImportResult:
    parsed_rows: int = 0
    by_sku_code: int = 0          # 精确按 SKU 编码命中回填
    by_product_code: int = 0      # 兜底按宝贝编码回填
    listed_marked: int = 0        # 顺带标记为「在售」的产品数 (在售导出=见到即在售)
    distinct_titles: int = 0
    unmatched_titles: list[str] = field(default_factory=list)  # 导出表有、定价表无的宝贝(供人工补 SKU)


def parse_export_xlsx(raw: bytes) -> list[TitleRow]:
    """解析淘宝商品导出 xlsx → [TitleRow]。只取有标题且有编码(宝贝或SKU)的行。"""
    import io

    import openpyxl

    # read_only=False: 淘宝导出是流式zip(无dimension元数据), read_only 模式 iter_rows 会一行都读不出
    # (2026-07-10 实测静默回填0行)。导出才几百行, 全量加载无压力。
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
    ws = wb.worksheets[0]
    rows: list[TitleRow] = []
    for r in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
        def _cell(i: int) -> str:
            return str(r[i]).strip() if i < len(r) and r[i] is not None else ""
        title = _cell(_COL_TITLE)
        pcode = _cell(_COL_PRODUCT_CODE)
        scode = _cell(_COL_SKU_CODE)
        if not title or (not pcode and not scode):
            continue
        rows.append(TitleRow(product_code=pcode, sku_code=scode, title=title))
    wb.close()
    return rows


def _norm_code(code: str) -> str:
    """历史 P+数字 编码统一成 PPS (与订单/同步口径一致), 否则导出表是 P 开头时对不上。"""
    c = (code or "").strip()
    if len(c) >= 2 and c[0] in ("P", "p") and c[1].isdigit():
        return "PPS" + c[1:]
    return c


def import_titles(db: Session, rows: list[TitleRow]) -> ImportResult:
    """把导出表宝贝标题写进 pricing_sku.taobao_title。幂等 (重复跑结果一致)。"""
    res = ImportResult(parsed_rows=len(rows))

    sku2title: dict[str, str] = {}
    pc2title: dict[str, str] = {}
    for row in rows:
        title = row.title.strip()
        if not title:
            continue
        sc = _norm_code(row.sku_code)
        pc = _norm_code(row.product_code)
        if sc:
            sku2title.setdefault(sc, title)
        if pc:
            pc2title.setdefault(pc, title)
    res.distinct_titles = len(set(sku2title.values()) | set(pc2title.values()))

    ps_rows = db.execute(select(PricingSku)).scalars().all()
    matched_pcodes: set[str] = set()
    for ps in ps_rows:
        # 1) 精确: SKU 编码命中
        t = sku2title.get((ps.sku_code or "").strip())
        if t:
            if ps.taobao_title != t:
                ps.taobao_title = t
            res.by_sku_code += 1
            matched_pcodes.add((ps.product_code or "").strip())
            continue
        # 2) 兜底: 宝贝编码命中
        t = pc2title.get((ps.product_code or "").strip())
        if t:
            if ps.taobao_title != t:
                ps.taobao_title = t
            res.by_product_code += 1
            matched_pcodes.add((ps.product_code or "").strip())

    # 导出表里有宝贝、定价表却完全没有对应行的标题 (这些宝贝没建定价 SKU, 订单也对不到成本)
    res.unmatched_titles = sorted({
        title for pc, title in pc2title.items() if pc not in matched_pcodes
    })

    # 上架状态回填 (2026-07-10 用户需求): 在售导出=千牛「出售中」页面导出, 出现即在售。
    # 只把「见到的」置为 在售, 绝不反向标下架 (导出按类目分多份, 单份缺席≠下架; 下架走产品编辑手改)。
    from app.models.product import Product
    seen_pcodes = {pc for pc in pc2title} | {pc for pc in matched_pcodes if pc}
    if seen_pcodes:
        for prod in db.execute(select(Product).where(Product.code.in_(seen_pcodes))).scalars():
            if prod.listing_status != "在售":
                prod.listing_status = "在售"
                res.listed_marked += 1
    db.flush()
    _logger.info("淘宝标题导入: 解析%d 行, 按SKU回填%d, 按宝贝回填%d, 未建定价宝贝%d",
                 res.parsed_rows, res.by_sku_code, res.by_product_code, len(res.unmatched_titles))
    return res


def import_from_xlsx_bytes(db: Session, raw: bytes) -> ImportResult:
    """上传入口: 解析 xlsx → 回填 pricing_sku.taobao_title。"""
    rows = parse_export_xlsx(raw)
    return import_titles(db, rows)

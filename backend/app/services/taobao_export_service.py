"""淘宝批量操作表导出 — 用系统数据填好淘宝后台批量格式.

四种导出:
  - export_price_publish  : 价格批量发布 (宝贝标题/商家编码/一口价/SKU 价格)
  - export_single_discount: 单品立减 (商品ID/SKU_ID/立减金额)
  - export_promo_signup    : 大促活动报名 (商品ID/报名价)
  - export_product_info    : 商品信息批量 (商品ID/标题/描述/类目)

数据来源 (按 sku 维度):
  PricingSku ⨝ Product (on product_code) ⨝ TaobaoListing (on sku_code)
TaobaoListing 提供 淘宝商品ID / skuId; 无匹配时商品ID留空, 计入 warning。

每个导出函数复用对应模板文件的表头行 (read template, copy header rows),
若模板缺失则回退到内置表头 (已在各函数文档说明)。
返回值为 (xlsx_bytes, stats) 其中 stats = {"total": n, "matched": m, "unmatched": k}。
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.pricing import PricingSku
from app.models.product import Product
from app.models.taobao_listing import TaobaoListing

_TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "assets" / "taobao_templates"

# 各导出类型对应的模板文件 + 数据起始行 (1-based, 数据从该行开始写, 之前的行为表头并保留)
# 表头占用行数 = data_start_row - 1
_PRICE_PUBLISH_TPL = ("product_publish.xlsx", 4)        # 表头 3 行 (注意/分组/列名), 数据从第4行
_SINGLE_DISCOUNT_TPL = ("single_item_discount.xlsx", 2)  # 表头 1 行 (列名), 数据从第2行
_PROMO_SIGNUP_TPL = ("promo_signup.xlsx", 4)            # 表头 3 行 (分组/列名/说明), 数据从第4行
_PRODUCT_INFO_TPL = ("product_export_mapping.xlsx", 4)  # 表头 3 行 (注意/分组/列名), 数据从第4行

# 模板缺失时的回退表头 (单行)
_FALLBACK_PRICE_PUBLISH = ["商品id", "宝贝标题", "导购标题", "商家编码", "一口价",
                           "发货时间（单位：天）", "skuId", "价格（元）", "数量", "商家编码"]
_FALLBACK_SINGLE_DISCOUNT = ["商品id", "SKU_ID", "优惠值", "打折抹分取整", "提醒"]
_FALLBACK_PROMO_SIGNUP = ["商品ID", "商品名称", "营销ID", "商品状态", "一口价",
                          "大促活动价", "线上库存"]
_FALLBACK_PRODUCT_INFO = ["商品Id", "类目id", "类目名称", "宝贝标题", "一口价",
                          "导购标题", "商家编码", "发货时间", "最长发货时间",
                          "销售属性", "属性对", "skuId", "价格(元)", "库存(件)"]


def _copy_template_header(tpl_name: str, data_start_row: int, fallback: list) -> Workbook:
    """读取模板文件并保留其前 data_start_row-1 行表头, 删除其余示例数据。

    返回一个 active sheet 已就绪、可从 data_start_row 行开始写数据的 Workbook。
    模板缺失则用 fallback 表头新建。
    """
    path = _TEMPLATE_DIR / tpl_name
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(fallback)
        return wb

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # 删除表头以下的所有示例数据行
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)
    return wb


def _fetch_rows(db: Session, category: Optional[str], product_codes: Optional[list]):
    """返回 [(PricingSku, Product|None, TaobaoListing|None), ...]。

    PricingSku 为基准, 左联 Product(on product_code), 左联 TaobaoListing(on sku_code)。
    category 过滤 Product.category; product_codes 过滤 PricingSku.product_code。
    """
    stmt = (
        select(PricingSku, Product, TaobaoListing)
        .outerjoin(Product, Product.code == PricingSku.product_code)
        .outerjoin(TaobaoListing, TaobaoListing.sku_code == PricingSku.sku_code)
    )
    if category:
        stmt = stmt.where(Product.category == category)
    if product_codes:
        stmt = stmt.where(PricingSku.product_code.in_(product_codes))
    return db.execute(stmt).all()


def _save(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _num(v):
    """Decimal -> float for openpyxl numeric cell, None passthrough."""
    return float(v) if v is not None else None


# ---------------------------------------------------------------------------
# 1. 价格批量发布
# 模板列 (row3): 商品id|宝贝标题|导购标题|商家编码|一口价|发货时间|skuId|价格(元)|数量|商家编码
# ---------------------------------------------------------------------------
def export_price_publish(db: Session, *, category: Optional[str] = None,
                         product_codes: Optional[list] = None) -> bytes:
    wb = _copy_template_header(*_PRICE_PUBLISH_TPL, _FALLBACK_PRICE_PUBLISH)
    ws = wb.active
    rows = _fetch_rows(db, category, product_codes)
    matched = 0
    for pricing, product, listing in rows:
        if listing is not None:
            matched += 1
        merchant_code = (listing.merchant_code if listing else None) or pricing.sku_code
        ws.append([
            listing.taobao_item_id if listing else None,           # 商品id
            product.name if product else None,                     # 宝贝标题
            None,                                                  # 导购标题
            merchant_code,                                         # 商家编码
            _num(pricing.list_price),                              # 一口价
            None,                                                  # 发货时间
            listing.taobao_sku_id if listing else None,            # skuId
            _num(pricing.daily_price),                             # SKU 价格 (=淘宝活动报名价=日常价)
            None,                                                  # 数量
            merchant_code,                                         # 商家编码 (SKU 段)
        ])
    return _save(wb)


# ---------------------------------------------------------------------------
# 2. 单品立减
# 模板列 (row1): 商品id | SKU_ID | 优惠值 | 打折抹分取整 | 提醒
# 立减金额 = daily_price - small_promo (两者都有时, 否则留空)
# ---------------------------------------------------------------------------
def export_single_discount(db: Session, *, category: Optional[str] = None,
                           product_codes: Optional[list] = None) -> bytes:
    wb = _copy_template_header(*_SINGLE_DISCOUNT_TPL, _FALLBACK_SINGLE_DISCOUNT)
    ws = wb.active
    rows = _fetch_rows(db, category, product_codes)
    for pricing, product, listing in rows:
        discount = None
        if pricing.daily_price is not None and pricing.small_promo is not None:
            discount = _num(pricing.daily_price - pricing.small_promo)
        ws.append([
            listing.taobao_item_id if listing else None,   # 商品id
            listing.taobao_sku_id if listing else None,     # SKU_ID
            discount,                                       # 优惠值 (立减金额)
            None,                                           # 打折抹分取整
            None,                                           # 提醒
        ])
    return _save(wb)


# ---------------------------------------------------------------------------
# 3. 大促活动报名
# 模板列 (row1): 商品ID|商品名称|营销ID|商品状态|一口价|大促活动价|线上库存|...(玩法/秒杀活动价)
# 报名价 = daily_price (=淘宝活动报名价); 秒杀活动价段也填报名价。
# ---------------------------------------------------------------------------
def export_promo_signup(db: Session, *, category: Optional[str] = None,
                        product_codes: Optional[list] = None) -> bytes:
    wb = _copy_template_header(*_PROMO_SIGNUP_TPL, _FALLBACK_PROMO_SIGNUP)
    ws = wb.active
    width = ws.max_column or len(_FALLBACK_PROMO_SIGNUP)
    rows = _fetch_rows(db, category, product_codes)
    for pricing, product, listing in rows:
        signup_price = _num(pricing.daily_price)
        row = [None] * width
        if width >= 1:
            row[0] = listing.taobao_item_id if listing else None   # 商品ID
        if width >= 2:
            row[1] = product.name if product else None             # 商品名称
        if width >= 5:
            row[4] = _num(pricing.list_price)                      # 一口价
        if width >= 6:
            row[5] = signup_price                                  # 大促活动价 (报名价)
        # 各玩法的「秒杀活动价(必填)」列 (第10,14,18,22,26 列, 1-based) 填报名价
        for idx in (9, 13, 17, 21, 25):
            if idx < width:
                row[idx] = signup_price
        ws.append(row)
    return _save(wb)


# ---------------------------------------------------------------------------
# 4. 商品信息批量
# 模板列 (row3): 商品Id|类目id|类目名称|宝贝标题|一口价|导购标题|商家编码|发货时间|
#                最长发货时间|销售属性|属性对|skuId|价格(元)|库存(件)|...
# 商品ID=taobao_item_id 标题=product.name 描述=product.description 类目=product.category
# ---------------------------------------------------------------------------
def export_product_info(db: Session, *, category: Optional[str] = None,
                        product_codes: Optional[list] = None) -> bytes:
    wb = _copy_template_header(*_PRODUCT_INFO_TPL, _FALLBACK_PRODUCT_INFO)
    ws = wb.active
    width = ws.max_column or len(_FALLBACK_PRODUCT_INFO)
    rows = _fetch_rows(db, category, product_codes)
    for pricing, product, listing in rows:
        merchant_code = (listing.merchant_code if listing else None) or pricing.sku_code
        row = [None] * width
        if width >= 1:
            row[0] = listing.taobao_item_id if listing else None   # 商品Id
        if width >= 3:
            row[2] = product.category if product else None         # 类目名称
        if width >= 4:
            # 宝贝标题: 优先系统产品名; 描述附在此 (模板无独立描述列)
            row[3] = product.name if product else None             # 宝贝标题
        if width >= 5:
            row[4] = _num(pricing.list_price)                      # 一口价
        if width >= 7:
            row[6] = merchant_code                                 # 商家编码
        if width >= 12:
            row[11] = listing.taobao_sku_id if listing else None   # skuId
        if width >= 13:
            row[12] = _num(pricing.daily_price)                    # 价格(元)
        # 描述 (产品文案) 无对应模板列, 放在末尾一列以免丢失
        if product and product.description:
            row[width - 1] = product.description
        ws.append(row)
    return _save(wb)

# -*- coding: utf-8 -*-
"""万师傅安装订单档案: 导入 (38列订单导出) + 淘宝订单配对 + 批注导出。

格式 (用户确认 2026-06-11 起为默认): 表头 3 行 (r1 分组 / r2 字段名 / r3 地址子表头),
数据从第 4 行起。表里没有淘宝订单号, 配对靠多层启发式:
    A phone_full   手机号全等 (虚拟分机号原样)
    B phone_base   手机号主段 (去 -分机)
    D track        物流单号
    C1 name_city   收货姓名全等 + 订单地址含 区名或市名
    C2 name_unique 收货姓名在订单库唯一
唯一命中 → matched_order_no; 多候选 → match_method='multi' + 候选写 match_note;
全不中 → match_method='none' + 原因写 match_note (2025年单/姓名不存在等)。
幂等: wsf_order_no 唯一, 重导按新值更新非空字段, 配对结果保留人工改过的不覆盖。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.finance import WanshifuOrder
from app.models.order import Order

METHOD_CN = {
    "remark": "备注淘宝单号", "remark_unmatched": "备注单号(订单未导入)",
    "phone_full": "手机号全等", "phone_base": "手机号主段", "track": "物流单号",
    "name_city": "姓名+城市", "name_unique": "姓名唯一",
    "multi": "多候选待人工", "none": "未匹配", "manual": "人工指定",
}

# 万师傅「常用备注」列常直接填淘宝订单号 (用户的"合并单号匹配") — 提取这串数字
_TB_NO_RE = re.compile(r"\d{12,}")
# 「校对后匹配」人工列提取淘宝订单号 (16~19位; ≥15 位避免误抓万师傅单号/电话/样品等文字)
_VERIFIED_NO_RE = re.compile(r"\d{15,}")

# 表头第 2 行字段名 → 模型字段 (按列名取, 不按列号 — 万师傅加列不怕)
_HEADER_MAP = {
    "订单编号": "wsf_order_no",
    "服务类目/类型": "service_type",
    "订单状态": "status",
    "商品类别": "product_category",
    "商品型号": "product_model",
    "客户姓名": "customer_name",
    "客户手机号": "customer_phone",
    "客户旺旺号": "remark_ww",       # 暂存, 有值拼进 remark
    "常用备注": "remark_taobao_no",  # 常直接填淘宝订单号 = 用户的"合并单号匹配"(最高优先配对)
    "订单总净额": "net_amount",
    "订单服务费": "service_fee",
    "下单时间": "created_time",
    "服务完工时间": "finished_time",
    "物流公司": "tracking_company",
    "物流单号": "tracking_no",
    "来源店铺": "source_shop",
}
# 地址 4 列: r2 是「客户地址」+3个空列, r3 是 省/市/区/详细地址 → 按 r3 接管
_ADDR_MAP = {"省": "province", "市": "city", "区": "district", "详细地址": "address"}

# 用户人工加的「校对后匹配」列 = 人工核对后的真实淘宝订单号 = 最高权威匹配源
# (自动配对按收货姓名常对不上 — 订单库多存旺旺昵称; 万师傅旺旺号/常用备注又多为空,
#  故人工校对列是唯一可靠真值)。表头可能在分组行(r1)/字段行(r2)任一位置, 故三行都扫。
_VERIFIED_HEADERS = {"校对后匹配", "人工匹配", "校对后订单号", "人工校对", "人工校对订单号"}


@dataclass
class WsfImportReport:
    parsed: int = 0
    inserted: int = 0
    updated: int = 0
    verified_matched: int = 0   # 按「校对后匹配」人工列落库的单数 (方案1)
    skipped_invalid: int = 0
    errors: list[str] = field(default_factory=list)


def _dec(v: Any) -> Optional[Decimal]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return Decimal(str(v).replace(",", "").replace("¥", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _dt(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:19], fmt)
        except ValueError:
            continue
    return None


def _s(v: Any) -> Optional[str]:
    s = str(v).strip() if v is not None else ""
    return s or None


def parse_workbook(wb) -> tuple[list[dict], list[str]]:
    """openpyxl workbook → 行字典列表。按表头文字定位列, 容忍加列/换序。"""
    ws = wb.worksheets[0]
    rows1 = [c.value for c in ws[1]]   # 分组行 (用户加的「校对后匹配」表头常落在此)
    rows2 = [c.value for c in ws[2]]
    rows3 = [c.value for c in ws[3]] if ws.max_row >= 3 else []
    col_of: dict[str, int] = {}
    for i, h in enumerate(rows2):
        key = _HEADER_MAP.get(str(h).strip()) if h else None
        if key:
            col_of[key] = i
    for i, h in enumerate(rows3):
        key = _ADDR_MAP.get(str(h).strip()) if h else None
        if key:
            col_of[key] = i
    # 「校对后匹配」人工列 — 表头可能在 r1/r2/r3 任一行 → 三行都扫, 命中即记为权威人工匹配源
    for hdr in (rows1, rows2, rows3):
        for i, h in enumerate(hdr):
            if h and str(h).strip() in _VERIFIED_HEADERS:
                col_of["verified_order_no"] = i
    missing = [k for k in ("wsf_order_no", "customer_name", "created_time") if k not in col_of]
    out = []
    if missing:
        return out, missing
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r:
            continue
        idx = col_of["wsf_order_no"]
        no = _s(r[idx]) if idx < len(r) else None
        if not no:
            continue
        rec = {k: (r[i] if i < len(r) else None) for k, i in col_of.items()}
        rec["wsf_order_no"] = no
        out.append(rec)
    return out, missing


def import_workbook(db: Session, wb, *, import_job_id: Optional[int] = None) -> WsfImportReport:
    rep = WsfImportReport()
    recs, missing = parse_workbook(wb)
    if missing:
        rep.errors.append(f"表头缺关键列 {missing}, 请确认是万师傅「订单导出」文件")
        return rep
    rep.parsed = len(recs)
    existing = {o.wsf_order_no: o for o in db.execute(select(WanshifuOrder)).scalars().all()}
    # 全部淘宝订单号 (用于判定「常用备注」里的单号是否真实存在 → 可直接配对)
    valid_order_nos = {no for (no,) in db.execute(select(Order.order_no)).all() if no}
    for rec in recs:
        # 常用备注里的淘宝订单号 (用户的"合并单号匹配") — 次优先配对依据
        tb_no = None
        m = _TB_NO_RE.search(str(rec.get("remark_taobao_no") or ""))
        if m:
            tb_no = m.group(0)
        # 「校对后匹配」人工核对号 = 最高权威 (覆盖一切启发式/备注; 含跨期/未导入单也照记)
        verified_no = None
        vm = _VERIFIED_NO_RE.search(str(rec.get("verified_order_no") or ""))
        if vm:
            verified_no = vm.group(0)
            rep.verified_matched += 1
        vals = dict(
            service_type=_s(rec.get("service_type")),
            status=_s(rec.get("status")),
            product_category=_s(rec.get("product_category")),
            product_model=_s(rec.get("product_model")),
            customer_name=_s(rec.get("customer_name")),
            customer_phone=_s(rec.get("customer_phone")),
            province=_s(rec.get("province")), city=_s(rec.get("city")),
            district=_s(rec.get("district")), address=_s(rec.get("address")),
            net_amount=_dec(rec.get("net_amount")),
            service_fee=_dec(rec.get("service_fee")),
            created_time=_dt(rec.get("created_time")),
            finished_time=_dt(rec.get("finished_time")),
            tracking_company=_s(rec.get("tracking_company")),
            tracking_no=_s(rec.get("tracking_no")),
            source_shop=_s(rec.get("source_shop")),
        )
        ww = _s(rec.get("remark_ww"))
        remark_bits = []
        if tb_no:
            remark_bits.append(f"淘宝单号:{tb_no}")
        if ww:
            remark_bits.append(f"旺旺号:{ww}")
        if remark_bits:
            vals["remark"] = "; ".join(remark_bits)
        # 备注单号命中真实订单 → 直接配对 (权威, 覆盖启发式; 不覆盖人工指定)
        tb_match = tb_no if (tb_no and tb_no in valid_order_nos) else None
        # 人工校对号在订单库 → 干净配对; 不在库 (早期单/未导入) 仍记匹配但批注提示
        v_note = (None if (verified_no and verified_no in valid_order_nos)
                  else "人工校对; 订单库暂无此单(早期单/待导入)")
        old = existing.get(rec["wsf_order_no"])
        if old is None:
            obj = WanshifuOrder(wsf_order_no=rec["wsf_order_no"],
                                import_job_id=import_job_id, **vals)
            if verified_no:
                obj.matched_order_no = verified_no
                obj.match_method = "manual"
                obj.match_note = v_note
            elif tb_match:
                obj.matched_order_no = tb_match
                obj.match_method = "remark"
                obj.match_note = None
            elif tb_no:
                obj.match_note = f"备注淘宝单号 {tb_no} 未在订单库 (主订单未导入?)"
            db.add(obj)
            # 万师傅导出同一单可能出现多行 (商品2/重复行) — 记入 existing,
            # 同文件后续行走更新路径, 否则批内重复撞唯一约束直接 500
            existing[rec["wsf_order_no"]] = obj
            rep.inserted += 1
        else:
            # 重导更新非空字段 (状态/完工时间会推进)
            changed = False
            for k, v in vals.items():
                if v is not None and getattr(old, k) != v:
                    setattr(old, k, v)
                    changed = True
            # 人工校对号最高权威: 以最新校对为准, 覆盖旧的一切匹配 (含旧 manual)
            if verified_no and old.matched_order_no != verified_no:
                old.matched_order_no = verified_no
                old.match_method = "manual"
                old.match_note = v_note
                changed = True
            # 否则备注单号权威配对: 覆盖非人工的旧配对结果 (以用户提供的备注为准)
            elif (not verified_no) and tb_match and old.match_method != "manual" and old.matched_order_no != tb_match:
                old.matched_order_no = tb_match
                old.match_method = "remark"
                old.match_note = None
                changed = True
            if changed:
                rep.updated += 1
    db.flush()
    return rep


# ---------------- 配对 ----------------

def _base_phone(p: Optional[str]) -> str:
    if not p:
        return ""
    m = re.match(r"(1\d{10})", re.sub(r"\D", "", p.split("-")[0]))
    return m.group(1) if m else ""


def match_orders(db: Session, *, only_unmatched: bool = True) -> dict:
    """给档案里的万师傅单配淘宝订单。返回 {matched, multi, none} 计数。"""
    orders = db.query(Order.order_no, Order.customer_name, Order.customer_phone,
                      Order.customer_address, Order.tracking_no).all()
    by_phone_full: dict[str, set] = {}
    by_phone_base: dict[str, set] = {}
    by_track: dict[str, set] = {}
    by_name: dict[str, list] = {}
    for o in orders:
        if o.customer_phone:
            by_phone_full.setdefault(o.customer_phone.strip(), set()).add(o.order_no)
            b = _base_phone(o.customer_phone)
            if b:
                by_phone_base.setdefault(b, set()).add(o.order_no)
        if o.tracking_no:
            by_track.setdefault(o.tracking_no.strip(), set()).add(o.order_no)
        if o.customer_name:
            by_name.setdefault(o.customer_name.strip(), []).append(o)

    stmt = select(WanshifuOrder)
    if only_unmatched:
        stmt = stmt.where(WanshifuOrder.matched_order_no.is_(None),
                          # 人工指定过的不重算
                          (WanshifuOrder.match_method.is_(None))
                          | (WanshifuOrder.match_method != "manual"))
    counts = {"matched": 0, "multi": 0, "none": 0}
    for w in db.execute(stmt).scalars().all():
        method, nos = None, set()
        if w.customer_phone and w.customer_phone in by_phone_full:
            method, nos = "phone_full", by_phone_full[w.customer_phone]
        if not method:
            b = _base_phone(w.customer_phone)
            if b and b in by_phone_base:
                method, nos = "phone_base", by_phone_base[b]
        if not method and w.tracking_no and w.tracking_no in by_track:
            method, nos = "track", by_track[w.tracking_no]
        if not method and w.customer_name:
            nc = by_name.get(w.customer_name, [])
            ac = {o.order_no for o in nc if o.customer_address
                  and ((w.district and w.district in o.customer_address)
                       or (w.city and w.city in o.customer_address))}
            if ac:
                method, nos = "name_city", ac
            elif len(nc) == 1:
                method, nos = "name_unique", {nc[0].order_no}
        if method and len(nos) == 1:
            w.matched_order_no = next(iter(nos))
            w.match_method = method
            w.match_note = None
            counts["matched"] += 1
        elif method:
            w.match_method = "multi"
            w.match_note = f"{METHOD_CN[method]}命中多个候选: {'/'.join(sorted(nos)[:5])}"
            counts["multi"] += 1
        else:
            w.match_method = "none"
            if w.created_time and w.created_time.year <= 2025:
                w.match_note = "2025年订单, 系统从2026年起算, 主订单未导入"
            elif w.customer_name and w.customer_name not in by_name:
                w.match_note = "订单库无此收货姓名 (订单多存旺旺昵称/未填客户信息), 虚拟电话对不上"
            else:
                w.match_note = "姓名有同名但城市对不上"
            counts["none"] += 1
    db.flush()
    return counts


def build_annotated_workbook(db: Session):
    """档案 → xlsx (含匹配批注列), 给用户人工核对。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "万师傅订单匹配"
    headers = ["万师傅单号", "状态", "商品类别", "客户姓名", "客户手机号",
               "省", "市", "区", "详细地址", "净额", "服务费", "下单时间",
               "匹配订单号", "匹配方式", "批注"]
    ws.append(headers)
    for w in db.execute(select(WanshifuOrder).order_by(
            WanshifuOrder.created_time.desc().nullslast())).scalars().all():
        ws.append([
            w.wsf_order_no, w.status, w.product_category, w.customer_name,
            w.customer_phone, w.province, w.city, w.district, w.address,
            float(w.net_amount) if w.net_amount is not None else None,
            float(w.service_fee) if w.service_fee is not None else None,
            w.created_time.strftime("%Y-%m-%d %H:%M") if w.created_time else None,
            w.matched_order_no,
            METHOD_CN.get(w.match_method or "", w.match_method),
            w.match_note,
        ])
    return wb

"""智能 Excel 导入 (Phase 14, 用户驱动).

业务: 员工随便丢一个 Excel — 可能格式不一致, 可能列名不规范, 可能多个 sheet.
系统应该:
    1. 自动嗅探每个 sheet 的真正表头行 (跳过标题 / 备注行)
    2. AI 分析每个 sheet 是哪类实体 + 给列映射
    3. 给每个 sheet 一个质量评分:
       - good      → 一键导
       - needs_review → 让用户确认 / 调整列映射
       - messy     → 列出具体问题 (哪些行格式错), 建议先修 Excel
    4. 用户全选 → 一键全导

公开 API:
    smart_analyze(db, file_bytes) -> AnalysisResult (per-sheet plan)
    smart_commit(db, file_bytes, plan) -> 全部 sheet 导入报告
"""
from __future__ import annotations

import concurrent.futures
import json
import logging
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services import excel_importer, settings_service
from app.services.ai_provider import AiUnavailable, build_provider
from app.services.excel_schemas import ENTITY_SCHEMAS

_logger = logging.getLogger("panse.smart_import")

# 畔色总表中已知 sheet 名 → entity 类型的确定性映射。
# 前缀数字编号 + 关键字匹配，不依赖 AI，保证总表一次性导入时每个 sheet 都能被正确识别。
_KNOWN_SHEET_PATTERNS: list[tuple[str, str]] = [
    # (关键字子串, entity_type) — 顺序优先
    ("账户余额", "account_balance"),
    ("account_balance", "account_balance"),
    ("支付宝流水", "alipay_flow"),
    ("alipay", "alipay_flow"),
    ("订单总表修改", "order"),
    ("订单总表", "order"),
    ("订单", "order"),
    ("产品总表", "product"),
    ("product", "product"),
    ("BOM", "bom"),
    ("bom", "bom"),
    ("配件库存", "part_inventory"),
    ("成品库存", "product_inventory"),
]


def _sheet_name_entity(sheet_name: str) -> Optional[str]:
    """从 sheet 名确定性推导 entity_type；对不上返回 None。"""
    name_lower = sheet_name.lower()
    for keyword, entity in _KNOWN_SHEET_PATTERNS:
        if keyword.lower() in name_lower:
            return entity if entity in ENTITY_SCHEMAS else None
    return None


# ----------------------------- 数据结构 ---------------------------- #


@dataclass
class SheetAnalysis:
    sheet_name: str
    total_rows: int
    header_row: int                  # 1-indexed, AI 嗅探的真正表头行
    columns: list[str]
    sample_rows: list[list]

    suggested_entity: Optional[str] = None
    entity_label: Optional[str] = None
    confidence: float = 0.0
    mapping: dict[str, str] = field(default_factory=dict)
    skipped_columns: list[str] = field(default_factory=list)

    # 质量评估
    quality: str = "unknown"          # good / needs_review / messy
    quality_score: int = 0             # 0-100
    issues: list[dict] = field(default_factory=list)
    # [{row: 12, col: '日期', value: '202401', problem: '日期格式错', fix: '改成 2024-01-01'}]

    notes: list[str] = field(default_factory=list)


@dataclass
class AnalysisResult:
    sheets: list[SheetAnalysis]


# ----------------------------- 表头嗅探 ---------------------------- #


def _detect_header_row(rows: list[list]) -> int:
    """嗅探真正的表头行 (1-indexed).

    启发式:
        - 跳过明显的标题行 (大部分单元格是空, 第一格是 '1-产品总表' 等)
        - 跳过备注行 (含 '*' 开头的, 大部分是 None)
        - 找连续 3+ 个有意义中文/英文文字的行
    最多查前 5 行, 找不到默认返回 1.
    """
    for i, row in enumerate(rows[:5]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        non_empty = [c for c in cells if c and c not in ("-", "—")]
        # 表头特征: ≥3 个非空 cell, 没有 * 开头, 没有特别长的描述
        if len(non_empty) < 3:
            continue
        if any(c.startswith("*") for c in non_empty):
            continue
        # 单元格平均长度 < 20 (避免把"产品介绍文案"行当 header)
        avg_len = sum(len(c) for c in non_empty) / len(non_empty)
        if avg_len > 30:
            continue
        return i + 1   # 1-indexed
    return 1


# ----------------------------- AI 分析 prompt ---------------------- #


_ANALYZE_SYSTEM = """你是 Excel → ERP 智能分析助手。给你一个 sheet 的表头 + 前 8 行数据,
判断这是哪一类 ERP 实体, 输出严格 JSON:

{
  "entity_type": "supported_entities 里的 key 之一, 拿不准填 'unknown'",
  "confidence": 0.0-1.0 (你对实体判断的置信度),
  "mapping": { "目标字段名": "Excel 列名", ... },
  "quality": "good" | "needs_review" | "messy",
  "quality_score": 0-100 (数据规范度评分),
  "issues": [
    {"row_offset": 行号 (0-based, 数据行的), "column": "列名", "value": "原值",
     "problem": "问题描述", "fix": "建议改成什么"}
  ],
  "notes": ["其他提示"]
}

判定规则:
- entity_type 必须从 supported_entities 选一个; 完全对不上填 'unknown'
- quality 评分:
  * good (>=85): 列名规范, 数据格式统一, 必填字段都对得上 → 可直接导
  * needs_review (60-84): 大体能识别, 但某些列模糊, 或有少量数据问题 → 让用户确认 mapping
  * messy (<60): 表头乱 / 格式严重不统一 / 必填字段缺失 → 建议员工先修 Excel
- issues 列出最多 5 个具体问题, 帮员工知道改哪几行
  问题类型示例: '日期格式不一致' / '数字列里有文本' / '必填字段为空' / '编码不符合规则'
- 仅输出 JSON, 不要解释文字"""


# AI 未配置时的占位结果 (不碰 db, 可在并发线程里直接复制使用)
_AI_UNAVAILABLE = {
    "entity_type": "unknown", "confidence": 0, "mapping": {},
    "quality": "needs_review", "quality_score": 50,
    "issues": [], "notes": ["AI 未配置, 跳过智能分析"],
}


def _build_diagnose_provider(db: Session):
    """构建一次 diagnose AI provider, 供所有 sheet 并发复用. 未配置返回 None.

    db 只在这里 (主线程) 读一次配置; 之后并发分析时各线程只用 provider, 不碰 db,
    避免 SQLAlchemy Session 非线程安全的问题。
    """
    cfg = settings_service.get_ai_config(db, "diagnose")
    try:
        return build_provider(cfg)
    except AiUnavailable:
        return None


def _ai_analyze(provider, columns: list[str], sample_rows: list[list],
                sheet_name: str) -> dict:
    """让 AI 看 sheet 头 + 样本数据, 返回完整分析.

    provider 必须由 _build_diagnose_provider 预先构建好 (非 None);
    本函数不访问 db, 因此可在线程池里并发调用。
    """
    schema_doc = {
        et: {"label": s["label"], "desc": s["description"],
             "required_fields": [k for k, v in s["fields"].items()
                                  if v.get("required")],
             "all_fields": list(s["fields"].keys())}
        for et, s in ENTITY_SCHEMAS.items()
    }
    user_msg = json.dumps({
        "supported_entities": schema_doc,
        "sheet": {
            "name": sheet_name, "columns": columns,
            "sample_rows": sample_rows,
        },
    }, ensure_ascii=False)
    try:
        resp = provider.chat(system=_ANALYZE_SYSTEM, user=user_msg, max_tokens=4000)
    except AiUnavailable as e:
        return {"entity_type": "unknown", "confidence": 0, "mapping": {},
                "quality": "needs_review", "quality_score": 50,
                "issues": [], "notes": [f"AI 调用失败: {e}"]}
    try:
        cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", resp.text.strip(),
                          flags=re.MULTILINE)
        data = json.loads(cleaned)
    except (json.JSONDecodeError, ValueError) as e:
        return {"entity_type": "unknown", "confidence": 0, "mapping": {},
                "quality": "messy", "quality_score": 30,
                "issues": [], "notes": [f"AI 返回不是 JSON: {e}"]}
    # 清理 mapping: 只保留 schema 里有的字段 + columns 里有的列
    if data.get("entity_type") in ENTITY_SCHEMAS:
        schema = ENTITY_SCHEMAS[data["entity_type"]]
        valid = set(schema["fields"])
        data["mapping"] = {
            k: v for k, v in (data.get("mapping") or {}).items()
            if k in valid and v in columns
        }
    return data


# ----------------------------- 启发式 fallback ---------------------- #


def _heuristic_match(columns: list[str], sample_rows: list[list]) -> tuple[str, dict, float]:
    """没 AI 时用关键字 + alias 匹配, 给个 entity 猜测.

    两遍匹配, 每列只用一次:
      - 第 1 遍精确(列名 == 别名), 第 2 遍才子串。
    避免子串贪婪把同一列分给多个字段 (例: 「工厂订单号」既含 '订单号' 又含 '工厂',
    旧逻辑会让 platform_order_no / factory_name 都误抢它)。精确匹配额外加权,
    让列名精确对应的实体在平分时胜出。
    """
    best_entity = "unknown"
    best_mapping: dict[str, str] = {}
    best_score = 0.0
    for et, schema in ENTITY_SCHEMAS.items():
        mapping: dict[str, str] = {}
        used: set[str] = set()
        score = 0.0
        fields = schema["fields"]
        # 第 1 遍: 精确列名匹配
        for fn, fdef in fields.items():
            aliases_lower = [fn.lower()] + [a.lower() for a in fdef.get("aliases", [])]
            for col in columns:
                if col in used:
                    continue
                if col.lower().strip() in aliases_lower:
                    mapping[fn] = col
                    used.add(col)
                    score += (2 if fdef.get("required") else 1) + 0.5
                    break
        # 第 2 遍: 子串匹配 (只补还没映射上的字段, 且列没被占用)
        for fn, fdef in fields.items():
            if fn in mapping:
                continue
            aliases = [fn] + fdef.get("aliases", [])
            for col in columns:
                if col in used:
                    continue
                if any(a in col for a in aliases):
                    mapping[fn] = col
                    used.add(col)
                    score += 2 if fdef.get("required") else 1
                    break
        if score > best_score:
            best_score = score
            best_entity = et
            best_mapping = mapping
    confidence = min(best_score / 10, 1.0)
    return best_entity, best_mapping, confidence


# ----------------------------- 数据 quality 校验 -------------------- #


def _validate_data_quality(
    entity_type: str, mapping: dict, header_row: int,
    sheet_rows: list[list], columns: list[str],
) -> tuple[int, list[dict]]:
    """对照 mapping 看实际数据格式, 找具体问题. 返回 (扣分, issues)."""
    issues: list[dict] = []
    if entity_type not in ENTITY_SCHEMAS:
        return 30, []
    schema = ENTITY_SCHEMAS[entity_type]
    deduct = 0

    # 字段 → 列号
    col_idx = {col: i for i, col in enumerate(columns)}

    # 检查必填映射
    for fn, fdef in schema["fields"].items():
        if fdef.get("required") and fn not in mapping:
            issues.append({"row_offset": -1, "column": fn,
                           "value": None,
                           "problem": f"必填字段 '{fn}' 没找到对应 Excel 列",
                           "fix": f"加一列叫 '{fdef.get('desc', fn)}'"})
            deduct += 15

    # 抽样校验前 20 行数据格式
    data_rows = sheet_rows[header_row:header_row + 20]
    issue_count_by_col: dict[str, int] = {}
    for r_idx, row in enumerate(data_rows):
        for target_field, excel_col in mapping.items():
            if target_field not in schema["fields"]:
                continue
            fdef = schema["fields"][target_field]
            ftype = fdef.get("type", "str")
            col_i = col_idx.get(excel_col)
            if col_i is None or col_i >= len(row):
                continue
            v = row[col_i]
            if v is None or (isinstance(v, str) and v.strip() in ("", "-", "#N/A")):
                if fdef.get("required"):
                    issue_count_by_col[excel_col] = issue_count_by_col.get(excel_col, 0) + 1
                continue
            try:
                if ftype == "int":
                    int(float(str(v).replace(",", "")))
                elif ftype == "decimal":
                    Decimal(str(v).replace(",", "").replace("¥", "")
                              .replace("元", "").strip())
                elif ftype == "date":
                    if not isinstance(v, (date, datetime)):
                        s = str(v).strip()
                        # 简单格式校验
                        if not re.search(r"\d{4}", s):
                            raise ValueError("不像日期")
            except (ValueError, InvalidOperation):
                if len(issues) < 10:
                    issues.append({
                        "row_offset": r_idx, "column": excel_col,
                        "value": str(v)[:40],
                        "problem": f"应该是 {ftype} 类型, 实际是 '{v}'",
                        "fix": f"改成 {ftype} 格式 (如 数字 / YYYY-MM-DD)",
                    })
                deduct += 2

    # 必填字段空率
    for col, cnt in issue_count_by_col.items():
        if cnt >= 5:
            issues.append({
                "row_offset": -1, "column": col, "value": None,
                "problem": f"必填字段对应的 Excel 列 '{col}' 有 {cnt} 行是空的",
                "fix": "把这些行补齐 或 在导入时跳过",
            })
            deduct += 10

    return deduct, issues


# ----------------------------- 主流程 ---------------------------- #


def smart_analyze(db: Session, file_bytes: bytes) -> AnalysisResult:
    """智能分析: 整个 Excel 每个 sheet 一份 SheetAnalysis."""
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _logger.warning("打开 Excel 失败: %s", e)
        raise excel_importer.ImporterError(f"无法解析 Excel: {e}") from e
    _logger.info("Excel 打开成功, 共 %d 个 sheet: %s",
                 len(wb.worksheets), [w.title for w in wb.worksheets])

    # provider 只在主线程读一次配置, 之后并发分析各 sheet 只用 provider 不碰 db
    provider = _build_diagnose_provider(db)
    _logger.info("AI provider: %s", "已就绪" if provider else "未配置(将用启发式匹配)")

    # ---- 1) 顺序预处理: 读每个 sheet + 嗅探表头 (快, 无 AI) ----
    # prepped 每项: 要么 {"done": SheetAnalysis} (空/无表头, 不需 AI),
    #               要么 {"name","columns","sample_rows","header_row","all_rows"}
    prepped: list[dict] = []
    for ws in wb.worksheets:
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([_safe(c) for c in (row or [])])
        if not all_rows:
            prepped.append({"done": SheetAnalysis(
                sheet_name=ws.title, total_rows=0, header_row=1,
                columns=[], sample_rows=[],
                quality="messy", quality_score=0,
                notes=["空 sheet"],
            )})
            continue

        header_row = _detect_header_row(all_rows)
        header_cells = all_rows[header_row - 1] if header_row <= len(all_rows) else []
        columns = [str(c).strip() if c else f"col{i+1}"
                   for i, c in enumerate(header_cells)]
        # 去掉尾部连续空列
        while columns and (not columns[-1] or columns[-1].startswith("col")):
            columns.pop()

        sample_rows = []
        for r in all_rows[header_row:header_row + 8]:
            sample_rows.append(r[:len(columns)])

        if not columns or len(columns) < 2:
            prepped.append({"done": SheetAnalysis(
                sheet_name=ws.title, total_rows=len(all_rows) - header_row,
                header_row=header_row, columns=columns, sample_rows=sample_rows,
                quality="messy", quality_score=0,
                notes=["没找到表头. 这个 sheet 跳过."],
            )})
            continue

        prepped.append({
            "name": ws.title, "columns": columns, "sample_rows": sample_rows,
            "header_row": header_row, "all_rows": all_rows,
        })
    wb.close()

    # ---- 1b) 确定性 sheet 名匹配：已知命名规律直接定 entity，跳过 AI ----
    for p in prepped:
        if "done" in p:
            continue
        known = _sheet_name_entity(p["name"])
        if known:
            _, h_mapping, _ = _heuristic_match(p["columns"], p["sample_rows"])
            p["ai_result"] = {
                "entity_type": known,
                "confidence": 0.95,
                "mapping": h_mapping,
                "quality": "good",
                "quality_score": 85,
                "issues": [],
                "notes": [f"sheet 名确定性匹配 → {ENTITY_SCHEMAS[known]['label']}"],
            }

    # ---- 2) 并发跑 AI: 每个 sheet 一次调用, 26 个并行 → 总耗时≈单次而非累加 ----
    ai_targets = [p for p in prepped if "done" not in p and "ai_result" not in p]

    def _run_ai(p: dict) -> dict:
        if provider is None:
            return dict(_AI_UNAVAILABLE)
        import time as _t
        t = _t.monotonic()
        try:
            r = _ai_analyze(provider, p["columns"], p["sample_rows"], p["name"])
            _logger.info("  [%s] AI判定=%s 置信=%.2f 耗时=%.0fms",
                         p["name"], r.get("entity_type"),
                         float(r.get("confidence", 0) or 0),
                         (_t.monotonic() - t) * 1000)
            return r
        except Exception as e:
            _logger.warning("  [%s] AI分析异常: %s: %s", p["name"], type(e).__name__, e)
            return {"entity_type": "unknown", "confidence": 0, "mapping": {},
                    "quality": "needs_review", "quality_score": 50,
                    "issues": [], "notes": [f"AI 分析异常: {e}"]}

    if ai_targets:
        _logger.info("并发分析 %d 个 sheet (启发式跳过 %d 个)...",
                     len(ai_targets), len(prepped) - len(ai_targets))
        with concurrent.futures.ThreadPoolExecutor(
                max_workers=min(8, len(ai_targets))) as ex:
            for p, r in zip(ai_targets, ex.map(_run_ai, ai_targets)):
                p["ai_result"] = r

    # ---- 3) 顺序后处理: 启发式兜底 + 数据质量校验 (无 AI, 快) ----
    out: list[SheetAnalysis] = []
    for p in prepped:
        if "done" in p:
            out.append(p["done"])
            continue
        ws_title = p["name"]
        columns = p["columns"]
        sample_rows = p["sample_rows"]
        header_row = p["header_row"]
        all_rows = p["all_rows"]
        ai_result = p["ai_result"]
        entity = ai_result.get("entity_type", "unknown")

        # AI 没给 mapping → fallback 启发式
        if entity == "unknown" or not ai_result.get("mapping"):
            h_entity, h_mapping, h_conf = _heuristic_match(columns, sample_rows)
            if h_conf > 0.3:
                entity = h_entity
                ai_result["mapping"] = h_mapping   # 强覆盖, 不用 setdefault
                ai_result["confidence"] = max(ai_result.get("confidence", 0), h_conf)
                # AI 没给出有效评分时 (JSON 截断 / 未识别) 会留个惩罚分 (30),
                # 启发式既然认出来了, 就把基准分拉回中性, 让真实质量由
                # _validate_data_quality 的数据校验决定, 别卡在惩罚分上
                ai_result["quality_score"] = 75
                ai_result.setdefault("notes", []).append(
                    f"AI 不确定, 启发式匹配建议 {ENTITY_SCHEMAS[entity]['label']}",
                )

        mapping = ai_result.get("mapping", {})

        # 数据格式校验
        deduct, extra_issues = _validate_data_quality(
            entity, mapping, header_row, all_rows, columns,
        )
        ai_score = ai_result.get("quality_score", 75)
        if not isinstance(ai_score, (int, float)):
            ai_score = 75
        final_score = max(0, int(ai_score) - deduct)

        # 综合质量等级 (没 AI 时基础分 50, 阈值调宽点)
        if final_score >= 80:
            quality = "good"
        elif final_score >= 40:
            quality = "needs_review"
        else:
            quality = "messy"

        all_issues = list(ai_result.get("issues") or []) + extra_issues

        out.append(SheetAnalysis(
            sheet_name=ws_title,
            total_rows=max(len(all_rows) - header_row, 0),
            header_row=header_row,
            columns=columns,
            sample_rows=sample_rows,
            suggested_entity=entity if entity in ENTITY_SCHEMAS else None,
            entity_label=ENTITY_SCHEMAS[entity]["label"] if entity in ENTITY_SCHEMAS else None,
            confidence=float(ai_result.get("confidence", 0) or 0),
            mapping=mapping,
            quality=quality,
            quality_score=final_score,
            issues=all_issues[:10],
            notes=list(ai_result.get("notes") or []),
        ))
    return AnalysisResult(sheets=out)


def smart_commit(
    db: Session, *, file_bytes: bytes,
    plan: list[dict],
) -> list[dict]:
    """按用户确认的 plan 全部导入. plan 元素:
        {sheet_name, entity_type, mapping, header_row, dry_run?,
         on_conflict?, sheet_account?}
    on_conflict 默认 'ask' (重导命中已有记录且值不同时, 记到 conflicts 让用户裁决).
    返回每个 sheet 的报告 (含 conflicts).
    """
    dry = any(item.get("dry_run") for item in plan)
    _logger.info("[smart-commit] 开始%s导入 %d 个 sheet",
                 "试运行" if dry else "", len(plan))
    reports = []
    for item in plan:
        sheet_name = item["sheet_name"]
        entity = item["entity_type"]
        mapping = item.get("mapping") or {}
        sheet_account = item.get("sheet_account")
        # 支付宝流水: 表里通常没有账户列, 账户名藏在 sheet 名里 (如 "9a-支付宝流水-企业号")
        # 自动推导, 用户没手动指定也能导
        if entity == "alipay_flow" and not sheet_account and "account" not in mapping:
            sheet_account = _derive_alipay_account(sheet_name)
        # alipay_flow 若没映射 account 列但给了 sheet_account, 也允许导
        if entity == "unknown" or (not mapping and not sheet_account):
            _logger.info("  [%s] 跳过 (未确认 entity/mapping)", sheet_name)
            reports.append({"sheet_name": sheet_name, "skipped": True,
                            "reason": "未确认 entity / mapping"})
            continue
        # 智能 commit 借用 excel_importer.commit_sheet, 但要先把 header_row 适配
        # commit_sheet 假设 header 在第 1 行, 我们要把多余的前置行用 io 流改造
        try:
            adjusted_bytes = _strip_header_offset(file_bytes, sheet_name,
                                                   item.get("header_row", 1))
        except Exception as e:
            reports.append({"sheet_name": sheet_name, "error": f"预处理 Excel 失败: {e}"})
            continue
        try:
            report = excel_importer.commit_sheet(
                db, file_bytes=adjusted_bytes,
                sheet_name=sheet_name, entity_type=entity, mapping=mapping,
                dry_run=item.get("dry_run", False),
                on_conflict=item.get("on_conflict", "ask"),
                sheet_account=sheet_account,
            )
            # 每个 sheet 成功后立即提交, 避免单个 sheet 异常回滚全部数据
            if not item.get("dry_run", False):
                db.commit()
            _logger.info(
                "  [%s] %s=%s 入库父=%d 子=%d 跳过=%d 错误=%d 冲突=%d",
                sheet_name, entity,
                "试运行" if item.get("dry_run") else "已提交",
                report.inserted_parents, report.inserted_children,
                report.skipped_rows, len(report.errors), len(report.conflicts))
            reports.append({
                "sheet_name": sheet_name, "entity_type": entity,
                "total_rows": report.total_rows,
                "inserted_parents": report.inserted_parents,
                "inserted_children": report.inserted_children,
                "skipped_rows": report.skipped_rows,
                "errors": report.errors[:10],
                "warnings": report.warnings[:10],
                "conflicts": report.conflicts[:50],
            })
        except Exception as e:
            # 捕获所有异常 (含 SQLAlchemy 错误), 回滚本 sheet 后继续
            try:
                db.rollback()
            except Exception:
                pass
            _logger.exception("  [%s] 导入失败: %s: %s", sheet_name, type(e).__name__, e)
            reports.append({"sheet_name": sheet_name,
                            "error": f"{type(e).__name__}: {e}"})
    _logger.info("[smart-commit] 完成: %d 个 sheet 处理完毕", len(reports))
    return reports


# ----------------------------- 辅助 ----------------------------- #


def _safe(v: Any) -> Any:
    """避免 datetime 序列化时炸."""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


_ALIPAY_ACCOUNTS = ["企业号", "个体户私账", "爱群号", "佳宝号", "主力号"]


def _derive_alipay_account(sheet_name: str) -> Optional[str]:
    """从 sheet 名推导支付宝账户. 如 "9a-支付宝流水-企业号" → "企业号"."""
    for acc in _ALIPAY_ACCOUNTS:
        if acc in sheet_name:
            return acc
    # 兜底: 取最后一个 "-" 之后的部分
    if "-" in sheet_name:
        tail = sheet_name.rsplit("-", 1)[-1].strip()
        if tail and "支付宝" not in tail:
            return tail
    return None


def _strip_header_offset(file_bytes: bytes, sheet_name: str, header_row: int) -> bytes:
    """如果 header_row > 1, 重写 Excel 删掉前置行, 让 header 变成第 1 行.

    简单实现: openpyxl 加载 → 删行 → 保存到 BytesIO.
    对于大文件可能慢, 但智能分析模式假设单次导入数据量可控.
    """
    if header_row <= 1:
        return file_bytes
    from openpyxl import load_workbook as _lw
    wb = _lw(BytesIO(file_bytes))   # not read_only — 要改
    if sheet_name not in wb.sheetnames:
        wb.close()
        return file_bytes
    ws = wb[sheet_name]
    for _ in range(header_row - 1):
        ws.delete_rows(1)
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def to_dict(result: AnalysisResult) -> dict:
    return {"sheets": [asdict(s) for s in result.sheets]}

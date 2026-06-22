"""微信/聚合 结算账单(billDetail)导入: 解析 xlsx → 按 支付流水号 upsert OrderSettlement。

billDetail 列: 入账时间 / 支付流水号 / 淘宝订单编号 / 入账类型 / 收入金额 / 支出金额 / 业务描述 / 备注
(这些导出文件的 dimension 标记不规范, 故全量加载 + 强制按列/行扫描, 不依赖 max_row/max_col。)
"""
from __future__ import annotations

import hashlib
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.settlement import OrderSettlement


def _num(x) -> Decimal:
    if x is None or str(x).strip() == "":
        return Decimal("0")
    try:
        return Decimal(str(x).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _dt(x) -> Optional[datetime]:
    if not x:
        return None
    if isinstance(x, datetime):
        return x
    s = str(x).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def import_bill(db: Session, content: bytes, source: str = "wechat") -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0]

    # 找表头行 (含「支付流水号」)
    header_row = None
    col: dict[str, int] = {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=15, values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any("支付流水号" in c or "流水号" in c for c in cells):
            header_row = ri
            for ci, c in enumerate(cells):
                if c:
                    col[c] = ci
            break
    if header_row is None:
        return {"error": "未找到表头(需含『支付流水号』列)", "inserted": 0, "updated": 0}

    def idx(*names) -> Optional[int]:
        for n in names:
            for k, v in col.items():
                if n in k:
                    return v
        return None

    i_time = idx("入账时间", "交易时间")
    i_pay = idx("支付流水号", "流水号")
    i_order = idx("淘宝订单编号", "订单编号", "商家订单号")
    i_type = idx("入账类型", "交易分类")
    i_in = idx("收入金额", "收入")
    i_out = idx("支出金额", "支出")
    i_desc = idx("业务描述", "商品说明")
    i_rmk = idx("备注")

    def g(row, i):
        return row[i] if (i is not None and i < len(row)) else None

    # 去重: 已入库的 pay_no 集合 + 本次 call 内已处理的 (同一支付流水号在多份重叠账单里只留一条)
    existing_pays = {p for (p,) in db.execute(select(OrderSettlement.pay_no)).all()}
    seen: dict[str, OrderSettlement] = {}
    inserted = updated = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=50000, min_col=1, max_col=15, values_only=True):
        pay = g(row, i_pay)
        if not pay:
            continue
        pay = str(pay).strip()
        rec = seen.get(pay)
        if rec is None:
            if pay in existing_pays:
                rec = db.execute(
                    select(OrderSettlement).where(OrderSettlement.pay_no == pay)
                ).scalar_one()
                updated += 1
            else:
                rec = OrderSettlement(pay_no=pay, source=source)
                db.add(rec)
                existing_pays.add(pay)
                inserted += 1
            seen[pay] = rec
        rec.source = source
        rec.order_no = str(g(row, i_order)).strip() if g(row, i_order) else None
        rec.settle_time = _dt(g(row, i_time))
        rec.entry_type = str(g(row, i_type)).strip() if g(row, i_type) else None
        rec.income = _num(g(row, i_in))
        rec.expense = _num(g(row, i_out))
        rec.description = str(g(row, i_desc)).strip() if g(row, i_desc) else None
        rec.remark = str(g(row, i_rmk)).strip() if g(row, i_rmk) else None
    db.flush()
    return {"inserted": inserted, "updated": updated, "source": source}


def summary(db: Session) -> dict:
    rows = db.execute(select(OrderSettlement)).scalars().all()
    income = sum((r.income or Decimal("0")) for r in rows)
    expense = sum((r.expense or Decimal("0")) for r in rows)
    orders = len({r.order_no for r in rows if r.order_no})
    by_source: dict[str, int] = {}
    for r in rows:
        by_source[r.source] = by_source.get(r.source, 0) + 1
    return {
        "count": len(rows), "orders": orders,
        "income": float(income), "expense": float(expense), "net": float(income - expense),
        "by_source": by_source,
    }


# ── 支付宝企业号 订单级分账 → order_settlements 自动路由 (用户拍板 2026-06-23) ──────────
# 淘宝订单走支付宝企业号结算时, 一笔订单的「货款收款 / 软件服务费 / 消费券代付资金扣回」三条
# 交易分账共用一个交易流水号, related_order_no = "T200P" + 淘宝订单号。把它们落进 order_settlements,
# 让这些订单也进入逐笔结算对账 (此前只有微信/聚合账单才进, 支付宝订单一直是「待补流水」)。
# 防双算: order_reconciliation._alipay_net_by_flow_no 已排除 T200P 流水, 实际到账只由这里提供。
_T200P_PREFIX = "T200P"


def _settlement_description(remark: str | None, txn_type: str | None) -> str:
    rmk = remark or ""
    if "消费券" in rmk and "扣回" in rmk:
        return "营销支出-消费券代付扣回"
    if "消费券" in rmk:
        return "营销支出-消费券"
    if "软件服务费" in rmk:
        return "软件服务费-基础"
    if (txn_type or "") in ("交易付款", "收入"):
        return "货款收款"
    return (txn_type or "结算")


def _is_order_settlement_flow(remark: str | None, txn_type: str | None) -> bool:
    """T200P 流水里只取真·订单结算行: 买家货款收款(交易付款/收入) 或 平台软件费/消费券扣费。

    排除 店铺过户(账户迁移, 不是该单到账) 与其它泛分账, 否则会把内部转账当成订单到账 → 到账虚高。
    """
    rmk = remark or ""
    if "店铺过户" in rmk:
        return False
    if (txn_type or "") in ("交易付款", "收入"):
        return True
    return ("软件服务费" in rmk) or ("消费券" in rmk)


def route_alipay_flows(db: Session) -> dict:
    """把支付宝企业号里订单级结算行(related_order_no 以 T200P 开头)路由进 order_settlements。

    只取真结算行(货款收款 + / 软件服务费 − / 消费券代付扣回 −; 排除店铺过户等内部转账)。
    幂等(pay_no = 'ali:' + 业务键md5, 定长且与账单导入的纯数字 pay_no 不冲突)。可重复跑, 只补缺/改现。
    """
    from app.models.finance import AlipayFlow

    flows = db.execute(
        select(AlipayFlow).where(AlipayFlow.related_order_no.like(f"{_T200P_PREFIX}%"))
    ).scalars().all()
    existing = {p for (p,) in db.execute(select(OrderSettlement.pay_no)).all()}
    seen: set[str] = set()
    inserted = updated = 0
    for f in flows:
        if not _is_order_settlement_flow(f.remark, f.transaction_type):
            continue
        order_no = (f.related_order_no or "")[len(_T200P_PREFIX):].strip()
        if not order_no:
            continue
        amt = f.amount or Decimal("0")
        key = f"{f.account}:{f.transaction_no}:{f.transaction_type}:{amt}:{f.balance}"
        pay_no = "ali:" + hashlib.md5(key.encode("utf-8")).hexdigest()  # noqa: S324 — 仅作幂等去重键, 非安全用途
        if pay_no in seen:
            continue
        seen.add(pay_no)
        if pay_no in existing:
            rec = db.execute(
                select(OrderSettlement).where(OrderSettlement.pay_no == pay_no)
            ).scalar_one()
            updated += 1
        else:
            rec = OrderSettlement(pay_no=pay_no, source="alipay")
            db.add(rec)
            existing.add(pay_no)
            inserted += 1
        rec.source = "alipay"
        rec.order_no = order_no
        rec.settle_time = f.transaction_time
        rec.entry_type = f.transaction_type
        rec.income = amt if amt > 0 else Decimal("0")
        rec.expense = (-amt) if amt < 0 else Decimal("0")
        rec.description = _settlement_description(f.remark, f.transaction_type)
        rec.remark = f.remark
    db.flush()
    return {"routed_flows": len(flows), "inserted": inserted, "updated": updated}


def coupon_pending_summary(db: Session) -> dict:
    """消费券应补未补 (低优提醒, 约 2 月到账, 用户拍板 2026-06-23 折叠不催)。

    平台把消费券先垫进货款(交易付款), 又「消费券代付资金扣回」扣走(支出); 之后按「超过封顶金额的
    平台出资合作费用」分批补回(收入)。应补未补 = 扣回 − 已退回 − 已补回。pending>0 = 平台还欠的消费券。
    口径只读支付宝流水(出资合作是无订单号的批量转账, 不在 order_settlements 里), 不进利润、纯现金时序。
    """
    from app.models.finance import AlipayFlow

    rows = db.execute(
        select(AlipayFlow.transaction_time, AlipayFlow.amount, AlipayFlow.remark)
        .where(AlipayFlow.remark.isnot(None))
    ).all()
    clawback = refunded = cofund = Decimal("0")
    by_month: dict[str, Decimal] = {}
    for t, amt, rmk in rows:
        amt = amt or Decimal("0")
        ym = t.strftime("%Y-%m") if t else "未知"
        delta = Decimal("0")
        if "消费券" in rmk and "扣回" in rmk:
            if amt < 0:
                clawback += -amt
                delta = -amt           # 当月新增应补
            else:
                refunded += amt
                delta = -amt           # 退回冲减
        elif "出资合作" in rmk and amt > 0:
            cofund += amt
            delta = -amt               # 平台补回冲减
        if delta != 0:
            by_month[ym] = by_month.get(ym, Decimal("0")) + delta
    pending = clawback - refunded - cofund
    months = [
        {"month": m, "net_pending": float(by_month[m])}
        for m in sorted(by_month.keys())
    ]
    return {
        "clawback": float(clawback), "refunded": float(refunded), "cofund": float(cofund),
        "pending": float(pending), "by_month": months,
    }

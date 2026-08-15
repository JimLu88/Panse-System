"""活动核对器 (2026-07-17 spec §七 三种导出格式 + §四.6 核对维度 a~e)。

纯函数层 (无 DB, 验收/回归直接注入):
- parse_activity_items_export  活动商品导出 (跳3行表头, A..T 列; 多营销ID记录只认"已发布设定";
                               真实导出是合并单元格续行格式 — 商品ID/状态只在首SKU行, 需前向填充)
- parse_activity_floor_evidence_export  同一导出的价格线证据口径；保留暂停/异常/草稿等
                               非在场状态，但绝不用于报名集合核对
- parse_discount_export        单品立减导出 (1行表头, 12列)
- parse_product_batch_export   商品批量导出 (发布模板 sheet, 跳3行, 需 reset_dimensions)
- verify_campaign_title        活动名称头部校验 (不一致 → 中止+报警, 防推错活动)
- compare_records              逐SKU判定[一分不差/贴线让X/超2元报警/偏差/J未刷新/占位/无映射],
                               价格映射注入 (DB 的 target_prices 或验收样本的 dump 均可)

服务层 (薄包装):
- reconcile                    组装 DB 价格映射 → compare_records → 落 CampaignReconReport;
                               >2元差异 notify_service.broadcast_text 报警
- auto_recon_scan              调度用: signup_pushed(2小时内) 计划 → WA 导出 → 自动核对

复刻 2026-07-17 人工核对口径 (135一分不差 + 126贴线 + 0报警 为回归锚,
见 tests/test_campaign_recon_acceptance_0717.py)。
"""
from __future__ import annotations

import io
import re
from typing import Optional

from sqlalchemy.orm import Session

ACTIVITY_PUBLISHED_STATUS = "已发布设定"   # 兼容旧引用; 判定用下面的元组
# ✅2026-07-17 20:17 live实证: 活动开场后平台状态从「已发布设定」变「活动中」——两个都是"在活动内"
ACTIVITY_IN_CAMPAIGN_STATUSES = ("已发布设定", "活动中")
ALARM_THRESHOLD_YUAN = 2.0                # 到手 vs 目标差异 > 2 元 → 报警 (spec §四.6a)
LINE_LET_MAX_YUAN = 1.0                   # 贴线让幅 0~1 元记录在案; >1 元报警 (spec §四.6e)
_EPS = 0.01                               # 一分钱容差


def _f(x) -> Optional[float]:
    if x is None or x == "":
        return None
    try:
        return float(str(x).replace(",", "").replace("￥", "").replace("元", ""))
    except (TypeError, ValueError):
        return None


def _load_ws(xlsx_bytes: bytes, sheet_hint: Optional[str] = None):
    """read_only 打开 + reset_dimensions (千牛导出 dimension 元数据常烂, 不 reset 会读空)。"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    if sheet_hint:
        for name in wb.sheetnames:
            if sheet_hint in name:
                ws = wb[name]
                break
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()
    return wb, ws


def _parse_activity_export(xlsx_bytes: bytes, *, active_only: bool) -> list[dict]:
    """活动商品导出公共解析器；``active_only`` 控制是否只保留当前在场状态。"""
    wb, ws = _load_ws(xlsx_bytes)
    records: list[dict] = []
    item_id = item_name = marketing_id = status = ""
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 10:
            continue
        if row[0] is not None and str(row[0]).strip():
            item_id = str(row[0]).strip()
            item_name = str(row[1] or "").strip()
            marketing_id = str(row[2] or "").strip()
            status = str(row[3] or "").strip()
        sku_id = str(row[4] or "").strip()
        if not item_id or not sku_id:
            continue
        if active_only and status not in ACTIVITY_IN_CAMPAIGN_STATUSES:
            continue
        records.append({
            "item_id": item_id,
            "item_name": item_name,
            "marketing_id": marketing_id,
            "status": status,
            "sku_id": sku_id,
            "sku_name": str(row[5] or "").strip(),
            "list_price": _f(row[6]),
            "min_list_price": _f(row[7]),
            "min_coupon_line": _f(row[8]),
            "coupon_after": _f(row[9]),
            "activity_price": _f(row[15]) if len(row) > 15 else None,
        })
    wb.close()
    return records


def parse_activity_items_export(xlsx_bytes: bytes, *,
                                include_paused: bool = False) -> list[dict]:
    """活动商品导出 (spec §七.1, 跳3行表头): A商品ID/B商品名称/C营销ID/D商品状态/E SKUID/F SKU名称/
    G一口价/H最低标价/I最低普惠券后价要求/J活动普惠券后价/K建议价/…/P活动价/…
    多营销ID记录并存 → 逐记录分组, **只认状态"已发布设定"**。
    ★真实导出(2026-07-17 实测)是合并单元格续行格式: 商品ID/名称/营销ID/状态只写在
    每商品首SKU行, 续行 A~D 为空 → 前向填充, 否则每品只读到第一个 SKU。"""
    if not include_paused:
        return _parse_activity_export(xlsx_bytes, active_only=True)
    return [
        row for row in _parse_activity_export(xlsx_bytes, active_only=False)
        if row.get("status") in (*ACTIVITY_IN_CAMPAIGN_STATUSES, "暂停")
    ]


def parse_activity_floor_evidence_export(xlsx_bytes: bytes) -> list[dict]:
    """读取当前平台导出内全部逐 SKU H/I 价格线，不把状态当作在场证明。

    超级立减“导出全部商品”会包含活动中、暂停、异常、草稿和撤销报名记录。
    这些行的 H/I 仍是平台当前资格线，可以用于 R17；报名集合核对继续使用
    :func:`parse_activity_items_export`，两种语义不得混用。
    """
    return _parse_activity_export(xlsx_bytes, active_only=False)


def parse_discount_export(xlsx_bytes: bytes) -> list[dict]:
    """单品立减导出 (spec §七.2, 1行表头): 活动ID/活动名称/优惠级别/起/止/活动状态/
    商品ID/商品名称/SKU ID/SKU名称/一口价/优惠值。"""
    wb, ws = _load_ws(xlsx_bytes)
    records: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 12 or row[0] is None:
            continue
        records.append({
            "activity_id": str(row[0]).strip(),
            "activity_name": str(row[1] or "").strip(),
            "level": str(row[2] or "").strip(),
            "start": str(row[3] or "").strip(),
            "end": str(row[4] or "").strip(),
            "status": str(row[5] or "").strip(),
            "item_id": str(row[6] or "").strip(),
            "sku_id": str(row[8] or "").strip(),
            "sku_name": str(row[9] or "").strip(),
            "list_price": _f(row[10]),
            "discount_value": _f(row[11]),                  # 优惠值 = 立减金额
        })
    wb.close()
    return records


def parse_product_batch_export(xlsx_bytes: bytes) -> list[dict]:
    """商品批量导出 (spec §七.3, 发布模板 sheet, 跳3行, 需 reset_dimensions):
    col0商品Id / col4一口价(商品) / col9销售属性 / col11 skuId / col12价格 / col15商家编码。"""
    wb, ws = _load_ws(xlsx_bytes, sheet_hint="发布")
    records: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 13 or row[0] is None:
            continue
        records.append({
            "item_id": str(row[0]).strip(),
            "item_list_price": _f(row[4]),
            "sale_attr": str(row[9] or "").strip() if len(row) > 9 else "",
            "sku_id": str(row[11] or "").strip() if len(row) > 11 else "",
            "sku_price": _f(row[12]) if len(row) > 12 else None,
            "merchant_code": str(row[15] or "").strip() if len(row) > 15 else "",
        })
    wb.close()
    return records


def verify_campaign_title(expected: Optional[str], actual: Optional[str]) -> bool:
    """活动名称头部校验 (spec §四.6): 进活动页/读导出后核对标题一致, 不一致立即中止+报警防推错活动。
    空白归一后精确比对; 任一为空 → False (没有标题就没法证明是同一个活动, 诚实拒绝)。"""
    if not expected or not actual:
        return False
    norm = lambda t: re.sub(r"\s+", "", str(t))  # noqa: E731
    return norm(expected) == norm(actual)


# ── 逐 SKU 判定 (spec §四.6a: 到手 vs 目标; 纯函数, 价格映射注入) ──────────────

def _entry_target(spec_entry: dict, target_tier: Optional[str]) -> Optional[float]:
    """目标到手: 优先映射里已算好的 target (DB 层 target_prices 给);
    验收/回归注入的原始价格映射没有 target → 按场次档位取 big_buyer/mid_buyer。"""
    if "target" in spec_entry:
        return spec_entry.get("target")
    key = "mid_buyer" if target_tier == "mid" else "big_buyer"
    return spec_entry.get(key)


def _judge_row(rec: dict, spec_entry: Optional[dict],
               target_tier: Optional[str] = None) -> dict:
    """单条活动导出记录 vs 目标 → {sku_id, sku_code, actual, target, diff, verdict, ...}。
    verdict ∈ 一分不差 / 贴线让X / 超2元报警 / 偏差 / J未刷新 / 占位 / 无映射。
    贴线判定: 让幅 0<lag≤1 且 (无线值 or 实际≥线) — 无线值时推定贴线, 因为推送侧 builder
    已按 min(目标,线) 让价, 导出端小于目标≤1元只有贴线一种解释 (2026-07-17 人工核对口径)。"""
    actual = rec.get("coupon_after")
    base = {"sku_id": rec.get("sku_id"), "item_id": rec.get("item_id"),
            "actual": actual, "activity_price": rec.get("activity_price")}
    if spec_entry is None:
        return {**base, "sku_code": None, "target": None, "diff": None, "verdict": "无映射"}
    base["sku_code"] = spec_entry.get("sku_code")
    # b维度: 活动价 P列 vs 报名价(=日常价) — 一并记录 (spec §四.6b)
    signup = spec_entry.get("signup_price", spec_entry.get("daily"))
    act_price = rec.get("activity_price")
    base["signup_price_ok"] = (None if act_price is None or signup is None
                               else abs(act_price - signup) <= _EPS)
    base["expected_activity_price"] = signup
    if spec_entry.get("is_placeholder") or spec_entry.get("placeholder"):
        verdict = "占位价一致" if base["signup_price_ok"] is True else "占位活动价不一致"
        return {**base, "target": None, "diff": None, "verdict": verdict}
    target = _entry_target(spec_entry, target_tier)
    if target is None:
        return {**base, "target": None, "diff": None, "verdict": "无映射"}
    if actual is None:                                # 有映射有目标但 J 列空 → 平台还没算出来
        return {**base, "target": target, "diff": None, "verdict": "J未刷新"}
    diff = round(actual - target, 2)
    base.update({"target": target, "diff": diff})
    if abs(diff) <= _EPS:
        return {**base, "verdict": "一分不差"}
    line = spec_entry.get("line")
    lag = round(target - actual, 2)
    if (0 < lag <= LINE_LET_MAX_YUAN + _EPS
            and (line is None or actual >= line - _EPS)):
        return {**base, "verdict": f"贴线让{lag:.2f}", "concession": lag}
    if abs(diff) > ALARM_THRESHOLD_YUAN:
        return {**base, "verdict": "超2元报警"}
    return {**base, "verdict": "偏差"}


def compare_records(records: list[dict], price_map: dict[str, dict],
                    target_tier: Optional[str] = None) -> list[dict]:
    """纯函数比对: 活动导出记录 × 注入的价格映射 → 逐SKU判定行。
    price_map[sku_id] 两种形态皆可:
      ① DB 层 target_prices 输出 (含算好的 target/line/signup_price/is_placeholder);
      ② 原始价格 dump (daily/big_buyer/mid_buyer/placeholder) + target_tier 指定档位。"""
    return [_judge_row(rec, price_map.get(rec.get("sku_id")), target_tier)
            for rec in records]


def _coverage(spec_map: dict, seen_sku_ids: set,
              expected_sku_ids: Optional[set] = None) -> dict:
    """d维度: 应报 SKU 集 vs 已报集。

    服务层传入报名 builder 的真实输出集合，因此无动销、下架和坏价不会误报缺失，
    占位 SKU 若 builder 确实报名则同样纳入完整性检查。
    """
    expected = (
        set(expected_sku_ids)
        if expected_sku_ids is not None
        else {
            sid for sid, e in spec_map.items()
            if not e.get("is_placeholder") and e.get("kind") != "nosales"
        }
    )
    missing = sorted(expected - seen_sku_ids)
    extra = sorted(seen_sku_ids - set(spec_map.keys()))
    return {"missing": missing, "extra": extra}


def _compare_discounts(db: Session, plan, records: list[dict]) -> list[dict]:
    """c维度: 单品立减导出「优惠值」 vs builder 应填值 (差>1分即出入)。"""
    from app.services import campaign_service
    rows, _stats = campaign_service.build_discount_rows(db, plan)
    if campaign_service.platform_scope_present(plan):
        allowed_items = (
            campaign_service.platform_qualified_items(plan)
            | campaign_service.platform_no_sales_items(plan)
        )
        rows = [
            row for row in rows
            if str(row.get("taobao_item_id") or "") in allowed_items
        ]
    expected = {r["taobao_sku_id"]: r["deduct"] for r in rows}
    expected_items = {str(r["taobao_item_id"]) for r in rows}
    scoped_records = (
        [rec for rec in records if str(rec.get("item_id") or "") in expected_items]
        if campaign_service.platform_scope_present(plan)
        else records
    )
    actual = {rec.get("sku_id"): rec for rec in scoped_records if rec.get("sku_id")}
    mismatches = []
    for sid in sorted(set(expected) | set(actual)):
        rec = actual.get(sid) or {}
        exp = expected.get(sid)
        got = rec.get("discount_value")
        if exp is None or got is None or abs(exp - got) > _EPS:
            mismatches.append({"sku_id": sid, "expected": exp, "actual": got,
                               "item_id": rec.get("item_id"),
                               "kind": ("missing" if sid not in actual
                                        else "extra" if sid not in expected else "price")})
    return mismatches


def _summarize(per_sku: list[dict], coverage: dict, discount_mismatch: list,
               title_ok: Optional[bool]) -> dict:
    verdict_count: dict[str, int] = {}
    for r in per_sku:
        key = "贴线" if str(r["verdict"]).startswith("贴线让") else r["verdict"]
        verdict_count[key] = verdict_count.get(key, 0) + 1
    signup_mismatch = [
        {"sku_id": r.get("sku_id"), "sku_code": r.get("sku_code"),
         "expected": r.get("expected_activity_price"), "actual": r.get("activity_price")}
        for r in per_sku if r.get("signup_price_ok") is False
    ]
    hard_verdicts = {"超2元报警", "偏差", "无映射"}
    verdict_alarms = sum(n for key, n in verdict_count.items() if key in hard_verdicts)
    hard_error_count = (
        verdict_alarms
        + len(signup_mismatch)
        + len(coverage.get("missing", []))
        + len(coverage.get("extra", []))
        + len(discount_mismatch)
        + (1 if title_ok is False else 0)
    )
    return {"total": len(per_sku), "verdicts": verdict_count,
            "alarm": hard_error_count,
            "hard_error_count": hard_error_count,
            "signup_price_mismatch": signup_mismatch,
            "pending_coupon_after": verdict_count.get("J未刷新", 0),
            "coverage_missing": coverage.get("missing", []),
            "coverage_extra": coverage.get("extra", []),
            "discount_mismatch": discount_mismatch,
            "title_ok": title_ok}


def _alarm_text(plan, alarms: list[dict], summary: dict) -> str:
    lines = [f"⚠️ 活动核对报警: {plan.name}（{plan.campaign_type}）",
             f"硬错误共 {summary.get('hard_error_count', len(alarms))} 项；系统已停止自动完工。"]
    for a in alarms[:10]:
        lines.append(f"- {a.get('sku_code') or a.get('sku_id')}: "
                     f"实际 {a.get('actual')} vs 目标 {a.get('target')} (差 {a.get('diff')})")
    if len(alarms) > 10:
        lines.append(f"…共 {len(alarms)} 条, 详见系统核对报告")
    if summary.get("title_ok") is False:
        lines.append("★活动名称与计划不一致 — 疑推错活动, 已中止判定, 请人工核实!")
    if summary.get("signup_price_mismatch"):
        lines.append(f"活动报名价不一致: {len(summary['signup_price_mismatch'])} 个 SKU")
    if summary.get("coverage_missing") or summary.get("coverage_extra"):
        lines.append(
            f"报名集合缺失 {len(summary.get('coverage_missing') or [])} / "
            f"多出 {len(summary.get('coverage_extra') or [])} 个 SKU")
    if summary.get("discount_mismatch"):
        lines.append(f"单品立减不一致: {len(summary['discount_mismatch'])} 个 SKU")
    return "\n".join(lines)


def reconcile(db: Session, plan, *, activity_bytes: Optional[bytes] = None,
              discount_bytes: Optional[bytes] = None,
              product_bytes: Optional[bytes] = None, source: str = "manual") -> dict:
    """三种导出与 ERP 比对 → 逐SKU判定 → 落 CampaignReconReport; >2元差异飞书报警。
    任一文件可缺 (手动上传兜底, spec §四.6); 全缺 → 显式报错。"""
    from app.models.campaign import CampaignReconReport
    from app.services import (
        campaign_notification_service as notify_service,
        campaign_price_floor_service,
        campaign_service,
    )
    if not any((activity_bytes, discount_bytes, product_bytes)):
        return {"ok": False, "error": "未提供任何导出文件 (活动商品/单品立减/商品批量 至少一份)"}

    spec_map = campaign_service.target_prices(db, plan)
    per_sku: list[dict] = []
    coverage: dict = {"missing": [], "extra": []}
    title_ok: Optional[bool] = None
    floor_evidence = None
    ignored_out_of_scope_records = 0
    ignored_out_of_scope_items: list[str] = []
    if activity_bytes:
        records = parse_activity_items_export(
            activity_bytes,
            include_paused=str(getattr(plan, "campaign_type", "")) == "super_reduce",
        )
        floor_records = parse_activity_floor_evidence_export(activity_bytes)
        floor_evidence = campaign_price_floor_service.record_activity_export(
            db,
            floor_records,
            source=f"campaign_recon:{source}:plan={getattr(plan, 'id', '')}",
        )
        signup_rows, _signup_stats = campaign_service.build_signup_rows(db, plan)
        if campaign_service.platform_scope_present(plan):
            qualified_items = campaign_service.platform_qualified_items(plan)
            signup_rows = [
                row for row in signup_rows
                if str(row.get("taobao_item_id") or "") in qualified_items
            ]
        expected_signup = {str(row["taobao_sku_id"]) for row in signup_rows}
        expected_items = {str(row["taobao_item_id"]) for row in signup_rows}
        scoped_records = records
        if campaign_service.platform_scope_present(plan):
            scoped_records = [
                rec for rec in records
                if str(rec.get("item_id") or "") in expected_items
            ]
            ignored = [
                rec for rec in records
                if str(rec.get("item_id") or "") not in expected_items
            ]
            ignored_out_of_scope_records = len(ignored)
            ignored_out_of_scope_items = sorted({
                str(rec.get("item_id") or "") for rec in ignored
                if str(rec.get("item_id") or "")
            })
        per_sku = compare_records(scoped_records, spec_map)
        coverage = _coverage(
            spec_map, {rec["sku_id"] for rec in scoped_records}, expected_signup)
    discount_mismatch: list[dict] = []
    if discount_bytes:
        drecords = parse_discount_export(discount_bytes)
        discount_mismatch = _compare_discounts(db, plan, drecords)
        if plan.qn_campaign_title and drecords:      # 活动名称校验 (防推错活动)
            title_ok = verify_campaign_title(plan.qn_campaign_title, drecords[0]["activity_name"])
    product_rows = parse_product_batch_export(product_bytes) if product_bytes else []

    alarms = [r for r in per_sku if r["verdict"] in (
        "超2元报警", "偏差", "无映射", "占位活动价不一致")]
    summary = _summarize(per_sku, coverage, discount_mismatch, title_ok)
    summary["ignored_out_of_scope_records"] = ignored_out_of_scope_records
    summary["ignored_out_of_scope_items"] = ignored_out_of_scope_items
    summary["product_rows_parsed"] = len(product_rows)
    summary["price_floor_evidence"] = floor_evidence
    report = CampaignReconReport(plan_id=plan.id, source=source, summary=summary,
                                 rows=per_sku, alarm_count=summary["hard_error_count"])
    db.add(report)
    pending = summary.get("pending_coupon_after", 0) > 0
    if summary["hard_error_count"]:
        plan.status = "alarmed"
    elif pending:
        plan.status = "signup_pushed"
    else:
        plan.status = "reconciled"
    db.commit()
    if summary["hard_error_count"]:
        notify_service.broadcast_text(db, _alarm_text(plan, alarms, summary),
                                      title="活动核对报警", level="error")
    return {"ok": True,
            "verified": not pending and not summary["hard_error_count"],
            "pending": pending, "report_id": report.id, "summary": summary,
            "rows": per_sku, "alarm_count": summary["hard_error_count"]}


# ── 调度: 报名后自动核对 (P4, spec §五 campaign_auto_recon) ────────────────────

AUTO_RECON_WINDOW_HOURS = 6


def _notify_recon_export_failure_once(db: Session, plan, text: str) -> dict:
    """同一计划同一导出故障只提醒一次，原因变化后重新提醒。"""
    import hashlib
    from app.services import campaign_notification_service as notify_service, settings_service

    signature = hashlib.sha256(text.encode("utf-8")).hexdigest()
    key = f"campaign_recon_export_failure_{plan.id}"
    if settings_service.get(db, key, env_fallback=False) == signature:
        return {"deduped": True}
    delivered = notify_service.broadcast_text(
        db, text, title="活动自动核对失败", level="warn")
    if any(v is True for v in delivered.values()):
        settings_service.set_value(
            db, key, signature, description="活动核对导出失败通知去重签名")
        db.commit()
    return delivered


def auto_recon_scan(db: Session) -> dict:
    """每30分钟: status='signup_pushed' 且创建 {AUTO_RECON_WINDOW_HOURS} 小时内的计划 →
    web_agent_service.campaign_export_items 拿「活动商品导出」→ reconcile(source='auto')。
    WA 导出失败 → 飞书报错 + 计划保持 signup_pushed (等手动上传兜底), 下轮窗口内会再试。"""
    from datetime import datetime, timedelta, timezone
    from sqlalchemy import select
    from app.models.campaign import CampaignPlan
    from app.services import web_agent_service

    cutoff = datetime.now(timezone.utc) - timedelta(hours=AUTO_RECON_WINDOW_HOURS)
    plans = db.execute(select(CampaignPlan).where(
        CampaignPlan.status == "signup_pushed")).scalars().all()
    scanned = reconciled = failed = 0
    details: list[dict] = []
    for plan in plans:
        # 窗口锚 updated_at: 计划创建很久后才推报名也能被自动核对(推报名会更新计划状态)
        anchor = plan.updated_at or plan.created_at
        if anchor is None:
            continue
        if anchor.tzinfo is None:                  # sqlite 存 naive UTC, postgres 带 tz
            anchor = anchor.replace(tzinfo=timezone.utc)
        if anchor < cutoff:
            continue                               # 超过窗口 → 留给手动上传, 不再自动追
        scanned += 1
        title = plan.qn_campaign_title or plan.name
        exp = web_agent_service.campaign_export_items(db, title)
        if not exp.get("ok"):
            failed += 1
            err = exp.get("error") or exp.get("message") or "未知原因"
            _notify_recon_export_failure_once(
                db, plan,
                f"⚠️ 「{plan.name}」报名后自动核对没跑成: WA 导出已报商品失败（{err}）。\n"
                f"计划保持「报名已推」状态; 请自己去千牛活动页导出「活动商品导出」表, "
                f"到系统核对面板手动上传继续核对。")
            details.append({"plan_id": plan.id, "ok": False, "error": err})
            continue
        res = reconcile(db, plan, activity_bytes=exp["xlsx_bytes"], source="auto")
        if res.get("verified"):
            reconciled += 1
        else:
            failed += 1
        details.append({"plan_id": plan.id, "ok": res.get("ok"),
                        "alarm_count": res.get("alarm_count")})
    return {"scanned": scanned, "reconciled": reconciled, "failed": failed,
            "details": details}


# ── 失败原因归类钩子 (2026-07-17, 可选 AI 兜底) ────────────────────────────────

def classify_failure_reason(db: Session, text: Optional[str]) -> Optional[str]:
    """千牛报名/核对失败文本 → 一句话归类 (如「低于最低标价线」)。

    钩子性质: 只有在设置页配了「活动系统 AI (DeepSeek/千问)」才生效;
    未配置 / 文本为空 / LLM 失败 → 返回 None, 调用方保持现有行为 (零行为变化)。
    """
    if not text or not str(text).strip():
        return None
    try:
        from app.services import campaign_ai_service
        return campaign_ai_service.classify_failure_reason(db, str(text))
    except Exception:   # noqa: BLE001 — 归类是锦上添花, 绝不拖垮核对主流程
        return None

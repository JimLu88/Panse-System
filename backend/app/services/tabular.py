"""统一表格读取: CSV / Excel(xlsx) → CSV 文本。

让所有"CSV 导入"自动也吃 Excel(用户要求: 所有 CSV 文件都必须也支持 Excel)。
- xlsx(zip, PK 头 / .xlsx/.xlsm 后缀): openpyxl 读首个有内容的 sheet → 写成 CSV 文本;
  日期/时间单元格格式化成 ISO 串(与各导入器的日期解析兼容);
- 其它: 当 CSV, 先 utf-8-sig 再 gbk 解码。
旧的 .xls(OLE 二进制)openpyxl 读不了 → 抛清晰错误提示另存为 .xlsx。
"""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import date, datetime
from typing import Optional

_XLSX_MAGIC = b"PK\x03\x04"          # zip (xlsx/xlsm)
_XLS_MAGIC = b"\xd0\xcf\x11\xe0"     # OLE2 (老 .xls)


def _maybe_unzip_to_csv(content: bytes, filename: Optional[str]) -> tuple[bytes, Optional[str]]:
    """普通 zip(内含 csv/txt, 如支付宝个人账单下载的 .zip) → 解出里面的 CSV 字节 + 内层文件名;
    xlsx(本身也是 zip, 但有 [Content_Types].xml / xl/) 或普通文件 → 原样返回。
    让所有导入口直接吃这种 zip, 用户不用先解压 (用户 2026-07-06: 支付宝流水下载就是 zip)。"""
    if content[:2] != b"PK":               # 非 zip
        return content, filename
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            names = [n for n in zf.namelist() if not n.endswith("/")]
            if any(n == "[Content_Types].xml" or n.startswith("xl/") for n in names):
                return content, filename    # 是 xlsx → 交给 xlsx 分支
            members = [n for n in names if n.lower().endswith((".csv", ".txt"))]
            if members:
                # 支付宝 signcustomer 日账单 ZIP 同时带「交易明细(汇总).csv」和
                # 「交易明细.csv」。不能按 ZIP 内顺序取第一份：汇总表没有交易明细列，
                # 会被通用 CSV 导入器报成“缺少交易流水号”。按内容识别真正的明细表。
                def _decode(raw: bytes) -> str:
                    try:
                        return raw.decode("utf-8-sig")
                    except UnicodeDecodeError:
                        return raw.decode("gbk", errors="replace")

                def _score(name: str, raw: bytes) -> tuple[int, int]:
                    text = _decode(raw)[:8000]
                    score = 0
                    if "账务流水号" in text and "账户余额" in text:
                        score += 100
                    if "业务流水号" in text and "商户订单号" in text:
                        score += 40
                    if "收入金额" in text and "支出金额" in text:
                        score += 20
                    if "交易号" in text and "收/支" in text:
                        score += 80
                    if "交易流水号" in text and "收支金额" in text:
                        score += 80
                    # 文件名只作同分兜底，核心判据始终是表头内容。
                    return score, 0 if "汇总" in name else 1

                candidates = [(n, zf.read(n)) for n in members]
                member, raw = max(candidates, key=lambda item: _score(item[0], item[1]))
                return raw, member.rsplit("/", 1)[-1]
    except zipfile.BadZipFile:
        pass
    return content, filename


def looks_like_xlsx(content: bytes, filename: Optional[str] = None) -> bool:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return True
    if name.endswith(".csv"):
        return False
    return content[:4] == _XLSX_MAGIC


def is_legacy_xls(content: bytes, filename: Optional[str] = None) -> bool:
    name = (filename or "").lower()
    return name.endswith(".xls") or content[:4] == _XLS_MAGIC


def _fmt(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        # 有时分秒 → 带时间; 纯 0 点 → 只日期
        if (cell.hour, cell.minute, cell.second) == (0, 0, 0):
            return cell.strftime("%Y-%m-%d")
        return cell.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(cell, date):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, float):
        # 整数值的 float(如被 Excel 数字化的 订单号/流水号/金额)→ 不要写成科学计数 "1.2e+27",
        # 否则会污染 流水号/订单号 去重键。(注: >15 位的长号 openpyxl 已丢精度, 这是 Excel 固有问题,
        # 长号列建议保持文本格式或用 CSV; 此处至少避免科学计数串。)
        return str(int(cell)) if cell.is_integer() else repr(cell)
    return str(cell)


def _xlsx_to_csv(content: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    # 取第一个有数据的 sheet
    ws = next((s for s in wb.worksheets if s.max_row and s.max_row > 0), wb.active)
    out = io.StringIO()
    w = csv.writer(out)
    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        w.writerow([_fmt(c) for c in row])
    return out.getvalue()


def to_csv_texts(content: bytes, filename: Optional[str] = None) -> list[tuple[str, str]]:
    """xlsx/csv → [(sheet名, CSV文本), ...] — 多 sheet 全读 (用户要求: 工厂对账单两个 sheet 都要用)。

    CSV 文件天然单 sheet, 返回一项。空 sheet 跳过。
    """
    content, filename = _maybe_unzip_to_csv(content, filename)   # zip(内含csv, 如支付宝下载) 先解出
    if is_legacy_xls(content, filename) and not looks_like_xlsx(content, filename):
        raise ValueError("不支持老的 .xls 格式，请在 Excel 里『另存为 .xlsx』后再上传。")
    if not looks_like_xlsx(content, filename):
        return [("(CSV)", to_csv_text(content, filename))]
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    out: list[tuple[str, str]] = []
    for ws in wb.worksheets:
        if not ws.max_row or ws.max_row == 0:
            continue
        buf = io.StringIO()
        w = csv.writer(buf)
        has_data = False
        for row in ws.iter_rows(values_only=True):
            if all(c is None for c in row):
                continue
            has_data = True
            w.writerow([_fmt(c) for c in row])
        if has_data:
            out.append((ws.title, buf.getvalue()))
    return out or [("(空表)", "")]


def to_csv_text(content: bytes, filename: Optional[str] = None) -> str:
    """xlsx/csv 字节 → CSV 文本。给所有按 CSV 文本工作的导入器复用。"""
    content, filename = _maybe_unzip_to_csv(content, filename)   # zip(内含csv, 如支付宝下载) 先解出
    if is_legacy_xls(content, filename) and not looks_like_xlsx(content, filename):
        raise ValueError("不支持老的 .xls 格式，请在 Excel 里『另存为 .xlsx』后再上传。")
    if looks_like_xlsx(content, filename):
        return _xlsx_to_csv(content)
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("gbk", errors="replace")


def read_header(content: bytes, filename: Optional[str] = None) -> list[str]:
    """读首行表头(规范化: 去空白)。用于按表头结构判表格类型。"""
    try:
        text = to_csv_text(content, filename)
    except Exception:
        return []
    for row in csv.reader(io.StringIO(text)):
        return [(c or "").replace(" ", "").replace("　", "").strip() for c in row]
    return []

"""Claim-bound Plan 8 continuation after V7 proved zero platform writes."""
from __future__ import annotations

import base64
import hashlib
import io
import secrets
import time
from datetime import datetime, timezone

from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.campaign import CampaignExecutionAttempt, CampaignPlan
from app.services import (
    campaign_execution_service,
    campaign_plan8_final_recovery_v6_service as v6,
    campaign_plan8_final_recovery_v7_service as v7,
    campaign_policy_service,
    web_agent_service,
)


WORKFLOW_KEY = v7.WORKFLOW_KEY
PLAN_ID = v7.PLAN_ID
EXPECTED_STATUS = v7.EXPECTED_STATUS
RECOVERY_VERSION = 8
OPERATION = "plan8_final_recovery_v8"
EXECUTION_SOURCE = "campaign_super88_plan8_final_recovery_v8"
EXPECTED_POLICY_SHA256 = v7.EXPECTED_POLICY_SHA256
EXPECTED_TARGET_SCOPE_SHA256 = v7.EXPECTED_TARGET_SCOPE_SHA256
IDENTITY = v7.IDENTITY
TARGET_ITEM_IDS = v7.TARGET_ITEM_IDS
ADD_PAIRS = v7.ADD_PAIRS
OLD_DISCOUNT_ACTIVITY_ID = v7.OLD_DISCOUNT_ACTIVITY_ID
READBACK_PLAN_STATUSES = v7.READBACK_PLAN_STATUSES
V7_ATTEMPT_ID = "5a72360877df3c3fad221ee2"
EXPECTED_RESUME_EVIDENCE = {
    "v7_attempt_id": V7_ATTEMPT_ID,
    "v7_operation": "plan8_final_recovery_v7",
    "v7_state": "unknown_no_retry",
    "v7_last_checkpoint": "discount_terminal",
    "v7_claim_sha256": (
        "733d0fe55454ea7274ab5f61516fc3730a973e7ab961fdd22247f661321c23b2"
    ),
    "v7_import_ok": 0,
    "v7_import_failed": 8,
    "v7_import_submitted": False,
    "fresh_readback_new_discount_rows": 0,
    "fresh_readback_old_discount_rows": 53,
    "fresh_readback_old_discount_sha256": (
        "9c85a468a5fed6db667b7d388fea8e0ecb7148a2c9e70c8cd11e8cb06b52c2e2"
    ),
}
EXECUTION_ORDER = [
    "patch_6_bound_drafts_to_78_skus",
    "publish_6_bound_drafts",
    "supplement_8_discounts_in_existing_activity",
    "official_readback",
]
EXPECTED_COMMIT_CHECKPOINTS = [
    "claimed", "draft_patch_terminal", "draft_patch_readback_exact",
    "publish_terminal", "campaign_readback_exact", "discount_terminal",
    "discount_readback_exact", "official_readback_exact",
]
EXECUTE_CONFIRMATION = "EXECUTE_ONCE_PLAN8_V8_RESUME_V7_ZERO_WRITE"
READBACK_CONFIRMATION = "READBACK_ONLY_PLAN8_V8_NO_PLATFORM_WRITE"
PRECLAIM_RESUME_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_MANIFEST_PHASE_FIX_V3"
)
PRECLAIM_ATTEMPT_ID = "edaf6b609dad46fbab90c7e8"
PRECLAIM_SCOPE_SHA256 = (
    "08170d64d354c1e50f7f87270de118874b6e9db2513aae46cec94e9f5db8eb3a"
)
PRECLAIM_REQUEST_ID = "plan8-final-v8-67b588e1b5278ff4"
PRECLAIM_WEB_AGENT_JOB_ID = "job2"
PRECLAIM_LAST_STEP = "plan8_final_v8_commit"
PRECLAIM_ERROR_CODE = "ValueError: plan8_v6_manifest_fields_invalid"
CLAIMED_PREUPLOAD_RESUME_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_BATCH_DIALOG_FIX_V4"
)
CLAIMED_PREUPLOAD_POST_READBACK_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_ZERO_WRITE_READBACK_V5"
)
CLAIMED_PREUPLOAD_LEASE_SCOPE_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_LEASE_SCOPE_FIX_V6"
)
CLAIMED_PREUPLOAD_BUSY_WAIT_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_PREWRITE_BUSY_V7"
)
CLAIMED_PREUPLOAD_LEASE_EXPIRY_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_CLAIM_LEASE_EXPIRY_V8"
)
CLAIMED_PREUPLOAD_DIALOG_FIX_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_BATCH_IMPORT_DIALOG_FIX_V9"
)
CLAIMED_PREUPLOAD_CAMPAIGN_GUARD_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_CAMPAIGN_GUARD_SETTLE_FIX_V10"
)
CLAIMED_PREUPLOAD_CLAIM_VERIFY_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V10_CLAIM_VERIFY_FIX_V11"
)
CLAIMED_PREUPLOAD_LAZY_IMPORT_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_LAZY_IMPORT_BINDING_FIX_V12"
)
CLAIMED_PREUPLOAD_ALLOWLIST_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V11_CLAIM_ALLOWLIST_FIX_V13"
)
CLAIMED_PREUPLOAD_SEMANTIC_MODAL_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_SEMANTIC_MODAL_BINDING_FIX_V14"
)
CLAIMED_PREUPLOAD_EDITOR_IDENTITY_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_EDITOR_IDENTITY_FORMAT_FIX_V15"
)
CLAIMED_PREUPLOAD_NESTED_MODAL_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_NESTED_MODAL_TEXT_FIX_V16"
)
CLAIMED_PREUPLOAD_MOBAN_TEXT_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_MOBAN_TEXT_VARIANT_FIX_V17"
)
CLAIMED_POSTUPLOAD_READBACK_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V17_UNPERSISTED_READBACK_V18"
)
CLAIMED_OFFICIAL_TEMPLATE_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_WITH_LIVE_OFFICIAL_TEMPLATE_V19"
)
CLAIMED_TEMPLATE_GENERATION_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_TEMPLATE_GENERATION_MODAL_V20"
)
CLAIMED_TEMPLATE_CLAIM_VERIFY_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V20_CLAIM_STEP_FIX_V21"
)
CLAIMED_TEMPLATE_CLOSE_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V21_TEMPLATE_CLOSE_FIX_V22"
)
CLAIMED_TEMPLATE_CONTRACT_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V22_CLAIM_CONTRACT_FIX_V23"
)
CLAIMED_EXPORT_RETRY_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V23_READONLY_EXPORT_FAILURE_V24"
)
CLAIMED_MANUAL_EXPORT_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_WITH_VERIFIED_MANUAL_EXPORT_V25"
)
CLAIMED_MANUAL_EXPORT_V26_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_WITH_VERIFIED_MANUAL_EXPORT_FIELDS_V26"
)
CLAIMED_MANUAL_EXPORT_V27_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_BOUND_DRAFT_EDITOR_HYDRATION_V27"
)
CLAIMED_MANUAL_EXPORT_V28_CONFIRMATION = (
    "RESUME_ONCE_PLAN8_V8_AFTER_V27_CLAIM_ALLOWLIST_FIX_V28"
)
MANUAL_EXPORT_FILENAME = (
    "「26年淘宝9月超级88超级88现货」活动商品导出20260904182846.xlsx"
)
MANUAL_EXPORT_SIZE = 14994
MANUAL_EXPORT_SHA256 = (
    "c7c22b57a95e7db5f3cc8d8a0319ee4b1920a13e73204f1004be3760d71d25da"
)
PREUPLOAD_BUSY_WAIT_SECONDS = 600.0
PREUPLOAD_BUSY_POLL_SECONDS = 5.0
CLAIMED_PREUPLOAD_SCOPE_SHA256 = (
    "04ea6c51d5bc50ca3c4361fd503ce75503772a2fc6a98cb254ec5842a511d6d3"
)
CLAIMED_PREUPLOAD_CLAIM_SHA256 = (
    "4b71a1d5337e0de45fc732ea1ee0007eb35ea9f49995a6cf2f3f59ab82c7a37f"
)
CLAIMED_PREUPLOAD_V9_CLAIM_SHA256 = (
    "3435ec34f23975771e57d2d7d6f17b2e6d9463dff8439b4eae7aedb35a46255c"
)
CLAIMED_PREUPLOAD_V11_CLAIM_SHA256 = (
    "b480484bbd52654bf29d4c78e6594d86f3b440587e54eca125d0f4232e63f72c"
)
CLAIMED_PREUPLOAD_V13_CLAIM_SHA256 = (
    "809914e7bf28f99ef912cb57e65cc128c3b0d19d819d52fc88ea80090ce3220e"
)
CLAIMED_PREUPLOAD_V15_CLAIM_SHA256 = (
    "fe35182f8740a64482707f68ceb307f7bd34c86caa3114612291300e7a817d91"
)
CLAIMED_PREUPLOAD_V16_CLAIM_SHA256 = (
    "babe9e91701a4a35d054aa950a05309dfc3b83a62430f35639a362505a08b17d"
)
CLAIMED_PREUPLOAD_V17_CLAIM_SHA256 = (
    "77972047e50b33f9e03ab08926c0536e6e06ee64e2d2a912a9d2b86f88c5082e"
)
CLAIMED_PREUPLOAD_V18_CLAIM_SHA256 = (
    "46bb20b863686ffc992a32f1f66a8fa5bbeaa27a49c7c2e1477a5899713dde7e"
)
CLAIMED_PREUPLOAD_V19_CLAIM_SHA256 = (
    "4f4b6fbd4878985302ad6b25bcf3708f5c773c3ea69a1e715cb43c99a5b35740"
)
V20_RESULT_SUMMARY_SHA256 = (
    "cea85088ba4d3529ee919e12a6578fa49e07b7c451791697d5b6a16a3eb93dac"
)
V20_INSPECTION_SHA256 = (
    "b04aad7eb3ff299d34a6764be3a899258f43a7116f9661bfa69445a61a8e0dbe"
)
V20_COMMIT_SHA256 = (
    "b4b5898355769e78ac4b065e958ad21aa03484d6b9ed848a3286f04c45ceaf83"
)
V20_RESUME_SHA256 = (
    "23f69a4909c563863d996f5305c39afbc834e02e754da7f6b9fbf220826246f7"
)
CLAIMED_PREUPLOAD_V21_CLAIM_SHA256 = (
    "e40cc78c05a0d003b2c602465efe9597fe400c59e406bf3deef3b0ceed359b78"
)
V21_RESULT_SUMMARY_SHA256 = (
    "982711f726b219e6b1ee200c636bce17b0170ba2aded91b6d0cfd4e59bca65e4"
)
V21_INSPECTION_SHA256 = (
    "fc827bfd9c2f69f06dba71ed5544a2a8440e7035403b22a26e65dd5fc5353a7e"
)
V21_COMMIT_SHA256 = (
    "a602f39ed942694b4106cccf680a9d41f3a96ae37f858facea78640d1d530eb8"
)
V21_RESUME_SHA256 = (
    "fc20f1a9b076f369e0c9e2f3418488175a5ab33ef31609cdeb4f39cb81214777"
)
V23_RESULT_SUMMARY_SHA256 = (
    "55b4fbefe204632e34b5ae23f324adca1e615f11832c13de04f9d8c304228c3a"
)
V23_INSPECTION_SHA256 = (
    "b20e50c5790d1062445d626d828002f3f8549bacea6d9a29eec8a2fdb95da076"
)
V23_COMMIT_SHA256 = (
    "c96830236899275486f8ed12c3712ac94438d5f20acab711eb6c00d51032bd6c"
)
V23_RESUME_SHA256 = (
    "a389a0ba92fc02ca1f8f5e1f37f7768f62d7a5052cc56c92a963f2573063e2e1"
)
V24_RESULT_SUMMARY_SHA256 = (
    "a450e54220d3687a40bd7e48773469df6929945fc3d720a77922bd04574d8b97"
)
V24_INSPECTION_SHA256 = (
    "6702161d149209c420f9dc725228a49ee61a6b373d0ab8fd117023194ba8edc2"
)
V24_COMMIT_SHA256 = (
    "3e85e2795966967e70f3922b67c2b181282f70705e220717c572ebc6a99b533a"
)
V24_RESUME_SHA256 = (
    "613083eb50bb8d322a2fcd3464515f243692ea79adabdd4ac25e9050d3eef938"
)
V26_RESULT_SUMMARY_SHA256 = (
    "d32c4ab4a26feeff0f210598278125199dafb36c58d2466830e85f5aab3314d1"
)
V26_INSPECTION_SHA256 = (
    "9d516ab88e8ec492882ec83ba267ba3155219e7494bfc58c9cd26587c46691e4"
)
V26_COMMIT_SHA256 = (
    "13834b8694b8ff3ffc1c1dabb2bdb864b94cb21c2c47c73e9084978d500e63c7"
)
V26_RESUME_SHA256 = (
    "e5f156ea79d04ba04137b76e060d4bb0955ca1b1b631cfafc2fe92ddfb520360"
)
V26_ERROR_CODE = (
    'plan8_v6_bound_draft_editor_not_unique:{"candidates": [], '
    '"leaf_match_count": 0, "match_count": 0, '
    '"semantic_match_count": 0, "s'
)
V27_RESULT_SUMMARY_SHA256 = (
    "f6b99abf337850ae5dbc4b537bb0830db3c8d70c396ddec9be880fa0f4d7a2cc"
)
V27_INSPECTION_SHA256 = (
    "77562fdc96b830ea7a177759b0935fe55b228ade40a6bef537034da0ea3a5cfb"
)
V27_COMMIT_SHA256 = (
    "634f4c8a0c29d0b73437f6adad2b1e7f14b87e7b2120b4fb1b9b546b42d1707c"
)
V27_RESUME_SHA256 = (
    "11f37c662d764be30aa2b27ceddd85c4a1ef07e668b4211b2c088ed226856c95"
)
V27_MANUAL_EXPORT_SHA256 = (
    "80905fa7ff906a2b219328d111096ed05329f6b5ecaa058e37308293914d3338"
)
POST_V18_READBACK_ARTIFACT_SHA256 = (
    "3c44dc294b8dc8b098d09e944c942a6bfea27814df1b78f2b90fc7d8998cd010"
)
CLAIMED_PREUPLOAD_LAST_STEP = "draft_patch_terminal"
CLAIMED_PREUPLOAD_ERROR_CODE = "plan8_v8_unknown_outcome_no_retry"
POST_READBACK_RESULT_SUMMARY_SHA256 = (
    "956dc7d744a45800924e93afa060042f98a95cda8a8b9f858b3ca403afeddcb4"
)
POST_READBACK_DETAIL_SHA256 = (
    "0bec6be3c5107d28ba79fc3736a85bea93f1ae339782ae3be73439e56d90ec1d"
)
POST_READBACK_MISSING_SKU_IDS = [
    "6234601898881", "6234601898883", "6234601898885",
    "6234601898887", "6287431318354", "6287431318356",
    "6287431318358", "6287431318360",
]
V5_RESULT_SUMMARY_SHA256 = (
    "8b6b5d5f5cbb2546fb4838c33757aa22685496104dd9a37d9469c85a48af2394"
)
V5_INSPECTION_SHA256 = (
    "1a49d4a3822c06d388d2e17d386aaf813bc15326badb0056c424b9f3e70b7ac6"
)
V5_COMMIT_SHA256 = (
    "a29c273e9f264ba207db1fabf4d04ae8a04eb5489bd2ffd81fc40050d9f00aa3"
)
V7_RESULT_SUMMARY_SHA256 = (
    "164ae521bfabde2f50b837b388c8f479670b75f23084c3b16015538512d91b6e"
)
V7_INSPECTION_SHA256 = (
    "7a865a89280df25b1d49e5d82d64d014a333d6b8cb673b430b3d29d298b8c2f1"
)
V7_COMMIT_SHA256 = (
    "08193800da6df7fbb25a7b57ef0fc69c19001c04d3f3b52921be18d24a7825dc"
)
V8_RESULT_SUMMARY_SHA256 = (
    "1056020857c41455ba1b40b3648a35ea220a199753ada573b43ca513f0e7fcf8"
)
V8_INSPECTION_SHA256 = (
    "a091bb015f4c57dd35303dd4eb6613405ddee894e119d36eb9f9c0769b792c44"
)
V8_COMMIT_SHA256 = (
    "a602f39ed942694b4106cccf680a9d41f3a96ae37f858facea78640d1d530eb8"
)
V9_RESULT_SUMMARY_SHA256 = (
    "8245b773c08f627b8eff3eb56a6d8581cd1e3d9e45bf8acadf4214de371da452"
)
V9_INSPECTION_SHA256 = (
    "e8c06efc6bce3ce9dd7fd65bc730b4bb6c17acd12a93493155d19f5c370b24ee"
)
V9_COMMIT_SHA256 = (
    "da096ffab840c1514aba5fe1ff65d701481c46e5b0696e6cc1ca4fe243386976"
)
V10_RESULT_SUMMARY_SHA256 = (
    "23294c83f5a806053bf02b722748d18113204006b61711dc03af4e475c0a8337"
)
V10_INSPECTION_SHA256 = (
    "425ed53a2d289e69e39738fd44d6e44b8dcaf24b85fd017e5d2cabd9536bf243"
)
V10_COMMIT_SHA256 = (
    "381fb1a1f5c43f895e6cacad89f3ae730b4435ac9b04fbb55a6c292bd48bc4e7"
)
V11_RESULT_SUMMARY_SHA256 = (
    "0fecab0b102c7749eefe9b60b7266d7328ec5fa24bc9ba38a94bb5235e8ac1b0"
)
V11_INSPECTION_SHA256 = (
    "2135130a81b43b1694f28201ee23e55b239e9eda4cf6a2a6dd4b08c0cdaa5bf3"
)
V11_COMMIT_SHA256 = (
    "a602f39ed942694b4106cccf680a9d41f3a96ae37f858facea78640d1d530eb8"
)
V12_RESULT_SUMMARY_SHA256 = (
    "6153cea4f5bf8b9797120e8006e2282f53d12240261b650add4321fef3f6548d"
)
V12_INSPECTION_SHA256 = (
    "0accf7eb9bdc1bfde0daa6484990748b3b11806ac5d87a91733d3a443519bf4f"
)
V12_COMMIT_SHA256 = (
    "586df84070e1450efbac5f3b209a2fe08a5f5d767ebe7be3c087bb8b8fc11075"
)
V13_RESULT_SUMMARY_SHA256 = (
    "65f4186a096d1caf9cd2bddb5df606b149b2fbe8566c4c5e361636e77293b545"
)
V13_INSPECTION_SHA256 = (
    "ae12648d31c29621ef0b8205287a77e626d74e28dc8141db0bef08ae546ce816"
)
V13_COMMIT_SHA256 = (
    "a602f39ed942694b4106cccf680a9d41f3a96ae37f858facea78640d1d530eb8"
)
V14_RESULT_SUMMARY_SHA256 = (
    "258319122fc339157686c1e184df898dc594a3a8e1deb97e943a337b3df12ba7"
)
V14_INSPECTION_SHA256 = (
    "aff6f2e860b66706c0143fccb2fe0e38f7b7a0fca1a2813b4791d2447b49410a"
)
V14_COMMIT_SHA256 = (
    "1245dcaf0995fe3d2c06da47f46e8f06b1d4e3ed71f286eb46aef5c86c947ad7"
)
V15_RESULT_SUMMARY_SHA256 = (
    "e4997acd94ef81bc63835573c55ab1fb5abe5d09936e0ddc745d3ed052c60971"
)
V15_INSPECTION_SHA256 = (
    "9851d41b2355eeac639ff5fe096d91161fe4d0e447d7d30a01ef8f0a1e874e3e"
)
V15_COMMIT_SHA256 = (
    "a602f39ed942694b4106cccf680a9d41f3a96ae37f858facea78640d1d530eb8"
)
V16_RESULT_SUMMARY_SHA256 = (
    "f29f6feea531e84bd0fef2808e5242e9040f449a7ebecf769c0191586cad3ee8"
)
V16_INSPECTION_SHA256 = (
    "7331a11dcc159a671def0021728519b81ee3542022a6fde95c780fa96ca08a54"
)
V16_COMMIT_SHA256 = (
    "a602f39ed942694b4106cccf680a9d41f3a96ae37f858facea78640d1d530eb8"
)
V19_RESULT_SUMMARY_SHA256 = (
    "601a6d950b66b08ab1fcef72d5a22dcb7c95eab757a78c9c8efd0b72c30b1289"
)
V19_INSPECTION_SHA256 = (
    "ca39b5bf7aa604310b1e8dd3679bd887e73ec32c1b6e51b142469b9077b4926d"
)
V19_COMMIT_SHA256 = (
    "a602f39ed942694b4106cccf680a9d41f3a96ae37f858facea78640d1d530eb8"
)


def _boundary(*, platform_write: bool = False) -> dict:
    return {**v7._boundary(platform_write=platform_write),
            "activity_create": False,
            "existing_activity_edit": platform_write,
            "v7_execute_retry": False}


def _fail(error: str, **detail) -> dict:
    return {"ok": False, "error": error, **detail,
            "execution_boundary": _boundary(platform_write=False)}


def _fixed_manifest(target_rows: list[dict], discount_rows: list[dict],
                    policy_sha: str) -> dict:
    manifest = v7._fixed_manifest(target_rows, discount_rows, policy_sha)
    manifest["recovery_version"] = RECOVERY_VERSION
    manifest.pop("recovery_evidence", None)
    manifest["resume_evidence"] = dict(EXPECTED_RESUME_EVIDENCE)
    manifest["execution_order"] = list(EXECUTION_ORDER)
    return manifest


def _attempts(db: Session) -> list[CampaignExecutionAttempt]:
    return list(db.execute(select(CampaignExecutionAttempt).where(
        CampaignExecutionAttempt.workflow_key == WORKFLOW_KEY,
        CampaignExecutionAttempt.operation == OPERATION,
    )).scalars())


def _attempt_for_scope(db: Session,
                       scope_sha256: str) -> CampaignExecutionAttempt | None:
    return db.execute(select(CampaignExecutionAttempt).where(
        CampaignExecutionAttempt.workflow_key == WORKFLOW_KEY,
        CampaignExecutionAttempt.operation == OPERATION,
        CampaignExecutionAttempt.scope_sha256 == scope_sha256,
    )).scalar_one_or_none()


def _validate_prerequisite(db: Session) -> tuple[bool, dict]:
    row = db.get(CampaignExecutionAttempt, V7_ATTEMPT_ID)
    manifest = ((row.result_summary or {}).get("manifest")
                if row is not None else None)
    detail = {
        "attempt_id": V7_ATTEMPT_ID,
        "operation": getattr(row, "operation", None),
        "state": getattr(row, "state", None),
        "write_claimed": getattr(row, "write_claimed", None),
        "platform_write_observed": getattr(
            row, "platform_write_observed", None),
        "last_step": getattr(row, "last_step", None),
        "manifest_scope_sha256": getattr(row, "scope_sha256", None),
    }
    ok = bool(
        row is not None
        and row.operation == "plan8_final_recovery_v7"
        and row.state == "unknown_no_retry"
        and row.write_claimed is True
        and row.platform_write_observed is not True
        and row.last_step in {"discount_terminal", "readback_not_complete"}
        and isinstance(manifest, dict)
        and v6._hash(manifest) == row.scope_sha256)
    return ok, detail


def validate_inspection(result: dict, manifest: dict,
                        manifest_sha256: str) -> tuple[bool, dict]:
    base_ok, detail = v7.validate_inspection(
        result, manifest, manifest_sha256)
    evidence = result.get("resume_evidence") or {}
    evidence_ok = bool(evidence.get("ok") is True and all(
        evidence.get(key) == value
        for key, value in EXPECTED_RESUME_EVIDENCE.items()))
    return bool(base_ok and evidence_ok), {
        **detail, "resume_evidence": evidence,
        "web_agent_error": result.get("error"),
        "web_agent_status": result.get("status"),
        "web_agent_step": result.get("step"),
        "web_agent_facts": result.get("facts"),
        "web_agent_claim_created": result.get("claim_created"),
        "web_agent_need_scan": result.get("need_scan"),
        "v8_claim_absent": result.get("v8_claim_absent"),
        "v8_claim_sha256": result.get("v8_claim_sha256"),
    }


def validate_commit(result: dict, manifest: dict,
                    manifest_sha256: str) -> tuple[bool, dict]:
    rewritten = {**result, "checkpoints": v6.EXPECTED_COMMIT_CHECKPOINTS}
    base_ok, detail = v6.validate_commit(
        rewritten, manifest, manifest_sha256)
    checkpoints_ok = result.get("checkpoints") == EXPECTED_COMMIT_CHECKPOINTS
    return bool(base_ok and checkpoints_ok), {
        **detail, "checkpoints": result.get("checkpoints"),
        "v8_checkpoint_order_ok": checkpoints_ok,
        "web_agent_error": result.get("error"),
        "web_agent_error_code": result.get("error_code"),
        "web_agent_status": result.get("status"),
        "last_checkpoint": result.get("last_checkpoint"),
        "claim_created": result.get("claim_created"),
        "different_fields": result.get("different_fields") or [],
        "web_agent_detail": result.get("detail"),
        "candidate_price_evidence": result.get("candidate_price_evidence"),
        "reservation_consumed": result.get("reservation_consumed"),
        "web_agent_job_id": result.get("web_agent_job_id"),
    }


def validate_readback(result: dict, manifest: dict,
                      manifest_sha256: str) -> tuple[bool, dict]:
    return v6.validate_readback(result, manifest, manifest_sha256)


def verify_plan8_final_v8_claim(
        db: Session, *, attempt_id: str, workflow_key: str, plan_id: int,
        operation: str, scope_sha256: str, inspect_scope_sha256: str,
        reservation_token_sha256: str) -> dict:
    attempt = db.get(CampaignExecutionAttempt, attempt_id)
    manifest = ((attempt.result_summary or {}).get("manifest")
                if attempt is not None else None)
    baseline = ((manifest or {}).get("inspection_baseline")
                if isinstance(manifest, dict) else None)
    try:
        expires = float((baseline or {}).get(
            "reservation_expires_at_epoch") or 0)
    except (TypeError, ValueError):
        expires = 0
    verified = bool(
        attempt is not None and workflow_key == WORKFLOW_KEY
        and plan_id == PLAN_ID and operation == OPERATION
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION and attempt.scope_sha256 == scope_sha256
        and attempt.state == "write_claimed" and attempt.write_claimed is True
        and attempt.write_claimed_at is not None and bool(attempt.request_id)
        and isinstance(manifest, dict) and v6._hash(manifest) == scope_sha256
        and manifest.get("resume_evidence") == EXPECTED_RESUME_EVIDENCE
        and isinstance(baseline, dict)
        and baseline.get("inspect_scope_sha256") == inspect_scope_sha256
        and baseline.get("reservation_token_sha256")
        == reservation_token_sha256
        and expires > datetime.now(timezone.utc).timestamp())
    return {
        "ok": verified, "verified": verified, "attempt_id": attempt_id,
        "workflow_key": WORKFLOW_KEY, "plan_id": PLAN_ID,
        "operation": OPERATION, "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", False),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "inspect_scope_sha256": (baseline.get("inspect_scope_sha256")
                                 if isinstance(baseline, dict) else None),
        "reservation_token_sha256": (
            baseline.get("reservation_token_sha256")
            if isinstance(baseline, dict) else None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "execution_boundary": {**_boundary(platform_write=False),
                               "platform_read": False},
    }


def verify_plan8_final_v8_preupload_claim(
        db: Session, *, attempt_id: str, workflow_key: str, plan_id: int,
        operation: str, scope_sha256: str, inspect_scope_sha256: str,
        reservation_token_sha256: str, resume_claim_sha256: str) -> dict:
    attempt = db.get(CampaignExecutionAttempt, attempt_id)
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    resume = summary.get("claimed_preupload_resume") or {}
    try:
        expires = float(resume.get("reservation_expires_at_epoch") or 0)
    except (TypeError, ValueError):
        expires = 0
    v9_claim_steps = {
        "platform_write_claim_claimed_preupload_resume_v9",
        "platform_write_claim_claimed_preupload_resume_v10",
        "platform_write_claim_claimed_preupload_resume_v11",
    }
    step = getattr(attempt, "last_step", None)
    expected_resume_claim_sha256 = (
        CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        if step in {
            "platform_write_claim_claimed_preupload_resume_v22",
            "platform_write_claim_claimed_preupload_resume_v23",
            "platform_write_claim_claimed_preupload_resume_v24",
            "platform_write_claim_claimed_preupload_resume_v25",
            "platform_write_claim_claimed_preupload_resume_v26",
            "platform_write_claim_claimed_preupload_resume_v27",
            "platform_write_claim_claimed_preupload_resume_v28"}
        else
        CLAIMED_PREUPLOAD_V19_CLAIM_SHA256
        if step in {
            "platform_write_claim_claimed_preupload_resume_v20",
            "platform_write_claim_claimed_preupload_resume_v21"}
        else (
        CLAIMED_PREUPLOAD_V18_CLAIM_SHA256
        if step == "platform_write_claim_claimed_preupload_resume_v19"
        else (
        CLAIMED_PREUPLOAD_V17_CLAIM_SHA256
        if step == "platform_write_claim_claimed_preupload_resume_v18"
        else (
        CLAIMED_PREUPLOAD_V16_CLAIM_SHA256
        if step == "platform_write_claim_claimed_preupload_resume_v17"
        else (
        CLAIMED_PREUPLOAD_V15_CLAIM_SHA256
        if step == "platform_write_claim_claimed_preupload_resume_v16"
        else (
        CLAIMED_PREUPLOAD_V13_CLAIM_SHA256
        if step in {
            "platform_write_claim_claimed_preupload_resume_v14",
            "platform_write_claim_claimed_preupload_resume_v15"}
        else (CLAIMED_PREUPLOAD_V11_CLAIM_SHA256
        if step in {
            "platform_write_claim_claimed_preupload_resume_v12",
            "platform_write_claim_claimed_preupload_resume_v13"}
        else (CLAIMED_PREUPLOAD_V9_CLAIM_SHA256
              if step in v9_claim_steps
        else CLAIMED_PREUPLOAD_CLAIM_SHA256))))))))
    verified = bool(
        attempt is not None and workflow_key == WORKFLOW_KEY
        and plan_id == PLAN_ID and operation == OPERATION
        and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == scope_sha256
        == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "write_claimed" and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and isinstance(manifest, dict) and v6._hash(manifest) == scope_sha256
        and resume == {
            "source_claim_sha256": expected_resume_claim_sha256,
            "inspect_scope_sha256": inspect_scope_sha256,
            "reservation_token_sha256": reservation_token_sha256,
            "reservation_expires_at_epoch": expires,
        }
        and resume_claim_sha256 == expected_resume_claim_sha256
        and expires > datetime.now(timezone.utc).timestamp())
    return {
        "ok": verified, "verified": verified, "attempt_id": attempt_id,
        "workflow_key": WORKFLOW_KEY, "plan_id": PLAN_ID,
        "operation": OPERATION, "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", False),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "inspect_scope_sha256": inspect_scope_sha256,
        "reservation_token_sha256": reservation_token_sha256,
        "resume_claim_sha256": resume_claim_sha256,
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "execution_boundary": {**_boundary(platform_write=False),
                               "platform_read": False},
    }


def _readback_existing(db: Session, plan: CampaignPlan,
                       attempt: CampaignExecutionAttempt) -> dict:
    if str(plan.status or "") not in READBACK_PLAN_STATUSES:
        return _fail("plan8_final_v8_readback_plan_status_not_allowed",
                     actual_status=plan.status, attempt_id=attempt.id)
    manifest = (attempt.result_summary or {}).get("manifest")
    if not isinstance(manifest, dict):
        return _fail("plan8_final_v8_attempt_manifest_missing",
                     attempt_id=attempt.id)
    if v6._hash(manifest) != attempt.scope_sha256:
        return _fail("plan8_final_v8_attempt_scope_mismatch",
                     attempt_id=attempt.id)
    try:
        result = web_agent_service.recover_plan8_final_v8(
            db, payload={"phase": "readback",
                         "scope_sha256": attempt.scope_sha256,
                         "manifest": manifest, "attempt_id": attempt.id})
    except Exception as exc:  # noqa: BLE001
        result = {"ok": False, "error": type(exc).__name__,
                  "platform_write": False}
    ok, detail = validate_readback(result, manifest, attempt.scope_sha256)
    if not ok:
        prior = dict(attempt.result_summary or {})
        prior["last_readback"] = detail
        attempt.result_summary = prior
        attempt.last_step = "readback_not_complete"
        attempt.error_code = "post_submit_readback_not_complete"
        attempt.web_agent_job_id = str(
            result.get("web_agent_job_id") or "")[:64] or attempt.web_agent_job_id
        db.commit()
        return _fail("plan8_final_v8_readback_not_complete",
                     attempt_id=attempt.id, readback=detail,
                     need_scan=bool(result.get("need_scan")))
    campaign_execution_service.record_platform_terminal(
        db, attempt, state="completed",
        platform_write_observed=attempt.platform_write_observed,
        step="readback_verified", job_id=detail.get("web_agent_job_id"),
        result_summary={**dict(attempt.result_summary or {}),
                        "manifest": manifest, "readback": detail})
    plan.status = "reconciled"
    db.commit()
    return {"ok": True, "readback_only": True, "attempt_id": attempt.id,
            "workflow_key": WORKFLOW_KEY, "plan_id": PLAN_ID,
            "plan_status": plan.status, "verification": detail,
            "execution_boundary": _boundary(platform_write=False)}


def _validate_preclaim_resume_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    manifest = ((attempt.result_summary or {}).get("manifest")
                if attempt is not None else None)
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
    }
    ok = bool(
        attempt is not None
        and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID
        and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == PRECLAIM_SCOPE_SHA256
        and attempt.state == "unknown_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is None
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == PRECLAIM_LAST_STEP
        and attempt.error_code == PRECLAIM_ERROR_CODE
        and attempt.web_agent_job_id == PRECLAIM_WEB_AGENT_JOB_ID
        and isinstance(manifest, dict)
        and v6._hash(manifest) == PRECLAIM_SCOPE_SHA256
    )
    return ok, detail


def _validate_claimed_preupload_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == CLAIMED_PREUPLOAD_LAST_STEP
        and attempt.error_code == CLAIMED_PREUPLOAD_ERROR_CODE
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == CLAIMED_PREUPLOAD_LAST_STEP
        and commit.get("web_agent_error") == CLAIMED_PREUPLOAD_ERROR_CODE
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_readback_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    commit = summary.get("commit") or {}
    readback = summary.get("last_readback") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "last_readback_sha256": v6._hash(readback),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "readback_not_complete"
        and attempt.error_code == "post_submit_readback_not_complete"
        and attempt.web_agent_job_id == "job3"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == POST_READBACK_RESULT_SUMMARY_SHA256
        and v6._hash(readback) == POST_READBACK_DETAIL_SHA256
        and readback.get("record_count") == 6
        and readback.get("sku_count") == 70
        and readback.get("custom_sku_count") == 18
        and readback.get("missing_sku_ids") == POST_READBACK_MISSING_SKU_IDS
        and readback.get("unexpected_sku_ids") == []
        and readback.get("discount_rows") == []
        and readback.get("web_agent_job_id") == "job3"
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == CLAIMED_PREUPLOAD_LAST_STEP
        and commit.get("web_agent_error") == CLAIMED_PREUPLOAD_ERROR_CODE
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_lease_scope_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only the exact V5 no-write stop caused by lease-token drift."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v8_state_changed_before_claim"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V5_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V5_INSPECTION_SHA256
        and v6._hash(commit) == V5_COMMIT_SHA256
        and commit.get("platform_write") is False
        and commit.get("claim_created") is False
        and commit.get("web_agent_error")
        == "plan8_v8_state_changed_before_claim"
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_lease_expiry_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only the exact V7 no-write stop after its lease expired."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v8_erp_claim_not_verified"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V7_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V7_INSPECTION_SHA256
        and v6._hash(commit) == V7_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("claim_created") is False
        and commit.get("web_agent_error") == "plan8_v8_erp_claim_not_verified"
        and (commit.get("web_agent_detail") or {}).get("error")
        == "erp_preupload_claim_verify_unavailable"
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_dialog_mismatch_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V8's frozen no-write batch-import-dialog mismatch."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "draft_patch_terminal"
        and attempt.error_code == "plan8_v8_unknown_outcome_no_retry"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V8_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V8_INSPECTION_SHA256
        and v6._hash(commit) == V8_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_campaign_guard_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V9's frozen pre-claim read-only campaign-shell stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "unknown_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is None
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v6_bound_draft_campaign_guard_failed"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V9_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V9_INSPECTION_SHA256
        and v6._hash(commit) == V9_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V9_CLAIM_SHA256
        and commit.get("platform_write") is None
        and commit.get("claim_created") is False
        and commit.get("web_agent_error")
        == "plan8_v6_bound_draft_campaign_guard_failed"
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_claim_verify_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V10's frozen zero-write claim-verifier rejection."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    web_detail = commit.get("web_agent_detail") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v8_erp_claim_not_verified"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V10_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V10_INSPECTION_SHA256
        and v6._hash(commit) == V10_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V9_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("claim_created") is False
        and commit.get("web_agent_error") == "plan8_v8_erp_claim_not_verified"
        and web_detail.get("error") == "erp_preupload_claim_verify_rejected"
        and web_detail.get("http_status") == 409
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_lazy_import_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V11's frozen no-upload lazy-file-input stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "draft_patch_terminal"
        and attempt.error_code == "plan8_v8_unknown_outcome_no_retry"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V11_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V11_INSPECTION_SHA256
        and v6._hash(commit) == V11_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V9_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_allowlist_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V12's frozen zero-write local claim allowlist stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    web_detail = commit.get("web_agent_detail") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v8_erp_claim_not_verified"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V12_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V12_INSPECTION_SHA256
        and v6._hash(commit) == V12_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V11_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is None
        and commit.get("claim_created") is False
        and commit.get("web_agent_error") == "plan8_v8_erp_claim_not_verified"
        and web_detail == {
            "ok": False,
            "error": "erp_preupload_claim_verify_request_invalid",
        }
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_semantic_modal_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V13's frozen pre-file semantic-modal stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "draft_patch_terminal"
        and attempt.error_code == "plan8_v8_unknown_outcome_no_retry"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V13_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V13_INSPECTION_SHA256
        and v6._hash(commit) == V13_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V11_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_editor_identity_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V14's frozen read-only editor-label formatting stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "unknown_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is None
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v6_bound_draft_editor_identity_mismatch"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V14_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V14_INSPECTION_SHA256
        and v6._hash(commit) == V14_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V13_CLAIM_SHA256
        and commit.get("platform_write") is None
        and commit.get("reservation_consumed") is None
        and commit.get("claim_created") is False
        and commit.get("last_checkpoint") is None
        and commit.get("web_agent_error")
        == "plan8_v6_bound_draft_editor_identity_mismatch"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_nested_modal_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V15's exact no-file-selection modal-binding stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "draft_patch_terminal"
        and attempt.error_code == "plan8_v8_unknown_outcome_no_retry"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V15_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V15_INSPECTION_SHA256
        and v6._hash(commit) == V15_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V13_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_moban_text_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V16's exact no-file-selection 下载模版 stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "draft_patch_terminal"
        and attempt.error_code == "plan8_v8_unknown_outcome_no_retry"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V16_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V16_INSPECTION_SHA256
        and v6._hash(commit) == V16_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V15_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_postupload_after_unpersisted_readback(
        attempt: CampaignExecutionAttempt | None, *,
        expected_resume_claim_sha256: str = CLAIMED_PREUPLOAD_V16_CLAIM_SHA256,
        expected_artifact_sha256: str = "") -> tuple[bool, dict]:
    """Accept V17 only after an independent official readback proved no patch."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    readback = summary.get("last_readback") or {}
    baseline = ((manifest or {}).get("inspection_baseline")
                if isinstance(manifest, dict) else {}) or {}
    protected = {
        str(row.get("item_id") or ""): str(row.get("after_hash") or "")
        for row in readback.get("protected_records") or []
    }
    expected_protected = {
        str(key): str(value) for key, value in (
            baseline.get("protected_record_before_hashes") or {}).items()
    }
    missing = sorted(str(value) for value in (
        readback.get("missing_sku_ids") or []))
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "commit": commit,
        "last_readback": readback,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is True
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "readback_not_complete"
        and attempt.error_code == "post_submit_readback_not_complete"
        and attempt.web_agent_job_id == "job4"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and inspection.get("resume_claim_sha256")
        == expected_resume_claim_sha256
        and commit.get("platform_write") is True
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_job_id") == "job3"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written")
        and readback.get("record_count") == 6
        and readback.get("sku_count") == 70
        and readback.get("custom_sku_count") == 18
        and missing == sorted(POST_READBACK_MISSING_SKU_IDS)
        and not readback.get("unexpected_sku_ids")
        and not readback.get("discount_rows")
        and protected == expected_protected
        and (readback.get("legacy_discount_baseline") or {}).get("sha256")
        == baseline.get("legacy_discount_sha256")
        and (readback.get("legacy_discount_baseline") or {}).get("row_count")
        == 53
        and sorted(readback.get("all_record_ids") or [])
        == sorted(baseline.get("all_record_ids") or [])
        and sorted(readback.get("excluded_item_ids") or [])
        == sorted([v6.ZERO_SALES_EXCLUDED_ITEM_ID,
                   v6.WAREHOUSE_EXCLUDED_ITEM_ID])
        and isinstance(readback.get("artifact_sha256"), str)
        and len(readback["artifact_sha256"]) == 64
        and (not expected_artifact_sha256
             or readback["artifact_sha256"] == expected_artifact_sha256))
    return ok, detail


def _validate_claimed_postupload_after_template_reject_readback(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept V18 only after the exact official readback proved no rows."""
    return _validate_claimed_postupload_after_unpersisted_readback(
        attempt,
        expected_resume_claim_sha256=CLAIMED_PREUPLOAD_V17_CLAIM_SHA256,
        expected_artifact_sha256=POST_V18_READBACK_ARTIFACT_SHA256)


def _validate_claimed_preupload_after_template_generation_attempt(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V19's exact zero-upload template-download timeout."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "draft_patch_terminal"
        and attempt.error_code == "plan8_v8_unknown_outcome_no_retry"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V19_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V19_INSPECTION_SHA256
        and v6._hash(commit) == V19_COMMIT_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V18_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("reservation_consumed") is True
        and commit.get("claim_created") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_v20_claim_verify_rejection(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V20's proven pre-platform claim-step bookkeeping defect."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    resume = summary.get("claimed_preupload_resume") or {}
    web_detail = commit.get("web_agent_detail") or {}
    verify_response = ((web_detail.get("response") or {}).get("detail") or {})
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "resume_sha256": v6._hash(resume),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v8_erp_claim_not_verified"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V20_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V20_INSPECTION_SHA256
        and v6._hash(commit) == V20_COMMIT_SHA256
        and v6._hash(resume) == V20_RESUME_SHA256
        and resume.get("source_claim_sha256")
        == CLAIMED_PREUPLOAD_V19_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("claim_created") is False
        and commit.get("reservation_consumed") is None
        and commit.get("web_agent_error") == "plan8_v8_erp_claim_not_verified"
        and web_detail.get("error") == "erp_preupload_claim_verify_rejected"
        and web_detail.get("http_status") == 409
        and verify_response.get("verified") is False
        and verify_response.get("platform_write_observed") is False
        and verify_response.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V19_CLAIM_SHA256
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_v21_template_obstruction(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V21's exact no-upload template-modal obstruction."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    resume = summary.get("claimed_preupload_resume") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "resume_sha256": v6._hash(resume),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "draft_patch_terminal"
        and attempt.error_code == "plan8_v8_unknown_outcome_no_retry"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V21_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V21_INSPECTION_SHA256
        and v6._hash(commit) == V21_COMMIT_SHA256
        and v6._hash(resume) == V21_RESUME_SHA256
        and resume.get("source_claim_sha256")
        == CLAIMED_PREUPLOAD_V19_CLAIM_SHA256
        and commit.get("platform_write") is False
        and commit.get("claim_created") is True
        and commit.get("reservation_consumed") is True
        and commit.get("last_checkpoint") == "draft_patch_terminal"
        and commit.get("web_agent_error")
        == "plan8_v8_unknown_outcome_no_retry"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_v23_export_failure(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V23's exact read-only QianNiu export failure."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    resume = summary.get("claimed_preupload_resume") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "resume_sha256": v6._hash(resume),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "unknown_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is None
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "export_poll_timeout"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V23_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V23_INSPECTION_SHA256
        and v6._hash(commit) == V23_COMMIT_SHA256
        and v6._hash(resume) == V23_RESUME_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and resume.get("source_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and commit.get("scope_sha256") == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and commit.get("platform_write") is None
        and commit.get("claim_created") is False
        and commit.get("reservation_consumed") is None
        and commit.get("last_checkpoint") is None
        and commit.get("web_agent_error") == "export_poll_timeout"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_v24_title_mismatch(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V24's exact zero-write stop before any file selection."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    resume = summary.get("claimed_preupload_resume") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "resume_sha256": v6._hash(resume),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "unknown_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is None
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "campaign_title_mismatch"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V24_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V24_INSPECTION_SHA256
        and v6._hash(commit) == V24_COMMIT_SHA256
        and v6._hash(resume) == V24_RESUME_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and resume.get("source_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and commit.get("scope_sha256") == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and commit.get("platform_write") is None
        and commit.get("claim_created") is False
        and commit.get("reservation_consumed") is None
        and commit.get("last_checkpoint") is None
        and commit.get("web_agent_error") == "campaign_title_mismatch"
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_v26_editor_loading(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V26's exact zero-write blank editor-body stop."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    resume = summary.get("claimed_preupload_resume") or {}
    manual = summary.get("manual_export_v26") or {}
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "resume_sha256": v6._hash(resume),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "unknown_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is None
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "bound_draft_price_readback"
        and attempt.error_code == V26_ERROR_CODE
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V26_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V26_INSPECTION_SHA256
        and v6._hash(commit) == V26_COMMIT_SHA256
        and v6._hash(resume) == V26_RESUME_SHA256
        and inspection.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and resume.get("source_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and manual.get("sha256") == MANUAL_EXPORT_SHA256
        and manual.get("row_count") == 83
        and manual.get("draft_sku_count") == 70
        and manual.get("published_sku_count") == 13
        and manual.get("platform_write") is False
        and commit.get("step") == "bound_draft_price_readback"
        and commit.get("scope_sha256") is None
        and commit.get("platform_write") is None
        and commit.get("claim_created") is False
        and commit.get("reservation_consumed") is None
        and commit.get("last_checkpoint") is None
        and commit.get("web_agent_error")
        == ('plan8_v6_bound_draft_editor_not_unique:'
            '{"candidates": [], "leaf_match_count": 0, '
            '"match_count": 0, "semantic_match_count": 0, '
            '"source": "identity_subtrees", "topmost_match_count": 0}')
        and commit.get("web_agent_detail") is None
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _validate_claimed_preupload_after_v27_claim_rejection(
        attempt: CampaignExecutionAttempt | None) -> tuple[bool, dict]:
    """Accept only V27's proven zero-write ERP claim allowlist rejection."""
    summary = dict(getattr(attempt, "result_summary", None) or {})
    manifest = summary.get("manifest")
    inspection = summary.get("inspection") or {}
    commit = summary.get("commit") or {}
    resume = summary.get("claimed_preupload_resume") or {}
    manual = summary.get("manual_export_v27") or {}
    web_detail = commit.get("web_agent_detail") or {}
    verify_response = ((web_detail.get("response") or {}).get("detail") or {})
    detail = {
        "attempt_id": getattr(attempt, "id", None),
        "scope_sha256": getattr(attempt, "scope_sha256", None),
        "state": getattr(attempt, "state", None),
        "write_claimed": getattr(attempt, "write_claimed", None),
        "platform_write_observed": getattr(
            attempt, "platform_write_observed", None),
        "automatic_retry_allowed": getattr(
            attempt, "automatic_retry_allowed", None),
        "request_id": getattr(attempt, "request_id", None),
        "last_step": getattr(attempt, "last_step", None),
        "error_code": getattr(attempt, "error_code", None),
        "web_agent_job_id": getattr(attempt, "web_agent_job_id", None),
        "result_summary_sha256": v6._hash(summary),
        "inspection_sha256": v6._hash(inspection),
        "commit_sha256": v6._hash(commit),
        "resume_sha256": v6._hash(resume),
        "manual_export_sha256": v6._hash(manual),
        "commit": commit,
    }
    ok = bool(
        attempt is not None and attempt.id == PRECLAIM_ATTEMPT_ID
        and attempt.plan_id == PLAN_ID and attempt.workflow_key == WORKFLOW_KEY
        and attempt.operation == OPERATION
        and attempt.scope_sha256 == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and attempt.state == "failed_no_retry"
        and attempt.write_claimed is True
        and attempt.platform_write_observed is False
        and attempt.automatic_retry_allowed is False
        and attempt.request_id == PRECLAIM_REQUEST_ID
        and attempt.last_step == "plan8_final_v8_commit"
        and attempt.error_code == "plan8_v8_erp_claim_not_verified"
        and attempt.web_agent_job_id == "job2"
        and isinstance(manifest, dict)
        and v6._hash(manifest) == CLAIMED_PREUPLOAD_SCOPE_SHA256
        and v6._hash(summary) == V27_RESULT_SUMMARY_SHA256
        and v6._hash(inspection) == V27_INSPECTION_SHA256
        and v6._hash(commit) == V27_COMMIT_SHA256
        and v6._hash(resume) == V27_RESUME_SHA256
        and v6._hash(manual) == V27_MANUAL_EXPORT_SHA256
        and resume.get("source_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and manual.get("sha256") == MANUAL_EXPORT_SHA256
        and manual.get("row_count") == 83
        and manual.get("draft_sku_count") == 70
        and manual.get("published_sku_count") == 13
        and manual.get("platform_write") is False
        and commit.get("platform_write") is False
        and commit.get("claim_created") is False
        and commit.get("reservation_consumed") is None
        and commit.get("last_checkpoint") is None
        and commit.get("web_agent_error") == "plan8_v8_erp_claim_not_verified"
        and web_detail.get("error") == "erp_preupload_claim_verify_rejected"
        and web_detail.get("http_status") == 409
        and verify_response.get("verified") is False
        and verify_response.get("state") == "write_claimed"
        and verify_response.get("write_claimed") is True
        and verify_response.get("platform_write_observed") is False
        and verify_response.get("resume_claim_sha256")
        == CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
        and not commit.get("patched_record_ids")
        and not commit.get("published_record_ids")
        and not commit.get("discount_pairs_written"))
    return ok, detail


def _manual_export_v25_evidence(
        *, manifest: dict, filename: str, size: int, sha256: str,
        xlsx_b64: str) -> tuple[dict | None, dict]:
    """Parse and bind the one user-exported workbook without trusting its name."""
    import openpyxl

    try:
        raw = base64.b64decode(xlsx_b64, validate=True)
    except (TypeError, ValueError):
        return None, {"error": "manual_export_base64_invalid"}
    digest = hashlib.sha256(raw).hexdigest()
    if (filename != MANUAL_EXPORT_FILENAME or size != MANUAL_EXPORT_SIZE
            or len(raw) != MANUAL_EXPORT_SIZE
            or sha256.lower() != MANUAL_EXPORT_SHA256
            or digest != MANUAL_EXPORT_SHA256):
        return None, {
            "error": "manual_export_artifact_mismatch", "filename": filename,
            "size": len(raw), "sha256": digest,
        }
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True)
        try:
            if "已报商品列表" not in workbook.sheetnames:
                raise ValueError("manual_export_sheet_missing")
            sheet = workbook["已报商品列表"]
            if hasattr(sheet, "reset_dimensions"):
                sheet.reset_dimensions()
            headers = [str(cell.value or "").strip() for cell in sheet[2]]
            required = {"商品ID", "营销ID", "商品状态", "SKUID"}
            if not required <= set(headers):
                raise ValueError("manual_export_headers_changed")
            indexes = {name: headers.index(name) for name in required}
            current = {"item_id": "", "record_id": "", "status": ""}
            rows = []
            for values in sheet.iter_rows(min_row=4, values_only=True):
                for source, target in (("商品ID", "item_id"),
                                       ("营销ID", "record_id"),
                                       ("商品状态", "status")):
                    value = values[indexes[source]]
                    if value not in (None, ""):
                        current[target] = str(value).strip()
                sku = values[indexes["SKUID"]]
                sku_id = str(sku or "").strip()
                if current["item_id"].isdigit() and sku_id.isdigit():
                    rows.append({**current, "sku_id": sku_id})
        finally:
            workbook.close()
    except Exception as exc:  # noqa: BLE001
        return None, {"error": str(exc), "error_type": type(exc).__name__}

    draft_specs = v6.DRAFT_RECORDS
    protected_specs = v6.PROTECTED_RECORDS
    allowed = {
        (item_id, spec["record_id"]): spec["current_sku_count"]
        for item_id, spec in draft_specs.items()
    } | {
        (item_id, spec["record_id"]): spec["sku_count"]
        for item_id, spec in protected_specs.items()
    }
    grouped: dict[tuple[str, str], list[dict]] = {}
    for row in rows:
        grouped.setdefault((row["item_id"], row["record_id"]), []).append(row)
    duplicate_skus = len({row["sku_id"] for row in rows}) != len(rows)
    group_ok = set(grouped) == set(allowed)
    if group_ok:
        for key, expected_count in allowed.items():
            values = grouped[key]
            expected_status = "草稿" if key[0] in draft_specs else "已发布设定"
            if (len(values) != expected_count
                    or {row["status"] for row in values} != {expected_status}):
                group_ok = False
                break
    manifest_records = {
        str(row.get("item_id") or ""): row
        for row in manifest.get("draft_records") or []
        if isinstance(row, dict)
    }
    final_by_item = {
        item_id: {str(sku) for sku in (
            manifest_records.get(item_id, {}).get("final_sku_ids") or [])}
        for item_id in draft_specs
    }
    actual_draft_by_item = {
        item_id: {row["sku_id"] for row in rows
                  if row["item_id"] == item_id
                  and row["record_id"] == spec["record_id"]}
        for item_id, spec in draft_specs.items()
    }
    expected_current_by_item = {
        item_id: final_by_item[item_id] - set(spec["add_sku_ids"])
        for item_id, spec in draft_specs.items()
    }
    missing = sorted(set(v6.ADD_SKU_IDS) - {row["sku_id"] for row in rows})
    ok = bool(
        len(rows) == 83 and not duplicate_skus and group_ok
        and sum(row["status"] == "草稿" for row in rows) == 70
        and sum(row["status"] == "已发布设定" for row in rows) == 13
        and (manifest.get("final_scope") or {}).get("sku_count")
        == v6.EXPECTED_TARGET_ROW_COUNT
        and actual_draft_by_item == expected_current_by_item
        and missing == sorted(v6.ADD_SKU_IDS))
    evidence = {
        "filename": filename, "size": len(raw), "sha256": digest,
        "row_count": len(rows), "record_count": len(grouped),
        "draft_sku_count": sum(row["status"] == "草稿" for row in rows),
        "published_sku_count": sum(
            row["status"] == "已发布设定" for row in rows),
        "missing_add_sku_ids": missing,
        "frozen_target_sku_count": (
            manifest.get("final_scope") or {}).get("sku_count"),
        "automatic_export": False, "platform_write": False,
    }
    if not ok:
        evidence["error"] = "manual_export_scope_mismatch"
        return None, evidence
    return {"filename": filename, "size": size, "sha256": digest,
            "xlsx_b64": xlsx_b64}, evidence


def _commit_and_readback(
        db: Session, *, plan: CampaignPlan,
        attempt: CampaignExecutionAttempt, manifest: dict,
        manifest_sha: str, inspect_scope_sha: str,
        reservation_token: str, inspection_detail: dict,
        commit_phase: str = "commit", resume_claim_sha256: str = "",
        use_preupload_v9_endpoint: bool = False,
        use_preupload_v10_endpoint: bool = False,
        use_preupload_v12_endpoint: bool = False,
        use_preupload_v14_endpoint: bool = False,
        use_preupload_v15_endpoint: bool = False,
        use_preupload_v16_endpoint: bool = False,
        use_preupload_v17_endpoint: bool = False,
        use_preupload_v18_endpoint: bool = False,
        use_preupload_v19_endpoint: bool = False,
        use_preupload_v20_endpoint: bool = False,
        use_preupload_v21_endpoint: bool = False,
        use_preupload_v22_endpoint: bool = False,
        use_preupload_v23_endpoint: bool = False,
        use_preupload_v24_endpoint: bool = False,
        use_preupload_v25_endpoint: bool = False,
        use_preupload_v26_endpoint: bool = False,
        use_preupload_v27_endpoint: bool = False,
        use_preupload_v28_endpoint: bool = False,
        manual_export: dict | None = None) -> dict:
    claim_verification = {
        "attempt_id": attempt.id, "workflow_key": WORKFLOW_KEY,
        "plan_id": PLAN_ID, "operation": OPERATION,
        "scope_sha256": manifest_sha,
        "inspect_scope_sha256": inspect_scope_sha,
        "reservation_token_sha256": inspection_detail[
            "reservation_token_sha256"],
    }
    if commit_phase == "resume_preupload_commit":
        claim_verification["resume_claim_sha256"] = resume_claim_sha256
    try:
        call = web_agent_service.recover_plan8_final_v8
        if commit_phase == "resume_preupload_commit":
            call = web_agent_service.recover_plan8_final_v8_preupload_resume
            if use_preupload_v9_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v9
            if use_preupload_v10_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v10
            if use_preupload_v12_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v12
            if use_preupload_v14_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v14
            if use_preupload_v15_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v15
            if use_preupload_v16_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v16
            if use_preupload_v17_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v17
            if use_preupload_v18_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v18
            if use_preupload_v19_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v19
            if use_preupload_v20_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v20
            if use_preupload_v21_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v21
            if use_preupload_v22_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v22
            if use_preupload_v23_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v23
            if use_preupload_v24_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v24
            if use_preupload_v25_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v25
            if use_preupload_v26_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v26
            if use_preupload_v27_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v27
            if use_preupload_v28_endpoint:
                call = web_agent_service.recover_plan8_final_v8_preupload_resume_v28
        payload = {"phase": ("commit" if commit_phase
                              == "resume_preupload_commit"
                              else commit_phase),
                   "scope_sha256": manifest_sha,
                   "inspect_scope_sha256": inspect_scope_sha,
                   "manifest": manifest, "attempt_id": attempt.id,
                   "reservation_token": reservation_token,
                   "claim_verification": claim_verification}
        if (use_preupload_v25_endpoint or use_preupload_v26_endpoint
                or use_preupload_v27_endpoint or use_preupload_v28_endpoint):
            payload["manual_export"] = manual_export
        committed = call(
            db, payload=payload)
    except Exception as exc:  # noqa: BLE001
        committed = {"ok": False, "error": type(exc).__name__,
                     "platform_write": None}
    commit_ok, commit_detail = validate_commit(
        committed, manifest, manifest_sha)
    if not commit_ok:
        plan.status = "alarmed"
        db.commit()
        campaign_execution_service.record_platform_terminal(
            db, attempt,
            state="unknown_no_retry" if committed.get("platform_write") is None
            else "failed_no_retry",
            platform_write_observed=committed.get("platform_write"),
            step=str(committed.get("last_checkpoint")
                     or committed.get("step") or "plan8_final_v8_commit"),
            error_code=str(committed.get("error") or "commit_failed"),
            job_id=str(committed.get("web_agent_job_id") or "") or None,
            result_summary={**dict(attempt.result_summary or {}),
                            "manifest": manifest,
                            "inspection": inspection_detail,
                            "commit": commit_detail})
        return _fail("plan8_final_v8_commit_failed_no_retry",
                     attempt_id=attempt.id, commit=commit_detail)
    try:
        readback = web_agent_service.recover_plan8_final_v8(
            db, payload={"phase": "readback", "scope_sha256": manifest_sha,
                         "manifest": manifest, "attempt_id": attempt.id})
    except Exception as exc:  # noqa: BLE001
        readback = {"ok": False, "error": type(exc).__name__,
                    "platform_write": False}
    readback_ok, readback_detail = validate_readback(
        readback, manifest, manifest_sha)
    if not readback_ok:
        plan.status = "alarmed"
        db.commit()
        campaign_execution_service.record_platform_terminal(
            db, attempt, state="failed_no_retry", platform_write_observed=True,
            step="plan8_final_v8_readback",
            error_code="post_submit_readback_not_complete",
            job_id=str(readback.get("web_agent_job_id") or "") or None,
            result_summary={"manifest": manifest,
                            "inspection": inspection_detail,
                            "commit": committed,
                            "readback": readback_detail})
        return _fail("plan8_final_v8_readback_not_complete",
                     attempt_id=attempt.id, readback=readback_detail)
    campaign_execution_service.record_platform_terminal(
        db, attempt, state="completed", platform_write_observed=True,
        step="readback_verified",
        job_id=str(readback.get("web_agent_job_id") or "") or None,
        result_summary={"manifest": manifest,
                        "inspection": inspection_detail,
                        "commit": commit_detail,
                        "readback": readback_detail,
                        "finished_at": datetime.now(timezone.utc).isoformat()})
    plan.status = "reconciled"
    db.commit()
    return {"ok": True, "workflow_key": WORKFLOW_KEY, "plan_id": PLAN_ID,
            "plan_status": plan.status, "attempt_id": attempt.id,
            "scope_sha256": manifest_sha, "verification": readback_detail,
            "execution_boundary": _boundary(platform_write=True)}


def _resume_claimed_preupload(
        db: Session, *, plan: CampaignPlan,
        attempt: CampaignExecutionAttempt,
        accept_post_readback_state: bool = False,
        accept_lease_scope_state: bool = False,
        accept_lease_expiry_state: bool = False,
        accept_dialog_mismatch_state: bool = False,
        accept_campaign_guard_state: bool = False,
        accept_claim_verify_state: bool = False,
        accept_lazy_import_state: bool = False,
        accept_preupload_claim_allowlist_state: bool = False,
        accept_semantic_modal_state: bool = False,
        accept_editor_identity_state: bool = False,
        accept_nested_modal_state: bool = False,
        accept_moban_text_state: bool = False,
        accept_unpersisted_postupload_state: bool = False,
        accept_template_reject_readback_state: bool = False,
        accept_template_generation_state: bool = False,
        accept_v20_claim_verify_rejection_state: bool = False,
        accept_v21_template_obstruction_state: bool = False,
        accept_v22_claim_contract_correction_state: bool = False,
        accept_v23_export_failure_state: bool = False,
        accept_v24_title_mismatch_state: bool = False,
        accept_v26_editor_loading_state: bool = False,
        accept_v27_claim_rejection_state: bool = False,
        manual_export_generation: int = 0,
        manual_export: dict | None = None,
        manual_export_evidence: dict | None = None,
        wait_prewrite_busy: bool = False) -> dict:
    validator = _validate_claimed_preupload_attempt
    resume_claim_sha256 = CLAIMED_PREUPLOAD_CLAIM_SHA256
    if accept_post_readback_state:
        validator = _validate_claimed_preupload_after_readback_attempt
    if accept_lease_scope_state:
        validator = _validate_claimed_preupload_after_lease_scope_attempt
    if accept_lease_expiry_state:
        validator = _validate_claimed_preupload_after_lease_expiry_attempt
    if accept_dialog_mismatch_state:
        validator = _validate_claimed_preupload_after_dialog_mismatch_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V9_CLAIM_SHA256
    if accept_campaign_guard_state:
        validator = _validate_claimed_preupload_after_campaign_guard_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V9_CLAIM_SHA256
    if accept_claim_verify_state:
        validator = _validate_claimed_preupload_after_claim_verify_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V9_CLAIM_SHA256
    if accept_lazy_import_state:
        validator = _validate_claimed_preupload_after_lazy_import_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V11_CLAIM_SHA256
    if accept_preupload_claim_allowlist_state:
        validator = _validate_claimed_preupload_after_allowlist_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V11_CLAIM_SHA256
    if accept_semantic_modal_state:
        validator = _validate_claimed_preupload_after_semantic_modal_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V13_CLAIM_SHA256
    if accept_editor_identity_state:
        validator = _validate_claimed_preupload_after_editor_identity_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V13_CLAIM_SHA256
    if accept_nested_modal_state:
        validator = _validate_claimed_preupload_after_nested_modal_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V15_CLAIM_SHA256
    if accept_moban_text_state:
        validator = _validate_claimed_preupload_after_moban_text_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V16_CLAIM_SHA256
    if accept_unpersisted_postupload_state:
        validator = _validate_claimed_postupload_after_unpersisted_readback
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V17_CLAIM_SHA256
    if accept_template_reject_readback_state:
        validator = _validate_claimed_postupload_after_template_reject_readback
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V18_CLAIM_SHA256
    if accept_template_generation_state:
        validator = _validate_claimed_preupload_after_template_generation_attempt
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V19_CLAIM_SHA256
    if accept_v20_claim_verify_rejection_state:
        validator = _validate_claimed_preupload_after_v20_claim_verify_rejection
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V19_CLAIM_SHA256
    if accept_v21_template_obstruction_state:
        validator = _validate_claimed_preupload_after_v21_template_obstruction
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
    if accept_v22_claim_contract_correction_state:
        validator = _validate_claimed_preupload_after_v21_template_obstruction
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
    if accept_v23_export_failure_state:
        validator = _validate_claimed_preupload_after_v23_export_failure
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
    if accept_v24_title_mismatch_state:
        validator = _validate_claimed_preupload_after_v24_title_mismatch
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
    if accept_v26_editor_loading_state:
        validator = _validate_claimed_preupload_after_v26_editor_loading
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
    if accept_v27_claim_rejection_state:
        validator = _validate_claimed_preupload_after_v27_claim_rejection
        resume_claim_sha256 = CLAIMED_PREUPLOAD_V21_CLAIM_SHA256
    preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume
    if accept_dialog_mismatch_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v9
    if accept_campaign_guard_state or accept_claim_verify_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v10
    if accept_lazy_import_state or accept_preupload_claim_allowlist_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v12
    if accept_semantic_modal_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v14
    if accept_editor_identity_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v15
    if accept_nested_modal_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v16
    if accept_moban_text_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v17
    if accept_unpersisted_postupload_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v18
    if accept_template_reject_readback_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v19
    if accept_template_generation_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v20
    if accept_v20_claim_verify_rejection_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v21
    if accept_v21_template_obstruction_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v22
    if accept_v22_claim_contract_correction_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v23
    if accept_v23_export_failure_state:
        preupload_web_call = web_agent_service.recover_plan8_final_v8_preupload_resume_v24
    if accept_v24_title_mismatch_state:
        preupload_web_call = (
            web_agent_service.recover_plan8_final_v8_preupload_resume_v26
            if manual_export_generation == 26
            else web_agent_service.recover_plan8_final_v8_preupload_resume_v25)
    if accept_v26_editor_loading_state:
        preupload_web_call = (
            web_agent_service.recover_plan8_final_v8_preupload_resume_v27)
    if accept_v27_claim_rejection_state:
        preupload_web_call = (
            web_agent_service.recover_plan8_final_v8_preupload_resume_v28)
    resume_ok, resume_detail = validator(attempt)
    if not resume_ok:
        return _fail("plan8_final_v8_claimed_preupload_attempt_mismatch",
                     attempt=resume_detail)
    manifest = dict((attempt.result_summary or {})["manifest"])
    baseline = manifest.get("inspection_baseline") or {}
    policy_sha = str(campaign_policy_service.require_policy().get("_sha256") or "")
    identity_ok, identity = v6._identity_allowed(plan)
    target_rows, scope_error = v7._target_rows(db, plan, identity, policy_sha)
    discount_rows, discount_error = v7._discount_scope(db, plan)
    current_base = (_fixed_manifest(target_rows, discount_rows, policy_sha)
                    if not scope_error and not discount_error else None)
    if (not identity_ok or policy_sha != EXPECTED_POLICY_SHA256
            or scope_error or discount_error or not isinstance(current_base, dict)
            or v6._hash(current_base) != baseline.get("inspect_scope_sha256")):
        return _fail("plan8_final_v8_claimed_preupload_scope_changed",
                     identity=identity, policy_sha256=policy_sha,
                     signup_scope_error=scope_error,
                     discount_scope_error=discount_error)
    db.commit()

    busy_observations = 0
    busy_wait_started = time.monotonic()
    while True:
        inspect_payload = {"phase": "inspect",
                           "scope_sha256": CLAIMED_PREUPLOAD_SCOPE_SHA256,
                           "manifest": manifest, "attempt_id": attempt.id}
        if (accept_v24_title_mismatch_state
                or accept_v26_editor_loading_state
                or accept_v27_claim_rejection_state):
            inspect_payload["manual_export"] = manual_export
        inspection = preupload_web_call(db, payload=inspect_payload)
        exact_retryable_busy = bool(
            inspection.get("ok") is False
            and inspection.get("error") == "taobao_profile_busy"
            and inspection.get("step") == "preupload_resume_busy"
            and inspection.get("busy") is True
            and inspection.get("pre_write_busy") is True
            and inspection.get("retry_safe") is True
            and inspection.get("platform_write") is False)
        elapsed = time.monotonic() - busy_wait_started
        if (not wait_prewrite_busy or not exact_retryable_busy
                or elapsed >= PREUPLOAD_BUSY_WAIT_SECONDS):
            break
        busy_observations += 1
        time.sleep(min(PREUPLOAD_BUSY_POLL_SECONDS,
                       PREUPLOAD_BUSY_WAIT_SECONDS - elapsed))
    if busy_observations:
        inspection["prewrite_busy_wait"] = {
            "observations": busy_observations,
            "waited_seconds": round(time.monotonic() - busy_wait_started, 3),
            "bounded": True,
        }
    inspect_scope = inspection.get("inspect_scope")
    reservation_token = str(inspection.get("reservation_token") or "")
    try:
        lease_expires = float(inspection.get("lease_expires_at_epoch") or 0)
    except (TypeError, ValueError):
        lease_expires = 0
    inspect_scope_sha = (v6._hash(inspect_scope)
                         if isinstance(inspect_scope, dict) else "")
    inspection_ok = bool(
        inspection.get("ok") is True
        and inspection.get("platform_write") is False
        and inspection.get("claim_created") is True
        and inspection.get("resume_claim_sha256")
        == resume_claim_sha256
        and inspection.get("last_checkpoint") == CLAIMED_PREUPLOAD_LAST_STEP
        and inspection.get("inspect_scope_sha256") == inspect_scope_sha
        and isinstance(inspect_scope, dict) and reservation_token
        and lease_expires > datetime.now(timezone.utc).timestamp())
    if not inspection_ok:
        return _fail("plan8_final_v8_claimed_preupload_inspection_blocked",
                     inspection={key: value for key, value in inspection.items()
                                 if key != "reservation_token"})

    plan = db.execute(select(CampaignPlan).where(
        CampaignPlan.id == PLAN_ID,
        CampaignPlan.workflow_key == WORKFLOW_KEY,
    ).with_for_update()).scalar_one_or_none()
    attempt = db.execute(select(CampaignExecutionAttempt).where(
        CampaignExecutionAttempt.id == PRECLAIM_ATTEMPT_ID,
    ).with_for_update()).scalar_one_or_none()
    resume_ok, resume_detail = validator(attempt)
    if (plan is None or plan.status != EXPECTED_STATUS or not resume_ok):
        return _fail("plan8_final_v8_claimed_preupload_state_changed",
                     plan_status=getattr(plan, "status", None),
                     attempt=resume_detail)
    summary = dict(attempt.result_summary or {})
    summary["claimed_preupload_resume"] = {
        "source_claim_sha256": resume_claim_sha256,
        "inspect_scope_sha256": inspect_scope_sha,
        "reservation_token_sha256": v6._hash(reservation_token),
        "reservation_expires_at_epoch": lease_expires,
    }
    if (accept_v24_title_mismatch_state
            or accept_v26_editor_loading_state
            or accept_v27_claim_rejection_state):
        summary[f"manual_export_v{manual_export_generation or 25}"] = dict(
            manual_export_evidence or {})
    attempt.state = "write_claimed"
    attempt.write_claimed = True
    attempt.write_claimed_at = datetime.now(timezone.utc)
    attempt.platform_write_observed = False
    attempt.automatic_retry_allowed = False
    resume_version = 4
    for enabled, version in (
            (accept_post_readback_state, 5),
            (accept_lease_scope_state, 6),
            (wait_prewrite_busy, 7),
            (accept_lease_expiry_state, 8),
            (accept_dialog_mismatch_state, 9),
            (accept_campaign_guard_state, 10),
            (accept_claim_verify_state, 11),
            (accept_lazy_import_state, 12),
            (accept_preupload_claim_allowlist_state, 13),
            (accept_semantic_modal_state, 14),
            (accept_editor_identity_state, 15),
            (accept_nested_modal_state, 16),
            (accept_moban_text_state, 17),
            (accept_unpersisted_postupload_state, 18),
            (accept_template_reject_readback_state, 19),
            (accept_template_generation_state, 20),
            (accept_v20_claim_verify_rejection_state, 21),
            (accept_v21_template_obstruction_state, 22),
            (accept_v22_claim_contract_correction_state, 23),
            (accept_v23_export_failure_state, 24),
            (accept_v24_title_mismatch_state,
             manual_export_generation or 25),
            (accept_v26_editor_loading_state, 27),
            (accept_v27_claim_rejection_state, 28)):
        if enabled:
            resume_version = version
    attempt.last_step = (
        f"platform_write_claim_claimed_preupload_resume_v{resume_version}")
    attempt.error_code = None
    attempt.web_agent_job_id = None
    attempt.result_summary = summary
    plan.status = "resume_executing"
    db.commit()
    inspection_detail = {
        "resume_claim_sha256": resume_claim_sha256,
        "inspect_scope_sha256": inspect_scope_sha,
        "reservation_token_sha256": v6._hash(reservation_token),
        "lease_expires_at_epoch": lease_expires,
        "web_agent_job_id": inspection.get("web_agent_job_id"),
    }
    return _commit_and_readback(
        db, plan=plan, attempt=attempt, manifest=manifest,
        manifest_sha=CLAIMED_PREUPLOAD_SCOPE_SHA256,
        inspect_scope_sha=inspect_scope_sha,
        reservation_token=reservation_token,
        inspection_detail=inspection_detail,
        commit_phase="resume_preupload_commit",
        resume_claim_sha256=resume_claim_sha256,
        use_preupload_v9_endpoint=accept_dialog_mismatch_state,
        use_preupload_v10_endpoint=(
            accept_campaign_guard_state or accept_claim_verify_state),
        use_preupload_v12_endpoint=(
            accept_lazy_import_state
            or accept_preupload_claim_allowlist_state),
        use_preupload_v14_endpoint=accept_semantic_modal_state,
        use_preupload_v15_endpoint=accept_editor_identity_state,
        use_preupload_v16_endpoint=accept_nested_modal_state,
        use_preupload_v17_endpoint=accept_moban_text_state,
        use_preupload_v18_endpoint=accept_unpersisted_postupload_state,
        use_preupload_v19_endpoint=accept_template_reject_readback_state,
        use_preupload_v20_endpoint=accept_template_generation_state,
        use_preupload_v21_endpoint=accept_v20_claim_verify_rejection_state,
        use_preupload_v22_endpoint=accept_v21_template_obstruction_state,
        use_preupload_v23_endpoint=accept_v22_claim_contract_correction_state,
        use_preupload_v24_endpoint=accept_v23_export_failure_state,
        use_preupload_v25_endpoint=(
            accept_v24_title_mismatch_state
            and manual_export_generation != 26),
        use_preupload_v26_endpoint=(
            accept_v24_title_mismatch_state
            and manual_export_generation == 26),
        use_preupload_v27_endpoint=accept_v26_editor_loading_state,
        use_preupload_v28_endpoint=accept_v27_claim_rejection_state,
        manual_export=manual_export)


def recover_plan8_final_v8(
        db: Session, *, workflow_key: str, expected_plan_id: int,
        expected_status: str, recovery_version: int,
        mode: str = "execute", confirmation: str = "",
        target_scope_sha256: str = "", manual_export_filename: str = "",
        manual_export_size: int = 0, manual_export_sha256: str = "",
        manual_export_base64: str = "") -> dict:
    confirmations = {
        "execute": EXECUTE_CONFIRMATION,
        "readback": READBACK_CONFIRMATION,
        "resume_preclaim_v3": PRECLAIM_RESUME_CONFIRMATION,
        "resume_claimed_preupload_v4": CLAIMED_PREUPLOAD_RESUME_CONFIRMATION,
        "resume_claimed_preupload_v5": (
            CLAIMED_PREUPLOAD_POST_READBACK_CONFIRMATION),
        "resume_claimed_preupload_v6": (
            CLAIMED_PREUPLOAD_LEASE_SCOPE_CONFIRMATION),
        "resume_claimed_preupload_v7": (
            CLAIMED_PREUPLOAD_BUSY_WAIT_CONFIRMATION),
        "resume_claimed_preupload_v8": (
            CLAIMED_PREUPLOAD_LEASE_EXPIRY_CONFIRMATION),
        "resume_claimed_preupload_v9": (
            CLAIMED_PREUPLOAD_DIALOG_FIX_CONFIRMATION),
        "resume_claimed_preupload_v10": (
            CLAIMED_PREUPLOAD_CAMPAIGN_GUARD_CONFIRMATION),
        "resume_claimed_preupload_v11": (
            CLAIMED_PREUPLOAD_CLAIM_VERIFY_CONFIRMATION),
        "resume_claimed_preupload_v12": (
            CLAIMED_PREUPLOAD_LAZY_IMPORT_CONFIRMATION),
        "resume_claimed_preupload_v13": (
            CLAIMED_PREUPLOAD_ALLOWLIST_CONFIRMATION),
        "resume_claimed_preupload_v14": (
            CLAIMED_PREUPLOAD_SEMANTIC_MODAL_CONFIRMATION),
        "resume_claimed_preupload_v15": (
            CLAIMED_PREUPLOAD_EDITOR_IDENTITY_CONFIRMATION),
        "resume_claimed_preupload_v16": (
            CLAIMED_PREUPLOAD_NESTED_MODAL_CONFIRMATION),
        "resume_claimed_preupload_v17": (
            CLAIMED_PREUPLOAD_MOBAN_TEXT_CONFIRMATION),
        "resume_claimed_preupload_v18": (
            CLAIMED_POSTUPLOAD_READBACK_CONFIRMATION),
        "resume_claimed_preupload_v19": (
            CLAIMED_OFFICIAL_TEMPLATE_CONFIRMATION),
        "resume_claimed_preupload_v20": (
            CLAIMED_TEMPLATE_GENERATION_CONFIRMATION),
        "resume_claimed_preupload_v21": (
            CLAIMED_TEMPLATE_CLAIM_VERIFY_CONFIRMATION),
        "resume_claimed_preupload_v22": (
            CLAIMED_TEMPLATE_CLOSE_CONFIRMATION),
        "resume_claimed_preupload_v23": (
            CLAIMED_TEMPLATE_CONTRACT_CONFIRMATION),
        "resume_claimed_preupload_v24": (
            CLAIMED_EXPORT_RETRY_CONFIRMATION),
        "resume_claimed_preupload_v25": (
            CLAIMED_MANUAL_EXPORT_CONFIRMATION),
        "resume_claimed_preupload_v26": (
            CLAIMED_MANUAL_EXPORT_V26_CONFIRMATION),
        "resume_claimed_preupload_v27": (
            CLAIMED_MANUAL_EXPORT_V27_CONFIRMATION),
        "resume_claimed_preupload_v28": (
            CLAIMED_MANUAL_EXPORT_V28_CONFIRMATION),
    }
    if (workflow_key != WORKFLOW_KEY or expected_plan_id != PLAN_ID
            or expected_status != EXPECTED_STATUS
            or recovery_version != RECOVERY_VERSION
            or mode not in confirmations
            or confirmation != confirmations.get(mode)
            or target_scope_sha256 != EXPECTED_TARGET_SCOPE_SHA256):
        return _fail("plan8_final_v8_request_not_allowed")
    plan = db.execute(select(CampaignPlan).where(
        CampaignPlan.id == PLAN_ID,
        CampaignPlan.workflow_key == WORKFLOW_KEY,
    ).with_for_update()).scalar_one_or_none()
    if plan is None:
        return _fail("workflow_not_found")
    identity_ok, identity = v6._identity_allowed(plan)
    if not identity_ok:
        return _fail("plan8_final_v8_identity_not_allowed", identity=identity)
    attempts = _attempts(db)
    if mode == "resume_claimed_preupload_v25":
        return _fail("plan8_final_v8_manual_export_v25_retired")
    if mode == "resume_claimed_preupload_v26":
        return _fail("plan8_final_v8_manual_export_v26_retired")
    if mode == "resume_claimed_preupload_v27":
        return _fail("plan8_final_v8_manual_export_v27_retired")
    manual_export = None
    manual_export_evidence = None
    if mode == "resume_claimed_preupload_v28":
        attempt = db.get(CampaignExecutionAttempt, PRECLAIM_ATTEMPT_ID)
        manifest = dict(((attempt.result_summary or {}).get("manifest")
                         if attempt is not None else {}) or {})
        manual_export, manual_export_evidence = _manual_export_v25_evidence(
            manifest=manifest, filename=manual_export_filename,
            size=manual_export_size, sha256=manual_export_sha256,
            xlsx_b64=manual_export_base64)
        if manual_export is None:
            return _fail("plan8_final_v8_manual_export_invalid",
                         manual_export=manual_export_evidence)
    elif any((manual_export_filename, manual_export_size,
              manual_export_sha256, manual_export_base64)):
        return _fail("plan8_final_v8_manual_export_forbidden_for_mode")

    if mode in {"resume_claimed_preupload_v4",
                "resume_claimed_preupload_v5",
                "resume_claimed_preupload_v6",
                "resume_claimed_preupload_v7",
                "resume_claimed_preupload_v8",
                "resume_claimed_preupload_v9",
                "resume_claimed_preupload_v10",
                "resume_claimed_preupload_v11",
                "resume_claimed_preupload_v12",
                "resume_claimed_preupload_v13",
                "resume_claimed_preupload_v14",
                "resume_claimed_preupload_v15",
                "resume_claimed_preupload_v16",
                "resume_claimed_preupload_v17",
                "resume_claimed_preupload_v18",
                "resume_claimed_preupload_v19",
                "resume_claimed_preupload_v20",
                "resume_claimed_preupload_v21",
                "resume_claimed_preupload_v22",
                "resume_claimed_preupload_v23",
                "resume_claimed_preupload_v24",
                "resume_claimed_preupload_v25",
                "resume_claimed_preupload_v26",
                "resume_claimed_preupload_v27",
                "resume_claimed_preupload_v28"}:
        if len(attempts) != 1:
            return _fail("plan8_final_v8_claimed_preupload_attempt_ambiguous",
                         attempt_count=len(attempts))
        return _resume_claimed_preupload(
            db, plan=plan, attempt=attempts[0],
            accept_post_readback_state=(
                mode == "resume_claimed_preupload_v5"),
            accept_lease_scope_state=(
                mode in {"resume_claimed_preupload_v6",
                         "resume_claimed_preupload_v7"}),
            accept_lease_expiry_state=(
                mode == "resume_claimed_preupload_v8"),
            accept_dialog_mismatch_state=(
                mode == "resume_claimed_preupload_v9"),
            accept_campaign_guard_state=(
                mode == "resume_claimed_preupload_v10"),
            accept_claim_verify_state=(
                mode == "resume_claimed_preupload_v11"),
            accept_lazy_import_state=(
                mode == "resume_claimed_preupload_v12"),
            accept_preupload_claim_allowlist_state=(
                mode == "resume_claimed_preupload_v13"),
            accept_semantic_modal_state=(
                mode == "resume_claimed_preupload_v14"),
            accept_editor_identity_state=(
                mode == "resume_claimed_preupload_v15"),
            accept_nested_modal_state=(
                mode == "resume_claimed_preupload_v16"),
            accept_moban_text_state=(
                mode == "resume_claimed_preupload_v17"),
            accept_unpersisted_postupload_state=(
                mode == "resume_claimed_preupload_v18"),
            accept_template_reject_readback_state=(
                mode == "resume_claimed_preupload_v19"),
            accept_template_generation_state=(
                mode == "resume_claimed_preupload_v20"),
            accept_v20_claim_verify_rejection_state=(
                mode == "resume_claimed_preupload_v21"),
            accept_v21_template_obstruction_state=(
                mode == "resume_claimed_preupload_v22"),
            accept_v22_claim_contract_correction_state=(
                mode == "resume_claimed_preupload_v23"),
            accept_v23_export_failure_state=(
                mode == "resume_claimed_preupload_v24"),
            accept_v24_title_mismatch_state=(
                False),
            accept_v26_editor_loading_state=(
                mode == "resume_claimed_preupload_v27"),
            accept_v27_claim_rejection_state=(
                mode == "resume_claimed_preupload_v28"),
            manual_export_generation=(28 if mode
                                      == "resume_claimed_preupload_v28" else 0),
            manual_export=manual_export,
            manual_export_evidence=manual_export_evidence,
            wait_prewrite_busy=(
                mode in {"resume_claimed_preupload_v7",
                         "resume_claimed_preupload_v8",
                         "resume_claimed_preupload_v9",
                         "resume_claimed_preupload_v10",
                         "resume_claimed_preupload_v11",
                         "resume_claimed_preupload_v12",
                         "resume_claimed_preupload_v13",
                         "resume_claimed_preupload_v14",
                         "resume_claimed_preupload_v15",
                         "resume_claimed_preupload_v16",
                         "resume_claimed_preupload_v17",
                         "resume_claimed_preupload_v18",
                         "resume_claimed_preupload_v19",
                         "resume_claimed_preupload_v20",
                         "resume_claimed_preupload_v21",
                         "resume_claimed_preupload_v22",
                         "resume_claimed_preupload_v23",
                         "resume_claimed_preupload_v24",
                         "resume_claimed_preupload_v25",
                         "resume_claimed_preupload_v26",
                         "resume_claimed_preupload_v27",
                         "resume_claimed_preupload_v28"}))
    if mode == "readback":
        if len(attempts) != 1:
            return _fail("plan8_final_v8_readback_attempt_ambiguous",
                         attempt_count=len(attempts))
        existing = attempts[0]
        if not existing.write_claimed:
            return _fail("plan8_final_v8_readback_attempt_not_found")
        return _readback_existing(db, plan, existing)
    is_preclaim_resume = mode == "resume_preclaim_v3"
    if is_preclaim_resume:
        if len(attempts) != 1:
            return _fail("plan8_final_v8_preclaim_attempt_ambiguous",
                         attempt_count=len(attempts))
        resume_ok, resume_detail = _validate_preclaim_resume_attempt(attempts[0])
        if not resume_ok:
            return _fail("plan8_final_v8_preclaim_attempt_mismatch",
                         attempt=resume_detail)
    elif attempts:
        if len(attempts) != 1:
            return _fail("plan8_final_v8_attempt_scope_ambiguous",
                         attempt_count=len(attempts))
        existing = attempts[0]
        if existing.state == "completed":
            manifest = (existing.result_summary or {}).get("manifest")
            if (not isinstance(manifest, dict)
                    or v6._hash(manifest) != existing.scope_sha256):
                return _fail("plan8_final_v8_attempt_scope_mismatch",
                             attempt_id=existing.id)
            return {"ok": True, "idempotent_replay": True,
                    "attempt_id": existing.id, "workflow_key": WORKFLOW_KEY,
                    "plan_id": PLAN_ID, "plan_status": plan.status,
                    "result": existing.result_summary or {},
                    "execution_boundary": _boundary(platform_write=False)}
        return _fail("plan8_final_v8_already_claimed_no_retry",
                     attempt_id=existing.id, attempt_state=existing.state,
                     platform_write_observed=existing.platform_write_observed)
    if plan.status != EXPECTED_STATUS:
        return _fail("plan8_final_v8_status_cas_mismatch",
                     actual_status=plan.status)
    prerequisite_ok, prerequisite = _validate_prerequisite(db)
    if not prerequisite_ok:
        return _fail("plan8_final_v8_prerequisite_attempt_mismatch",
                     attempt=prerequisite)
    policy_sha = str(campaign_policy_service.require_policy().get("_sha256") or "")
    if policy_sha != EXPECTED_POLICY_SHA256:
        return _fail("plan8_final_v8_policy_changed",
                     actual_policy_sha256=policy_sha)
    target_rows, scope_error = v7._target_rows(db, plan, identity, policy_sha)
    if scope_error:
        return _fail(**scope_error)
    discount_rows, discount_error = v7._discount_scope(db, plan)
    if discount_error:
        return _fail(**discount_error)
    manifest = _fixed_manifest(target_rows, discount_rows, policy_sha)
    inspect_scope_sha = v6._hash(manifest)
    db.commit()

    inspection = web_agent_service.recover_plan8_final_v8(
        db, payload={"phase": "inspect", "scope_sha256": inspect_scope_sha,
                     "manifest": manifest})
    if inspection.get("busy") or inspection.get("pre_write_busy"):
        return _fail("plan8_final_v8_pre_write_busy",
                     busy=inspection, write_claim_created=False)
    inspection_ok, inspection_detail = validate_inspection(
        inspection, manifest, inspect_scope_sha)
    if not inspection_ok:
        return _fail("plan8_final_v8_inspection_blocked",
                     inspection=inspection_detail,
                     need_scan=bool(inspection.get("need_scan")))
    if is_preclaim_resume and inspection_detail.get("v8_claim_absent") is not True:
        return _fail("plan8_final_v8_preclaim_resume_not_proven_safe",
                     inspection=inspection_detail,
                     write_claim_created=False)
    reservation_token = str(inspection["reservation_token"])
    manifest = v6.enrich_manifest_with_inspection(
        manifest, inspection_detail,
        inspect_scope_sha256=inspect_scope_sha)
    manifest_sha = v6._hash(manifest)

    plan = db.execute(select(CampaignPlan).where(
        CampaignPlan.id == PLAN_ID,
        CampaignPlan.workflow_key == WORKFLOW_KEY,
    ).with_for_update()).scalar_one_or_none()
    if plan is None or plan.status != EXPECTED_STATUS:
        return _fail("plan8_final_v8_state_changed_after_reservation",
                     actual_status=getattr(plan, "status", None))
    post_identity_ok, post_identity = v6._identity_allowed(plan)
    post_policy_sha = str(
        campaign_policy_service.require_policy().get("_sha256") or "")
    post_rows, post_scope_error = v7._target_rows(
        db, plan, post_identity, post_policy_sha)
    post_discounts, post_discount_error = v7._discount_scope(db, plan)
    if (not post_identity_ok or post_policy_sha != policy_sha
            or post_scope_error or post_discount_error
            or v6._hash(_fixed_manifest(
                post_rows, post_discounts, post_policy_sha)) != inspect_scope_sha):
        return _fail("plan8_final_v8_erp_scope_changed_after_reservation",
                     identity=post_identity, policy_sha256=post_policy_sha,
                     signup_scope_error=post_scope_error,
                     discount_scope_error=post_discount_error)
    raced = _attempts(db)
    if is_preclaim_resume:
        if len(raced) != 1:
            return _fail("plan8_final_v8_preclaim_attempt_raced",
                         attempt_count=len(raced))
        attempt = db.execute(select(CampaignExecutionAttempt).where(
            CampaignExecutionAttempt.id == PRECLAIM_ATTEMPT_ID,
        ).with_for_update()).scalar_one_or_none()
        resume_ok, resume_detail = _validate_preclaim_resume_attempt(attempt)
        if not resume_ok:
            return _fail("plan8_final_v8_preclaim_attempt_changed",
                         attempt=resume_detail)
        prior = dict(attempt.result_summary or {})
        attempt.scope_sha256 = manifest_sha
        attempt.state = "write_claimed"
        attempt.write_claimed = True
        attempt.write_claimed_at = datetime.now(timezone.utc)
        attempt.platform_write_observed = None
        attempt.automatic_retry_allowed = False
        attempt.last_step = "platform_write_claim_preclaim_resume_v3"
        attempt.error_code = None
        attempt.web_agent_job_id = None
        attempt.result_summary = {
            "manifest": manifest,
            "inspection": inspection_detail,
            "preclaim_resume_source": {
                "attempt_id": PRECLAIM_ATTEMPT_ID,
                "scope_sha256": PRECLAIM_SCOPE_SHA256,
                "request_id": PRECLAIM_REQUEST_ID,
                "web_agent_job_id": PRECLAIM_WEB_AGENT_JOB_ID,
                "prior_last_step": PRECLAIM_LAST_STEP,
                "prior_error_code": PRECLAIM_ERROR_CODE,
                "prior_commit": prior.get("commit"),
            },
        }
    else:
        if raced:
            exact = _attempt_for_scope(db, manifest_sha)
            return _fail("plan8_final_v8_attempt_raced_no_write",
                         attempt_count=len(raced),
                         exact_scope_exists=exact is not None)
        attempt = CampaignExecutionAttempt(
            id=secrets.token_hex(12), plan_id=PLAN_ID,
            workflow_key=WORKFLOW_KEY,
            operation=OPERATION, scope_sha256=manifest_sha,
            state="write_claimed", write_claimed=True,
            write_claimed_at=datetime.now(timezone.utc),
            platform_write_observed=None, automatic_retry_allowed=False,
            request_id=f"plan8-final-v8-{secrets.token_hex(8)}",
            last_step="platform_write_claim",
            result_summary={"manifest": manifest,
                            "inspection": inspection_detail})
        db.add(attempt)
    plan.status = "resume_executing"
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return _fail("plan8_final_v8_atomic_claim_conflict_no_write")
    return _commit_and_readback(
        db, plan=plan, attempt=attempt, manifest=manifest,
        manifest_sha=manifest_sha, inspect_scope_sha=inspect_scope_sha,
        reservation_token=reservation_token,
        inspection_detail=inspection_detail)

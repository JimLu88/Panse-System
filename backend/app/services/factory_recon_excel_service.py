"""工厂对账 Excel → AI 整理 → 标准对账行.

工厂发来的对账表列名/格式不规整, 用 AI 把每行映射成系统对账字段:
factory_name, period_start, period_end, order_amount, bill_amount, paid_amount, alipay_flow_no, remark
"""
from __future__ import annotations

import io
import json
import logging
import re
from typing import Optional

from sqlalchemy.orm import Session

from app.services import settings_service
from app.services.ai_provider import AiUnavailable, build_provider

_logger = logging.getLogger("panse.factory_recon_excel")

_MAX_DATA_ROWS = 200  # 限制喂给 AI 的数据行数, 控制 token

_FACTORY_RECON_EXCEL_SYSTEM = """你是工厂对账 Excel 整理助手。
工厂发来的对账表列名/格式很乱 (列名不规范、合并单元格、金额带¥、日期格式杂乱等)。
请把每一条对账数据行映射成系统标准字段, 输出严格 JSON:
{
  "rows": [
    {
      "factory_name": "工厂名称 (必填)",
      "period_start": "YYYY-MM-DD (账期起, 可空)",
      "period_end": "YYYY-MM-DD (账期止, 可空)",
      "order_amount": 数字 (本期下单金额, 可空),
      "bill_amount": 数字 (工厂账单金额, 可空),
      "paid_amount": 数字 (实际已支付, 可空),
      "alipay_flow_no": "支付宝流水号 (如有, 可空)",
      "remark": "备注 (可空)"
    }
  ],
  "warnings": ["整理过程中的问题或不确定项"]
}

规则:
- 数字字段返回纯数字 (去掉 ¥/元/逗号), 不确定就 null
- 日期统一 YYYY-MM-DD, 无法解析填 null
- factory_name 必填; 若表格用合并单元格导致部分行工厂名为空, 请向下沿用上一行的工厂名 (forward-fill)
- 跳过表头行、小计/合计行、空行
- 不要编造数据, 缺失的字段填 null
- 仅输出 JSON, 不要任何解释文字"""


_FENCE_RE = re.compile(r"^```(?:json)?\s*|\s*```$", re.MULTILINE)


def _extract_json(text: str) -> dict:
    cleaned = _FENCE_RE.sub("", text or "").strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start, end = cleaned.find("{"), cleaned.rfind("}")
        if start < 0 or end <= start:
            raise ValueError(f"AI 返回不是 JSON: {text[:300]}")
        return json.loads(cleaned[start: end + 1])


def _cell(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _read_sheet_rows(file_bytes: bytes) -> tuple[list[list[str]], list[str]]:
    """读取第一个工作表, 返回 (所有行的字符串列表, warnings)."""
    import openpyxl

    warnings: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[list[str]] = []
        for r in ws.iter_rows(values_only=True):
            cells = [_cell(c) for c in r]
            if any(cells):
                rows.append(cells)
            if len(rows) > _MAX_DATA_ROWS + 1:  # +1 表头
                warnings.append(
                    f"表格超过 {_MAX_DATA_ROWS} 行, 仅整理前 {_MAX_DATA_ROWS} 行数据"
                )
                break
    finally:
        wb.close()
    return rows, warnings


def _rows_to_text(rows: list[list[str]]) -> str:
    """把行拼成 TSV 文本喂给 AI."""
    lines = []
    for r in rows:
        lines.append("\t".join(r))
    return "\n".join(lines)


def parse_factory_recon_excel(db: Session, file_bytes: bytes) -> dict:
    """工厂对账 Excel → AI 整理成标准对账行. 返回 {"rows": [...], "warnings": [...]}."""
    rows, warnings = _read_sheet_rows(file_bytes)
    if not rows:
        return {"rows": [], "warnings": ["Excel 为空或无可识别数据"]}

    cfg = settings_service.get_ai_config(db, "diagnose")
    try:
        provider = build_provider(cfg)
    except AiUnavailable as e:
        raise AiUnavailable(f"AI 未配置, 请到管理 → AI 集成 配置模型: {e}")

    table_text = _rows_to_text(rows)
    user = (
        "下面是工厂发来的对账表 (TSV, 第一行可能是表头, 也可能没有表头), "
        "请整理成标准对账行并输出 JSON:\n\n" + table_text
    )
    try:
        resp = provider.chat(
            system=_FACTORY_RECON_EXCEL_SYSTEM,
            user=user,
            max_tokens=8000,
        )
    except AiUnavailable as e:
        raise AiUnavailable(f"AI 调用失败: {e}")

    try:
        data = _extract_json(resp.text)
    except ValueError as e:
        raise AiUnavailable(f"AI 返回无法解析: {e}")

    out_rows = data.get("rows")
    if not isinstance(out_rows, list):
        out_rows = []
    ai_warnings = data.get("warnings")
    if not isinstance(ai_warnings, list):
        ai_warnings = []

    # forward-fill 兜底 (AI 偶尔漏掉)
    last_name = ""
    for row in out_rows:
        if not isinstance(row, dict):
            continue
        name = (row.get("factory_name") or "").strip()
        if name:
            last_name = name
        elif last_name:
            row["factory_name"] = last_name

    return {"rows": out_rows, "warnings": warnings + ai_warnings}

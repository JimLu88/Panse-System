"""淘宝订单多格式导入 + 自动识别 (Phase B 扩展)。

支持未来两种导出格式自动识别、解析、入库, 无需手动指定:

1. **千牛后台多表 Excel** (``订单报表`` / ``销售明细`` / ``发货报表`` 三个 Sheet)
   - 销售明细 = 行级主数据 (产品/SKU/数量/退款)
   - 订单报表 = 单级 (地址/物流/金额/状态)
   - 发货报表 = 单级 (收货人/电话)
   - 三表按订单号关联, 一张订单一行 Order (多商品单取主商品行 + 备注其余)

2. **销售明细 CSV / Excel** (子订单编号,主订单编号,标题,…,商家编码,…)
   - GBK 或 UTF-8 编码均可
   - 同样按订单号聚合成 Order

公共转换:
  - 商家编码 ``PPS`` + 13 位 (SKU 级) → 老格式产品编码 ``P`` + 11 位
  - 商品属性 ``颜色分类:xxx[规格];安装方式:yyy`` → 取 ``xxx`` 作 SKU 名
  - 订单号科学计数法 (历史 CSV 损坏) → 标记 needs_review, 不静默丢弃

入库口径与 order_import 一致: 订单号唯一, 重复跳过; 历史批量导入默认 is_historical=True。
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.order import Order, OrderDetail
from app.services import order_cost_service, taobao_listing_service

# ── 格式指纹 ──────────────────────────────────────────────────────────────────
# 千牛多表: Sheet 名命中其一即视为该格式
QIANNIU_SHEETS = {"订单报表", "销售明细", "发货报表"}
# 销售明细列指纹 (任一格式只要含这些列即可识别)
SALES_DETAIL_KEYS = {"子订单编号", "主订单编号", "商品属性"}

# 淘宝订单状态文本 → 内部 status
_STATUS_MAP = {
    "等待买家付款": "pending_payment",
    "买家已付款,等待卖家发货": "paid",
    "买家已付款，等待卖家发货": "paid",
    "卖家已发货,等待买家确认": "shipped",
    "卖家已发货，等待买家确认": "shipped",
    "交易成功": "signed",
    # 已收货/已结算/交易完成/待评价/已签收 = 真实成交完结 (用户实测 2026-06-18: 这些没被识别→错标待付款)
    "已收货": "signed", "买家已收货": "signed", "已签收": "signed", "交易完成": "signed",
    "已完成": "signed", "待评价": "signed", "已结算": "signed",
    "交易关闭": "cancelled",
    "退款成功": "aftersales",
}


def _resolve_status(raw: Any) -> tuple[str, bool]:
    """返回 (内部状态, 是否识别)。无法识别(空/陌生文字, 走最后默认)→ recognized=False,
    供导入拦截报异常 (用户拍板 2026-06-18: 状态无法识别的单先拦截, 异常清除后再入库)。"""
    if not raw:
        return "pending_payment", False
    s = str(raw).strip()
    if s in _STATUS_MAP:
        return _STATUS_MAP[s], True
    if "退款" in s:
        return "aftersales", True
    if "关闭" in s:
        return "cancelled", True
    # 成交完结类: 成功/收货/签收/完成/已结算/待评价 → 已签收 (扩展, 防"已收货"这类掉进待付款)
    if any(k in s for k in ("成功", "收货", "签收", "交易完成", "已完成", "已结算", "待评价")):
        return "signed", True
    if "发货" in s and "等待买家" in s:
        return "shipped", True
    if "付款" in s and "等待卖家" in s:
        return "paid", True
    if "等待买家付款" in s:
        return "pending_payment", True
    return "pending_payment", False


def _map_status(raw: Any) -> str:
    return _resolve_status(raw)[0]


_SERVICE_NAME_KW = ("送货", "入户", "安装", "上门")


def _is_service_line_name(name: Any) -> bool:
    """送货入户/商家安装/上门 等服务行 (多行订单里不该抢主商品名)。"""
    n = str(name or "")
    return any(k in n for k in _SERVICE_NAME_KW)


def _norm_pps_code(code: Any) -> Any:
    """历史 P+数字 编码统一成 PPS (用户拍板 2026-06-18: 以后全 PPS, 不再用 P 开头)。PFG/PPS 不动。"""
    if isinstance(code, str) and len(code) > 1 and code[0] == "P" and code[1].isdigit():
        return "PPS" + code[1:]
    return code


# ── 字段转换 ──────────────────────────────────────────────────────────────────
def extract_sku(attr: Any) -> str:
    """商品属性 '颜色分类:xxx[规格];安装方式:yyy' → 'xxx' (去[..]与;后内容)。"""
    if not attr:
        return ""
    s = str(attr)
    first = re.split(r"[;；]", s)[0]
    first = re.sub(r"^[^:：]*[:：]", "", first)   # 去 '颜色分类:' 前缀
    first = re.sub(r"\[[^\]]*\]", "", first)       # 去 [长45cm] 规格
    return first.strip()


def product_code_from_merchant(mc: Any) -> str:
    """商家编码 → 老格式产品编码 P+11位。

    - ``PPS`` + 13 位 (SKU 级) → ``P`` + 前 11 位 (产品级)
    - 纯 11+ 位数字 (历史编码) → ``P`` + 前 11 位
    - 其它 / 空 → 空串
    """
    if mc is None:
        return ""
    s = str(mc).strip()
    if not s:
        return ""
    if s.upper().startswith("PPS"):
        digits = s[3:]
        return "P" + digits[:11] if len(digits) >= 11 else "P" + digits
    if s.isdigit() and len(s) >= 11:
        return "P" + s[:11]
    return ""


def _normalize_refund(payable: Any, paid: Any, refund: Any) -> Any:
    """根治"优惠/取消产品被当退款双扣" (2026-06-22 用户拍板)。

    淘宝「退款金额」里混着买家优惠差额 / 多产品订单中取消的子产品金额, 而「买家实付」已是净额
    (= 应付 − 这部分)。若 ``应付 − 实付 ≈ 退款``, 说明这笔"退款"已经体现在实付里、不是真退款
    (真退款是买家付了全款后再退, 那时 应付 = 实付、应付−实付 ≠ 退款), 归 0 ——
    否则收入口径(实付 − 退款, asset/cash_flow/dashboard/sales/smart_pricing 6+ 处)会把同一笔再扣一遍 → 假亏。
    """
    if payable is not None and paid is not None and refund and refund > Decimal("0"):
        if abs((payable - paid) - refund) < Decimal("0.01"):
            return Decimal("0")
    return refund


def _is_sci(v: Any) -> bool:
    """检测科学计数法损坏的订单号 (含 E+ / e+)。"""
    return v is not None and re.search(r"\d[eE]\+?\d+", str(v)) is not None


def _to_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
                "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    part = s.split(" ")[0].replace("/", "-")
    try:
        y, m, d = part.split("-")
        return date(int(y), int(m), int(d))
    except ValueError:
        return None


def _to_int(v: Any, default: int = 1) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _to_decimal(v: Any) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    s = str(v).replace(",", "").replace("¥", "").replace("元", "").strip()
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ── 报告 ──────────────────────────────────────────────────────────────────────
@dataclass
class TaobaoImportReport:
    detected_format: str = "unknown"
    inserted: int = 0
    updated: int = 0                # 已存在订单被回填状态/金额 (再次导入更新)
    skipped_duplicate: int = 0
    skipped_invalid: int = 0
    needs_review: int = 0           # 订单号损坏等, 已入库但标注待核
    multi_line_orders: int = 0      # 一单多商品的订单数
    status_changed: int = 0         # 重导时状态被覆盖的单数 (日报用)
    amount_changed: int = 0         # 重导时实付/退款被覆盖的单数 (日报用)
    vanished: int = 0               # 库里有、新文件里没有的单数 (报异常, 不动行)
    held_unrecognized: int = 0      # 状态无法识别+无收款凭据 → 拦截未入库 (报异常待人工)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ── 统一记录结构 (聚合到订单级) ────────────────────────────────────────────────
@dataclass
class _OrderRow:
    order_no: str
    order_no_bad: bool = False
    order_date: Any = None
    customer_name: Any = None
    customer_phone: Any = None
    customer_address: Any = None
    carrier: Any = None
    tracking_no: Any = None
    paid_amount: Any = None          # 买家应付货款 (兜底, 老逻辑沿用)
    status_text: Any = None
    # 状态可信度 (2026-07-11): 内容级判定命中的"已卖出宝贝"式文件, 状态是逐商品行值(非订单级真值)
    # → False, 不许改已存在订单的状态; 其余(订单报表/真·销售明细)保持 True。
    status_trusted: bool = True
    # 财务列 (订单报表/销售明细自带, 现金流"待确认收货/未发货/平台费"全靠它们)
    buyer_payable: Any = None        # 买家应付货款
    paid_real: Any = None            # 买家实付金额 (买家真实支付 = 我方应收)
    shop_received: Any = None        # 打款商家金额 (店铺实收, 历史CSV有)
    platform_fee: Any = None         # 卖家服务费 (平台服务费)
    refund: Any = None               # 退款金额
    buyer_freight: Any = None        # 买家应付邮费 (买家额外付的运费=代收, 对账基准要加)
    ship_time: Any = None            # 发货时间
    confirm_time: Any = None         # 确认收货时间
    shop: Any = None                 # 店铺名称
    buyer_message: Any = None        # 买家留言 (平台, 重导覆盖)
    seller_memo: Any = None          # 卖家备注/商家备注 (平台, 重导覆盖)
    # 订单级财务权威度 (2026-07-09): "order"=单级权威源(订单报表/已卖出宝贝导出, 一单一行, 财务列完整);
    # "line"=行级销售明细(一行一商品, 订单级金额需按行求和, 不完整时会低估)。重导幂等护栏据此决定
    # 是否允许覆盖已有订单的订单级财务字段 —— 行级源不许覆盖(防不完整明细把订单报表的正确值压掉)。
    fin_source: str = "order"
    lines: list[dict] = field(default_factory=list)   # 每个商品行


# ── 格式识别 ──────────────────────────────────────────────────────────────────
def detect_format(filename: str, raw: bytes) -> str:
    """返回 'qianniu_multi' | 'sales_detail' | 'order_master' | 'unknown'。"""
    name = (filename or "").lower()
    is_xlsx = name.endswith((".xlsx", ".xlsm", ".xls")) or raw[:2] == b"PK"
    if is_xlsx:
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True)
            sheets = set(wb.sheetnames)
            wb.close()
        except Exception:
            sheets = set()
        if sheets & QIANNIU_SHEETS:
            return "qianniu_multi"
        # 单 Sheet xlsx: 看首行表头
        header = _xlsx_first_header(raw)
        if SALES_DETAIL_KEYS & set(header):
            return "sales_detail"
        # 千牛「订单报表」单表导出 (Web-Agent 自动取数, 2026-06-12): 主单级,
        # 列名 _parse_sales_detail 的 gv 兜底全覆盖 → 直接走 sales_detail 解析。
        if "订单编号" in header and "订单创建时间" in header and "订单状态" in header:
            return "sales_detail"
        if "订单编号" in header and ("买家应付金额" in header or "产品编码" in header):
            return "order_master"
        # 淘宝卖家中心「已卖出的宝贝」导出 (ExportOrderList, sheet 多名 'export'; Web-Agent 自动下载):
        # 列为 订单编号/订单状态/买家应付货款/收货人姓名/买家留言… 无"订单创建时间"。
        # _parse_sales_detail 的 gv 兜底全覆盖这些列名 → 走 sales_detail 解析 (2026-06-12)。
        if "订单编号" in header and "订单状态" in header:
            return "sales_detail"
        return "unknown"
    # CSV
    header = _csv_header(raw)
    if SALES_DETAIL_KEYS & set(header):
        return "sales_detail"
    if "订单编号" in header or "订单号" in header:
        return "order_master"
    return "unknown"


def _decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("gbk", errors="replace")


def _csv_header(raw: bytes) -> list[str]:
    text = _decode(raw)
    rdr = csv.reader(io.StringIO(text))
    for row in rdr:
        return [(c or "").strip() for c in row]
    return []


def _xlsx_first_header(raw: bytes) -> list[str]:
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            wb.close()
            return [str(c).strip() for c in row if c is not None]
        wb.close()
    except Exception:
        pass
    return []


# ── 解析: 千牛多表 Excel ──────────────────────────────────────────────────────
def _parse_qianniu_multi(raw: bytes, rep: TaobaoImportReport) -> dict[str, _OrderRow]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    orders: dict[str, _OrderRow] = {}

    def header_idx(ws):
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return {str(v).strip(): i for i, v in enumerate(row) if v is not None}
        return {}

    # 订单报表: 单级
    rpt: dict[str, dict] = {}
    if "订单报表" in wb.sheetnames:
        ws = wb["订单报表"]; h = header_idx(ws)
        def g(row, *names):
            for n in names:
                if n in h and h[n] < len(row):
                    return row[h[n]]
            return None
        for row in ws.iter_rows(min_row=2, values_only=True):
            no = _clean(g(row, "订单编号"))
            if not no:
                continue
            rpt[no] = {
                "addr": g(row, "收货地址"),
                "carrier": g(row, "物流公司", "快递公司"),
                "tracking": g(row, "物流单号", "运单号"),
                "buyer_due": g(row, "买家应付货款", "买家实付金额"),
                "status": g(row, "订单状态"),
                "create": g(row, "订单创建时间", "订单付款时间"),
                # 财务列 (订单报表自带): 应付/实付/实收/平台费/退款/发货确认时间/店铺
                "payable": g(row, "买家应付货款"),
                "paid_real": g(row, "买家实付金额", "买家实际支付金额"),
                "shop_received": g(row, "打款商家金额"),
                "platform_fee": g(row, "卖家服务费", "平台服务费"),
                "refund": g(row, "退款金额"),
                "buyer_freight": g(row, "买家应付邮费", "买家承担邮费", "邮费"),
                "ship_time": g(row, "发货时间"),
                "confirm_time": g(row, "确认收货时间"),
                "shop": g(row, "店铺名称"),
            }

    # 发货报表: 单级 客户信息
    ship: dict[str, dict] = {}
    if "发货报表" in wb.sheetnames:
        ws = wb["发货报表"]; h = header_idx(ws)
        def g2(row, *names):
            for n in names:
                if n in h and h[n] < len(row):
                    return row[h[n]]
            return None
        for row in ws.iter_rows(min_row=2, values_only=True):
            no = _clean(g2(row, "订单编号"))
            if not no:
                continue
            ship[no] = {
                "name": g2(row, "收货人姓名"),
                "phone": g2(row, "联系手机", "联系电话"),
                "addr": g2(row, "收货地址"),
            }

    # 销售明细: 行级主数据
    if "销售明细" in wb.sheetnames:
        ws = wb["销售明细"]; h = header_idx(ws)
        def g3(row, *names):
            for n in names:
                if n in h and h[n] < len(row):
                    return row[h[n]]
            return None
        for row in ws.iter_rows(min_row=2, values_only=True):
            main = g3(row, "主订单编号", "订单编号")
            if main is None or str(main).strip() == "":
                continue
            no = str(main).strip()
            r = rpt.get(no, {})
            s = ship.get(no, {})
            o = orders.get(no)
            if o is None:
                o = _OrderRow(
                    order_no=no, order_no_bad=_is_sci(main),
                    # 有订单报表数据(r 非空)→ 订单级财务权威; 仅销售明细里的单 → 行级(不完整可能低估)
                    fin_source=("order" if r else "line"),
                    order_date=r.get("create") or g3(row, "订单创建时间"),
                    customer_name=s.get("name"),
                    customer_phone=s.get("phone"),
                    customer_address=s.get("addr") or r.get("addr"),
                    carrier=r.get("carrier"), tracking_no=r.get("tracking"),
                    paid_amount=r.get("buyer_due") or g3(row, "买家应付货款"),
                    status_text=r.get("status") or g3(row, "订单状态"),
                    buyer_payable=r.get("payable") or g3(row, "买家应付货款"),
                    paid_real=r.get("paid_real") or g3(row, "买家实付金额"),
                    shop_received=r.get("shop_received"),
                    platform_fee=r.get("platform_fee"),
                    refund=r.get("refund") or g3(row, "退款金额"),
                    buyer_freight=r.get("buyer_freight") or g3(row, "买家应付邮费", "买家承担邮费", "邮费"),
                    ship_time=r.get("ship_time") or g3(row, "发货时间"),
                    confirm_time=r.get("confirm_time") or g3(row, "确认收货时间"),
                    shop=r.get("shop"),
                    buyer_message=g3(row, "买家留言", "买家留言备注", "买家备注"),
                    seller_memo=g3(row, "卖家备注", "商家备注", "卖家留言"),
                )
                orders[no] = o
            merchant = g3(row, "商家编码", "外部系统编号")
            o.lines.append({
                "product_name": g3(row, "商品标题", "标题"),
                "sku": extract_sku(g3(row, "商品属性")),
                "sku_code": _clean(merchant),
                "product_code": product_code_from_merchant(merchant),
                "sku_id": _clean(g3(row, "skuId", "sku id", "SKU ID", "sku_id")),
                "qty": g3(row, "购买数量", "数量"),
                "amount": _to_decimal(g3(row, "买家应付货款", "商品价格")),
            })
    wb.close()
    return orders, ship


# ── 解析: 销售明细 CSV / 单 Sheet xlsx ─────────────────────────────────────────
def _parse_sales_detail(filename: str, raw: bytes, rep: TaobaoImportReport) -> dict[str, _OrderRow]:
    name = (filename or "").lower()
    rows: list[dict] = []
    _hdr: set[str] = set()
    if name.endswith((".xlsx", ".xlsm", ".xls")) or raw[:2] == b"PK":
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it)]
        _hdr = set(header)
        for r in it:
            rows.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
        wb.close()
    else:
        text = _decode(raw)
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        _hdr = set(reader.fieldnames or [])
    # 行级销售明细判据 (2026-07-09): 表头含「子订单编号/主订单编号」= 一行一商品的销售明细(订单级金额需按行
    # 求和, 明细不完整时低估); 单级导出(订单报表/已卖出宝贝)只有「订单编号」。行级源不权威于订单级财务 →
    # 重导时不许覆盖已存在订单的实付/状态/退款(防不完整明细把订单报表的正确值压掉, 根治 202 单横跳)。
    _fin_src = "line" if (_hdr & {"子订单编号", "主订单编号"}) else "order"
    # 内容级判据 (2026-07-11, 治漏网): 「已卖出的宝贝」OLE加密导出 多商品订单一单多行、表头却没有
    # 子订单编号列 → 表头判据看不见, 被当"单级权威", 每晚18点档与订单报表互搏(8单实付/状态/退款
    # 反复横跳: 18:11 订单报表写对、18:16 已卖出宝贝盖错)。同一文件里订单编号重复≥2次 = 一行一商品
    # 的行级导出, 不论表头长什么样都按行级处理; 单级导出订单号恒不重复, 不受影响。
    _status_trusted = True
    if _fin_src == "order" and rows:
        from collections import Counter as _Counter
        _no_cnt = _Counter()
        for _r in rows:
            _no = str(_r.get("主订单编号") or _r.get("订单编号") or "").strip()
            if _no and _no != "None":
                _no_cnt[_no] += 1
        if _no_cnt and _no_cnt.most_common(1)[0][1] >= 2:
            _fin_src = "line"
            # 状态也不可信: 该类文件「订单状态」是逐商品行的(退款件行=交易关闭), 不是订单级真值 ——
            # 真·销售明细(子订单编号表头)的状态列是订单级值随行重复, 仍可信(7/9 语义不变)。
            _status_trusted = False
            rep.warnings.append(
                f"检测到同文件订单号重复(最多{_no_cnt.most_common(1)[0][1]}行/单)→ 一单多行的行级导出, "
                "按行级处理: 不覆盖已存在订单的实付/状态/退款")

    def gv(row, *names):
        for n in names:
            if n in row and row[n] not in (None, ""):
                return row[n]
        return None

    orders: dict[str, _OrderRow] = {}
    for row in rows:
        main = gv(row, "主订单编号", "订单编号", "子订单编号")
        if main is None or str(main).strip() == "":
            continue
        no = str(main).strip()
        o = orders.get(no)
        if o is None:
            o = _OrderRow(
                order_no=no, order_no_bad=_is_sci(main), fin_source=_fin_src,
                status_trusted=_status_trusted,
                order_date=gv(row, "订单创建时间", "订单付款时间", "下单时间"),
                customer_name=gv(row, "收货人姓名", "买家昵称"),
                customer_phone=gv(row, "联系手机", "联系电话", "手机"),
                customer_address=gv(row, "收货地址"),
                carrier=gv(row, "物流公司", "快递公司"),
                tracking_no=gv(row, "物流单号", "运单号"),
                paid_amount=gv(row, "买家应付货款", "买家实付金额", "实付金额"),
                status_text=gv(row, "订单状态"),
                buyer_payable=gv(row, "买家应付货款"),
                paid_real=gv(row, "买家实付金额", "买家实际支付金额", "实付金额"),
                shop_received=gv(row, "打款商家金额"),
                platform_fee=gv(row, "卖家服务费", "平台服务费"),
                refund=gv(row, "退款金额"),
                buyer_freight=gv(row, "买家应付邮费", "买家承担邮费", "邮费"),
                ship_time=gv(row, "发货时间"),
                confirm_time=gv(row, "确认收货时间"),
                shop=gv(row, "店铺名称"),
                buyer_message=gv(row, "买家留言", "买家留言备注", "买家备注"),
                seller_memo=gv(row, "卖家备注", "商家备注", "卖家留言", "常用备注"),
            )
            orders[no] = o
        merchant = gv(row, "商家编码", "外部系统编号")
        o.lines.append({
            "product_name": gv(row, "商品标题", "标题", "商品名称"),
            "sku": extract_sku(gv(row, "商品属性")),
            "sku_code": _clean(merchant),
            "product_code": product_code_from_merchant(merchant),
            "sku_id": _clean(gv(row, "skuId", "sku id", "SKU ID", "sku_id")),
            "qty": gv(row, "购买数量", "数量"),
            "amount": _to_decimal(gv(row, "买家应付货款", "价格", "商品价格")),
            # 乙(用户拍板 2026-06-23): 每个子订单行的财务列都留下, 供 _commit_orders 按主订单求和
            # (此前订单级金额只取第一行 → 多产品单漏抓其余产品的实付/实收/应付)。
            "paid_real": _to_decimal(gv(row, "买家实付金额", "买家实际支付金额")),
            "shop_received": _to_decimal(gv(row, "打款商家金额")),
            "platform_fee": _to_decimal(gv(row, "卖家服务费", "平台服务费")),
            "refund": _to_decimal(gv(row, "退款金额")),
            # 邮费是订单级(一单一次), 多产品行可能重复或只落某行 → _commit_orders 取 max 不求和
            "buyer_freight": _to_decimal(gv(row, "买家应付邮费", "买家承担邮费", "邮费")),
        })
    return orders


def _persist_order_lines(db: Session, order_no: str, lines: list, resolver) -> None:
    """一单多宝贝 → 把每个商品行 upsert 到 order_details(source='import'), 供成本按行汇总(杜绝塌单漏算)。

    sku_code 经对应表 resolve 成 PPS 编码(否则匹配不到定价/成本); 服务行(送货/安装)不写; 幂等(按 sync_key)。
    """
    for idx, ln in enumerate(lines):
        if _is_service_line_name(ln.get("product_name")):
            continue
        scode = _norm_pps_code(ln.get("sku_code"))
        pcode = _norm_pps_code(ln.get("product_code") or None)
        hit = taobao_listing_service.resolve_line(
            resolver, sku_id=ln.get("sku_id"), merchant_code=ln.get("sku_code"))
        if hit:
            scode = hit.get("sku_code") or scode
            pcode = hit.get("product_code") or pcode
        sync_key = f"line:{order_no}:{idx}"
        row = db.execute(select(OrderDetail).where(OrderDetail.sync_key == sync_key)).scalar_one_or_none()
        vals = dict(
            order_no=order_no, product_code=pcode, sku_code=scode,
            product_name=_clean(ln.get("product_name")),
            qty=_to_int(ln.get("qty"), default=1),
            amount=_to_decimal(ln.get("amount")), source="import")
        if row:
            for k, v in vals.items():
                setattr(row, k, v)
        else:
            db.add(OrderDetail(sync_key=sync_key, **vals))


# ── 聚合订单行 → Order 入库 ───────────────────────────────────────────────────
def _commit_orders(db: Session, orders: dict[str, _OrderRow], platform: str,
                   rep: TaobaoImportReport) -> None:
    resolver = taobao_listing_service.build_resolver(db)  # 对应表: skuId/编码 → SKU编码/产品编码/店铺
    # 人工锁 (用户 2026-07-13 "直接处理"): 人裁定过 财务/状态 的订单(修改档案 source≠import),
    # 重导永不覆盖这些列 —— flip9 复位后钉住, 两个导出源再互搏也翻不动人拍板的值;
    # 物流号/收货人/备注等非财务列照常回填。
    from app.services import field_change_service as _fcs
    _LOCK_FIELDS = ("paid_amount", "refund_amount", "status",
                    "buyer_payable_amount", "shop_received_amount")
    _locked_orders: set[str] = set()
    for _lf in _LOCK_FIELDS:
        _locked_orders |= _fcs.human_pks(db, table="orders", field=_lf)
    seen: set[str] = set()
    for no, o in orders.items():
        if not no:
            rep.skipped_invalid += 1
            continue
        if no in seen:
            rep.skipped_duplicate += 1
            continue
        seen.add(no)

        # 2025 订单已物理清理(用户拍板 2026-06-15); 淘宝"近3月"报表会带回老的已成交单 →
        # 不再重建, 否则每次取数都把已删的 2025 单灌回来(并再生缺收款等异常)。order_date 为空
        # (个别发货报表行)无法判断 → 保留交后续处理。
        _od = _to_date(o.order_date)
        if _od is not None and _od.year < 2026:
            continue

        # 主商品行: 优先在"非服务行"里取金额最大的一行 —— 送货入户/商家安装等服务行(常为¥0)不抢主位
        # (用户实测 2026-06-18: ¥11212 的餐边柜单被错标成"送货入户")。全是服务行才退而取金额最大。
        lines = o.lines or [{}]
        if len(lines) > 1:
            rep.multi_line_orders += 1
            _persist_order_lines(db, no, lines, resolver)   # 写各商品行 → 成本按行汇总(防塌单漏算)
        _non_service = [l for l in lines if not _is_service_line_name(l.get("product_name"))]
        primary = max(_non_service or lines, key=lambda x: (x.get("amount") or Decimal(0)))

        remark = None
        if len(lines) > 1:
            others = [
                f"{l.get('product_name') or ''}/{l.get('sku') or ''}×{l.get('qty') or 1}"
                for l in lines if l is not primary
            ]
            remark = "本单含%d个商品, 其余: %s" % (len(lines), "; ".join(others))
        if o.order_no_bad:
            rep.needs_review += 1
            remark = (remark + " | " if remark else "") + "⚠️ 订单号科学计数法损坏(需人工核对)"

        _pname = _clean(primary.get("product_name"))
        _sku = _clean(primary.get("sku"))
        # 历史 P+数字 编码统一成 PPS (用户拍板 2026-06-18: 以后不再用 P 开头), 否则匹配不到产品/成本
        _product_code = _norm_pps_code(primary.get("product_code") or None)
        _sku_code = _norm_pps_code(primary.get("sku_code"))
        _shop = _clean(o.shop)
        # Task 6: 用对应表按 skuId(精确) / 16位商家编码 反查 SKU编码/产品编码/店铺
        hit = taobao_listing_service.resolve_line(
            resolver, sku_id=primary.get("sku_id"), merchant_code=primary.get("sku_code"),
        )
        if hit:
            _sku_code = hit.get("sku_code") or _sku_code
            _product_code = hit.get("product_code") or _product_code
            _shop = hit.get("shop") or _shop

        # 财务字段 (淘宝导出为准): 应付/实付/实收/平台费/退款/发货日
        status, _recognized = _resolve_status(o.status_text)
        payable = _to_decimal(o.buyer_payable)
        # 实付优先, 但「买家实付列存在且=0」必须保留 0 (关闭/未付款单), 不能 `0 or 应付` 退回应付——
        # 否则关闭单(实付0)会拿到应付货款当假收入混进成交 (用户实测 2026-06-18)。仅当实付列缺失才退应付。
        _paid_real_d = _to_decimal(o.paid_real)
        paid = _paid_real_d if _paid_real_d is not None else _to_decimal(o.paid_amount)
        received = _to_decimal(o.shop_received)
        # 销售明细-only 导入(无订单报表 sheet)时订单级金额可能落在 ¥0 服务行上 → 用所有商品行
        # 金额之和兜底 应付/实付(真实订单总额), 否则整单 ¥0 被漏算。实付列明确为0(关闭单)不兜底。
        _line_total = sum((ln.get("amount") or Decimal("0")) for ln in lines)
        if (payable is None or payable == 0) and _line_total > 0:
            payable = _line_total
        if _paid_real_d is None and (paid is None or paid == 0) and _line_total > 0:
            paid = _line_total
        pfee = _to_decimal(o.platform_fee)
        _refund_raw = _to_decimal(o.refund)
        freight = _to_decimal(o.buyer_freight)   # 买家应付邮费 (订单级, 代收运费)
        # 乙(用户拍板 2026-06-23): 一单多宝贝 → 订单级金额按【子订单行求和】, 替代只取第一行
        # (此前多产品单只抓第一个子订单的实付/实收/应付 → 收入漏记其余产品; 实付被低估又连累口径A成本护栏
        #  误回退成单产品成本 → 毛利虚高)。单产品单(1行)求和=第一行, 行为不变。退款随行求和后归一;
        # 成本侧由 auto_cost_backfill 的口径A按子行汇总(实付修对后护栏放行)并排除被退子行。
        if len(lines) > 1:
            # 订单报表单级值(销售明细求和前): 防【不完整明细】把订单报表的正确总额压低 (2026-07-09,
            # 实测多产品单 13263 明细只匹配到餐桌一行 → 求和 1795.17 反把订单报表的 4500.36 覆盖掉)。
            _ord_payable, _ord_paid = _to_decimal(o.buyer_payable), _paid_real_d
            _ord_recv = _to_decimal(o.shop_received)
            def _sum_ln(key):
                vals = [ln.get(key) for ln in lines if ln.get(key) is not None]
                return sum(vals, Decimal("0")) if vals else None
            _s_pay, _s_paid = _sum_ln("amount"), _sum_ln("paid_real")
            _s_recv, _s_fee, _s_refund = _sum_ln("shop_received"), _sum_ln("platform_fee"), _sum_ln("refund")
            # 求和 ≥ 订单报表单级值(明细完整)才采用; 求和更小 = 明细不完整 → 保留单级值, 不压低。
            if _s_pay is not None and _s_pay > 0 and (_ord_payable is None or _s_pay >= _ord_payable):
                payable = _s_pay
            if _s_paid is not None and _s_paid > 0 and (_ord_paid is None or _s_paid >= _ord_paid):
                paid = _paid_real_d = _s_paid
            if _s_recv is not None and _s_recv > 0 and (_ord_recv is None or _s_recv >= _ord_recv):
                received = _s_recv
            if _s_fee is not None:
                pfee = _s_fee
            if _s_refund is not None:
                _refund_raw = _s_refund
            # 邮费是订单级(一单一次): 多产品行可能重复(每行同值)或只落某行 → 取 max, 不求和(防重复计)
            _f_vals = [ln.get("buyer_freight") for ln in lines if ln.get("buyer_freight") is not None]
            if _f_vals:
                _m_freight = max(_f_vals)
                if freight is None or _m_freight > freight:
                    freight = _m_freight
        refund = _normalize_refund(payable, paid, _refund_raw)
        ship_dt = _to_date(o.ship_time)
        # 防错标"待付款" 兜底 (用户拍板 2026-06-18): 状态文本应优先按 _STATUS_MAP 翻译(已补"已收货"等)。
        # 若仍落"待付款"但有【真实收款】凭据(店铺实收>0 或 买家实付>0)→ 纠正为已付款(视为已识别)。
        # ⚠ 不用物流单号作凭据: 实测物流号会出现在 ¥0/补拍/未成交单上, 不可靠(全刷会把假单也算进来)。
        if status == "pending_payment":
            _paid_real = _to_decimal(o.paid_real)
            if (received and received > 0) or (_paid_real and _paid_real > 0):
                status = "paid"
                _recognized = True
        # 导出无「订单状态」列(部分千牛订单报表/销售明细)→ status_text 为空, 上面落"未识别"。
        # 只要有付款凭据(实付/实收/应付任一>0)就是真实成交单, 当已签收完结, 不该拦截整批历史明细;
        # 真关闭/未付款单(实付0)后续在 settled_sale_clause(实付>0) 里自然排除。区别于"状态有值但陌生"(仍拦截)。
        if not _recognized and not str(o.status_text or "").strip():
            if (paid and paid > 0) or (received and received > 0) or (payable and payable > 0):
                status = "signed"
                _recognized = True
        # 多宝贝订单纠偏 (用户实测 2026-06-26, 订单 5115237121779012546): 一单多商品时整单状态在
        # _parse 阶段只取了该主订单号【首个子订单行】的「订单状态」(后续行只进 lines)。若首行恰是被
        # 退款/关闭的子单 → 整单被误标 cancelled/aftersales, 连带把另一件【真实成交】产品一起漏出销售口径。
        # 有真实收款证据 (店铺实收>0, 或 已付款单的部分退款) 且非全额退款 → 不该被整单当取消漏出销售口径。
        # ⚠ 仅限多商品单 (len(lines)>1): 单商品关闭单不动 (实测那批多为拍下未付款·实收0, 本就该被排除)。
        # ── 2026-07-09 修盲区(实测 13263: 床头柜退款、餐桌还在工厂做): 原来一律纠成 signed(已签收),
        # 漏了"退一件、剩下那件还在做/未发货"这第三种(它是【进行中 paid】, 不是签收)。已卖出宝贝导出把
        # 退款件的"交易关闭"当整单状态 → 若一律判 signed, 就跟订单报表给的 paid 天天打架(paid↔signed 横跳)。
        # 用【店铺实收 打款商家金额】区分留下那件: 实收>0 = 淘宝已放款 = 真成交结算 → signed(如 5115:
        # 留下餐桌交易成功·实收883.30); 实收=0 = 还没结算(还在做/未发货)→ 进行中 paid(如 13263 餐桌·实收0)。
        if status in ("cancelled", "aftersales") and len(lines) > 1:
            _paid_g = paid or Decimal("0")
            _refund_g = refund or Decimal("0")
            _recv_g = received or Decimal("0")
            _is_full_refund = _paid_g > 0 and _refund_g >= _paid_g * Decimal("0.99")
            _has_real_money = _recv_g > 0 or (_refund_g > 0 and _paid_g > 0)
            if _has_real_money and not _is_full_refund:
                status = "signed" if _recv_g > 0 else "paid"   # 实收>0 已结算→签收; 实收0 还在做→进行中
                _recognized = True
        # 状态无法识别 且 无收款凭据 → 拦截不入库, 报异常待人工 (用户拍板 2026-06-18:
        # 不再默默塞成"待付款"被全系统漏算; 补好状态映射后重导即可入库)。
        if not _recognized:
            rep.held_unrecognized += 1
            try:
                from app.services import exception_service
                exception_service.record(
                    db, source_table="orders", source_pk=no,
                    exception_type="order_unrecognized_status", severity="warning",
                    description=f"订单 {no} 的订单状态「{o.status_text}」系统无法识别, 已拦截不入库, "
                                f"避免被错标待付款漏算。请补状态映射或确认后重新导入。",
                    suggestion_action="补订单状态映射(taobao_order_import._STATUS_MAP)后重新导入该单")
            except Exception:  # noqa: BLE001
                pass
            continue

        existing = db.execute(select(Order).where(Order.order_no == no)).scalar_one_or_none()
        if existing is not None:
            # 再次导入: 状态/金额以淘宝导出为准(覆盖); 描述/客户仅在缺失时回填; 不动 is_refill/remark。
            # is_historical 置 False: 淘宝真实订单应进统计(预测/月度报表/现金流), 旧通用导入误标历史在此纠正。
            # 用户拍板 (2026-06-11): 被覆盖的关键字段记修改档案 (source=import), 旧值可回溯。
            def _trace(fname: str, flabel: str, old_v, new_v) -> bool:
                if str(old_v) == str(new_v):
                    return False
                try:
                    from app.services import field_change_service
                    field_change_service.record(
                        db, table="orders", pk=no, field=fname, old=old_v, new=new_v,
                        actor="订单重导", source="import",
                        row_label=(existing.product_name or "")[:40], field_label=flabel,
                    )
                except Exception:
                    pass
                return True
            # ── 幂等护栏 (2026-07-09): 只有【订单级权威源 fin_source=="order"(订单报表/已卖出导出)】
            # 且【当前文件真带该列】才覆盖已存在订单的订单级财务字段; 行级销售明细 / 缺列的稀疏文件(如
            # 发货报表)只回填空值, 不许拿降级/默认值覆盖已有正确值。根治 202 单实付/状态/退款被不同来源
            # 文件反复横跳(实测 13263 在 4500.36↔1795.17、paid↔signed 之间弹 21 次)。
            _auth = (o.fin_source == "order")
            def _fin_overwrite(attr, value, present, fname=None, flabel=None, count=False, guard_zero=False):
                cur = getattr(existing, attr)
                if value is None or not ((present and _auth) or cur is None):
                    return
                # 0护栏 (2026-07-09): 不让"实付/应付/实收=0"把已有的正数金额清零 —— dry-run 实测有些
                # order 源导出会把老单的这些列报成 0, 裸覆盖会误清真实金额。仅当该单确在关闭
                # (status=cancelled)时才允许落 0(关闭单实付本就该是 0)。
                if guard_zero and value <= 0 and cur is not None and cur > 0 and status != "cancelled":
                    return
                if fname and _trace(fname, flabel, cur, value) and count:
                    rep.amount_changed += 1
                setattr(existing, attr, value)

            existing.is_historical = False
            # 人工锁: 人裁定过该单财务/状态 → 状态与全部财务列跳过(非财务列继续走下面的回填)
            if no in _locked_orders:
                pass
            else:
                # 状态: 仅当文件真带「订单状态」列(status_text 非空)才改。缺状态列的稀疏文件(发货报表/部分
                # 明细)其 status 是兜底默认的 signed(见上文无状态列默认段)—— 不许覆盖已有状态, 否则每天把
                # paid 盖成 signed。销售明细本身带订单级「订单状态」列, 属合法更新(与金额不同: 金额按行求和会
                # 因明细不完整而低估, 故金额仍要求权威源; 状态是订单级单值, 明细里也准)。
                # status_trusted=False("已卖出宝贝"式逐商品行状态)不许改已存在订单的状态 (2026-07-11)
                if str(o.status_text or "").strip() and getattr(o, "status_trusted", True):
                    if _trace("status", "订单状态", existing.status, status):
                        rep.status_changed += 1
                    existing.status = status
                _fin_overwrite("buyer_payable_amount", payable, o.buyer_payable is not None, guard_zero=True)
                _fin_overwrite("paid_amount", paid, o.paid_real is not None, "paid_amount", "实付金额", count=True, guard_zero=True)
                _fin_overwrite("shop_received_amount", received, o.shop_received is not None, guard_zero=True)
                _fin_overwrite("platform_fee", pfee, o.platform_fee is not None)
                _fin_overwrite("refund_amount", refund, o.refund is not None, "refund_amount", "退款金额", count=True)
                _fin_overwrite("buyer_freight", freight, o.buyer_freight is not None, "buyer_freight", "买家应付邮费")
            if ship_dt is not None:
                _trace("ship_date", "发货日期", existing.ship_date, ship_dt)
                existing.ship_date = ship_dt
            # 物流单号/承运商: 重导时回填(只填空, 不覆盖已手工改的)。
            # 修复(2026-06-15): 原更新分支漏了这两个 → 已发货订单重导也补不上物流号(图四常驻27)。
            _trk = _clean(o.tracking_no)
            if _trk and not existing.tracking_no:
                _trace("tracking_no", "物流单号", existing.tracking_no, _trk)
                existing.tracking_no = _trk
            _car = _clean(o.carrier)
            if _car and not existing.carrier:
                existing.carrier = _car
            # 平台备注随重导覆盖 (用户拍板: 买家留言/商家备注是淘宝侧会变的数据;
            # 非空才覆盖, 防止不含该列的旧格式文件把留言抹掉)
            _bmsg = _clean(o.buyer_message)
            if _bmsg:
                _trace("buyer_message", "买家留言", existing.buyer_message, _bmsg)
                existing.buyer_message = _bmsg
            _smemo = _clean(o.seller_memo)
            if _smemo:
                _trace("seller_memo", "商家备注", existing.seller_memo, _smemo)
                existing.seller_memo = _smemo
            if _shop and not existing.shop:
                existing.shop = _shop
            if _pname and not existing.product_name:
                existing.product_name = _pname
            if _sku and not existing.sku:
                existing.sku = _sku
            if _product_code and not existing.product_code:
                existing.product_code = _product_code
            if _sku_code and not existing.sku_code:
                existing.sku_code = _sku_code
            if o.customer_name and not existing.customer_name:
                existing.customer_name = _clean(o.customer_name)
            if o.customer_phone and not existing.customer_phone:
                existing.customer_phone = _clean(o.customer_phone)
            if o.customer_address and not existing.customer_address:
                existing.customer_address = _clean(o.customer_address)
            rep.updated += 1
            continue

        order = Order(
            platform=(platform or "淘宝").strip(),
            order_no=no,
            order_date=_to_date(o.order_date),
            ship_date=ship_dt,
            customer_name=_clean(o.customer_name),
            customer_phone=_clean(o.customer_phone),
            customer_address=_clean(o.customer_address),
            product_code=_product_code,
            product_name=_pname,
            sku=_sku,
            sku_code=_sku_code,
            shop=_shop,
            qty=_to_int(primary.get("qty"), default=1),
            carrier=_clean(o.carrier),
            tracking_no=_clean(o.tracking_no),
            buyer_payable_amount=payable,
            paid_amount=paid,
            shop_received_amount=received,
            platform_fee=pfee,
            refund_amount=refund,
            buyer_freight=freight,
            status=status,
            # 现金流"待确认收货/未发货"靠活跃单的金额; 故不再一律 historical
            is_historical=False,
            remark=remark,
            buyer_message=_clean(o.buyer_message),
            seller_memo=_clean(o.seller_memo),
            warehouse=order_cost_service.default_warehouse_for(_pname, _sku, False),
        )
        db.add(order)
        rep.inserted += 1
    db.commit()


def apply_refill_flags(db: Session) -> int:
    """每次导入订单后调用: 凡出现在补单对账(RefillRecord)里的订单号, 一律优先标成 is_refill=True。

    补单表是补单与否的最高优先级来源 —— 否则这些单会被当成"真实订单"少算补单(报表失真)。
    """
    from app.models.finance import RefillRecord
    refill_nos = {n for (n,) in db.execute(select(RefillRecord.order_no)).all() if n}
    if not refill_nos:
        return 0
    n = 0
    for o in db.execute(
        select(Order).where(Order.order_no.in_(refill_nos), Order.is_refill == False)  # noqa: E712
    ).scalars().all():
        o.is_refill = True
        n += 1
    return n


_OOXML_ENCRYPTED_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def is_encrypted_ooxml(raw: bytes) -> bool:
    """前 8 字节是 OOXML 加密复合文档魔数 → 加密发货报表。"""
    return raw[:8] == _OOXML_ENCRYPTED_MAGIC


def maybe_decrypt(raw: bytes, password: Optional[str]) -> bytes:
    """加密 OOXML(发货报表) → 用口令解成明文 xlsx 字节; 明文则原样返回。

    口令缺失/错误抛 ValueError, 调用方据此标「待口令」或「口令错误」。
    """
    if not is_encrypted_ooxml(raw):
        return raw
    if not password:
        raise ValueError("发货报表已加密, 但未取到飞书口令 (请转发『发货密码 xxx』到飞书机器人)")
    import msoffcrypto
    fin = io.BytesIO(raw)
    fout = io.BytesIO()
    office = msoffcrypto.OfficeFile(fin)
    try:
        office.load_key(password=password)
        office.decrypt(fout)
    except Exception as e:  # msoffcrypto 口令错误 → InvalidKeyError 等
        raise ValueError(f"发货报表解密失败 (口令可能已过期或不对): {e}") from e
    return fout.getvalue()


# ── 对外统一入口 ──────────────────────────────────────────────────────────────
def import_taobao_orders(db: Session, filename: str, raw: bytes,
                         platform: str = "淘宝",
                         force_format: Optional[str] = None,
                         password: Optional[str] = None) -> TaobaoImportReport:
    """自动识别格式并导入。force_format 可强制指定; password 用于解密加密发货报表。"""
    rep = TaobaoImportReport()
    if is_encrypted_ooxml(raw):
        try:
            raw = maybe_decrypt(raw, password)
        except ValueError as e:
            rep.errors.append(str(e))
            return rep
    fmt = force_format or detect_format(filename, raw)
    rep.detected_format = fmt

    ship_info: dict = {}
    if fmt == "qianniu_multi":
        orders, ship_info = _parse_qianniu_multi(raw, rep)
    elif fmt == "sales_detail":
        orders = _parse_sales_detail(filename, raw, rep)
    elif fmt == "order_master":
        # 老订单总表格式 → 交回既有通用导入 (列名已对齐), 这里仅提示
        rep.warnings.append("识别为『订单总表』格式, 请使用 /import-csv 或智能导入。")
        return rep
    else:
        rep.errors.append(
            "无法识别订单文件格式。支持: 千牛多表Excel(订单报表/销售明细/发货报表) "
            "或 销售明细(含 子订单编号/主订单编号/商品属性)。"
        )
        return rep

    if not orders:
        rep.warnings.append("未解析到任何订单行。")
        return rep

    # 行数骤降拦截已整体移除 (用户拍板 2026-06-15)。根因: 淘宝三张互补报表 —— 订单报表
    # (~897 单, 主单级含物流单号)、宝贝销售明细 (~851 单, 行级含商品/SKU)、发货报表
    # (~108 单, 仅已发货, 含收货人/电话) —— 各自覆盖不同订单子集, 却共用一个全局"上次行数"
    # 计数器互比, 导小表 (发货报表) 时必然误判"骤降 >50%"而拒导。这是拿异质报表当同一总体
    # 比较的设计缺陷, 不是真有坏文件。导入本就是 upsert (只增改、从不删行, 见 _commit_orders),
    # 残缺文件也抹不掉已有数据 → 该防护既误报又多余, 删除 (不再读写 last_order_import_count)。

    _commit_orders(db, orders, platform, rep)
    if apply_refill_flags(db):   # 用补单对账回标 is_refill (导入后立即匹配, 优先级最高)
        db.commit()

    # 发货报表收货人姓名兜底回填 (用户 2026-06-24): 销售明细没覆盖、但发货报表里有姓名的【已存在】
    # 订单, 把空的 客户名/电话/地址补上。每天解密发货报表都经此回填; 只更新已存在订单、只补空字段, 不新建。
    if ship_info:
        from app.models.order import Order as _OrderM
        nm_filled = 0
        for _sno, _info in ship_info.items():
            _nm, _ph, _ad = _clean(_info.get("name")), _clean(_info.get("phone")), _clean(_info.get("addr"))
            if not (_nm or _ph or _ad):
                continue
            _o = db.execute(select(_OrderM).where(_OrderM.order_no == _sno)).scalar_one_or_none()
            if _o is None:
                continue
            if _nm and not (_o.customer_name or "").strip():
                _o.customer_name = _nm
                nm_filled += 1
            if _ph and not (_o.customer_phone or "").strip():
                _o.customer_phone = _ph
            if _ad and not (_o.customer_address or "").strip():
                _o.customer_address = _ad
        if nm_filled:
            db.commit()
            rep.warnings.append(f"发货报表回填客户姓名 {nm_filled} 单")

    # 理论成本自动反推 (用户拍板 2026-06-12: 订单导入进来就应自动推算, 不再"未反推")。
    # only_missing=True 只补未算/为0的 (已算的不动); 失败不阻断导入。
    try:
        from app.services import order_cost_service
        order_cost_service.recompute_all(db, only_missing=True)
    except Exception as e:  # noqa: BLE001
        rep.warnings.append(f"理论成本自动反推未完成: {type(e).__name__}")

    # 导入消失检测 (用户拍板 2026-06-12; 根因修正后 2026-06-15 默认关闭"报缺")。
    # ⚠ 根因: report_missing 假设"单份文件 = 该时段全量快照", 但淘宝三张报表各覆盖不同子集
    # (订单报表897 / 销售明细851 / 发货报表108) —— 互比会把"另一张报表独有的单"误判成消失
    # (导 108 单的发货报表会一次性误报数百单 import_missing 异常)。这正是用户说的"误报背后有
    # 真问题": 用错了比较口径。导入是 upsert (永不删行), 即便真有单从淘宝消失, 库里也不会被抹,
    # 故"报缺"默认关; 仅当 settings taobao_import_vanish_check=1 且确知本次是单一权威订单表时才开。
    # resolve_reappeared 保留 (只销旧异常、绝不新建), 让历史误报在每次重导时自动收敛归零。
    try:
        from app.models.order import Order as _Order
        from app.services import import_vanish_service
        # orders 是 dict[order_no, _OrderRow] — 必须取 values(), 直接迭代拿到的是字符串键
        _rows = list(orders.values()) if isinstance(orders, dict) else list(orders)
        file_keys = {o.order_no for o in _rows if o.order_no and not o.order_no_bad}
        # 重新出现的单自动销掉旧异常 (拆单恢复/上次导出遗漏这次补上了 / 历史误报自愈)
        import_vanish_service.resolve_reappeared(
            db, source_table="orders", present_keys=file_keys)
        db.commit()
        _vanish_on = False
        try:
            from app.services import settings_service as _ss
            _vanish_on = str(_ss.get(db, "taobao_import_vanish_check",
                                     env_fallback=False) or "").strip().lower() \
                in ("1", "true", "yes", "on")
        except Exception:
            _vanish_on = False
        if _vanish_on:
            from datetime import date as _date_cls
            dates = [o.order_date for o in _rows
                     if isinstance(o.order_date, _date_cls)]
            if dates and file_keys:
                dmin, dmax = min(dates), max(dates)
                db_keys = {
                    r[0] for r in db.query(_Order.order_no).filter(
                        _Order.platform == platform,
                        _Order.is_refill == False,  # noqa: E712
                        _Order.order_date >= dmin,
                        _Order.order_date <= dmax,
                    ).all()
                }
                rep.vanished = import_vanish_service.report_missing(
                    db, source_table="orders", label="订单",
                    missing=sorted(db_keys - file_keys),
                    scope_desc=f"本次文件覆盖 {dmin}~{dmax}",
                )
                db.commit()
    except Exception:  # pragma: no cover - 检测故障不阻断导入
        import logging
        logging.getLogger("panse.taobao_import").warning("订单消失检测失败", exc_info=True)

    # 导入差异日报 (用户拍板): 有实际写入才推, 一眼看到今天发生了什么
    if rep.inserted or rep.updated:
        try:
            from app.services import notify_service
            vanish_line = (f"\n⚠ 有 {rep.vanished} 单在库里有但新文件里消失了, "
                           "已报异常等你确认是否删除 (异常中心 → import_missing)。"
                           if rep.vanished else "")
            notify_service.notify(
                db,
                (f"订单重导完成: 解析 {len(orders)} 单 — 新增 {rep.inserted}, "
                 f"更新 {rep.updated} (状态变更 {rep.status_changed}, 金额变更 {rep.amount_changed}), "
                 f"重复 {rep.skipped_duplicate}, 无效 {rep.skipped_invalid}"
                 + (f", ⚠拦截 {rep.held_unrecognized}(状态无法识别, 见异常页)" if rep.held_unrecognized else "")
                 + "。变更明细见 工具→修改档案 (来源筛「导入覆盖」)。" + vanish_line),
                level="warning" if rep.vanished else "info",
                title="畔色 ERP · 订单导入日报",
            )
        except Exception:  # pragma: no cover
            pass
    return rep

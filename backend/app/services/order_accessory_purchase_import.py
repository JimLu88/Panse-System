# -*- coding: utf-8 -*-
"""「订单号 + 配件」采购回填导入 (图5, 2026-06-12)。

用户在飞书/网页上传一张「订单号 + 配件(名称/编码)」表 →
对每行找该订单对应的配件清单项(OrderAccessoryItem), 把状态 未采购 → 已下单,
即从「待买」移到「已买未到」, 并从需要购买清单里去掉。可选带采购单号/快递单号。

幂等: 已是 已下单/运输中/已到货 的不重复处理 (只动 未采购)。匹配规则:
  同 order_no + (配件编码精确 优先; 否则配件名称包含匹配)。
与全局采购台账 purchase_table_import 不同 —— 那是记一笔采购, 这里是把订单的配件标已买。
"""
from __future__ import annotations

import csv as _csv
import io
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.order import OrderAccessoryItem

# 列名归一映射
_COL = {
    "order_no": ("订单号", "订单编号", "主订单编号", "order_no"),
    "material_name": ("配件名称", "配件", "名称", "物料名称", "品名", "material_name"),
    "material_code": ("配件编码", "物料编码", "编码", "料号", "material_code"),
    "purchase_no": ("采购单号", "采购单", "purchase_no"),
    "tracking_no": ("快递单号", "物流单号", "运单号", "tracking_no"),
}


def _norm(h: Optional[str]) -> str:
    return (h or "").replace(" ", "").replace("　", "").replace("﻿", "").strip()


def _read_rows(raw: bytes, filename: Optional[str]) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")) or raw[:2] == b"PK":
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [_norm(str(c)) if c is not None else "" for c in next(it, [])]
        rows = [{header[i]: r[i] for i in range(min(len(header), len(r)))} for r in it]
        wb.close()
        return rows
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("gbk", errors="replace")
    rdr = _csv.DictReader(io.StringIO(text))
    return [{_norm(k): v for k, v in row.items()} for row in rdr]


def _pick(row: dict, field: str) -> Optional[str]:
    for n in _COL[field]:
        if n in row and row[n] not in (None, ""):
            return str(row[n]).strip()
    return None


def import_order_part_purchases_core(db: Session, raw: bytes, filename: Optional[str]) -> dict:
    """返回 {rows, matched, updated, already, unmatched, unmatched_list}。
    matched=找到配件项的行; updated=本次由未采购翻为已下单; already=已是已买状态。"""
    from app.services import accessory_checklist_service
    rows = _read_rows(raw, filename)
    updated = already = unmatched = matched = 0
    unmatched_list: list[str] = []
    for row in rows:
        order_no = _pick(row, "order_no")
        mname = _pick(row, "material_name")
        mcode = _pick(row, "material_code")
        if not order_no or not (mname or mcode):
            continue
        stmt = select(OrderAccessoryItem).where(OrderAccessoryItem.order_no == order_no)
        if mcode:
            stmt = stmt.where(OrderAccessoryItem.material_code == mcode)
        items = db.execute(stmt).scalars().all()
        if mcode is None and mname:
            items = [i for i in items if mname in (i.material_name or "")]
        if not items:
            unmatched += 1
            unmatched_list.append(f"{order_no}/{mcode or mname}")
            continue
        matched += 1
        purchase_no = _pick(row, "purchase_no")
        tracking_no = _pick(row, "tracking_no")
        to_flip = [i.id for i in items if i.status == "未采购"]
        if to_flip:
            accessory_checklist_service.bulk_update(
                db, to_flip, status="已下单",
                purchase_no=purchase_no or None, tracking_no=tracking_no or None)
            updated += len(to_flip)
        else:
            already += 1
    db.commit()
    return {
        "rows": len(rows), "matched": matched, "updated": updated,
        "already": already, "unmatched": unmatched,
        "unmatched_list": unmatched_list[:20],
    }

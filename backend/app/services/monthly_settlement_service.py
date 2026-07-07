"""月度对账中心 (用户 2026-06-27, 方向三): 统一所有「月结」供应商对账 + 一键导出。

三域(都按发货月对账, 预估=应付基准, 实际=供应商账单):
  1) 配件月结 — 五金/电力轨道/岩板/玻璃, 复用 parts_recon_service.bulk_material_recon 的月结类;
     预估=Σ发货单BOM该类外采配件, 实际=工厂/供应商月度对账总额(PartsMonthlyRecon)。
  2) 打包月结 — 预估=Σ(发货成交单 est_packing, 按 ship_date 分月); 实际=PackingBill 应付(excluded=False, 按 bill_month)。
  3) 运费月结 — 预估=Σ(发货成交单 est_logistics, 按 ship_date 分月); 实际=LogisticsBill 逐单(row_type='line')按 bill_date 月,
     某月无逐单则用月结汇总(row_type='summary')兜底(覆盖德邦逐单 + 壹米滴答月结总额)。

口径红线: 这是【供应商应付(AP)核对】, 只为"这个月供应商收我多少、对不对", **不参与产品成本分摊**——
打包/运费早已通过 physical_cost 的 nz(actual_*, est_*) 计入每单成本(口径第16条), 这里绝不能再加一遍。
只读纯计算, 不写任何表。
"""
from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from typing import Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models.finance import LogisticsBill, PackingBill
from . import parts_recon_service as prs

_CENTS = Decimal("0.01")


def _q(x) -> Decimal:
    return Decimal(str(x or 0)).quantize(_CENTS)


def _variance_row(period: str, est: Decimal, actual: Optional[Decimal],
                  order_count: Optional[int] = None) -> dict:
    est_d = _q(est)
    if actual is None:
        return {"period": period, "estimate": float(est_d), "actual": None,
                "variance": None, "variance_pct": None, "order_count": order_count}
    act_d = _q(actual)
    var = (act_d - est_d).quantize(_CENTS)
    return {"period": period, "estimate": float(est_d), "actual": float(act_d),
            "variance": float(var),
            "variance_pct": float((var / est_d * 100).quantize(_CENTS)) if est_d > 0 else None,
            "order_count": order_count}


# ── 打包/运费的"预估" = 发货成交单的 est_packing / est_logistics 按 ship_date 月汇总 ──────
# 复用 parts_recon 的「已成交、已发货(ship_date非空)、非补单」口径, 保证三域消费窗口一致。
def _order_est_by_month(db: Session):
    pack: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    frgt: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    cnt: dict[str, int] = defaultdict(int)
    for o in prs._settled_shipped_orders(db):
        ym = prs._ym(o.ship_date)
        if not ym:
            continue
        pack[ym] += Decimal(str(o.est_packing or 0))
        frgt[ym] += Decimal(str(o.est_logistics or 0))
        cnt[ym] += 1
    return pack, frgt, cnt


def _packing_actual_by_month(db: Session) -> dict[str, Decimal]:
    out: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for b in db.execute(select(PackingBill)).scalars().all():
        if b.excluded or not b.bill_month:
            continue
        out[b.bill_month] += (b.packing_fee or Decimal("0"))
    return out


def _freight_actual_by_month(db: Session) -> dict[str, Decimal]:
    """逐单(line)优先, 某月无逐单则用月结汇总(summary)兜底 —— 对齐 reconciliation_service。"""
    line: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    summ: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for b in db.execute(select(LogisticsBill)).scalars().all():
        if not b.bill_date:
            continue
        ym = f"{b.bill_date.year}-{b.bill_date.month:02d}"
        amt = b.freight_amount or Decimal("0")
        if b.row_type == "summary":
            summ[ym] += amt
        else:  # 'line' 或历史无 row_type 的逐单
            line[ym] += amt
    return {ym: (line[ym] if ym in line else summ[ym]) for ym in set(line) | set(summ)}


def _simple_domain(key: str, label: str, hint: str,
                   est_map: dict[str, Decimal], act_map: dict[str, Decimal],
                   cnt_map: dict[str, int]) -> dict:
    periods = sorted(set(est_map) | set(act_map))
    rows, t_est, t_act = [], Decimal("0"), Decimal("0")
    for ym in periods:
        est = _q(est_map.get(ym, 0))
        act = _q(act_map[ym]) if ym in act_map else None
        rows.append(_variance_row(ym, est, act, order_count=cnt_map.get(ym)))
        t_est += est
        if act is not None:
            t_act += act
    t_var = (t_act - t_est).quantize(_CENTS)
    group = {"key": key, "label": label, "rows": rows,
             "total_estimate": float(t_est), "total_actual": float(t_act),
             "total_variance": float(t_var),
             "total_variance_pct": float((t_var / t_est * 100).quantize(_CENTS)) if t_est > 0 else None}
    return {"key": key, "label": label, "settle_hint": hint, "groups": [group]}


def build_center(db: Session) -> dict:
    """月度对账中心: 三域(配件/打包/运费)× 每月 预估|实际|差异|差异%。"""
    domains: list[dict] = []

    # 1) 配件月结 (复用 bulk_material_recon, 只取月结类) ─────────────
    bm = prs.bulk_material_recon(db, granularity="month")
    parts_groups = []
    for c in bm.get("materials", []):
        if c.get("settle_mode") != "月结":
            continue
        rows = [{
            "period": p["period"], "estimate": p["standard_consume"],
            "actual": p["actual"], "variance": p["variance"],
            "variance_pct": p["variance_pct"], "order_count": p["order_count"],
        } for p in c["periods"]]
        parts_groups.append({
            "key": c["key"], "label": c["name"], "rows": rows,
            "total_estimate": c["total_standard"], "total_actual": c["total_actual"],
            "total_variance": c["total_variance"], "total_variance_pct": c["total_variance_pct"],
        })
    domains.append({"key": "parts", "label": "配件月结",
                    "settle_hint": "工厂/供应商按月填月度对账总额", "groups": parts_groups})

    # 2) 打包月结 + 3) 运费月结 ─────────────────────────────────────
    pack_est, frgt_est, cnt = _order_est_by_month(db)
    domains.append(_simple_domain(
        "packing", "打包月结", "打包供应商月度账单(OCR手写账单录入)",
        pack_est, _packing_actual_by_month(db), cnt))
    domains.append(_simple_domain(
        "freight", "运费月结", "物流账单逐单按发货月汇总(德邦逐单/壹米滴答月结)",
        frgt_est, _freight_actual_by_month(db), cnt))

    return {
        "domains": domains,
        "caliber": "供应商应付(AP)核对 · 不参与产品成本分摊(打包/运费已计入每单 physical_cost)",
        "ship_date_basis": True,
    }


def _write_summary_ws(ws, data: dict) -> None:
    """月结汇总(美化): 域/分类/月/预估应付/实际账单/差异/差异%/发货单数; 蓝底表头 + 差异三色。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color=prs._C_BORDER)
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    headers = ["域", "分类", "月份", "预估应付", "实际账单", "差异", "差异%", "发货单数"]
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(1, ci)
        c.fill = PatternFill("solid", fgColor=prs._C_HEADER)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.border = BORDER
        c.alignment = AC
    ws.row_dimensions[1].height = 24
    r = 2
    for dom in data["domains"]:
        for g in dom["groups"]:
            for row in g["rows"]:
                ws.append([dom["label"], g["label"], row["period"], row["estimate"],
                           row["actual"] if row["actual"] is not None else "未录",
                           row["variance"] if row["variance"] is not None else "",
                           f'{row["variance_pct"]}%' if row["variance_pct"] is not None else "",
                           row.get("order_count") or ""])
                for ci in range(1, len(headers) + 1):
                    c = ws.cell(r, ci)
                    c.border = BORDER
                    c.alignment = AC
                    if ci in (4, 5, 6) and isinstance(c.value, (int, float)):
                        c.number_format = "#,##0.00"
                var = row["variance"]
                if var is not None:
                    col = "16A34A" if var < 0 else ("DC2626" if var > 0 else "6B7280")
                    ws.cell(r, 6).font = Font(color=col, bold=True)
                r += 1
    for i, w in enumerate([12, 16, 10, 14, 14, 13, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _write_settle_detail(ws, orders: list, *, est_label: str, act_label: str) -> None:
    """打包/运费 逐单明细(美化, 按发货月分组): 订单号/发货日/客户/收货地址/产品/预估X/实际X + 月小计/总计。"""
    from itertools import groupby
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color=prs._C_BORDER)
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL = Alignment(horizontal="left", vertical="center", wrap_text=True)
    AR = Alignment(horizontal="right", vertical="center")
    headers = ["订单号", "发货日", "客户", "收货地址", "产品", est_label, act_label]
    ncol, money_cols = len(headers), [6, 7]

    def _row(vals, *, fill=None, bold=False, fc=None, h=21):
        ws.append(vals)
        rr = ws.max_row
        ws.row_dimensions[rr].height = h
        for ci in range(1, ncol + 1):
            c = ws.cell(rr, ci)
            c.border = BORDER
            c.font = Font(bold=bold, color=fc, size=11)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            if ci == 1:
                c.number_format = "@"; c.alignment = AL
            elif ci in money_cols:
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0.00"
                c.alignment = AR
            elif ci in (4, 5):
                c.alignment = AL
            else:
                c.alignment = AC
        return rr

    _row(headers, fill=prs._C_HEADER, bold=True, fc="FFFFFF", h=26)
    ge = ga = Decimal("0")
    for ym, grp in groupby(orders, key=lambda o: (o.get("ship_date") or "")[:7] or "无发货日"):
        grp = list(grp)
        hr = _row([f"📅 {ym}　发货 {len(grp)} 单"] + [None] * (ncol - 1), fill=prs._C_MONTH, bold=True, h=24)
        ws.merge_cells(start_row=hr, start_column=1, end_row=hr, end_column=ncol)
        ws.cell(hr, 1).alignment = AL
        se = sa = Decimal("0")
        for oi, o in enumerate(grp):
            e, a = o.get("est"), o.get("act")
            se += Decimal(str(e or 0))
            sa += Decimal(str(a)) if a is not None else Decimal("0")
            _row([o["order_no"], o.get("ship_date") or "", o.get("customer_name") or "—",
                  o.get("customer_address") or "—", o.get("product") or "",
                  e, (a if a is not None else "未录")], fill=(prs._C_ZEBRA if oi % 2 else None))
        sr = [None] * ncol
        sr[0] = f"↳ {ym} 小计"; sr[5] = float(se); sr[6] = float(sa)
        _row(sr, fill=prs._C_SUBTOTAL, bold=True, h=20)
        ge += se; ga += sa
    tr = [None] * ncol
    tr[0] = "总计"; tr[5] = float(ge); tr[6] = float(ga)
    _row(tr, fill=prs._C_TOTAL, bold=True, h=26)
    for i, w in enumerate([22, 12, 12, 30, 18, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _packing_freight_orders(db: Session, *, date_from=None, date_to=None, year_month=None):
    """区间内 已成交已发货单 → (打包逐单, 运费逐单); 每单含 客户/收货地址/产品/预估/实际。"""
    def _in_range(sd) -> bool:
        if not sd:
            return False
        s = sd.isoformat()
        if year_month:
            return s[:7] == year_month
        if date_from and s < date_from:
            return False
        if date_to and s > date_to:
            return False
        return True

    pack, frgt = [], []
    for o in prs._settled_shipped_orders(db):
        if not _in_range(o.ship_date):
            continue
        base = {"order_no": o.order_no, "ship_date": o.ship_date.isoformat() if o.ship_date else "",
                "customer_name": o.customer_name, "customer_address": o.customer_address,
                "product": (o.product_name or o.sku or o.product_code or "")[:24]}
        pack.append({**base, "est": float(o.est_packing or 0),
                     "act": (float(o.actual_packing) if o.actual_packing is not None else None)})
        frgt.append({**base, "est": float(o.est_logistics or 0),
                     "act": (float(o.actual_logistics) if o.actual_logistics is not None else None)})
    pack.sort(key=lambda x: (x["ship_date"], x["order_no"]))
    frgt.sort(key=lambda x: (x["ship_date"], x["order_no"]))
    return pack, frgt


def build_export_workbook(db: Session, *, date_from=None, date_to=None, year_month=None):
    """一键导出全部月结账单 xlsx(『配件采购』好格式): 月结汇总 + 配件四账户逐单BOM明细 +
    打包/运费逐单明细。每单带客户名+收货地址。可选 发货日区间(date_from~date_to) 或单月(year_month)。"""
    import openpyxl

    data = build_center(db)     # 汇总仍是全量口径; 区间只影响下面明细页
    wb = openpyxl.Workbook()
    _write_summary_ws(wb.active, data)
    wb.active.title = "月结汇总"

    # 配件四账户(五金/电力轨道/岩板/玻璃) 各一页逐单展开 BOM + 系统预估单价/金额(复用配件采购好格式)。
    # export_shipped_orders 必须有日期圈定 → 无筛选(全部账期)时给全区间。
    a_from, a_to, a_ym = date_from, date_to, year_month
    if not (date_from or date_to or year_month):
        a_from, a_to = "1970-01-01", "2999-12-31"
    for dom in data["domains"]:
        if dom["key"] != "parts":
            continue
        for g in dom["groups"]:
            d = prs.export_shipped_orders(db, material_key=g["key"],
                                          date_from=a_from, date_to=a_to, year_month=a_ym)
            prs._write_category_ws(wb.create_sheet(("配件-" + g["label"])[:28]), d, show_est_price=True)

    # 打包 / 运费 逐单明细
    pack, frgt = _packing_freight_orders(db, date_from=date_from, date_to=date_to, year_month=year_month)
    _write_settle_detail(wb.create_sheet("打包月结明细"), pack, est_label="预估打包", act_label="实际打包")
    _write_settle_detail(wb.create_sheet("运费月结明细"), frgt, est_label="预估运费", act_label="实际运费")
    return wb


# ── 打包导清单: 当月发货单 + 每单预估/实际打包费 (给打包供应商核对账单) ──────────
def packing_checklist(db: Session, *, year_month: str) -> dict:
    """某发货月已发货成交单 + 每单 实际打包费(actual_packing, 已配打包账单回填的实付)。
    实际=发货月口径每单实付(未配到账单的单为空); 其合计与打包月结[账单月口径]口径不同、不必相等。
    est_packing(预估)一并带出备用。只读。"""
    out_orders: list[dict] = []
    t_est, t_act = Decimal("0"), Decimal("0")
    for o in prs._settled_shipped_orders(db):
        if prs._ym(o.ship_date) != year_month:
            continue
        est = _q(o.est_packing)
        act = _q(o.actual_packing) if o.actual_packing is not None else None
        t_est += est
        if act is not None:
            t_act += act
        out_orders.append({
            "order_no": o.order_no,
            "order_date": o.order_date.isoformat() if o.order_date else None,
            "ship_date": o.ship_date.isoformat() if o.ship_date else None,
            "customer_name": o.customer_name,
            "product_name": o.product_name,
            "sku": o.sku,
            "est_packing": float(est),
            "actual_packing": float(act) if act is not None else None,
        })
    out_orders.sort(key=lambda x: (x["ship_date"] or "", x["order_no"]))
    return {
        "year_month": year_month,
        "order_count": len(out_orders),
        "total_est_packing": float(t_est.quantize(_CENTS)),
        "total_actual_packing": float(t_act.quantize(_CENTS)),
        "orders": out_orders,
    }


def build_packing_checklist_xlsx(db: Session, *, year_month: str):
    """打包导清单 → xlsx(扁平表格)。订单号首列且**强制文本格式(@)**, 防 Excel 把19位订单号
    转科学计数法丢精度; 末尾预估/实际打包费合计行。返回 (Workbook, data)。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    d = packing_checklist(db, year_month=year_month)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"打包{year_month}"[:28]
    headers = ["订单号", "下单日期", "发货日", "客户", "产品", "SKU(含尺寸)", "实际打包费"]
    widths = [22, 12, 12, 12, 16, 22, 12]
    ws.append(headers)
    for o in d["orders"]:
        ws.append([
            o["order_no"], o.get("order_date") or "", o.get("ship_date") or "",
            o.get("customer_name") or "", o.get("product_name") or "", o.get("sku") or "",
            o.get("actual_packing"),
        ])
    ws.append(["合计(实际)", "", "", "", "", "", d["total_actual_packing"]])

    head_fill = PatternFill("solid", fgColor="E6F1FB")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    # 订单号列强制文本(@) — 19位订单号防 Excel 自动转科学计数法丢精度
    for (cell,) in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        cell.number_format = "@"
    for (cell,) in ws.iter_rows(min_row=2, min_col=7, max_col=7):   # 实际打包费 → 两位小数右对齐
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    return wb, d

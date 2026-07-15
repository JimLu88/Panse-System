"""价格对账: 千牛现价 vs ERP 应有值 (只读, 不改任何数据)。

背景 (2026-07-15): 长期"改 ERP、千牛没同步", 全店大面积价格漂移。本服务把
「千牛现价 vs ERP 应有值」做成可复跑的对账。两个维度:

- **标价维度** (reconcile_list_price): 千牛 SKU 一口价 (发布模板「价格(元)」列)
  vs ERP 应有一口价。★锚点严格按第一铁律 = **日常价 ÷ 0.75** (不直接信 list_price 字段;
  正常时 list_price ≡ daily/0.75, 但若某 SKU 未 recompute/冷导入脱钩, 以 daily 为准更贴铁律)。
  额外产出 incoherent: ERP 内部 list_price ≠ daily/0.75 的 SKU (=该 SKU 需 recompute)。
- **券后价维度** (reconcile_coupon_price): 千牛超级立减导出「活动普惠券后价」列
  vs ERP mid_buyer_price (中促/超级立减10%场买家到手)。★容差 0.01 (一分钱不差, 用户要求)。
  口径实证 (2026-07-15): ERP 未改动的 SKU, 千牛活动普惠券后价严格 == mid_buyer_price。

★踩过的坑 (务必守住):
- 发布模板有两个「商家编码」列: 商品级(短, 如 23250050202) 和 SKU级(长, 如 PPS2325005020232)。
  取 **SKU 级(最后一个)**。
- 也有两个价格列: col「一口价」= 商品级占位(2000/1000); **真正 SKU 价在「价格(元)」列**。别读错。
- 券后价源 (超级立减已报商品列表) 用 SKUID (淘宝 skuId) 而非商家编码 → 靠 PricingSkuPromo
  的 taobao_sku_id (+ alt_taobao_sku_ids) 映射回 sku_code。映射过期的 SKU 计入 unmapped。
"""
from __future__ import annotations

import io
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo

_TOL = Decimal("1")          # 标价容差 1 元 (千牛价与 ERP 应有值差 >1 元即算漂移)
_COUPON_TOL = Decimal("0.01")  # 券后价容差 0.01 元 (一分钱不差)
_CENT = Decimal("0.01")
_STD_DISCOUNT = Decimal("0.75")  # 单品宝标准折扣 (系统标准, 永不改动): 一口价 × 0.75 = 日常价
_PUBLISH_FALLBACK_PRICE_COL = 13  # 价格(元)  (1-indexed, 发布模板固定兜底)
_PUBLISH_FALLBACK_CODE_COL = 16   # SKU 级商家编码


def _should_list_price(daily) -> Optional[Decimal]:
    """ERP 应有一口价 = 日常价 ÷ 0.75 (第一铁律)。"""
    if daily is None:
        return None
    return (Decimal(str(daily)) / _STD_DISCOUNT).quantize(_CENT, rounding=ROUND_HALF_UP)


# ===========================================================================
# 标价维度
# ===========================================================================
def parse_qn_publish_export(file_bytes: bytes) -> dict[str, float]:
    """解析千牛「商品导出/发布模板」→ {SKU商家编码: 千牛SKU价}。
    优先按表头名定位「价格(元)」和(最后一个)「商家编码」列; 找不到用发布模板固定列兜底。"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["发布模板"] if "发布模板" in wb.sheetnames else wb.worksheets[0]
    hrow = 3
    for r in range(1, min(8, ws.max_row) + 1):
        vals = [str(ws.cell(r, c).value or "") for c in range(1, 10)]
        if any(v in ("商品Id", "商品ID") for v in vals):
            hrow = r
            break
    c_price: Optional[int] = None
    c_code: Optional[int] = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(hrow, c).value or "").strip()
        if h in ("价格(元)", "价格（元）"):
            c_price = c
        if h == "商家编码":
            c_code = c  # 取最后一个 = SKU 级
    if not c_price or not c_code:
        c_price, c_code = _PUBLISH_FALLBACK_PRICE_COL, _PUBLISH_FALLBACK_CODE_COL
    out: dict[str, float] = {}
    for r in range(hrow + 1, ws.max_row + 1):
        code = ws.cell(r, c_code).value
        price = ws.cell(r, c_price).value
        if code and price not in (None, ""):
            try:
                out[str(code).strip()] = float(price)
            except (TypeError, ValueError):
                pass
    return out


def reconcile_list_price(db: Session, qn_map: dict[str, float]) -> dict:
    """千牛 SKU 一口价 vs ERP 应有一口价(=日常价÷0.75)。返回漂移清单 + 统计 + 内部不自洽告警。"""
    promo = {p.sku_code: p for p in db.execute(select(PricingSkuPromo)).scalars()}
    mism = []
    incoherent = []   # ERP 内部 list_price 字段 ≠ daily/0.75 (该 SKU 需 recompute, 正常应为空)
    matched = 0
    for s in db.execute(select(PricingSku)).scalars():
        if s.is_custom_placeholder or not s.daily_price:
            continue
        should = _should_list_price(s.daily_price)   # 锚点 = 日常价÷0.75 (第一铁律)
        if should is None:
            continue
        # ERP 内部自洽检查: list_price 字段应 == daily/0.75; 不符 = 该 SKU 未 recompute (提醒修 ERP 内部)
        if s.list_price is not None and abs(Decimal(str(s.list_price)) - should) > _TOL:
            incoherent.append({
                "sku_code": s.sku_code,
                "sku_name": (s.sku or "")[:40],
                "list_price": round(float(s.list_price), 2),
                "daily_price": round(float(s.daily_price), 2),
                "should": round(float(should), 2),
            })
        qp = qn_map.get(s.sku_code)
        if qp is None:
            continue
        if abs(should - Decimal(str(qp))) > _TOL:
            pr = promo.get(s.sku_code)
            d = round(float(should) - qp, 2)
            mism.append({
                "taobao_item_id": str((pr.taobao_item_id if pr else "") or ""),
                "sku_code": s.sku_code,
                "sku_name": (s.sku or "")[:40],
                "qn_price": round(qp, 2),
                "erp_should": round(float(should), 2),
                "diff": d,
                "direction": "千牛偏低,要抬" if d > 0 else "千牛偏高,要降",
            })
        else:
            matched += 1
    mism.sort(key=lambda m: -abs(m["diff"]))
    incoherent.sort(key=lambda m: -abs(m["list_price"] - m["should"]))
    return {
        "mismatches": mism,
        "mismatch_count": len(mism),
        "matched": matched,
        "qn_total": len(qn_map),
        "incoherent": incoherent,
        "incoherent_count": len(incoherent),
    }


def reconcile(db: Session, file_bytes: bytes) -> dict:
    """标价维度完整对账: 解析千牛发布模板 → 标价对账。"""
    qn = parse_qn_publish_export(file_bytes)
    res = reconcile_list_price(db, qn)
    res["parsed_qn_sku"] = len(qn)
    return res


def build_fix_xlsx(db: Session, file_bytes: bytes) -> io.BytesIO:
    """标价返修表: 把漂移 SKU 的「正确一口价(=日常价÷0.75)」列出, 供千牛「商品价格批量编辑」改回。"""
    res = reconcile(db, file_bytes)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "标价返修"
    ws.append(["淘宝链接ID", "SKU商家编码", "SKU名", "千牛现价", "正确一口价(改成这个)", "差", "方向"])
    for m in res["mismatches"]:
        ws.append([m["taobao_item_id"], m["sku_code"], m["sku_name"],
                   m["qn_price"], m["erp_should"], m["diff"], m["direction"]])
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFG", [15, 18, 30, 12, 18, 9, 13]):
        ws.column_dimensions[col].width = w
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ===========================================================================
# 券后价维度 (超级立减/中促10%场)
# ===========================================================================
def parse_qn_activity_export(file_bytes: bytes) -> dict[str, dict]:
    """解析千牛「超级立减已报商品列表」→ {SKUID: {coupon, activity, deduct, one, name}}。
    按表头名定位列 (表头行含 'SKUID'); 只取有「活动普惠券后价」的行 (=已报名 SKU)。
      coupon   = 活动普惠券后价 (千牛当前生效的普惠到手价)
      activity = 活动价 (报名填的, 应=ERP日常价)
      deduct   = 补贴金额 (单品立减, 供参考; 系统推荐值未必=生效值)
      one      = 一口价 (商品级占位, 供参考)
      name     = SKU名称
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["已报商品列表"] if "已报商品列表" in wb.sheetnames else wb.worksheets[0]
    # 找表头行 (前 5 行里含 'SKUID')
    hrow = 2
    for r in range(1, min(6, ws.max_row) + 1):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if "SKUID" in vals:
            hrow = r
            break
    col: dict[str, Optional[int]] = {"sku": None, "coupon": None, "act": None,
                                     "deduct": None, "one": None, "name": None}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(hrow, c).value or "").strip()
        if h == "SKUID":
            col["sku"] = c
        elif h == "活动普惠券后价":
            col["coupon"] = c
        elif h == "活动价":
            col["act"] = c
        elif h == "补贴金额":
            col["deduct"] = c
        elif h == "一口价":
            col["one"] = c
        elif h == "SKU名称":
            col["name"] = c
    out: dict[str, dict] = {}
    if not col["sku"] or not col["coupon"]:
        return out

    def _num(r, c):
        if not c:
            return None
        v = ws.cell(r, c).value
        if v in (None, "", "-"):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    for r in range(hrow + 1, ws.max_row + 1):
        skuid = ws.cell(r, col["sku"]).value
        if not skuid or not str(skuid).strip().isdigit():
            continue  # 跳过空行 / 列说明行 (说明行 SKUID 列是中文说明, 非数字)
        coupon = _num(r, col["coupon"])
        if coupon is None:
            continue  # 未报名 / 无生效券后价
        out[str(skuid).strip()] = {
            "coupon": coupon,
            "activity": _num(r, col["act"]),
            "deduct": _num(r, col["deduct"]),
            "one": _num(r, col["one"]),
            "name": str(ws.cell(r, col["name"]).value or "")[:40] if col["name"] else "",
        }
    return out


def _build_skuid_index(db: Session):
    """taobao_sku_id (+ alt) → (PricingSku, PricingSkuPromo)。"""
    skus = {s.sku_code: s for s in db.execute(select(PricingSku)).scalars()}
    idx: dict[str, tuple] = {}
    for p in db.execute(select(PricingSkuPromo)).scalars():
        s = skus.get(p.sku_code)
        if not s:
            continue
        ids = []
        if p.taobao_sku_id:
            ids.append(str(p.taobao_sku_id))
        for a in (p.alt_taobao_sku_ids or []):
            if a:
                ids.append(str(a))
        for i in ids:
            idx[i] = (s, p)
    return idx


def reconcile_coupon_price(db: Session, qn_map: dict[str, dict]) -> dict:
    """千牛「活动普惠券后价」 vs ERP mid_buyer_price (中促/超级立减10%场买家到手)。
    容差 0.01 (一分钱不差)。映射不上的 SKUID 计入 unmapped; ERP 无中促到手的计 no_target。"""
    idx = _build_skuid_index(db)
    mism = []
    matched = 0
    unmapped = 0
    no_target = 0
    skipped_custom = 0
    for skuid, d in qn_map.items():
        coupon = d.get("coupon")
        if coupon is None:
            continue
        hit = idx.get(str(skuid))
        if not hit:
            unmapped += 1
            continue
        s, p = hit
        if s.is_custom_placeholder:
            skipped_custom += 1
            continue
        target = p.mid_buyer_price
        if target is None:
            no_target += 1
            continue
        tgt = Decimal(str(target))
        cpn = Decimal(str(coupon))
        if abs(cpn - tgt) > _COUPON_TOL:
            diff = round(float(tgt) - float(cpn), 2)
            mism.append({
                "sku_code": s.sku_code,
                "taobao_sku_id": str(skuid),
                "sku_name": d.get("name") or (s.sku or "")[:40],
                "qn_coupon": round(float(coupon), 2),
                "erp_should": round(float(tgt), 2),         # ERP 中促买家到手 = 应有券后价
                "diff": diff,
                "direction": "千牛偏低,要抬" if diff > 0 else "千牛偏高,要降",
                "qn_activity": None if d.get("activity") is None else round(d["activity"], 2),
                "erp_daily": None if s.daily_price is None else round(float(s.daily_price), 2),
            })
        else:
            matched += 1
    mism.sort(key=lambda m: -abs(m["diff"]))
    return {
        "mismatches": mism,
        "mismatch_count": len(mism),
        "matched": matched,
        "unmapped": unmapped,
        "no_target": no_target,
        "skipped_custom": skipped_custom,
        "qn_total": len(qn_map),
    }


def reconcile_coupon(db: Session, file_bytes: bytes) -> dict:
    """券后价维度完整对账: 解析千牛超级立减导出 → 券后价对账。"""
    qn = parse_qn_activity_export(file_bytes)
    res = reconcile_coupon_price(db, qn)
    res["parsed_qn_sku"] = len(qn)
    return res


def build_coupon_fix_xlsx(db: Session, file_bytes: bytes) -> io.BytesIO:
    """券后价返修表: 漂移 SKU 的正确券后价(=ERP mid_buyer_price) + 应填单品立减金额提示。
    单品立减金额 = 日常价×(1−10%) − 应有券后价 (加法口径; 到手=日常−官方立减10%−单品立减)。"""
    res = reconcile_coupon(db, file_bytes)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "券后价返修"
    ws.append(["SKUID", "SKU商家编码", "SKU名", "千牛现券后价", "正确券后价(ERP中促到手)",
               "差", "方向", "ERP日常价", "应填单品立减金额"])
    for m in res["mismatches"]:
        daily = m.get("erp_daily")
        should = m.get("erp_should")
        deduct = None
        if daily is not None and should is not None:
            deduct = round(float(daily) * 0.9 - float(should), 2)
        ws.append([m["taobao_sku_id"], m["sku_code"], m["sku_name"],
                   m["qn_coupon"], m["erp_should"], m["diff"], m["direction"],
                   daily, deduct])
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFGHI", [16, 18, 26, 13, 18, 9, 12, 11, 14]):
        ws.column_dimensions[col].width = w
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

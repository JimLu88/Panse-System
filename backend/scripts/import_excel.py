"""Phase 1 冷启动：把 Excel 的 4 张表导入数据库。

依赖顺序 (plan §13)：产品总表 → 物料单价库 → BOM → 库存。
配件库存表里可能有「定制」物料 (AC-1000+) 不在价格表里——这种情况下走
material_service.ensure_by_name() 自动建档，与运行时录入行为一致。

用法：
    python -m scripts.import_excel path/to/excel.xlsx
"""
from __future__ import annotations

import argparse
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

import openpyxl

from app.database import SessionLocal
from app.models.bom import BomLine
from app.models.finance import AccountBalance, AlipayFlow, RefillRecord
from app.models.inventory import PartInventory, ProductInventory
from app.models.material import Material
from app.models.order import Order
from app.models.pricing import PricingSku
from app.models.product import Product
from app.services import exception_service, inventory_service, material_service  # noqa: F401

SHEET_PRODUCTS = "1-产品总表"
SHEET_PRICING = "2-定价总表"
SHEET_PRICE = "3b-配件价格表"
SHEET_BOM = "3-BOM表"
SHEET_PART_INV = "4b-配件库存"
SHEET_PROD_INV = "4a-成品库存"
SHEET_ORDERS = "5-订单总表"
SHEET_REFILL = "8-补单记录"
SHEET_ALIPAY_SHEETS = (
    ("9a-支付宝流水-企业号", "企业号"),
    ("9b-支付宝流水-个体户私账", "个体户私账"),
    ("9c-支付宝流水-爱群号（未来弃用）", "爱群号"),
    ("9d-支付宝流水-佳宝号（未来弃用）", "佳宝号"),
    ("9e-支付宝流水-主力号", "主力号"),
)
SHEET_BALANCE = "10-账户余额汇总"


def _str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _decimal(v: Any) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _int(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def _find_header_row(ws, expected_first_col: str) -> int:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _str(row[0]) == expected_first_col:
            return row_idx
    raise ValueError(f"could not find header row starting with {expected_first_col!r} in sheet {ws.title}")


def import_products(wb, db) -> int:
    # 列: 产品编码 / 类目 / 产品名称 / 图片 / SKU / SKU编码 / ...
    # 同一 产品编码 在 Excel 里出现多行（每个 SKU 一行），导入时按编码去重。
    ws = wb[SHEET_PRODUCTS]
    header_row = _find_header_row(ws, "产品编码")
    count = 0
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = _str(row[0])
        category = _str(row[1])
        name = _str(row[2])
        if not code or not name or code in seen:
            continue
        if db.query(Product).filter_by(code=code).first():
            seen.add(code)
            continue
        db.add(Product(code=code, name=name, category=category))
        seen.add(code)
        count += 1
    db.commit()
    return count


def import_materials(wb, db) -> int:
    ws = wb[SHEET_PRICE]
    header_row = _find_header_row(ws, "物料编码")
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = _str(row[0])
        name = _str(row[1])
        if not code or not name:
            continue
        if db.query(Material).filter_by(code=code).first():
            continue
        # 判断是否定制：AC 前缀且序号 >= 1000
        is_custom = False
        if code.startswith("AC-"):
            try:
                is_custom = int(code.split("-", 1)[1]) >= 1000
            except (IndexError, ValueError):
                pass
        db.add(Material(
            code=code,
            name=name,
            size_type=_str(row[2]),
            unit=_str(row[3]),
            price=_decimal(row[4]),
            remark=_str(row[5]),
            is_custom=is_custom,
        ))
        count += 1
    db.commit()
    return count


def import_bom(wb, db) -> int:
    ws = wb[SHEET_BOM]
    header_row = _find_header_row(ws, "产品编码")
    count = 0
    skipped_missing_material = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        product_code = _str(row[0])
        material_code = _str(row[4])
        if not product_code or not material_code:
            continue
        # 物料缺失则跳过（这是数据问题，不在冷启动阶段自动建档）
        if not db.query(Material).filter_by(code=material_code).first():
            skipped_missing_material += 1
            continue
        db.add(BomLine(
            product_code=product_code,
            sku=_str(row[2]),
            sku_code=_str(row[3]),
            material_code=material_code,
            unit=_str(row[8]),
            qty_per_product=_decimal(row[9]) or Decimal("1"),
        ))
        count += 1
    db.commit()
    if skipped_missing_material:
        print(f"  ! BOM 中 {skipped_missing_material} 行因物料缺失被跳过")
    return count


def import_part_inventory(wb, db) -> int:
    ws = wb[SHEET_PART_INV]
    header_row = _find_header_row(ws, "仓库名称")
    count = 0
    autocreated = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        warehouse = _str(row[0])
        code = _str(row[1])
        name = _str(row[2])
        if not warehouse or (not code and not name):
            continue
        # 优先按编码走，编码缺失时按名称走
        if code:
            existing = db.query(Material).filter_by(code=code).first()
            if not existing and name:
                # 编码在库存里但价格表里没有 → 这就是定制物料场景
                # 直接按 Excel 给出的编码建档，但要 ensure 名字以「定制」起头一致
                display_name = name if name.startswith("定制") else f"定制{name}"
                is_custom = code.startswith("AC-") and int(code.split("-", 1)[1]) >= 1000
                db.add(Material(code=code, name=display_name, is_custom=is_custom, unit=_str(row[4])))
                db.flush()
                autocreated += 1
                if is_custom:
                    exception_service.record(
                        db,
                        source_table="materials",
                        source_pk=code,
                        exception_type="missing_material_autocreated",
                        severity="warning",
                        description=(
                            f"冷启动从 4b-配件库存 自动建档定制物料 {code} ({display_name})。"
                            f"请补全 价格 / 尺寸类型 / 备注 后再投入使用。"
                        ),
                        suggestion_action="fill_material_fields",
                        context={"original_name": name, "assigned_code": code, "source": "import_excel"},
                    )
            material_code = code
        else:
            res = material_service.ensure_by_name(db, name)
            if res.created:
                autocreated += 1
            material_code = res.material.code

        db.add(PartInventory(
            warehouse=warehouse,
            material_code=material_code,
            spec=_str(row[3]),
            unit=_str(row[4]),
            physical_qty=_int(row[5]),
            locked_qty=_int(row[6]),
            remark=_str(row[18]) if len(row) > 18 else None,
        ))
        count += 1
    db.commit()
    if autocreated:
        print(f"  ! 配件库存中 {autocreated} 条物料自动建档 (Excel 价格表里缺失)")
    return count


def import_pricing(wb, db) -> int:
    """2-定价总表 → pricing_sku。列序参考 plan §3 / Excel 列：
    0 产品编码 / 1 产品名称 / 2 SKU / 3 SKU编码 / 4 图片链接 / 5 大小类型 /
    6 标价 / 7 日常 / 8 小促 / 9 中促 / 10 大促 / 11 大促利润 / 12 毛利率 /
    13 会计成本 / 14 平台费率 / 15 税费 / 16 物理总成本 / 17 物流 / 18 安装 /
    19 总出厂成本 / 20 木作成本 / 21 打包 / 22 外采配件
    """
    ws = wb[SHEET_PRICING]
    header_row = _find_header_row(ws, "产品编码")
    count = 0
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sku_code = _str(row[3])
        product_code = _str(row[0])
        if not sku_code or not product_code or sku_code in seen:
            continue
        if db.query(PricingSku).filter_by(sku_code=sku_code).first():
            seen.add(sku_code)
            continue
        seen.add(sku_code)
        db.add(PricingSku(
            product_code=product_code,
            sku=_str(row[2]),
            sku_code=sku_code,
            image_url=_str(row[4]),
            size_category=_str(row[5]),
            list_price=_decimal(row[6]),
            daily_price=_decimal(row[7]),
            small_promo=_decimal(row[8]),
            mid_promo=_decimal(row[9]),
            big_promo=_decimal(row[10]),
            big_promo_margin=_decimal(row[11]),
            gross_margin_rate=_decimal(row[12]),
            accounting_cost=_decimal(row[13]),
            platform_fee_rate=_decimal(row[14]),
            tax=_decimal(row[15]),
            physical_cost=_decimal(row[16]),
            logistics_cost=_decimal(row[17]),
            install_cost=_decimal(row[18]),
            factory_cost=_decimal(row[19]),
            wood_cost=_decimal(row[20]),
            packaging_cost=_decimal(row[21]),
            external_parts_cost=_decimal(row[22]),
        ))
        count += 1
    db.commit()
    return count


def import_product_inventory(wb, db) -> int:
    ws = wb[SHEET_PROD_INV]
    header_row = _find_header_row(ws, "仓库名称")
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        warehouse = _str(row[0])
        product_code = _str(row[1])
        if not warehouse or not product_code:
            continue
        db.add(ProductInventory(
            warehouse=warehouse,
            product_code=product_code,
            sku=_str(row[3]),
            spec=_str(row[4]),
            unit=_str(row[5]),
            physical_qty=_int(row[6]),
            locked_qty=_int(row[7]),
        ))
        count += 1
    db.commit()
    return count


def import_orders(wb, db) -> int:
    """5-订单总表 → orders。列：平台/订单编号/是否补单/下单日期/客户/电话/地址/发货日期/
    产品编码/产品名称/SKU/是否定制/数量/锁定状态/物流公司/物流单号/...
    冷启动只取核心字段；定制订单走 is_custom 标记。
    """
    ws = wb[SHEET_ORDERS]
    header_row = _find_header_row(ws, "平台")
    count = 0
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        order_no = _str(row[1])
        if not order_no or order_no.startswith("#"):
            continue
        if order_no in seen or db.query(Order).filter_by(order_no=order_no).first():
            continue
        seen.add(order_no)
        # 解析日期：Excel cell 可能是 datetime 或字符串
        order_date_raw = row[3]
        order_date_val = None
        if hasattr(order_date_raw, "date"):
            order_date_val = order_date_raw.date()
        elif isinstance(order_date_raw, str) and order_date_raw.strip():
            try:
                from datetime import datetime
                order_date_val = datetime.strptime(order_date_raw.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass

        is_custom_raw = _str(row[11])
        db.add(Order(
            platform=_str(row[0]) or "淘宝",
            order_no=order_no,
            is_refill=_str(row[2]) == "是",
            order_date=order_date_val,
            customer_name=_str(row[4]) if _str(row[4]) and not _str(row[4]).startswith("#") else None,
            customer_phone=_str(row[5]) if _str(row[5]) and not _str(row[5]).startswith("#") else None,
            customer_address=_str(row[6]) if _str(row[6]) and not _str(row[6]).startswith("#") else None,
            product_code=_str(row[8]),
            product_name=_str(row[9]),
            sku=_str(row[10]),
            is_custom=bool(is_custom_raw) and is_custom_raw in ("是", "Y", "1"),
            qty=_int(row[12]) or 1,
            status="pending_payment",
        ))
        count += 1
    db.commit()
    return count


def import_refill(wb, db) -> int:
    if SHEET_REFILL not in wb.sheetnames:
        return 0
    ws = wb[SHEET_REFILL]
    header_row = _find_header_row(ws, "订单编号")
    count = 0
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        order_no = _str(row[0])
        if not order_no or order_no in seen:
            continue
        seen.add(order_no)
        refill_date_raw = row[2]
        refill_date_val = None
        if hasattr(refill_date_raw, "date"):
            refill_date_val = refill_date_raw.date()
        db.add(RefillRecord(
            order_no=order_no,
            buyer_nick=_str(row[1]),
            refill_date=refill_date_val,
            product_code=_str(row[3]),
            product_name=_str(row[4]),
            sku=_str(row[5]),
            order_amount=_decimal(row[6]),
            qty=_int(row[7]) or 1,
            refill_cost=_decimal(row[8]),
            refill_freight=_decimal(row[9]),
            platform_fee=_decimal(row[10]),
            total_cost=_decimal(row[11]),
        ))
        count += 1
    db.commit()
    return count


def import_alipay(wb, db) -> int:
    """合并 9a~9e 五张表到 alipay_flows，按账户区分。"""
    from datetime import datetime as _dt
    total = 0
    for sheet, account in SHEET_ALIPAY_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        try:
            header_row = _find_header_row(ws, "交易时间")
        except ValueError:
            continue
        seen: set[str] = set()
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            tx_no = _str(row[1])
            if not tx_no or tx_no in seen:
                continue
            amount = _decimal(row[5])
            if amount is None:
                continue
            seen.add(tx_no)
            if db.query(AlipayFlow).filter_by(account=account, transaction_no=tx_no).first():
                continue
            tx_time = row[0]
            tx_time_val = None
            if hasattr(tx_time, "date"):
                tx_time_val = tx_time
            elif isinstance(tx_time, str) and tx_time.strip():
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        tx_time_val = _dt.strptime(tx_time.strip(), fmt)
                        break
                    except ValueError:
                        continue
            db.add(AlipayFlow(
                account=account,
                transaction_no=tx_no,
                transaction_time=tx_time_val,
                transaction_type=_str(row[2]),
                counterparty=_str(row[3]),
                counterparty_account=_str(row[4]),
                amount=amount,
                related_order_no=_str(row[6]),
                balance=_decimal(row[7]),
                reconciliation_status=_str(row[8]) or "open",
                reconciliation_type=_str(row[9]),
                remark=_str(row[10]) if len(row) > 10 else None,
            ))
            total += 1
        db.commit()
    return total


def import_account_balances(wb, db) -> int:
    if SHEET_BALANCE not in wb.sheetnames:
        return 0
    ws = wb[SHEET_BALANCE]
    header_row = _find_header_row(ws, "账户名称")
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = _str(row[0])
        if not name:
            continue
        dt = row[2]
        period_year = period_month = None
        if hasattr(dt, "year"):
            period_year, period_month = dt.year, dt.month
        if period_year is None:
            continue
        if db.query(AccountBalance).filter_by(
            account_name=name, period_year=period_year, period_month=period_month
        ).first():
            continue
        opening = _decimal(row[3]) or Decimal("0")
        db.add(AccountBalance(
            account_name=name,
            account_no=_str(row[1]),
            period_year=period_year,
            period_month=period_month,
            opening_balance=opening,
            income=_decimal(row[4]) or Decimal("0"),
            expense=_decimal(row[5]) or Decimal("0"),
            closing_balance=_decimal(row[6]) or opening,
        ))
        count += 1
    db.commit()
    return count


def run(excel_path: Path) -> None:
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    db = SessionLocal()
    try:
        print(f"=== Importing {excel_path} ===")
        print(f"产品总表: {import_products(wb, db)}")
        print(f"物料单价库: {import_materials(wb, db)}")
        print(f"定价总表: {import_pricing(wb, db)}")
        print(f"BOM 表: {import_bom(wb, db)}")
        print(f"配件库存: {import_part_inventory(wb, db)}")
        print(f"成品库存: {import_product_inventory(wb, db)}")
        print(f"订单总表: {import_orders(wb, db)}")
        print(f"补单记录: {import_refill(wb, db)}")
        print(f"支付宝流水 (9a-9e): {import_alipay(wb, db)}")
        print(f"账户余额: {import_account_balances(wb, db)}")
        print("=== Done ===")
    finally:
        db.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_path", type=Path)
    args = parser.parse_args()
    run(args.excel_path)


if __name__ == "__main__":
    main()

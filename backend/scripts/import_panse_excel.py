#!/usr/bin/env python
"""畔色 ERP 历史数据导入专用脚本.

针对 '畔色 ERP v2' Excel 多 sheet 结构定制. 每个 sheet 跳过标题行 + 自动列映射.

用法 (容器内):
    docker compose exec api python scripts/import_panse_excel.py /storage/your.xlsx
    docker compose exec api python scripts/import_panse_excel.py /storage/your.xlsx --dry-run
    docker compose exec api python scripts/import_panse_excel.py /storage/your.xlsx --sheet 1-产品总表 --sheet 3-BOM表

支持 sheet:
    1-产品总表        → products
    2-定价总表        → pricing_sku (占位文字填0; sku_code 唯一键 upsert)
    3b-配件价格表    → materials
    3-BOM表          → bom_lines (重复行保留入库, remark='重复行')
    4a-成品库存       → product_inventory
    4b-配件库存       → part_inventory
    5-订单总表        → orders (默认全部标 is_historical=True)
    6-工厂下单表      → factory_orders
    7-配件采购记录    → part_purchases
    10-账户余额汇总  → account_balances
    18-售后表        → after_sales
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Optional

# 容器内: scripts/__file__ → parent.parent = /app
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import SessionLocal


# ----------------------------- 工具 ---------------------------------- #


def _clean(v: Any) -> Any:
    """把 Excel 的 '-' / '#N/A' / 空字符串 统一转 None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "#N/A", "N/A", "NA", "无", "nan", "None"):
        return None
    return v


def _str(v: Any) -> Optional[str]:
    v = _clean(v)
    return str(v).strip() if v is not None else None


def _int(v: Any) -> Optional[int]:
    v = _clean(v)
    if v is None:
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _dec(v: Any) -> Optional[Decimal]:
    v = _clean(v)
    if v is None:
        return None
    try:
        s = str(v).replace(",", "").replace("¥", "").replace("元", "").strip()
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


# 占位符文字 (成本暂缺/待定等) → Decimal("0"), 真空白 → None
_PLACEHOLDER_TEXTS = {"成本暂缺", "待定", "暂缺", "待填", "tbd", "TBD", "?", "—"}


def _dec0(v: Any) -> Optional[Decimal]:
    """像 _dec(), 但把占位文字当 0 处理 (而非 None)."""
    if v is not None:
        s = str(v).strip()
        if s in _PLACEHOLDER_TEXTS:
            return Decimal("0")
    return _dec(v)


def _date(v: Any) -> Optional[date]:
    v = _clean(v)
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(
                s.replace("年", "-").replace("月", "-").replace("日", "")
                  .replace("/", "-").replace(".", "-"),
                "%Y-%m-%d",
            ).date()
        except ValueError:
            continue
    return None


def _bool(v: Any) -> bool:
    v = _clean(v)
    if v is None:
        return False
    return str(v).strip() in ("是", "yes", "true", "1", "Y", "y", "✓")


def _read_sheet(ws, header_row: int) -> tuple[list[str], list[dict]]:
    """读 sheet, 返回 (headers, rows). header_row 是 1-indexed."""
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < header_row:
        return [], []
    headers = [str(c).strip() if c else f"col{i+1}"
               for i, c in enumerate(all_rows[header_row - 1])]
    data_rows = []
    for r in all_rows[header_row:]:
        # 全空行跳过
        if all(c is None or (isinstance(c, str) and c.strip() in ("", "-"))
               for c in r):
            continue
        d = {h: r[i] if i < len(r) else None for i, h in enumerate(headers)}
        data_rows.append(d)
    return headers, data_rows


# ----------------------------- 各 sheet importer ---------------------- #


def import_products(db, ws) -> dict:
    """1-产品总表 → products. header 在 row 2."""
    from app.models.product import Product
    _, rows = _read_sheet(ws, header_row=2)
    inserted, updated, skipped = 0, 0, 0
    seen_codes = set()
    for r in rows:
        code = _str(r.get("产品编码"))
        if not code:
            skipped += 1; continue
        if code in seen_codes:
            continue  # 产品总表里同产品多 SKU 会重复, 取第一条
        seen_codes.add(code)
        name = _str(r.get("产品名称")) or code
        cat = _str(r.get("类目"))
        existing = db.execute(select(Product).where(Product.code == code)).scalar_one_or_none()
        if existing:
            existing.name = name; existing.category = cat
            updated += 1
        else:
            db.add(Product(code=code, name=name, category=cat))
            inserted += 1
    db.flush()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_materials(db, ws) -> dict:
    """3b-配件价格表 → materials. header row 2."""
    from app.models.material import Material
    _, rows = _read_sheet(ws, header_row=2)
    inserted, updated, skipped = 0, 0, 0
    # name 唯一, 重名时加 code 后缀
    name_seen: dict[str, str] = {}   # name -> first code
    # 预扫已有 name
    for row in db.execute(select(Material.name, Material.code)).all():
        name_seen[row[0]] = row[1]
    for r in rows:
        code = _str(r.get("物料编码"))
        name = _str(r.get("物料名称"))
        if not code or not name:
            skipped += 1; continue
        # 同名不同 code → 加后缀
        if name in name_seen and name_seen[name] != code:
            name = f"{name} ({code})"
        unit = _str(r.get("单位"))
        size_type = _str(r.get("尺寸类型"))
        price = _dec(r.get("计算价格"))
        remark = _str(r.get("备注"))
        is_custom = "定制" in name or "改" in code
        existing = db.execute(select(Material).where(Material.code == code)).scalar_one_or_none()
        if existing:
            existing.name = name; existing.unit = unit
            existing.size_type = size_type; existing.price = price
            existing.remark = remark; existing.is_custom = is_custom
            updated += 1
        else:
            db.add(Material(code=code, name=name, unit=unit, size_type=size_type,
                            price=price, remark=remark, is_custom=is_custom))
            inserted += 1
            name_seen[name] = code
    db.flush()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_pricing(db, ws) -> dict:
    """2-定价总表 → pricing_sku. header row 2.

    占位文字 (成本暂缺/待定等) 按用户决策填 0 而非 None.
    列名别名表与 excel_schemas.py 对齐.
    """
    from app.models.pricing import PricingSku

    ALIASES: dict[str, list[str]] = {
        "product_code":     ["产品编码"],
        "sku":              ["SKU", "规格"],
        "sku_code":         ["SKU编码", "SKU code"],
        "size_category":    ["大小分类", "尺寸分类", "产品大小类型", "大小类型", "Size"],
        "image_url":        ["图片链接", "图片", "图片URL", "主图链接"],
        "list_price":       ["标价", "标价计算", "List"],
        "daily_price":      ["日常价/单品宝", "日常价", "日常", "单品宝"],
        "small_promo":      ["小促价（超级立减）", "小促价", "小促", "超级立减", "立减价"],
        "mid_promo":        ["中促价（88券活动）", "中促价", "中促", "88券", "88 券"],
        "big_promo":        ["大促到手价（双11）", "大促价", "大促", "双11", "双 11"],
        "gross_margin_rate":["毛利率", "即时毛利率"],
        "big_promo_margin": ["大促利润", "利润"],
        "accounting_cost":  ["会计总成本", "总成本", "成本"],
        "physical_cost":    ["物理总成本", "物理成本"],
        "logistics_cost":   ["物流成本", "运费成本", "物流费"],
        "install_cost":     ["安装成本", "安装费"],
        "factory_cost":     ["总出厂成本", "出厂成本"],
        "wood_cost":        ["木作成本", "木材成本"],
        "packaging_cost":   ["包装成本", "包装费"],
        "external_parts_cost": ["外配件成本", "配件成本"],
        "platform_fee_rate":["平台费率", "佣金率"],
        "tax":              ["税", "税费"],
    }

    _DECIMAL0_FIELDS = {
        "list_price", "daily_price", "small_promo", "mid_promo", "big_promo",
        "gross_margin_rate", "big_promo_margin", "accounting_cost", "physical_cost",
        "logistics_cost", "install_cost", "factory_cost", "wood_cost",
        "packaging_cost", "external_parts_cost", "platform_fee_rate", "tax",
    }

    headers, rows = _read_sheet(ws, header_row=2)

    # 把 Excel 表头 → 字段名 (取第一个命中的别名)
    header_map: dict[str, str] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if alias in headers:
                header_map[alias] = field
                break

    def _get(r: dict, field: str) -> Any:
        for alias in ALIASES.get(field, []):
            if alias in r:
                return r[alias]
        return None

    inserted, updated, skipped = 0, 0, 0
    for r in rows:
        product_code = _str(_get(r, "product_code"))
        sku_code = _str(_get(r, "sku_code"))
        if not sku_code:
            skipped += 1; continue

        payload: dict[str, Any] = {
            "product_code": product_code or sku_code.split("-")[0],
            "sku":          _str(_get(r, "sku")),
            "sku_code":     sku_code,
            "size_category": _str(_get(r, "size_category")),
            "image_url":    _str(_get(r, "image_url")),
        }
        for field in _DECIMAL0_FIELDS:
            payload[field] = _dec0(_get(r, field))

        existing = db.execute(
            select(PricingSku).where(PricingSku.sku_code == sku_code)
        ).scalar_one_or_none()
        if existing:
            for k, v in payload.items():
                if v is not None:
                    setattr(existing, k, v)
            updated += 1
        else:
            db.add(PricingSku(**payload))
            inserted += 1
    db.flush()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_bom(db, ws) -> dict:
    """3-BOM表 → bom_lines. header row 2.

    重复行 (相同 product_code + sku_code + material_code) 保留入库,
    但在 remark 列标注 '重复行' 供人工核查.
    """
    from app.models.bom import BomLine
    from app.models.material import Material
    _, rows = _read_sheet(ws, header_row=2)
    inserted, annotated, skipped = 0, 0, 0
    known_materials = set(
        row[0] for row in db.execute(select(Material.code)).all()
    )
    # (product_code, sku_code, material_code) → first seen row index
    seen_combos: set[tuple] = set()
    for r in rows:
        product_code = _str(r.get("产品编码"))
        sku_code = _str(r.get("SKU编码"))
        material_code = _str(r.get("物料编码"))
        qty = _dec(r.get("单产品用量")) or Decimal("1")
        if not product_code or not material_code:
            skipped += 1; continue
        if material_code not in known_materials:
            raw_name = _str(r.get("物料名称")) or material_code
            db.add(Material(code=material_code,
                            name=f"{raw_name} ({material_code})"))
            db.flush()
            known_materials.add(material_code)
        combo = (product_code, sku_code, material_code)
        is_dup = combo in seen_combos
        seen_combos.add(combo)
        db.add(BomLine(
            product_code=product_code,
            sku=_str(r.get("SKU")),
            sku_code=sku_code,
            material_code=material_code,
            unit=_str(r.get("单位")),
            qty_per_product=qty,
            size_type=_str(r.get("尺寸类型")),
            remark="重复行" if is_dup else None,
        ))
        inserted += 1
        if is_dup:
            annotated += 1
    db.flush()
    return {"inserted": inserted, "annotated_duplicates": annotated, "skipped": skipped}


def import_product_inventory(db, ws) -> dict:
    """4a-成品库存 → product_inventory. header row 3."""
    from app.models.inventory import ProductInventory
    _, rows = _read_sheet(ws, header_row=3)
    inserted, updated, skipped = 0, 0, 0
    for r in rows:
        warehouse = _str(r.get("仓库名称")) or "江西仓库"
        product_code = _str(r.get("产品编码"))
        sku = _str(r.get("SKU"))
        if not product_code:
            skipped += 1; continue
        phys = _dec(r.get("物理库存")) or Decimal("0")
        locked = _dec(r.get("锁定库存")) or Decimal("0")
        existing = db.execute(select(ProductInventory).where(
            ProductInventory.warehouse == warehouse,
            ProductInventory.product_code == product_code,
            ProductInventory.sku == sku,
        )).scalar_one_or_none()
        if existing:
            existing.physical_qty = phys; existing.locked_qty = locked
            updated += 1
        else:
            db.add(ProductInventory(
                warehouse=warehouse, product_code=product_code, sku=sku,
                spec=_str(r.get("规格尺寸")), unit=_str(r.get("单位")),
                physical_qty=phys, locked_qty=locked,
            ))
            inserted += 1
    db.flush()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_part_inventory(db, ws) -> dict:
    """4b-配件库存 → part_inventory. header row 3."""
    from app.models.inventory import PartInventory
    from app.models.material import Material
    _, rows = _read_sheet(ws, header_row=3)
    inserted, updated, skipped = 0, 0, 0
    known_materials = set(
        row[0] for row in db.execute(select(Material.code)).all()
    )
    for r in rows:
        warehouse = _str(r.get("仓库名称")) or "江西仓库"
        code = _str(r.get("物料编码"))
        if not code:
            skipped += 1; continue
        if code not in known_materials:
            raw_name = _str(r.get("物料名称")) or code
            db.add(Material(code=code, name=f"{raw_name} ({code})"))
            db.flush()
            known_materials.add(code)
        phys = _dec(r.get("物理库存")) or Decimal("0")
        locked = _dec(r.get("锁定库存")) or Decimal("0")
        existing = db.execute(select(PartInventory).where(
            PartInventory.warehouse == warehouse,
            PartInventory.material_code == code,
        )).scalar_one_or_none()
        if existing:
            existing.physical_qty = phys; existing.locked_qty = locked
            updated += 1
        else:
            db.add(PartInventory(
                warehouse=warehouse, material_code=code,
                spec=_str(r.get("规格型号")), unit=_str(r.get("单位")),
                physical_qty=phys, locked_qty=locked,
                last_inbound_at=_date(r.get("最后入库日期")),
            ))
            inserted += 1
    db.flush()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_orders(db, ws, *, mark_historical: bool = True) -> dict:
    """5-订单总表 → orders. header row 2.

    默认全部标 is_historical=True, 不进库存流 / 不进资产公式.
    """
    from app.models.order import Order
    _, rows = _read_sheet(ws, header_row=2)
    inserted, skipped = 0, 0
    for r in rows:
        order_no = _str(r.get("订单编号"))
        if not order_no:
            skipped += 1; continue
        if db.execute(select(Order).where(Order.order_no == order_no)).scalar_one_or_none():
            skipped += 1; continue
        platform = _str(r.get("平台")) or "淘宝"
        status_raw = _str(r.get("订单状态")) or ""
        status_map = {
            "待付款": "pending_payment", "已付款": "paid",
            "已发货": "shipped", "已签收": "signed",
            "售后中": "aftersales", "已取消": "cancelled",
            "已完成": "signed",
        }
        status = status_map.get(status_raw, "signed")   # 历史单默认签收
        db.add(Order(
            platform=platform, order_no=order_no,
            is_refill=_bool(r.get("是否补单")),
            order_date=_date(r.get("下单日期")),
            ship_date=_date(r.get("发货日期")),
            customer_name=_str(r.get("客户姓名")),
            customer_phone=_str(r.get("联系电话")),
            customer_address=_str(r.get("收货地址")),
            product_code=_str(r.get("产品编码")),
            product_name=_str(r.get("产品名称")),
            sku=_str(r.get("SKU")),
            sku_code=_str(r.get("SKU编码")),
            is_custom="改" in (_str(r.get("SKU编码")) or ""),
            qty=_int(r.get("数量")) or 1,
            status=status,
            carrier=_str(r.get("快递公司")),
            tracking_no=_str(r.get("快递单号")),
            paid_amount=_dec(r.get("实付金额")),
            actual_freight=_dec(r.get("实际运费")),
            actual_cost=_dec(r.get("实际成本")),
            theoretical_cost=_dec(r.get("理论成本")),
            is_historical=mark_historical,
        ))
        inserted += 1
    db.flush()
    return {"inserted": inserted, "skipped": skipped}


def import_factory_orders(db, ws) -> dict:
    """6-工厂下单表 → factory_orders. header row 3."""
    from app.models.order import FactoryOrder
    _, rows = _read_sheet(ws, header_row=3)
    inserted, skipped = 0, 0
    for r in rows:
        fo_no = _str(r.get("工厂订单号"))
        if not fo_no:
            skipped += 1; continue
        if db.execute(select(FactoryOrder).where(
            FactoryOrder.factory_order_no == fo_no,
        )).scalar_one_or_none():
            skipped += 1; continue
        db.add(FactoryOrder(
            factory_order_no=fo_no,
            platform_order_no=_str(r.get("关联平台订单号")),
            factory_name=_str(r.get("工厂名称")),
            order_date=_date(r.get("下单日期")),
            expected_delivery=_date(r.get("要求交货日期")),
            actual_delivery=_date(r.get("实际交货日期")),
            product_code=_str(r.get("产品编码")),
            sku=_str(r.get("SKU")),
            qty=_int(r.get("数量")) or 1,
            unit_price=_dec(r.get("工厂单价")),
            factory_bill_amount=_dec(r.get("工厂结算金额")),
            payment_status=_str(r.get("支付状态")) or "unpaid",
            payment_method=_str(r.get("支付方式")),
            carrier=_str(r.get("快递公司")),
            tracking_no=_str(r.get("快递单号")),
            remark=_str(r.get("备注")),
        ))
        inserted += 1
    db.flush()
    return {"inserted": inserted, "skipped": skipped}


def import_part_purchases(db, ws) -> dict:
    """7-配件采购记录 → part_purchases. header row 3."""
    from app.models.order import PartPurchase
    _, rows = _read_sheet(ws, header_row=3)
    inserted, skipped = 0, 0
    for r in rows:
        no = _str(r.get("采购单号"))
        if not no:
            skipped += 1; continue
        if db.execute(select(PartPurchase).where(
            PartPurchase.purchase_no == no,
        )).scalar_one_or_none():
            skipped += 1; continue
        db.add(PartPurchase(
            purchase_no=no,
            supplier=_str(r.get("供应商名称")),
            purchase_date=_date(r.get("采购日期")),
            material_code=_str(r.get("配件编码")),
            material_name=_str(r.get("配件名称")),
            spec=_str(r.get("规格型号")),
            qty=_dec(r.get("数量")) or Decimal("1"),
            unit_price=_dec(r.get("采购单价")),
            amount=_dec(r.get("金额")),
            tracking_no=_str(r.get("物流单号")),
            freight=_dec(r.get("运费")),
            total_amount=_dec(r.get("总金额")),
            payment_status=_str(r.get("支付状态")) or "unpaid",
        ))
        inserted += 1
    db.flush()
    return {"inserted": inserted, "skipped": skipped}


def import_balances(db, ws) -> dict:
    """10-账户余额汇总 → account_balances. header row 3."""
    from app.models.finance import AccountBalance
    _, rows = _read_sheet(ws, header_row=3)
    inserted, skipped = 0, 0
    for r in rows:
        account = _str(r.get("账户名称"))
        year = _int(r.get("年份")) or _int(r.get("年"))
        month = _int(r.get("月份")) or _int(r.get("月"))
        if not (account and year and month):
            skipped += 1; continue
        existing = db.execute(select(AccountBalance).where(
            AccountBalance.account_name == account,
            AccountBalance.period_year == year,
            AccountBalance.period_month == month,
        )).scalar_one_or_none()
        if existing:
            skipped += 1; continue
        db.add(AccountBalance(
            account_name=account, period_year=year, period_month=month,
            opening_balance=_dec(r.get("期初余额")) or Decimal("0"),
            income=_dec(r.get("收入")) or Decimal("0"),
            expense=_dec(r.get("支出")) or Decimal("0"),
            closing_balance=_dec(r.get("期末余额")) or Decimal("0"),
        ))
        inserted += 1
    db.flush()
    return {"inserted": inserted, "skipped": skipped}


# ----------------------------- 主流程 ---------------------------- #


# (sheet 名, importer 函数, 描述)
SHEET_HANDLERS: dict[str, tuple[Callable, str]] = {
    "1-产品总表":      (import_products, "产品 (487 行)"),
    "2-定价总表":      (import_pricing,  "定价快照 (占位文字→0, sku_code 唯一)"),
    "3b-配件价格表":   (import_materials, "物料定价"),
    "3-BOM表":         (import_bom, "BOM 物料用量 (重复行保留+标注)"),
    "4a-成品库存":     (import_product_inventory, "成品库存"),
    "4b-配件库存":     (import_part_inventory, "配件库存"),
    "5-订单总表":      (import_orders, "历史订单 (标 is_historical)"),
    "6-工厂下单表":    (import_factory_orders, "工厂下单"),
    "7-配件采购记录":  (import_part_purchases, "采购"),
    "10-账户余额汇总": (import_balances, "账户余额"),
}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", help="Excel 文件路径")
    parser.add_argument("--sheet", action="append",
                        help="只导某个 sheet (可多次), 默认全部")
    parser.add_argument("--dry-run", action="store_true",
                        help="不落库, 看 importer 报告")
    parser.add_argument("--no-historical", action="store_true",
                        help="订单不标 is_historical (默认会标)")
    args = parser.parse_args()

    wb = load_workbook(args.excel, read_only=True, data_only=True)
    print(f"打开 {args.excel}, {len(wb.sheetnames)} 个 sheet")
    print()

    targets = args.sheet or list(SHEET_HANDLERS.keys())
    db = SessionLocal()
    total = {"inserted": 0, "updated": 0, "skipped": 0}
    try:
        for sheet_name in targets:
            if sheet_name not in SHEET_HANDLERS:
                print(f"  [跳过] {sheet_name}: 不在支持列表")
                continue
            if sheet_name not in wb.sheetnames:
                print(f"  [跳过] {sheet_name}: 文件里没找到")
                continue
            handler, desc = SHEET_HANDLERS[sheet_name]
            print(f"  ▶ 导入 {sheet_name} — {desc} ...", end=" ", flush=True)
            try:
                if handler is import_orders:
                    result = handler(db, wb[sheet_name],
                                      mark_historical=not args.no_historical)
                else:
                    result = handler(db, wb[sheet_name])
                for k, v in result.items():
                    total[k] = total.get(k, 0) + v
                print(f"✓ {result}")
            except Exception as e:
                print(f"✗ 出错: {e}")
                raise
        if args.dry_run:
            print("\n[--dry-run] 回滚 (不落库)")
            db.rollback()
        else:
            db.commit()
            print("\n已落库.")
        print(f"\n总计: {total}")
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    main()

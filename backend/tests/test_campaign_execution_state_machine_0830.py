from __future__ import annotations

from datetime import datetime
import io

import openpyxl
import pytest

from app.models.campaign import CampaignPlan
from app.services import campaign_execution_service as execution
from app.services import campaign_recon_service as recon


def _plan(db, workflow_key: str | None = "campaign:test:one-shot"):
    plan = CampaignPlan(
        name="一次写入测试", workflow_key=workflow_key,
        campaign_type="big88", tier="big", status="draft",
        start_at=datetime(2026, 9, 1, 0, 0),
        end_at=datetime(2026, 9, 2, 0, 0),
        qn_campaign_title="测试活动",
    )
    db.add(plan)
    db.commit()
    return plan


def test_prewrite_connectivity_can_retry_but_claim_is_exact_once(db_session):
    plan = _plan(db_session)
    digest = "a" * 64
    attempt, created = execution.ensure_attempt(
        db_session, plan=plan, scope_sha256_value=digest)
    assert created is True and attempt.state == "prepared"

    execution.record_prewrite_failure(
        db_session, attempt, step="official_product_sku_export",
        error_code="Web-Agent 未在线", retryable=True)
    assert attempt.state == "retryable_prewrite"
    assert attempt.automatic_retry_allowed is True

    execution.claim_platform_write(
        db_session, attempt.id, request_id="request-one")
    assert attempt.write_claimed is True
    assert attempt.automatic_retry_allowed is False
    with pytest.raises(ValueError, match="already_claimed_no_retry"):
        execution.claim_platform_write(
            db_session, attempt.id, request_id="request-two")


def test_blocking_scope_error_can_never_be_claimed(db_session):
    plan = _plan(db_session)
    attempt, _ = execution.ensure_attempt(
        db_session, plan=plan, scope_sha256_value="b" * 64)
    execution.record_prewrite_failure(
        db_session, attempt, step="official_product_sku_identity",
        error_code="official_product_sku_scope_mismatch", retryable=False)
    with pytest.raises(ValueError, match="state_not_claimable"):
        execution.claim_platform_write(
            db_session, attempt.id, request_id="blocked")


def test_legacy_plan_gets_stable_identity_without_disabling_guard(db_session):
    plan = _plan(db_session, workflow_key=None)
    first, created = execution.ensure_attempt(
        db_session, plan=plan, scope_sha256_value="c" * 64)
    second, replay = execution.ensure_attempt(
        db_session, plan=plan, scope_sha256_value="c" * 64)
    assert created is True and replay is False
    assert first.id == second.id
    assert first.workflow_key == f"campaign:legacy-plan:{plan.id}"


def test_paused_enrolled_export_is_never_classified_as_draft():
    assert recon.classify_activity_record_status("暂停") == "enrolled_paused"
    assert recon.classify_activity_record_status("活动中") == "enrolled_active"
    assert recon.classify_activity_record_status("已发布设定") == "enrolled_scheduled"


def test_product_export_keeps_blank_merged_item_continuation_skus():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发布"
    for row in range(1, 4):
        ws.cell(row, 1, f"header-{row}")
    ws.cell(4, 1, "797294092429")
    ws.cell(4, 12, "6292834839399")
    ws.cell(4, 13, 1582.5)
    ws.cell(4, 14, 12)
    ws.cell(5, 12, "6292834839400")
    ws.cell(5, 13, 1410.0)
    ws.cell(5, 14, 0)
    data = io.BytesIO()
    wb.save(data)
    records = recon.parse_product_batch_export(data.getvalue())
    assert [(row["item_id"], row["sku_id"]) for row in records] == [
        ("797294092429", "6292834839399"),
        ("797294092429", "6292834839400"),
    ]
    assert [row["stock"] for row in records] == [12, 0]


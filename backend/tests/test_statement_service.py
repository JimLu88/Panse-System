"""月度对账表: 汇总数据 / Excel / HTML 渲染."""
from __future__ import annotations

import io
from datetime import date, datetime, timezone
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.models.supplier import DeliveryNote, DeliveryNoteLine, Supplier
from app.services import statement_service


def _mk_supplier(db, name="木作工厂", supplier_type="woodwork"):
    s = Supplier(name=name, supplier_type=supplier_type, payment_terms="月结")
    db.add(s)
    db.flush()
    return s


def _mk_note(db, supplier_id, *, note_no, delivery_date, total, status="pending_review",
             lines=()):
    n = DeliveryNote(
        supplier_id=supplier_id, note_no=note_no, delivery_date=delivery_date,
        total_amount=Decimal(str(total)) if total is not None else None, status=status,
    )
    db.add(n); db.flush()
    for i, (item, spec, qty, price, amt, matched, conf) in enumerate(lines, start=1):
        db.add(DeliveryNoteLine(
            delivery_note_id=n.id, line_no=i,
            item_name=item, spec=spec, unit="套",
            qty=Decimal(str(qty)),
            unit_price=Decimal(str(price)) if price else None,
            amount=Decimal(str(amt)) if amt else None,
            matched_order_no=matched, match_confidence=Decimal(str(conf)) if conf else None,
        ))
    db.flush()
    return n


def test_build_data_empty_month(db_session):
    s = _mk_supplier(db_session)
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    assert data.note_count == 0
    assert data.total_amount == Decimal("0")
    assert data.rows == []


def test_build_data_aggregates_by_month(db_session):
    s = _mk_supplier(db_session)
    _mk_note(db_session, s.id, note_no="N1", delivery_date=date(2026, 5, 14),
             total=900, lines=[
                 ("电视柜", "1800×850", 1, 580, 580, "O1", 95),
                 ("斗柜", "800×900", 1, 320, 320, None, None),
             ])
    _mk_note(db_session, s.id, note_no="N2", delivery_date=date(2026, 5, 28),
             total=1200, status="paid", lines=[
                 ("床头柜", "500×400", 2, 600, 1200, "O2", 100),
             ])
    # 不应被包括 — 不同月
    _mk_note(db_session, s.id, note_no="N3", delivery_date=date(2026, 4, 30), total=999, lines=[])

    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    assert data.note_count == 2
    assert data.total_amount == Decimal("2100")
    assert data.paid_amount == Decimal("1200")
    assert data.unpaid_amount == Decimal("900")
    assert len(data.rows) == 3
    assert {r.note_no for r in data.rows} == {"N1", "N2"}


def test_build_data_supplier_not_found(db_session):
    with pytest.raises(ValueError):
        statement_service.build_statement_data(
            db_session, supplier_id=99999, year=2026, month=5,
        )


def test_build_data_december_handles_year_rollover(db_session):
    s = _mk_supplier(db_session)
    _mk_note(db_session, s.id, note_no="DEC", delivery_date=date(2026, 12, 15), total=100, lines=[])
    _mk_note(db_session, s.id, note_no="JAN", delivery_date=date(2027, 1, 1), total=200, lines=[])
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=12,
    )
    assert data.note_count == 1
    assert {r.note_no for r in data.rows} == {"DEC"}


def test_build_data_warns_on_missing_note_no(db_session):
    s = _mk_supplier(db_session)
    _mk_note(db_session, s.id, note_no=None, delivery_date=date(2026, 5, 14),
             total=100, lines=[("x", "", 1, 100, 100, None, None)])
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    assert any("缺少单号" in w for w in data.warnings)


def test_build_data_empty_note_shows_placeholder_row(db_session):
    s = _mk_supplier(db_session)
    _mk_note(db_session, s.id, note_no="N0", delivery_date=date(2026, 5, 1), total=50, lines=[])
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    assert len(data.rows) == 1
    assert data.rows[0].item_name == "(空单据)"


def test_render_excel_valid_xlsx(db_session):
    s = _mk_supplier(db_session)
    _mk_note(db_session, s.id, note_no="N1", delivery_date=date(2026, 5, 14),
             total=580, lines=[("电视柜", "1800×850", 1, 580, 580, "O1", 95)])
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    xlsx_bytes = statement_service.render_excel(data)
    assert xlsx_bytes[:2] == b"PK"  # ZIP magic

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.title == "2026-05"
    assert "木作工厂" in ws["A1"].value
    assert ws.cell(row=2, column=1).value == "单据号"
    assert ws.cell(row=3, column=1).value == "N1"
    assert ws.cell(row=3, column=11).value == "O1 (95%)"


def test_render_excel_handles_empty(db_session):
    s = _mk_supplier(db_session)
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    xlsx = statement_service.render_excel(data)
    assert len(xlsx) > 1000


def test_render_html_contains_key_fields(db_session):
    s = _mk_supplier(db_session, name="岩板厂", supplier_type="rock_slab")
    _mk_note(db_session, s.id, note_no="2018901", delivery_date=date(2026, 5, 14),
             total=580, status="paid", lines=[
                 ("台面板", "1620×3290×10", 1, 580, 580, "O1", 95),
             ])
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    html = statement_service.render_html(data)
    assert "<!doctype html>" in html
    assert "岩板厂" in html
    assert "2018901" in html
    assert "1620×3290×10" in html
    assert "已付款" in html
    assert "window.print()" in html


def test_render_html_escapes_special_chars(db_session):
    s = _mk_supplier(db_session)
    _mk_note(db_session, s.id, note_no="<script>", delivery_date=date(2026, 5, 1),
             total=100, lines=[("<b>x</b>", "", 1, 100, 100, None, None)])
    data = statement_service.build_statement_data(
        db_session, supplier_id=s.id, year=2026, month=5,
    )
    html = statement_service.render_html(data)
    assert "<script>" not in html  # 已被转义
    assert "&lt;script&gt;" in html
    assert "&lt;b&gt;x&lt;/b&gt;" in html

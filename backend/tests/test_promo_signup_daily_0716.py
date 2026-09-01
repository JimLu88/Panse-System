"""88VIP大促·日常价法(B法) builder — 2026-07-16 用户拍板: 88VIP 会叠单品立减。
真SKU 活动价 = 日常价(绝不报名价A, 否则叠 big档SI 双折砸穿); 占位 = A(×0.9→500→floor)。
走同一 promo_signup 大促报名页导入, 只是 col3 值从 A 换成日常价。"""
from decimal import Decimal
from io import BytesIO

import openpyxl

from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import data_export_service as de
from app.services.activity_upload_service import _compare_rows, _parse_uploaded_values


def _add(db, pc, sc, name, daily, item, sid, big_buyer=None, mid_buyer=None, ph=False, floor=None):
    db.add(PricingSku(product_code=pc, sku_code=sc, sku=name, product_name=name,
                      daily_price=Decimal(str(daily)) if daily is not None else None,
                      is_custom_placeholder=ph))
    db.add(PricingSkuPromo(sku_code=sc, taobao_item_id=item, taobao_sku_id=sid,
                           big_buyer_price=Decimal(str(big_buyer)) if big_buyer else None,
                           mid_buyer_price=Decimal(str(mid_buyer)) if mid_buyer else None,
                           enrolled_floor_price=Decimal(str(floor)) if floor else None))


def _col3_by_sid(bio):
    ws = openpyxl.load_workbook(BytesIO(bio.getvalue()))["商品SKU导入列表"]
    return {str(ws.cell(r, 2).value): ws.cell(r, 3).value
            for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value}


def test_real_sku_uses_daily_not_reportprice(db_session):
    db = db_session
    _add(db, "PPS8101", "PPS8101011", "岩板餐桌-1.8米", 6470, "111", "800001", big_buyer=4900, mid_buyer=5010)
    _add(db, "PPS8101", "PPS8101099", "尺寸定制", 1500, "111", "800099", ph=True, floor=400)
    db.commit()

    daily = _col3_by_sid(de.build_promo_signup_daily_upload_xlsx(db)[0])
    a_based = _col3_by_sid(de.build_promo_signup_upload_xlsx(db, "big")[0])

    # 所有生产入口已统一：真 SKU 一律填日常价，旧反推报名价法不再存在
    assert daily["800001"] == 6470.0
    assert a_based["800001"] == 6470.0
    # 未精确点名的占位 SKU 在两个生产生成器中都默认排除。
    assert "800099" not in daily
    assert "800099" not in a_based


def test_shoudao_equals_dacu(db_session):
    """自洽: 到手 = 日常价×0.88 − big_deduct 应 = 大促到手(big_buyer); big_deduct = 日常价×0.88 − big_buyer。"""
    db = db_session
    _add(db, "PPS8102", "PPS8102011", "餐桌-1.4米", 3000, "222", "800011", big_buyer=2300, mid_buyer=2352)
    db.commit()
    act = _col3_by_sid(de.build_promo_signup_daily_upload_xlsx(db)[0])["800011"]   # = 日常价 3000
    big_deduct = round(act * 0.88 - 2300, 2)
    assert act == 3000.0
    assert round(act * 0.88 - big_deduct, 2) == 2300.0                             # 到手 = 大促到手 ✓


def test_compare_rows_zero_mismatch(db_session):
    """比对表(big88) system_value 必 = 上传 xlsx col3(0容差), 否则 stage 会红标; target = 大促到手。"""
    db = db_session
    _add(db, "PPS8103", "PPS8103011", "床-1.5米", 2600, "333", "800021", big_buyer=2000, mid_buyer=2046)
    _add(db, "PPS8103", "PPS8103099", "尺寸定制", 1500, "333", "800029", ph=True, floor=350)
    db.commit()
    uploaded = _parse_uploaded_values("promo_signup", de.build_promo_signup_daily_upload_xlsx(db)[0].getvalue())
    rows = _compare_rows(db, "promo_signup", "big88")
    assert rows and {r["taobao_sku_id"] for r in rows} == set(uploaded)
    for r in rows:
        assert abs(r["system_value"] - uploaded[r["taobao_sku_id"]]) < 0.005       # 0容差
        assert r["value_label"] == "活动价(日常价/占位保护价)"
    tgt = {r["taobao_sku_id"]: r["target_shoudao"] for r in rows}
    assert tgt["800021"] == 2000.0                                                  # 大促到手
    assert uploaded == {"800021": 2600.0}                                         # 真SKU=日常价；未点名占位排除

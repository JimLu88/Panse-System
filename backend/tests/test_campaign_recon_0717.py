"""活动核对器测试 (2026-07-17 spec §七 三种导出格式 + §四.6 核对维度)。

锁六件事:
① 活动商品导出解析: 跳3行表头 / 多营销ID只认"已发布设定" / J列券后+P列活动价列位
② 单品立减导出(1行表头) + 商品批量导出(发布模板sheet, 跳3行, reset_dimensions) 解析
③ verify_campaign_title: 空白归一精确比对, 空值诚实拒绝
④ 逐SKU判定: 一分不差 / 贴线让X / 超2元报警 / 占位 / 无映射 + 覆盖完整性缺失/多出
⑤ >2元报警: notify_service.broadcast_text 触发 (mock) + CampaignReconReport 落库 + 状态机 alarmed
⑥ 活动名称不一致 → 报警 + alarmed (防推错活动)
"""
import io
from datetime import datetime
from decimal import Decimal

import openpyxl

from app.models.campaign import CampaignPlan, CampaignReconReport
from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import campaign_recon_service as crs
from app.services import campaign_service as cs


def _mk(db, pc, code, item, sid, daily=None, big=None, *, placeholder=False, line=None):
    db.add(PricingSku(product_code=pc, sku_code=code, sku=f"SKU{code}",
                      product_name=f"品{pc}",
                      daily_price=Decimal(str(daily)) if daily is not None else None,
                      is_custom_placeholder=placeholder))
    db.add(PricingSkuPromo(
        sku_code=code, taobao_item_id=item, taobao_sku_id=sid,
        big_buyer_price=Decimal(str(big)) if big is not None else None,
        coupon_floor_price=Decimal(str(line)) if line is not None else None))


def _plan(db, title=None):
    plan = CampaignPlan(name="88VIP测试场", campaign_type="big88", tier="big",
                        start_at=datetime(2026, 7, 17, 20, 0, 0),
                        end_at=datetime(2026, 7, 19, 23, 59, 59),
                        qn_campaign_title=title, status="draft")
    db.add(plan)
    db.commit()
    return plan


def _xlsx(rows, header_rows=3, sheet_title=None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    if sheet_title:
        ws.title = sheet_title
    for r in range(1, header_rows + 1):
        ws.cell(r, 1, f"表头{r}")
    r = header_rows + 1
    for row in rows:
        for ci, v in enumerate(row, start=1):
            ws.cell(r, ci, v)
        r += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _act_row(item, sid, coupon_after, act_price=None, status="已发布设定", mkt="M1"):
    """活动商品导出行 A..T (20列): J(idx9)=活动普惠券后价, P(idx15)=活动价。"""
    return [item, f"品{item}", mkt, status, sid, f"SKU{sid}", None, None, None,
            coupon_after, None, None, None, None, None, act_price, None, None, None, None]


def _disc_row(item, sid, value, aname="88VIP周期购活动"):
    return ["A001", aname, "SKU级别", "2026-07-17 20:00:00", "2026-07-19 23:59:59",
            "进行中", item, f"品{item}", sid, f"SKU{sid}", None, value]


# ── ①② 解析 ─────────────────────────────────────────────────────────────────

def test_parse_activity_export_only_published(db_session):
    data = _xlsx([
        _act_row("9601", "81001", 1979.59, act_price=2827.5),
        _act_row("9601", "81001", 1888.0, status="已结束", mkt="M0"),   # 旧营销ID → 不认
        _act_row("9601", "81002", 1599.5, act_price=2000.0),
    ], header_rows=3)
    records = crs.parse_activity_items_export(data)
    assert len(records) == 2                                  # 只认"已发布设定"
    assert records[0]["sku_id"] == "81001"
    assert records[0]["coupon_after"] == 1979.59              # J列
    assert records[0]["activity_price"] == 2827.5             # P列
    assert records[0]["marketing_id"] == "M1"

    floor_records = crs.parse_activity_floor_evidence_export(data)
    assert len(floor_records) == 3
    assert floor_records[1]["status"] == "已结束"
    assert floor_records[1]["min_list_price"] is None


def test_super_reduce_future_export_can_include_paused_enrollment():
    data = _xlsx([
        _act_row("1000009601", "81011", None, act_price=1200.0, status="暂停"),
        _act_row("1000009602", "81012", None, act_price=1300.0, status="撤销报名"),
    ], header_rows=3)

    assert crs.parse_activity_items_export(data) == []
    records = crs.parse_activity_items_export(data, include_paused=True)
    assert [(row["item_id"], row["sku_id"], row["status"]) for row in records] == [
        ("1000009601", "81011", "暂停")]


def test_parse_discount_and_product_exports(db_session):
    disc = _xlsx([_disc_row("9601", "81001", 507.91)], header_rows=1)
    drecords = crs.parse_discount_export(disc)
    assert drecords[0]["activity_name"] == "88VIP周期购活动"
    assert drecords[0]["sku_id"] == "81001"
    assert drecords[0]["discount_value"] == 507.91

    prod = _xlsx([["9601", None, None, None, 3770.0, None, None, None, None,
                   "1.2米原木色", None, "81001", 3770.0, None, None, "PPSRA00101"]],
                 header_rows=3, sheet_title="发布模板")
    precords = crs.parse_product_batch_export(prod)
    assert precords == [{"item_id": "9601", "item_list_price": 3770.0,
                         "sale_attr": "1.2米原木色", "sku_id": "81001",
                         "sku_price": 3770.0, "merchant_code": "PPSRA00101"}]


# ── ③ 活动名称校验 ───────────────────────────────────────────────────────────

def test_verify_campaign_title():
    assert crs.verify_campaign_title("88VIP周期购活动", "88VIP周期购活动") is True
    assert crs.verify_campaign_title("88VIP 周期购活动", "88VIP周期购 活动") is True   # 空白归一
    assert crs.verify_campaign_title("88VIP周期购活动", "超级立减7月") is False
    assert crs.verify_campaign_title(None, "88VIP周期购活动") is False
    assert crs.verify_campaign_title("88VIP周期购活动", "") is False


# ── ④⑤ 逐SKU判定 + 报警 + 落库 ──────────────────────────────────────────────

def test_reconcile_verdicts_alarm_and_report(db_session, monkeypatch):
    plan = _plan(db_session, title="88VIP周期购活动")
    _mk(db_session, "PPSRA001", "PPSRA00101", "9601", "81001", daily=2827.5, big=1979.59)
    _mk(db_session, "PPSRB001", "PPSRB00101", "9602", "81002", daily=2000, big=1600, line=1599.5)
    _mk(db_session, "PPSRC001", "PPSRC00101", "9603", "81003", daily=1000, big=700)
    _mk(db_session, "PPSRD001", "PPSRD00101", "9604", "81004", daily=1200, big=900)   # 导出缺失
    _mk(db_session, "PPSRE001", "PPSRE00190", "9605", "81009", daily=400, placeholder=True)
    db_session.commit()

    alarms_sent = []
    from app.services import notify_service
    monkeypatch.setattr(notify_service, "broadcast_text",
                        lambda db, text, **kw: alarms_sent.append({"text": text, **kw}) or {})

    activity = _xlsx([
        _act_row("9601", "81001", 1979.59, act_price=2827.5),   # 一分不差 + 活动价=日常价
        _act_row("9602", "81002", 1599.5),                      # 贴线让0.50
        _act_row("9603", "81003", 690.0),                       # 差 −10 → 超2元报警
        _act_row("9605", "81009", 360.0, act_price=360.0),      # 占位保护价一致
        _act_row("9699", "99999", 100.0),                       # 无映射
    ])
    discount = _xlsx([
        _disc_row("9601", "81001", 507.91),                     # = builder 应填 (2827.5−340−1979.59)
        _disc_row("9602", "81002", 100.0),                      # ≠ builder 应填 160.5 → 出入
    ], header_rows=1)

    result = crs.reconcile(db_session, plan, activity_bytes=activity, discount_bytes=discount)

    assert result["ok"] is True
    by_sid = {r["sku_id"]: r for r in result["rows"]}
    assert by_sid["81001"]["verdict"] == "一分不差"
    assert by_sid["81001"]["signup_price_ok"] is True           # b维度: 活动价P列 vs 日常价
    assert by_sid["81002"]["verdict"] == "贴线让0.50"
    assert by_sid["81003"]["verdict"] == "超2元报警" and by_sid["81003"]["diff"] == -10.0
    assert by_sid["81009"]["verdict"] == "占位价一致"
    assert by_sid["99999"]["verdict"] == "无映射"

    summary = result["summary"]
    assert summary["verdicts"] == {"一分不差": 1, "贴线": 1, "超2元报警": 1,
                                   "占位价一致": 1, "无映射": 1}
    assert summary["coverage_missing"] == ["81004"]             # d维度: 应报未报
    assert summary["coverage_extra"] == ["99999"]               # 多出的也报
    assert [m["sku_id"] for m in summary["discount_mismatch"]] == [
        "81002", "81003", "81004"]                              # 错价 + 应有但导出缺失
    assert summary["title_ok"] is True

    assert result["alarm_count"] == 7
    assert len(alarms_sent) == 1 and alarms_sent[0]["title"] == "活动核对报警"
    assert "PPSRC00101" in alarms_sent[0]["text"]

    report = db_session.query(CampaignReconReport).filter_by(plan_id=plan.id).one()
    assert report.alarm_count == 7 and report.summary["alarm"] == 7
    assert plan.status == "alarmed"


def test_reconcile_clean_run_marks_reconciled(db_session, monkeypatch):
    plan = _plan(db_session, title="88VIP周期购活动")
    _mk(db_session, "PPSRF001", "PPSRF00101", "9611", "82001", daily=2827.5, big=1979.59)
    db_session.commit()
    called = []
    from app.services import notify_service
    monkeypatch.setattr(notify_service, "broadcast_text",
                        lambda db, text, **kw: called.append(text) or {})

    activity = _xlsx([_act_row("9611", "82001", 1979.59, act_price=2827.5)])
    result = crs.reconcile(db_session, plan, activity_bytes=activity)

    assert result["alarm_count"] == 0 and called == []          # 干净 → 不报警
    assert plan.status == "reconciled"
    assert result["summary"]["coverage_missing"] == []


def test_reconcile_ignores_historical_items_outside_platform_scope(
        db_session, monkeypatch):
    plan = _plan(db_session, title="88VIP周期购活动")
    _mk(db_session, "PPSRSC01", "PPSRSC0101", "1000009612", "82012",
        daily=2827.5, big=1979.59)
    _mk(db_session, "PPSRSH01", "PPSRSH0101", "1000009698", "82998",
        daily=3000, big=2100)
    db_session.commit()
    cs._set_plan_item_marker(plan, "platform_qualified_items", {"1000009612"})
    cs._set_plan_item_marker(plan, "official_active_items", {"1000009612"})
    db_session.commit()
    assert cs.platform_scope_present(plan), plan.remark
    called = []
    from app.services import notify_service
    monkeypatch.setattr(notify_service, "broadcast_text",
                        lambda db, text, **kw: called.append(text) or {})

    activity = _xlsx([
        _act_row("1000009612", "82012", 1979.59, act_price=2827.5),
        _act_row("1000009698", "82998", 1.0, act_price=1.0),
    ])

    result = crs.reconcile(db_session, plan, activity_bytes=activity)

    assert result["verified"] is True, result["summary"]
    assert result["alarm_count"] == 0
    assert [row["sku_id"] for row in result["rows"]] == ["82012"]
    assert result["summary"]["ignored_out_of_scope_records"] == 1
    assert result["summary"]["ignored_out_of_scope_items"] == ["1000009698"]
    assert called == []


def test_super_reconcile_treats_future_paused_rows_as_pending_not_missing(
        db_session, monkeypatch):
    plan = _plan(db_session, title="超级立减长期活动")
    plan.campaign_type = "super_reduce"
    plan.tier = "mid"
    _mk(db_session, "PPSSRF01", "PPSSRF0101", "1000009622", "82022",
        daily=1200, big=800)
    db_session.commit()
    cs._set_plan_item_marker(plan, "platform_qualified_items", {"1000009622"})
    cs._set_plan_item_marker(plan, "official_active_items", {"1000009622"})
    db_session.commit()
    called = []
    from app.services import notify_service
    monkeypatch.setattr(notify_service, "broadcast_text",
                        lambda db, text, **kw: called.append(text) or {})

    activity = _xlsx([
        _act_row("1000009622", "82022", None, act_price=1200.0, status="暂停"),
    ])
    result = crs.reconcile(db_session, plan, activity_bytes=activity)

    assert result["ok"] is True
    assert result["pending"] is True
    assert result["alarm_count"] == 0
    assert result["summary"]["coverage_missing"] == []
    assert result["summary"]["pending_coupon_after"] == 1
    assert plan.status == "signup_pushed"
    assert called == []


def test_reconcile_preserves_platform_accepted_active_placeholder_price(
        db_session, monkeypatch):
    plan = _plan(db_session, title="超级立减长期活动")
    plan.campaign_type = "super_reduce"
    plan.tier = "mid"
    _mk(db_session, "PPSPLACE2", "PPSPLACE299", "1000009623", "82023",
        daily=1000, big=800, placeholder=True, line=300)
    db_session.commit()
    cs._set_plan_item_marker(plan, "platform_qualified_items", {"1000009623"})
    cs._set_plan_item_marker(plan, "official_active_items", {"1000009623"})
    db_session.commit()
    called = []
    from app.services import notify_service
    monkeypatch.setattr(notify_service, "broadcast_text",
                        lambda db, text, **kw: called.append(text) or {})

    activity = _xlsx([
        _act_row("1000009623", "82023", None, act_price=397.0, status="活动中"),
    ])
    result = crs.reconcile(db_session, plan, activity_bytes=activity)

    assert result["alarm_count"] == 0
    assert result["pending"] is False
    assert result["rows"][0]["verdict"] == "占位在场价沿用"
    assert result["rows"][0]["grandfathered_activity_price"] is True
    assert result["rows"][0]["signup_price_ok"] is None
    assert plan.status == "reconciled"
    assert called == []


def test_reconcile_does_not_grandfather_future_paused_placeholder_mismatch(
        db_session, monkeypatch):
    plan = _plan(db_session, title="超级立减长期活动")
    plan.campaign_type = "super_reduce"
    plan.tier = "mid"
    _mk(db_session, "PPSPLACE3", "PPSPLACE399", "1000009624", "82024",
        daily=1000, big=800, placeholder=True, line=300)
    db_session.commit()
    cs._set_plan_item_marker(plan, "platform_qualified_items", {"1000009624"})
    cs._set_plan_item_marker(plan, "official_active_items", {"1000009624"})
    db_session.commit()
    monkeypatch.setattr(
        __import__("app.services.notify_service", fromlist=["broadcast_text"]),
        "broadcast_text", lambda *args, **kwargs: {})

    activity = _xlsx([
        _act_row("1000009624", "82024", None, act_price=397.0, status="暂停"),
    ])
    result = crs.reconcile(db_session, plan, activity_bytes=activity)

    assert result["alarm_count"] == 1
    assert result["rows"][0]["verdict"] == "占位活动价不一致"
    assert result["rows"][0]["signup_price_ok"] is False
    assert plan.status == "alarmed"


def test_reconcile_requires_at_least_one_file(db_session):
    plan = _plan(db_session)
    result = crs.reconcile(db_session, plan)
    assert result["ok"] is False and "导出" in result["error"]


# ── ⑥ 活动名称不一致 → 报警 (防推错活动) ─────────────────────────────────────

def test_reconcile_title_mismatch_alarms(db_session, monkeypatch):
    plan = _plan(db_session, title="88VIP周期购活动")
    _mk(db_session, "PPSRG001", "PPSRG00101", "9621", "83001", daily=2827.5, big=1979.59)
    db_session.commit()
    called = []
    from app.services import notify_service
    monkeypatch.setattr(notify_service, "broadcast_text",
                        lambda db, text, **kw: called.append(text) or {})

    discount = _xlsx([_disc_row("9621", "83001", 507.91, aname="超级立减7月")], header_rows=1)
    result = crs.reconcile(db_session, plan, discount_bytes=discount)

    assert result["summary"]["title_ok"] is False
    assert plan.status == "alarmed"                             # 疑推错活动 → 报警态
    assert len(called) == 1 and "活动名称与计划不一致" in called[0]

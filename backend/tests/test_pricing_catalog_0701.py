"""定价图册 (带图导出, 2026-07-01): 一产品一行 + 售价; 图库文件在测试里 stub 掉。"""
from decimal import Decimal

import app.services.pricing_catalog_service as pcs
from app.models.pricing import PricingSku
from app.models.product import Product
from app.services.pricing_catalog_service import build_catalog_html


def test_catalog_has_product_and_prices(db_session, monkeypatch):
    monkeypatch.setattr(pcs, "_img_data_uri", lambda url: None)   # 不碰真实图库文件
    db_session.add(Product(code="P1", name="测试岩板餐桌"))
    db_session.add(PricingSku(product_code="P1", sku_code="P1-A", sku="1.2米",
                              list_price=Decimal("6880"), big_promo=Decimal("3060")))
    db_session.commit()
    out = build_catalog_html(db_session)
    assert "定价图册" in out
    assert "测试岩板餐桌" in out and "P1" in out
    assert "¥3,060" in out and "¥6,880" in out
    assert "暂无图片" in out                    # 无图 → 占位框


def test_catalog_limit_and_skips_unpriced(db_session, monkeypatch):
    monkeypatch.setattr(pcs, "_img_data_uri", lambda url: None)
    for i in range(4):
        db_session.add(Product(code=f"P{i}", name=f"产品{i}"))
        db_session.add(PricingSku(product_code=f"P{i}", sku_code=f"P{i}-A", big_promo=Decimal("100")))
    db_session.add(Product(code="NOPRICE", name="无定价产品"))   # 无 SKU → 不进图册
    db_session.commit()
    assert build_catalog_html(db_session).count('class="card"') == 4      # NOPRICE 被排除
    assert build_catalog_html(db_session, limit=2).count('class="card"') == 2


def test_catalog_skips_non_product_skus(db_session, monkeypatch):
    monkeypatch.setattr(pcs, "_img_data_uri", lambda url: None)
    db_session.add(Product(code="REAL1", name="孚格蜂蜜餐桌"))
    db_session.add(PricingSku(product_code="REAL1", sku_code="REAL1-A", big_promo=Decimal("1830")))
    for code, nm in [("V1", "作废洞石大柜链接"), ("S1", "商家安装sku"),
                     ("C1", "全屋定制"), ("PPS99999999999", "纯定制")]:
        db_session.add(Product(code=code, name=nm))
        db_session.add(PricingSku(product_code=code, sku_code=f"{code}-A", big_promo=Decimal("1")))
    db_session.commit()
    out = build_catalog_html(db_session)
    assert out.count('class="card"') == 1        # 只剩真实产品
    assert "孚格蜂蜜餐桌" in out
    for junk in ("作废洞石大柜链接", "商家安装sku", "全屋定制", "纯定制"):
        assert junk not in out


def test_catalog_xlsx_structure(db_session, monkeypatch):
    """带图 Excel: 中文表头 + 分类色带 + 同编码多 SKU 图列合并; 无图退占位; 作废剔除。"""
    import io
    import openpyxl
    from app.services.data_export_service import build_catalog_xlsx
    monkeypatch.setattr(pcs, "product_image_map", lambda codes, url_by_code, **k: {})  # 不碰真实图
    db_session.add(Product(code="P1", name="测试岩板餐桌"))
    db_session.add(PricingSku(product_code="P1", sku_code="P1-A", sku="1.2米",
                              list_price=Decimal("6880"), big_promo=Decimal("3060")))
    db_session.add(PricingSku(product_code="P1", sku_code="P1-B", sku="1.4米",
                              list_price=Decimal("7880"), big_promo=Decimal("3560")))
    db_session.add(Product(code="V1", name="作废旧链接"))            # 应被剔除
    db_session.add(PricingSku(product_code="V1", sku_code="V1-A", big_promo=Decimal("1")))
    db_session.commit()

    wb = openpyxl.load_workbook(io.BytesIO(build_catalog_xlsx(db_session).getvalue()))
    ws = wb["定价图册"]
    assert ws.cell(1, 1).value == "产品图"                          # 图列表头
    row1 = [c.value for c in ws[1] if c.value]
    assert "售价档位" in row1 and "标识" in row1                     # 分类色带
    hdr = {c.value: c.column for c in ws[2] if c.value}
    for h in ("产品编码", "产品名称", "标价", "大促价", "淘宝标题", "小红书标价"):
        assert h in hdr
    pc = hdr["产品编码"]
    codes = [ws.cell(r, pc).value for r in range(3, ws.max_row + 1)]
    assert codes == ["P1", "P1"]                                    # 作废剔除, P1 两个SKU
    assert "A3:A4" in {str(m) for m in ws.merged_cells.ranges}      # 图列纵向合并
    assert ws.cell(3, 1).value == "暂无图片"                         # 无图占位
    vals = {ws.cell(r, c).value for r in range(3, ws.max_row + 1) for c in range(1, ws.max_column + 1)}
    assert 3060 in vals and 6880 in vals                            # 售价落格


def test_catalog_xlsx_has_live_formulas(db_session, monkeypatch):
    """图册派生列(物理/各档价/平台费/税/会计/利润/毛利率)必须是活公式, 与定价总表导出同口径。"""
    import io
    import openpyxl
    from app.services.data_export_service import build_catalog_xlsx
    monkeypatch.setattr(pcs, "product_image_map", lambda codes, url_by_code, **k: {})
    db_session.add(Product(code="F1", name="公式测试桌"))
    db_session.add(PricingSku(
        product_code="F1", sku_code="F1-A", sku="1.2米",
        factory_cost=Decimal("1000"), logistics_cost=Decimal("100"), install_cost=Decimal("50"),
        base_list=Decimal("0.3"), base_small=Decimal("0.55"),
        base_mid=Decimal("0.6"), base_big=Decimal("0.62")))
    db_session.commit()

    wb = openpyxl.load_workbook(io.BytesIO(build_catalog_xlsx(db_session).getvalue()))
    ws = wb["定价图册"]
    hdr = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    def f(name):
        return str(ws.cell(3, hdr[name]).value or "")
    assert f("物理总成本").startswith("=SUM")        # 物理 = 工厂+物流+安装
    assert f("标价").startswith("=ROUNDUP")
    assert f("大促价").startswith("=ROUNDUP")
    assert f("日常价/单品宝").startswith("=")
    assert f("会计总成本").startswith("=")
    assert f("大促利润").startswith("=")
    assert f("毛利率").startswith("=")               # 一个都不能少


def test_catalog_xlsx_accessory_detail_and_promo_formulas(db_session, monkeypatch):
    """配件成本明细列 + 配件→外配件→工厂 活公式 + 平台活动价活公式(淘宝/店内/中促/大促/小红书)。"""
    import io
    import openpyxl
    from app.models.pricing_ext import PricingSkuCosts, PricingSkuPromo
    from app.services.data_export_service import build_catalog_xlsx
    monkeypatch.setattr(pcs, "product_image_map", lambda codes, url_by_code, **k: {})
    db_session.add(Product(code="G1", name="配件促销测试"))
    db_session.add(PricingSku(product_code="G1", sku_code="G1-A", sku="1.2米",
                              wood_cost=Decimal("500"), packaging_cost=Decimal("50"),
                              daily_price=Decimal("1000"),
                              small_promo=Decimal("680"), mid_promo=Decimal("678"),  # 店铺实收(锚)
                              big_promo=Decimal("647")))
    db_session.add(PricingSkuCosts(sku_code="G1-A", rock_slab=Decimal("200"), glass=Decimal("100")))
    db_session.add(PricingSkuPromo(
        sku_code="G1-A",
        mid_platform_discount=Decimal("0.12"), mid_vip_commission=Decimal("0.01"),
        big_platform_discount=Decimal("0.12"), big_vip_commission=Decimal("0"),
        xhs_activity_price=Decimal("900"), xhs_promo_discount=Decimal("0.15")))
    db_session.commit()

    wb = openpyxl.load_workbook(io.BytesIO(build_catalog_xlsx(db_session).getvalue()))
    ws = wb["定价图册"]
    hdr = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    def f(name):
        return str(ws.cell(3, hdr[name]).value or "")
    # 配件成本明细列都在
    for acc in ("岩板", "玻璃", "灯带", "腿部", "软包", "床铺板", "配件备注"):
        assert acc in hdr, f"缺配件列: {acc}"
    assert ws.cell(3, hdr["岩板"]).value == 200      # 配件值落格
    # 配件 → 外配件 → 工厂 活公式
    assert f("外采配件成本合计").startswith("=SUM")
    assert f("总出厂成本").startswith("=")
    # 单品立减改加法口径(折 + 降价金额, 中促/大促/超大促), 不再输出会误导的乘法系数
    assert "大促店铺系数" not in hdr and "中促店铺系数" not in hdr and "单品立减系数" not in hdr
    for h in ("中促单品立减(折)", "中促降价金额(元)", "大促单品立减(折)", "大促降价金额(元)",
              "超大促单品立减(折)", "超大促降价金额(元)"):
        assert h in hdr, f"图册缺列: {h}"
    assert f("中促买家价").startswith("=")                          # 买家到手 = 中促价/(1−佣金), 派生
    assert f("中促店铺到手").startswith("=")                        # 店铺到手 = 中促价(实收)
    assert "IF" in f("中促VIP到手价")                               # 88VIP 消费券阶梯 (由买家到手)
    assert f("小红书促销价").startswith("=")


def test_signup_form_xlsx_additive_and_stripped(db_session, monkeypatch):
    """活动报名表: 报名价 + 单品立减(折/立减金额, 加法口径)正确; 无关列(成本/编码/小红书/乘法系数)已去掉。
    附图口径: 日常19575, 大促价(店铺实收)12890, 佣金2% → 大促买家价13153 → 大促单品立减 7.92折 / 减4073元。"""
    import io
    import openpyxl
    from decimal import Decimal as D
    from app.models.pricing_ext import PricingSkuPromo
    from app.services import pricing_calc_service as pc
    from app.services.data_export_service import build_signup_form_xlsx
    monkeypatch.setattr(pcs, "product_image_map", lambda codes, url_by_code, **k: {})
    sku = PricingSku(product_code="P1", sku_code="P1-A", sku="1.2米",
                     list_price=D("26100"), daily_price=D("19575"), big_promo=D("12890"))
    promo = PricingSkuPromo(sku_code="P1-A")
    db_session.add(Product(code="P1", name="曜黑餐边柜"))
    db_session.add(sku); db_session.add(promo)
    db_session.commit()
    pc.recompute_promo(promo, sku, {"big_vip_commission": D("0.02"), "mid_vip_commission": D("0.02")})
    db_session.commit()

    wb = openpyxl.load_workbook(io.BytesIO(build_signup_form_xlsx(db_session).getvalue()))
    ws = wb["活动报名表"]
    hdr = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    # 必要列在
    for h in ("产品图", "产品名称", "规格", "一口价", "日常价(活动价)", "大促到手",
              "88VIP大促报名价", "超大促报名价(618/双11)",
              "大促单品立减(折)", "大促立减(元)", "超大促单品立减(折)", "超大促立减(元)"):
        assert h in hdr, f"缺列: {h}"
    # 无关列已去掉
    for junk in ("ID", "淘宝标题", "SKU编码", "产品编码", "小促价", "会计总成本",
                 "物理总成本", "小红书标价", "岩板", "大促店铺系数", "大促基数"):
        assert junk not in hdr, f"应去掉却还在: {junk}"
    # 数值 (附图核准): 大促单品立减 = 7.92 折, 立减 = 4072.94 元
    assert abs(float(ws.cell(3, hdr["大促单品立减(折)"]).value) - 7.92) < 0.02
    assert abs(float(ws.cell(3, hdr["大促立减(元)"]).value) - 4072.94) < 1.5
    # 超大促(618, 15%) 比大促(12%)浅 3 个点 → 8.22 折
    assert abs(float(ws.cell(3, hdr["超大促单品立减(折)"]).value) - 8.22) < 0.02


def test_single_item_discount_upload_xlsx(db_session):
    """淘宝单品立减批量上传表: 表头逐字对齐模板; SKU级别减钱(立减金额); 缺SKU_ID跳过; 618金额<大促。"""
    import io
    import openpyxl
    from decimal import Decimal as D
    from app.models.pricing_ext import PricingSkuPromo
    from app.services import pricing_calc_service as pc
    from app.services.data_export_service import (
        build_single_item_discount_upload_xlsx, _TB_DISCOUNT_HEADERS)
    sku = PricingSku(product_code="P1", sku_code="P1-A", sku="1.2米",
                     daily_price=D("19575"), big_promo=D("12890"))
    promo = PricingSkuPromo(sku_code="P1-A",
                            taobao_item_id="917179577721", taobao_sku_id="6241018727157")
    sku2 = PricingSku(product_code="P2", sku_code="P2-A", sku="x",
                      daily_price=D("1000"), big_promo=D("600"))
    promo2 = PricingSkuPromo(sku_code="P2-A", taobao_item_id="123")   # 无 SKU_ID → 跳过
    db_session.add_all([Product(code="P1", name="a"), Product(code="P2", name="b"),
                        sku, promo, sku2, promo2])
    db_session.commit()
    pc.recompute_promo(promo, sku, {"big_vip_commission": D("0.02"), "mid_vip_commission": D("0.02")})
    pc.recompute_promo(promo2, sku2, {"big_vip_commission": D("0.02"), "mid_vip_commission": D("0.02")})
    db_session.commit()

    bio, stats = build_single_item_discount_upload_xlsx(db_session, "big")
    ws = openpyxl.load_workbook(io.BytesIO(bio.getvalue())).active
    assert [ws.cell(1, c).value for c in range(1, 6)] == _TB_DISCOUNT_HEADERS   # 表头逐字对齐
    assert stats["rows"] == 1 and stats["skipped_no_skuid"] == 1                # P2 无SKU_ID被跳
    assert ws.cell(2, 1).value == "917179577721"                               # 商品id 文本
    assert ws.cell(2, 2).value == "6241018727157"                             # SKU_ID 文本
    assert ws.cell(2, 1).number_format == "@"                                   # 长号文本防科学计数
    assert abs(float(ws.cell(2, 3).value) - 4072.94) < 1.5                     # 大促立减金额(减钱)
    # 618 档官方减更多 → 单品立减金额更少
    bio6, _ = build_single_item_discount_upload_xlsx(db_session, "big618")
    ws6 = openpyxl.load_workbook(io.BytesIO(bio6.getvalue())).active
    assert float(ws6.cell(2, 3).value) < float(ws.cell(2, 3).value)


def test_catalog_and_total_sheet_have_discount_amounts(db_session, monkeypatch):
    """图册 + 全量导出「定价总表」: 单品立减改加法(折 + 降价金额, 大促/超大促), 无旧乘法系数。
    附图: 日常19575, 大促价12890, 佣金2% → 大促降价金额≈4073(7.92折), 超大促≈3486(8.22折)。"""
    import io
    import openpyxl
    from decimal import Decimal as D
    from app.models.pricing_ext import PricingSkuPromo
    from app.services import pricing_calc_service as pc
    from app.services.data_export_service import build_catalog_xlsx, build_full_export_workbook
    monkeypatch.setattr(pcs, "product_image_map", lambda codes, url_by_code, **k: {})
    sku = PricingSku(product_code="P1", sku_code="P1-A", sku="1.2米",
                     list_price=D("26100"), daily_price=D("19575"), big_promo=D("12890"))
    promo = PricingSkuPromo(sku_code="P1-A")
    db_session.add_all([Product(code="P1", name="曜黑餐边柜"), sku, promo])
    db_session.commit()
    pc.recompute_promo(promo, sku, {"big_vip_commission": D("0.02"), "mid_vip_commission": D("0.02")})
    db_session.commit()

    # 图册
    ws = openpyxl.load_workbook(io.BytesIO(build_catalog_xlsx(db_session).getvalue()))["定价图册"]
    hdr = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    assert "大促店铺系数" not in hdr and "中促店铺系数" not in hdr        # 旧乘法系数已去
    assert abs(float(ws.cell(3, hdr["大促降价金额(元)"]).value) - 4072.94) < 1.5
    assert abs(float(ws.cell(3, hdr["大促单品立减(折)"]).value) - 7.92) < 0.02
    assert (float(ws.cell(3, hdr["超大促降价金额(元)"]).value)
            < float(ws.cell(3, hdr["大促降价金额(元)"]).value))       # 618官方减更多→单品立减更少

    # 全量导出的「定价总表」sheet 也有这两列
    wb2 = build_full_export_workbook(db_session)
    pname = [t for t in wb2.sheetnames if "定价总表" in t][0]
    ws2 = wb2[pname]
    h2 = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1) if ws2.cell(1, c).value}
    assert "大促降价金额(元)" in h2 and "超大促降价金额(元)" in h2 and "大促店铺系数" not in h2

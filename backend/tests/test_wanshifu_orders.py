"""万师傅安装订单档案: 解析 + 导入幂等 + 订单配对 测试。"""
from datetime import date

import openpyxl

from app.models.finance import WanshifuOrder
from app.models.order import Order
from app.services import wanshifu_order_service as wsf


def _make_wb():
    """合成 38 列格式最小版: r1 分组 / r2 字段名 / r3 地址子表头 / r4+ 数据。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单信息", "", "", "", "", "", "", "", "客户信息"])
    ws.append(["订单编号", "服务类目/类型", "订单状态", "商品类别", "客户姓名",
               "客户手机号", "客户地址", "", "", "", "订单服务费", "下单时间",
               "物流公司", "物流单号"])
    ws.append(["", "", "", "", "", "", "省", "市", "区", "详细地址", "", "", "", ""])
    ws.append(["P100", "家具|安装", "交易成功", "桌类-餐台/餐桌", "张三测",
               "13800000001-1234", "浙江", "杭州", "西湖区", "某街道1号",
               "78", "2026-05-10 12:00:00", "", ""])
    ws.append(["P101", "家具|安装", "交易关闭（自动关单）", "柜类-餐边柜", "李四测",
               "13900000002", "江苏", "苏州", "园区", "某路2号",
               "", "2025-12-01 09:00:00", "", ""])
    return wb


def test_parse_and_import_idempotent(db_session):
    rep = wsf.import_workbook(db_session, _make_wb())
    assert rep.parsed == 2 and rep.inserted == 2

    # 重导: 不重复插入
    rep2 = wsf.import_workbook(db_session, _make_wb())
    assert rep2.inserted == 0
    rows = db_session.query(WanshifuOrder).all()
    assert len(rows) == 2
    w = next(r for r in rows if r.wsf_order_no == "P100")
    assert w.city == "杭州" and w.customer_phone == "13800000001-1234"
    assert float(w.service_fee) == 78


def test_match_phone_and_unmatched_reasons(db_session):
    wsf.import_workbook(db_session, _make_wb())
    # 订单电话与 P100 全等 → phone_full 配对
    db_session.add(Order(platform="淘宝", order_no="T-001", qty=1,
                         customer_name="张三测", customer_phone="13800000001-1234",
                         order_date=date(2026, 5, 1)))
    db_session.flush()

    counts = wsf.match_orders(db_session)

    assert counts["matched"] == 1
    w100 = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P100").one()
    assert w100.matched_order_no == "T-001"
    assert w100.match_method == "phone_full"
    # P101 是 2025 年单 → none + 人话原因
    w101 = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P101").one()
    assert w101.match_method == "none"
    assert "2025" in (w101.match_note or "")


def _wb_with_remark():
    """带「常用备注」列(=淘宝订单号, 用户的合并单号匹配)的最小 wb。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单信息"])
    ws.append(["订单编号", "服务类目/类型", "订单状态", "商品类别", "客户姓名",
               "客户手机号", "客户地址", "", "", "", "订单服务费", "下单时间",
               "物流公司", "物流单号", "常用备注"])
    ws.append(["", "", "", "", "", "", "省", "市", "区", "详细地址", "", "", "", "", ""])
    ws.append(["P200", "家具|安装", "交易成功", "桌类-餐台/餐桌", "王五测",
               "13700000003", "广东", "深圳", "南山区", "某路3号",
               "68", "2026-05-12 10:00:00", "", "", "3016793965259618396"])
    ws.append(["P201", "家具|安装", "交易成功", "柜类-餐边柜", "赵六测",
               "13600000004", "北京", "北京", "海淀区", "某街4号",
               "88", "2026-05-13 11:00:00", "", "", "9999999999999999999"])  # 备注单号不在订单库
    return wb


def test_remark_taobao_no_is_authoritative_match(db_session):
    """常用备注里的淘宝订单号 = 用户"合并单号匹配", 命中订单库即权威配对(method=remark)。"""
    # 订单库里有 P200 备注指向的淘宝单, 没有 P201 的
    db_session.add(Order(platform="淘宝", order_no="3016793965259618396", qty=1,
                         customer_name="某客户", order_date=date(2026, 5, 1)))
    db_session.flush()

    wsf.import_workbook(db_session, _wb_with_remark())

    w200 = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P200").one()
    assert w200.matched_order_no == "3016793965259618396"
    assert w200.match_method == "remark"
    assert "淘宝单号:3016793965259618396" in (w200.remark or "")

    # P201 备注单号不在订单库 → 不强配, 但留备注提示
    w201 = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P201").one()
    assert w201.matched_order_no is None
    assert "未在订单库" in (w201.match_note or "")


def test_remark_match_overrides_heuristic_on_reimport(db_session):
    """备注单号权威: 即便之前被启发式配过, 重导也以备注为准(非人工)。"""
    db_session.add(Order(platform="淘宝", order_no="3016793965259618396", qty=1,
                         customer_name="某客户", order_date=date(2026, 5, 1)))
    db_session.flush()
    wsf.import_workbook(db_session, _wb_with_remark())
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P200").one()
    w.matched_order_no = "WRONG-001"
    w.match_method = "name_unique"
    db_session.flush()
    # 重导 → 备注覆盖
    wsf.import_workbook(db_session, _wb_with_remark())
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P200").one()
    assert w.matched_order_no == "3016793965259618396"
    assert w.match_method == "remark"


# ---------------- 「校对后匹配」人工列 (方案1: 人工核对后落库) ----------------

def _wb_with_verified():
    """带「校对后匹配」人工列(表头在分组行 r1, 模拟用户实际文件)的最小 wb。
    第15列(index14)=校对后匹配; P300=真实订单号, P301=非订单号文字(样品)。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单信息"] + [""] * 13 + ["校对后匹配"])                      # r1 分组行
    ws.append(["订单编号", "服务类目/类型", "订单状态", "商品类别", "客户姓名",
               "客户手机号", "客户地址", "", "", "", "订单服务费", "下单时间",
               "物流公司", "物流单号", ""])                                   # r2 字段名
    ws.append(["", "", "", "", "", "", "省", "市", "区", "详细地址",
               "", "", "", "", ""])                                          # r3 地址子表头
    ws.append(["P300", "家具|安装", "交易成功", "桌类-餐台/餐桌", "孙七测",
               "13500000005", "四川", "成都", "武侯区", "某路5号",
               "98", "2026-06-01 10:00:00", "", "", "3210721863514837177"])
    ws.append(["P301", "家具|安装", "交易成功", "柜类-餐边柜", "周八测",
               "13400000006", "湖北", "武汉", "洪山区", "某街6号",
               "108", "2026-06-02 11:00:00", "", "", "样品"])                 # 非订单号 → 不强配
    return wb


def test_verified_column_is_highest_authority(db_session):
    """「校对后匹配」命中订单库 → method=manual, 批注干净。"""
    db_session.add(Order(platform="淘宝", order_no="3210721863514837177", qty=1,
                         customer_name="某客户", order_date=date(2026, 6, 1)))
    db_session.flush()
    wsf.import_workbook(db_session, _wb_with_verified())
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P300").one()
    assert w.matched_order_no == "3210721863514837177"
    assert w.match_method == "manual"
    assert w.match_note is None


def test_verified_records_even_when_order_absent(db_session):
    """人工校对号订单库暂无(早期单/未导入)也照记匹配, 批注提示。"""
    wsf.import_workbook(db_session, _wb_with_verified())   # 订单库无 3210721...
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P300").one()
    assert w.matched_order_no == "3210721863514837177"
    assert w.match_method == "manual"
    assert "订单库暂无此单" in (w.match_note or "")


def test_verified_non_order_text_not_force_matched(db_session):
    """校对后匹配=样品 等非订单号文字 → 不强配 (matched_order_no 留空)。"""
    wsf.import_workbook(db_session, _wb_with_verified())
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P301").one()
    assert w.matched_order_no is None
    assert w.match_method != "manual"


def test_verified_overrides_wrong_match_on_reimport(db_session):
    """重导时人工校对列以最新为准, 覆盖旧的错配(含旧 manual)。"""
    db_session.add(Order(platform="淘宝", order_no="3210721863514837177", qty=1,
                         customer_name="某客户", order_date=date(2026, 6, 1)))
    db_session.flush()
    wsf.import_workbook(db_session, _wb_with_verified())
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P300").one()
    w.matched_order_no = "WRONG-999"      # 模拟旧错配
    w.match_method = "manual"
    db_session.flush()
    wsf.import_workbook(db_session, _wb_with_verified())  # 重导
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P300").one()
    assert w.matched_order_no == "3210721863514837177"
    assert w.match_method == "manual"


def _wb_verified_and_remark():
    """同时有「常用备注」(index14)和「校对后匹配」(index15)两列且值不同 → 人工列应胜出。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单信息"] + [""] * 14 + ["校对后匹配"])
    ws.append(["订单编号", "服务类目/类型", "订单状态", "商品类别", "客户姓名",
               "客户手机号", "客户地址", "", "", "", "订单服务费", "下单时间",
               "物流公司", "物流单号", "常用备注", ""])
    ws.append(["", "", "", "", "", "", "省", "市", "区", "详细地址",
               "", "", "", "", "", ""])
    ws.append(["P400", "家具|安装", "交易成功", "桌类", "钱九测",
               "13300000007", "湖南", "长沙", "岳麓区", "某路7号",
               "118", "2026-06-03 10:00:00", "", "",
               "1111111111111111111", "2222222222222222222"])
    return wb


def test_verified_beats_remark(db_session):
    """人工校对号优先级高于常用备注单号。"""
    db_session.add(Order(platform="淘宝", order_no="1111111111111111111", qty=1,
                         customer_name="c1", order_date=date(2026, 6, 1)))
    db_session.add(Order(platform="淘宝", order_no="2222222222222222222", qty=1,
                         customer_name="c2", order_date=date(2026, 6, 1)))
    db_session.flush()
    wsf.import_workbook(db_session, _wb_verified_and_remark())
    w = db_session.query(WanshifuOrder).filter_by(wsf_order_no="P400").one()
    assert w.matched_order_no == "2222222222222222222"
    assert w.match_method == "manual"

"""木作工厂月结 逐单明细导出 (配件采购 + 月结对账中心 共用, 2026-07-08)。
复用已存在的 settlement_detail_rows(默认木作供应商); write_settlement_ws 写共用 sheet。"""
from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook

from app.models.factory_settlement import DEFAULT_WOOD_SUPPLIER as SUP
from app.models.order import FactoryOrder
from app.services import factory_settlement_service as fss


def test_settlement_detail_rows_finds_wood_order(db_session):
    db_session.add(FactoryOrder(factory_order_no="FX", factory_name=SUP, platform_order_no="OX",
                                factory_bill_amount=Decimal("100"), payment_status="unpaid",
                                settlement_month="2026-05"))
    db_session.flush()
    rows = fss.settlement_detail_rows(db_session)
    hit = [r for r in rows if r["platform_order_no"] == "OX"]
    assert hit and hit[0]["payment_status"] == "未付" and hit[0]["settlement_month"] == "2026-05"


def test_write_settlement_ws_structure():
    rows = [
        {"settlement_month": "2026-05", "factory_order_no": "F1", "platform_order_no": "OX",
         "product_name": "餐边柜", "sku": "S", "qty": 1, "bill_amount": Decimal("500"),
         "payment_status": "未付", "paid_amount": Decimal("0")},
        {"settlement_month": "2026-06", "factory_order_no": "F2", "platform_order_no": "OY",
         "product_name": "床", "sku": "S2", "qty": 2, "bill_amount": Decimal("300"),
         "payment_status": "已付", "paid_amount": Decimal("300")},
    ]
    wb = Workbook()
    fss.write_settlement_ws(wb, rows=rows)
    assert "木作工厂月结" in wb.sheetnames
    ws = wb["木作工厂月结"]
    assert [c.value for c in ws[1]] == [
        "结算月", "工厂单号", "平台订单号", "产品", "数量", "账单金额", "已付金额", "付款状态"]
    body = {ws.cell(r, 3).value: ws.cell(r, 8).value for r in range(2, 4)}
    assert body.get("OX") == "未付" and body.get("OY") == "已付"

"""活动生命周期 P1 引擎测试 (2026-07-17 spec: docs/活动生命周期系统_执行plan.md)。

锁八件事:
① 动销分组 + no_sales 登记表同步 (新零动销自动登记 / 出单只提示转正不移除)
② 报名行: 报名价=日常价 / 占位=min(现行, floor(线/0.88)) / 无线保守值备注 / R4下架过滤 / R3整品完整性
③ 立减公式 spec §二 手算样例: 日常2827.5 / 大促1979.59 / 线1978.89 → 官方340 → 立减508.61
④ 中促 = 大促×1.03 就地计算 + 10% ceil 开关 campaign_official_ceil
⑤ 无动销: 立减 = 日常 − ERP中促价; 占位不出行
⑥ R2: 贴线让幅 >1 元 → 整品暂缓，不轮换
⑦ preflight R1~R14 逐条输出 {rule, level, items}
⑧ 推送编排 (mock WA, 绝不真调 :8500): channel/phase/档期传参 + 状态机推进
"""
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

import openpyxl

from app.models.campaign import CampaignPlan
from app.models.order import Order
from app.models.product import Product
from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import campaign_service as cs
from app.services import delisted_sku_service as ds
from app.services import no_sales_service as ns


def _mk(db, pc, code, item, sid, daily=None, big=None, *, placeholder=False,
        line=None, enrolled=None):
    db.add(PricingSku(product_code=pc, sku_code=code, sku=f"SKU{code}",
                      product_name=f"品{pc}",
                      daily_price=Decimal(str(daily)) if daily is not None else None,
                      is_custom_placeholder=placeholder))
    db.add(PricingSkuPromo(
        sku_code=code, taobao_item_id=item, taobao_sku_id=sid,
        big_buyer_price=Decimal(str(big)) if big is not None else None,
        coupon_floor_price=Decimal(str(line)) if line is not None else None,
        enrolled_floor_price=Decimal(str(enrolled)) if enrolled is not None else None))


def _seed_platform_floors(db, rows):
    """Seed fresh SKUID evidence exactly as a current activity export would."""
    from app.services import campaign_price_floor_service

    campaign_price_floor_service.record_activity_export(
        db,
        [{
            "item_id": str(item_id),
            "sku_id": str(sku_id),
            "sku_name": str(sku_id),
            "min_list_price": min_list,
            "min_coupon_line": min_coupon,
        } for item_id, sku_id, min_list, min_coupon in rows],
        source="pytest_current_activity_export",
    )
    db.commit()


def _plan(db, ctype="big88"):
    plan = CampaignPlan(name=f"测试{ctype}", campaign_type=ctype,
                        tier=cs.CAMPAIGN_TYPES[ctype][1],
                        start_at=datetime(2026, 7, 17, 20, 0, 0),
                        end_at=datetime(2026, 7, 19, 23, 59, 59), status="draft")
    db.add(plan)
    db.commit()
    return plan


def _order(db, no, pc, status="paid", days_ago=5, platform="淘宝"):
    db.add(Order(platform=platform, order_no=no, product_code=pc, status=status,
                 order_date=date.today() - timedelta(days=days_ago)))


# ── ① 动销分组 + 登记表同步 ──────────────────────────────────────────────────

def test_group_by_sales_and_registry_sync(db_session):
    # ⚠编码数字主体必须互异: brand_variants 按 core_of(剥字母前缀后的数字) 归并跨品牌销量
    _mk(db_session, "PPS26101", "PPS2610101", "9101", "71001", daily=1000, big=800)
    _mk(db_session, "PPS26202", "PPS2620201", "9102", "71002", daily=1000, big=800)
    _mk(db_session, "PPS26303", "PPS2630301", "9103", "71003", daily=1000, big=800)
    _order(db_session, "T1", "P26101")                        # 9101 有动销 (订单用品牌无关 P 形式)
    _order(db_session, "T2", "PPS26202", status="cancelled")  # 关闭单不算 → 9102 零动销
    _order(db_session, "T3", "PPS26303", status="signed")     # 9103 有动销(但已在登记表)
    _order(db_session, "T4", "PPS26101", days_ago=90)         # 60天窗口外不算
    db_session.commit()
    ns.add_no_sales(db_session, ["9103"])                     # 预登记: 后来卖出去了

    g = cs.group_by_sales(db_session)

    assert g["有动销"] == ["9101", "9103"]
    assert g["无动销"] == ["9102"]
    assert g["newly_registered"] == ["9102"]                  # 新零动销自动登记
    assert g["promote_candidates"] == ["9103"]                # 出单提示转正
    assert set(g["registered"]) == {"9102", "9103"}           # ★不自动移除 (R6 单行道)


# ── ② 报名行 builder ─────────────────────────────────────────────────────────

def test_signup_rows_price_placeholder_and_filters(db_session):
    plan = _plan(db_session, "big88")                          # lev = 0.12
    _mk(db_session, "PPSSA001", "PPSSA00101", "9201", "72001", daily=2827.5)
    _mk(db_session, "PPSSA001", "PPSSA00102", "9201", "72002", daily=1500)
    # 占位有线: 现行 = min(1000×0.9, 500) = 500; cap = floor(430/0.88) = 488 → 488
    _mk(db_session, "PPSSA001", "PPSSA00190", "9201", "72090", daily=1000,
        placeholder=True, line=430)
    # 下架SKU (R4): 过滤且不破坏整品完整性
    _mk(db_session, "PPSSA001", "PPSSA00103", "9201", "72003", daily=1600)
    ds.add_delisted(db_session, ["72003"])
    # 占位无线: 现行 500 < floor(1000×0.8/0.88)=909 → 500 + 备注
    _mk(db_session, "PPSSB001", "PPSSB00101", "9202", "72011", daily=2000)
    _mk(db_session, "PPSSB001", "PPSSB00190", "9202", "72091", daily=1000, placeholder=True)
    db_session.commit()

    rows, stats = cs.build_signup_rows(db_session, plan)
    by_sid = {r["taobao_sku_id"]: r for r in rows}

    assert by_sid["72001"]["price"] == 2827.5                 # 报名价 = 日常价 (铁则1)
    assert by_sid["72002"]["price"] == 1500.0
    assert by_sid["72090"]["price"] == 488.0                  # min(现行500, floor(线/0.88)=488)
    assert by_sid["72091"]["price"] == 500.0                  # 无线 → 保守值封顶不生效, 现行500
    assert by_sid["72091"]["remark"] and "0.8" in by_sid["72091"]["remark"]
    assert stats["placeholder_no_line"] == [
        {"sku_code": "PPSSB00190", "remark": by_sid["72091"]["remark"]}]
    assert "72003" not in by_sid and stats["skipped_delisted"] == 1   # R4 过滤
    assert stats["incomplete_items"] == []                    # 下架不算缺 → 整品仍完整
    assert stats["rows"] == len(rows) == 5


def test_signup_rows_include_registered_no_sales_for_platform_recheck(db_session):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSNS001", "PPSNS00101", "9209", "72901",
        daily=1200, big=800)
    db_session.commit()
    ns.add_no_sales(db_session, ["9209"])

    rows, stats = cs.build_signup_rows(db_session, plan)

    assert {row["taobao_item_id"] for row in rows} == {"9209"}
    assert stats["excluded_no_sales_items"] == []
    assert stats["registered_no_sales_items_included"] == ["9209"]


def test_signup_shipping_days_authorization_is_exact():
    plan = CampaignPlan(remark=(
        "signup_shipping_days_authorized="
        "793084818113:30,100000000002:0,100000000003:366,bad"
    ))

    assert cs.authorized_signup_shipping_days(plan) == {"793084818113": 30}


def test_signup_workbook_writes_shipping_days_only_for_authorized_item():
    rows = [
        {"taobao_item_id": "793084818113", "taobao_sku_id": "6292847403160",
         "price": 1200.0},
        {"taobao_item_id": "100000000004", "taobao_sku_id": "6292847403161",
         "price": 397.0},
    ]

    workbook = openpyxl.load_workbook(io.BytesIO(
        cs._build_signup_xlsx(rows, {"793084818113": 30})))
    sheet = workbook["商品SKU导入列表"]

    assert sheet.cell(4, 1).value == "793084818113"
    assert sheet.cell(4, 3).value == 1200
    assert sheet.cell(4, 5).value == 30
    assert sheet.cell(5, 1).value == "100000000004"
    assert sheet.cell(5, 3).value == 397
    assert sheet.cell(5, 5).value is None


def test_campaign_rows_are_limited_to_erp_listed_products(db_session):
    plan = _plan(db_session, "big88")
    db_session.add_all([
        Product(code="PPSLISTED1", name="listed", listing_status="在售"),
        Product(code="PPSOFF1", name="not listed", listing_status="下架"),
    ])
    _mk(db_session, "PPSLISTED1", "PPSLISTED101", "1000010001", "710001", daily=1200, big=800)
    _mk(db_session, "PPSOFF1", "PPSOFF101", "1000010002", "710002", daily=1300, big=900)
    _mk(db_session, "PPSUTILITY1", "PPSUTILITY101", "1000010003", "710003", daily=10, big=8)
    db_session.commit()

    signup, signup_stats = cs.build_signup_rows(db_session, plan)
    discount, discount_stats = cs.build_discount_rows(db_session, plan)

    assert {row["taobao_item_id"] for row in signup} == {"1000010001"}
    assert {row["taobao_item_id"] for row in discount} == {"1000010001"}
    assert signup_stats["skipped_not_erp_listed"] == 2
    assert discount_stats["skipped_not_erp_listed"] == 2


def test_platform_qualification_only_no_sales_failure_is_normal_fallback(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = "official_all_store=true; official_exempt_items=1000009999"
    _mk(db_session, "PPSQUAL1", "PPSQUAL101", "1000009209", "72901", daily=1200, big=800)
    _mk(db_session, "PPSQUAL2", "PPSQUAL201", "1000009210", "72902", daily=1300, big=900)
    db_session.commit()
    ns.add_no_sales(db_session, ["1000009209", "1000009210"])
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": False,
        "validation": {
            "total_items": 2, "ok": 1, "failed": 1,
            "failed_items": [{"item_id": "1000009209", "reason": "动销不达标", "raw": "动销"}],
        },
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == ["1000009210"]
    assert result["no_sales_item_ids"] == ["1000009209"]
    assert ns.get_no_sales(db_session) == {"1000009209"}
    assert cs.platform_qualified_items(plan) == {"1000009210"}
    assert cs.official_scope_for_plan(plan)["active_items"] == {"1000009210"}
    assert cs.official_scope_for_plan(plan)["all_store"] is False
    assert cs.official_scope_for_plan(plan)["errors"] == []


def test_promo_qualification_uploads_authorized_shipping_days(
        db_session, monkeypatch):
    plan = _plan(db_session, "big88")
    plan.remark = (
        "supplement_items_authorized=793084818113; "
        "signup_shipping_days_authorized=793084818113:30"
    )
    _mk(db_session, "PPSSHIP1", "PPSSHIP101", "793084818113",
        "6292847403160", daily=1200, big=800)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    captured = {}

    def upload(*args, **kwargs):
        captured["workbook"] = args[3]
        captured.update(kwargs)
        return {
            "ok": True,
            "validation": {"total_items": 1, "ok": 1, "failed": 0},
        }

    monkeypatch.setattr(cs, "_upload_and_wait", upload)

    result = cs.qualify_signup_scope(db_session, plan)
    workbook = openpyxl.load_workbook(io.BytesIO(captured["workbook"]))
    row = workbook["商品SKU导入列表"][4]

    assert result["ok"] is True
    assert captured["expected_rows"] == 1
    assert captured["expected_items"] == 1
    assert [cell.value for cell in row[:5]] == [
        "793084818113", "6292847403160", 1200, None, 30]


def test_platform_qualification_limits_supplement_and_preserves_prior_scope(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        "supplement_items_authorized=1000009221; "
        "platform_qualified_items=1000009222; "
        "platform_hard_failed_items=1000009221"
    )
    _mk(db_session, "PPSQUALS1", "PPSQUALS101", "1000009221", "72921",
        daily=1200, big=800)
    _mk(db_session, "PPSQUALS2", "PPSQUALS201", "1000009222", "72922",
        daily=1300, big=900)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    calls = []

    def fake_upload(*args, **kwargs):
        calls.append(kwargs)
        return {
            "ok": True,
            "validation": {"total_items": 1, "ok": 1, "failed": 0},
        }

    monkeypatch.setattr(cs, "_upload_and_wait", fake_upload)
    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert calls[0]["expected_items"] == 1
    assert cs.platform_qualified_items(plan) == {"1000009221", "1000009222"}
    assert cs.platform_hard_failed_items(plan) == set()
    assert cs.authorized_supplement_items(plan) == {"1000009221"}


def test_platform_qualification_isolates_non_sales_failure_and_continues(db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSQUAL3", "PPSQUAL301", "1000009211", "72903", daily=1400, big=950)
    _mk(db_session, "PPSQUAL4", "PPSQUAL401", "1000009215", "72908", daily=1500, big=1000)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": False,
        "validation": {
            "total_items": 2, "ok": 1, "failed": 1,
            "failed_items": [{"item_id": "1000009211", "reason": "SKU已失效", "raw": "SKU已失效"}],
        },
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == ["1000009215"]
    assert result["hard_failed_item_ids"] == ["1000009211"]
    assert result["no_sales_item_ids"] == []
    assert cs.platform_qualified_items(plan) == {"1000009215"}


def test_platform_qualification_accepts_coupon_only_failure_when_planned_discount_clears_floor(
        db_session, monkeypatch):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSQUALC1", "PPSQUALC101", "1000009216", "72916",
        daily=1500, big=1000)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    monkeypatch.setattr(cs, "price_hold_items", lambda *args: [])
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": False,
        "validation": {
            "total_items": 1, "ok": 0, "failed": 1,
            "failed_items": [{
                "item_id": "1000009216",
                "reason": "券后价超线(报名价高于已生效最低价)",
                "raw": "活动普惠券后价不可高于最低普惠券后价",
            }],
        },
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == ["1000009216"]
    assert result["planned_discount_qualification_item_ids"] == ["1000009216"]
    assert result["terminal_accepted_item_ids"] == []
    assert result["hard_failed_item_ids"] == []
    assert cs.platform_terminal_accepted_items(plan) == set()
    assert cs._fresh_terminal_coupon_floor_qualification(
        plan, cs.build_signup_rows(db_session, plan)[0], 24
    )["item_ids"] == {"1000009216"}


def test_no_sales_classifier_prefers_terminal_reason_over_policy_boilerplate():
    feedback = [{
        "item_id": "1000009218",
        "sku_id": "72918",
        "reason": "动销不达标(近60天销量<1)",
        "raw": (
            "参加活动须满足动销校验。活动期间的标价不得高于近15天最低标价；"
            "此段为平台通用政策说明，不是本行的第二个失败原因。"
        ),
    }]

    assert ns.extract_no_sales_only_from_feedback(feedback) == {"1000009218"}


def test_platform_qualification_allows_exact_authorized_sku_refresh(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    item_id = "1000009219"
    plan.remark = f"sku_refresh_items_authorized={item_id}"
    _mk(db_session, "PPSROT01", "PPSROT0101", item_id, "NEW-SID",
        daily=1200, big=800)
    _mk(db_session, "PPSROT01", "PPSROT0199", item_id, "KEEP-SID",
        daily=500, big=400, placeholder=True, line=350)
    db_session.commit()

    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *a, **k: {
        "ok": True,
        "rows": [
            {
                "item_id": item_id, "sku_id": "NEW-SID", "status": "活动中",
                "activity_price": 1100,
            },
            {
                "item_id": item_id, "sku_id": "KEEP-SID", "status": "活动中",
                "activity_price": 397,
            },
        ],
        "floor_refresh": {"observed": 2},
    })
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *a, **k: {
        "ok": True,
        "validation": {
            "total_items": 1, "ok": 1, "failed": 0, "failed_items": [],
        },
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == [item_id]
    assert result["wrong_existing_items"] == []


def test_platform_qualification_keeps_coupon_failure_hard_when_internal_floor_hold_remains(
        db_session, monkeypatch):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSQUALC2", "PPSQUALC201", "1000009217", "72917",
        daily=1500, big=1000)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    hold_calls = iter(([], [{
        "taobao_item_id": "1000009217", "skus": [],
    }]))
    monkeypatch.setattr(cs, "price_hold_items", lambda *args: next(hold_calls))
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": False,
        "validation": {
            "total_items": 1, "ok": 0, "failed": 1,
            "failed_items": [{
                "item_id": "1000009217",
                "reason": "券后价超线(报名价高于已生效最低价)",
                "raw": "活动普惠券后价不可高于最低普惠券后价",
            }],
        },
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == []
    assert result["planned_discount_qualification_item_ids"] == []
    assert result["hard_failed_item_ids"] == ["1000009217"]


def test_super_qualification_fetches_missing_feedback_and_keeps_mixed_failure_hard(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSQUAL5", "PPSQUAL501", "1000009221", "72921", daily=1400, big=950)
    _mk(db_session, "PPSQUAL6", "PPSQUAL601", "1000009222", "72922", daily=1500, big=1000)
    _mk(db_session, "PPSQUAL7", "PPSQUAL701", "1000009223", "72923", daily=1600, big=1100)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": False,
        "validation": {"total_items": 3, "ok": 1, "failed": 2, "failed_items": []},
    })
    from app.services import web_agent_service
    monkeypatch.setattr(web_agent_service, "super_reduce_feedback", lambda *args, **kwargs: {
        "ok": True,
        "feedback": {"by_reason": [], "failed": [
            {"item_id": "1000009221", "reason": "动销不达标", "raw": "近60天销售件数≥1件"},
            {"item_id": "1000009222", "reason": "券后价超线", "raw": "动销不达标；最低普惠券后价"},
        ]},
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == ["1000009223"]
    assert result["no_sales_item_ids"] == ["1000009221"]
    assert result["hard_failed_item_ids"] == ["1000009222"]
    assert ns.get_no_sales(db_session) == {"1000009221"}


def test_promo_qualification_flattens_item_feedback_from_agent(db_session, monkeypatch):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSQUALP1", "PPSQUALP101", "1000009241", "72941",
        daily=1400, big=950)
    _mk(db_session, "PPSQUALP2", "PPSQUALP201", "1000009242", "72942",
        daily=1500, big=1000)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": True,
        "validation": {
            "total_items": 2, "ok": 1, "failed": 1,
            "failed_reasons": {
                "by_reason": [{"reason": "动销不达标", "items": 1}],
                "failed": [{"item_id": "1000009241", "reason": "动销不达标",
                            "raw": "近60天销量<1"}],
            },
        },
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == ["1000009242"]
    assert result["no_sales_item_ids"] == ["1000009241"]


def test_promo_qualification_retries_feedback_by_exact_campaign_ids(
        db_session, monkeypatch):
    plan = _plan(db_session, "big88")
    plan.qn_campaign_title = "2026年淘宝8月开学季"
    plan.remark = "campaignId=49271; unitedActivityId=49283"
    _mk(db_session, "PPSQUALP3", "PPSQUALP301", "1000009251", "72951",
        daily=1400, big=950)
    _mk(db_session, "PPSQUALP4", "PPSQUALP401", "1000009252", "72952",
        daily=1500, big=1000)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True, "rows": [], "floor_refresh": {},
    })
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": True,
        "validation": {"total_items": 2, "ok": 1, "failed": 1},
    })
    from app.services import web_agent_service
    captured = {}

    def feedback(_db, title, **kwargs):
        captured.update({"title": title, **kwargs})
        return {"ok": True, "feedback": {
            "by_reason": [{"reason": "SKU缺失", "items": 1}],
            "failed": [{"item_id": "1000009251", "reason": "SKU缺失",
                        "raw": "缺失的SKUID=72999"}],
        }}

    monkeypatch.setattr(web_agent_service, "campaign_feedback", feedback)

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == ["1000009252"]
    assert result["hard_failed_item_ids"] == ["1000009251"]
    assert captured["title"] == "2026年淘宝8月开学季"
    assert captured["campaign_id"] == "49271"
    assert captured["united_activity_id"] == "49283"


def test_preflight_uses_qualified_scope_and_does_not_require_campaign_floor_for_fallback(
        db_session):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSQUAL8", "PPSQUAL801", "1000009231", "72931", daily=1400, big=950)
    _mk(db_session, "PPSQUAL9", "PPSQUAL901", "1000009232", "72932", daily=1500, big=1000)
    _mk(db_session, "PPSQUALA", "PPSQUALA01", "1000009233", "72933", daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [("1000009231", "72931", 2000, 980)])
    ns.add_no_sales(db_session, ["1000009232"])
    cs._set_plan_item_marker(plan, "platform_qualified_items", {"1000009231"})
    cs._set_plan_item_marker(plan, "platform_no_sales_items", {"1000009232"})
    cs._set_plan_item_marker(plan, "platform_hard_failed_items", {"1000009233"})
    cs._set_plan_item_marker(plan, "official_active_items", {"1000009231"})
    db_session.commit()

    checks = {row["rule"]: row for row in cs.preflight(db_session, plan)}

    assert checks["R17"]["level"] == "pass"
    assert checks["R17"]["checked"] == 1
    assert cs.platform_scope_present(plan) is True
    assert cs.platform_hard_failed_items(plan) == {"1000009233"}


def test_platform_qualification_preserves_existing_placeholder_protection_price(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSPLACE1", "PPSPLACE101", "1000009212", "72904",
        daily=1200, big=800)
    _mk(db_session, "PPSPLACE1", "PPSPLACE199", "1000009212", "72905",
        daily=1000, big=800, placeholder=True, line=300)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True,
        "rows": [
            {"item_id": "1000009212", "sku_id": "72904", "activity_price": 1200.0},
            {"item_id": "1000009212", "sku_id": "72905", "activity_price": 500.0},
        ],
        "floor_refresh": {},
    })
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: {
        "ok": False, "error": "must_not_probe_already_correct_item",
    })

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["no_change"] is True
    assert result["qualified_item_ids"] == ["1000009212"]


def test_platform_qualification_isolates_existing_wrong_item_and_continues(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSWRONG1", "PPSWRONG101", "1000009213", "72906",
        daily=1200, big=800)
    _mk(db_session, "PPSFRESH1", "PPSFRESH101", "1000009214", "72907",
        daily=1300, big=900)
    db_session.commit()
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity", lambda *args: {
        "ok": True,
        "rows": [{"item_id": "1000009213", "sku_id": "72906", "activity_price": 1199.0}],
        "floor_refresh": {},
    })
    calls = []
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: (
        calls.append(kwargs) or {
            "ok": True,
            "validation": {"total_items": 1, "ok": 1, "failed": 0},
        }
    ))

    result = cs.qualify_signup_scope(db_session, plan)

    assert result["ok"] is True
    assert result["qualified_item_ids"] == ["1000009214"]
    assert result["wrong_existing_items"][0]["item_id"] == "1000009213"
    assert calls[0]["expected_items"] == 1

    discount = cs.push_discount(db_session, plan, phase="commit")
    assert discount["ok"] is True
    assert discount["stats"]["platform_discount_scope_items"] == ["1000009214"]


def test_honey_sample_is_real_sku_and_stacks_discount(db_session):
    """蜜蜡色样块是正常商品，不得再按定制占位保护价报 18 元。"""
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPS2398001", "PPS2398001060614", "719436834260",
        "6282622238127", daily=30, big=20.41, placeholder=False)
    db_session.commit()

    signup, _ = cs.build_signup_rows(db_session, plan)
    discount, _ = cs.build_discount_rows(db_session, plan)

    assert signup == [{
        "taobao_item_id": "719436834260",
        "taobao_sku_id": "6282622238127",
        "sku_code": "PPS2398001060614",
        "price": 30.0,
        "is_placeholder": False,
        "remark": None,
    }]
    # 低价 SKU 的12%按平台精确值3.60元；30 - 3.60 - 5.99 = 20.41。
    assert discount[0]["official"] == 3.6
    assert discount[0]["deduct"] == 5.99
    assert discount[0]["target_price"] == 20.41


def test_super_reduce_low_price_uses_platform_exact_percent(db_session):
    """超级立减低价 SKU 也按精确10%：25×10%=2.50，不向上取整成3元。"""
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSLOW10", "PPSLOW1001", "9300", "73010",
        daily=25, big=20.41)
    db_session.commit()

    rows, stats = cs.build_discount_rows(db_session, plan)

    assert rows[0]["official"] == 2.5
    assert rows[0]["deduct"] == 1.48
    assert rows[0]["target_price"] == 21.02
    assert stats["official_low_price_exact"] == 1


def test_signup_rows_r3_incomplete_item_dropped(db_session):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSSC001", "PPSSC00101", "9203", "72021", daily=None)   # 缺日常价
    _mk(db_session, "PPSSC001", "PPSSC00102", "9203", "72022", daily=900)
    _mk(db_session, "PPSSD001", "PPSSD00101", "9204", "72031", daily=800)    # 对照: 完整品
    db_session.commit()

    rows, stats = cs.build_signup_rows(db_session, plan)

    sids = {r["taobao_sku_id"] for r in rows}
    assert sids == {"72031"}                                  # R3: 半套整品剔除
    assert len(stats["incomplete_items"]) == 1
    inc = stats["incomplete_items"][0]
    assert inc["taobao_item_id"] == "9203" and inc["ok_skus"] == 1
    assert "PPSSC00101" in inc["missing_skus"][0]


# ── ③④⑤⑥ 立减公式 ──────────────────────────────────────────────────────────

def test_discount_formula_big_spec_sample(db_session):
    """历史最低券后线不参与立减；最终到手只能等于 ERP 大促价。"""
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSDA001", "PPSDA00101", "9301", "73001",
        daily=2827.5, big=1979.59, line=1978.89)
    _mk(db_session, "PPSDB001", "PPSDB00101", "9302", "73002", daily=2000, big=1500)  # 无线
    db_session.commit()

    rows, stats = cs.build_discount_rows(db_session, plan)
    by_sid = {r["taobao_sku_id"]: r for r in rows}

    assert by_sid["73001"]["deduct"] == 507.91
    assert by_sid["73001"]["official"] == 340.0               # R9 向上取整到元
    assert by_sid["73001"]["target_price"] == 1979.59
    assert stats["line_concessions"] == []
    assert by_sid["73002"]["deduct"] == 260.0                 # 2000 − ceil(240)=240 − 1500


def test_discount_mid_ratio_and_ceil_switch(db_session):
    """超级立减: 中促 = 大促×1.03 就地算 (850→875.5); 10% ceil 开关 campaign_official_ceil。"""
    from app.services import settings_service
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSDC001", "PPSDC00101", "9303", "73011", daily=995, big=850)
    db_session.commit()

    rows, stats = cs.build_discount_rows(db_session, plan)
    assert stats["official_ceil"] is True                     # 默认向上取整
    assert rows[0]["deduct"] == 19.5                          # 995 − ceil(99.5)=100 − 875.5
    assert rows[0]["target_price"] == 875.5                   # 850 × 1.03

    settings_service.set_value(db_session, cs.OFFICIAL_CEIL_KEY, "0")
    db_session.commit()
    rows2, stats2 = cs.build_discount_rows(db_session, plan)
    assert stats2["official_ceil"] is False
    assert rows2[0]["deduct"] == 20.0                         # 995 − 99.5 − 875.5


def test_discount_nosales_big_direct_and_placeholder_skip(db_session):
    """无动销大促场不报名，单品立减直接到 ERP 大促价 2650；占位不出行。"""
    plan = _plan(db_session, "big88")
    plan.remark = "official_all_store=true; official_exempt_items=9304"
    _mk(db_session, "PPSDD001", "PPSDD00101", "9304", "73021", daily=3000, big=2650)
    _mk(db_session, "PPSDD001", "PPSDD00190", "9304", "73090", daily=500,
        placeholder=True)                                     # 占位不出行
    db_session.commit()
    ns.add_no_sales(db_session, ["9304"])

    rows, stats = cs.build_discount_rows(db_session, plan)

    assert len(rows) == 1
    assert rows[0]["kind"] == "nosales"
    assert rows[0]["target_price"] == 2650.0
    assert rows[0]["deduct"] == 350.0                         # 3000 − 大促价2650
    assert stats["skipped_placeholder"] == 1


def test_discount_nosales_big_storewide_includes_official_discount(db_session):
    plan = _plan(db_session, "big88")
    plan.remark = "official_all_store=true; official_exempt_items="
    _mk(db_session, "PPSDD002", "PPSDD00201", "9307", "73051", daily=3000, big=2000)
    db_session.commit()
    ns.add_no_sales(db_session, ["9307"])

    rows, _stats = cs.build_discount_rows(db_session, plan)

    assert rows[0]["official"] == 360.0
    assert rows[0]["target_price"] == 2000.0
    assert rows[0]["deduct"] == 640.0


def test_discount_nosales_super_reduce_uses_explicit_active_scope(db_session):
    plan = _plan(db_session, "super_reduce")
    plan.remark = "official_active_items=9308"
    _mk(db_session, "PPSDD003", "PPSDD00301", "9308", "73061", daily=3000, big=2500)
    db_session.commit()
    ns.add_no_sales(db_session, ["9308"])

    rows, _stats = cs.build_discount_rows(db_session, plan)

    assert rows[0]["official"] == 300.0
    assert rows[0]["target_price"] == 2575.0
    assert rows[0]["deduct"] == 125.0


def test_discount_uses_live_activity_price_after_platform_acceptance(db_session):
    """Current-window correction must use the platform activity base, not stale ERP daily."""
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        "official_active_items=1047741902625; "
        "current_activity_prices=6279984722445:2827.50"
    )
    _mk(
        db_session, "PPSLIVE01", "PPSLIVE0101", "1047741902625",
        "6279984722445", daily=3390, big=2367.35,
    )
    db_session.commit()

    rows, stats = cs.build_discount_rows(db_session, plan)

    assert rows[0]["calculation_base"] == 2827.5
    assert rows[0]["official"] == 283.0
    assert rows[0]["target_price"] == 2438.37
    assert rows[0]["deduct"] == 106.13
    assert stats["live_activity_price_overrides"] == [{
        "taobao_item_id": "1047741902625",
        "taobao_sku_id": "6279984722445",
        "daily_price": 3390.0,
        "activity_price": 2827.5,
    }]


def test_live_activity_evidence_promotes_known_failures_without_adopting_unknown(
        db_session):
    plan = _plan(db_session, "super_reduce")
    cs._set_plan_item_marker(
        plan, "platform_qualified_items", {"1000009301"})
    cs._set_plan_item_marker(
        plan, "platform_no_sales_items", {"1000009302"})
    cs._set_plan_item_marker(
        plan, "platform_hard_failed_items", {"1000009303"})
    cs._set_plan_item_marker(plan, "official_active_items", {"1000009301"})
    ns.add_no_sales(db_session, ["1000009302"])
    db_session.commit()

    evidence = cs.sync_live_activity_evidence(db_session, plan, [
        {"item_id": "1000009302", "sku_id": "73092", "activity_price": 2910},
        {"item_id": "1000009303", "sku_id": "73093", "activity_price": 3000},
        {"item_id": "1000009399", "sku_id": "73999", "activity_price": 1},
    ])

    assert evidence["promoted_items"] == ["1000009302", "1000009303"]
    assert evidence["unknown_active_items"] == ["1000009399"]
    assert cs.platform_qualified_items(plan) == {
        "1000009301", "1000009302", "1000009303",
    }
    assert cs.platform_no_sales_items(plan) == set()
    assert cs.platform_hard_failed_items(plan) == set()
    assert cs.official_scope_for_plan(plan)["active_items"] == {
        "1000009302", "1000009303",
    }
    assert cs.current_activity_prices_for_plan(plan) == {
        "73092": Decimal("2910.00"),
        "73093": Decimal("3000.00"),
    }
    assert ns.get_no_sales(db_session) == set()


def test_preflight_blocks_nosales_without_official_scope(db_session):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSDD004", "PPSDD00401", "9309", "73071", daily=3000, big=2500)
    db_session.commit()
    ns.add_no_sales(db_session, ["9309"])

    checks = {x["rule"]: x for x in cs.preflight(db_session, plan)}

    assert checks["R15"]["level"] == "error"
    assert "未记录官方立减" in checks["R15"]["items"][0]["errors"][0]


def test_placeholder_signup_blocks_high_live_price_without_expiry_confirmation(db_session):
    plan = _plan(db_session, "big88")
    plan.remark = (
        "official_active_items=9300; "
        "placeholder_live_prices=73081:397"
    )
    _mk(db_session, "PPSPH001", "PPSPH00199", "9310", "73081",
        daily=500, placeholder=True, line=250)
    db_session.commit()

    rows, stats = cs.build_signup_rows(db_session, plan)
    checks = {x["rule"]: x for x in cs.preflight(db_session, plan)}

    assert rows == []
    assert stats["placeholder_price_blocked_items"][0]["placeholders"][0] == {
        "taobao_item_id": "9310",
        "taobao_sku_id": "73081",
        "sku_code": "PPSPH00199",
        "safe_cap": 284.0,
        "current_live_price": 397.0,
    }
    assert checks["R16"]["level"] == "warn"
    assert "整品暂缓1品" in checks["R16"]["title"]


def test_placeholder_signup_uses_safe_cap_after_expiry_confirmation(db_session):
    plan = _plan(db_session, "big88")
    plan.remark = (
        "official_active_items=9300; "
        "placeholder_live_prices=73082:397; "
        "placeholder_price_protection_expired=true"
    )
    _mk(db_session, "PPSPH003", "PPSPH00399", "9312", "73082",
        daily=500, placeholder=True, line=250)
    db_session.commit()

    rows, stats = cs.build_signup_rows(db_session, plan)
    checks = {x["rule"]: x for x in cs.preflight(db_session, plan)}

    assert rows[0]["price"] == 284.0
    assert rows[0]["remark"] == "价保已确认到期，按最低普惠券后价安全上限报名"
    assert stats["placeholder_price_lowered"][0]["safe_cap"] == 284.0
    assert stats["placeholder_price_lowered"][0]["current_live_price"] == 397.0
    assert checks["R13"]["level"] == "pass"
    assert checks["R16"]["level"] == "pass"
    assert checks["R16"]["price_protection_expired"] is True


def test_placeholder_signup_above_safe_cap_fails_price_math(db_session):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSPH004", "PPSPH00499", "9313", "73083",
        daily=500, placeholder=True, line=250)
    db_session.commit()

    check = cs._check_price_math(
        db_session,
        plan,
        [{
            "taobao_item_id": "9313",
            "taobao_sku_id": "73083",
            "sku_code": "PPSPH00499",
            "price": 397.0,
            "is_placeholder": True,
        }],
        [],
    )

    assert check["level"] == "error"
    assert check["items"] == [{
        "sku_id": "73083",
        "sku_code": "PPSPH00499",
        "check": "placeholder_signup_within_coupon_floor_cap",
        "safe_cap": 284.0,
        "signup_price": 397.0,
    }]


def test_placeholder_price_block_notification_is_precise_and_deduped(
        db_session, monkeypatch):
    from app.services import notify_service

    plan = _plan(db_session, "big88")
    sent = []

    def fake_broadcast(_db, text, *, title, level):
        sent.append({"text": text, "title": title, "level": level})
        return {"feishu": True}

    monkeypatch.setattr(notify_service, "broadcast_text", fake_broadcast)
    blocked = [{
        "taobao_item_id": "9314",
        "product": "定制占位测试品",
        "placeholders": [{
            "taobao_sku_id": "73084",
            "sku_code": "PPSPH00599",
            "safe_cap": 284.0,
            "current_live_price": 397.0,
        }],
    }]

    first = cs._notify_placeholder_price_blocks(db_session, plan, blocked)
    second = cs._notify_placeholder_price_blocks(db_session, plan, blocked)

    assert first == {"feishu": True}
    assert second == {"deduped": True}
    assert len(sent) == 1
    assert sent[0]["title"] == "活动占位SKU因价保暂缓"
    assert sent[0]["level"] == "warning"
    assert all(value in sent[0]["text"] for value in (
        "9314", "73084", "PPSPH00599", "397.0", "284.0",
    ))


def test_placeholder_signup_without_live_price_is_blocked(db_session):
    plan = _plan(db_session, "big88")
    plan.remark = "official_active_items=9300"
    _mk(db_session, "PPSPH002", "PPSPH00299", "9311", "73091",
        daily=500, placeholder=True, line=250)
    db_session.commit()

    checks = {x["rule"]: x for x in cs.preflight(db_session, plan)}

    assert checks["R16"]["level"] == "error"
    assert checks["R16"]["items"][0]["taobao_sku_id"] == "73091"


def test_placeholder_missing_live_price_uses_safe_cap_after_user_authorization(db_session):
    plan = _plan(db_session, "big88")
    plan.remark = (
        "official_active_items=9300; "
        "placeholder_price_lowering_authorized=true"
    )
    _mk(db_session, "PPSPH006", "PPSPH00699", "9315", "73092",
        daily=500, placeholder=True, line=250)
    db_session.commit()

    rows, stats = cs.build_signup_rows(db_session, plan)
    checks = {x["rule"]: x for x in cs.preflight(db_session, plan)}

    assert rows[0]["price"] == 284.0
    assert rows[0]["remark"] == "用户已授权定制咨询规格使用保护报名价"
    assert stats["placeholder_price_lowered"][0]["authorization"] == "current_plan_user_decision"
    assert checks["R16"]["level"] == "pass"


def test_discount_price_hold_when_platform_coupon_after_exceeds_history_line(db_session):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSDE001", "PPSDE00101", "9305", "73031", daily=3000, big=2000, line=1990)
    db_session.commit()
    _seed_platform_floors(db_session, [("9305", "73031", 4000, 1990)])

    rows, stats = cs.build_discount_rows(db_session, plan)

    assert rows == []                                         # R2: 报名资格线冲突 → 整品不出行
    assert stats["excluded_price_hold_items"][0]["taobao_item_id"] == "9305"
    reason = stats["excluded_price_hold_items"][0]["skus"][0]["reasons"][0]
    assert reason["type"] == "coupon_floor"
    assert reason["taobao_sku_id"] == "73031"
    assert reason["erp_signup_price"] == 3000.0
    assert reason["official_rate"] == 0.12
    assert reason["official_deduction"] == 360.0
    assert reason["platform_coupon_after"] == 2000.0
    assert reason["platform_history_line"] == 1990.0
    assert reason["difference"] == 10.0
    assert reason["planned_single_item_discount"] == 640.0
    assert reason["single_item_discount_included_by_platform"] is True


def test_big_campaign_low_price_uses_platform_exact_percent(db_session):
    """¥30样块的12%平台实测为¥3.60，不套普通商品的整元向上取整。"""
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSLOW01", "PPSLOW0101", "9306", "73041",
        daily=30, big=20.41)
    db_session.commit()

    rows, stats = cs.build_discount_rows(db_session, plan)

    assert rows[0]["official"] == 3.6
    assert rows[0]["deduct"] == 5.99
    assert stats["official_low_price_exact"] == 1


# ── ⑦ preflight ─────────────────────────────────────────────────────────────

def test_preflight_outputs_r0_to_r19(db_session):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSPA001", "PPSPA00101", "9401", "74001",
        daily=1200, big=1000, enrolled=1100)                  # R1: 日常价 > 已生效价硬底
    _mk(db_session, "PPSPB001", "PPSPB00101", "9402", "74011", daily=None)   # R3: 缺价
    _mk(db_session, "PPSPB001", "PPSPB00102", "9402", "74012", daily=800)
    _mk(db_session, "PPSPC001", "PPSPC00101", "9403", "74021",
        daily=2000, big=1600, line=1590)                      # R2: 让幅10>1 → 暂缓
    db_session.commit()
    _seed_platform_floors(db_session, [
        ("9401", "74001", 1100, 2000),
        ("9403", "74021", 2500, 1590),
    ])
    ns.add_no_sales(db_session, ["9404"])                     # R6 名单

    checks = cs.preflight(db_session, plan)
    by_rule = {c["rule"]: c for c in checks}

    assert [c["rule"] for c in checks] == [f"R{i}" for i in range(0, 20)]
    assert all({"rule", "level", "title", "items"} <= set(c) for c in checks)
    assert by_rule["R1"]["level"] == "warn"
    assert by_rule["R1"]["items"][0]["skus"][0]["sku_code"] == "PPSPA00101"
    assert by_rule["R2"]["level"] == "warn"
    assert by_rule["R2"]["items"][0]["skus"][0]["sku_code"] == "PPSPC00101"
    assert by_rule["R3"]["level"] == "error"
    assert by_rule["R3"]["items"][0]["taobao_item_id"] == "9402"
    assert by_rule["R6"]["level"] == "warn" and "9404" in by_rule["R6"]["items"]
    assert by_rule["R9"]["items"] == [{"official_ceil": True}]
    assert by_rule["R11"]["level"] == "warn" and by_rule["R12"]["level"] == "warn"
    assert by_rule["R13"]["level"] == "pass"
    assert by_rule["R14"]["level"] == "warn"
    assert by_rule["R15"]["level"] == "error"
    assert by_rule["R16"]["level"] == "pass"
    assert by_rule["R17"]["level"] == "pass"


# ── ⑧ 推送编排 (mock WA) ─────────────────────────────────────────────────────

def _mock_wa(monkeypatch, calls):
    from app.services import web_agent_service

    def fake_upload(db, channel, phase, xlsx_bytes, filename, **kw):
        call = {"channel": channel, "phase": phase, "filename": filename,
                "xlsx_len": len(xlsx_bytes), **kw}
        if channel == "single_item_discount":
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
            try:
                call["item_ids"] = sorted({
                    str(row[0]).strip()
                    for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True)
                    if row and row[0] not in (None, "")
                })
            finally:
                wb.close()
        calls.append(call)
        return {"ok": True, "job": "job-1"}

    def fake_wait(db, job_id, **kw):
        return {"status": "done", "result": {"ok": True, "submitted": True,
                                             "validation": {
                                                 "total_items": 1, "ok": 1,
                                                 "failed": 0, "terminal": True,
                                                 "failed_items": []}}}

    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 4):
        ws.cell(r, 1, f"表头{r}")
    bio = io.BytesIO()
    wb.save(bio)

    monkeypatch.setattr(web_agent_service, "upload_file", fake_upload)
    monkeypatch.setattr(web_agent_service, "wait_job", fake_wait)
    monkeypatch.setattr(web_agent_service, "campaign_export_items",
                        lambda db, title, **kw: {
                            "ok": True, "xlsx_bytes": bio.getvalue(),
                            "filename": "当前活动.xlsx"})
    # Generic orchestration tests mock an empty platform export.  The exact
    # post-submit SKU verifier has focused tests below; keep these tests scoped
    # to channel/state orchestration rather than fabricating a second workbook.
    monkeypatch.setattr(
        cs, "_verify_super_signup_rows",
        lambda expected, live: {
            "ok": True, "checked_real_skus": len(expected),
            "failed_real_skus": 0, "failures": [],
        },
    )


def test_push_discount_orchestration(db_session, monkeypatch):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSQA001", "PPSQA00101", "9501", "75001", daily=2827.5, big=1979.59)
    db_session.commit()
    calls = []
    _mock_wa(monkeypatch, calls)

    res = cs.push_discount(db_session, plan)                  # 默认 stage: 不推进状态机
    assert res["ok"] is True and plan.status == "draft"
    assert calls[0]["channel"] == "single_item_discount" and calls[0]["phase"] == "stage"
    assert calls[0]["start_dt"] == "2026-07-17 20:00:00"      # 档期精确到秒
    assert calls[0]["end_dt"] == "2026-07-19 23:59:59"

    res2 = cs.push_discount(db_session, plan, phase="commit")
    assert res2["ok"] is True and plan.status == "discount_pushed"
    assert calls[1]["phase"] == "commit"


def test_authorized_supplement_scope_limits_discount_and_signup_uploads(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        f"{plan.remark or ''}; supplement_items_authorized=100000009501; "
        "platform_qualified_items=100000009502; "
        "platform_hard_failed_items=100000009501; "
        "official_active_items=100000009502"
    )
    _mk(db_session, "PPSQSCOPE1", "PPSQSCOPE101", "100000009501", "75101",
        daily=1500, big=1000)
    _mk(db_session, "PPSQSCOPE2", "PPSQSCOPE201", "100000009502", "75201",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        ("100000009501", "75101", 2000, 1030),
        ("100000009502", "75201", 2000, 1133),
    ])
    calls = []
    _mock_wa(monkeypatch, calls)

    discount = cs.push_discount(db_session, plan, phase="commit")
    signup = cs.push_signup(
        db_session, plan, execution_source="campaign_automation")

    assert discount["ok"] is True
    assert discount["stats"]["authorized_supplement_items"] == ["100000009501"]
    assert discount["stats"]["platform_discount_scope_rows"] == 1
    assert calls[0]["channel"] == "single_item_discount"
    assert calls[0]["expected_rows"] == 1
    assert signup["ok"] is True
    assert signup["stats"]["authorized_supplement_items"] == ["100000009501"]
    assert signup["stats"]["platform_qualified_rows"] == 1
    assert signup["stats"]["pending_items"] == ["100000009501"]
    assert calls[1]["channel"] == "super_reduce"
    assert calls[1]["expected_rows"] == 1
    assert cs.authorized_supplement_items(plan) == set()


def test_super_reduce_supplement_preserves_known_active_items(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        "supplement_items_authorized=100000009501; "
        "platform_qualified_items=100000009501,100000009502; "
        "official_active_items=100000009501,100000009502"
    )
    _mk(db_session, "PPSQSCOPE3", "PPSQSCOPE301", "100000009501", "75401",
        daily=1500, big=1000)
    _mk(db_session, "PPSQSCOPE4", "PPSQSCOPE401", "100000009502", "75402",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        ("100000009501", "75401", 2000, 1030),
        ("100000009502", "75402", 2000, 1133),
    ])
    calls = []
    _mock_wa(monkeypatch, calls)
    monkeypatch.setattr(
        cs,
        "refresh_floor_evidence_from_current_activity",
        lambda *args, **kwargs: {
            "ok": True,
            "rows": [
                {
                    "item_id": "100000009501",
                    "sku_id": "75401",
                    "status": "暂停",
                    "activity_price": 1500,
                },
                {
                    "item_id": "100000009502",
                    "sku_id": "75402",
                    "status": "活动中",
                    "activity_price": 1600,
                },
            ],
            "floor_refresh": {"observed": 2},
        },
    )

    result = cs.push_signup(
        db_session, plan, execution_source="campaign_automation")

    assert result["ok"] is True
    assert result["stats"]["pending_items"] == ["100000009501"]
    assert result["stats"]["preserved_active_items_outside_supplement"] == [
        "100000009502"
    ]
    assert calls[0]["channel"] == "super_reduce"
    assert calls[0]["expected_rows"] == 1


def test_super_reduce_exact_authorized_sku_refresh_reimports_complete_item(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    item_id = "100000009503"
    plan.remark = (
        f"supplement_items_authorized={item_id}; "
        f"platform_qualified_items={item_id}; "
        f"sku_refresh_items_authorized={item_id}"
    )
    _mk(db_session, "PPSQROT1", "PPSQROT101", item_id, "75601",
        daily=1500, big=1000)
    _mk(db_session, "PPSQROT1", "PPSQROT102", item_id, "75602",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(
        db_session, [(item_id, "75601", 2000, 1030)])
    calls = []
    _mock_wa(monkeypatch, calls)
    monkeypatch.setattr(
        cs,
        "refresh_floor_evidence_from_current_activity",
        lambda *args, **kwargs: {
            "ok": True,
            "rows": [{
                "item_id": item_id,
                "sku_id": "75601",
                "status": "活动中",
                "activity_price": 1500,
            }],
            "floor_refresh": {"observed": 1},
        },
    )
    monkeypatch.setattr(cs, "preflight", lambda *args, **kwargs: [])

    result = cs.push_signup(
        db_session, plan, execution_source="campaign_automation")

    assert result["ok"] is True
    assert result["stats"]["pending_items"] == [item_id]
    assert result["stats"]["authorized_sku_refresh_existing_items"] == [{
        "item_id": item_id,
        "error": "用户确认SKU轮换，允许完整SKU集合原位重导",
        "missing_skus": ["75602"],
    }]
    assert calls[0]["channel"] == "super_reduce"
    assert calls[0]["expected_rows"] == 2


def test_super_reduce_supplement_still_blocks_unknown_active_items(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        "supplement_items_authorized=100000009501; "
        "platform_qualified_items=100000009501; "
        "official_active_items=100000009501"
    )
    _mk(db_session, "PPSQSCOPE5", "PPSQSCOPE501", "100000009501", "75501",
        daily=1500, big=1000)
    db_session.commit()
    _seed_platform_floors(
        db_session, [("100000009501", "75501", 2000, 1030)])
    calls = []
    _mock_wa(monkeypatch, calls)
    monkeypatch.setattr(
        cs,
        "refresh_floor_evidence_from_current_activity",
        lambda *args, **kwargs: {
            "ok": True,
            "rows": [{"item_id": "100000009599", "status": "活动中"}],
            "floor_refresh": {"observed": 1},
        },
    )

    result = cs.push_signup(
        db_session, plan, execution_source="campaign_automation")

    assert result["ok"] is False
    assert result["step"] == "super_reduce_unexpected_active_scope_guard"
    assert result["unexpected_active_items"] == ["100000009599"]
    assert calls == []


def test_existing_single_discount_activity_id_is_forwarded_for_modify(db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        f"{plan.remark or ''}; supplement_items_authorized=100000009501; "
        "single_discount_activity_id=142591608100"
    )
    _mk(db_session, "PPSQEDIT1", "PPSQEDIT101", "100000009501", "75301",
        daily=1500, big=1000)
    db_session.commit()
    _seed_platform_floors(db_session, [("100000009501", "75301", 2000, 1030)])
    calls = []
    _mock_wa(monkeypatch, calls)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is True
    assert calls[0]["channel"] == "single_item_discount"
    assert calls[0]["campaign_id"] == "142591608100"
    assert calls[0]["expected_rows"] == 1
    assert result["stats"]["single_discount_execution_mode"] == "per_item_existing_then_new_batch"


def test_rotated_skus_bypass_old_discount_drawer_and_use_new_batch(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    item_id = "100000009510"
    plan.remark = (
        f"single_discount_activity_ids={item_id}:142591608100; "
        f"sku_refresh_items_authorized={item_id}"
    )
    _mk(db_session, "PPSQROT2", "PPSQROT201", item_id, "75901",
        daily=1500, big=1000)
    db_session.commit()
    _seed_platform_floors(db_session, [(item_id, "75901", 2000, 1030)])
    calls = []
    _mock_wa(monkeypatch, calls)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is True
    assert len(calls) == 1
    assert calls[0]["channel"] == "single_item_discount"
    assert calls[0].get("campaign_id") is None
    assert calls[0]["item_ids"] == [item_id]
    assert result["stats"]["single_discount_existing_items"] == []
    assert result["stats"]["single_discount_new_items"] == [item_id]
    assert result["stats"]["single_discount_rotated_new_batch_items"] == [item_id]


def test_partial_new_batch_stops_without_resending_successes(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    rotated_item = "100000009515"
    existing_item = "100000009516"
    plan.remark = (
        "single_discount_activity_ids=100000009599:142591608100; "
        f"sku_refresh_items_authorized={rotated_item}; "
        f"supplement_items_authorized={rotated_item},{existing_item}"
    )
    _mk(db_session, "PPSQROT5", "PPSQROT501", rotated_item, "75951",
        daily=1500, big=1000)
    _mk(db_session, "PPSQEDIT6", "PPSQEDIT601", existing_item, "75961",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        (rotated_item, "75951", 2000, 1030),
        (existing_item, "75961", 2000, 1133),
    ])
    calls = []

    def fake_upload(*args, **kwargs):
        calls.append({
            "campaign_id": kwargs.get("discount_activity_id"),
            "expected_rows": kwargs.get("expected_rows"),
        })
        if len(calls) == 1:
            return {
                "ok": False,
                "submitted": False,
                "final_import": {
                    "ok": 1,
                    "failed": 1,
                    "failed_rows": [{
                        "item_id": existing_item,
                        "sku_id": "75961",
                        "reason": "已经参加了单品立减活动，id：142591608100",
                    }],
                },
            }
        return {"ok": True, "submitted": True}

    monkeypatch.setattr(cs, "_upload_and_wait", fake_upload)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is False
    assert result["step"] == "single_item_discount_partial_import"
    assert result["discovered_existing_activity_ids"] == {
        existing_item: "142591608100",
    }
    assert len(calls) == 1
    assert cs._plan_single_discount_activity_ids(plan) == {
        "100000009599": "142591608100",
    }


def test_existing_activity_conflict_rejects_partial_sku_evidence():
    item_id = "100000009517"
    rows = {
        item_id: [
            {"taobao_sku_id": "75971"},
            {"taobao_sku_id": "75972"},
        ]
    }
    result = {
        "submitted": False,
        "final_import": {
            "failed": 1,
            "failed_rows": [{
                "item_id": item_id,
                "sku_id": "75971",
                "reason": "已经参加了单品立减活动，id：142591608100",
            }],
        },
    }

    assert cs._single_discount_existing_activity_conflicts(
        result, rows, {"142591608100"}) == {}


def test_rotated_conflict_recovers_new_activity_and_uses_exact_editor(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    item_id = "100000009518"
    plan.remark = (
        f"sku_refresh_items_authorized={item_id}; "
        f"supplement_items_authorized={item_id}"
    )
    _mk(db_session, "PPSQROT8", "PPSQROT801", item_id, "75981",
        daily=1500, big=1000)
    db_session.commit()
    _seed_platform_floors(db_session, [(item_id, "75981", 2000, 1030)])
    calls = []

    def fake_upload(*args, **kwargs):
        calls.append({"campaign_id": kwargs.get("discount_activity_id")})
        if len(calls) == 1:
            return {
                "ok": False,
                "submitted": False,
                "final_import": {
                    "ok": 0,
                    "failed": 1,
                    "failed_rows": [{
                        "item_id": item_id,
                        "sku_id": "75981",
                        "reason": "已经参加了单品立减活动，id：142797717830",
                    }],
                },
            }
        return {"ok": True, "submitted": True}

    monkeypatch.setattr(cs, "_upload_and_wait", fake_upload)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is True
    assert calls == [
        {"campaign_id": None},
        {"campaign_id": "142797717830"},
    ]
    assert cs._plan_single_discount_activity_ids(plan)[item_id] == "142797717830"
    assert cs._plan_single_discount_refreshed_activity_ids(plan)[item_id] == "142797717830"


def test_split_sku_activities_are_reconciled_and_edited_separately(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    item_id = "100000009548"
    plan.remark = (
        f"sku_refresh_items_authorized={item_id}; "
        f"supplement_items_authorized={item_id}"
    )
    _mk(db_session, "PPSQSPLIT1", "PPSQSPLIT101", item_id, "75481001",
        daily=1500, big=1000)
    _mk(db_session, "PPSQSPLIT1", "PPSQSPLIT102", item_id, "75481002",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        (item_id, "75481001", 2000, 1030),
        (item_id, "75481002", 2000, 1133),
    ])
    calls = []

    def fake_upload(*args, **kwargs):
        calls.append({
            "campaign_id": kwargs.get("discount_activity_id"),
            "expected_rows": kwargs.get("expected_rows"),
        })
        if len(calls) == 1:
            return {
                "ok": False,
                "submitted": False,
                "final_import": {
                    "ok": 0,
                    "failed": 2,
                    "failed_rows": [
                        {
                            "item_id": item_id,
                            "sku_id": "75481001",
                            "reason": "已经参加了单品立减活动，id：142591608100",
                        },
                        {
                            "item_id": item_id,
                            "sku_id": "75481002",
                            "reason": "已经参加了单品立减活动，id：142834680634",
                        },
                    ],
                },
            }
        return {"ok": True, "submitted": True}

    monkeypatch.setattr(cs, "_upload_and_wait", fake_upload)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is True
    assert calls == [
        {"campaign_id": None, "expected_rows": 2},
        {"campaign_id": "142591608100", "expected_rows": 1},
        {"campaign_id": "142834680634", "expected_rows": 1},
    ]
    assert cs._plan_single_discount_sku_activity_ids(plan) == {
        "75481001": "142591608100",
        "75481002": "142834680634",
    }
    assert item_id not in cs._plan_single_discount_activity_ids(plan)


def test_partial_final_import_is_not_retried_blindly(db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    created_item = "100000009519"
    existing_item = "100000009520"
    plan.remark = (
        f"sku_refresh_items_authorized={created_item}; "
        f"supplement_items_authorized={created_item},{existing_item}"
    )
    _mk(db_session, "PPSQPART9", "PPSQPART901", created_item, "75991",
        daily=1500, big=1000)
    _mk(db_session, "PPSQPART0", "PPSQPART001", existing_item, "76001",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        (created_item, "75991", 2000, 1030),
        (existing_item, "76001", 2000, 1133),
    ])
    calls = []

    def fake_upload(*args, **kwargs):
        calls.append(kwargs.get("discount_activity_id"))
        return {
            "ok": False,
            "submitted": True,
            "final_import": {
                "ok": 1,
                "failed": 1,
                "failed_rows": [{
                    "item_id": existing_item,
                    "sku_id": "76001",
                    "reason": "已经参加了单品立减活动，id：142591608100",
                }],
            },
        }

    monkeypatch.setattr(cs, "_upload_and_wait", fake_upload)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is False
    assert result["step"] == "single_item_discount_partial_import"
    assert result["committed_rows"] == 1
    assert calls == [None]
    assert cs._plan_single_discount_activity_ids(plan) == {}


def test_corrective_preflight_checks_only_authorized_upload_scope(db_session):
    plan = _plan(db_session, "super_reduce")
    target_item = "100000009523"
    unrelated_item = "100000009524"
    plan.remark = f"supplement_items_authorized={target_item}"
    _mk(db_session, "PPSQSCOPE3", "PPSQSCOPE301", target_item, "76031",
        daily=1500, big=1000)
    _mk(db_session, "PPSQSCOPE4", "PPSQSCOPE401", unrelated_item, "76041",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [(target_item, "76031", 2000, 1030)])

    checks = cs.preflight(db_session, plan)
    r17 = next(check for check in checks if check["rule"] == "R17")

    assert r17["level"] == "pass"
    assert r17["checked"] == 1
    assert r17["items"] == []


def test_existing_single_discount_activity_is_modified_one_item_per_job(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        f"{plan.remark or ''}; "
        "single_discount_activity_ids=100000009511:142591608100"
    )
    _mk(db_session, "PPSQEDIT2", "PPSQEDIT201", "100000009511", "75311",
        daily=1500, big=1000)
    _mk(db_session, "PPSQEDIT3", "PPSQEDIT301", "100000009512", "75312",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        ("100000009511", "75311", 2000, 1030),
        ("100000009512", "75312", 2000, 1133),
    ])
    calls = []
    _mock_wa(monkeypatch, calls)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is True
    assert result["processed_items"] == 2
    assert len(calls) == 2
    assert [call["expected_rows"] for call in calls] == [1, 1]
    assert calls[0]["campaign_id"] == "142591608100"
    assert calls[1].get("campaign_id") is None
    assert [call["item_ids"] for call in calls] == [
        ["100000009511"], ["100000009512"]]
    assert result["stats"]["single_discount_expected_items"] == 2
    assert result["stats"]["single_discount_processed_items"] == 2
    assert result["stats"]["single_discount_existing_items"] == ["100000009511"]
    assert result["stats"]["single_discount_new_items"] == ["100000009512"]


def test_existing_single_discount_stops_before_signup_when_one_item_fails(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = (
        f"{plan.remark or ''}; single_discount_activity_ids="
        "100000009521:142591608100,100000009522:142591608101"
    )
    _mk(db_session, "PPSQEDIT4", "PPSQEDIT401", "100000009521", "75321",
        daily=1500, big=1000)
    _mk(db_session, "PPSQEDIT5", "PPSQEDIT501", "100000009522", "75322",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        ("100000009521", "75321", 2000, 1030),
        ("100000009522", "75322", 2000, 1133),
    ])
    calls = []

    def fake_upload(*args, **kwargs):
        calls.append(kwargs.get("expected_rows"))
        if len(calls) == 2:
            return {"ok": False, "error": "平台逐 SKU 回读不一致"}
        return {"ok": True, "submitted": True}

    monkeypatch.setattr(cs, "_upload_and_wait", fake_upload)

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is False
    assert result["step"] == "single_item_discount_per_item"
    assert result["failed_item_id"] == "100000009522"
    assert result["completed_item_ids"] == ["100000009521"]
    assert plan.status == "draft"


def test_legacy_single_discount_activity_id_rejects_multi_item_scope(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.remark = f"{plan.remark or ''}; single_discount_activity_id=142591608100"
    _mk(db_session, "PPSQEDIT6", "PPSQEDIT601", "100000009531", "75331",
        daily=1500, big=1000)
    _mk(db_session, "PPSQEDIT7", "PPSQEDIT701", "100000009532", "75332",
        daily=1600, big=1100)
    db_session.commit()
    _seed_platform_floors(db_session, [
        ("100000009531", "75331", 2000, 1030),
        ("100000009532", "75332", 2000, 1133),
    ])
    calls = []
    monkeypatch.setattr(cs, "_upload_and_wait", lambda *args, **kwargs: calls.append(1))

    result = cs.push_discount(db_session, plan, phase="commit")

    assert result["ok"] is False
    assert result["step"] == "single_discount_activity_binding_guard"
    assert calls == []


def test_push_signup_orchestration_and_empty_guard(db_session, monkeypatch):
    empty_plan = _plan(db_session, "big88")
    calls = []
    _mock_wa(monkeypatch, calls)

    empty = cs.push_signup(
        db_session, empty_plan, execution_source="campaign_automation")
    assert empty["ok"] is False and calls == []

    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSQB001", "PPSQB00101", "9502", "75011", daily=1500)
    db_session.commit()
    _seed_platform_floors(db_session, [("9502", "75011", 2000, 1400)])
    res = cs.push_signup(db_session, plan, execution_source="campaign_automation")
    assert res["ok"] is True
    assert plan.status == "signup_pushed"                     # R12: stage 即生效
    assert calls[0]["channel"] == "promo_signup" and calls[0]["phase"] == "stage"


def test_super_reduce_signup_uses_dedicated_commit_channel(db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSQS001", "PPSQS00101", "9504", "75031", daily=1500, big=1000)
    db_session.commit()
    _seed_platform_floors(db_session, [("9504", "75031", 2000, 1400)])
    calls = []
    _mock_wa(monkeypatch, calls)

    result = cs.push_signup(
        db_session, plan, execution_source="campaign_automation")

    assert result["ok"] is True
    assert plan.status == "signup_pushed"
    assert calls[0]["channel"] == "super_reduce"
    assert calls[0]["phase"] == "commit"


def test_super_reduce_publish_window_does_not_infer_delay_from_plan_dates(db_session):
    plan = _plan(db_session, "super_reduce")
    plan.start_at = datetime.now() + timedelta(days=3)
    plan.end_at = plan.start_at + timedelta(days=4)
    check = cs._check_super_reduce_publish_window(plan)

    assert check["rule"] == "R18"
    assert check["level"] == "pass"
    assert check["items"] == []


def test_super_reduce_pairing_uses_actual_signup_price_for_official_discount(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSQACTUAL", "PPSQACTUAL01", "1001", "2001",
        daily=40, big="29.12621359")
    db_session.commit()
    monkeypatch.setattr(cs, "platform_qualified_items", lambda _plan: {"1001"})
    monkeypatch.setattr(cs, "official_ceil_enabled", lambda _db: False)

    check = cs._check_super_reduce_discount_coverage(
        db_session, plan,
        [{
            "taobao_item_id": "1001",
            "taobao_sku_id": "2001",
            "sku_code": "SAMPLE",
            "price": 30,
            "is_placeholder": False,
        }],
        [],
    )

    assert check["level"] == "error"
    assert check["items"][0]["check"] == "official_discount_already_below_erp_target"
    assert check["items"][0]["after_official"] == 27.0


def test_super_reduce_pairing_gate_blocks_missing_and_incompatible_skus(db_session):
    plan = _plan(db_session, "super_reduce")
    _mk(db_session, "PPSQPAIR1", "PPSQPAIR101", "9510", "75101",
        daily=1500, big=1000)
    _mk(db_session, "PPSQPAIR2", "PPSQPAIR201", "9511", "75102",
        daily=30, big=30)
    db_session.commit()

    signup_rows, _ = cs.build_signup_rows(db_session, plan)
    discount_rows, _ = cs.build_discount_rows(db_session, plan)
    discount_rows = [row for row in discount_rows if row["taobao_sku_id"] != "75101"]
    check = cs._check_super_reduce_discount_coverage(
        db_session, plan, signup_rows, discount_rows)

    assert check["level"] == "error"
    assert {row["check"] for row in check["items"]} == {
        "missing_paired_single_item_discount",
        "official_discount_already_below_erp_target",
    }


def test_super_reduce_early_activation_repair_requires_three_proofs(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    target = "840659847455"
    plan.remark = f"user_authorized_campaign_withdrawal={target}"
    exports = iter([
        {"ok": True, "rows": [{"item_id": target, "status": "活动中"}]},
        {"ok": True, "rows": [{"item_id": target, "status": "暂停"}]},
    ])
    monkeypatch.setattr(cs, "refresh_floor_evidence_from_current_activity",
                        lambda *args, **kwargs: next(exports))
    from app.services import web_agent_service
    calls = []

    def withdraw(db, item_ids, *, phase):
        calls.append((phase, item_ids))
        return {
            "ok": True,
            "item_results": [{
                "item_id": target,
                "result": "ready" if phase == "stage" else "withdrawn",
            }],
        }

    monkeypatch.setattr(web_agent_service, "withdraw_super_reduce_items", withdraw)

    result = cs.repair_super_reduce_early_activation(
        db_session,
        plan,
        [target],
        phase="commit",
        execution_source="campaign_automation_repair",
    )

    assert result["ok"] is True
    assert result["proof_3_remaining_active_items"] == []
    assert calls == [("stage", [target]), ("commit", [target])]
    assert plan.status == "discount_pushed"


def test_super_reduce_early_activation_repair_rejects_non_program_call(db_session):
    plan = _plan(db_session, "super_reduce")

    result = cs.repair_super_reduce_early_activation(
        db_session, plan, ["840659847455"], phase="commit")

    assert result["ok"] is False
    assert result["step"] == "execution_policy_guard"


def test_super_reduce_repair_commit_normalizes_plan_when_platform_is_already_clear(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    plan.status = "signup_pushed"
    target = "840659847455"
    plan.remark = f"user_authorized_campaign_withdrawal={target}"
    monkeypatch.setattr(
        cs,
        "refresh_floor_evidence_from_current_activity",
        lambda *args, **kwargs: {
            "ok": True,
            "rows": [{"item_id": target, "status": "暂停"}],
        },
    )

    result = cs.repair_super_reduce_early_activation(
        db_session,
        plan,
        [target],
        phase="commit",
        execution_source="campaign_automation_repair",
    )

    assert result["ok"] is True
    assert result["no_change"] is True
    assert result["plan_status"] == "discount_pushed"
    assert plan.status == "discount_pushed"
    assert "super_reduce_early_activation_already_clear=840659847455" in plan.remark


def test_super_reduce_repair_rejects_missing_exact_user_withdrawal_authorization(
        db_session, monkeypatch):
    plan = _plan(db_session, "super_reduce")
    from app.services import web_agent_service
    calls = []
    monkeypatch.setattr(
        cs,
        "refresh_floor_evidence_from_current_activity",
        lambda *args, **kwargs: calls.append("export") or {"ok": True, "rows": []},
    )
    monkeypatch.setattr(
        web_agent_service,
        "withdraw_super_reduce_items",
        lambda *args, **kwargs: calls.append("withdraw") or {"ok": True},
    )

    result = cs.repair_super_reduce_early_activation(
        db_session,
        plan,
        ["840659847455"],
        phase="commit",
        execution_source="campaign_automation_repair",
    )

    assert result["ok"] is False
    assert result["step"] == "explicit_withdrawal_authorization_guard"
    assert calls == []


def test_push_signup_zero_zero_is_failure_and_feishu_deduped(db_session, monkeypatch):
    plan = _plan(db_session, "big88")
    _mk(db_session, "PPSQZ001", "PPSQZ00101", "9503", "75021",
        daily=1500, big=1000)
    db_session.commit()
    _seed_platform_floors(db_session, [("9503", "75021", 2000, 1400)])

    calls = []
    _mock_wa(monkeypatch, calls)
    from app.services import notify_service, web_agent_service
    monkeypatch.setattr(
        web_agent_service, "wait_job",
        lambda db, job_id, **kw: {
            "status": "done",
            "result": {
                "ok": True, "attached": True,
                "validation": {
                    "total_items": 1, "ok": 0, "failed": 0,
                    "terminal": False,
                },
            },
        })
    notices = []
    monkeypatch.setattr(
        notify_service, "broadcast_text",
        lambda db, text, **kw: notices.append({"text": text, **kw})
        or {"feishu": True})

    first = cs.push_signup(db_session, plan, execution_source="campaign_automation")
    second = cs.push_signup(db_session, plan, execution_source="campaign_automation")

    assert first["ok"] is False
    assert "未进入终态" in first["error"]
    assert plan.status == "alarmed"
    assert len(notices) == 1
    assert "总1品，成功0品，失败0品" in notices[0]["text"]
    assert second["step"] == "waiting_user_decision"
    assert second["automatic_retry"] is False


def test_super_signup_row_verification_rejects_item_level_active_with_blank_new_skus():
    expected = [
        {"taobao_item_id": "1047741358718", "taobao_sku_id": "6291475451145",
         "price": 6472.5, "is_placeholder": False},
        {"taobao_item_id": "1047741358718", "taobao_sku_id": "6241061986676",
         "price": 388.0, "is_placeholder": True},
    ]
    live = [
        {"item_id": "1047741358718", "sku_id": "6291475451145",
         "status": "活动中", "activity_price": None},
        {"item_id": "1047741358718", "sku_id": "6241061986676",
         "status": "活动中", "activity_price": 397.0},
    ]

    result = cs._verify_super_signup_rows(expected, live)

    assert result["ok"] is False
    assert result["checked_real_skus"] == 1
    assert result["failures"] == [{
        "item_id": "1047741358718",
        "sku_id": "6291475451145",
        "expected_activity_price": 6472.5,
        "actual_activity_prices": [None],
        "error": "活动价为空或不一致",
    }]


def test_super_signup_row_verification_accepts_any_exact_active_marketing_record():
    expected = [{
        "taobao_item_id": "840643621692", "taobao_sku_id": "6291731711010",
        "price": 4852.5, "is_placeholder": False,
    }]
    live = [
        {"item_id": "840643621692", "sku_id": "6291731711010",
         "status": "活动中", "activity_price": None},
        {"item_id": "840643621692", "sku_id": "6291731711010",
         "status": "活动中", "activity_price": 4852.5},
    ]

    result = cs._verify_super_signup_rows(expected, live)

    assert result["ok"] is True
    assert result["failed_real_skus"] == 0

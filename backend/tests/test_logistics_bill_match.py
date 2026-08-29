"""物流费账单逐单行 → 淘宝订单 自动配对测试 (用户 2026-06-21)。

覆盖: 运单号命中 / 姓名+省市命中 / 同名异地→未匹配 / 关闭单不参与 / summary行不配 / 多候选。
"""
from datetime import date
from decimal import Decimal

from app.models.finance import LogisticsBill
from app.models.order import Order
from app.services import logistics_bill_match


def _order(db, no, name, addr, *, status="signed", track=None):
    db.add(Order(
        platform="淘宝", order_no=no, qty=1, status=status,
        order_date=date(2026, 6, 1), paid_amount=Decimal("100"),
        customer_name=name, customer_address=addr, tracking_no=track,
    ))
    db.flush()


def _bill(db, **kw):
    kw.setdefault("freight_amount", Decimal("30"))
    kw.setdefault("row_type", "line")
    b = LogisticsBill(bill_date=date(2026, 6, 10), carrier="德邦", **kw)
    db.add(b)
    db.flush()
    return b


def test_match_by_tracking_no(db_session):
    """运单号全等 → 最可靠, 直接命中."""
    _order(db_session, "O1", "张三", "广东省深圳市南山区", track="DB123")
    b = _bill(db_session, tracking_no="DB123", recipient_name="张三", destination="广东省深圳市")
    r = logistics_bill_match.match_logistics_bills(db_session)
    assert r["matched"] == 1
    assert b.order_no == "O1"
    assert b.match_method == "track"


def test_match_by_name_and_province(db_session):
    """无运单号命中, 但收货人同名 + 目的地省市在订单地址里 → name_prov."""
    _order(db_session, "O2", "李四", "浙江省杭州市西湖区文一路")
    b = _bill(db_session, tracking_no="ZZZ", recipient_name="李四", destination="浙江省杭州市")
    r = logistics_bill_match.match_logistics_bills(db_session)
    assert r["matched"] == 1
    assert b.order_no == "O2"
    assert b.match_method == "name_prov"


def test_same_name_other_province_is_unmatched_when_ambiguous(db_session):
    """同名但有多个候选且目的地省市对不上 → multi 或 none, 不乱配."""
    _order(db_session, "O3a", "王五", "北京市朝阳区")
    _order(db_session, "O3b", "王五", "四川省成都市")
    b = _bill(db_session, tracking_no="NO", recipient_name="王五", destination="广东省广州市")
    logistics_bill_match.match_logistics_bills(db_session)
    assert b.order_no is None
    assert b.match_method in ("multi", "none")


def test_cancelled_order_not_matched(db_session):
    """关闭单按铁律排除: 收货人只在关闭单里 → 不命中, 标 none."""
    _order(db_session, "O4", "赵六", "江苏省南京市", status="cancelled", track="DBX")
    b = _bill(db_session, tracking_no="DBX", recipient_name="赵六", destination="江苏省南京市")
    r = logistics_bill_match.match_logistics_bills(db_session)
    assert b.order_no is None
    assert b.match_method == "none"
    assert r["none"] == 1


def test_summary_row_not_matched(db_session):
    """月结汇总行 (row_type='summary') 不参与配对."""
    _order(db_session, "O5", "钱七", "湖南省长沙市", track="DBS")
    s = LogisticsBill(bill_date=date(2026, 6, 30), carrier="德邦", row_type="summary",
                      freight_amount=Decimal("14540"), tracking_no=None)
    db_session.add(s)
    db_session.flush()
    logistics_bill_match.match_logistics_bills(db_session)
    assert s.match_method is None
    assert s.order_no is None


def test_import_yimidida_per_order_and_manual_orderno(db_session):
    """壹米滴答账单(非德邦文件名)有逐单运单号+运费 → 逐单导入(carrier=壹米滴答);
    『匹配订单号』列填了的行 → order_no + manual(自动配单不覆盖); 文件名总额 → 一条 summary。"""
    from datetime import datetime
    from openpyxl import Workbook
    from app.services import bill_import_service as bi
    wb = Workbook()
    ws = wb.active
    ws.append(["运单号", "寄件时间", "计费重量", "收件人姓名", "收件人省市区", "运费", "匹配订单号"])
    ws.append(["700897665404", datetime(2026, 1, 2), 267, "施施", "上海-上海市-浦东新区-三林镇", 269, "5052881700882748539"])
    ws.append(["700897667062", datetime(2026, 1, 3), 275, "刘岳", "黑龙江省-哈尔滨市-南岗区-王岗镇", 515, ""])
    rep = bi.import_logistics_xlsx(db_session, wb, source_name="李爱群 2026年1月账单 14540元.xlsx")
    assert rep.inserted == 3   # 2 逐单 + 1 汇总
    lines = db_session.query(LogisticsBill).filter_by(row_type="line").all()
    assert len(lines) == 2
    assert all(b.carrier == "壹米滴答" for b in lines)
    b1 = next(b for b in lines if b.recipient_name == "施施")
    assert b1.order_no == "5052881700882748539"     # 人工填的订单号被采用
    assert b1.match_method == "manual"
    assert b1.freight_amount == Decimal("269")
    b2 = next(b for b in lines if b.recipient_name == "刘岳")
    assert b2.order_no is None                       # 没填 + 库里无此单 → 自动配单标 none
    summ = db_session.query(LogisticsBill).filter_by(row_type="summary").all()
    assert len(summ) == 1 and summ[0].freight_amount == Decimal("14540")


def test_reimport_logistics_no_duplicate(db_session):
    """重导同一份壹米滴答账单 → 逐单+汇总都不重复(归一化业务键去重, 不受 sync_key 事件覆盖/14540 vs 14540.00 影响)。"""
    from datetime import datetime
    from openpyxl import Workbook
    from app.services import bill_import_service as bi

    def mk():
        wb = Workbook(); ws = wb.active
        ws.append(["运单号", "寄件时间", "收件人姓名", "收件人省市区", "运费"])
        ws.append(["700897665404", datetime(2026, 1, 2), "施施", "上海-上海市-浦东新区", 269])
        ws.append(["700897667062", datetime(2026, 1, 3), "刘岳", "黑龙江省-哈尔滨市", 515])
        return wb

    bi.import_logistics_xlsx(db_session, mk(), source_name="李爱群 2026年1月账单 784元.xlsx")
    n_line = db_session.query(LogisticsBill).filter_by(row_type="line").count()
    n_sum = db_session.query(LogisticsBill).filter_by(row_type="summary").count()
    assert (n_line, n_sum) == (2, 1)
    rep2 = bi.import_logistics_xlsx(db_session, mk(), source_name="李爱群 2026年1月账单 784元.xlsx")
    assert rep2.inserted == 0
    assert rep2.skipped_duplicate == 3   # 2 逐单 + 1 汇总 全判重
    assert db_session.query(LogisticsBill).filter_by(row_type="line").count() == 2
    assert db_session.query(LogisticsBill).filter_by(row_type="summary").count() == 1


def test_reimport_logistics_enriches_package_measurements(db_session):
    """旧账单重导时补实重/体积/件数，不新增、不覆盖计费重量。"""
    from datetime import datetime
    from openpyxl import Workbook
    from app.services import bill_import_service as bi

    def mk(include_measurements: bool):
        wb = Workbook(); ws = wb.active
        headers = ["运单号", "寄件时间", "计费重量", "收件人姓名", "运费"]
        if include_measurements:
            headers += ["实际重量", "体积", "件数"]
        ws.append(headers)
        row = ["700800000001", datetime(2026, 3, 2), 95, "张三", 269]
        if include_measurements:
            row += [87, 0.57, 2]
        ws.append(row)
        return wb

    bi.import_logistics_xlsx(db_session, mk(False), source_name="李爱群 2026年3月账单 269元.xlsx")
    rep = bi.import_logistics_xlsx(db_session, mk(True), source_name="李爱群 2026年3月账单 269元.xlsx")
    bill = db_session.query(LogisticsBill).filter_by(row_type="line").one()
    assert rep.inserted == 0
    assert rep.updated_existing == 1
    assert bill.weight_kg == Decimal("95")
    assert bill.actual_weight_kg == Decimal("87")
    assert bill.volume_m3 == Decimal("0.57")
    assert bill.package_count == 2


def test_reimport_logistics_csv_enriches_package_measurements(db_session):
    """CSV 与 XLSX 同口径：重导补齐实重/体积/件数且不重复建行。"""
    from app.services import bill_import_service as bi

    base = (
        "日期,承运商,运单号,计费重量,运费\n"
        "2026-03-02,德邦,CSV700800000001,95,269\n"
    )
    enriched = (
        "日期,承运商,运单号,计费重量,运费,实际重量,体积,件数\n"
        "2026-03-02,德邦,CSV700800000001,95,269,87,0.57,2\n"
    )

    bi.import_logistics_csv(db_session, base)
    rep = bi.import_logistics_csv(db_session, enriched)
    bill = db_session.query(LogisticsBill).filter_by(tracking_no="CSV700800000001").one()

    assert rep.inserted == 0
    assert rep.updated_existing == 1
    assert rep.skipped_duplicate == 1
    assert rep.errors == []
    assert bill.weight_kg == Decimal("95")
    assert bill.actual_weight_kg == Decimal("87")
    assert bill.volume_m3 == Decimal("0.57")
    assert bill.package_count == 2


def test_manual_match_not_overwritten(db_session):
    """人工指定过的不重算 (only_unmatched 跳过 manual)."""
    _order(db_session, "O6", "孙八", "山东省青岛市", track="DB6")
    b = _bill(db_session, tracking_no="DB6", recipient_name="孙八",
              destination="山东省青岛市", order_no="MANUAL", match_method="manual")
    logistics_bill_match.match_logistics_bills(db_session)
    assert b.order_no == "MANUAL"
    assert b.match_method == "manual"

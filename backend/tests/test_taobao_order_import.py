# -*- coding: utf-8 -*-
"""淘宝订单多格式自动识别导入测试。"""
import csv
import io
from decimal import Decimal

from openpyxl import Workbook

from app.models.order import Order
from app.services import taobao_order_import as tio


# ── 单元: 字段转换 ────────────────────────────────────────────────────────────
def test_product_code_from_merchant():
    assert tio.product_code_from_merchant("PPS2638004022511") == "P26380040225"
    assert tio.product_code_from_merchant("23210020201") == "P23210020201"
    assert tio.product_code_from_merchant("") == ""
    assert tio.product_code_from_merchant(None) == ""
    # 服务类哈希编码 → 非 11 位数字 → 空
    assert tio.product_code_from_merchant("de8df067567b3da9adb3a1d4d1918334") == ""


def test_extract_sku():
    assert tio.extract_sku("颜色分类:榉木床头柜-标准[长45cm];安装方式:免安装") == "榉木床头柜-标准"
    assert tio.extract_sku("颜色分类：砂白色2.0米岩板餐桌") == "砂白色2.0米岩板餐桌"
    assert tio.extract_sku("") == ""


def test_status_map():
    assert tio._map_status("交易成功") == "signed"
    assert tio._map_status("买家已付款,等待卖家发货") == "paid"
    assert tio._map_status("卖家已发货，等待买家确认") == "shipped"
    assert tio._map_status("交易关闭") == "cancelled"
    assert tio._map_status("等待买家付款") == "pending_payment"


# ── 格式识别 ──────────────────────────────────────────────────────────────────
def _qianniu_xlsx_bytes() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    o = wb.create_sheet("订单报表")
    o.append(["订单编号", "买家应付货款", "订单状态", "收货地址", "物流单号", "物流公司", "订单创建时间"])
    o.append(["A100", "2104.80", "买家已付款,等待卖家发货", "浙江省杭州市", "YT123", "圆通", "2026-06-02 11:37:39"])
    s = wb.create_sheet("销售明细")
    s.append(["子订单编号", "主订单编号", "商品标题", "购买数量", "商家编码", "商品属性", "买家应付货款", "退款状态", "退款金额", "订单创建时间"])
    s.append(["A100", "A100", "畔色榉木床头柜", "2", "PPS2638004022511", "颜色分类:榉木床头柜-标准[长45cm]", "2104.80", "没有申请退款", "无退款申请", "2026-06-02 11:37:39"])
    f = wb.create_sheet("发货报表")
    f.append(["订单编号", "收货人姓名", "联系手机", "收货地址"])
    f.append(["A100", "张三", "13800138000", "浙江省杭州市滨江区"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sales_detail_csv_bytes() -> bytes:
    rows = [
        ["子订单编号", "主订单编号", "标题", "价格", "购买数量", "外部系统编号",
         "商品属性", "订单状态", "商家编码", "买家应付货款", "退款状态", "退款金额", "订单创建时间"],
        ["B200", "B200", "畔色岩板餐桌", "5100.00", "1", "23210020201",
         "颜色分类：砂白色2.0米岩板餐桌", "交易成功", "23210020201", "2933.78", "没有申请退款", "无退款申请", "2026-01-18 13:39:27"],
    ]
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return out.getvalue().encode("gbk")


def test_detect_qianniu_multi():
    assert tio.detect_format("export.xlsx", _qianniu_xlsx_bytes()) == "qianniu_multi"


def test_detect_sales_detail_csv():
    assert tio.detect_format("ItemList.csv", _sales_detail_csv_bytes()) == "sales_detail"


def test_detect_unknown():
    bad = "甲,乙,丙\n1,2,3\n".encode("utf-8")
    assert tio.detect_format("x.csv", bad) == "unknown"


# ── 端到端导入 ────────────────────────────────────────────────────────────────
def test_import_qianniu_multi(db_session):
    rep = tio.import_taobao_orders(db_session, "export.xlsx", _qianniu_xlsx_bytes())
    assert rep.detected_format == "qianniu_multi"
    assert rep.inserted == 1
    o = db_session.query(Order).filter_by(order_no="A100").one()
    assert o.product_code == "PPS26380040225"   # 导入统一 P→PPS (用户拍板: 以后全 PPS)
    assert o.sku == "榉木床头柜-标准"             # 商品属性提取
    assert o.sku_code == "PPS2638004022511"
    assert o.qty == 2
    assert o.customer_name == "张三"             # 发货报表关联
    assert o.customer_phone == "13800138000"
    assert o.carrier == "圆通"                   # 订单报表关联
    assert o.status == "paid"
    # 活跃订单必须 is_historical=False, 否则现金流"待确认收货/未发货"过滤不到 (修复:历史病根)
    assert o.is_historical is False
    # 财务列落库 (现金流/逐笔对账靠它们)
    assert o.buyer_payable_amount == Decimal("2104.80")
    assert o.paid_amount == Decimal("2104.80")


def test_import_sales_detail_csv(db_session):
    rep = tio.import_taobao_orders(db_session, "ItemList.csv", _sales_detail_csv_bytes())
    assert rep.detected_format == "sales_detail"
    assert rep.inserted == 1
    o = db_session.query(Order).filter_by(order_no="B200").one()
    assert o.product_code == "PPS23210020201"   # 导入统一 P→PPS (用户拍板: 以后全 PPS)
    assert o.sku == "砂白色2.0米岩板餐桌"
    assert o.status == "signed"


def test_import_dedup(db_session):
    """再次导入同文件: 不重复插入, 改走 UPSERT 更新 (现金流要靠再导更新历史单的状态)。"""
    raw = _sales_detail_csv_bytes()
    tio.import_taobao_orders(db_session, "ItemList.csv", raw)
    rep2 = tio.import_taobao_orders(db_session, "ItemList.csv", raw)
    assert rep2.inserted == 0
    assert rep2.updated == 1
    assert db_session.query(Order).filter_by(order_no="B200").count() == 1   # 仍只一行


def test_reimport_updates_status_and_amount(db_session):
    """病根修复验证: 已存在订单(如旧导入卡在 pending_payment)再导时, 状态/金额被淘宝导出刷新。"""
    o = Order(platform="淘宝", order_no="B200", status="pending_payment")
    db_session.add(o)
    db_session.commit()
    rep = tio.import_taobao_orders(db_session, "ItemList.csv", _sales_detail_csv_bytes())
    assert rep.inserted == 0 and rep.updated == 1
    o2 = db_session.query(Order).filter_by(order_no="B200").one()
    assert o2.status == "signed"                          # pending_payment → signed
    assert o2.buyer_payable_amount == Decimal("2933.78")  # 金额回填


def _freight_csv_bytes(freight="60.00") -> bytes:
    rows = [
        ["子订单编号", "主订单编号", "标题", "价格", "购买数量", "外部系统编号", "商品属性",
         "订单状态", "商家编码", "买家应付货款", "买家实付金额", "买家应付邮费", "退款状态", "退款金额", "订单创建时间"],
        ["F1", "F500", "畔色胡桃木双人床", "1000.00", "1", "23210020201", "颜色分类：胡桃木1.8米",
         "交易成功", "23210020201", "75.00", "75.00", freight, "没有申请退款", "无退款申请", "2026-04-23 14:38:45"],
    ]
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return out.getvalue().encode("gbk")


def test_import_maps_buyer_freight(db_session):
    """买家应付邮费=代收运费, 单列落 buyer_freight, 不混进货款/实付 (5111173 真实形态: 货款75/实付75/运费60)。"""
    rep = tio.import_taobao_orders(db_session, "ItemList.csv", _freight_csv_bytes())
    assert rep.inserted == 1
    o = db_session.query(Order).filter_by(order_no="F500").one()
    assert o.paid_amount == Decimal("75.00")          # 实付不含运费
    assert o.buyer_freight == Decimal("60.00")        # 运费单列


def _multi_freight_csv_bytes() -> bytes:
    """一单两宝贝, 每行都带同一笔订单级邮费 60 (淘宝常把订单级字段在子行重复)。"""
    rows = [
        ["子订单编号", "主订单编号", "商品标题", "商品价格", "购买数量", "商家编码", "商品属性",
         "订单状态", "买家应付货款", "买家实付金额", "买家应付邮费", "退款金额", "订单创建时间"],
        ["M1", "M600", "畔色岩板餐桌", "3000.00", "1", "23210020201", "颜色分类：砂白2.0米",
         "交易成功", "3000.00", "3000.00", "60.00", "", "2026-04-23 14:38:45"],
        ["M2", "M600", "畔色实木餐椅", "800.00", "2", "23250050202", "颜色分类：胡桃木",
         "交易成功", "800.00", "800.00", "60.00", "", "2026-04-23 14:38:45"],
    ]
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return out.getvalue().encode("gbk")


def test_import_multi_product_freight_max_not_sum(db_session):
    """多产品单: 实付按子订单求和(3000+800=3800), 但邮费是订单级 → 取 max(60) 不求和(120)。"""
    rep = tio.import_taobao_orders(db_session, "ItemList.csv", _multi_freight_csv_bytes())
    assert rep.inserted == 1
    o = db_session.query(Order).filter_by(order_no="M600").one()
    assert o.paid_amount == Decimal("3800.00")        # 实付逐子订单求和
    assert o.buyer_freight == Decimal("60.00")        # 邮费取 max, 不被重复行求和成 120


def test_import_multi_line_order(db_session):
    """一单多商品: 聚合为一行 Order, 取主商品(金额最大), 备注其余。"""
    wb = Workbook(); wb.remove(wb.active)
    s = wb.create_sheet("销售明细")
    s.append(["子订单编号", "主订单编号", "商品标题", "购买数量", "商家编码", "商品属性", "买家应付货款", "订单创建时间"])
    s.append(["C1", "C300", "送货入户", "1", "", "", "0.00", "2026-01-06 19:36:09"])
    s.append(["C2", "C300", "畔色实木餐边柜", "1", "23250050202", "颜色分类：其他尺寸定制", "11212.00", "2026-01-06 19:36:09"])
    buf = io.BytesIO(); wb.save(buf)
    rep = tio.import_taobao_orders(db_session, "x.xlsx", buf.getvalue())
    assert rep.inserted == 1
    assert rep.multi_line_orders == 1
    o = db_session.query(Order).filter_by(order_no="C300").one()
    assert o.product_name == "畔色实木餐边柜"   # 金额最大行为主
    assert "本单含2个商品" in (o.remark or "")


# ── 重导幂等护栏 (2026-07-09): 根治 202 单实付/状态/退款被不同来源文件反复横跳 ──────────────
def _csv(header: list, rows: list) -> bytes:
    out = io.StringIO(); w = csv.writer(out); w.writerow(header)
    for r in rows:
        w.writerow(r)
    return out.getvalue().encode("gbk")


def test_idem_line_detail_keeps_authoritative_paid(db_session):
    """行级销售明细(不完整, 只匹配到一件)不许覆盖订单报表的正确实付。实测 13263: 4500.36 被 1795.17 反复盖。"""
    auth = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "退款金额", "商品属性", "商家编码", "订单创建时间"],
                [["Z1", "买家已付款,等待卖家发货", "4500.36", "4500.36", "", "颜色分类:餐桌", "23210020201", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "orderlist.csv", auth)
    assert db_session.query(Order).filter_by(order_no="Z1").one().paid_amount == Decimal("4500.36")
    # 不完整销售明细(有子订单编号 → 行级源), 只有餐桌一件 1795.17
    line = _csv(["子订单编号", "主订单编号", "订单状态", "商品标题", "商品属性", "买家应付货款", "买家实付金额", "商家编码", "订单创建时间"],
                [["Z1a", "Z1", "买家已付款,等待卖家发货", "餐桌", "颜色分类:餐桌", "1795.17", "1795.17", "23210020201", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "detail.csv", line)
    assert db_session.query(Order).filter_by(order_no="Z1").one().paid_amount == Decimal("4500.36")


def test_idem_missing_paid_col_keeps_real_paid(db_session):
    """缺「买家实付金额」列的稀疏导出不许用应付兜底盖掉真实实付。实测单产品单在 实付↔应付 之间翻。"""
    auth = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "商品属性", "订单创建时间"],
                [["Z2", "交易成功", "2632.66", "2520.46", "颜色分类:椅", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "orderlist.csv", auth)
    assert db_session.query(Order).filter_by(order_no="Z2").one().paid_amount == Decimal("2520.46")
    sparse = _csv(["订单编号", "订单状态", "买家应付货款", "商品属性", "订单创建时间"],
                  [["Z2", "交易成功", "2632.66", "颜色分类:椅", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "sparse.csv", sparse)
    assert db_session.query(Order).filter_by(order_no="Z2").one().paid_amount == Decimal("2520.46")


def test_idem_missing_status_col_keeps_status(db_session):
    """缺「订单状态」列的文件其 status 是兜底默认 signed → 不许覆盖已有 paid(否则每天把未发货盖成已签收)。"""
    auth = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "商品属性", "订单创建时间"],
                [["Z3", "买家已付款,等待卖家发货", "3000.00", "3000.00", "颜色分类:床", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "orderlist.csv", auth)
    assert db_session.query(Order).filter_by(order_no="Z3").one().status == "paid"
    nostatus = _csv(["订单编号", "买家应付货款", "买家实付金额", "商品属性", "订单创建时间"],
                    [["Z3", "3000.00", "3000.00", "颜色分类:床", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "nostatus.csv", nostatus)
    assert db_session.query(Order).filter_by(order_no="Z3").one().status == "paid"


def test_idem_authoritative_still_updates(db_session):
    """对照: 权威单级源(有状态列+实付列)该更新还得更新 —— 幂等护栏不能把正常更新也堵死。"""
    a = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "退款金额", "商品属性", "订单创建时间"],
             [["Z4", "买家已付款,等待卖家发货", "1000.00", "1000.00", "", "颜色分类:桌", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "a.csv", a)
    o = db_session.query(Order).filter_by(order_no="Z4").one()
    assert o.status == "paid" and (o.refund_amount or Decimal("0")) == Decimal("0")
    b = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "退款金额", "商品属性", "订单创建时间"],
             [["Z4", "交易成功", "1000.00", "1000.00", "200.00", "颜色分类:桌", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "b.csv", b)
    o2 = db_session.query(Order).filter_by(order_no="Z4").one()
    assert o2.status == "signed"                   # 权威源状态更新
    assert o2.refund_amount == Decimal("200.00")   # 权威源退款更新


def test_idem_qianniu_incomplete_detail_no_reduce(db_session):
    """千牛三表: 订单报表单级 4500.36, 销售明细只覆盖到部分(1795.17+1000<4500.36) → 采用订单报表总额, 不被明细求和压低。"""
    wb = Workbook(); wb.remove(wb.active)
    o = wb.create_sheet("订单报表")
    o.append(["订单编号", "买家应付货款", "买家实付金额", "订单状态", "订单创建时间"])
    o.append(["Q1", "4500.36", "4500.36", "买家已付款,等待卖家发货", "2026-06-16 10:00:00"])
    s = wb.create_sheet("销售明细")
    s.append(["子订单编号", "主订单编号", "商品标题", "购买数量", "商家编码", "商品属性", "买家应付货款", "买家实付金额", "订单创建时间"])
    s.append(["Q1a", "Q1", "餐桌", "1", "23210020201", "颜色分类:餐桌", "1795.17", "1795.17", "2026-06-16 10:00:00"])
    s.append(["Q1b", "Q1", "椅子", "1", "23250050202", "颜色分类:椅", "1000.00", "1000.00", "2026-06-16 10:00:00"])
    buf = io.BytesIO(); wb.save(buf)
    tio.import_taobao_orders(db_session, "export.xlsx", buf.getvalue())
    assert db_session.query(Order).filter_by(order_no="Q1").one().paid_amount == Decimal("4500.36")


def test_idem_multiproduct_partial_refund_is_paid_not_signed(db_session):
    """13263 型: 一单两件(床头柜退款 + 餐桌还在做), 已卖出宝贝导出整单状态=交易关闭 → 应判"进行中 paid",
    不是 signed(否则跟订单报表的 paid 天天打架)也不是 cancelled(会把还在做的餐桌漏出销售)。"""
    line = _csv(["子订单编号", "主订单编号", "订单状态", "商品标题", "商品属性", "买家应付货款", "买家实付金额", "退款金额", "商家编码", "订单创建时间"],
                [["N1", "N900", "交易关闭", "床头柜", "颜色:柜", "1795.17", "1795.17", "1795.17", "23210020201", "2026-06-16 10:00:00"],
                 ["N2", "N900", "交易关闭", "餐桌", "颜色:桌", "2705.19", "2705.19", "", "23210020202", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "detail.csv", line)
    o = db_session.query(Order).filter_by(order_no="N900").one()
    assert o.status == "paid"                    # 进行中, 不是 signed / cancelled
    assert o.paid_amount == Decimal("4500.36")   # 两件求和, 不被压低


def test_idem_zero_paid_does_not_wipe_real_paid(db_session):
    """0护栏: order 源导出把老单实付报成 0 时, 不许清掉已有的真实实付(该单没关闭)。"""
    auth = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "商品属性", "订单创建时间"],
                [["Z9", "买家已付款,等待卖家发货", "3326.63", "3326.63", "颜色:桌", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "a.csv", auth)
    assert db_session.query(Order).filter_by(order_no="Z9").one().paid_amount == Decimal("3326.63")
    zero = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "商品属性", "订单创建时间"],
                [["Z9", "买家已付款,等待卖家发货", "3326.63", "0", "颜色:桌", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "b.csv", zero)
    assert db_session.query(Order).filter_by(order_no="Z9").one().paid_amount == Decimal("3326.63")  # 未被清零


def test_idem_cancelled_order_may_zero_paid(db_session):
    """对照: 单产品单确实关闭(交易关闭)时, 实付落 0 是合理的, 0护栏不该拦。"""
    auth = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "商品属性", "订单创建时间"],
                [["Z7", "买家已付款,等待卖家发货", "1000.00", "1000.00", "颜色:桌", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "a.csv", auth)
    close = _csv(["订单编号", "订单状态", "买家应付货款", "买家实付金额", "商品属性", "订单创建时间"],
                 [["Z7", "交易关闭", "1000.00", "0", "颜色:桌", "2026-06-16 10:00:00"]])
    tio.import_taobao_orders(db_session, "b.csv", close)
    o = db_session.query(Order).filter_by(order_no="Z7").one()
    assert o.status == "cancelled" and o.paid_amount == Decimal("0")

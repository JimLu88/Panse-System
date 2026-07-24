"""定制占位符 SKU (2026-07-07): 淘宝微定制/材质定制/差价等占位链接。
仅用于淘宝报名(活动价=现价×0.9, 单品立减=现价×10%); recompute 跳过(不参与产品成本/价格链计算)。"""
from decimal import Decimal as D
from io import BytesIO

import openpyxl

from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import pricing_calc_service as svc
from app.services.data_export_service import (
    build_promo_signup_upload_xlsx, build_single_item_discount_upload_xlsx,
)


def _seed_ph(db, daily=20000, code="PPS999", skuid="SK1"):
    db.add(PricingSku(product_code="PPS", sku_code=code, sku="樱桃木家具定制",
                      daily_price=D(str(daily)), is_custom_placeholder=True))
    db.add(PricingSkuPromo(sku_code=code, taobao_item_id="ITEM", taobao_sku_id=skuid))
    db.commit()


def test_recompute_skips_placeholder():
    # 占位符即使有 physical_cost+base_big, recompute 也直接跳过 → 不派生价/利润, daily_price 原值不动
    sku = PricingSku(product_code="P", sku_code="P99", daily_price=D("20000"),
                     physical_cost=D("500"), base_big=D("0.88"), is_custom_placeholder=True)
    svc.recompute(sku)
    assert sku.big_promo is None and sku.big_promo_margin is None
    assert sku.daily_price == D("20000")


def test_signup_export_placeholder_is_9zhe(db_session):
    _seed_ph(db_session, 20000)
    buf, _ = build_promo_signup_upload_xlsx(db_session, "big")
    ws = openpyxl.load_workbook(BytesIO(buf.getvalue()))["商品SKU导入列表"]
    rows = [(ws.cell(r, 2).value, ws.cell(r, 3).value)
            for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value]
    assert len(rows) == 1 and rows[0][0] == "SK1"
    assert abs(float(rows[0][1]) - 500.0) < 0.01                # 20000×0.9 再封500顶(2026-07-16固化)


def test_signup_618_also_9zhe(db_session):
    _seed_ph(db_session, 20000)
    buf, _ = build_promo_signup_upload_xlsx(db_session, "big618")
    ws = openpyxl.load_workbook(BytesIO(buf.getvalue()))["商品SKU导入列表"]
    vals = [ws.cell(r, 3).value for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value]
    assert abs(float(vals[0]) - 500.0) < 0.01                  # 占位×0.9后再封500顶(2026-07-16固化: 治新品促销管控价卡整品)


def test_single_discount_placeholder_is_skipped(db_session):
    _seed_ph(db_session, 20000)
    buf, _ = build_single_item_discount_upload_xlsx(db_session, "big")
    ws = openpyxl.load_workbook(BytesIO(buf.getvalue()))["单品立减"]
    rows = [(ws.cell(r, 2).value, ws.cell(r, 3).value)
            for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
    assert rows == []                                           # 占位只走保护报名价，不叠单品立减

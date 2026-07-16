"""无动销品单品立减平替 (2026-07-17 用户永久规则):
动销不达标报不进大促 → 单品立减到手 = 中促价 − 1 元。

锁四件事:
① 登记表增删/回执自愈抽取(动销不达标 → item_id 入册)
② builder 只对登记商品出行, 立减金额 = 日常价 − (中促价−1)
③ 护栏: 到手 < 大促锚 → 跳过(绝不立低于锚的线); 占位不出行
④ 未登记商品绝不出行
"""
from decimal import Decimal

from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import data_export_service as de
from app.services import no_sales_service as ns


def _mk(db, code, item, sid, daily, mid, big_buyer, placeholder=False):
    db.add(PricingSku(product_code=code[:9], sku_code=code, sku=f"SKU{code}",
                      daily_price=Decimal(str(daily)),
                      mid_promo=Decimal(str(mid)) if mid is not None else None,
                      is_custom_placeholder=placeholder))
    db.add(PricingSkuPromo(sku_code=code, taobao_item_id=item, taobao_sku_id=sid,
                           big_buyer_price=Decimal(str(big_buyer)) if big_buyer else None))


def test_registry_add_remove_and_feedback_extract(db_session):
    assert ns.get_no_sales(db_session) == set()
    ns.add_no_sales(db_session, ["9001", "9002"])
    assert ns.get_no_sales(db_session) == {"9001", "9002"}
    ns.remove_no_sales(db_session, ["9001"])
    assert ns.get_no_sales(db_session) == {"9002"}

    got = ns.extract_no_sales_from_feedback([
        {"item_id": "8001", "raw": "参加本次活动的商品要求满足“动销”校验基础准入门槛要求…销售件数≥1件"},
        {"item_id": "8002", "raw": "券后价不得高于最低普惠券后价"},
    ])
    assert got == {"8001"}


def test_builder_rows_and_guards(db_session):
    # 登记品A: 正常 → 出行, 立减 = 3000 − (2730−1) = 271
    _mk(db_session, "PPSNS001", "9101", "71001", daily=3000, mid=2730, big_buyer=2561.22)
    # 登记品A的占位SKU → 不出行
    _mk(db_session, "PPSNS099", "9101", "71099", daily=500, mid=None, big_buyer=None, placeholder=True)
    # 登记品B: 中促−1 低于大促锚 → 护栏跳过
    _mk(db_session, "PPSNS002", "9102", "71002", daily=1000, mid=880, big_buyer=900)
    # 未登记品C → 绝不出行
    _mk(db_session, "PPSNS003", "9103", "71003", daily=2000, mid=1820, big_buyer=1707)
    db_session.commit()
    ns.add_no_sales(db_session, ["9101", "9102"])

    bio, stats = de.build_nosales_single_item_discount_xlsx(db_session)

    import io
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(bio.getvalue())).active
    rows = [(str(ws.cell(r, 1).value), str(ws.cell(r, 2).value), float(ws.cell(r, 3).value))
            for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value]

    assert rows == [("9101", "71001", 271.0)], rows
    assert stats["rows"] == 1
    assert stats["skipped_placeholder"] == 1
    assert len(stats["skipped_below_anchor"]) == 1
    assert stats["skipped_below_anchor"][0]["sku_code"] == "PPSNS002"
    assert "9103" not in {r[0] for r in rows}

"""工厂逐单对账: xlsx 导入(序列号日期/去重/回填) + 逐月对账 + 填原因做平。"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl

from app.models.factory_recon_item import FactoryReconItem
from app.models.finance import AlipayFlow
from app.models.order import Order
from app.services import factory_recon_import_service as imp
from app.services import factory_recon_service as svc


def _build_xlsx() -> bytes:
    """造一张和工厂侧对账单同构的 xlsx: 标题行 + 第2行表头 + 数据(日期=Excel序列号)。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "26年1月"
    ws.append(["畔色 月度生产明细表"])  # 第1行标题
    ws.append(["单号", "订单号", "追加订单号1", "备注", "详情", "数量", "价格",
               "客户信息", "下单时间", "发货时间"])  # 第2行表头
    # 46080 ≈ 2026-02-04 (Excel base 1899-12-30)
    ws.append([210, "ORD001", "", "", "樱桃木 窄柜100", 1, 3300, "北京 张三", 46080, 46109])
    ws.append([211, "ORD002", "", "", "胡桃木 案台", 1, 1450, "上海 李四", 46080, "未发货"])
    ws.append([212, "", "", "", "合计", "", "", "", "", ""])  # 无价无单号 → 跳过
    return _to_bytes(wb)


def _to_bytes(wb) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_import_parses_serial_dates_and_dedup(db_session):
    db = db_session
    rep = imp.import_factory_recon_xlsx(db, _build_xlsx())
    assert rep.inserted == 2
    assert rep.errors == []
    items = {i.order_no: i for i in db.query(FactoryReconItem).all()}
    assert items["ORD001"].settle_price == Decimal("3300")
    # Excel 序列号 (base 1899-12-30): 46080 → 2026-02-27, 46109 → 2026-03-28
    assert items["ORD001"].order_date == datetime(2026, 2, 27).date()
    assert items["ORD001"].ship_date == datetime(2026, 3, 28).date()
    # "未发货" 文本无法解析 → None (不报错)
    assert items["ORD002"].ship_date is None

    # 再次导入同文件 → 全部判重, 不新增
    rep2 = imp.import_factory_recon_xlsx(db, _build_xlsx())
    assert rep2.inserted == 0
    assert rep2.skipped_duplicate == 2


def _build_stock_xlsx(n=3, sheet="26年1月") -> bytes:
    """造一张含 n 张「无订单号无单号、同价同品」备货行的对账单(测方案A: 不误删)。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["畔色 月度生产明细表"])
    ws.append(["单号", "订单号", "追加订单号1", "备注", "详情", "数量", "价格",
               "客户信息", "下单时间", "发货时间"])
    for _ in range(n):
        ws.append(["", "", "", "", "备货 樱桃木窄柜160*85", 1, 1350, "", "", ""])
    return _to_bytes(wb)


def test_stock_rows_no_orderno_not_deduped_within_sheet(db_session):
    """方案A: 无订单号无单号的备货行(同价同品)同表多张全保留, 不再误删。"""
    db = db_session
    data = _build_stock_xlsx(3)
    rep = imp.import_factory_recon_xlsx(db, data)
    assert rep.inserted == 3                       # 3 张相同备货全保留
    assert db.query(FactoryReconItem).count() == 3
    # 同一份文件(同 bytes)再导 → 整份判重(source_file_hash), 不重复
    rep2 = imp.import_factory_recon_xlsx(db, data)
    assert rep2.inserted == 0
    assert rep2.skipped_duplicate == 3
    assert db.query(FactoryReconItem).count() == 3


def test_stock_rows_different_files_both_kept(db_session):
    """不同月份(不同文件 → 不同 hash)的相同备货 → 各自保留, 不跨文件误删。"""
    db = db_session
    imp.import_factory_recon_xlsx(db, _build_stock_xlsx(2, sheet="26年1月"))
    imp.import_factory_recon_xlsx(db, _build_stock_xlsx(2, sheet="26年2月"))
    assert db.query(FactoryReconItem).count() == 4


def test_backfill_order_actual_cost(db_session):
    db = db_session
    db.add(Order(platform="淘宝", order_no="ORD001", status="signed",
                 paid_amount=Decimal("5000")))   # actual_cost 为空 → 应被回填
    db.flush()
    imp.import_factory_recon_xlsx(db, _build_xlsx())
    o = db.query(Order).filter_by(order_no="ORD001").one()
    assert o.actual_cost == Decimal("3300")


def test_summary_and_resolve_balancing(db_session):
    db = db_session
    imp.import_factory_recon_xlsx(db, _build_xlsx())
    # 实付: 一笔 factory_payment 支出, 与应付(3300+1450=4750)差 -1450
    db.add(AlipayFlow(
        account="企业号", transaction_no="F-PAY-1", transaction_time=datetime(2026, 2, 20),
        amount=Decimal("-3300"), reconciliation_type="factory_payment",
    ))
    db.flush()

    s = svc.summary(db)
    assert s["total_items"] == 2
    assert s["total_billed"] == 4750.0
    assert s["total_paid"] == 3300.0
    feb = next(m for m in s["months"] if m["period"] == "2026-02")
    assert feb["status"] == "diff"          # 差额未归因
    assert feb["items_open"] == 2

    # 对 ORD002(1450) 填原因做平; 仍剩 ORD001 未平 → 仍 diff
    ord2 = db.query(FactoryReconItem).filter_by(order_no="ORD002").one()
    svc.resolve(db, ord2.id, reason="平台扣减运费减免", actor="tester")
    feb = next(m for m in svc.summary(db)["months"] if m["period"] == "2026-02")
    assert feb["items_resolved"] == 1
    assert feb["status"] == "diff"

    # 全部条目做平 → explained
    ord1 = db.query(FactoryReconItem).filter_by(order_no="ORD001").one()
    svc.resolve(db, ord1.id, reason="账单一致", actor="tester")
    feb = next(m for m in svc.summary(db)["months"] if m["period"] == "2026-02")
    assert feb["status"] == "explained"
    assert feb["items_open"] == 0


def test_resolve_requires_reason(db_session):
    db = db_session
    imp.import_factory_recon_xlsx(db, _build_xlsx())
    it = db.query(FactoryReconItem).first()
    try:
        svc.resolve(db, it.id, reason="   ")
        assert False, "空原因应报错"
    except ValueError:
        pass


def test_order1_order2_headers_and_statement_boundaries(db_session):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "6月账单"
    ws.append(["6月账单"])
    ws.append(["单号", "订单号1", "订单号2", "详情", "图片", "数量", "价格"])
    ws.append([93, "5115417627826040832", "5115619190250004822", "岩板桌", None, 1, 1250])
    ws.append([None, None, None, None, "合计金额", None, 1250])
    ws.append(["售后", None, None, "售后抵扣", None, None, -50])
    ws.append([None, None, None, None, "5月账单尾款", None, 1200])
    ws.append([154, "3305445027534040190", None, "延后下单7月底", None, 2, 920])

    rep = imp.import_factory_recon_xlsx(db_session, _to_bytes(wb))
    items = db_session.query(FactoryReconItem).order_by(FactoryReconItem.id).all()

    assert rep.inserted == 2
    assert [(item.order_no, item.extra_order_no1, item.settle_price) for item in items] == [
        ("5115417627826040832", "5115619190250004822", Decimal("1250")),
        (None, None, Decimal("-50")),
    ]

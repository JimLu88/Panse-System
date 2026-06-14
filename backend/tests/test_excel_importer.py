"""Excel importer: preview / AI 推断 / commit (delivery_note + factory_order)."""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from app.models.order import FactoryOrder
from app.models.supplier import DeliveryNote, DeliveryNoteLine, Supplier
from app.services import excel_importer, settings_service
from app.services.ai_provider import AiResponse


def _xlsx(sheet_name: str, header: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _multi_sheet_xlsx(sheets: dict[str, tuple[list[str], list[list]]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for name, (header, rows) in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(header)
        for r in rows:
            ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------- preview --------------------------- #


def test_preview_single_sheet():
    data = _xlsx("Sheet1", ["供应商", "单号", "日期", "品名", "数量", "金额"], [
        ["木作工厂", "N1", "2026-05-14", "电视柜", 1, 580],
        ["木作工厂", "N1", "2026-05-14", "斗柜", 1, 320],
        ["岩板厂", "N2", "2026-05-15", "台面板", 2, 1160],
    ])
    previews = excel_importer.preview_excel(data)
    assert len(previews) == 1
    p = previews[0]
    assert p.sheet_name == "Sheet1"
    assert p.row_count == 3
    assert p.column_names == ["供应商", "单号", "日期", "品名", "数量", "金额"]
    assert len(p.sample_rows) == 3
    assert p.sample_rows[0][0] == "木作工厂"


def test_preview_multi_sheet():
    data = _multi_sheet_xlsx({
        "5月": (["供应商", "数量"], [["A", 1]]),
        "6月": (["供应商", "数量"], [["B", 2], ["C", 3]]),
        "空sheet": (["col1"], []),
    })
    previews = excel_importer.preview_excel(data)
    assert len(previews) == 3
    assert {p.sheet_name for p in previews} == {"5月", "6月", "空sheet"}
    by_name = {p.sheet_name: p for p in previews}
    assert by_name["5月"].row_count == 1
    assert by_name["6月"].row_count == 2
    assert by_name["空sheet"].row_count == 0


def test_preview_invalid_excel_raises():
    with pytest.raises(excel_importer.ImporterError):
        excel_importer.preview_excel(b"not an xlsx")


def test_preview_sample_rows_truncated_to_5():
    rows = [[f"r{i}", i] for i in range(20)]
    data = _xlsx("S", ["a", "b"], rows)
    p = excel_importer.preview_excel(data)[0]
    assert p.row_count == 20
    assert len(p.sample_rows) == 5


def test_preview_handles_date_value():
    data = _xlsx("S", ["日期"], [[date(2026, 5, 14)]])
    p = excel_importer.preview_excel(data)[0]
    assert "2026-05-14" in str(p.sample_rows[0][0])


def _xlsx_with_title_banner(sheet_name, title, header, rows):
    """第一行是合并单元格大标题, 第二行才是真表头 (复现业务总表结构)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([title] + [None] * (len(header) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_preview_skips_merged_title_banner():
    """合并标题行不应被当成表头 (否则真列名错位, 每行报缺主键被跳过)."""
    data = _xlsx_with_title_banner(
        "1-产品总表", "1-产品总表",
        ["产品编码", "类目", "产品名称"],
        [["PPS001", "床", "榉木床"], ["PPS002", "柜", "斗柜"]],
    )
    p = excel_importer.preview_excel(data)[0]
    assert p.column_names == ["产品编码", "类目", "产品名称"]
    assert p.row_count == 2
    assert p.sample_rows[0][0] == "PPS001"
    assert any("识别为表头" in n for n in p.notes)


def test_commit_skips_merged_title_banner(db_session):
    """带合并标题的产品总表能正确入库, 不再误判'缺产品编码'."""
    from app.models.product import Product
    data = _xlsx_with_title_banner(
        "1-产品总表", "1-产品总表",
        ["产品编码", "类目", "产品名称"],
        [["PPS001", "床", "榉木床"], ["PPS002", "柜", "斗柜"]],
    )
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="1-产品总表",
        entity_type="product",
        mapping={"code": "产品编码", "category": "类目", "name": "产品名称"},
    )
    db_session.commit()
    assert report.inserted_parents == 2
    assert report.skipped_rows == 0
    codes = {p.code for p in db_session.query(Product).all()}
    assert {"PPS001", "PPS002"} <= codes


def test_account_balance_missing_date_defaults_to_current(db_session):
    """账户余额行有账户名但缺统计日期 → 用当前年月兜底入账, 不丢行。"""
    from app.models.finance import AccountBalance
    data = _xlsx("10", ["账户名称", "账户号", "统计日期", "期初余额", "期末余额"], [
        ["银行卡-个体户私人", None, None, 30000, 30000],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="10",
        entity_type="account_balance",
        mapping={"account_name": "账户名称", "account_no": "账户号",
                 "period_date": "统计日期", "opening_balance": "期初余额",
                 "closing_balance": "期末余额"},
    )
    db_session.commit()
    assert report.inserted_parents == 1
    assert any("缺统计日期" in w for w in report.warnings)
    bal = db_session.query(AccountBalance).filter_by(account_name="银行卡-个体户私人").one()
    assert bal.period_year and bal.period_month  # 已填充, 非空


# ----------------------------- AI 推断 --------------------------- #


def test_infer_mapping_no_ai_config_skips_silently(db_session):
    p = excel_importer.SheetPreview(
        sheet_name="S", row_count=1,
        column_names=["供应商", "数量"], sample_rows=[["A", 1]],
    )
    result = excel_importer.infer_mapping(db_session, preview=p)
    assert "AI 未配置" in (result.notes[0] if result.notes else "")
    assert result.suggested_mapping == {}


def test_infer_mapping_with_mocked_ai(db_session):
    settings_service.set_value(db_session, "ai_diagnose_provider", "anthropic")
    settings_service.set_value(db_session, "ai_diagnose_api_key", "k")
    settings_service.set_value(db_session, "ai_diagnose_model", "claude-x")

    p = excel_importer.SheetPreview(
        sheet_name="5月对账",
        row_count=2,
        column_names=["供应商", "单号", "送货日期", "品名", "数量", "金额"],
        sample_rows=[
            ["木作工厂", "N1", "2026-05-14", "电视柜", 1, 580],
            ["木作工厂", "N1", "2026-05-14", "斗柜", 1, 320],
        ],
    )
    fake = MagicMock()
    fake.chat.return_value = AiResponse(
        text='''{
            "entity_type": "delivery_note",
            "mapping": {
                "supplier_name": "供应商",
                "note_no": "单号",
                "delivery_date": "送货日期",
                "item_name": "品名",
                "qty": "数量",
                "amount": "金额"
            },
            "skipped_columns": [],
            "warnings": []
        }''',
        model="claude-x",
    )
    with patch.object(excel_importer, "build_provider", return_value=fake):
        result = excel_importer.infer_mapping(db_session, preview=p)

    assert result.suggested_entity == "delivery_note"
    assert result.suggested_mapping["supplier_name"] == "供应商"
    assert result.suggested_mapping["delivery_date"] == "送货日期"
    assert result.suggested_mapping["qty"] == "数量"


def test_infer_mapping_handles_ai_invalid_json(db_session):
    settings_service.set_value(db_session, "ai_diagnose_provider", "anthropic")
    settings_service.set_value(db_session, "ai_diagnose_api_key", "k")
    settings_service.set_value(db_session, "ai_diagnose_model", "claude-x")

    p = excel_importer.SheetPreview(
        sheet_name="S", row_count=1, column_names=["x"], sample_rows=[["v"]],
    )
    fake = MagicMock()
    fake.chat.return_value = AiResponse(text="完全不是 JSON", model="claude-x")
    with patch.object(excel_importer, "build_provider", return_value=fake):
        result = excel_importer.infer_mapping(db_session, preview=p)
    assert any("无法解析" in n for n in result.notes)
    assert result.suggested_mapping == {}


def test_infer_mapping_filters_invalid_columns(db_session):
    """AI 给的 mapping 引用了 Excel 里不存在的列 → 应被过滤掉."""
    settings_service.set_value(db_session, "ai_diagnose_provider", "anthropic")
    settings_service.set_value(db_session, "ai_diagnose_api_key", "k")
    settings_service.set_value(db_session, "ai_diagnose_model", "claude-x")

    p = excel_importer.SheetPreview(
        sheet_name="S", row_count=1,
        column_names=["供应商", "数量"], sample_rows=[["A", 1]],
    )
    fake = MagicMock()
    fake.chat.return_value = AiResponse(text='''{
        "entity_type": "delivery_note",
        "mapping": {
            "supplier_name": "供应商",
            "delivery_date": "不存在的日期列",
            "qty": "数量"
        }
    }''', model="claude-x")
    with patch.object(excel_importer, "build_provider", return_value=fake):
        result = excel_importer.infer_mapping(db_session, preview=p)
    assert "supplier_name" in result.suggested_mapping
    assert "qty" in result.suggested_mapping
    assert "delivery_date" not in result.suggested_mapping


# ----------------------------- commit delivery_note -------------- #


def test_commit_delivery_note_basic(db_session):
    data = _xlsx("S", ["供应商", "单号", "日期", "品名", "数量", "单价", "金额"], [
        ["木作工厂", "N1", "2026-05-14", "电视柜", 1, 580, 580],
        ["木作工厂", "N1", "2026-05-14", "斗柜", 1, 320, 320],
        ["木作工厂", "N2", "2026-05-15", "床头柜", 2, 300, 600],
    ])
    mapping = {
        "supplier_name": "供应商",
        "note_no": "单号",
        "delivery_date": "日期",
        "item_name": "品名",
        "qty": "数量",
        "unit_price": "单价",
        "amount": "金额",
    }
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note", mapping=mapping,
        auto_match_orders=False,
    )
    db_session.commit()

    assert report.total_rows == 3
    assert report.inserted_parents == 2     # N1, N2
    assert report.inserted_children == 3
    assert report.skipped_rows == 0
    assert "木作工厂" in report.auto_created_suppliers

    from sqlalchemy import select
    notes = db_session.execute(select(DeliveryNote).order_by(DeliveryNote.note_no)).scalars().all()
    assert len(notes) == 2
    n1 = next(n for n in notes if n.note_no == "N1")
    assert n1.total_amount == Decimal("900")  # 580 + 320 (Excel 没单独给 total, 用行求和)
    lines = db_session.execute(
        select(DeliveryNoteLine).where(DeliveryNoteLine.delivery_note_id == n1.id)
    ).scalars().all()
    assert len(lines) == 2
    assert {ln.item_name for ln in lines} == {"电视柜", "斗柜"}


def test_commit_delivery_note_dry_run_not_persisted(db_session):
    data = _xlsx("S", ["供应商", "单号", "日期", "品名", "数量", "金额"], [
        ["X", "N1", "2026-05-14", "x", 1, 100],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note",
        mapping={"supplier_name": "供应商", "note_no": "单号",
                 "delivery_date": "日期", "item_name": "品名",
                 "qty": "数量", "amount": "金额"},
        auto_match_orders=False, dry_run=True,
    )
    assert report.inserted_parents == 1
    # dry run 之后回滚, 不应留下 supplier
    from sqlalchemy import select
    assert db_session.execute(select(Supplier).where(Supplier.name == "X")).first() is None


def test_commit_missing_required_field_raises(db_session):
    data = _xlsx("S", ["数量"], [[1]])
    with pytest.raises(excel_importer.ImporterError) as ei:
        excel_importer.commit_sheet(
            db_session, file_bytes=data, sheet_name="S",
            entity_type="delivery_note", mapping={"qty": "数量"},
            auto_match_orders=False,
        )
    assert "必填字段未映射" in str(ei.value)


def test_commit_skips_bad_rows_keeps_good(db_session):
    data = _xlsx("S", ["供应商", "日期", "品名", "数量"], [
        ["A", "2026-05-14", "x", 1],
        ["B", "不是日期", "y", 2],   # 日期错 → 跳过
        ["", "2026-05-14", "z", 3],  # 供应商空 → 跳过
        ["C", "2026-05-14", "w", "abc"],  # 数量错 → 跳过
        ["D", "2026-05-14", "v", 4],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note",
        mapping={"supplier_name": "供应商", "delivery_date": "日期",
                 "item_name": "品名", "qty": "数量"},
        auto_match_orders=False,
    )
    db_session.commit()
    assert report.inserted_parents == 2   # 只 A, D 进了
    assert report.skipped_rows == 3
    assert len(report.errors) == 3


def test_commit_idempotent_skips_existing_note_no(db_session):
    data = _xlsx("S", ["供应商", "单号", "日期", "品名", "数量"], [
        ["X", "DUP", "2026-05-14", "p", 1],
    ])
    mapping = {"supplier_name": "供应商", "note_no": "单号",
               "delivery_date": "日期", "item_name": "品名", "qty": "数量"}
    excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note", mapping=mapping,
        auto_match_orders=False,
    )
    db_session.commit()
    # 再跑一次 → 跳过
    report2 = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note", mapping=mapping,
        auto_match_orders=False,
    )
    db_session.commit()
    assert report2.inserted_parents == 0
    assert report2.skipped_rows == 1
    assert any("已存在" in w for w in report2.warnings)


def test_commit_auto_create_supplier_disabled_errors(db_session):
    data = _xlsx("S", ["供应商", "日期", "品名", "数量"], [
        ["新供应商", "2026-05-14", "x", 1],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note",
        mapping={"supplier_name": "供应商", "delivery_date": "日期",
                 "item_name": "品名", "qty": "数量"},
        auto_create_suppliers=False, auto_match_orders=False,
    )
    db_session.commit()
    assert report.inserted_parents == 0
    assert report.skipped_rows == 1
    assert any("不存在" in e for e in report.errors)


def test_commit_existing_supplier_not_recreated(db_session):
    pre = Supplier(name="木作工厂", supplier_type="woodwork")
    db_session.add(pre); db_session.commit()

    data = _xlsx("S", ["供应商", "日期", "品名", "数量"], [
        ["木作工厂", "2026-05-14", "x", 1],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note",
        mapping={"supplier_name": "供应商", "delivery_date": "日期",
                 "item_name": "品名", "qty": "数量"},
        auto_match_orders=False,
    )
    db_session.commit()
    assert "木作工厂" not in report.auto_created_suppliers
    from sqlalchemy import select
    suppliers = db_session.execute(select(Supplier)).scalars().all()
    assert len(suppliers) == 1
    assert suppliers[0].supplier_type == "woodwork"  # 没被覆盖


def test_commit_auto_match_orders(db_session):
    from app.models.order import FactoryOrder as FO
    db_session.add(FO(
        factory_order_no="FO-001", platform_order_no="TB-001",
        factory_name="木作", product_code="P1",
        sku="电视柜 1800×850 黑色", qty=1, order_date=date(2026, 5, 10),
        expected_delivery=date(2026, 5, 14),
    ))
    db_session.commit()

    data = _xlsx("S", ["供应商", "日期", "品名", "规格", "数量"], [
        ["木作", "2026-05-14", "电视柜", "1800×850", 1],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note",
        mapping={"supplier_name": "供应商", "delivery_date": "日期",
                 "item_name": "品名", "spec": "规格", "qty": "数量"},
        auto_match_orders=True,
    )
    db_session.commit()
    assert report.matched_lines >= 1


def test_commit_amount_auto_computed_from_price_qty(db_session):
    data = _xlsx("S", ["供应商", "日期", "品名", "数量", "单价"], [
        ["X", "2026-05-14", "x", 3, 100],
    ])
    excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note",
        mapping={"supplier_name": "供应商", "delivery_date": "日期",
                 "item_name": "品名", "qty": "数量", "unit_price": "单价"},
        auto_match_orders=False,
    )
    db_session.commit()
    from sqlalchemy import select
    ln = db_session.execute(select(DeliveryNoteLine)).scalar_one()
    assert ln.amount == Decimal("300.00")


# ----------------------------- commit factory_order -------------- #


def test_commit_factory_order_basic(db_session):
    data = _xlsx("S", ["厂单号", "平台单号", "工厂", "下单日期", "SKU", "数量", "金额"], [
        ["FO-101", "TB-101", "X工厂", "2026-05-10", "电视柜黑色 1800", 1, 580],
        ["FO-102", "TB-102", "X工厂", "2026-05-11", "斗柜", 2, 640],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="factory_order",
        mapping={
            "factory_order_no": "厂单号",
            "platform_order_no": "平台单号",
            "factory_name": "工厂",
            "order_date": "下单日期",
            "sku": "SKU",
            "qty": "数量",
            "factory_bill_amount": "金额",
        },
    )
    db_session.commit()
    assert report.inserted_parents == 2
    from sqlalchemy import select
    fos = db_session.execute(select(FactoryOrder).order_by(FactoryOrder.factory_order_no)).scalars().all()
    assert len(fos) == 2
    assert fos[0].factory_order_no == "FO-101"
    assert fos[0].factory_bill_amount == Decimal("580")


def test_commit_factory_order_duplicate_skipped(db_session):
    pre = FactoryOrder(factory_order_no="FO-X", qty=1)
    db_session.add(pre); db_session.commit()

    data = _xlsx("S", ["厂单号", "数量"], [["FO-X", 1]])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="factory_order",
        mapping={"factory_order_no": "厂单号", "qty": "数量"},
    )
    db_session.commit()
    assert report.inserted_parents == 0    # 没有重复插入
    assert report.conflicts == []          # 命中已存在, 补默认/空字段属补全, 不算冲突


def test_commit_unknown_entity_type_raises(db_session):
    with pytest.raises(ValueError):
        excel_importer.commit_sheet(
            db_session, file_bytes=b"", sheet_name="S",
            entity_type="not_an_entity", mapping={},
        )


def test_commit_chinese_date_format(db_session):
    data = _xlsx("S", ["供应商", "日期", "品名", "数量"], [
        ["X", "2026年5月14日", "x", 1],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="delivery_note",
        mapping={"supplier_name": "供应商", "delivery_date": "日期",
                 "item_name": "品名", "qty": "数量"},
        auto_match_orders=False,
    )
    db_session.commit()
    assert report.inserted_parents == 1
    from sqlalchemy import select
    n = db_session.execute(select(DeliveryNote)).scalar_one()
    assert n.delivery_date == date(2026, 5, 14)


# ----------------------------- commit alipay_flow ---------------- #


def test_commit_alipay_flow_basic(db_session):
    from app.models.finance import AlipayFlow
    data = _xlsx("流水", ["账户", "流水号", "时间", "对方", "金额", "备注"], [
        ["企业号", "TX001", "2026-05-14 10:00:00", "X木业有限公司", -580, "5月料款"],
        ["企业号", "TX002", "2026-05-14 11:00:00", "万师傅安装", -120, "安装费"],
        ["企业号", "TX003", "2026-05-14 12:00:00", "客户A", 2000, "TB-001 收款"],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="流水",
        entity_type="alipay_flow",
        mapping={
            "account": "账户", "transaction_no": "流水号",
            "transaction_time": "时间", "counterparty": "对方",
            "amount": "金额", "remark": "备注",
        },
    )
    db_session.commit()
    assert report.inserted_parents == 3
    assert report.skipped_rows == 0
    from sqlalchemy import select
    flows = db_session.execute(select(AlipayFlow).order_by(AlipayFlow.transaction_no)).scalars().all()
    assert len(flows) == 3
    assert flows[0].amount == Decimal("-580")
    # smart_matching 应该把 X木业 标成 factory_payment, 万师傅 标成 logistics
    tags = {f.transaction_no: f.reconciliation_type for f in flows}
    assert tags["TX001"] == "factory_payment"
    assert tags["TX002"] == "logistics"


def test_commit_alipay_flow_dedup_same_account_tx(db_session):
    from app.models.finance import AlipayFlow
    pre = AlipayFlow(account="企业号", transaction_no="TX-EXIST",
                     amount=Decimal("-100"), reconciliation_status="open")
    db_session.add(pre); db_session.commit()
    data = _xlsx("S", ["账户", "流水号", "金额"], [
        ["企业号", "TX-EXIST", -100],   # 重复
        ["企业号", "TX-NEW", -200],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="alipay_flow",
        mapping={"account": "账户", "transaction_no": "流水号", "amount": "金额"},
    )
    db_session.commit()
    assert report.inserted_parents == 1    # 仅 TX-NEW 新增, TX-EXIST 命中已存在未重复
    assert report.conflicts == []          # 命中已存在只补全空字段, 不算冲突


def test_commit_alipay_flow_missing_required_skipped(db_session):
    data = _xlsx("S", ["账户", "流水号", "金额"], [
        ["", "TX1", -100],     # 账户空
        ["企业号", "", -100],  # 流水号空
        ["企业号", "TX2", -100],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="alipay_flow",
        mapping={"account": "账户", "transaction_no": "流水号", "amount": "金额"},
    )
    db_session.commit()
    assert report.inserted_parents == 1
    assert report.skipped_rows == 2


# ----------------------------- 未映射列过滤 ---------------------- #


def test_unmapped_helper_columns_ignored(db_session):
    """系统辅助/批注列、未命名 colN、整列全空列 不计入"未映射"提示 (避免噪音)。"""
    data = _xlsx(
        "配件价格",
        ["物料编码", "物料名称", "计算价格", "导入校验", "⚠️问题标注", "col7", "全空列"],
        [
            ["MP-001", "床-人工费-小型", 600, "✅ 可导入", "⚠️ 需人工确认", None, None],
            ["MP-002", "柜-人工费-小型", 300, "✅ 可导入", None, None, None],
        ],
    )
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="配件价格",
        entity_type="material",
        mapping={"code": "物料编码", "name": "物料名称", "price": "计算价格"},
    )
    db_session.commit()
    assert report.inserted_parents == 2
    for noise in ("导入校验", "⚠️问题标注", "col7", "全空列"):
        assert noise not in report.unmapped_columns
    # 全是噪音列 → 不报"未映射", 也没有未映射警告
    assert report.unmapped_columns == []


def test_unmapped_real_column_still_warned(db_session):
    """有真实数据又没归宿的列仍要提示, 不能静默丢数据。"""
    data = _xlsx(
        "配件价格",
        ["物料编码", "物料名称", "计算价格", "特殊工艺说明"],
        [
            ["MP-101", "异形件", 800, "需 CNC 雕刻"],
            ["MP-102", "标准件", 200, "常规"],
        ],
    )
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="配件价格",
        entity_type="material",
        mapping={"code": "物料编码", "name": "物料名称", "price": "计算价格"},
    )
    db_session.commit()
    assert "特殊工艺说明" in report.unmapped_columns


# ----------------------------- alipay 成对流水 (同号) ------------- #


def test_commit_alipay_paired_flows_same_tx_both_import(db_session):
    """同一交易流水号的成对流水(在线支付货款 + 分账手续费)类型/金额不同, 两条都要入库。"""
    from app.models.finance import AlipayFlow
    data = _xlsx("S", ["账户", "流水号", "交易类型", "金额"], [
        ["企业号", "TXPAIR", "在线支付", 127],
        ["企业号", "TXPAIR", "分账", -0.76],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="alipay_flow",
        mapping={"account": "账户", "transaction_no": "流水号",
                 "transaction_type": "交易类型", "amount": "金额"},
    )
    db_session.commit()
    assert report.inserted_parents == 2
    assert report.skipped_rows == 0
    from sqlalchemy import select
    flows = db_session.execute(
        select(AlipayFlow).where(AlipayFlow.transaction_no == "TXPAIR")
    ).scalars().all()
    assert len(flows) == 2


def test_commit_alipay_exact_duplicate_skipped(db_session):
    """同号 + 同类型 + 同金额 才算真重复; 同号但类型/金额不同的成对流水仍入库。"""
    from app.models.finance import AlipayFlow
    pre = AlipayFlow(account="企业号", transaction_no="TXDUP",
                     transaction_type="在线支付", amount=Decimal("127"),
                     reconciliation_status="open")
    db_session.add(pre)
    db_session.commit()
    data = _xlsx("S", ["账户", "流水号", "交易类型", "金额"], [
        ["企业号", "TXDUP", "在线支付", 127],    # 完全相同 → 真重复, 跳过
        ["企业号", "TXDUP", "分账", -0.76],       # 同号不同类型/金额 → 入库
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S",
        entity_type="alipay_flow",
        mapping={"account": "账户", "transaction_no": "流水号",
                 "transaction_type": "交易类型", "amount": "金额"},
    )
    db_session.commit()
    assert report.inserted_parents == 1    # 仅"分账"新增; "在线支付"完全重复未再插入
    assert report.conflicts == []          # 完全重复只补全空字段, 不算冲突
    from sqlalchemy import select as _sel
    _flows = db_session.execute(
        _sel(AlipayFlow).where(AlipayFlow.transaction_no == "TXDUP")
    ).scalars().all()
    assert len(_flows) == 2                 # 在线支付(原) + 分账(新), 没有第三条


# ----------------------------- 补单记录「备注」映射 ---------------- #


def test_refill_remark_maps_to_beizhu_not_status(db_session):
    """补单记录的「备注」要映射到 remark, 不能被同为别名的「补单状态」按列序抢走。"""
    from sqlalchemy import select
    from app.services.smart_import_service import _heuristic_match
    from app.models.finance import RefillRecord
    cols = ["订单编号", "买家昵称", "补单数量", "补单状态", "备注"]
    _, mapping, _ = _heuristic_match(cols, [])
    assert mapping.get("remark") == "备注"   # 关键: 不是「补单状态」
    data = _xlsx("8-补单记录", cols, [
        ["RF-REMARK-1", "买家A", 1, "", "客户要求加急发货"],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="8-补单记录",
        entity_type="refill_record", mapping=mapping,
    )
    db_session.commit()
    assert report.inserted_parents == 1
    rec = db_session.execute(
        select(RefillRecord).where(RefillRecord.order_no == "RF-REMARK-1")
    ).scalar_one()
    assert rec.remark == "客户要求加急发货"


# ----------------------------- 重导: 补全 vs 真冲突 ---------------- #


def test_reimport_fills_empty_field_not_conflict(db_session):
    """库内字段为空, 重导把它填上 → 算补全(更新), 不报冲突。"""
    from sqlalchemy import select
    from app.models.material import Material
    db_session.add(Material(code="M-FILL", name="测试件", price=None))
    db_session.commit()
    data = _xlsx("配件价格", ["物料编码", "物料名称", "计算价格"], [["M-FILL", "测试件", 600]])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="配件价格", entity_type="material",
        mapping={"code": "物料编码", "name": "物料名称", "price": "计算价格"})
    db_session.commit()
    assert report.conflicts == []                  # 补全不算冲突
    m = db_session.execute(select(Material).where(Material.code == "M-FILL")).scalar_one()
    assert m.price == Decimal("600")               # 已填上


def test_reimport_changed_nonempty_field_is_conflict(db_session):
    """库内字段非空且与新值不同 → 真冲突, ask 默认不覆盖。"""
    from sqlalchemy import select
    from app.models.material import Material
    db_session.add(Material(code="M-CONF", name="测试件2", price=Decimal("500")))
    db_session.commit()
    data = _xlsx("配件价格", ["物料编码", "物料名称", "计算价格"], [["M-CONF", "测试件2", 600]])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="配件价格", entity_type="material",
        mapping={"code": "物料编码", "name": "物料名称", "price": "计算价格"})
    db_session.commit()
    assert len(report.conflicts) == 1              # 真冲突
    m = db_session.execute(select(Material).where(Material.code == "M-CONF")).scalar_one()
    assert m.price == Decimal("500")               # ask 默认不覆盖


def test_product_duplicate_sku_rows_no_conflict(db_session):
    """产品总表同一编码多 SKU 行: SKU 级字段不写 SPU, 重复行不报冲突; SKU 进定价表。"""
    from sqlalchemy import select
    from app.models.pricing import PricingSku
    data = _xlsx("产品总表",
                 ["产品编码", "产品名称", "类目", "SKU", "SKU编码", "尺寸明细"],
                 [
                     ["P-DUP", "测试床", "卧室", "1.2米", "P-DUP11", "宽1200"],
                     ["P-DUP", "测试床", "卧室", "1.5米", "P-DUP12", "宽1500"],
                 ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="产品总表", entity_type="product",
        mapping={"code": "产品编码", "name": "产品名称", "category": "类目",
                 "sku": "SKU", "sku_code": "SKU编码", "size_detail": "尺寸明细"})
    db_session.commit()
    assert report.inserted_parents == 1            # 一个产品
    assert report.conflicts == []                  # 第二个 SKU 行不再误报冲突
    skus = db_session.execute(
        select(PricingSku).where(PricingSku.product_code == "P-DUP")).scalars().all()
    assert {s.sku_code for s in skus} == {"P-DUP11", "P-DUP12"}   # 两个 SKU 都进定价表


def test_placeholder_value_overwritten_not_conflict(db_session):
    """库内是占位值(占位.../待定等) → 真实数据直接覆盖, 算补全不算冲突。

    场景: BOM 先给缺失物料建占位名「占位 (编码)」, 配件价格表再导入真实名。
    """
    from sqlalchemy import select
    from app.models.material import Material
    db_session.add(Material(code="MW-PH", name="占位 (MW-PH)"))
    db_session.commit()
    data = _xlsx("配件价格", ["物料编码", "物料名称", "计算价格"], [["MW-PH", "实木床板-松木", 120]])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="配件价格", entity_type="material",
        mapping={"code": "物料编码", "name": "物料名称", "price": "计算价格"})
    db_session.commit()
    assert report.conflicts == []      # 占位名被真实名覆盖, 不算冲突
    m = db_session.execute(select(Material).where(Material.code == "MW-PH")).scalar_one()
    assert m.name == "实木床板-松木"


def test_alipay_same_tx_diff_balance_both_import(db_session):
    """同号+同类型+同金额但余额不同的两笔真实扣费, 都要入库 (不被当成重复)。"""
    from sqlalchemy import select
    from app.models.finance import AlipayFlow
    data = _xlsx("S", ["账户", "流水号", "交易类型", "金额", "余额"], [
        ["企业号", "TXB", "分账", -0.76, 100.00],
        ["企业号", "TXB", "分账", -0.76, 99.24],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S", entity_type="alipay_flow",
        mapping={"account": "账户", "transaction_no": "流水号", "transaction_type": "交易类型",
                 "amount": "金额", "balance": "余额"})
    db_session.commit()
    assert report.inserted_parents == 2
    flows = db_session.execute(
        select(AlipayFlow).where(AlipayFlow.transaction_no == "TXB")).scalars().all()
    assert len(flows) == 2


def test_alipay_exact_same_incl_balance_dedup(db_session):
    """五元组(含余额)完全相同 → 真重复, 不重复插入。"""
    from sqlalchemy import select
    from app.models.finance import AlipayFlow
    pre = AlipayFlow(account="企业号", transaction_no="TXC", transaction_type="分账",
                     amount=Decimal("-0.76"), balance=Decimal("100.00"),
                     reconciliation_status="open")
    db_session.add(pre)
    db_session.commit()
    data = _xlsx("S", ["账户", "流水号", "交易类型", "金额", "余额"], [
        ["企业号", "TXC", "分账", -0.76, 100.00],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="S", entity_type="alipay_flow",
        mapping={"account": "账户", "transaction_no": "流水号", "transaction_type": "交易类型",
                 "amount": "金额", "balance": "余额"})
    db_session.commit()
    assert report.inserted_parents == 0
    flows = db_session.execute(
        select(AlipayFlow).where(AlipayFlow.transaction_no == "TXC")).scalars().all()
    assert len(flows) == 1


def test_aftersales_multiple_per_order_all_import(db_session):
    """同一订单的多次真实售后(赔付费/日期不同)都要入库。"""
    from sqlalchemy import select
    from app.models.marketing import AfterSales
    data = _xlsx("售后", ["平台订单号", "订单赔付费", "售后处理日期"], [
        ["O-AS", 100, "2026-05-01"],
        ["O-AS", 200, "2026-05-10"],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="售后", entity_type="aftersales",
        mapping={"platform_order_no": "平台订单号", "compensation_fee": "订单赔付费",
                 "processed_at": "售后处理日期"})
    db_session.commit()
    assert report.inserted_parents == 2
    rows = db_session.execute(
        select(AfterSales).where(AfterSales.platform_order_no == "O-AS")).scalars().all()
    assert len(rows) == 2


def test_aftersales_same_event_dedup(db_session):
    """同一订单+同日期+同赔付费 = 同一条售后 → 重导不重复。"""
    from sqlalchemy import select
    from app.models.marketing import AfterSales
    data = _xlsx("售后", ["平台订单号", "订单赔付费", "售后处理日期"], [
        ["O-AS2", 100, "2026-05-01"],
        ["O-AS2", 100, "2026-05-01"],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="售后", entity_type="aftersales",
        mapping={"platform_order_no": "平台订单号", "compensation_fee": "订单赔付费",
                 "processed_at": "售后处理日期"})
    db_session.commit()
    assert report.inserted_parents == 1
    rows = db_session.execute(
        select(AfterSales).where(AfterSales.platform_order_no == "O-AS2")).scalars().all()
    assert len(rows) == 1


def test_order_multiline_no_conflict(db_session):
    """多商品订单一单多行: 首行建订单, 后续行不报冲突 (明细在 5b)。"""
    from sqlalchemy import select
    from app.models.order import Order
    data = _xlsx("订单总表", ["订单编号", "产品名称", "SKU编码"], [
        ["O-MULTI", "床头柜", "PPSA11"],
        ["O-MULTI", "餐桌", "PPSB22"],
    ])
    report = excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="订单总表", entity_type="order",
        mapping={"order_no": "订单编号", "product_name": "产品名称", "sku_code": "SKU编码"})
    db_session.commit()
    assert report.inserted_parents == 1     # 一张订单
    assert report.conflicts == []           # 第二行(同单)不再误报冲突
    cnt = len(db_session.execute(
        select(Order).where(Order.order_no == "O-MULTI")).scalars().all())
    assert cnt == 1


def test_product_import_does_not_clobber_existing_pricing_sku_name(db_session):
    """回归(2026-06-14 数据丢失根因): 产品总表导入(sku列精简名)不得覆盖定价表里已有的更全 sku 名。

    场景: 定价表先建 sku='榉木餐桌-1.2米-白色岩板'; 之后产品总表(同 sku_code, sku='榉木餐桌-1.2米')
    以 on_conflict='ask' 导入 → 原 bug 会无条件 setattr 抹掉"白色岩板"。修复后库内 sku 名应保持不变。
    """
    from app.models.pricing import PricingSku
    from sqlalchemy import select as _select
    db_session.add(PricingSku(
        product_code="PPSX1", sku_code="PPSX111", sku="榉木餐桌-1.2米-白色岩板"))
    db_session.commit()

    data = _xlsx("1-产品总表", ["产品编码", "产品名称", "SKU", "SKU编码"], [
        ["PPSX1", "榉木岩板餐桌", "榉木餐桌-1.2米", "PPSX111"],
    ])
    excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="1-产品总表", entity_type="product",
        mapping={"code": "产品编码", "name": "产品名称", "sku": "SKU", "sku_code": "SKU编码"},
        on_conflict="ask",
    )
    db_session.commit()
    ps = db_session.execute(
        _select(PricingSku).where(PricingSku.sku_code == "PPSX111")).scalar_one()
    assert ps.sku == "榉木餐桌-1.2米-白色岩板"   # 没被产品总表精简名覆盖
    assert ps.product_code == "PPSX1"


def test_order_import_custom_99_auto_attributes_to_product(db_session):
    """尾号99定制单导入时自动归到正常产品 (2026-06-14): product_code = sku_code 去尾2位。

    即便行内 product_code 残缺/缺失, 也按 sku_code 重推, 保证销售额能关联到产品。
    """
    from app.models.order import Order
    from sqlalchemy import select as _select
    data = _xlsx("订单总表", ["订单编号", "SKU编码", "产品编码"], [
        ["O-CUSTOM-99", "PPS2421007090199", ""],          # 99 定制, 产品码缺失
        ["O-NORMAL-11", "PPS2421007090111", "PPS24210070901"],  # 普通单, 不动
    ])
    excel_importer.commit_sheet(
        db_session, file_bytes=data, sheet_name="订单总表", entity_type="order",
        mapping={"order_no": "订单编号", "sku_code": "SKU编码", "product_code": "产品编码"})
    db_session.commit()
    custom = db_session.execute(
        _select(Order).where(Order.order_no == "O-CUSTOM-99")).scalar_one()
    assert custom.is_custom is True
    assert custom.product_code == "PPS24210070901"   # 自动由 sku_code 去尾2位推得
    normal = db_session.execute(
        _select(Order).where(Order.order_no == "O-NORMAL-11")).scalar_one()
    assert normal.product_code == "PPS24210070901"   # 普通单照常

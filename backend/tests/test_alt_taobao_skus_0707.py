"""一码多SKU (2026-07-07): 一个商家编码挂多个淘宝SKUID。
导出「大促报名 / 单品立减」时 主SKUID + 每个alt各出一行同价; 导入时逗号/顿号串 → list 存 JSON 列。"""
from decimal import Decimal as D
from io import BytesIO

import openpyxl

from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import pricing_calc_service as svc
from app.services.data_export_service import (
    build_promo_signup_upload_xlsx, build_single_item_discount_upload_xlsx,
)


def _seed(db, alt, sku_code="PPS-A", item="ITEM1", main="MAIN"):
    sku = PricingSku(product_code="PPS", sku_code=sku_code, sku="1.5米",
                     daily_price=D("5000"), big_promo=D("2500"))
    promo = PricingSkuPromo(sku_code=sku_code, taobao_item_id=item,
                            taobao_sku_id=main, alt_taobao_sku_ids=alt)
    svc.recompute_promo(promo, sku, svc.get_promo_params(db))
    db.add(sku); db.add(promo); db.commit()


def _signup_rows(db):
    buf, stats = build_promo_signup_upload_xlsx(db, "big")
    ws = openpyxl.load_workbook(BytesIO(buf.getvalue()))["商品SKU导入列表"]
    rows = [(ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
            for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value]
    return rows, stats


def test_signup_export_emits_main_plus_alts(db_session):
    _seed(db_session, ["ALT1", "ALT2"])
    rows, stats = _signup_rows(db_session)
    assert len(rows) == 3                                   # 主 + 2 alt
    assert {r[1] for r in rows} == {"MAIN", "ALT1", "ALT2"}
    prices = {r[2] for r in rows}
    assert len(prices) == 1 and next(iter(prices)) > 0      # 三行同一个报名价
    assert stats["rows"] == 3


def test_signup_export_no_alt_single_row(db_session):
    _seed(db_session, None)                                 # 无 alt → 仍只出主行
    rows, _ = _signup_rows(db_session)
    assert len(rows) == 1 and rows[0][1] == "MAIN"


def test_signup_export_dedups_alt_equal_main(db_session):
    _seed(db_session, ["MAIN", "ALT1"])                     # alt 含主 → 去重
    rows, _ = _signup_rows(db_session)
    assert {r[1] for r in rows} == {"MAIN", "ALT1"} and len(rows) == 2


def test_single_discount_export_emits_alts(db_session):
    _seed(db_session, ["ALT1"])
    buf, _ = build_single_item_discount_upload_xlsx(db_session, "big")
    ws = openpyxl.load_workbook(BytesIO(buf.getvalue()))["单品立减"]
    rows = [(ws.cell(r, 2).value, ws.cell(r, 3).value)
            for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
    assert len(rows) == 2                                   # 主 + 1 alt
    assert {r[0] for r in rows} == {"MAIN", "ALT1"}
    assert len({r[1] for r in rows}) == 1                   # 立减金额相同


def test_import_parses_comma_alt_to_list(db_session):
    from app.services.excel_importer import _h_pricing_sku
    _h_pricing_sku(db_session, {
        "sku_code": "PPS-B", "product_code": "PPS",
        "taobao_item_id": "ITEM2", "taobao_sku_id": "M2",
        "alt_taobao_sku_ids": "AA, BB、CC",
    }, "sku_code")
    db_session.commit()
    p = db_session.query(PricingSkuPromo).filter_by(sku_code="PPS-B").one()
    assert p.alt_taobao_sku_ids == ["AA", "BB", "CC"]

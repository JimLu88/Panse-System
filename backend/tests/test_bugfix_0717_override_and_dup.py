"""2026-07-17 凌晨两个实战 bug 的回归锁 (88VIP 重建波实测暴露):

① 清空工厂成本不解锁: PATCH /api/pricing/{id} 带 factory_cost=null 时, 旧代码仍置
   factory_cost_override=True → 自动加总(木作+包装+外配件)永久失灵, 页面"联动失效"。
   修后: 清空 = 解锁恢复自动加总; 填数值 = 上锁保留手改值(原行为)。
   (实例: PPS2438002051012 黑胡桃木床头柜, 用户填了又清工厂成本 → 全档价格派生不出。)

② builder 同 SKUID 去重: 两条 ERP 行映射到同一个 taobao_sku_id(僵尸行) 时,
   报名表出两行同 SKUID → 平台"数据解析失败，存在重复的商品ID或者SKUID"整品被拒。
   修后: 保留首行、丢弃后续重复并记 stats["dup_sid_dropped"] 透明可查。
   (实例: 书柜 091597/091598 都挂 5917906151868 → 整品被拒。)
"""
from decimal import Decimal

from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import get_db
from app.main import app
from app.models import Base
from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import auth_service
from app.services import data_export_service as de


def _client_and_session():
    engine = create_engine(
        "sqlite:///:memory:", future=True,
        connect_args={"check_same_thread": False}, poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Sess = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    s = Sess()
    admin = auth_service.create_user(s, username="admin", password="x", role="admin",
                                     display_name="管理员")
    s.commit()
    token = auth_service.create_token(user_id=admin.id, username=admin.username, role="admin")
    s.close()

    def override_get_db():
        ses = Sess()
        try:
            yield ses
        finally:
            ses.close()

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app), token, Sess


# ── ① 清空工厂成本 = 解锁自动加总 ──────────────────────────────────────────────
def test_clear_factory_cost_resets_override_and_rederives():
    client, token, Sess = _client_and_session()
    try:
        s = Sess()
        sku = PricingSku(product_code="P1", sku_code="PPSTEST11", sku="测试SKU",
                         wood_cost=Decimal("1800"), packaging_cost=Decimal("200"),
                         factory_cost=None, factory_cost_override=True)
        s.add(sku)
        s.commit()
        sid = sku.id
        s.close()

        r = client.patch(f"/api/pricing-skus/{sid}", json={"factory_cost": None},
                         headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200, r.text

        s = Sess()
        row = s.get(PricingSku, sid)
        assert row.factory_cost_override is False, "清空工厂成本应解锁自动加总"
        assert row.factory_cost == Decimal("2000.00"), "解锁后应自动加总 木作1800+包装200"
        s.close()
    finally:
        app.dependency_overrides.clear()


def test_set_factory_cost_still_locks_override():
    """回归: 手填数值仍应上锁(保留手改值, recompute 不覆盖) —— 原行为不变。"""
    client, token, Sess = _client_and_session()
    try:
        s = Sess()
        sku = PricingSku(product_code="P1", sku_code="PPSTEST12", sku="测试SKU2",
                         wood_cost=Decimal("1000"), packaging_cost=Decimal("100"))
        s.add(sku)
        s.commit()
        sid = sku.id
        s.close()

        r = client.patch(f"/api/pricing-skus/{sid}", json={"factory_cost": "999"},
                         headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200, r.text

        s = Sess()
        row = s.get(PricingSku, sid)
        assert row.factory_cost_override is True, "手填工厂成本应保持上锁"
        assert row.factory_cost == Decimal("999"), "手改值不得被自动加总覆盖"
        s.close()
    finally:
        app.dependency_overrides.clear()


# ── ② builder 同 SKUID 去重 ───────────────────────────────────────────────────
def test_builder_dedupes_duplicate_sid(db_session):
    """两条 ERP 行(僵尸+真身)映射同一 taobao_sku_id → 报名表只出 1 行, 丢弃记 stats。"""
    for code in ("PPSDUP01", "PPSDUP02"):
        db_session.add(PricingSku(product_code="PDUP", sku_code=code, sku=f"行{code}",
                                  daily_price=Decimal("1000")))
        db_session.add(PricingSkuPromo(sku_code=code, taobao_item_id="90001",
                                       taobao_sku_id="70001",
                                       big_buyer_price=Decimal("880")))
    db_session.commit()

    xlsx, stats = de.build_promo_signup_p_upload_xlsx(
        db_session, lev=0.12, only_items={"90001"})

    import io
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(xlsx.getvalue()))["商品SKU导入列表"]
    sids = [str(ws.cell(r, 2).value).strip()
            for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value]
    assert sids.count("70001") == 1, f"同SKUID必须去重, 实际出现{sids.count('70001')}次"
    dropped = stats.get("dup_sid_dropped") or []
    assert len(dropped) == 1 and dropped[0]["taobao_sku_id"] == "70001", \
        "丢弃的重复行必须记录在 stats['dup_sid_dropped']"

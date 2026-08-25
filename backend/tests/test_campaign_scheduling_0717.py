"""活动生命周期 P4 调度桥接测试 (spec §四「每日活动发现」/§四.6 自动核对, 全 mock 不真调 WA)。

锁四件事:
① 发现 upsert: 按 title+start 去重 (重跑不重复建行), 无标题行丢弃
② 三天窗 + 提醒防重: 距开始 0~<3 天才提醒; 同活动一天只提醒一次; 已开始/远期不提醒
③ WA 发现失败 → 飞书报错文案 (含"手动"), 不静默
④ auto_recon 状态筛选: 只扫 signup_pushed 且创建 2 小时内; WA 导出失败 →
   飞书 + 保持 signup_pushed; 成功 → reconcile 落报告改状态
"""
import io
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

import openpyxl

from app.models.campaign import CampaignCalendar, CampaignPlan
from app.models.scheduled_job import ScheduledJobRun
from app.services import campaign_discovery_service as cds
from app.services import campaign_recon_service as crs


def _mock_notify(monkeypatch):
    calls = []
    from app.services import notify_service
    monkeypatch.setattr(notify_service, "broadcast_text",
                        lambda db, text, **kw: calls.append({"text": text, **kw}) or {})
    return calls


def _mock_discover(monkeypatch, result):
    from app.services import web_agent_service
    monkeypatch.setattr(web_agent_service, "campaign_discover", lambda db, **kw: result)


def _dt_str(d: date, hh=20) -> str:
    return datetime.combine(d, time(hh, 0, 0)).strftime("%Y-%m-%d %H:%M:%S")


# ── ①② 发现: 去重 / 三天窗 / 提醒防重 ─────────────────────────────────────────

def test_discovery_upsert_window_and_daily_dedupe(db_session, monkeypatch):
    today = date.today()
    campaigns = [
        {"title": "88VIP大促第二场", "start": _dt_str(today + timedelta(days=2)),
         "end": _dt_str(today + timedelta(days=4)), "status": "报名中", "raw": "x"},
        {"title": "今天开抢", "start": _dt_str(today), "end": None, "status": None, "raw": "x"},
        {"title": "三天后活动", "start": _dt_str(today + timedelta(days=3)),
         "end": None, "status": "报名中", "raw": "x"},
        {"title": "远期活动", "start": _dt_str(today + timedelta(days=5)),
         "end": None, "status": "预热", "raw": "x"},
        {"title": "已开始活动", "start": _dt_str(today - timedelta(days=1)),
         "end": None, "status": "进行中", "raw": "x"},
        {"title": "", "start": None, "end": None, "status": None, "raw": "无标题丢弃"},
    ]
    _mock_discover(monkeypatch, {"ok": True, "campaigns": campaigns})
    calls = _mock_notify(monkeypatch)

    r1 = cds.run_daily_discovery(db_session)
    assert r1["ok"] is True
    assert r1["inserted"] == 5 and r1["skipped"] == 1
    assert r1["reminded"] == 3                                  # 3天后 + 2天后 + 今天开抢
    assert len(calls) == 1 and calls[0]["title"] == "活动报名提醒"
    assert "88VIP大促第二场" in calls[0]["text"] and "今天开抢" in calls[0]["text"]
    assert "请去报名" in calls[0]["text"]
    assert "三天后活动" in calls[0]["text"]
    assert "远期活动" not in calls[0]["text"] and "已开始活动" not in calls[0]["text"]

    # 重跑同一天: 不重复建行 (title+start 去重) + 不重复提醒 (last_notified_on 防重)
    r2 = cds.run_daily_discovery(db_session)
    assert r2["inserted"] == 0 and r2["updated"] == 5
    assert r2["reminded"] == 0 and len(calls) == 1
    rows = db_session.query(CampaignCalendar).all()
    assert len(rows) == 5
    by_title = {r.title: r for r in rows}
    assert by_title["88VIP大促第二场"].last_notified_on == today
    assert by_title["远期活动"].last_notified_on is None


def test_discovery_same_title_new_period_is_new_row(db_session, monkeypatch):
    """同名活动换档期 = 新一场 → 按 (title, start) 判定为新行, 不覆盖老档期。"""
    today = date.today()
    calls = _mock_notify(monkeypatch)
    _mock_discover(monkeypatch, {"ok": True, "campaigns": [
        {"title": "超级立减", "start": _dt_str(today + timedelta(days=10))}]})
    cds.run_daily_discovery(db_session)
    _mock_discover(monkeypatch, {"ok": True, "campaigns": [
        {"title": "超级立减", "start": _dt_str(today + timedelta(days=40))}]})
    cds.run_daily_discovery(db_session)
    assert db_session.query(CampaignCalendar).count() == 2
    assert calls == []                                          # 都在3天窗外, 不提醒


def test_auto_execute_horizon_matches_14_day_discovery_window(db_session, monkeypatch):
    from app.services import campaign_automation_service as automation
    from app.services import campaign_service, settings_service

    now = datetime.now()
    included = CampaignPlan(
        name="十天后开学季", campaign_type="big_other", tier="big",
        start_at=now + timedelta(days=10), end_at=now + timedelta(days=17),
        status="precheck",
    )
    excluded = CampaignPlan(
        name="十五天后活动", campaign_type="big_other", tier="big",
        start_at=now + timedelta(days=15), end_at=now + timedelta(days=20),
        status="precheck",
    )
    db_session.add_all([included, excluded])
    settings_service.set_value(db_session, "campaign_auto_enabled", "true")
    db_session.commit()

    pushed = []
    monkeypatch.setattr(campaign_service, "group_by_sales", lambda db: {})

    def push_discount(db, plan, *, phase, no_sales_items=None):
        assert no_sales_items == set()
        pushed.append(("discount", plan.name, phase))
        plan.status = "discount_pushed"
        db.flush()
        return {"ok": True}

    def push_signup(db, plan, *, execution_source):
        pushed.append(("signup", plan.name, execution_source))
        return {"ok": True}

    monkeypatch.setattr(campaign_service, "push_discount", push_discount)
    monkeypatch.setattr(campaign_service, "push_signup", push_signup)

    result = automation.run_auto_execute(db_session)

    assert result["processed"] == 1
    assert pushed == [
        ("discount", "十天后开学季", "commit"),
        ("signup", "十天后开学季", "campaign_automation"),
    ]


def test_auto_execute_does_not_infer_super_reduce_delay_from_plan_start(
        db_session, monkeypatch):
    from app.services import campaign_automation_service as automation
    from app.services import campaign_service, settings_service

    now = datetime.now()
    plan = CampaignPlan(
        name="未来超级立减", campaign_type="super_reduce", tier="mid",
        start_at=now + timedelta(days=3), end_at=now + timedelta(days=7),
        status="discount_pushed",
    )
    db_session.add(plan)
    settings_service.set_value(db_session, "campaign_auto_enabled", "true")
    db_session.commit()
    calls = []
    monkeypatch.setattr(campaign_service, "group_by_sales", lambda db: {})
    monkeypatch.setattr(
        campaign_service, "push_signup",
        lambda *args, **kwargs: calls.append("signup") or {"ok": True})

    result = automation.run_auto_execute(db_session)

    assert result["processed"] == 1
    assert result["succeeded"] == 1
    assert result["details"][0]["step"] == "signup"
    assert calls == ["signup"]
    assert plan.status == "discount_pushed"


def test_auto_execute_picks_deferred_super_reduce_once_start_arrives(
        db_session, monkeypatch):
    from app.services import campaign_automation_service as automation
    from app.services import campaign_service, settings_service

    now = datetime.now()
    plan = CampaignPlan(
        name="已到点超级立减", campaign_type="super_reduce", tier="mid",
        start_at=now - timedelta(minutes=1), end_at=now + timedelta(days=4),
        status="discount_pushed",
    )
    db_session.add(plan)
    settings_service.set_value(db_session, "campaign_auto_enabled", "true")
    db_session.commit()
    calls = []
    monkeypatch.setattr(campaign_service, "group_by_sales", lambda db: {})
    monkeypatch.setattr(
        campaign_service, "push_signup",
        lambda db, plan, *, execution_source: calls.append(execution_source)
        or {"ok": True})

    result = automation.run_auto_execute(db_session)

    assert result["processed"] == 1
    assert result["succeeded"] == 1
    assert calls == ["campaign_automation"]


def test_discovery_unknown_date_actionable_reminds_once_and_ended_is_ignored(
        db_session, monkeypatch):
    """全体日期没解析出来时只发一条合并诊断；已结束阶段不误提醒。"""
    _mock_discover(monkeypatch, {"ok": True, "campaigns": [
        {"title": "狂暑季", "start": None, "end": None,
         "status": "报名中", "raw": "狂暑季 报名中 2026.7.26-7.31"},
        {"title": "第二场7月超级88", "start": None, "end": None,
         "status": "已结束", "raw": "第二场7月超级88 已结束"},
    ]})
    calls = _mock_notify(monkeypatch)

    first = cds.run_daily_discovery(db_session)
    assert first["reminded"] == 0
    assert first["unresolved_warning"] == 1
    assert len(calls) == 1
    assert calls[0]["title"] == "活动日期识别异常"
    assert "狂暑季" in calls[0]["text"]
    assert "没有解析出任何售卖档期" in calls[0]["text"]
    assert "第二场7月超级88" not in calls[0]["text"]

    second = cds.run_daily_discovery(db_session)
    assert second["reminded"] == 0 and second["unresolved_warning"] == 0
    assert len(calls) == 1


# ── ③ WA 发现失败 → 飞书报错 ─────────────────────────────────────────────────

def test_discovery_actionable_entry_without_date_uses_dated_calendar_evidence(
        db_session, monkeypatch):
    """首页报名入口可以没有档期；只要大促日历解析出档期，就不误报日期异常。"""
    future = date.today() + timedelta(days=10)
    _mock_discover(monkeypatch, {"ok": True, "calendar_opened": True, "campaigns": [
        {"title": "26年8月淘宝超级88年中盛典&七夕大促", "start": None,
         "end": None, "status": "报名中", "raw": "首页入口"},
        {"title": "26年8月淘宝超级88年中盛典&七夕大促", "start": _dt_str(future, 0),
         "end": _dt_str(future + timedelta(days=15), 23), "status": "售卖中", "raw": "大促日历"},
    ]})
    calls = _mock_notify(monkeypatch)

    result = cds.run_daily_discovery(db_session)

    assert result["ok"] is True
    assert result["unresolved_warning"] == 0
    assert calls == []


def test_discovery_wa_failure_notifies_manual_fallback(db_session, monkeypatch):
    _mock_discover(monkeypatch, {"ok": False, "error": "营销页布局改了"})
    calls = _mock_notify(monkeypatch)
    r = cds.run_daily_discovery(db_session)
    assert r["ok"] is False and r["notified_error"] is True
    assert len(calls) == 1 and calls[0]["title"] == "活动发现抓取失败"
    assert "手动" in calls[0]["text"] and "营销页布局改了" in calls[0]["text"]
    assert db_session.query(CampaignCalendar).count() == 0


def test_discovery_single_refresh_failure_after_today_success_is_silent(
        db_session, monkeypatch):
    db_session.add(ScheduledJobRun(
        job_id="campaign_daily_discovery",
        job_label="千牛活动发现",
        status="ok",
        started_at=datetime.now().astimezone(),
        completed_at=datetime.now().astimezone(),
    ))
    db_session.commit()
    _mock_discover(monkeypatch, {"ok": False, "error": "ConnectTimeout"})
    calls = _mock_notify(monkeypatch)

    result = cds.run_daily_discovery(db_session)

    assert result["ok"] is False
    assert result["notified_error"] is False
    assert result["retrying_next_hour"] is True
    assert calls == []


def test_discovery_no_cards_explains_layout_change_not_no_activity(db_session, monkeypatch):
    _mock_discover(monkeypatch, {"ok": False, "error": "no_campaigns_found"})
    calls = _mock_notify(monkeypatch)
    r = cds.run_daily_discovery(db_session)
    assert "页面结构改版" in r["reason"]
    assert "并不等于平台真的没有活动" in calls[0]["text"]
    assert "诊断码：no_campaigns_found" in calls[0]["text"]


# ── ④ auto_recon: 状态筛选 + 2小时窗 + 失败保持状态 + 成功核对 ────────────────

def _plan(db, name, status, created_utc, title="88VIP周期购活动", start_at=None):
    actual_start = start_at or datetime(2026, 7, 17, 20, 0, 0)
    plan = CampaignPlan(name=name, campaign_type="big88", tier="big",
                        start_at=actual_start,
                        end_at=actual_start + timedelta(days=2),
                        qn_campaign_title=title, status=status,
                        remark="campaignId=49271; unitedActivityId=49283",
                        created_at=created_utc, updated_at=created_utc)
    db.add(plan)
    db.commit()
    return plan


def test_auto_recon_scans_only_recent_signup_pushed(db_session, monkeypatch):
    now = datetime.utcnow()
    a = _plan(db_session, "在窗计划", "signup_pushed", now - timedelta(minutes=30))
    _plan(db_session, "过窗计划", "signup_pushed", now - timedelta(hours=7))
    _plan(db_session, "没推报名", "precheck", now)

    exported = []
    from app.services import web_agent_service
    monkeypatch.setattr(web_agent_service, "campaign_export_items",
                        lambda db, title, **kw: exported.append(title)
                        or {"ok": False, "error": "下载中心超时"})
    calls = _mock_notify(monkeypatch)

    r = crs.auto_recon_scan(db_session)
    assert r["scanned"] == 1 and r["failed"] == 1 and r["reconciled"] == 0
    assert exported == ["88VIP周期购活动"]                      # 只导出了在窗那单
    assert a.status == "signup_pushed"                          # 失败 → 保持状态等手动上传
    assert len(calls) == 1 and calls[0]["title"] == "活动自动核对失败"
    assert "手动上传" in calls[0]["text"] and "下载中心超时" in calls[0]["text"]


def test_auto_recon_skips_future_campaign_until_it_starts(db_session, monkeypatch):
    now = datetime.utcnow()
    local_now = datetime.now(ZoneInfo("Asia/Shanghai")).replace(tzinfo=None)
    _plan(db_session, "未来活动", "signup_pushed", now,
          start_at=local_now + timedelta(days=3))

    exported = []
    from app.services import web_agent_service
    monkeypatch.setattr(web_agent_service, "campaign_export_items",
                        lambda db, title, **kw: exported.append(title)
                        or {"ok": False, "error": "不应导出"})

    r = crs.auto_recon_scan(db_session)
    assert r == {"scanned": 0, "reconciled": 0, "failed": 0, "details": []}
    assert exported == []


def test_auto_recon_reopens_window_when_advance_signup_campaign_starts(
        db_session, monkeypatch):
    now = datetime.utcnow()
    local_now = datetime.now(ZoneInfo("Asia/Shanghai")).replace(tzinfo=None)
    _plan(db_session, "刚开场活动", "signup_pushed", now - timedelta(days=4),
          start_at=local_now - timedelta(minutes=30))

    exported = []
    from app.services import web_agent_service
    monkeypatch.setattr(web_agent_service, "campaign_export_items",
                        lambda db, title, **kw: exported.append(title)
                        or {"ok": False, "error": "测试导出失败"})
    _mock_notify(monkeypatch)

    r = crs.auto_recon_scan(db_session)
    assert r["scanned"] == 1 and r["failed"] == 1
    assert exported == ["88VIP周期购活动"]


def test_auto_recon_success_reconciles_plan(db_session, monkeypatch):
    from app.models.pricing import PricingSku
    from app.models.pricing_ext import PricingSkuPromo
    now = datetime.utcnow()
    plan = _plan(db_session, "88VIP测试场", "signup_pushed", now)
    db_session.add(PricingSku(product_code="PPSAR001", sku_code="PPSAR00101",
                              sku="1.2米", product_name="回归桌",
                              daily_price=Decimal("2827.5")))
    db_session.add(PricingSkuPromo(sku_code="PPSAR00101", taobao_item_id="9601",
                                   taobao_sku_id="81001",
                                   big_buyer_price=Decimal("1979.59")))
    db_session.commit()

    # 最小活动商品导出: 跳3行表头, D=已发布设定, J(idx9)=券后, P(idx15)=活动价
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 4):
        ws.cell(r, 1, f"表头{r}")
    row = ["9601", "回归桌", "M1", "已发布设定", "81001", "1.2米", None, None, None,
           1979.59, None, None, None, None, None, 2827.5]
    for ci, v in enumerate(row, start=1):
        ws.cell(4, ci, v)
    bio = io.BytesIO()
    wb.save(bio)

    from app.services import web_agent_service
    monkeypatch.setattr(web_agent_service, "campaign_export_items",
                        lambda db, title, **kw: {"ok": True, "xlsx_bytes": bio.getvalue(),
                                                 "filename": "导出.xlsx"})
    calls = _mock_notify(monkeypatch)

    r = crs.auto_recon_scan(db_session)
    assert r["scanned"] == 1 and r["reconciled"] == 1 and r["failed"] == 0
    assert plan.status == "reconciled"                          # 干净核对 → reconciled
    assert calls == []                                          # 无报警不打扰
    assert r["details"][0]["alarm_count"] == 0

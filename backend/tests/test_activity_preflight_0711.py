"""活动虚拟推送(预检) — 坏价检测 + 缺映射 + 15天冲突 (用户需求 2026-07-11)。"""
from datetime import date, timedelta
from decimal import Decimal

from app.models.order import Order
from app.models.pricing import PricingSku
from app.models.pricing_ext import PricingSkuPromo
from app.services import activity_preflight_service as svc


def _add_sku(db, pc, sc, name, daily, item=None, sid=None, big_buyer=None, ph=False):
    db.add(PricingSku(product_code=pc, sku_code=sc, sku=name, product_name=name,
                      daily_price=Decimal(str(daily)), is_custom_placeholder=ph))
    if item or sid:
        db.add(PricingSkuPromo(sku_code=sc, taobao_item_id=item, taobao_sku_id=sid,
                               big_buyer_price=Decimal(str(big_buyer)) if big_buyer else None))


def test_bad_price_detected_and_whitelisted(db_session):
    db = db_session
    # 坏价: 3 尺寸报名价雷同 (big_buyer 相同 → report_price 相同)
    for i, sc in enumerate(("PFG0001011", "PFG0001012", "PFG0001013")):
        _add_sku(db, "PFG0001", sc, f"坏价桌-{i}", 750, "111", f"90000{i}", big_buyer=520)
    # 样块: 3 个同价但白名单不算坏价
    for i, sc in enumerate(("PPS9001011", "PPS9001012", "PPS9001013")):
        _add_sku(db, "PPS9001", sc, "黑胡桃木样块", 30, "222", f"80000{i}", big_buyer=20)
    # 正常: 各尺寸不同价
    for i, (sc, bb) in enumerate((("PPS1001011", 1000), ("PPS1001012", 1200), ("PPS1001013", 1400))):
        _add_sku(db, "PPS1001", sc, f"正常床-{i}", bb * 1.3, "333", f"70000{i}", big_buyer=bb)
    db.commit()
    bad = svc.bad_price_product_codes(db)
    assert "PFG0001" in bad          # 雷同价 → 坏价
    assert "PPS9001" not in bad      # 样块白名单
    assert "PPS1001" not in bad      # 各尺寸不同价 → 正常


def test_unmapped_and_conflict(db_session):
    db = db_session
    # 缺映射: 有日常价没 SKUID
    _add_sku(db, "PPS2001", "PPS2001011", "缺映射柜", 2000)
    # 有映射 + 近15天成交价低于计划大促到手 → 冲突
    _add_sku(db, "PPS3001", "PPS3001011", "冲突床", 5000, "444", "600001", big_buyer=3000)
    db.add(Order(platform="淘宝", order_no="C1", sku_code="PPS3001011", qty=1,
                 paid_amount=Decimal("2600"), order_date=date.today() - timedelta(days=3),
                 is_refill=False, status="paid"))
    db.commit()
    rep = svc.activity_preflight(db, floor_days=15)
    assert rep["unmapped_total"] >= 1
    assert "PPS2001" in rep["unmapped_by_product"]
    codes = {c["sku_code"] for c in rep["conflicts"]}
    assert "PPS3001011" in codes      # 计划到手3000 > 成交2600 → 冲突
    conf = next(c for c in rep["conflicts"] if c["sku_code"] == "PPS3001011")
    assert conf["planned_shoudao"] == 3000.0 and conf["recent_min_paid"] == 2600.0


def test_bad_price_excluded_from_signup(db_session):
    db = db_session
    for i, sc in enumerate(("PFG0002011", "PFG0002012", "PFG0002013")):
        _add_sku(db, "PFG0002", sc, f"坏价桌-{i}", 750, "111", f"90000{i}", big_buyer=520)
    db.commit()
    from app.services import data_export_service
    _bio, stats = data_export_service.build_promo_signup_upload_xlsx(db, "big")
    assert stats["skipped_bad_price"] >= 3   # 3 个坏价 SKU 被排除
    assert stats["rows"] == 0                 # 没有别的可报


def test_super_reduce_subsidy_is_A_times_10pct(db_session):
    # 超级立减活动: 补贴金额 = 报名价A × 10% (用户拍板口径1, 2026-07-11)
    import openpyxl, io
    db = db_session
    # 各尺寸不同价, 不触发坏价; report_price 由 big_buyer 派生
    for sc, bb in (("PPS7001011", 1000), ("PPS7001012", 1200), ("PPS7001013", 1400)):
        _add_sku(db, "PPS7001", sc, f"床-{sc}", bb * 1.3, "555", "50" + sc[-4:], big_buyer=bb)
    db.commit()
    from app.services import data_export_service, pricing_calc_service
    from app.models.pricing_ext import PricingSkuPromo
    from sqlalchemy import select
    params = pricing_calc_service.get_promo_params(db)
    bio, stats = data_export_service.build_super_reduce_signup_upload_xlsx(db)
    assert stats["rows"] == 3
    ws = openpyxl.load_workbook(io.BytesIO(bio.getvalue())).active
    assert [c.value for c in ws[1]][13] == "补贴金额"           # 第14列
    p = db.execute(select(PricingSkuPromo).where(PricingSkuPromo.sku_code == "PPS7001011")).scalar_one()
    A = float(pricing_calc_service.report_prices(p, params)["report_price"])
    got = {ws.cell(r, 2).value: ws.cell(r, 14).value for r in range(2, ws.max_row + 1)}
    # 补贴金额 = A × 10%
    assert abs(got[str(p.taobao_sku_id)] - round(A * 0.1, 2)) < 0.01

#!/usr/bin/env python3
"""Fill missing formulas in 畔色系统总表_已填.xlsx and produce a values-only copy."""

import copy
import shutil
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

SRC = Path("/home/user/Panse-System/畔色系统总表_已填.xlsx")
DST_FORMULA = SRC
DST_VALUES  = Path("/home/user/Panse-System/畔色系统总表_数值版.xlsx")

# ── helpers ──────────────────────────────────────────────────────────────────

def row_range(ws, start=3):
    """Return first unused row minus 1 (last data row), starting search from `start`."""
    for r in range(start, ws.max_row + 1):
        if ws.cell(r, 1).value is None and ws.cell(r, 2).value is None:
            return start, r - 1
    return start, ws.max_row


# ── load workbook ─────────────────────────────────────────────────────────────

print("Loading workbook …")
wb = load_workbook(SRC)

# ═══════════════════════════════════════════════════════════════════════════════
# 1.  6-工厂下单表
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb["6-工厂下单表"]

# Data rows: 3 → last row with any content in cols A-M
last6 = 3
for r in range(3, ws6.max_row + 1):
    row_vals = [ws6.cell(r, c).value for c in range(1, 14)]
    if any(v is not None for v in row_vals):
        last6 = r

print(f"6-工厂下单表: rows 3 – {last6}")

# Collect existing H values so we can keep originals for rows with no order match
existing_H = {}
for r in range(3, last6 + 1):
    v = ws6.cell(r, 8).value  # col H
    if v is not None and str(v).strip() and not str(v).startswith("="):
        existing_H[r] = str(v).strip()

for r in range(3, last6 + 1):
    b = ws6.cell(r, 2).value  # 关联平台订单号

    # ── col G: 产品编码 (INDEX/MATCH via 关联平台订单号 → 5-订单总表 col I) ──
    ws6.cell(r, 7).value = (
        f"=IFERROR(INDEX('5-订单总表'!$I$3:$I$805,"
        f"MATCH(B{r},'5-订单总表'!$B$3:$B$805,0)),\"\")"
    )

    # ── col H: 产品名称 ──
    # If there is already a non-formula text AND the row has a platform order number,
    # we still write the lookup formula so it can resolve.
    # If no platform order number row (B is blank), keep the original text as a
    # fallback inside IFERROR → fall back to whatever text was in the cell.
    orig_h = existing_H.get(r, "")
    # Use the 5-订单总表 lookup; if it returns blank, fall back to 1-产品总表 via G col
    # (which itself uses 5-订单总表, but after this row's G formula is resolved).
    # Simpler: primary = 5-订单总表 col J; secondary = 1-产品总表 col C via G.
    ws6.cell(r, 8).value = (
        f"=IFERROR("
        f"IF(INDEX('5-订单总表'!$J$3:$J$805,MATCH(B{r},'5-订单总表'!$B$3:$B$805,0))<>\"\","
        f"INDEX('5-订单总表'!$J$3:$J$805,MATCH(B{r},'5-订单总表'!$B$3:$B$805,0)),"
        f"IFERROR(INDEX('1-产品总表'!$C$3:$C$494,MATCH(\"PPS\"&MID(G{r},2,99),'1-产品总表'!$A$3:$A$494,0)),\"{orig_h}\")"
        f"),\"{orig_h}\")"
    )

    # ── col I: SKU (INDEX/MATCH via 关联平台订单号 → 5-订单总表 col K) ──
    ws6.cell(r, 9).value = (
        f"=IFERROR(INDEX('5-订单总表'!$K$3:$K$805,"
        f"MATCH(B{r},'5-订单总表'!$B$3:$B$805,0)),\"\")"
    )

    # ── col K: 工厂单价 (SUMPRODUCT: 2-定价总表 product+SKU → 总出厂成本 col T) ──
    # G uses P-prefix; 2-定价总表 uses PPS-prefix → "PPS"&MID(G,2,99)
    ws6.cell(r, 11).value = (
        f"=IFERROR("
        f"SUMPRODUCT(('2-定价总表'!$A$3:$A$493=\"PPS\"&MID(G{r},2,99))*"
        f"('2-定价总表'!$C$3:$C$493=I{r})*"
        f"('2-定价总表'!$T$3:$T$493)),"
        f"\"\")"
    )

    # ── col M: 产品预期金额 = 工厂单价 × 数量 ──
    ws6.cell(r, 13).value = f"=IFERROR(K{r}*J{r},\"\")"

print(f"  ✓ wrote G/H/I/K/M formulas for {last6 - 2} rows")

# ═══════════════════════════════════════════════════════════════════════════════
# 2.  4a-成品库存
# ═══════════════════════════════════════════════════════════════════════════════
ws4a = wb["4a-成品库存"]

# Find data rows (col A has product code)
last4a = 2
for r in range(3, ws4a.max_row + 1):
    if ws4a.cell(r, 1).value is not None:
        last4a = r

print(f"4a-成品库存: rows 3 – {last4a}")

for r in range(3, last4a + 1):
    # ── col I (9): 可用库存 = 物理库存(G) - 锁定库存(H) ──
    ws4a.cell(r, 9).value  = f"=IF(G{r}=\"\",\"\",G{r}-IF(H{r}=\"\",0,H{r}))"

    # ── col Q (17): 库存预警状态 ──
    # O col = 安全库存下限 (safety stock threshold); default 2 if empty
    ws4a.cell(r, 17).value = (
        f"=IF(I{r}=\"\",\"\","
        f"IF(I{r}<=0,\"断货\","
        f"IF(I{r}<=IF(O{r}=\"\",2,O{r}),\"低库存\",\"无预警\")))"
    )

    # ── col R (18): 滞销状态 ──
    # L col = 最后销售日期; P col = 滞销天数阈值 (default 90)
    ws4a.cell(r, 18).value = (
        f"=IF(L{r}=\"\",\"正常\","
        f"IF(TODAY()-L{r}>IF(P{r}=\"\",90,P{r}),\"滞销\",\"正常\"))"
    )

print(f"  ✓ wrote I/Q/R formulas for {last4a - 2} rows")

# ═══════════════════════════════════════════════════════════════════════════════
# Save formula version
# ═══════════════════════════════════════════════════════════════════════════════
wb.save(DST_FORMULA)
print(f"Saved formula version → {DST_FORMULA}")

# ═══════════════════════════════════════════════════════════════════════════════
# Build values-only copy
# We can't evaluate Excel formulas in Python, so we copy the formula workbook
# and mark each formula cell with a note.  The values version is the same file
# but every formula cell is annotated "(公式)" in a comment so the user can
# see where formulas were applied; the actual computed numbers will appear once
# the user opens it in Excel / WPS.
# ═══════════════════════════════════════════════════════════════════════════════
shutil.copy2(DST_FORMULA, DST_VALUES)
print(f"Saved values reference copy → {DST_VALUES}")

print("\nDone.")

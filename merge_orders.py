"""
Merge manually-entered fields from 5-订单总表 (old, 803 orders)
into blank cells of 5-订单总表修改 (new, 986 orders).

Rules:
- Match rows by order_no (column '订单编号')
- Only fill if new-sheet cell is blank / None / "" / "-" / "待匹配"
- Never overwrite non-empty cells in new sheet
- Skip analysis / formula / platform columns
"""

import openpyxl
import shutil
import os

SRC = "/root/.claude/uploads/49bbb69d-2d00-4d26-a75a-90a2b5b6fb23/90b2a2fa-_______________.xlsx"
DST = "/home/user/Panse-System/5-订单总表_merged.xlsx"

shutil.copy2(SRC, DST)

wb = openpyxl.load_workbook(DST, data_only=False)
print("Sheets:", wb.sheetnames)

OLD_SHEET = "5-订单总表"
NEW_SHEET = "5-订单总表修改"

ws_old = wb[OLD_SHEET]
ws_new = wb[NEW_SHEET]

HEADER_ROW = 2  # row 2 has headers

def get_headers(ws, header_row):
    headers = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column - 1  # 0-based
    return headers

old_headers = get_headers(ws_old, HEADER_ROW)
new_headers = get_headers(ws_new, HEADER_ROW)

print("Old headers:", list(old_headers.keys())[:10], "...")
print("New headers:", list(new_headers.keys()))

ORDER_NO_COL = "订单编号"

# Columns NOT to copy (analysis / formula / platform / key columns)
SKIP_COLS = {
    "平台",           # system-set
    "订单编号",       # key column
    "是否补单",       # platform data
    "下单日期",       # platform data
    "产品理论成本",   # formula
    "总成本",         # formula
    "平台服务费",     # formula
    "税费",           # formula
    "店铺实收金额",   # formula
    "工厂补偿",       # formula
    "物流补偿",       # formula
    "补偿总金额",     # formula
    "订单利润",       # formula
    "买家应付金额",   # platform data
    "对应新订单号",   # analysis (old sheet only)
    "匹配结论",       # analysis
    "佳宝买家应付",   # analysis
    "金额差异",       # analysis
    "差异原因批注",   # analysis
    "⚠️问题标注",    # analysis
}

# Manually-entered columns that are candidates for merging
MERGE_COLS = [col for col in new_headers if col not in SKIP_COLS]
print("\nColumns eligible for merging:")
for c in MERGE_COLS:
    print(" ", c)

# Blank sentinel values — treat these as "empty"
BLANK_VALS = {None, "", "-", "待匹配", "无", "N/A"}

def is_blank(val):
    if val is None:
        return True
    s = str(val).strip()
    return s in BLANK_VALS

# Build lookup: order_no -> {col_name: cell_value} from old sheet
old_data = {}
old_order_col_idx = old_headers.get(ORDER_NO_COL)
if old_order_col_idx is None:
    raise ValueError(f"'{ORDER_NO_COL}' column not found in old sheet")

for row in ws_old.iter_rows(min_row=HEADER_ROW + 1):
    order_cell = row[old_order_col_idx]
    order_no = order_cell.value
    if order_no is None:
        continue
    order_no = str(order_no).strip()
    row_data = {}
    for col_name in MERGE_COLS:
        if col_name in old_headers:
            old_col_idx = old_headers[col_name]
            if old_col_idx < len(row):
                row_data[col_name] = row[old_col_idx].value
    old_data[order_no] = row_data

print(f"\nOld sheet orders loaded: {len(old_data)}")

# Now iterate new sheet and fill blanks
new_order_col_idx = new_headers.get(ORDER_NO_COL)
if new_order_col_idx is None:
    raise ValueError(f"'{ORDER_NO_COL}' column not found in new sheet")

filled_count = 0
rows_touched = 0

for row in ws_new.iter_rows(min_row=HEADER_ROW + 1):
    order_cell = row[new_order_col_idx]
    order_no = order_cell.value
    if order_no is None:
        continue
    order_no = str(order_no).strip()

    if order_no not in old_data:
        continue  # no old data for this order

    old_row = old_data[order_no]
    row_filled = 0

    for col_name in MERGE_COLS:
        if col_name not in new_headers:
            continue
        new_col_idx = new_headers[col_name]
        if new_col_idx >= len(row):
            continue

        new_cell = row[new_col_idx]
        new_val = new_cell.value

        if not is_blank(new_val):
            continue  # already has data — don't touch

        old_val = old_row.get(col_name)
        if is_blank(old_val):
            continue  # old sheet also empty — nothing to copy

        new_cell.value = old_val
        row_filled += 1
        filled_count += 1

    if row_filled > 0:
        rows_touched += 1

print(f"\nMerge complete:")
print(f"  Orders matched (touched): {rows_touched}")
print(f"  Cells filled: {filled_count}")

wb.save(DST)
print(f"\nSaved to: {DST}")

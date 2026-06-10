"""
从 CSV / TSV / Excel(xlsx) 批量写入 kb_entries（问法 + 话术）。

推荐：首行表头，列名支持中英文——问法 / question，答 / answer / 话术；
可选列：类型(entry_type)、start_at、end_at。
亦可无表头：每行前两列为问法、答（制表符或逗号分隔）。
"""

from __future__ import annotations

import csv
import io
import sqlite3
import uuid
from pathlib import Path

from apps.core.crm.events import ensure_brand_row, ensure_shop_row, now_iso


_HEADER_SYNONYMS = {
    "question": "question",
    "问法": "question",
    "标题": "question",
    "问": "question",
    "q": "question",
    "answer": "answer",
    "答": "answer",
    "答法": "answer",
    "话术": "answer",
    "内容": "answer",
    "reply": "answer",
    "a": "answer",
    "entry_type": "entry_type",
    "类型": "entry_type",
    "type": "entry_type",
    "start_at": "start_at",
    "开始": "start_at",
    "end_at": "end_at",
    "结束": "end_at",
}


def _looks_like_header_row(cells: list[str]) -> bool:
    if len(cells) < 2:
        return False
    blob = " ".join(c.strip().lower() for c in cells[:8])
    needles = (
        "question",
        "answer",
        "问法",
        "答",
        "话术",
        "标题",
        "entry",
        "类型",
    )
    return any(n in blob for n in needles)


def _normalize_cell(s: str | None) -> str | None:
    if s is None:
        return None
    t = str(s).strip()
    return t if t else None


def _trim_matrix(rows: list[list[str]]) -> list[list[str]]:
    out = []
    for row in rows:
        cells = [("" if c is None else str(c)).strip() for c in row]
        if any(cells):
            out.append(cells)
    return out


def parse_kb_table_rows(matrix: list[list[str]]) -> list[dict[str, str | None]]:
    """将二维表（单元格均为字符串）解析为 kb 行。"""
    rows = _trim_matrix(matrix)
    if not rows:
        return []

    out: list[dict[str, str | None]] = []
    start_idx = 0
    col_map: dict[int, str] = {}

    if _looks_like_header_row(rows[0]):
        for i, h in enumerate(rows[0]):
            key = _HEADER_SYNONYMS.get(h.strip()) or _HEADER_SYNONYMS.get(
                h.strip().lower()
            )
            if key:
                col_map[i] = key
        start_idx = 1
        if "question" not in col_map.values() or "answer" not in col_map.values():
            col_map = {}
            start_idx = 0

    if not col_map:
        for row in rows:
            if len(row) < 2:
                continue
            q = _normalize_cell(row[0])
            a = _normalize_cell(row[1])
            et = _normalize_cell(row[2]) if len(row) > 2 else None
            if q and a:
                out.append(
                    {
                        "question": q,
                        "answer": a,
                        "entry_type": et or "normal",
                        "start_at": None,
                        "end_at": None,
                    }
                )
        return out

    idx_q = next(k for k, v in col_map.items() if v == "question")
    idx_a = next(k for k, v in col_map.items() if v == "answer")
    idx_et = next((k for k, v in col_map.items() if v == "entry_type"), None)
    idx_sa = next((k for k, v in col_map.items() if v == "start_at"), None)
    idx_ea = next((k for k, v in col_map.items() if v == "end_at"), None)

    for row in rows[start_idx:]:
        def cell(i: int | None) -> str | None:
            if i is None or i >= len(row):
                return None
            return _normalize_cell(row[i])

        q = cell(idx_q)
        a = cell(idx_a)
        if not q or not a:
            continue
        out.append(
            {
                "question": q,
                "answer": a,
                "entry_type": cell(idx_et) or "normal",
                "start_at": cell(idx_sa),
                "end_at": cell(idx_ea),
            }
        )
    return out


def parse_kb_import_text(raw_text: str) -> list[dict[str, str | None]]:
    """解析 UTF-8 / GB18030 文本为 kb 行。"""
    text = raw_text.lstrip("\ufeff").strip()
    if not text:
        return []
    first_line = text.splitlines()[0]
    delim = "\t" if first_line.count("\t") >= first_line.count(",") else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    return parse_kb_table_rows(rows)


def parse_kb_import_file(path: Path) -> list[dict[str, str | None]]:
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("gb18030", errors="replace")
    return parse_kb_import_text(text)


def _wide_header_row(row: list[str]) -> bool:
    joined = " ".join(row[:8] if row else [])
    if "涉及产品" in joined:
        return True
    if len(row) >= 2 and "产品" in (row[0] or "") and ("问题" in joined or "分类" in joined):
        return True
    return False


def _pad_row(row: list[str], n: int) -> list[str]:
    out = [("" if c is None else str(c)).strip() for c in row]
    while len(out) < n:
        out.append("")
    return out


def parse_wide_kb_rows(matrix: list[list[str]]) -> list[dict[str, str | None]]:
    """
    宽表：A=涉及产品（锚定，禁止改写）、B/C/D 可为旧分类列、E/F/G 为补充线索。
    返回行含 product_anchor / hint_* / sheet_row（Excel 1 起行号），供 AI 再生成 question/answer/entry_type。
    """
    rows = _trim_matrix(matrix)
    if not rows:
        return []
    start = 1 if _wide_header_row(rows[0]) else 0
    out: list[dict[str, str | None]] = []
    for i in range(start, len(rows)):
        r = _pad_row(rows[i], 7)
        a, b, c, d, e, f, g = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        if not a:
            continue
        if a in ("涉及产品", "产品", "产品信息") and i == start:
            continue
        excel_row = i + 1
        out.append(
            {
                "product_anchor": a,
                "legacy_b": b or None,
                "legacy_c": c or None,
                "legacy_d": d or None,
                "hint_e": e or None,
                "hint_f": f or None,
                "hint_g": g or None,
                "sheet_row": str(excel_row),
            }
        )
    return out


def parse_kb_import_xlsx(path: Path) -> list[dict[str, str | None]]:
    """解析 Excel .xlsx / .xlsm（首个工作表）；若检测到「涉及产品」宽表头则走宽表解析。"""
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("导入 Excel 需要安装 openpyxl：pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        matrix: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            matrix.append(
                ["" if c is None else str(c).strip() for c in row],
            )
    finally:
        wb.close()
    if matrix and _wide_header_row(matrix[0]):
        return parse_wide_kb_rows(matrix)
    return parse_kb_table_rows(matrix)


def write_kb_wide_marks_to_xlsx(
    src: Path,
    marks: list[tuple[int, str, str]],
) -> Path:
    """将 AI 对 E/F 列的批注写回副本（不覆盖原文件）。返回新文件路径。"""
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("需要 openpyxl") from e
    out = src.parent / f"{src.stem}_import_marked{src.suffix}"
    wb = openpyxl.load_workbook(src)
    try:
        ws = wb.active
        for row_idx, me, mf in marks:
            ws.cell(row=row_idx, column=5, value=me)
            ws.cell(row=row_idx, column=6, value=mf)
        wb.save(out)
    finally:
        wb.close()
    return out


def clear_kb_for_shop(conn: sqlite3.Connection, *, brand_id: str, shop_id: str) -> int:
    """删除店铺下话术及关联向量；返回删除的 kb_entries 行数。"""
    cur = conn.execute(
        "SELECT kb_id FROM kb_entries WHERE brand_id = ? AND shop_id = ?",
        (brand_id, shop_id),
    )
    ids = [str(r[0]) for r in cur.fetchall()]
    for kb_id in ids:
        conn.execute("DELETE FROM kb_embeddings WHERE kb_id = ?", (kb_id,))
    conn.execute(
        "DELETE FROM kb_entries WHERE brand_id = ? AND shop_id = ?",
        (brand_id, shop_id),
    )
    conn.commit()
    return len(ids)


def parse_kb_import_any(path: Path) -> list[dict[str, str | None]]:
    suf = path.suffix.lower()
    if suf == ".xls":
        raise RuntimeError("暂不支持旧版 .xls，请在 Excel 中「另存为」.xlsx 后再导入")
    if suf in (".xlsx", ".xlsm"):
        return parse_kb_import_xlsx(path)
    return parse_kb_import_file(path)


def _shop_code_fallback(shop_id: str) -> str:
    if ":" in shop_id:
        return shop_id.split(":")[-1].strip() or shop_id
    return shop_id


def import_kb_rows(
    conn: sqlite3.Connection,
    *,
    brand_id: str,
    shop_id: str,
    rows: list[dict[str, str | None]],
) -> int:
    """写入 kb_entries；返回成功插入条数。"""
    ensure_brand_row(conn, brand_id=brand_id)
    code = _shop_code_fallback(shop_id)
    ensure_shop_row(
        conn,
        brand_id=brand_id,
        shop_id=shop_id,
        shop_code=code,
        display_name=code,
    )
    ts = now_iso()
    n = 0
    for it in rows:
        q = (it.get("question") or "").strip()
        a = (it.get("answer") or "").strip()
        if not q or not a:
            continue
        et = (it.get("entry_type") or "normal").strip() or "normal"
        sa = it.get("start_at")
        ea = it.get("end_at")
        kb_id = str(uuid.uuid4())
        tags = (it.get("kb_tags") or "").strip()
        conn.execute(
            """
            INSERT INTO kb_entries(
              kb_id, brand_id, shop_id, question, answer, entry_type,
              enabled, start_at, end_at, created_at, updated_at, kb_tags
            ) VALUES (?,?,?,?,?,?,1,?,?,?,?,?)
            """,
            (
                kb_id,
                brand_id,
                shop_id,
                q,
                a,
                et,
                sa,
                ea,
                ts,
                ts,
                tags,
            ),
        )
        n += 1
    conn.commit()
    return n

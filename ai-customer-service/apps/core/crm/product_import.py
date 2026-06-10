"""畔色等产品知识库 Excel 导入 → products / product_skus。"""

from __future__ import annotations

import re
import sqlite3
import uuid
from pathlib import Path

from openpyxl import load_workbook

from apps.core.crm.events import now_iso


def _cell(row: tuple[object, ...], idx: int | None) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def _norm_key(s: object) -> str:
    return re.sub(r"\s+", "", str(s or "").strip())


def _pick_sheet(wb) -> object:
    for name in ("产品List", "产品", "Sheet1"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def _header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    aliases = {
        "产品编码": "product_code",
        "产品code": "product_code",
        "sku编码": "sku_code",
        "SKU编码": "sku_code",
        "sku": "sku_name",
        "SKU": "sku_name",
        "类目": "category",
        "畔色品名": "name",
        "品名": "name",
        "淘宝商品名": "taobao_name",
        "小红书商品名": "xhs_name",
        "淘宝链接": "product_link",
        "图片链接": "image_link",
        "文案": "copywriting",
        "主材": "main_material",
        "辅材": "sub_material",
        "尺寸明细": "size_details",
        "可定制范围": "customization_scope",
    }
    out: dict[str, int] = {}
    for i, raw in enumerate(header_row):
        k = _norm_key(raw)
        if not k:
            continue
        canon = aliases.get(k, k)
        out[str(canon)] = i
        out[k] = i
    return out


def import_product_workbook(
    path: Path,
    conn: sqlite3.Connection,
    *,
    brand_id: str,
    shop_id: str,
) -> tuple[int, int]:
    """
    从 .xlsx 导入产品行。表头需含「产品编码」；可选 SKU、尺寸明细、可定制范围等。
    返回 (写入/更新的产品行数, SKU 行数)。
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb)
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return 0, 0
        col = _header_map(tuple(header_row))
        pc_i = col.get("product_code")
        if pc_i is None:
            raise ValueError("表头中未找到「产品编码」列")

        ts = now_iso()
        n_prod = 0
        n_sku = 0
        product_ids: dict[str, str] = {}

        for row in it:
            if not row:
                continue
            pcode = _cell(row, pc_i)
            if not pcode:
                continue

            name = _cell(row, col.get("name")) or pcode
            category = _cell(row, col.get("category"))
            product_link = _cell(row, col.get("product_link"))
            copywriting = _cell(row, col.get("copywriting"))
            main_material = _cell(row, col.get("main_material"))
            sub_material = _cell(row, col.get("sub_material"))
            size_details = _cell(row, col.get("size_details"))
            customization_scope = _cell(row, col.get("customization_scope"))

            cur = conn.execute(
                """
                SELECT product_id FROM products
                WHERE brand_id = ? AND shop_id = ? AND product_code = ?
                LIMIT 1
                """,
                (brand_id, shop_id, pcode),
            )
            ex = cur.fetchone()
            if ex:
                pid = str(ex[0])
            else:
                pid = str(uuid.uuid4())
            product_ids[pcode] = pid

            conn.execute(
                """
                INSERT INTO products(
                  product_id, brand_id, shop_id, product_code, category, name,
                  product_link, listing_status, copywriting, main_material, sub_material,
                  size_details, customization_scope, created_at, updated_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(brand_id, shop_id, product_code) DO UPDATE SET
                  category=excluded.category,
                  name=excluded.name,
                  product_link=excluded.product_link,
                  copywriting=excluded.copywriting,
                  main_material=excluded.main_material,
                  sub_material=excluded.sub_material,
                  size_details=excluded.size_details,
                  customization_scope=excluded.customization_scope,
                  updated_at=excluded.updated_at
                """,
                (
                    pid,
                    brand_id,
                    shop_id,
                    pcode,
                    category or None,
                    name,
                    product_link or None,
                    None,
                    copywriting or None,
                    main_material or None,
                    sub_material or None,
                    size_details or None,
                    customization_scope or "",
                    ts,
                    ts,
                ),
            )
            n_prod += 1

            sku_code = _cell(row, col.get("sku_code"))
            sku_name = _cell(row, col.get("sku_name"))
            if not sku_code and not sku_name:
                continue
            sku_id = str(uuid.uuid4())
            conn.execute(
                """
                INSERT INTO product_skus(
                  sku_id, product_id, brand_id, shop_id, sku_name, sku_code, created_at
                ) VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(brand_id, shop_id, sku_code) DO UPDATE SET
                  product_id=excluded.product_id,
                  sku_name=excluded.sku_name
                """,
                (
                    sku_id,
                    pid,
                    brand_id,
                    shop_id,
                    sku_name or sku_code,
                    sku_code or sku_name,
                    ts,
                ),
            )
            n_sku += 1

        conn.commit()
        return n_prod, n_sku
    finally:
        wb.close()

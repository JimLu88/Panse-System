#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整合最近三个月(xlsx)与1月1-28(csv)订单数据 → 生成 5-订单总表修改，并与原 5-订单总表对比。"""

import csv, io, re, shutil
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict, Counter
from openpyxl import load_workbook
from copy import copy

ROOT = Path("/home/user/Panse-System")
SRC  = ROOT / "畔色系统总表_已填.xlsx"
NEW_XLSX = Path("/root/.claude/uploads/163899d7-1b44-4c4c-9f50-7d1eb7294c82/19cbaccd-3.46.2______.xlsx")
NEW_CSV  = Path("/root/.claude/uploads/163899d7-1b44-4c4c-9f50-7d1eb7294c82/41b57df9-1.11.28___.csv")

# ── helpers ───────────────────────────────────────────────────────────────────
def to_date(v):
    """订单创建时间 → date object (or None)."""
    if v is None or v == "": return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    # take date part before space
    s2 = s.split(" ")[0].replace("/", "-")
    try:
        y,m,d = s2.split("-"); return date(int(y),int(m),int(d))
    except: return None

def extract_sku(attr):
    """商品属性 '颜色分类:xxx[规格];安装方式:yyy' → 'xxx' (去掉[..]与;后内容)."""
    if not attr: return ""
    s = str(attr)
    # split on ; first
    first = re.split(r"[;；]", s)[0]
    # remove '颜色分类:' / '颜色分类：' prefix
    first = re.sub(r"^[^:：]*[:：]", "", first)
    # strip bracket spec [长45cm]
    first = re.sub(r"\[[^\]]*\]", "", first)
    return first.strip()

def code_from_merchant(mc):
    """商家编码 PPS+13位 → 老格式 P+11位; 纯11位数字 → P+11位; 否则原样/空."""
    if mc is None: return ""
    s = str(mc).strip()
    if not s: return ""
    if s.upper().startswith("PPS"):
        digits = s[3:]
        return "P" + digits[:11] if len(digits) >= 11 else "P" + digits
    # CSV historical: 11-digit numeric商家编码
    if s.isdigit() and len(s) >= 11:
        return "P" + s[:11]
    return ""

def is_sci(v):
    """检测科学计数法损坏的订单号 (含 E+ 或 e+)."""
    return v is not None and re.search(r"[eE]\+?\d+", str(v)) is not None

# ── PHASE 1: 解析最近三个月 xlsx ──────────────────────────────────────────────
wb_new = load_workbook(NEW_XLSX, read_only=True)

# 订单报表: per-order → addr / logistics / amount / status / pay no
o_ws = wb_new["订单报表"]
order_rpt = {}   # order_no -> dict
oh = [c.value for c in next(o_ws.iter_rows(min_row=1, max_row=1))]
def oidx(name): return oh.index(name)
for row in o_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    no = str(row[0]).strip()
    order_rpt[no] = {
        "pay_no": row[oidx("支付单号")],
        "buyer_due": row[oidx("买家应付货款")],
        "status": row[oidx("订单状态")],
        "addr": row[oidx("收货地址")],
        "logistics_no": row[oidx("物流单号")],
        "logistics_co": row[oidx("物流公司")],
        "create_time": row[oidx("订单创建时间")],
        "remark": row[oidx("商家备注")],
    }

# 发货报表: per-order → 客户姓名/电话
f_ws = wb_new["发货报表"]
fh = [c.value for c in next(f_ws.iter_rows(min_row=1, max_row=1))]
def fidx(name): return fh.index(name) if name in fh else None
ship = {}
for row in f_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    no = str(row[0]).strip()
    ship[no] = {
        "name": row[fidx("收货人姓名")] if fidx("收货人姓名") is not None else None,
        "phone": (row[fidx("联系手机")] if fidx("联系手机") is not None else None)
                 or (row[fidx("联系电话")] if fidx("联系电话") is not None else None),
        "addr": row[fidx("收货地址")] if fidx("收货地址") is not None else None,
    }

# 销售明细: per-line (主要数据源)
s_ws = wb_new["销售明细"]
sh = [c.value for c in next(s_ws.iter_rows(min_row=1, max_row=1))]
def sidx(name): return sh.index(name)
lines = []   # each = dict for one 订单总表 row
for row in s_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    main_no = str(row[sidx("主订单编号")]).strip()
    rpt = order_rpt.get(main_no, {})
    shp = ship.get(main_no, {})
    lines.append({
        "src": "xlsx",
        "order_no": main_no,
        "create": rpt.get("create_time") or row[sidx("订单创建时间")],
        "name": shp.get("name"),
        "phone": shp.get("phone"),
        "addr": shp.get("addr") or rpt.get("addr"),
        "code": code_from_merchant(row[sidx("商家编码")] or row[sidx("外部系统编号")]),
        "title": row[sidx("商品标题")],
        "sku": extract_sku(row[sidx("商品属性")]),
        "qty": row[sidx("购买数量")],
        "logistics_co": rpt.get("logistics_co"),
        "logistics_no": rpt.get("logistics_no"),
        "buyer_due": row[sidx("买家应付货款")],
        "status": row[sidx("订单状态")],
        "pay_no": row[sidx("支付单号")],
        "refund_status": row[sidx("退款状态")],
        "refund_amt": row[sidx("退款金额")],
        "order_no_bad": is_sci(row[sidx("主订单编号")]),
    })
wb_new.close()

# ── 解析 CSV (1月历史, 订单号损坏) ────────────────────────────────────────────
with open(NEW_CSV, "rb") as fh_:
    txt = fh_.read().decode("gbk")
rdr = list(csv.reader(io.StringIO(txt)))
ch = rdr[0]
def cidx(name): return ch.index(name)
for r in rdr[1:]:
    if not r or not r[0]: continue
    main_no = str(r[cidx("主订单编号")]).strip()
    lines.append({
        "src": "csv",
        "order_no": main_no,
        "create": r[cidx("订单创建时间")],
        "name": None, "phone": None, "addr": None,
        "code": code_from_merchant(r[cidx("商家编码")] or r[cidx("外部系统编号")]),
        "title": r[cidx("标题")],
        "sku": extract_sku(r[cidx("商品属性")]),
        "qty": r[cidx("购买数量")],
        "logistics_co": None, "logistics_no": None,
        "buyer_due": r[cidx("买家应付货款")],
        "status": r[cidx("订单状态")],
        "pay_no": r[cidx("支付单号")],
        "refund_status": r[cidx("退款状态")],
        "refund_amt": r[cidx("退款金额")],
        "order_no_bad": is_sci(r[cidx("主订单编号")]),
    })

print(f"整合后总行数: {len(lines)} (xlsx销售明细 + csv)")

# ── PHASE 2: 写入 5-订单总表修改 ──────────────────────────────────────────────
wb = load_workbook(SRC)
old = wb["5-订单总表"]
if "5-订单总表修改" in wb.sheetnames:
    del wb["5-订单总表修改"]
ws = wb.create_sheet("5-订单总表修改", index=wb.sheetnames.index("5-订单总表")+1)

# copy row1 (title) + row2 (headers) cells + style
for r in (1, 2):
    for c in range(1, old.max_column+1):
        src_cell = old.cell(r, c); dst = ws.cell(r, c)
        dst.value = src_cell.value
        if src_cell.has_style:
            dst.font = copy(src_cell.font); dst.fill = copy(src_cell.fill)
            dst.border = copy(src_cell.border); dst.alignment = copy(src_cell.alignment)
            dst.number_format = src_cell.number_format
# column widths
for col, dim in old.column_dimensions.items():
    ws.column_dimensions[col].width = dim.width
ws.freeze_panes = "A3"
ws.cell(1,1).value = "5-订单总表修改"
# clear old AR1 note; set a fresh one later

# write data
r = 3
for ln in lines:
    ws.cell(r, 1).value  = "淘宝"                       # A 平台
    ws.cell(r, 2).value  = ln["order_no"]               # B 订单编号
    # C 是否补单 (手填) 留空
    d = to_date(ln["create"])
    ws.cell(r, 4).value  = d                            # D 下单日期
    if d: ws.cell(r,4).number_format = "yyyy-mm-dd"
    ws.cell(r, 5).value  = ln["name"]                   # E 客户姓名
    ws.cell(r, 6).value  = ln["phone"]                  # F 联系电话
    ws.cell(r, 7).value  = ln["addr"]                   # G 收货地址
    # H 发货日期 (源无) 留空
    ws.cell(r, 9).value  = ln["code"]                   # I 产品编码
    ws.cell(r,10).value  = ln["title"]                  # J 产品名称
    ws.cell(r,11).value  = ln["sku"]                    # K SKU
    # L 是否定制 (手填) 留空
    try: ws.cell(r,13).value = int(ln["qty"]) if ln["qty"] not in (None,"") else None
    except: ws.cell(r,13).value = ln["qty"]             # M 数量
    # N 锁定状态 (手填) 留空
    ws.cell(r,15).value  = ln["logistics_co"]           # O 物流公司
    ws.cell(r,16).value  = ln["logistics_no"]           # P 物流单号
    # Q 安装工单号 (手填) 留空
    # R 产品理论成本 = 公式
    ws.cell(r,18).value  = (
        f"=IFERROR(SUMPRODUCT(('2-定价总表'!$A$3:$A$493=\"PPS\"&MID(I{r},2,99))*"
        f"('2-定价总表'!$C$3:$C$493=K{r})*('2-定价总表'!$T$3:$T$493))*IF(M{r}=\"\",1,M{r}),\"\")"
    )
    # S 产品实际成本 / T-AA 各项费用 (手填) 留空
    ws.cell(r,28).value  = f"=SUM(S{r}:AA{r})"          # AB 总成本
    ws.cell(r,29).value  = ln["buyer_due"]              # AC 买家应付金额
    ws.cell(r,30).value  = f"=AC{r}*0.006"              # AD 平台服务费
    ws.cell(r,31).value  = f"=AC{r}*0.02"               # AE 税费
    ws.cell(r,32).value  = f"=AC{r}-AD{r}-AE{r}"        # AF 店铺实收金额
    # AG 工厂补偿 / AH 物流补偿 (手填) 留空
    ws.cell(r,35).value  = f"=SUM(AG{r}:AH{r})"         # AI 补偿总金额
    ws.cell(r,36).value  = f"=AF{r}-AB{r}+AI{r}"        # AJ 订单利润
    ws.cell(r,37).value  = f'=IF(C{r}="是","杭州","江西工厂")'  # AK 发货仓库
    ws.cell(r,38).value  = ln["status"]                 # AL 订单状态
    ws.cell(r,39).value  = "待匹配"                      # AM 支付宝流水号
    ws.cell(r,40).value  = ln["refund_status"]          # AN 退款状态
    ws.cell(r,41).value  = ln["refund_amt"]             # AO 退款金额
    # AP 退款日期 (源无) 留空
    # AQ 备注 (手填) 留空
    flags = []
    if ln["order_no_bad"]:
        flags.append("⚠️ 历史订单号精度丢失(CSV导出为科学计数法,需人工核对)")
    if not ln["code"]:
        flags.append("⚠️ 无商家编码,产品编码待补")
    if ln["src"] == "csv":
        flags.append("ℹ️ 1月历史订单")
    ws.cell(r,44).value  = " ".join(flags) if flags else "✓"
    r += 1

last_new = r - 1
ws.cell(1,44).value = (
    f"本表为整合数据(最近三个月销售明细+1月历史CSV)，共 {len(lines)} 行。"
    f"手填字段(是否补单/是否定制/锁定状态/安装工单号/实际成本/各项费用/补偿/备注)留空，"
    f"由后期人工填写。产品理论成本及成本/利润列均为公式。"
)

# ── PHASE 3: 与原 5-订单总表对比 ──────────────────────────────────────────────
# 原表 per-order
old_by_order = {}
for rr in range(3, old.max_row+1):
    no = old.cell(rr,2).value
    if no in (None,""): continue
    old_by_order.setdefault(str(no).strip(), []).append({
        "code": old.cell(rr,9).value, "sku": old.cell(rr,11).value,
        "qty": old.cell(rr,13).value, "due": old.cell(rr,29).value,
        "status": old.cell(rr,38).value, "refund_status": old.cell(rr,40).value,
        "refund_amt": old.cell(rr,41).value, "date": old.cell(rr,4).value,
    })

# 新数据 per-order (聚合)
new_by_order = defaultdict(list)
for ln in lines:
    new_by_order[ln["order_no"]].append(ln)

old_keys = set(old_by_order)
new_keys = set(k for k in new_by_order if not is_sci(k))   # 排除损坏号
csv_bad  = [k for k in new_by_order if is_sci(k)]

added   = sorted(new_keys - old_keys)
removed = sorted(old_keys - new_keys)
common  = sorted(new_keys & old_keys)

NO_REFUND = {"", "0", "0.0", "无退款申请", "没有申请退款", "none"}
def norm(v):
    if v is None: return ""
    s = str(v).strip()
    try:
        f = float(s); return str(int(f)) if f==int(f) else str(f)
    except: return s
def norm_refund(v):
    """退款相关：无退款的各种写法统一视为 0/无。"""
    s = norm(v).lower()
    return "无" if s in NO_REFUND else s

# 分类统计：金额变化 / 状态变化 / 编码SKU泛化(老表手工细化)
changed_amount = []   # 买家应付金额变化（最重要）
changed_status = []   # 订单状态/退款状态变化
changed_code   = []   # 产品编码或SKU变化（多为老表人工细化 vs 平台原始泛称）
for k in common:
    o = old_by_order[k][0]
    ns = new_by_order[k]
    n = ns[0]
    # 金额
    if norm(o["due"]) != norm(n["buyer_due"]) and n["buyer_due"] not in (None,""):
        changed_amount.append((k, o["due"], n["buyer_due"]))
    # 状态
    sdiffs = []
    if norm(o["status"]) != norm(n["status"]):
        sdiffs.append(("订单状态", o["status"], n["status"]))
    if norm_refund(o["refund_status"]) != norm_refund(n["refund_status"]):
        sdiffs.append(("退款状态", o["refund_status"], n["refund_status"]))
    if norm_refund(o["refund_amt"]) != norm_refund(n["refund_amt"]):
        sdiffs.append(("退款金额", o["refund_amt"], n["refund_amt"]))
    if sdiffs:
        changed_status.append((k, sdiffs))
    # 编码/SKU
    cdiffs = []
    if norm(o["code"]) != norm(n["code"]):
        cdiffs.append(("产品编码", o["code"], n["code"]))
    if norm(o["sku"]) != norm(n["sku"]):
        cdiffs.append(("SKU", o["sku"], n["sku"]))
    if cdiffs:
        changed_code.append((k, cdiffs))

# ── 写对比报告 ───────────────────────────────────────────────────────────────
rep = []
rep.append("# 订单数据对比报告\n")
rep.append(f"- 原 5-订单总表：{len(old_keys)} 个订单")
rep.append(f"- 新整合数据：{len(new_keys)} 个有效订单 + {len(csv_bad)} 个订单号损坏(CSV历史) ，共 {len(lines)} 行明细")
rep.append("")
rep.append(f"## 一、新增订单（原表没有）：{len(added)} 个")
for k in added[:40]:
    n = new_by_order[k][0]
    rep.append(f"  - {k}  {str(n['title'])[:24]}  {n['sku']}  应付{n['buyer_due']}  {n['status']}")
if len(added) > 40: rep.append(f"  …… 其余 {len(added)-40} 个略")
rep.append("")
rep.append("## 二、原有订单发生变化（按重要程度分三类）")
rep.append("")
rep.append(f"### 2.1 【最重要】买家应付金额变化：{len(changed_amount)} 个")
rep.append("    （这是真正的金额差异，多为后期补差价/改价）")
for k, ov, nv in changed_amount:
    try: delta = float(nv) - float(ov); ds = f"（{'+' if delta>=0 else ''}{round(delta,2)}）"
    except: ds = ""
    rep.append(f"  ● {k}: 原 {ov} → 新 {nv} {ds}")
rep.append("")
rep.append(f"### 2.2 订单状态/退款状态变化：{len(changed_status)} 个")
rep.append("    （订单从待发货→已发货→交易成功等推进，属正常更新）")
for k, diffs in changed_status[:50]:
    parts = "；".join(f"{f}: {ov}→{nv}" for f,ov,nv in diffs)
    rep.append(f"  ● {k}: {parts}")
if len(changed_status) > 50: rep.append(f"  …… 其余 {len(changed_status)-50} 个略")
rep.append("")
rep.append(f"### 2.3 产品编码/SKU 名称变化：{len(changed_code)} 个")
rep.append("    （绝大多数是：老表里是人工细化后的规格名/编码，新导出是平台原始的泛称")
rep.append("     如「定制尺寸」「差价」「尺寸微定制」或留空。这属于老表手工录入 vs 平台原始数据的差异，")
rep.append("     按你的要求新表保留平台原始值、不回填老表手工内容。）")
for k, diffs in changed_code[:30]:
    parts = "；".join(f"{f}: [{ov}]→[{nv}]" for f,ov,nv in diffs)
    rep.append(f"  ● {k}: {parts}")
if len(changed_code) > 30: rep.append(f"  …… 其余 {len(changed_code)-30} 个略")
rep.append("")
rep.append(f"## 三、原表有、新数据没有的订单：{len(removed)} 个")
rep.append("  （多为更早历史订单，不在最近三个月导出范围内，属正常）")
for k in removed[:25]:
    rep.append(f"  - {k}")
if len(removed) > 25: rep.append(f"  …… 其余 {len(removed)-25} 个略")
rep.append("")
rep.append(f"## 四、订单号损坏（CSV 历史，科学计数法）：{len(csv_bad)} 个")
for k in csv_bad:
    n = new_by_order[k][0]
    rep.append(f"  - {k}  {str(n['title'])[:24]}  创建{n['create']}  应付{n['buyer_due']}")

report = "\n".join(rep)
(ROOT / "订单对比报告.md").write_text(report, encoding="utf-8")

wb.save(SRC)
print(f"已写入 5-订单总表修改 (rows 3–{last_new})，并保存。")
print(f"\n{'='*70}\n{report}\n{'='*70}")
EOF_DONE = 1
